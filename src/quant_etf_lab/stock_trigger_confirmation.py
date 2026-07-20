"""Build a manual review queue for targets that also have fresh trigger signals."""

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Mapping

import pandas as pd


CONFIRMATION_COLUMNS = [
    "code",
    "name",
    "primary_confirmation",
    "pretrade_decision",
    "manual_review_checklist",
    "portfolio_target_weight",
    "trigger_run_time",
    "trigger_signal_type",
    "trigger_action",
    "trigger_score",
    "trigger_score_level",
    "trigger_pct",
    "trigger_last",
    "trigger_reason",
    "support",
    "pressure",
    "stop_loss",
    "source_strategy_profile",
    "risk_filter_status",
    "execution_gate_action",
    "target_action",
    "unrealized_return",
    "market_cap_yi",
    "review_bucket",
    "review_stage",
    "review_reason",
    "manual_review_state",
    "broker_action",
    "research_only",
]

DECISION_TEMPLATE_USER_FILL_COLUMNS = [
    "manual_decision_to_fill",
    "pullback_confirmed_to_fill",
    "risk_ok_to_fill",
    "position_ok_to_fill",
    "stop_loss_ok_to_fill",
    "position_weight_to_fill",
    "manual_note_to_fill",
    "next_review_date_to_fill",
    "reviewed_by_to_fill",
]

DECISION_TEMPLATE_COLUMNS = [
    "confirmation_rank",
    "code",
    "name",
    *DECISION_TEMPLATE_USER_FILL_COLUMNS,
    "allowed_manual_decisions",
    "allowed_yes_no_values",
    "portfolio_target_weight",
    "trigger_score",
    "trigger_score_level",
    "trigger_action",
    "trigger_pct",
    "trigger_last",
    "support",
    "pressure",
    "stop_loss",
    "manual_review_checklist",
    "pretrade_decision",
    "review_bucket",
    "review_reason",
    "broker_action",
    "research_only",
]

DECISION_TEMPLATE_CHINESE_HEADERS = {
    "confirmation_rank": "确认排序",
    "code": "代码",
    "name": "名称",
    "manual_decision_to_fill": "人工决策（填写）",
    "pullback_confirmed_to_fill": "是否回踩确认（填写）",
    "risk_ok_to_fill": "风险是否可接受（填写）",
    "position_ok_to_fill": "仓位规则是否确认（填写）",
    "stop_loss_ok_to_fill": "止损规则是否确认（填写）",
    "position_weight_to_fill": "目标仓位（填写）",
    "manual_note_to_fill": "人工备注（填写）",
    "next_review_date_to_fill": "下次复核日期（填写）",
    "reviewed_by_to_fill": "复核人（填写）",
    "allowed_manual_decisions": "允许人工决策值",
    "allowed_yes_no_values": "允许确认值",
    "portfolio_target_weight": "模型目标仓位",
    "trigger_score": "触发分数",
    "trigger_score_level": "触发等级",
    "trigger_action": "触发动作",
    "trigger_pct": "当日涨跌幅",
    "trigger_last": "最新价",
    "support": "支撑位",
    "pressure": "压力位",
    "stop_loss": "止损位",
    "manual_review_checklist": "人工复核清单",
    "pretrade_decision": "交易前决策",
    "review_bucket": "复核分组",
    "review_reason": "复核原因",
    "broker_action": "券商动作",
    "research_only": "仅研究标记",
}

SHADOW_JOURNAL_COLUMNS = [
    "trade_date",
    "code",
    "name",
    "entry_date",
    "entry_price",
    "close_price",
    "high_price",
    "low_price",
    "planned_shadow_weight",
    "stop_loss",
    "close_return_pct",
    "portfolio_close_return_pct",
    "result_status",
    "shadow_review_decision",
    "trigger_pct",
    "trigger_pct_bucket",
    "trigger_score",
    "trigger_score_level",
    "market_data_source_kind",
    "market_data_source",
    "journal_event_id",
    "broker_action",
    "research_only",
]


@dataclass(frozen=True)
class StockTriggerConfirmationResult:
    output_dir: Path
    table_path: Path
    snapshot_path: Path
    report_path: Path
    decision_template_path: Path
    decision_template_json_path: Path
    decision_template_report_path: Path
    decision_template_xlsx_path: Path
    decision_template_zh_xlsx_path: Path
    snapshot: dict[str, Any]
    table: pd.DataFrame
    decision_template: pd.DataFrame


@dataclass(frozen=True)
class StockTriggerConfirmationDecisionApplyResult:
    template_path: Path
    output_dir: Path
    applied_path: Path
    snapshot_path: Path
    report_path: Path
    snapshot: dict[str, Any]
    applied: pd.DataFrame


@dataclass(frozen=True)
class StockTriggerConfirmationShadowBuyPlanResult:
    applied_path: Path
    template_path: Path
    output_dir: Path
    plan_path: Path
    snapshot_path: Path
    report_path: Path
    snapshot: dict[str, Any]
    plan: pd.DataFrame


@dataclass(frozen=True)
class StockTriggerConfirmationNextOpenCheckResult:
    plan_path: Path
    output_dir: Path
    check_path: Path
    snapshot_path: Path
    report_path: Path
    snapshot: dict[str, Any]
    checks: pd.DataFrame


@dataclass(frozen=True)
class StockTriggerConfirmationShadowTrackingResult:
    check_path: Path
    output_dir: Path
    tracking_path: Path
    snapshot_path: Path
    report_path: Path
    snapshot: dict[str, Any]
    tracking: pd.DataFrame


@dataclass(frozen=True)
class StockTriggerConfirmationCloseReviewResult:
    tracking_path: Path
    output_dir: Path
    review_path: Path
    bucket_summary_path: Path
    snapshot_path: Path
    report_path: Path
    snapshot: dict[str, Any]
    review: pd.DataFrame
    bucket_summary: pd.DataFrame


@dataclass(frozen=True)
class StockTriggerConfirmationDailyFlowResult:
    output_dir: Path
    journal_path: Path
    snapshot_path: Path
    report_path: Path
    snapshot: dict[str, Any]
    decision_apply_result: StockTriggerConfirmationDecisionApplyResult
    shadow_plan_result: StockTriggerConfirmationShadowBuyPlanResult
    open_check_result: StockTriggerConfirmationNextOpenCheckResult
    shadow_tracking_result: StockTriggerConfirmationShadowTrackingResult
    close_review_result: StockTriggerConfirmationCloseReviewResult
    journal: pd.DataFrame


def _clean_code(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits.zfill(6) if digits else text.zfill(6)


def _text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return default
    return text


def _number(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or pd.isna(value):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _xlsx_scalar(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def _template_header_to_machine_columns(frame: pd.DataFrame) -> pd.DataFrame:
    inverse = {label: column for column, label in DECISION_TEMPLATE_CHINESE_HEADERS.items()}
    rename = {column: inverse.get(column, column) for column in frame.columns}
    return frame.rename(columns=rename)


def _load_stock_trigger_confirmation_decision_template(path: str | Path) -> pd.DataFrame:
    resolved = Path(path)
    suffix = resolved.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        frame = pd.read_excel(resolved, sheet_name=0, dtype={"code": str, "代码": str})
    else:
        try:
            frame = pd.read_csv(resolved, dtype={"code": str, "代码": str}, encoding="utf-8-sig")
        except UnicodeDecodeError:
            frame = pd.read_csv(resolved, dtype={"code": str, "代码": str}, encoding="gbk")
    if frame.empty:
        return pd.DataFrame(columns=DECISION_TEMPLATE_COLUMNS)
    frame = _template_header_to_machine_columns(frame)
    if "code" in frame.columns:
        frame = frame.copy()
        frame["code"] = frame["code"].map(_clean_code)
    return frame


def _normalise_manual_decision(value: Any) -> str:
    text = _text(value).lower()
    aliases = {
        "approve_shadow": "approve_shadow",
        "buy": "approve_shadow",
        "shadow_buy": "approve_shadow",
        "approve": "approve_shadow",
        "买入": "approve_shadow",
        "批准买入": "approve_shadow",
        "同意买入": "approve_shadow",
        "允许买入": "approve_shadow",
        "进入影子复核": "approve_shadow",
        "watch": "watch",
        "观察": "watch",
        "继续观察": "watch",
        "reject": "reject",
        "不买": "reject",
        "放弃": "reject",
        "拒绝": "reject",
        "needs_retrigger": "needs_retrigger",
        "retrigger": "needs_retrigger",
        "重新触发": "needs_retrigger",
        "等待重新触发": "needs_retrigger",
    }
    return aliases.get(text, "")


def _normalise_yes_no_unknown(value: Any) -> str:
    text = _text(value).lower()
    aliases = {
        "yes": "yes",
        "y": "yes",
        "true": "yes",
        "1": "yes",
        "是": "yes",
        "确认": "yes",
        "通过": "yes",
        "no": "no",
        "n": "no",
        "false": "no",
        "0": "no",
        "否": "no",
        "不通过": "no",
        "unknown": "unknown",
        "unk": "unknown",
        "不确定": "unknown",
        "未知": "unknown",
        "待确认": "unknown",
    }
    return aliases.get(text, "")


def _parse_position_weight(value: Any) -> float | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    if not text:
        return None
    try:
        if text.endswith("%"):
            return float(text[:-1].strip()) / 100.0
        return float(value)
    except (TypeError, ValueError):
        return None


def _decision_action(decision: str) -> str:
    return {
        "approve_shadow": "allow_live_shadow_review",
        "watch": "keep_watch",
        "reject": "reject_candidate",
        "needs_retrigger": "wait_for_retrigger",
    }.get(decision, "invalid_decision")


def _bool_value(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return _text(value).lower() in {"true", "1", "yes", "y", "是"}


def _instrument_check_status(code: str, name: Any = "") -> tuple[str, str]:
    clean = _clean_code(code)
    upper_name = _text(name).upper()
    fund_prefixes = ("50", "51", "52", "56", "58", "15", "16", "18")
    fund_name_tokens = ("ETF", "LOF", "FUND", "REIT", "基金")
    if clean.startswith(fund_prefixes) or any(token in upper_name for token in fund_name_tokens):
        return "blocked_non_stock_fund_or_etf", "no_etf_or_fund_rule"
    return "ok_stock", ""


def _trigger_pct_bucket(value: Any) -> str:
    pct = _number(value, default=float("nan"))
    if pd.isna(pct):
        return "unknown"
    if pct < 5.0:
        return "below_5"
    if pct < 7.0:
        return "gain_5_to_7"
    if pct < 9.8:
        return "gain_7_to_9_8"
    return "limit_up_or_near"


def _load_stock_trigger_confirmation_applied(path: str | Path) -> pd.DataFrame:
    resolved = Path(path)
    suffix = resolved.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        frame = pd.read_excel(resolved, sheet_name=0, dtype={"code": str})
    else:
        try:
            frame = pd.read_csv(resolved, dtype={"code": str}, encoding="utf-8-sig")
        except UnicodeDecodeError:
            frame = pd.read_csv(resolved, dtype={"code": str}, encoding="gbk")
    if frame.empty:
        return frame
    frame = frame.copy()
    if "code" in frame.columns:
        frame["code"] = frame["code"].map(_clean_code)
    return frame


def _load_stock_trigger_confirmation_shadow_plan(path: str | Path) -> pd.DataFrame:
    resolved = Path(path)
    suffix = resolved.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        frame = pd.read_excel(resolved, sheet_name=0, dtype={"code": str})
    else:
        try:
            frame = pd.read_csv(resolved, dtype={"code": str}, encoding="utf-8-sig")
        except UnicodeDecodeError:
            frame = pd.read_csv(resolved, dtype={"code": str}, encoding="gbk")
    if frame.empty:
        return frame
    frame = frame.copy()
    if "code" in frame.columns:
        frame["code"] = frame["code"].map(_clean_code)
    return frame


def _load_stock_trigger_confirmation_next_open_check(path: str | Path) -> pd.DataFrame:
    resolved = Path(path)
    suffix = resolved.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        frame = pd.read_excel(resolved, sheet_name=0, dtype={"code": str})
    else:
        try:
            frame = pd.read_csv(resolved, dtype={"code": str}, encoding="utf-8-sig")
        except UnicodeDecodeError:
            frame = pd.read_csv(resolved, dtype={"code": str}, encoding="gbk")
    if frame.empty:
        return frame
    frame = frame.copy()
    if "code" in frame.columns:
        frame["code"] = frame["code"].map(_clean_code)
    return frame


def _load_stock_trigger_confirmation_shadow_tracking(path: str | Path) -> pd.DataFrame:
    resolved = Path(path)
    suffix = resolved.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        frame = pd.read_excel(resolved, sheet_name=0, dtype={"code": str})
    else:
        try:
            frame = pd.read_csv(resolved, dtype={"code": str}, encoding="utf-8-sig")
        except UnicodeDecodeError:
            frame = pd.read_csv(resolved, dtype={"code": str}, encoding="gbk")
    if frame.empty:
        return frame
    frame = frame.copy()
    if "code" in frame.columns:
        frame["code"] = frame["code"].map(_clean_code)
    return frame


def _read_open_prices_from_snapshot_csv(path: Path) -> dict[str, float]:
    try:
        frame = pd.read_csv(path, dtype={"security_code": str}, encoding="utf-8-sig")
    except (OSError, UnicodeDecodeError, pd.errors.EmptyDataError, pd.errors.ParserError):
        return {}
    if frame.empty or "security_code" not in frame.columns or "open_price" not in frame.columns:
        return {}
    prices: dict[str, float] = {}
    for raw in frame.to_dict(orient="records"):
        code = _clean_code(raw.get("security_code"))
        price = _number(raw.get("open_price"))
        if code and price > 0:
            prices[code] = price
    return prices


def _load_market_rows_by_code(
    trade_date: str | None,
    *,
    market_data_dir: str | Path = "D:/codex/daily-market-data",
    ingest_project_dir: str | Path = "D:/codex/2026-06-15-exchange-data-ingest",
) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    if not trade_date:
        return {}, {"market_data_status": "not_requested", "market_data_source": None}
    try:
        from .market_data_source import load_market_snapshot_rows

        result = load_market_snapshot_rows(
            trade_date=trade_date,
            market="all",
            daily_data_dir=market_data_dir,
            ingest_project_dir=ingest_project_dir,
            require_success=True,
        )
    except Exception as error:
        return {}, {
            "market_data_status": "load_failed",
            "market_data_source": None,
            "market_data_error": str(error),
            "market_data_trade_date": pd.Timestamp(pd.to_datetime(trade_date)).strftime("%Y-%m-%d"),
        }
    rows_by_code: dict[str, dict[str, Any]] = {}
    for raw in result.rows:
        code = _clean_code(raw.get("security_code"))
        if code:
            rows_by_code[code] = raw
    fetch_status = result.fetch_status
    return rows_by_code, {
        "market_data_status": "ok" if rows_by_code else "empty",
        "market_data_source_kind": result.source_kind,
        "market_data_source": str(result.source_path) if result.source_path is not None else None,
        "market_data_trade_date": result.trade_date,
        "fetch_status": getattr(fetch_status, "status", None) if fetch_status is not None else None,
        "fetch_run_time": getattr(fetch_status, "run_time", None) if fetch_status is not None else None,
        "fetch_row_count": getattr(fetch_status, "row_count", None) if fetch_status is not None else None,
    }


def _load_market_data_utils(ingest_project_dir: Path) -> Any | None:
    script_path = ingest_project_dir / "scripts" / "market_data_utils.py"
    if not script_path.exists():
        return None
    try:
        import importlib.util

        spec = importlib.util.spec_from_file_location("_quant_etf_lab_market_data_utils", script_path)
        if spec is None or spec.loader is None:
            return None
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module
    except Exception:
        return None


def _load_open_prices_from_market_data(
    trade_date: str | None,
    *,
    market_data_dir: str | Path = "D:/codex/daily-market-data",
    ingest_project_dir: str | Path = "D:/codex/2026-06-15-exchange-data-ingest",
) -> tuple[dict[str, float], dict[str, Any]]:
    if not trade_date:
        return {}, {"market_data_status": "not_requested", "market_data_source": None}
    try:
        from .market_data_source import load_market_snapshot_rows

        result = load_market_snapshot_rows(
            trade_date=trade_date,
            market="all",
            daily_data_dir=market_data_dir,
            ingest_project_dir=ingest_project_dir,
            require_success=True,
        )
    except Exception as error:
        return {}, {
            "market_data_status": "load_failed",
            "market_data_source": None,
            "market_data_error": str(error),
            "market_data_trade_date": pd.Timestamp(pd.to_datetime(trade_date)).strftime("%Y-%m-%d"),
        }
    prices = {}
    for raw in result.rows:
        code = _clean_code(raw.get("security_code"))
        price = _number(raw.get("open_price"))
        if code and price > 0:
            prices[code] = price
    fetch_status = result.fetch_status
    return prices, {
        "market_data_status": "ok" if prices else "empty",
        "market_data_source_kind": result.source_kind,
        "market_data_source": str(result.source_path) if result.source_path is not None else None,
        "market_data_trade_date": result.trade_date,
        "fetch_status": getattr(fetch_status, "status", None) if fetch_status is not None else None,
        "fetch_run_time": getattr(fetch_status, "run_time", None) if fetch_status is not None else None,
        "fetch_row_count": getattr(fetch_status, "row_count", None) if fetch_status is not None else None,
    }


def _records_from_template_by_code(template_path: str | Path) -> dict[str, dict[str, Any]]:
    template = _load_stock_trigger_confirmation_decision_template(template_path)
    if template.empty or "code" not in template.columns:
        return {}
    out: dict[str, dict[str, Any]] = {}
    for raw in template.to_dict(orient="records"):
        code = _clean_code(raw.get("code"))
        if code:
            out[code] = raw
    return out


def _render_shadow_buy_plan_report(plan: pd.DataFrame, snapshot: dict[str, Any]) -> str:
    if plan.empty:
        rows = "| - | - | - | - | - | - | - |\n"
    else:
        lines = []
        for row in plan.to_dict(orient="records"):
            lines.append(
                "| `{code}` | {name} | {weight:.2%} | {max_open:.2f} | {stop:.2f} | `{status}` | `{warning}` |".format(
                    code=_text(row.get("code")),
                    name=_text(row.get("name")),
                    weight=_number(row.get("planned_shadow_weight")),
                    max_open=_number(row.get("max_allowed_open_price")),
                    stop=_number(row.get("stop_loss")),
                    status=_text(row.get("execution_status")),
                    warning=_text(row.get("plan_warnings")),
                )
            )
        rows = "\n".join(lines) + "\n"
    return f"""# 影子买入计划
<!-- stock_trigger_confirmation_shadow_buy_plan -->

本计划仅用于实盘前置观察和模拟复核；不会生成券商订单，不会自动下单。

## 摘要

| 项目 | 值 |
| --- | ---: |
| 人工决策应用表 | `{snapshot.get("applied_path")}` |
| 确认模板 | `{snapshot.get("template_path")}` |
| 影子候选数 | {snapshot.get("shadow_candidate_count")} |
| 等待开盘检查数 | {snapshot.get("ready_for_open_check_count")} |
| 阻断数 | {snapshot.get("blocked_count")} |
| 单票影子仓位上限 | {snapshot.get("max_single_position_weight"):.2%} |
| 跳空上限 | {snapshot.get("max_open_gap_pct"):.2%} |
| 券商动作 | `{snapshot.get("broker_action")}` |

## 明细

| 代码 | 名称 | 计划影子仓位 | 最高允许开盘价 | 止损位 | 状态 | 警告 |
| --- | --- | ---: | ---: | ---: | --- | --- |
{rows}

## 下一步

- 明日开盘后补入实际开盘价，对比最高允许开盘价。
- 若开盘跳空过高、止损位无效或仓位超限，取消影子买入复核。
- 通过检查后也只进入影子复核，不代表券商下单。
"""


def _render_next_open_check_report(checks: pd.DataFrame, snapshot: dict[str, Any]) -> str:
    if checks.empty:
        rows = "| - | - | - | - | - | - | - |\n"
    else:
        lines = []
        for row in checks.to_dict(orient="records"):
            lines.append(
                "| `{code}` | {name} | `{status}` | `{decision}` | {open_price:.2f} | {max_open:.2f} | `{broker}` |".format(
                    code=_text(row.get("code")),
                    name=_text(row.get("name")),
                    status=_text(row.get("open_check_status")),
                    decision=_text(row.get("shadow_review_decision")),
                    open_price=_number(row.get("actual_open_price")),
                    max_open=_number(row.get("max_allowed_open_price")),
                    broker=_text(row.get("broker_action")),
                )
            )
        rows = "\n".join(lines) + "\n"
    return f"""# 影子买入下一交易日开盘检查
<!-- stock_trigger_confirmation_next_open_check -->

本检查只用于判断影子买入复核是否继续；不会生成券商订单，不会自动下单。

## 摘要

| 项目 | 值 |
| --- | ---: |
| 影子计划 | `{snapshot.get("plan_path")}` |
| 交易日 | `{snapshot.get("trade_date")}` |
| 市场数据状态 | `{snapshot.get("market_data_status")}` |
| 待数据行数 | {snapshot.get("pending_market_data_count")} |
| 通过行数 | {snapshot.get("passed_count")} |
| 跳空阻断行数 | {snapshot.get("blocked_open_gap_count")} |
| 其他阻断行数 | {snapshot.get("blocked_other_count")} |
| 券商动作 | `{snapshot.get("broker_action")}` |

## 明细

| 代码 | 名称 | 检查状态 | 影子复核决策 | 实际开盘价 | 最高允许开盘价 | 券商动作 |
| --- | --- | --- | --- | ---: | ---: | --- |
{rows}
"""


def _render_shadow_tracking_report(tracking: pd.DataFrame, snapshot: dict[str, Any]) -> str:
    if tracking.empty:
        rows = "| - | - | - | - | - | - |\n"
    else:
        lines = []
        for row in tracking.to_dict(orient="records"):
            lines.append(
                "| `{code}` | {name} | {entry:.2f} | {weight:.2%} | {stop:.2f} | {risk:.2%} |".format(
                    code=_text(row.get("code")),
                    name=_text(row.get("name")),
                    entry=_number(row.get("entry_price")),
                    weight=_number(row.get("planned_shadow_weight")),
                    stop=_number(row.get("stop_loss")),
                    risk=_number(row.get("portfolio_risk_at_stop_pct")),
                )
            )
        rows = "\n".join(lines) + "\n"
    return f"""# 影子买入跟踪台账
<!-- stock_trigger_confirmation_shadow_tracking -->

本台账只记录研究用影子持仓，不会生成券商订单，不会自动下单。

## 摘要

| 项目 | 值 |
| --- | ---: |
| 开盘检查表 | `{snapshot.get("check_path")}` |
| 活跃影子持仓数 | {snapshot.get("active_shadow_position_count")} |
| 计划影子总仓位 | {snapshot.get("planned_shadow_total_weight"):.2%} |
| 止损组合风险 | {snapshot.get("portfolio_risk_at_stop_pct"):.2%} |
| 券商动作 | `{snapshot.get("broker_action")}` |

## 明细

| 代码 | 名称 | 影子入场价 | 影子仓位 | 止损位 | 组合止损风险 |
| --- | --- | ---: | ---: | ---: | ---: |
{rows}
"""


def _bucket_summary_from_close_review(review: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "trigger_pct_bucket",
        "row_count",
        "avg_close_return_pct",
        "win_rate",
        "stop_loss_touched_count",
        "avg_portfolio_close_return_pct",
    ]
    if review.empty:
        return pd.DataFrame(columns=columns)
    grouped_rows: list[dict[str, Any]] = []
    for bucket, group in review.groupby("trigger_pct_bucket", dropna=False):
        close_returns = pd.to_numeric(group["close_return_pct"], errors="coerce")
        portfolio_returns = pd.to_numeric(group["portfolio_close_return_pct"], errors="coerce")
        grouped_rows.append(
            {
                "trigger_pct_bucket": _text(bucket, "unknown"),
                "row_count": int(len(group)),
                "avg_close_return_pct": float(close_returns.mean()) if not close_returns.empty else 0.0,
                "win_rate": float((close_returns > 0).mean()) if not close_returns.empty else 0.0,
                "stop_loss_touched_count": int((group["close_review_status"] == "stop_loss_touched").sum()),
                "avg_portfolio_close_return_pct": float(portfolio_returns.mean()) if not portfolio_returns.empty else 0.0,
            }
        )
    return pd.DataFrame(grouped_rows, columns=columns).sort_values("trigger_pct_bucket").reset_index(drop=True)


def _render_close_review_report(
    review: pd.DataFrame,
    bucket_summary: pd.DataFrame,
    snapshot: dict[str, Any],
) -> str:
    if review.empty:
        rows = "| - | - | - | - | - | - | - |\n"
    else:
        lines = []
        for row in review.to_dict(orient="records"):
            lines.append(
                "| `{code}` | {name} | {entry:.2f} | {close:.2f} | {ret:.2%} | `{status}` | `{decision}` |".format(
                    code=_text(row.get("code")),
                    name=_text(row.get("name")),
                    entry=_number(row.get("entry_price")),
                    close=_number(row.get("close_price")),
                    ret=_number(row.get("close_return_pct")),
                    status=_text(row.get("close_review_status")),
                    decision=_text(row.get("shadow_review_decision")),
                )
            )
        rows = "\n".join(lines) + "\n"
    if bucket_summary.empty:
        bucket_rows = "| - | - | - | - | - |\n"
    else:
        bucket_lines = []
        for row in bucket_summary.to_dict(orient="records"):
            bucket_lines.append(
                "| `{bucket}` | {count} | {ret:.2%} | {win:.2%} | {stops} |".format(
                    bucket=_text(row.get("trigger_pct_bucket")),
                    count=int(_number(row.get("row_count"))),
                    ret=_number(row.get("avg_close_return_pct")),
                    win=_number(row.get("win_rate")),
                    stops=int(_number(row.get("stop_loss_touched_count"))),
                )
            )
        bucket_rows = "\n".join(bucket_lines) + "\n"
    return f"""# 影子持仓收盘复盘
<!-- stock_trigger_confirmation_close_review -->

本复盘只用于研究影子持仓表现；不会生成券商订单，不会自动下单。ETF/fund candidates are excluded by rule.

## 摘要

| 项目 | 值 |
| --- | ---: |
| 跟踪台账 | `{snapshot.get("tracking_path")}` |
| 交易日 | `{snapshot.get("trade_date")}` |
| 复盘行数 | {snapshot.get("review_count")} |
| ETF/基金排除数 | {snapshot.get("non_stock_excluded_count")} |
| 触及止损数 | {snapshot.get("stop_loss_touched_count")} |
| 平均收盘收益 | {snapshot.get("avg_close_return_pct"):.2%} |
| 组合影子收益 | {snapshot.get("portfolio_close_return_pct"):.2%} |
| 行情来源 | `{snapshot.get("market_data_source")}` |
| 券商动作 | `{snapshot.get("broker_action")}` |

## 明细

| 代码 | 名称 | 入场价 | 收盘价 | 收盘收益 | 状态 | 决策 |
| --- | --- | ---: | ---: | ---: | --- | --- |
{rows}

## 分桶收益

| 触发涨幅桶 | 样本数 | 平均收益 | 胜率 | 触及止损数 |
| --- | ---: | ---: | ---: | ---: |
{bucket_rows}
"""


def build_stock_trigger_confirmation_shadow_buy_plan(
    applied_path: str | Path,
    template_path: str | Path,
    *,
    output_dir: str | Path | None = None,
    max_single_position_weight: float = 0.20,
    max_open_gap_pct: float = 0.03,
) -> StockTriggerConfirmationShadowBuyPlanResult:
    resolved_applied = Path(applied_path)
    resolved_template = Path(template_path)
    resolved_output = Path(output_dir) if output_dir is not None else resolved_applied.parent
    resolved_output.mkdir(parents=True, exist_ok=True)
    applied = _load_stock_trigger_confirmation_applied(resolved_applied)
    template_by_code = _records_from_template_by_code(resolved_template)
    rows: list[dict[str, Any]] = []
    if not applied.empty:
        for raw in applied.to_dict(orient="records"):
            code = _clean_code(raw.get("code"))
            if not code:
                continue
            if _text(raw.get("manual_decision_normalized")) != "approve_shadow":
                continue
            if not _bool_value(raw.get("live_shadow_review_allowed")):
                continue
            source = template_by_code.get(code, {})
            name = _text(raw.get("name") or source.get("name"))
            instrument_status, instrument_reason = _instrument_check_status(code, name)
            requested_weight = _number(raw.get("position_weight_normalized"))
            planned_weight = min(max(requested_weight, 0.0), max_single_position_weight)
            weight_status = (
                "capped_to_single_position_limit"
                if requested_weight > max_single_position_weight
                else "ok"
            )
            trigger_last = _number(source.get("trigger_last"))
            stop_loss = _number(source.get("stop_loss"))
            support = _number(source.get("support"))
            pressure = _number(source.get("pressure"))
            max_allowed_open = round(trigger_last * (1.0 + max_open_gap_pct), 4) if trigger_last > 0 else 0.0
            stop_loss_status = "ok" if stop_loss > 0 and (trigger_last <= 0 or stop_loss < trigger_last) else "invalid_stop_loss"
            warnings = []
            if weight_status != "ok":
                warnings.append("high_position_weight_capped")
            if stop_loss_status != "ok":
                warnings.append(stop_loss_status)
            if trigger_last <= 0:
                warnings.append("missing_trigger_last")
            if instrument_status != "ok_stock":
                warnings.append(instrument_reason)
            blocked = stop_loss_status != "ok" or trigger_last <= 0 or planned_weight <= 0
            rows.append(
                {
                    "code": code,
                    "name": name,
                    "instrument_check_status": instrument_status,
                    "instrument_exclusion_reason": instrument_reason,
                    "shadow_plan_action": (
                        "blocked_non_stock_fund_or_etf"
                        if instrument_status != "ok_stock"
                        else "prepare_shadow_buy_review"
                    ),
                    "execution_status": (
                        "blocked_non_stock_fund_or_etf"
                        if instrument_status != "ok_stock"
                        else "blocked_precheck"
                        if blocked
                        else "pending_next_open_check"
                    ),
                    "next_open_check_status": "pending_next_open",
                    "requested_position_weight": requested_weight,
                    "planned_shadow_weight": planned_weight,
                    "max_single_position_weight": float(max_single_position_weight),
                    "weight_check_status": weight_status,
                    "trigger_pct": _number(source.get("trigger_pct")),
                    "trigger_score": _number(source.get("trigger_score")),
                    "trigger_score_level": _text(source.get("trigger_score_level")),
                    "trigger_last": trigger_last,
                    "max_open_gap_pct": float(max_open_gap_pct),
                    "max_allowed_open_price": max_allowed_open,
                    "support": support,
                    "pressure": pressure,
                    "stop_loss": stop_loss,
                    "stop_loss_check_status": stop_loss_status,
                    "manual_note": _text(raw.get("manual_note")),
                    "next_review_date": _text(raw.get("next_review_date")),
                    "reviewed_by": _text(raw.get("reviewed_by")),
                    "plan_warnings": "|".join(warnings),
                    "broker_action": "none",
                    "research_only": True,
                }
            )
    plan = pd.DataFrame(
        rows,
        columns=[
            "code",
            "name",
            "instrument_check_status",
            "instrument_exclusion_reason",
            "shadow_plan_action",
            "execution_status",
            "next_open_check_status",
            "requested_position_weight",
            "planned_shadow_weight",
            "max_single_position_weight",
            "weight_check_status",
            "trigger_pct",
            "trigger_score",
            "trigger_score_level",
            "trigger_last",
            "max_open_gap_pct",
            "max_allowed_open_price",
            "support",
            "pressure",
            "stop_loss",
            "stop_loss_check_status",
            "manual_note",
            "next_review_date",
            "reviewed_by",
            "plan_warnings",
            "broker_action",
            "research_only",
        ],
    )
    snapshot = {
        "schema_version": "stock-trigger-confirmation-shadow-buy-plan-v1",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "applied_path": str(resolved_applied),
        "template_path": str(resolved_template),
        "plan_row_count": int(len(plan)),
        "shadow_candidate_count": int((plan["execution_status"] == "pending_next_open_check").sum()) if not plan.empty else 0,
        "ready_for_open_check_count": int((plan["execution_status"] == "pending_next_open_check").sum()) if not plan.empty else 0,
        "blocked_count": int((plan["execution_status"] == "blocked_precheck").sum()) if not plan.empty else 0,
        "non_stock_excluded_count": int((plan["execution_status"] == "blocked_non_stock_fund_or_etf").sum()) if not plan.empty else 0,
        "max_single_position_weight": float(max_single_position_weight),
        "max_open_gap_pct": float(max_open_gap_pct),
        "broker_action": "none",
        "research_only": True,
    }
    plan_path = resolved_output / "stock_trigger_confirmation_shadow_buy_plan.csv"
    snapshot_path = resolved_output / "stock_trigger_confirmation_shadow_buy_plan.json"
    report_path = resolved_output / "stock_trigger_confirmation_shadow_buy_plan.md"
    plan.to_csv(plan_path, index=False, encoding="utf-8-sig")
    snapshot_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    report_path.write_text(_render_shadow_buy_plan_report(plan, snapshot), encoding="utf-8")
    return StockTriggerConfirmationShadowBuyPlanResult(
        applied_path=resolved_applied,
        template_path=resolved_template,
        output_dir=resolved_output,
        plan_path=plan_path,
        snapshot_path=snapshot_path,
        report_path=report_path,
        snapshot=snapshot,
        plan=plan,
    )


def build_stock_trigger_confirmation_next_open_check(
    plan_path: str | Path,
    *,
    output_dir: str | Path | None = None,
    open_prices: Mapping[str, float] | None = None,
    trade_date: str | None = None,
    market_data_dir: str | Path = "D:/codex/daily-market-data",
    ingest_project_dir: str | Path = "D:/codex/2026-06-15-exchange-data-ingest",
) -> StockTriggerConfirmationNextOpenCheckResult:
    resolved_plan = Path(plan_path)
    resolved_output = Path(output_dir) if output_dir is not None else resolved_plan.parent
    resolved_output.mkdir(parents=True, exist_ok=True)
    plan = _load_stock_trigger_confirmation_shadow_plan(resolved_plan)
    if open_prices is None:
        loaded_prices, market_payload = _load_open_prices_from_market_data(
            trade_date,
            market_data_dir=market_data_dir,
            ingest_project_dir=ingest_project_dir,
        )
    else:
        loaded_prices = {_clean_code(code): float(price) for code, price in open_prices.items()}
        market_payload = {
            "market_data_status": "manual_open_prices",
            "market_data_source": "open_prices_argument",
            "market_data_trade_date": trade_date,
        }

    rows: list[dict[str, Any]] = []
    if not plan.empty:
        for raw in plan.to_dict(orient="records"):
            code = _clean_code(raw.get("code"))
            if not code:
                continue
            plan_status = _text(raw.get("execution_status"))
            actual_open = loaded_prices.get(code)
            max_allowed = _number(raw.get("max_allowed_open_price"))
            planned_weight = _number(raw.get("planned_shadow_weight"))
            stop_status = _text(raw.get("stop_loss_check_status"))
            if plan_status != "pending_next_open_check":
                check_status = "skipped_not_pending"
                decision = "no_action"
            elif actual_open is None or actual_open <= 0:
                check_status = "pending_market_data"
                decision = "wait_for_next_open"
            elif stop_status != "ok":
                check_status = "blocked_stop_loss"
                decision = "cancel_shadow_buy_review"
            elif planned_weight <= 0:
                check_status = "blocked_weight"
                decision = "cancel_shadow_buy_review"
            elif max_allowed > 0 and actual_open > max_allowed:
                check_status = "blocked_open_gap"
                decision = "cancel_shadow_buy_review"
            else:
                check_status = "passed"
                decision = "allow_shadow_tracking"
            rows.append(
                {
                    "code": code,
                    "name": _text(raw.get("name")),
                    "trade_date": pd.Timestamp(pd.to_datetime(trade_date)).strftime("%Y-%m-%d") if trade_date else "",
                    "open_check_status": check_status,
                    "shadow_review_decision": decision,
                    "actual_open_price": actual_open if actual_open is not None else None,
                    "max_allowed_open_price": max_allowed,
                    "planned_shadow_weight": planned_weight,
                    "stop_loss": _number(raw.get("stop_loss")),
                    "trigger_pct": _number(raw.get("trigger_pct")),
                    "trigger_score": _number(raw.get("trigger_score")),
                    "trigger_score_level": _text(raw.get("trigger_score_level")),
                    "source_execution_status": plan_status,
                    "broker_action": "none",
                    "research_only": True,
                }
            )
    checks = pd.DataFrame(
        rows,
        columns=[
            "code",
            "name",
            "trade_date",
            "open_check_status",
            "shadow_review_decision",
            "actual_open_price",
            "max_allowed_open_price",
            "planned_shadow_weight",
            "stop_loss",
            "trigger_pct",
            "trigger_score",
            "trigger_score_level",
            "source_execution_status",
            "broker_action",
            "research_only",
        ],
    )
    snapshot = {
        "schema_version": "stock-trigger-confirmation-next-open-check-v1",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "plan_path": str(resolved_plan),
        "trade_date": pd.Timestamp(pd.to_datetime(trade_date)).strftime("%Y-%m-%d") if trade_date else None,
        "check_count": int(len(checks)),
        "pending_market_data_count": int((checks["open_check_status"] == "pending_market_data").sum()) if not checks.empty else 0,
        "passed_count": int((checks["open_check_status"] == "passed").sum()) if not checks.empty else 0,
        "blocked_open_gap_count": int((checks["open_check_status"] == "blocked_open_gap").sum()) if not checks.empty else 0,
        "blocked_other_count": int(checks["open_check_status"].isin(["blocked_stop_loss", "blocked_weight"]).sum()) if not checks.empty else 0,
        "broker_action": "none",
        "research_only": True,
        **market_payload,
    }
    check_path = resolved_output / "stock_trigger_confirmation_next_open_check.csv"
    snapshot_path = resolved_output / "stock_trigger_confirmation_next_open_check.json"
    report_path = resolved_output / "stock_trigger_confirmation_next_open_check.md"
    checks.to_csv(check_path, index=False, encoding="utf-8-sig")
    snapshot_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    report_path.write_text(_render_next_open_check_report(checks, snapshot), encoding="utf-8")
    return StockTriggerConfirmationNextOpenCheckResult(
        plan_path=resolved_plan,
        output_dir=resolved_output,
        check_path=check_path,
        snapshot_path=snapshot_path,
        report_path=report_path,
        snapshot=snapshot,
        checks=checks,
    )


def build_stock_trigger_confirmation_shadow_tracking(
    check_path: str | Path,
    *,
    output_dir: str | Path | None = None,
) -> StockTriggerConfirmationShadowTrackingResult:
    resolved_check = Path(check_path)
    resolved_output = Path(output_dir) if output_dir is not None else resolved_check.parent
    resolved_output.mkdir(parents=True, exist_ok=True)
    checks = _load_stock_trigger_confirmation_next_open_check(resolved_check)
    rows: list[dict[str, Any]] = []
    if not checks.empty:
        for raw in checks.to_dict(orient="records"):
            code = _clean_code(raw.get("code"))
            if not code:
                continue
            if _text(raw.get("open_check_status")) != "passed":
                continue
            if _text(raw.get("shadow_review_decision")) != "allow_shadow_tracking":
                continue
            entry_price = _number(raw.get("actual_open_price"))
            planned_weight = _number(raw.get("planned_shadow_weight"))
            stop_loss = _number(raw.get("stop_loss"))
            entry_to_stop = (stop_loss / entry_price - 1.0) if entry_price > 0 and stop_loss > 0 else 0.0
            portfolio_risk = planned_weight * entry_to_stop
            rows.append(
                {
                    "code": code,
                    "name": _text(raw.get("name")),
                    "shadow_position_status": "active_shadow_tracking",
                    "entry_date": _text(raw.get("trade_date")),
                    "entry_price": entry_price,
                    "planned_shadow_weight": planned_weight,
                    "stop_loss": stop_loss,
                    "trigger_pct": _number(raw.get("trigger_pct")),
                    "trigger_score": _number(raw.get("trigger_score")),
                    "trigger_score_level": _text(raw.get("trigger_score_level")),
                    "entry_to_stop_loss_pct": entry_to_stop,
                    "portfolio_risk_at_stop_pct": portfolio_risk,
                    "latest_price": entry_price,
                    "unrealized_return_pct": 0.0,
                    "next_action": "track_close_and_stop_loss",
                    "broker_action": "none",
                    "research_only": True,
                }
            )
    tracking = pd.DataFrame(
        rows,
        columns=[
            "code",
            "name",
            "shadow_position_status",
            "entry_date",
            "entry_price",
            "planned_shadow_weight",
            "stop_loss",
            "trigger_pct",
            "trigger_score",
            "trigger_score_level",
            "entry_to_stop_loss_pct",
            "portfolio_risk_at_stop_pct",
            "latest_price",
            "unrealized_return_pct",
            "next_action",
            "broker_action",
            "research_only",
        ],
    )
    snapshot = {
        "schema_version": "stock-trigger-confirmation-shadow-tracking-v1",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "check_path": str(resolved_check),
        "active_shadow_position_count": int(len(tracking)),
        "planned_shadow_total_weight": float(tracking["planned_shadow_weight"].sum()) if not tracking.empty else 0.0,
        "portfolio_risk_at_stop_pct": float(tracking["portfolio_risk_at_stop_pct"].sum()) if not tracking.empty else 0.0,
        "broker_action": "none",
        "research_only": True,
    }
    tracking_path = resolved_output / "stock_trigger_confirmation_shadow_tracking.csv"
    snapshot_path = resolved_output / "stock_trigger_confirmation_shadow_tracking.json"
    report_path = resolved_output / "stock_trigger_confirmation_shadow_tracking.md"
    tracking.to_csv(tracking_path, index=False, encoding="utf-8-sig")
    snapshot_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    report_path.write_text(_render_shadow_tracking_report(tracking, snapshot), encoding="utf-8")
    return StockTriggerConfirmationShadowTrackingResult(
        check_path=resolved_check,
        output_dir=resolved_output,
        tracking_path=tracking_path,
        snapshot_path=snapshot_path,
        report_path=report_path,
        snapshot=snapshot,
        tracking=tracking,
    )


def build_stock_trigger_confirmation_close_review(
    tracking_path: str | Path,
    *,
    output_dir: str | Path | None = None,
    trade_date: str | None = None,
    market_data_dir: str | Path = "D:/codex/daily-market-data",
    ingest_project_dir: str | Path = "D:/codex/2026-06-15-exchange-data-ingest",
) -> StockTriggerConfirmationCloseReviewResult:
    resolved_tracking = Path(tracking_path)
    resolved_output = Path(output_dir) if output_dir is not None else resolved_tracking.parent
    resolved_output.mkdir(parents=True, exist_ok=True)
    tracking = _load_stock_trigger_confirmation_shadow_tracking(resolved_tracking)
    market_rows, market_payload = _load_market_rows_by_code(
        trade_date,
        market_data_dir=market_data_dir,
        ingest_project_dir=ingest_project_dir,
    )
    rows: list[dict[str, Any]] = []
    non_stock_excluded_count = 0
    if not tracking.empty:
        for raw in tracking.to_dict(orient="records"):
            code = _clean_code(raw.get("code"))
            if not code:
                continue
            name = _text(raw.get("name"))
            instrument_status, instrument_reason = _instrument_check_status(code, name)
            if instrument_status != "ok_stock":
                non_stock_excluded_count += 1
                continue
            if _text(raw.get("shadow_position_status")) != "active_shadow_tracking":
                continue
            market_row = market_rows.get(code, {})
            entry_price = _number(raw.get("entry_price"))
            planned_weight = _number(raw.get("planned_shadow_weight"))
            stop_loss = _number(raw.get("stop_loss"))
            close_price = _number(market_row.get("close_price") or market_row.get("last_price"))
            high_price = _number(market_row.get("high_price"))
            low_price = _number(market_row.get("low_price"))
            if close_price <= 0:
                close_review_status = "pending_market_data"
                decision = "wait_for_close_data"
                close_return = 0.0
            else:
                close_return = close_price / entry_price - 1.0 if entry_price > 0 else 0.0
                if stop_loss > 0 and low_price > 0 and low_price <= stop_loss:
                    close_review_status = "stop_loss_touched"
                    decision = "review_shadow_exit"
                elif stop_loss > 0 and close_price <= stop_loss:
                    close_review_status = "close_below_stop_loss"
                    decision = "review_shadow_exit"
                else:
                    close_review_status = "continue_tracking"
                    decision = "continue_shadow_tracking"
            portfolio_return = planned_weight * close_return
            trigger_pct = _number(raw.get("trigger_pct"), default=float("nan"))
            rows.append(
                {
                    "code": code,
                    "name": name,
                    "trade_date": pd.Timestamp(pd.to_datetime(trade_date)).strftime("%Y-%m-%d") if trade_date else "",
                    "entry_date": _text(raw.get("entry_date")),
                    "entry_price": entry_price,
                    "close_price": close_price if close_price > 0 else None,
                    "high_price": high_price if high_price > 0 else None,
                    "low_price": low_price if low_price > 0 else None,
                    "planned_shadow_weight": planned_weight,
                    "stop_loss": stop_loss,
                    "close_return_pct": close_return,
                    "portfolio_close_return_pct": portfolio_return,
                    "close_review_status": close_review_status,
                    "shadow_review_decision": decision,
                    "trigger_pct": trigger_pct if not pd.isna(trigger_pct) else None,
                    "trigger_pct_bucket": _trigger_pct_bucket(trigger_pct),
                    "trigger_score": _number(raw.get("trigger_score")),
                    "trigger_score_level": _text(raw.get("trigger_score_level")),
                    "instrument_check_status": instrument_status,
                    "instrument_exclusion_reason": instrument_reason,
                    "broker_action": "none",
                    "research_only": True,
                }
            )
    review = pd.DataFrame(
        rows,
        columns=[
            "code",
            "name",
            "trade_date",
            "entry_date",
            "entry_price",
            "close_price",
            "high_price",
            "low_price",
            "planned_shadow_weight",
            "stop_loss",
            "close_return_pct",
            "portfolio_close_return_pct",
            "close_review_status",
            "shadow_review_decision",
            "trigger_pct",
            "trigger_pct_bucket",
            "trigger_score",
            "trigger_score_level",
            "instrument_check_status",
            "instrument_exclusion_reason",
            "broker_action",
            "research_only",
        ],
    )
    bucket_summary = _bucket_summary_from_close_review(review)
    close_returns = pd.to_numeric(review["close_return_pct"], errors="coerce") if not review.empty else pd.Series(dtype=float)
    snapshot = {
        "schema_version": "stock-trigger-confirmation-close-review-v1",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "tracking_path": str(resolved_tracking),
        "trade_date": pd.Timestamp(pd.to_datetime(trade_date)).strftime("%Y-%m-%d") if trade_date else None,
        "review_count": int(len(review)),
        "bucket_count": int(len(bucket_summary)),
        "non_stock_excluded_count": int(non_stock_excluded_count),
        "stop_loss_touched_count": int((review["close_review_status"] == "stop_loss_touched").sum()) if not review.empty else 0,
        "pending_market_data_count": int((review["close_review_status"] == "pending_market_data").sum()) if not review.empty else 0,
        "avg_close_return_pct": float(close_returns.mean()) if not close_returns.empty else 0.0,
        "portfolio_close_return_pct": float(review["portfolio_close_return_pct"].sum()) if not review.empty else 0.0,
        "broker_action": "none",
        "research_only": True,
        **market_payload,
    }
    review_path = resolved_output / "stock_trigger_confirmation_close_review.csv"
    bucket_summary_path = resolved_output / "stock_trigger_confirmation_close_review_buckets.csv"
    snapshot_path = resolved_output / "stock_trigger_confirmation_close_review.json"
    report_path = resolved_output / "stock_trigger_confirmation_close_review.md"
    review.to_csv(review_path, index=False, encoding="utf-8-sig")
    bucket_summary.to_csv(bucket_summary_path, index=False, encoding="utf-8-sig")
    snapshot_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    report_path.write_text(_render_close_review_report(review, bucket_summary, snapshot), encoding="utf-8")
    return StockTriggerConfirmationCloseReviewResult(
        tracking_path=resolved_tracking,
        output_dir=resolved_output,
        review_path=review_path,
        bucket_summary_path=bucket_summary_path,
        snapshot_path=snapshot_path,
        report_path=report_path,
        snapshot=snapshot,
        review=review,
        bucket_summary=bucket_summary,
    )


def _journal_rows_from_close_review(
    close_review: StockTriggerConfirmationCloseReviewResult,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    market_source_kind = close_review.snapshot.get("market_data_source_kind")
    market_source = close_review.snapshot.get("market_data_source")
    for raw in close_review.review.to_dict(orient="records"):
        code = _clean_code(raw.get("code"))
        trade_date = _text(raw.get("trade_date"))
        entry_date = _text(raw.get("entry_date"))
        if not code or not trade_date:
            continue
        event_id = f"{trade_date}:{code}:{entry_date or '-'}"
        rows.append(
            {
                "trade_date": trade_date,
                "code": code,
                "name": _text(raw.get("name")),
                "entry_date": entry_date,
                "entry_price": _number(raw.get("entry_price")),
                "close_price": _number(raw.get("close_price")),
                "high_price": _number(raw.get("high_price")),
                "low_price": _number(raw.get("low_price")),
                "planned_shadow_weight": _number(raw.get("planned_shadow_weight")),
                "stop_loss": _number(raw.get("stop_loss")),
                "close_return_pct": _number(raw.get("close_return_pct")),
                "portfolio_close_return_pct": _number(raw.get("portfolio_close_return_pct")),
                "result_status": _text(raw.get("close_review_status")),
                "shadow_review_decision": _text(raw.get("shadow_review_decision")),
                "trigger_pct": _number(raw.get("trigger_pct")),
                "trigger_pct_bucket": _text(raw.get("trigger_pct_bucket")),
                "trigger_score": _number(raw.get("trigger_score")),
                "trigger_score_level": _text(raw.get("trigger_score_level")),
                "market_data_source_kind": _text(market_source_kind),
                "market_data_source": _text(market_source),
                "journal_event_id": event_id,
                "broker_action": "none",
                "research_only": True,
            }
        )
    return pd.DataFrame(rows, columns=SHADOW_JOURNAL_COLUMNS)


def _read_shadow_journal(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=SHADOW_JOURNAL_COLUMNS)
    try:
        frame = pd.read_csv(path, dtype={"code": str}, encoding="utf-8-sig")
    except (OSError, UnicodeDecodeError, pd.errors.EmptyDataError, pd.errors.ParserError):
        return pd.DataFrame(columns=SHADOW_JOURNAL_COLUMNS)
    for column in SHADOW_JOURNAL_COLUMNS:
        if column not in frame.columns:
            frame[column] = None
    frame = frame[SHADOW_JOURNAL_COLUMNS].copy()
    if "code" in frame.columns:
        frame["code"] = frame["code"].map(_clean_code)
    return frame


def _write_shadow_journal(
    journal_path: Path,
    close_review: StockTriggerConfirmationCloseReviewResult,
) -> tuple[pd.DataFrame, int]:
    existing = _read_shadow_journal(journal_path)
    new_rows = _journal_rows_from_close_review(close_review)
    existing_keys = set(existing["journal_event_id"].dropna().astype(str)) if not existing.empty else set()
    new_keys = set(new_rows["journal_event_id"].dropna().astype(str)) if not new_rows.empty else set()
    appended_count = len(new_keys - existing_keys)
    if new_rows.empty:
        combined = existing.copy()
    elif existing.empty:
        combined = new_rows.copy()
    else:
        combined = pd.concat([existing, new_rows], ignore_index=True)
    if not combined.empty:
        combined = combined.drop_duplicates(subset=["journal_event_id"], keep="last")
        combined = combined.sort_values(["trade_date", "code", "entry_date"], kind="stable").reset_index(drop=True)
    combined = combined.reindex(columns=SHADOW_JOURNAL_COLUMNS)
    journal_path.parent.mkdir(parents=True, exist_ok=True)
    combined.to_csv(journal_path, index=False, encoding="utf-8-sig")
    return combined, appended_count


def _render_daily_flow_report(snapshot: dict[str, Any]) -> str:
    return f"""# 个股触发确认日流程
<!-- stock_trigger_confirmation_daily_flow -->

本流程只用于研究影子持仓闭环；不会生成券商订单，不会自动下单。ETF/基金候选按规则排除。

## 摘要

| 项目 | 值 |
| --- | ---: |
| 交易日 | `{snapshot.get("trade_date")}` |
| 模板应用行数 | {snapshot.get("applied_count")} |
| 影子计划候选 | {snapshot.get("shadow_candidate_count")} |
| ETF/基金排除数 | {snapshot.get("non_stock_excluded_count")} |
| 开盘检查通过 | {snapshot.get("open_check_passed_count")} |
| 影子跟踪数 | {snapshot.get("active_shadow_position_count")} |
| 收盘复盘数 | {snapshot.get("close_review_count")} |
| 触及止损数 | {snapshot.get("stop_loss_touched_count")} |
| Journal 新增行数 | {snapshot.get("journal_appended_count")} |
| Journal 总行数 | {snapshot.get("journal_row_count")} |
| 券商动作 | `{snapshot.get("broker_action")}` |

## 文件

- 决策应用：`{snapshot.get("decision_applied_path")}`
- 影子计划：`{snapshot.get("shadow_plan_path")}`
- 开盘检查：`{snapshot.get("open_check_path")}`
- 影子跟踪：`{snapshot.get("shadow_tracking_path")}`
- 收盘复盘：`{snapshot.get("close_review_path")}`
- 累计样本库：`{snapshot.get("journal_path")}`
"""


def run_stock_trigger_confirmation_daily_flow(
    *,
    template_path: str | Path,
    output_dir: str | Path,
    trade_date: str,
    journal_path: str | Path | None = None,
    market_data_dir: str | Path = "D:/codex/daily-market-data",
    ingest_project_dir: str | Path = "D:/codex/2026-06-15-exchange-data-ingest",
    max_single_position_weight: float = 0.20,
    max_open_gap_pct: float = 0.03,
) -> StockTriggerConfirmationDailyFlowResult:
    resolved_output = Path(output_dir)
    resolved_output.mkdir(parents=True, exist_ok=True)
    resolved_journal = Path(journal_path) if journal_path is not None else resolved_output / "stock_trigger_confirmation_shadow_journal.csv"

    decision_apply = apply_stock_trigger_confirmation_decision_template(
        template_path,
        output_dir=resolved_output,
    )
    shadow_plan = build_stock_trigger_confirmation_shadow_buy_plan(
        decision_apply.applied_path,
        template_path,
        output_dir=resolved_output,
        max_single_position_weight=max_single_position_weight,
        max_open_gap_pct=max_open_gap_pct,
    )
    open_check = build_stock_trigger_confirmation_next_open_check(
        shadow_plan.plan_path,
        output_dir=resolved_output,
        trade_date=trade_date,
        market_data_dir=market_data_dir,
        ingest_project_dir=ingest_project_dir,
    )
    shadow_tracking = build_stock_trigger_confirmation_shadow_tracking(
        open_check.check_path,
        output_dir=resolved_output,
    )
    close_review = build_stock_trigger_confirmation_close_review(
        shadow_tracking.tracking_path,
        output_dir=resolved_output,
        trade_date=trade_date,
        market_data_dir=market_data_dir,
        ingest_project_dir=ingest_project_dir,
    )
    journal, appended_count = _write_shadow_journal(resolved_journal, close_review)

    normalized_trade_date = pd.Timestamp(pd.to_datetime(trade_date)).strftime("%Y-%m-%d")
    snapshot = {
        "schema_version": "stock-trigger-confirmation-daily-flow-v1",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "trade_date": normalized_trade_date,
        "template_path": str(template_path),
        "output_dir": str(resolved_output),
        "decision_applied_path": str(decision_apply.applied_path),
        "shadow_plan_path": str(shadow_plan.plan_path),
        "open_check_path": str(open_check.check_path),
        "shadow_tracking_path": str(shadow_tracking.tracking_path),
        "close_review_path": str(close_review.review_path),
        "journal_path": str(resolved_journal),
        "applied_count": decision_apply.snapshot.get("applied_count", 0),
        "shadow_candidate_count": shadow_plan.snapshot.get("shadow_candidate_count", 0),
        "non_stock_excluded_count": shadow_plan.snapshot.get("non_stock_excluded_count", 0),
        "open_check_passed_count": open_check.snapshot.get("passed_count", 0),
        "active_shadow_position_count": shadow_tracking.snapshot.get("active_shadow_position_count", 0),
        "close_review_count": close_review.snapshot.get("review_count", 0),
        "stop_loss_touched_count": close_review.snapshot.get("stop_loss_touched_count", 0),
        "portfolio_close_return_pct": close_review.snapshot.get("portfolio_close_return_pct", 0.0),
        "journal_appended_count": int(appended_count),
        "journal_row_count": int(len(journal)),
        "market_data_source_kind": close_review.snapshot.get("market_data_source_kind"),
        "market_data_source": close_review.snapshot.get("market_data_source"),
        "broker_action": "none",
        "research_only": True,
    }
    snapshot_path = resolved_output / "stock_trigger_confirmation_daily_flow.json"
    report_path = resolved_output / "stock_trigger_confirmation_daily_flow.md"
    snapshot_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    report_path.write_text(_render_daily_flow_report(snapshot), encoding="utf-8")
    return StockTriggerConfirmationDailyFlowResult(
        output_dir=resolved_output,
        journal_path=resolved_journal,
        snapshot_path=snapshot_path,
        report_path=report_path,
        snapshot=snapshot,
        decision_apply_result=decision_apply,
        shadow_plan_result=shadow_plan,
        open_check_result=open_check,
        shadow_tracking_result=shadow_tracking,
        close_review_result=close_review,
        journal=journal,
    )


def _render_decision_apply_report(applied: pd.DataFrame, snapshot: dict[str, Any]) -> str:
    if applied.empty:
        rows = "| - | - | - | - | - | - |\n"
    else:
        lines = []
        for row in applied.to_dict(orient="records"):
            lines.append(
                "| `{code}` | {name} | `{raw}` | `{decision}` | `{gate}` | {weight:.2%} |".format(
                    code=_text(row.get("code")),
                    name=_text(row.get("name")),
                    raw=_text(row.get("manual_decision_raw")),
                    decision=_text(row.get("manual_decision_normalized")),
                    gate=_text(row.get("confirmation_gate")),
                    weight=_number(row.get("position_weight_normalized")),
                )
            )
        rows = "\n".join(lines) + "\n"
    return f"""# 个股触发确认人工决策应用
<!-- stock_trigger_confirmation_decision_apply -->

本报告只把人工确认表转为研究系统可读的审计结果；不会生成券商订单，不会自动下单。

## 摘要

| 项目 | 值 |
| --- | ---: |
| 模板路径 | `{snapshot.get("template_path")}` |
| 应用行数 | {snapshot.get("applied_count")} |
| 允许进入影子复核 | {snapshot.get("approved_shadow_count")} |
| 影子复核可用行数 | {snapshot.get("live_shadow_review_allowed_count")} |
| 警告行数 | {snapshot.get("warning_count")} |
| 无效行数 | {snapshot.get("invalid_count")} |
| 券商动作 | `{snapshot.get("broker_action")}` |

## 决策明细

| 代码 | 名称 | 原始决策 | 归一化决策 | 确认门 | 目标仓位 |
| --- | --- | --- | --- | --- | ---: |
{rows}
"""


def apply_stock_trigger_confirmation_decision_template(
    template_path: str | Path,
    *,
    output_dir: str | Path | None = None,
) -> StockTriggerConfirmationDecisionApplyResult:
    resolved_template = Path(template_path)
    resolved_output = Path(output_dir) if output_dir is not None else resolved_template.parent
    resolved_output.mkdir(parents=True, exist_ok=True)
    template = _load_stock_trigger_confirmation_decision_template(resolved_template)
    rows: list[dict[str, Any]] = []
    for raw in template.to_dict(orient="records"):
        code = _clean_code(raw.get("code"))
        if not code:
            continue
        decision_raw = _text(raw.get("manual_decision_to_fill"))
        decision = _normalise_manual_decision(decision_raw)
        pullback = _normalise_yes_no_unknown(raw.get("pullback_confirmed_to_fill"))
        risk_ok = _normalise_yes_no_unknown(raw.get("risk_ok_to_fill"))
        position_ok = _normalise_yes_no_unknown(raw.get("position_ok_to_fill"))
        stop_loss_ok = _normalise_yes_no_unknown(raw.get("stop_loss_ok_to_fill"))
        weight = _parse_position_weight(raw.get("position_weight_to_fill"))
        issues: list[str] = []
        warnings: list[str] = []
        if not decision:
            issues.append("invalid_manual_decision")
        for label, value in {
            "pullback_confirmed": pullback,
            "risk_ok": risk_ok,
            "position_ok": position_ok,
            "stop_loss_ok": stop_loss_ok,
        }.items():
            if not value:
                issues.append(f"invalid_{label}")
        if weight is None:
            issues.append("invalid_position_weight")
        elif weight < 0 or weight > 1:
            issues.append("position_weight_out_of_range")
        elif weight > 0.2:
            warnings.append("high_position_weight")

        live_shadow_allowed = False
        if issues:
            gate = "invalid"
        elif decision == "approve_shadow":
            if all(value == "yes" for value in [pullback, risk_ok, position_ok, stop_loss_ok]) and _number(weight) > 0:
                live_shadow_allowed = True
                gate = "passed_with_warnings" if warnings else "passed"
            else:
                gate = "blocked_confirmation_not_all_yes"
        elif decision == "watch":
            gate = "watch"
        elif decision == "reject":
            gate = "rejected"
        else:
            gate = "wait_for_retrigger"

        rows.append(
            {
                "code": code,
                "name": _text(raw.get("name")),
                "manual_decision_raw": decision_raw,
                "manual_decision_normalized": decision,
                "decision_action": _decision_action(decision),
                "pullback_confirmed": pullback,
                "risk_ok": risk_ok,
                "position_ok": position_ok,
                "stop_loss_ok": stop_loss_ok,
                "position_weight_normalized": weight if weight is not None else None,
                "manual_note": _text(raw.get("manual_note_to_fill")),
                "next_review_date": _text(raw.get("next_review_date_to_fill")),
                "reviewed_by": _text(raw.get("reviewed_by_to_fill")),
                "confirmation_gate": gate,
                "live_shadow_review_allowed": bool(live_shadow_allowed),
                "warnings": "|".join(warnings),
                "validation_issues": "|".join(issues),
                "broker_action": "none",
                "research_only": True,
            }
        )
    applied = pd.DataFrame(
        rows,
        columns=[
            "code",
            "name",
            "manual_decision_raw",
            "manual_decision_normalized",
            "decision_action",
            "pullback_confirmed",
            "risk_ok",
            "position_ok",
            "stop_loss_ok",
            "position_weight_normalized",
            "manual_note",
            "next_review_date",
            "reviewed_by",
            "confirmation_gate",
            "live_shadow_review_allowed",
            "warnings",
            "validation_issues",
            "broker_action",
            "research_only",
        ],
    )
    snapshot = {
        "schema_version": "stock-trigger-confirmation-decision-apply-v1",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "template_path": str(resolved_template),
        "applied_count": int(len(applied)),
        "approved_shadow_count": int((applied["manual_decision_normalized"] == "approve_shadow").sum()) if not applied.empty else 0,
        "live_shadow_review_allowed_count": int(applied["live_shadow_review_allowed"].sum()) if not applied.empty else 0,
        "warning_count": int((applied["warnings"].astype(str).str.strip() != "").sum()) if not applied.empty else 0,
        "invalid_count": int((applied["confirmation_gate"] == "invalid").sum()) if not applied.empty else 0,
        "broker_action": "none",
        "research_only": True,
    }
    applied_path = resolved_output / "stock_trigger_confirmation_decision_applied.csv"
    snapshot_path = resolved_output / "stock_trigger_confirmation_decision_applied.json"
    report_path = resolved_output / "stock_trigger_confirmation_decision_applied.md"
    applied.to_csv(applied_path, index=False, encoding="utf-8-sig")
    snapshot_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    report_path.write_text(_render_decision_apply_report(applied, snapshot), encoding="utf-8")
    return StockTriggerConfirmationDecisionApplyResult(
        template_path=resolved_template,
        output_dir=resolved_output,
        applied_path=applied_path,
        snapshot_path=snapshot_path,
        report_path=report_path,
        snapshot=snapshot,
        applied=applied,
    )


def _read_csv(path: str | Path | None) -> pd.DataFrame:
    if path is None:
        return pd.DataFrame()
    resolved = Path(path)
    if not resolved.exists():
        return pd.DataFrame()
    try:
        frame = pd.read_csv(resolved, dtype={"code": str}, encoding="utf-8-sig")
    except UnicodeDecodeError:
        frame = pd.read_csv(resolved, dtype={"code": str}, encoding="gbk")
    except (OSError, pd.errors.EmptyDataError, pd.errors.ParserError):
        return pd.DataFrame()
    if frame.empty:
        return frame
    if "code" in frame.columns:
        frame = frame.copy()
        frame["code"] = frame["code"].map(_clean_code)
        frame = frame[frame["code"] != ""].copy()
    return frame


def _records_by_code(frame: pd.DataFrame) -> dict[str, dict[str, Any]]:
    if frame.empty or "code" not in frame.columns:
        return {}
    out: dict[str, dict[str, Any]] = {}
    for raw in frame.to_dict(orient="records"):
        code = _clean_code(raw.get("code"))
        if code:
            out[code] = raw
    return out


def _manual_checklist(alignment: dict[str, Any], signal: dict[str, Any], target: dict[str, Any]) -> str:
    action = _text(alignment.get("action") or signal.get("action"))
    risk = _text(target.get("risk_filter_status"))
    parts = [
        "确认回踩不破或下一交易日仍满足触发结构",
        "核对组合目标仓位、单票上限和现金约束",
        "复核止损/压力/支撑位置是否清晰",
    ]
    if action:
        parts.append(f"触发动作要求：{action}")
    if risk:
        parts.append(f"当前风险状态：{risk}")
    return "；".join(parts)


def _confirmation_row(
    alignment: dict[str, Any],
    target: dict[str, Any],
    signal: dict[str, Any],
    review: dict[str, Any],
) -> dict[str, Any]:
    return {
        "code": _clean_code(alignment.get("code")),
        "name": _text(alignment.get("name") or target.get("name") or signal.get("name")),
        "primary_confirmation": "model_target_with_fresh_trigger",
        "pretrade_decision": "manual_review_only",
        "manual_review_checklist": _manual_checklist(alignment, signal, target),
        "portfolio_target_weight": _number(alignment.get("portfolio_target_weight") or target.get("portfolio_target_weight")),
        "trigger_run_time": _text(alignment.get("trigger_run_time") or signal.get("run_time")),
        "trigger_signal_type": _text(alignment.get("signal_type") or signal.get("signal_type")),
        "trigger_action": _text(alignment.get("action") or signal.get("action")),
        "trigger_score": _number(alignment.get("score") or signal.get("score")),
        "trigger_score_level": _text(alignment.get("score_level") or signal.get("score_level")),
        "trigger_pct": _number(alignment.get("pct") or signal.get("pct")),
        "trigger_last": _number(alignment.get("last") or signal.get("last")),
        "trigger_reason": _text(signal.get("reason") or signal.get("score_reason")),
        "support": _number(signal.get("support")),
        "pressure": _number(signal.get("pressure")),
        "stop_loss": _number(signal.get("stop_loss")),
        "source_strategy_profile": _text(target.get("source_strategy_profile")),
        "risk_filter_status": _text(target.get("risk_filter_status")),
        "execution_gate_action": _text(target.get("execution_gate_action"), "manual_review_only"),
        "target_action": _text(target.get("target_action")),
        "unrealized_return": _number(target.get("unrealized_return")),
        "market_cap_yi": _number(target.get("market_cap_yi")),
        "review_bucket": _text(review.get("review_bucket"), "not_in_review_queue"),
        "review_stage": _text(review.get("review_stage")),
        "review_reason": _text(review.get("review_reason")),
        "manual_review_state": _text(review.get("manual_review_state")),
        "broker_action": "none",
        "research_only": True,
    }


def _render_report(snapshot: dict[str, Any], table: pd.DataFrame) -> str:
    if table.empty:
        rows = "| - | - | - | - | - | - |\n"
        details = "- 暂无模型目标和 fresh 触发同时确认的个股。"
    else:
        lines = []
        detail_lines = []
        for row in table.to_dict(orient="records"):
            lines.append(
                "| `{code}` | {name} | {weight:.2f}% | {score:.1f} | `{decision}` | {checklist} |".format(
                    code=_text(row.get("code")),
                    name=_text(row.get("name")),
                    weight=_number(row.get("portfolio_target_weight")) * 100,
                    score=_number(row.get("trigger_score")),
                    decision=_text(row.get("pretrade_decision")),
                    checklist=_text(row.get("manual_review_checklist")),
                )
            )
            detail_lines.append(
                "- `{code}` {name}：目标仓位 {weight:.2f}%，触发分数 {score:.1f}，动作 `{action}`；只进入人工复核，不生成券商订单。".format(
                    code=_text(row.get("code")),
                    name=_text(row.get("name")),
                    weight=_number(row.get("portfolio_target_weight")) * 100,
                    score=_number(row.get("trigger_score")),
                    action=_text(row.get("trigger_action")),
                )
            )
        rows = "\n".join(lines) + "\n"
        details = "\n".join(detail_lines)
    return f"""# 个股目标-fresh触发确认复核
<!-- stock_trigger_confirmation -->

生成时间：`{snapshot.get("generated_at")}`

本文件只用于实盘前置人工复核，不连接券商、不自动下单、不构成投资建议。

## 摘要

| 项目 | 值 |
| --- | ---: |
| 截止日期 | {snapshot.get("as_of_date")} |
| 确认状态 | `{snapshot.get("confirmation_status")}` |
| 确认触发行数 | {snapshot.get("confirmed_target_trigger_count")} |
| 高分确认行数 | {snapshot.get("high_score_confirmed_count")} |
| 需人工复核 | {snapshot.get("manual_review_required_count")} |
| 券商动作 | `{snapshot.get("broker_action")}` |

## 复核表

| 代码 | 名称 | 目标仓位 | 触发分数 | 预交易决策 | 人工复核清单 |
| --- | --- | ---: | ---: | --- | --- |
{rows}
## 逐项说明

{details}

## 来源

- 对齐审计：`{snapshot.get("alignment_path")}`
- 个股目标：`{snapshot.get("stock_targets_path")}`
- 最新触发信号：`{snapshot.get("trigger_signal_path")}`
- 个股复核：`{snapshot.get("stock_target_review_path")}`
- 明细 CSV：`{snapshot.get("table_path")}`
- 决策模板 CSV：`{snapshot.get("decision_template_path")}`
- 决策模板 XLSX：`{snapshot.get("decision_template_xlsx_path")}`
"""


def build_stock_trigger_confirmation_decision_template(table: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not table.empty:
        ranked = table.copy()
        ranked["_trigger_score"] = pd.to_numeric(ranked["trigger_score"], errors="coerce").fillna(0.0)
        ranked = ranked.sort_values(["_trigger_score", "portfolio_target_weight"], ascending=[False, False])
        for rank, raw in enumerate(ranked.to_dict(orient="records"), start=1):
            rows.append(
                {
                    "confirmation_rank": rank,
                    "code": _clean_code(raw.get("code")),
                    "name": _text(raw.get("name")),
                    "manual_decision_to_fill": "",
                    "pullback_confirmed_to_fill": "",
                    "risk_ok_to_fill": "",
                    "position_ok_to_fill": "",
                    "stop_loss_ok_to_fill": "",
                    "position_weight_to_fill": "",
                    "manual_note_to_fill": "",
                    "next_review_date_to_fill": "",
                    "reviewed_by_to_fill": "",
                    "allowed_manual_decisions": "approve_shadow|watch|reject|needs_retrigger",
                    "allowed_yes_no_values": "yes|no|unknown",
                    "portfolio_target_weight": _number(raw.get("portfolio_target_weight")),
                    "trigger_score": _number(raw.get("trigger_score")),
                    "trigger_score_level": _text(raw.get("trigger_score_level")),
                    "trigger_action": _text(raw.get("trigger_action")),
                    "trigger_pct": _number(raw.get("trigger_pct")),
                    "trigger_last": _number(raw.get("trigger_last")),
                    "support": _number(raw.get("support")),
                    "pressure": _number(raw.get("pressure")),
                    "stop_loss": _number(raw.get("stop_loss")),
                    "manual_review_checklist": _text(raw.get("manual_review_checklist")),
                    "pretrade_decision": _text(raw.get("pretrade_decision")),
                    "review_bucket": _text(raw.get("review_bucket")),
                    "review_reason": _text(raw.get("review_reason")),
                    "broker_action": "none",
                    "research_only": True,
                }
            )
    template = pd.DataFrame(rows, columns=DECISION_TEMPLATE_COLUMNS)
    payload = {
        "decision_template_count": int(len(template)),
        "decision_template_blank_decision_count": int(
            (template["manual_decision_to_fill"].astype(str).str.strip() == "").sum()
        )
        if not template.empty
        else 0,
        "decision_template_user_fill_columns": DECISION_TEMPLATE_USER_FILL_COLUMNS,
        "allowed_manual_decisions": "approve_shadow|watch|reject|needs_retrigger",
        "allowed_yes_no_values": "yes|no|unknown",
        "broker_action": "none",
        "research_only": True,
    }
    return template, payload


def write_stock_trigger_confirmation_decision_template_xlsx(
    template: pd.DataFrame,
    payload: dict[str, Any],
    output_path: str | Path,
    *,
    header_labels: Mapping[str, str] | None = None,
    sheet_title: str = "decision_template",
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as error:  # pragma: no cover
        raise RuntimeError("openpyxl is required to write the highlighted confirmation decision template.") from error

    resolved = Path(output_path)
    resolved.parent.mkdir(parents=True, exist_ok=True)
    columns = list(template.columns) if not template.empty else list(DECISION_TEMPLATE_COLUMNS)
    display_columns = [header_labels.get(column, column) for column in columns] if header_labels else columns
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    sheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="FF1F4E78")
    fill_header_fill = PatternFill("solid", fgColor="FFFFC000")
    required_fill = PatternFill("solid", fgColor="FFFFE699")
    optional_fill = PatternFill("solid", fgColor="FFFFF2CC")
    neutral_fill = PatternFill("solid", fgColor="FFFFFFFF")
    header_font = Font(color="FFFFFFFF", bold=True)
    thin = Side(style="thin", color="FFD9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.append(display_columns)
    for record in template.to_dict(orient="records"):
        sheet.append([_xlsx_scalar(record.get(column)) for column in columns])

    max_row = max(sheet.max_row, 2)
    max_col = len(columns)
    sheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    fill_column_indexes = {
        column: columns.index(column) + 1
        for column in DECISION_TEMPLATE_USER_FILL_COLUMNS
        if column in columns
    }

    for col_idx, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = fill_header_fill if column in fill_column_indexes else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        if column in fill_column_indexes:
            cell.comment = Comment("需要你填写。黄色列会被后续人工复核流程读取。", "quant_etf_lab")
        width = 18
        if column in {"manual_review_checklist", "manual_note_to_fill", "review_reason"}:
            width = 42
        elif column in fill_column_indexes:
            width = 24
        elif column in {"code", "name"}:
            width = 14
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx in range(2, max_row + 1):
        for col_idx, column in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=column in fill_column_indexes or column in {"manual_review_checklist", "review_reason"})
            if column == "manual_decision_to_fill":
                cell.fill = required_fill
            elif column in fill_column_indexes:
                cell.fill = optional_fill
            else:
                cell.fill = neutral_fill
            if column in {"portfolio_target_weight", "position_weight_to_fill"}:
                cell.number_format = "0.00%"
            elif column in {"trigger_last", "support", "pressure", "stop_loss"}:
                cell.number_format = "0.00"
        sheet.row_dimensions[row_idx].height = 34

    if "manual_decision_to_fill" in fill_column_indexes:
        decision_col = get_column_letter(fill_column_indexes["manual_decision_to_fill"])
        validation = DataValidation(
            type="list",
            formula1='"approve_shadow,watch,reject,needs_retrigger"',
            allow_blank=True,
        )
        validation.error = "Use one of: approve_shadow, watch, reject, needs_retrigger."
        validation.errorTitle = "Invalid manual decision"
        validation.prompt = "Choose a manual decision."
        validation.promptTitle = "Manual decision"
        sheet.add_data_validation(validation)
        validation.add(f"{decision_col}2:{decision_col}{max_row}")

    yes_no_columns = {
        "pullback_confirmed_to_fill",
        "risk_ok_to_fill",
        "position_ok_to_fill",
        "stop_loss_ok_to_fill",
    }
    for column in yes_no_columns:
        if column in fill_column_indexes:
            col = get_column_letter(fill_column_indexes[column])
            validation = DataValidation(type="list", formula1='"yes,no,unknown"', allow_blank=True)
            validation.error = "Use one of: yes, no, unknown."
            validation.errorTitle = "Invalid confirmation"
            sheet.add_data_validation(validation)
            validation.add(f"{col}2:{col}{max_row}")

    sheet["A1"].comment = Comment(
        f"Fill highlighted columns only. Rows: {payload.get('decision_template_count', 0)}. No broker order is generated.",
        "quant_etf_lab",
    )
    workbook.save(resolved)
    return resolved


def _render_decision_template_report(template: pd.DataFrame, payload: dict[str, Any], output_dir: Path) -> str:
    return f"""# 个股目标-fresh触发人工决策模板
<!-- stock_trigger_confirmation_decision_template -->

此模板用于填写确认触发样本的人工复核结果；不会自动下单。

## 填写说明

- 黄色列为需要人工填写的字段。
- `manual_decision_to_fill` 可填：`approve_shadow` / `watch` / `reject` / `needs_retrigger`。
- 回踩、风控、仓位、止损确认列可填：`yes` / `no` / `unknown`。
- `approve_shadow` 只表示允许进入后续实盘影子复核，不代表券商下单。

## 摘要

| 项目 | 值 |
| --- | ---: |
| 模板行数 | {payload.get("decision_template_count")} |
| 待填写决策数 | {payload.get("decision_template_blank_decision_count")} |
| 券商动作 | `{payload.get("broker_action")}` |

## 文件

- 决策模板 CSV：`{output_dir / "stock_trigger_confirmation_decision_template.csv"}`
- 高亮填写工作簿：`{output_dir / "stock_trigger_confirmation_decision_template.xlsx"}`
- 中文表头填写工作簿：`{payload.get("decision_template_zh_xlsx_path", output_dir / "stock_trigger_confirmation_decision_template_zh.xlsx")}`
- 决策模板 JSON：`{output_dir / "stock_trigger_confirmation_decision_template.json"}`
"""


def run_stock_trigger_confirmation(
    *,
    alignment_path: str | Path,
    stock_targets_path: str | Path,
    trigger_signal_path: str | Path,
    output_dir: str | Path,
    stock_target_review_path: str | Path | None = None,
    as_of_date: str | pd.Timestamp,
    high_score_threshold: float = 90.0,
) -> StockTriggerConfirmationResult:
    resolved_output = Path(output_dir)
    resolved_output.mkdir(parents=True, exist_ok=True)

    alignment = _read_csv(alignment_path)
    targets = _read_csv(stock_targets_path)
    signals = _read_csv(trigger_signal_path)
    review = _read_csv(stock_target_review_path)
    target_by_code = _records_by_code(targets)
    signal_by_code = _records_by_code(signals)
    review_by_code = _records_by_code(review)

    rows: list[dict[str, Any]] = []
    if not alignment.empty and "alignment_bucket" in alignment.columns:
        confirmed = alignment[alignment["alignment_bucket"].astype(str).str.strip() == "target_with_fresh_trigger"].copy()
        for raw in confirmed.to_dict(orient="records"):
            code = _clean_code(raw.get("code"))
            rows.append(
                _confirmation_row(
                    raw,
                    target_by_code.get(code, {}),
                    signal_by_code.get(code, {}),
                    review_by_code.get(code, {}),
                )
            )

    table = pd.DataFrame(rows, columns=CONFIRMATION_COLUMNS)
    confirmed_count = int(len(table))
    high_score_count = int((pd.to_numeric(table["trigger_score"], errors="coerce").fillna(0.0) >= high_score_threshold).sum()) if not table.empty else 0
    status = "manual_review_required" if confirmed_count > 0 else "no_confirmed_target_trigger"

    table_path = resolved_output / "stock_trigger_confirmation.csv"
    snapshot_path = resolved_output / "stock_trigger_confirmation_snapshot.json"
    report_path = resolved_output / "stock_trigger_confirmation.md"
    decision_template_path = resolved_output / "stock_trigger_confirmation_decision_template.csv"
    decision_template_json_path = resolved_output / "stock_trigger_confirmation_decision_template.json"
    decision_template_report_path = resolved_output / "stock_trigger_confirmation_decision_template.md"
    decision_template_xlsx_path = resolved_output / "stock_trigger_confirmation_decision_template.xlsx"
    decision_template_zh_xlsx_path = resolved_output / "stock_trigger_confirmation_decision_template_zh.xlsx"
    decision_template, decision_template_payload = build_stock_trigger_confirmation_decision_template(table)
    decision_template_payload["decision_template_zh_xlsx_path"] = str(decision_template_zh_xlsx_path)
    snapshot: dict[str, Any] = {
        "schema_version": "stock-trigger-confirmation-v1",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "as_of_date": pd.Timestamp(pd.to_datetime(as_of_date)).strftime("%Y-%m-%d"),
        "confirmation_status": status,
        "confirmed_target_trigger_count": confirmed_count,
        "manual_review_required_count": confirmed_count,
        "high_score_confirmed_count": high_score_count,
        "high_score_threshold": float(high_score_threshold),
        "alignment_path": str(alignment_path),
        "stock_targets_path": str(stock_targets_path),
        "trigger_signal_path": str(trigger_signal_path),
        "stock_target_review_path": str(stock_target_review_path) if stock_target_review_path is not None else None,
        "table_path": str(table_path),
        "snapshot_path": str(snapshot_path),
        "report_path": str(report_path),
        "decision_template_path": str(decision_template_path),
        "decision_template_json_path": str(decision_template_json_path),
        "decision_template_report_path": str(decision_template_report_path),
        "decision_template_xlsx_path": str(decision_template_xlsx_path),
        "decision_template_zh_xlsx_path": str(decision_template_zh_xlsx_path),
        **decision_template_payload,
        "research_only": True,
        "broker_action": "none",
    }
    table.to_csv(table_path, index=False, encoding="utf-8-sig")
    decision_template.to_csv(decision_template_path, index=False, encoding="utf-8-sig")
    decision_template_json_path.write_text(json.dumps(decision_template_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    decision_template_report_path.write_text(
        _render_decision_template_report(decision_template, decision_template_payload, resolved_output),
        encoding="utf-8",
    )
    write_stock_trigger_confirmation_decision_template_xlsx(
        decision_template,
        decision_template_payload,
        decision_template_xlsx_path,
    )
    write_stock_trigger_confirmation_decision_template_xlsx(
        decision_template,
        decision_template_payload,
        decision_template_zh_xlsx_path,
        header_labels=DECISION_TEMPLATE_CHINESE_HEADERS,
        sheet_title="中文填写模板",
    )
    snapshot_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    report_path.write_text(_render_report(snapshot, table), encoding="utf-8")
    return StockTriggerConfirmationResult(
        output_dir=resolved_output,
        table_path=table_path,
        snapshot_path=snapshot_path,
        report_path=report_path,
        decision_template_path=decision_template_path,
        decision_template_json_path=decision_template_json_path,
        decision_template_report_path=decision_template_report_path,
        decision_template_xlsx_path=decision_template_xlsx_path,
        decision_template_zh_xlsx_path=decision_template_zh_xlsx_path,
        snapshot=snapshot,
        table=table,
        decision_template=decision_template,
    )
