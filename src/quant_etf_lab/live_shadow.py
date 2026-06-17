"""Live shadow-account reconciliation for manual pre-trade review."""

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from pandas.errors import EmptyDataError


DEFAULT_OUTPUT_DIR = Path("outputs/research/live_shadow_latest")
DEFAULT_PRE_FLIGHT_OUTPUT_DIR = Path("outputs/research/live_shadow_preflight_latest")
DEFAULT_IMPORT_TEMPLATE_OUTPUT_DIR = Path("outputs/research/live_shadow_import_latest")
DEFAULT_REVIEW_OUTPUT_DIR = Path("outputs/research/live_shadow_review_latest")
DEFAULT_REVIEW_DECISION_OUTPUT_DIR = Path("outputs/research/live_shadow_review_decisions_latest")
DEFAULT_TARGETS_FILE = Path("outputs/research/paper_account_latest/stock_targets.csv")
TARGET_COLUMNS = [
    "code",
    "name",
    "raw_target_weight",
    "target_value_hint",
    "target_price",
    "layer",
    "target_action",
    "target_explanation",
]
LIVE_SHADOW_HOLDINGS_TEMPLATE_COLUMNS = [
    "as_of_date",
    "code",
    "name",
    "quantity",
    "available_quantity",
    "current_price",
    "market_value",
    "source_account",
    "manual_note",
    "target_weight_reference",
    "target_value_hint_reference",
    "target_price_reference",
    "layer_reference",
    "target_action_reference",
    "broker_action",
    "research_only",
]
LIVE_SHADOW_CASH_TEMPLATE_COLUMNS = [
    "as_of_date",
    "cash",
    "source_account",
    "manual_note",
    "broker_action",
    "research_only",
]
LIVE_SHADOW_HOLDINGS_REQUIRED_FILL_COLUMNS = {"code", "quantity", "current_price"}
LIVE_SHADOW_HOLDINGS_OPTIONAL_FILL_COLUMNS = {
    "available_quantity",
    "market_value",
    "source_account",
    "manual_note",
}
LIVE_SHADOW_CASH_REQUIRED_FILL_COLUMNS = {"cash"}
LIVE_SHADOW_CASH_OPTIONAL_FILL_COLUMNS = {"source_account", "manual_note"}
LIVE_SHADOW_REVIEW_QUEUE_COLUMNS = [
    "as_of_date",
    "review_rank",
    "code",
    "name",
    "side",
    "action",
    "order_quantity",
    "price",
    "estimated_amount",
    "current_quantity",
    "target_quantity",
    "current_value",
    "target_value",
    "delta_value",
    "layer",
    "portfolio_target_weight",
    "market_cap_yi",
    "tracking_rule_status",
    "tracking_excluded",
    "tracking_exclusion_reason",
    "review_gate",
    "review_status_to_fill",
    "manual_note_to_fill",
    "allowed_review_statuses",
    "broker_action",
    "research_only",
]
LIVE_SHADOW_REVIEW_REQUIRED_FILL_COLUMNS = {"review_status_to_fill"}
LIVE_SHADOW_REVIEW_OPTIONAL_FILL_COLUMNS = {"manual_note_to_fill"}
LIVE_SHADOW_REVIEW_ALLOWED_STATUSES = {
    "reviewed_for_manual_consideration",
    "watch_only",
    "skip",
    "needs_data",
}
LIVE_SHADOW_REVIEW_DECISION_COLUMNS = [
    "as_of_date",
    "review_rank",
    "code",
    "name",
    "side",
    "action",
    "order_quantity",
    "price",
    "estimated_amount",
    "review_gate",
    "review_status",
    "review_decision",
    "manual_note",
    "tracking_exclusion_reason",
    "broker_action",
    "research_only",
]
LIVE_SHADOW_COLUMN_LABELS = {
    "as_of_date": "日期",
    "code": "股票代码",
    "name": "股票名称",
    "quantity": "持仓数量",
    "available_quantity": "可卖数量",
    "current_price": "当前价格",
    "market_value": "市值",
    "source_account": "账户来源",
    "manual_note": "备注",
    "target_weight_reference": "目标权重参考",
    "target_value_hint_reference": "目标市值参考",
    "target_price_reference": "目标价格参考",
    "layer_reference": "层级参考",
    "target_action_reference": "目标动作参考",
    "cash": "可用现金",
    "review_rank": "复核序号",
    "side": "方向",
    "action": "差异动作",
    "order_quantity": "差异数量",
    "price": "价格",
    "estimated_amount": "估算金额",
    "current_quantity": "当前数量",
    "target_quantity": "目标数量",
    "current_value": "当前市值",
    "target_value": "目标市值",
    "delta_value": "市值差异",
    "layer": "组合层级",
    "portfolio_target_weight": "组合目标权重",
    "market_cap_yi": "总市值(亿元)",
    "tracking_rule_status": "跟踪规则状态",
    "tracking_excluded": "是否暂停跟踪",
    "tracking_exclusion_reason": "暂停跟踪原因",
    "review_gate": "复核门禁",
    "review_status_to_fill": "复核状态（请填写）",
    "manual_note_to_fill": "备注（请填写）",
    "allowed_review_statuses": "允许复核状态",
    "broker_action": "券商动作",
    "research_only": "仅研究",
}
LIVE_SHADOW_LABEL_COLUMNS = {value: key for key, value in LIVE_SHADOW_COLUMN_LABELS.items()}


@dataclass(frozen=True)
class LiveShadowPreflightResult:
    output_dir: Path
    snapshot_path: Path
    report_path: Path
    holdings_path: Path
    targets_path: Path
    prices_path: Path | None
    snapshot: dict[str, Any]


@dataclass(frozen=True)
class LiveShadowResult:
    output_dir: Path
    snapshot_path: Path
    report_path: Path
    orders_path: Path
    reconcile_path: Path
    holdings_path: Path
    targets_path: Path
    prices_path: Path | None
    snapshot: dict[str, Any]
    orders: pd.DataFrame
    reconcile: pd.DataFrame


@dataclass(frozen=True)
class LiveShadowImportTemplateResult:
    output_dir: Path
    holdings_template_path: Path
    cash_template_path: Path
    workbook_path: Path
    snapshot_path: Path
    report_path: Path
    snapshot: dict[str, Any]
    holdings_template: pd.DataFrame
    cash_template: pd.DataFrame


@dataclass(frozen=True)
class LiveShadowReviewQueueResult:
    output_dir: Path
    review_queue_path: Path
    workbook_path: Path
    snapshot_path: Path
    report_path: Path
    snapshot: dict[str, Any]
    review_queue: pd.DataFrame


@dataclass(frozen=True)
class LiveShadowReviewDecisionResult:
    output_dir: Path
    decisions_path: Path
    snapshot_path: Path
    report_path: Path
    snapshot: dict[str, Any]
    decisions: pd.DataFrame


def _resolve(project_root: Path, path: str | Path | None, default: Path | None = None) -> Path:
    if path is None:
        if default is None:
            raise ValueError("path is required")
        raw = default
    else:
        raw = Path(path)
    return raw if raw.is_absolute() else project_root / raw


def _parse_date(value: str | date | None) -> str:
    if value is None:
        return datetime.now().date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if len(text) == 8 and text.isdigit():
        parsed = pd.to_datetime(text, format="%Y%m%d", errors="coerce")
    else:
        parsed = pd.to_datetime(text, errors="coerce")
    if pd.isna(parsed):
        raise ValueError(f"Invalid as_of_date: {value}")
    return pd.Timestamp(parsed).date().isoformat()


def _clean_code(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits.zfill(6) if digits else text


def _as_float(value: Any, default: float = 0.0) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    if not pd.notna(number):
        return default
    return number


def _first_number(row: pd.Series, columns: tuple[str, ...], default: float = 0.0) -> float:
    for column in columns:
        if column in row:
            number = _as_float(row.get(column), default=float("nan"))
            if pd.notna(number):
                return number
    return default


def _first_text(row: pd.Series, columns: tuple[str, ...], default: str = "") -> str:
    for column in columns:
        if column in row:
            value = row.get(column)
            if value is not None and not pd.isna(value):
                text = str(value).strip()
                if text:
                    return text
    return default


def _read_csv(path: Path, required: tuple[str, ...], sheet_name: str | int | None = None) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(path)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        try:
            frame = pd.read_excel(path, sheet_name=sheet_name or 0, dtype={"code": str})
        except ValueError:
            frame = pd.read_excel(path, sheet_name=0, dtype={"code": str})
    else:
        frame = pd.read_csv(path, dtype={"code": str})
    frame = frame.rename(columns={column: LIVE_SHADOW_LABEL_COLUMNS.get(str(column).strip(), column) for column in frame.columns})
    missing = [column for column in required if column not in frame.columns]
    if missing:
        raise ValueError(f"{path} is missing required columns: {', '.join(missing)}")
    return frame


def _normalize_holdings(path: Path) -> pd.DataFrame:
    raw = _read_csv(path, ("code", "quantity"), sheet_name="holdings_to_fill")
    rows: list[dict[str, Any]] = []
    for row in raw.to_dict(orient="records"):
        series = pd.Series(row)
        code = _clean_code(series.get("code"))
        if not code:
            continue
        quantity = max(_as_float(series.get("quantity")), 0.0)
        available_quantity = min(
            max(_first_number(series, ("available_quantity", "available", "sellable_quantity"), quantity), 0.0),
            quantity,
        )
        price = _first_number(series, ("current_price", "price", "last_price", "close_price", "close"), default=float("nan"))
        market_value = _first_number(series, ("market_value", "current_value"), default=float("nan"))
        if quantity > 0 and pd.notna(price) and price > 0:
            market_value = quantity * price
        if not pd.notna(price) and quantity > 0 and pd.notna(market_value):
            price = market_value / quantity
        rows.append(
            {
                "code": code,
                "name": _first_text(series, ("name", "security_name"), code),
                "quantity": quantity,
                "available_quantity": available_quantity,
                "current_price": price if pd.notna(price) else 0.0,
                "current_value": market_value if pd.notna(market_value) else 0.0,
            }
        )
    frame = pd.DataFrame(rows)
    if frame.empty:
        return pd.DataFrame(columns=["code", "name", "quantity", "available_quantity", "current_price", "current_value"])
    grouped = (
        frame.groupby("code", as_index=False)
        .agg(
            {
                "name": "first",
                "quantity": "sum",
                "available_quantity": "sum",
                "current_value": "sum",
                "current_price": "last",
            }
        )
        .reset_index(drop=True)
    )
    grouped.loc[grouped["quantity"] > 0, "current_price"] = (
        grouped.loc[grouped["quantity"] > 0, "current_value"] / grouped.loc[grouped["quantity"] > 0, "quantity"]
    ).fillna(grouped.loc[grouped["quantity"] > 0, "current_price"])
    return grouped


def _empty_targets() -> pd.DataFrame:
    return pd.DataFrame(columns=TARGET_COLUMNS)


def _normalize_targets(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(path)
    try:
        raw = pd.read_csv(path, dtype={"code": str})
    except EmptyDataError:
        return _empty_targets()
    if "code" not in raw.columns:
        if raw.empty:
            return _empty_targets()
        raise ValueError(f"{path} is missing required columns: code")
    rows: list[dict[str, Any]] = []
    for row in raw.to_dict(orient="records"):
        series = pd.Series(row)
        weight = _first_number(series, ("portfolio_target_weight", "target_weight", "weight"), 0.0)
        target_value = _first_number(series, ("portfolio_target_value", "target_value"), 0.0)
        price = _first_number(series, ("close_price", "current_price", "price", "last_price", "close"), 0.0)
        rows.append(
            {
                "code": _clean_code(series.get("code")),
                "name": _first_text(series, ("name", "security_name"), _clean_code(series.get("code"))),
                "raw_target_weight": max(weight, 0.0),
                "target_value_hint": max(target_value, 0.0),
                "target_price": max(price, 0.0),
                "layer": _first_text(series, ("layer",), ""),
                "target_action": _first_text(series, ("target_action",), ""),
                "target_explanation": _first_text(series, ("target_explanation",), ""),
            }
        )
    frame = pd.DataFrame(rows)
    if frame.empty:
        return _empty_targets()
    return (
        frame.groupby("code", as_index=False)
        .agg(
            {
                "name": "first",
                "raw_target_weight": "sum",
                "target_value_hint": "sum",
                "target_price": "max",
                "layer": lambda values: ",".join(sorted({str(value) for value in values if str(value)})),
                "target_action": "first",
                "target_explanation": "first",
            }
        )
        .reset_index(drop=True)
    )


def _normalize_prices(path: Path | None) -> dict[str, float]:
    if path is None:
        return {}
    raw = _read_csv(path, ("code",), sheet_name="prices")
    prices: dict[str, float] = {}
    for row in raw.to_dict(orient="records"):
        series = pd.Series(row)
        code = _clean_code(series.get("code"))
        if not code:
            continue
        price = _first_number(series, ("price", "current_price", "last_price", "close", "close_price"), 0.0)
        if price > 0:
            prices[code] = price
    return prices


def _round_down_to_lot(quantity: float, lot_size: int) -> int:
    if quantity <= 0:
        return 0
    if lot_size <= 1:
        return int(quantity)
    return int(quantity // lot_size * lot_size)


def _pct(value: Any) -> str:
    return f"{_as_float(value) * 100:.2f}%"


def _num(value: Any, digits: int = 2) -> str:
    return f"{_as_float(value):.{digits}f}"


def _json_records(frame: pd.DataFrame) -> list[dict[str, Any]]:
    records = frame.where(pd.notna(frame), None).to_dict(orient="records")
    return records


def _render_report(snapshot: dict[str, Any], orders: pd.DataFrame) -> str:
    if orders.empty:
        order_lines = "- No non-zero shadow orders."
    else:
        lines = ["| 股票代码 | 股票名称 | 方向 | 数量 | 价格 | 金额 | 备注 |", "| --- | --- | ---: | ---: | ---: | ---: | --- |"]
        for row in orders.head(30).to_dict(orient="records"):
            lines.append(
                "| `{code}` | {name} | `{side}` | {qty} | {price} | {amount} | {note} |".format(
                    code=str(row.get("code", "")),
                    name=str(row.get("name", "")),
                    side=str(row.get("side", "")),
                    qty=_num(row.get("order_quantity"), 0),
                    price=_num(row.get("price"), 3),
                    amount=_num(row.get("estimated_amount"), 2),
                    note=str(row.get("notes", "")),
                )
            )
        order_lines = "\n".join(lines)

    warning_lines = "\n".join(f"- {item}" for item in snapshot.get("warnings", [])) or "- No shadow-account warnings."
    return f"""# Live Shadow Account Plan

Generated at: `{snapshot.get("generated_at")}`

This is a manual-review shadow plan only. It does not connect to brokers, place orders, cancel orders, or provide investment advice.

## Status

| 项目 | 值 |
| --- | ---: |
| 计划状态 | `{snapshot.get("trade_plan_status")}` |
| 券商动作 | `{snapshot.get("broker_action")}` |
| 仅研究 | `{snapshot.get("research_only")}` |
| 日期 | {snapshot.get("as_of_date")} |
| 当前权益 | {_num(snapshot.get("current_equity"), 2)} |
| 可用现金 | {_num(snapshot.get("cash"), 2)} |
| 目标总仓位 | {_pct(snapshot.get("target_gross_weight"))} |
| 差异笔数 | {snapshot.get("order_count")} |
| 买入 / 卖出笔数 | {snapshot.get("buy_count")} / {snapshot.get("sell_count")} |
| 计划买入金额 | {_num(snapshot.get("planned_buy_amount"), 2)} |
| 计划卖出金额 | {_num(snapshot.get("planned_sell_amount"), 2)} |
| 影子调整后估算现金 | {_num(snapshot.get("estimated_cash_after_orders"), 2)} |

## Manual Shadow Orders

{order_lines}

## Warnings

{warning_lines}

## Files

- Orders CSV: `{snapshot.get("orders_path")}`
- Reconcile CSV: `{snapshot.get("reconcile_path")}`
- Snapshot JSON: `{snapshot.get("snapshot_path")}`
"""


def _build_shadow_plan(
    holdings: pd.DataFrame,
    targets: pd.DataFrame,
    prices: dict[str, float],
    cash: float,
    as_of: str,
    lot_size: int,
    min_trade_value: float,
    max_position_weight: float,
    max_gross_exposure: float,
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any], list[str]]:
    if lot_size <= 0:
        raise ValueError("lot_size must be positive.")
    if cash < 0:
        raise ValueError("cash must be non-negative.")
    if max_position_weight <= 0 or max_gross_exposure <= 0:
        raise ValueError("max_position_weight and max_gross_exposure must be positive.")
    if min_trade_value < 0:
        raise ValueError("min_trade_value must be non-negative.")

    holdings_by_code = holdings.set_index("code").to_dict(orient="index") if not holdings.empty else {}
    targets_by_code = targets.set_index("code").to_dict(orient="index") if not targets.empty else {}
    current_equity = float(cash) + sum(_as_float(item.get("current_value")) for item in holdings_by_code.values())
    if current_equity <= 0:
        raise ValueError("cash plus holding market value must be positive.")

    raw_weights: dict[str, float] = {}
    warnings: list[str] = []
    for code, target in targets_by_code.items():
        raw_weight = _as_float(target.get("raw_target_weight"))
        if raw_weight <= 0 and _as_float(target.get("target_value_hint")) > 0:
            raw_weight = _as_float(target.get("target_value_hint")) / current_equity
        capped = min(max(raw_weight, 0.0), max_position_weight)
        if raw_weight > max_position_weight:
            warnings.append(f"{code} target weight capped from {_pct(raw_weight)} to {_pct(max_position_weight)}.")
        raw_weights[code] = capped
    gross_weight_before_scale = sum(raw_weights.values())
    scale = min(1.0, max_gross_exposure / gross_weight_before_scale) if gross_weight_before_scale > 0 else 1.0
    if scale < 1.0:
        warnings.append(
            f"Target gross weight scaled from {_pct(gross_weight_before_scale)} to {_pct(max_gross_exposure)}."
        )

    codes = sorted(set(holdings_by_code) | set(targets_by_code))
    rows: list[dict[str, Any]] = []
    for code in codes:
        holding = holdings_by_code.get(code, {})
        target = targets_by_code.get(code, {})
        quantity = _as_float(holding.get("quantity"))
        available_quantity = _as_float(holding.get("available_quantity"), quantity)
        current_value = _as_float(holding.get("current_value"))
        price = (
            prices.get(code)
            or _as_float(holding.get("current_price"))
            or _as_float(target.get("target_price"))
        )
        target_weight = raw_weights.get(code, 0.0) * scale
        target_value = target_weight * current_equity
        target_quantity = _round_down_to_lot(target_value / price, lot_size) if price > 0 else 0
        side = "HOLD"
        action = "hold"
        order_quantity = 0
        block_reason = ""
        if price <= 0 and abs(target_value - current_value) > 1e-8:
            action = "blocked_missing_price"
            block_reason = "missing_price"
        elif target_quantity > quantity:
            order_quantity = _round_down_to_lot(target_quantity - quantity, lot_size)
            side = "BUY" if order_quantity > 0 else "HOLD"
            action = "buy_to_target" if order_quantity > 0 else "hold_rounding"
        elif target_quantity < quantity:
            desired_sell = max(quantity - target_quantity, 0.0)
            order_quantity = int(min(desired_sell, available_quantity))
            side = "SELL" if order_quantity > 0 else "HOLD"
            action = "sell_to_target" if order_quantity > 0 else "hold_unavailable"
            if desired_sell > available_quantity:
                block_reason = "available_quantity_limits_sell"
        estimated_amount = order_quantity * price if price > 0 else 0.0
        if order_quantity > 0 and estimated_amount < min_trade_value:
            side = "HOLD"
            action = "hold_small_delta"
            order_quantity = 0
            estimated_amount = 0.0
        name = str(target.get("name") or holding.get("name") or code)
        current_weight = current_value / current_equity if current_equity > 0 else 0.0
        rows.append(
            {
                "as_of_date": as_of,
                "code": code,
                "name": name,
                "side": side,
                "action": action,
                "current_quantity": quantity,
                "target_quantity": target_quantity,
                "order_quantity": order_quantity,
                "available_quantity": available_quantity,
                "price": price,
                "current_value": current_value,
                "target_value": target_value,
                "delta_value": target_value - current_value,
                "estimated_amount": estimated_amount,
                "current_weight": current_weight,
                "target_weight": target_weight,
                "weight_gap": target_weight - current_weight,
                "manual_review_required": True,
                "block_reason": block_reason,
                "notes": "manual_review_only; no broker connection or automatic order",
            }
        )

    reconcile = pd.DataFrame(rows)
    if not reconcile.empty:
        reconcile = reconcile.sort_values(["side", "code"], ascending=[False, True]).reset_index(drop=True)
    orders = reconcile[reconcile["order_quantity"] > 0].copy() if not reconcile.empty else pd.DataFrame()
    planned_buy = float(orders.loc[orders["side"] == "BUY", "estimated_amount"].sum()) if not orders.empty else 0.0
    planned_sell = float(orders.loc[orders["side"] == "SELL", "estimated_amount"].sum()) if not orders.empty else 0.0
    snapshot = {
        "as_of_date": as_of,
        "current_equity": current_equity,
        "holding_count": int(len(holdings)),
        "target_count": int(len(targets)),
        "reconcile_count": int(len(reconcile)),
        "order_count": int(len(orders)),
        "buy_count": int((orders["side"] == "BUY").sum()) if not orders.empty else 0,
        "sell_count": int((orders["side"] == "SELL").sum()) if not orders.empty else 0,
        "planned_buy_amount": planned_buy,
        "planned_sell_amount": planned_sell,
        "estimated_cash_after_orders": float(cash) + planned_sell - planned_buy,
        "target_gross_weight": float(sum(raw_weights.values()) * scale),
        "gross_weight_before_scale": float(gross_weight_before_scale),
        "max_position_weight": float(max_position_weight),
        "max_gross_exposure": float(max_gross_exposure),
        "lot_size": int(lot_size),
        "min_trade_value": float(min_trade_value),
        "warning_count": int(len(warnings)),
        "warnings": warnings,
    }
    return reconcile, orders, snapshot, warnings


def _render_preflight_report(snapshot: dict[str, Any]) -> str:
    blockers = snapshot.get("blockers") or []
    warnings = snapshot.get("warnings") or []
    if snapshot.get("status") != "passed":
        warnings_text = "\n".join(f"- {item}" for item in warnings or [])
        blockers_text = "\n".join(f"- {item}" for item in blockers) if blockers else "- None"
    else:
        warnings_text = "\n".join(f"- {item}" for item in warnings) if warnings else "- None"
        blockers_text = "- None"
    return f"""# Live Shadow Preflight

Generated at: `{snapshot.get("generated_at")}`

This check is for research workflow review only. It does not connect to a broker.

## Status

| 项目 | 值 |
| --- | ---: |
| 检查结果 | `{snapshot.get("status")}` |
| 日期 | {snapshot.get("as_of_date")} |
| 当前权益 | {_num(snapshot.get("current_equity"), 2)} |
| 可用现金 | {_num(snapshot.get("cash"), 2)} |
| 差异笔数 | {snapshot.get("order_count", 0)} |
| 买入 / 卖出笔数 | {snapshot.get("buy_count", 0)} / {snapshot.get("sell_count", 0)} |
| 计划买入金额 | {_num(snapshot.get("planned_buy_amount"), 2)} |
| 计划卖出金额 | {_num(snapshot.get("planned_sell_amount"), 2)} |
| 影子调整后估算现金 | {_num(snapshot.get("estimated_cash_after_orders"), 2)} |

## Blocking Items

{blockers_text}

## Warnings

{warnings_text}

## Files

- Live shadow preflight snapshot: `{snapshot.get("snapshot_path")}`
- Holdings: `{snapshot.get("holdings_file")}`
- Targets: `{snapshot.get("targets_file")}`
- Prices: `{snapshot.get("prices_file") or "not provided"}`
"""


def _xlsx_scalar(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def _localized_columns(columns: list[str] | pd.Index) -> list[str]:
    return [LIVE_SHADOW_COLUMN_LABELS.get(str(column), str(column)) for column in columns]


def _localized_frame(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    result.columns = _localized_columns(list(result.columns))
    return result


def _localized_names(columns: set[str]) -> set[str]:
    return {LIVE_SHADOW_COLUMN_LABELS.get(column, column) for column in columns}


def _build_live_shadow_import_templates(
    targets: pd.DataFrame,
    as_of: str,
    blank_rows: int,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if blank_rows < 0:
        raise ValueError("blank_rows must be non-negative.")
    rows: list[dict[str, Any]] = []
    for target in targets.to_dict(orient="records"):
        rows.append(
            {
                "as_of_date": as_of,
                "code": _clean_code(target.get("code")),
                "name": target.get("name"),
                "quantity": "",
                "available_quantity": "",
                "current_price": "",
                "market_value": "",
                "source_account": "",
                "manual_note": "",
                "target_weight_reference": _as_float(target.get("raw_target_weight")),
                "target_value_hint_reference": _as_float(target.get("target_value_hint")),
                "target_price_reference": _as_float(target.get("target_price")),
                "layer_reference": target.get("layer"),
                "target_action_reference": target.get("target_action"),
                "broker_action": "none",
                "research_only": True,
            }
        )
    for _ in range(blank_rows):
        rows.append(
            {
                "as_of_date": as_of,
                "code": "",
                "name": "",
                "quantity": "",
                "available_quantity": "",
                "current_price": "",
                "market_value": "",
                "source_account": "",
                "manual_note": "",
                "target_weight_reference": "",
                "target_value_hint_reference": "",
                "target_price_reference": "",
                "layer_reference": "",
                "target_action_reference": "",
                "broker_action": "none",
                "research_only": True,
            }
        )
    holdings = pd.DataFrame(rows, columns=LIVE_SHADOW_HOLDINGS_TEMPLATE_COLUMNS)
    cash = pd.DataFrame(
        [
            {
                "as_of_date": as_of,
                "cash": "",
                "source_account": "",
                "manual_note": "",
                "broker_action": "none",
                "research_only": True,
            }
        ],
        columns=LIVE_SHADOW_CASH_TEMPLATE_COLUMNS,
    )
    return holdings, cash


def _style_template_sheet(
    sheet: Any,
    required_fill_columns: set[str],
    optional_fill_columns: set[str],
) -> None:
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="FF1F4E78")
    header_font = Font(color="FFFFFFFF", bold=True)
    fill_header_fill = PatternFill("solid", fgColor="FFFFC000")
    required_fill = PatternFill("solid", fgColor="FFFFE699")
    optional_fill = PatternFill("solid", fgColor="FFFFF2CC")
    neutral_fill = PatternFill("solid", fgColor="FFFFFFFF")
    thin = Side(style="thin", color="FFD9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    columns = [sheet.cell(row=1, column=column).value for column in range(1, sheet.max_column + 1)]
    fill_columns = required_fill_columns | optional_fill_columns
    max_row = max(sheet.max_row, 2)
    max_col = len(columns)
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = True
    sheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    for col_idx, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = fill_header_fill if column in fill_columns else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        if column in required_fill_columns:
            cell.comment = Comment("Need human input here before running live-shadow.", "quant_etf_lab")
        elif column in optional_fill_columns:
            cell.comment = Comment("Optional human input for reconciliation context.", "quant_etf_lab")
        width = max(12, min(36, len(str(column)) + 4))
        if column in {"manual_note", "target_action_reference"}:
            width = 34
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    for row_idx in range(2, max_row + 1):
        for col_idx, column in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=column in {"manual_note", "target_action_reference"})
            if column in required_fill_columns:
                cell.fill = required_fill
            elif column in optional_fill_columns:
                cell.fill = optional_fill
            else:
                cell.fill = neutral_fill
            if column in {"target_weight_reference"}:
                cell.number_format = "0.00%"
            elif column in {
                "quantity",
                "available_quantity",
                "current_price",
                "market_value",
                "target_value_hint_reference",
                "target_price_reference",
                "cash",
            }:
                cell.number_format = "0.00"
        sheet.row_dimensions[row_idx].height = 28


def _write_live_shadow_import_template_xlsx(
    holdings: pd.DataFrame,
    cash: pd.DataFrame,
    output_path: str | Path,
) -> Path:
    try:
        from openpyxl import Workbook
    except ImportError as error:  # pragma: no cover - dependency is covered by requirements and CI env.
        raise RuntimeError("openpyxl is required to write the highlighted live-shadow import workbook.") from error

    resolved = Path(output_path)
    resolved.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    display_holdings = _localized_frame(holdings)
    display_cash = _localized_frame(cash)
    holdings_sheet = workbook.active
    holdings_sheet.title = "holdings_to_fill"
    holdings_sheet.append(list(display_holdings.columns))
    for record in display_holdings.to_dict(orient="records"):
        holdings_sheet.append([_xlsx_scalar(record.get(column)) for column in display_holdings.columns])
    _style_template_sheet(
        holdings_sheet,
        _localized_names(LIVE_SHADOW_HOLDINGS_REQUIRED_FILL_COLUMNS),
        _localized_names(LIVE_SHADOW_HOLDINGS_OPTIONAL_FILL_COLUMNS),
    )

    cash_sheet = workbook.create_sheet("cash_to_fill")
    cash_sheet.append(list(display_cash.columns))
    for record in display_cash.to_dict(orient="records"):
        cash_sheet.append([_xlsx_scalar(record.get(column)) for column in display_cash.columns])
    _style_template_sheet(
        cash_sheet,
        _localized_names(LIVE_SHADOW_CASH_REQUIRED_FILL_COLUMNS),
        _localized_names(LIVE_SHADOW_CASH_OPTIONAL_FILL_COLUMNS),
    )

    instructions = workbook.create_sheet("instructions")
    instructions["A1"] = "实盘影子账户导入模板"
    instructions["A2"] = "只填写高亮单元格。本工作簿仅用于人工核对，不连接券商。"
    instructions["A3"] = "持仓必填字段：股票代码、持仓数量、当前价格；可卖数量和市值为可选字段。"
    instructions["A4"] = "现金必填字段：可用现金。人工复核后，可将本工作簿作为 live-shadow 的持仓输入。"
    instructions.column_dimensions["A"].width = 120
    workbook.save(resolved)
    return resolved


def _write_live_shadow_review_queue_xlsx(review: pd.DataFrame, output_path: str | Path) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as error:  # pragma: no cover - dependency is covered by requirements and CI env.
        raise RuntimeError("openpyxl is required to write the highlighted live-shadow review workbook.") from error

    resolved = Path(output_path)
    resolved.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    display_review = _localized_frame(review)
    sheet = workbook.active
    sheet.title = "review_queue"
    sheet.append(list(display_review.columns))
    for record in display_review.to_dict(orient="records"):
        sheet.append([_xlsx_scalar(record.get(column)) for column in display_review.columns])
    _style_template_sheet(
        sheet,
        _localized_names(LIVE_SHADOW_REVIEW_REQUIRED_FILL_COLUMNS),
        _localized_names(LIVE_SHADOW_REVIEW_OPTIONAL_FILL_COLUMNS),
    )
    review_status_label = LIVE_SHADOW_COLUMN_LABELS["review_status_to_fill"]
    if review_status_label in display_review.columns:
        from openpyxl.utils import get_column_letter

        col = get_column_letter(list(display_review.columns).index(review_status_label) + 1)
        allowed = "reviewed_for_manual_consideration,watch_only,skip,needs_data"
        validation = DataValidation(type="list", formula1=f'"{allowed}"', allow_blank=True)
        validation.error = "Use one of: reviewed_for_manual_consideration, watch_only, skip, needs_data."
        validation.errorTitle = "Invalid review status"
        validation.prompt = "Choose a manual review status. This is not an order approval."
        validation.promptTitle = "Manual review status"
        sheet.add_data_validation(validation)
        validation.add(f"{col}2:{col}{max(sheet.max_row, 2)}")
    instructions = workbook.create_sheet("instructions")
    instructions["A1"] = "实盘影子账户人工复核队列"
    instructions["A2"] = "只在人工复核后填写高亮的“复核状态（请填写）”。本工作簿不是下单单据。"
    instructions["A3"] = "复核门禁为 blocked_by_tracking_rule 的行，表示与本地跟踪规则冲突，例如大市值暂停跟踪。"
    instructions["A4"] = "允许状态：reviewed_for_manual_consideration, watch_only, skip, needs_data."
    instructions.column_dimensions["A"].width = 120
    workbook.save(resolved)
    return resolved


def _render_live_shadow_import_template_report(snapshot: dict[str, Any]) -> str:
    return f"""# Live Shadow Import Template

Generated at: `{snapshot.get("generated_at")}`

This is a research-only broker-export fill-in template for manual reconciliation. It does not connect to brokers, place orders, cancel orders, or provide investment advice.

## Summary

| 项目 | 值 |
| --- | ---: |
| 日期 | {snapshot.get("as_of_date")} |
| 目标行数 | {snapshot.get("target_row_count")} |
| 空白补充行 | {snapshot.get("blank_row_count")} |
| 券商动作 | `{snapshot.get("broker_action")}` |
| 仅研究 | `{snapshot.get("research_only")}` |

## Fill-In Fields

- Holdings sheet: fill highlighted `股票代码`, `持仓数量`, and `当前价格`; optionally fill `可卖数量`, `市值`, `账户来源`, and `备注`.
- Cash sheet: fill highlighted `可用现金`.
- The target reference columns are copied from local paper-account outputs for human comparison only.

## Next Manual Step

After filling the workbook, run live-shadow with the filled workbook as `--holdings-file` and the verified cash value as `--cash`. The output remains a manual-review shadow plan only.

## Files

- Holdings CSV template: `{snapshot.get("holdings_template_path")}`
- Cash CSV template: `{snapshot.get("cash_template_path")}`
- Highlighted workbook: `{snapshot.get("workbook_path")}`
- Snapshot JSON: `{snapshot.get("snapshot_path")}`
"""


def _boolish(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    try:
        if pd.isna(value):
            return False
    except (TypeError, ValueError):
        pass
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def _load_target_review_context(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(path)
    try:
        frame = pd.read_csv(path, dtype={"code": str})
    except EmptyDataError:
        frame = pd.DataFrame(columns=["code"])
    if "code" not in frame.columns:
        raise ValueError(f"{path} is missing required columns: code")
    frame = frame.copy()
    frame["code"] = frame["code"].map(_clean_code)
    wanted = [
        "code",
        "layer",
        "portfolio_target_weight",
        "market_cap_yi",
        "tracking_rule_status",
        "tracking_excluded",
        "tracking_exclusion_reason",
    ]
    for column in wanted:
        if column not in frame.columns:
            frame[column] = ""
    return frame[wanted].drop_duplicates("code", keep="first")


def _build_live_shadow_review_queue(
    orders: pd.DataFrame,
    target_context: pd.DataFrame,
    as_of: str,
) -> pd.DataFrame:
    required = {"code", "side", "action", "order_quantity", "price", "estimated_amount"}
    missing = sorted(required - set(orders.columns))
    if missing:
        raise ValueError(f"orders file is missing required columns: {', '.join(missing)}")
    frame = orders.copy()
    frame["code"] = frame["code"].map(_clean_code)
    merged = frame.merge(target_context, on="code", how="left")
    rows: list[dict[str, Any]] = []
    for rank, raw in enumerate(merged.to_dict(orient="records"), start=1):
        tracking_status = _first_text(pd.Series(raw), ("tracking_rule_status",), "unknown")
        tracking_excluded = _boolish(raw.get("tracking_excluded")) or tracking_status not in {
            "",
            "tracking_allowed",
            "unknown",
        }
        review_gate = "blocked_by_tracking_rule" if tracking_excluded else "manual_review_required"
        rows.append(
            {
                "as_of_date": as_of,
                "review_rank": rank,
                "code": _clean_code(raw.get("code")),
                "name": _first_text(pd.Series(raw), ("name",), _clean_code(raw.get("code"))),
                "side": _first_text(pd.Series(raw), ("side",), ""),
                "action": _first_text(pd.Series(raw), ("action",), ""),
                "order_quantity": _as_float(raw.get("order_quantity")),
                "price": _as_float(raw.get("price")),
                "estimated_amount": _as_float(raw.get("estimated_amount")),
                "current_quantity": _as_float(raw.get("current_quantity")),
                "target_quantity": _as_float(raw.get("target_quantity")),
                "current_value": _as_float(raw.get("current_value")),
                "target_value": _as_float(raw.get("target_value")),
                "delta_value": _as_float(raw.get("delta_value")),
                "layer": _first_text(pd.Series(raw), ("layer",), ""),
                "portfolio_target_weight": _as_float(raw.get("portfolio_target_weight")),
                "market_cap_yi": _as_float(raw.get("market_cap_yi")),
                "tracking_rule_status": tracking_status,
                "tracking_excluded": tracking_excluded,
                "tracking_exclusion_reason": _first_text(pd.Series(raw), ("tracking_exclusion_reason",), ""),
                "review_gate": review_gate,
                "review_status_to_fill": "",
                "manual_note_to_fill": "",
                "allowed_review_statuses": "reviewed_for_manual_consideration|watch_only|skip|needs_data",
                "broker_action": "none",
                "research_only": True,
            }
        )
    return pd.DataFrame(rows, columns=LIVE_SHADOW_REVIEW_QUEUE_COLUMNS)


def _render_live_shadow_review_queue_report(snapshot: dict[str, Any], review: pd.DataFrame) -> str:
    if review.empty:
        rows = "| N/A | N/A | N/A | 0 | 0 | N/A | N/A |"
    else:
        lines = []
        for row in review.head(40).to_dict(orient="records"):
            lines.append(
                "| `{code}` | {name} | `{side}` | {qty:.0f} | {amount:.2f} | `{gate}` | {reason} |".format(
                    code=str(row.get("code", "")),
                    name=str(row.get("name", "")),
                    side=str(row.get("side", "")),
                    qty=_as_float(row.get("order_quantity")),
                    amount=_as_float(row.get("estimated_amount")),
                    gate=str(row.get("review_gate", "")),
                    reason=str(row.get("tracking_exclusion_reason", "")) or "N/A",
                )
            )
        rows = "\n".join(lines)
    blocked = snapshot.get("tracking_blocked_codes") or []
    blocked_text = "\n".join(f"- `{code}`" for code in blocked) if blocked else "- None"
    return f"""# Live Shadow Review Queue

Generated at: `{snapshot.get("generated_at")}`

This is a research-only manual review queue. It does not connect to brokers, place orders, cancel orders, or create executable order instructions.

## Summary

| 项目 | 值 |
| --- | ---: |
| 状态 | `{snapshot.get("status")}` |
| 日期 | {snapshot.get("as_of_date")} |
| 复核行数 | {snapshot.get("review_row_count")} |
| 跟踪规则拦截行数 | {snapshot.get("tracking_blocked_count")} |
| 人工复核行数 | {snapshot.get("manual_review_count")} |
| 券商动作 | `{snapshot.get("broker_action")}` |

## Tracking-Rule Blockers

{blocked_text}

## Review Queue

| 股票代码 | 股票名称 | 方向 | 数量 | 金额 | 复核门禁 | 跟踪原因 |
| --- | --- | --- | ---: | ---: | --- | --- |
{rows}

## Files

- Review queue CSV: `{snapshot.get("review_queue_path")}`
- Highlighted workbook: `{snapshot.get("workbook_path")}`
- Snapshot JSON: `{snapshot.get("snapshot_path")}`
"""


def _review_decision_for_row(review_gate: str, review_status: str) -> str:
    if not review_status:
        return "incomplete_review"
    if review_status not in LIVE_SHADOW_REVIEW_ALLOWED_STATUSES:
        return "invalid_review_status"
    if review_gate == "blocked_by_tracking_rule":
        return "blocked_by_tracking_rule"
    if review_status == "reviewed_for_manual_consideration":
        return "manual_considered"
    if review_status == "watch_only":
        return "watch_only"
    if review_status == "skip":
        return "skipped"
    if review_status == "needs_data":
        return "needs_data"
    return "invalid_review_status"


def _build_live_shadow_review_decisions(review: pd.DataFrame, as_of: str) -> pd.DataFrame:
    required = {"code", "review_gate", "review_status_to_fill"}
    missing = sorted(required - set(review.columns))
    if missing:
        raise ValueError(f"review file is missing required columns: {', '.join(missing)}")
    rows: list[dict[str, Any]] = []
    for rank, raw in enumerate(review.to_dict(orient="records"), start=1):
        series = pd.Series(raw)
        review_status = _first_text(series, ("review_status_to_fill",), "").strip()
        review_gate = _first_text(series, ("review_gate",), "").strip() or "manual_review_required"
        rows.append(
            {
                "as_of_date": _first_text(series, ("as_of_date",), as_of) or as_of,
                "review_rank": _as_float(raw.get("review_rank"), float(rank)),
                "code": _clean_code(raw.get("code")),
                "name": _first_text(series, ("name",), _clean_code(raw.get("code"))),
                "side": _first_text(series, ("side",), ""),
                "action": _first_text(series, ("action",), ""),
                "order_quantity": _as_float(raw.get("order_quantity")),
                "price": _as_float(raw.get("price")),
                "estimated_amount": _as_float(raw.get("estimated_amount")),
                "review_gate": review_gate,
                "review_status": review_status,
                "review_decision": _review_decision_for_row(review_gate, review_status),
                "manual_note": _first_text(series, ("manual_note_to_fill", "manual_note"), ""),
                "tracking_exclusion_reason": _first_text(series, ("tracking_exclusion_reason",), ""),
                "broker_action": "none",
                "research_only": True,
            }
        )
    return pd.DataFrame(rows, columns=LIVE_SHADOW_REVIEW_DECISION_COLUMNS)


def _live_shadow_review_decision_status(decisions: pd.DataFrame) -> str:
    if decisions.empty:
        return "empty_review"
    if int((decisions["review_decision"] == "invalid_review_status").sum()):
        return "invalid_review_status"
    if int((decisions["review_decision"] == "incomplete_review").sum()):
        return "incomplete_review"
    if int((decisions["review_decision"] == "blocked_by_tracking_rule").sum()):
        return "review_completed_with_tracking_blockers"
    return "review_completed"


def _render_live_shadow_review_decisions_report(snapshot: dict[str, Any], decisions: pd.DataFrame) -> str:
    if decisions.empty:
        rows = "| N/A | N/A | N/A | 0 | 0 | N/A | N/A | N/A |"
    else:
        lines = []
        for row in decisions.head(40).to_dict(orient="records"):
            lines.append(
                "| `{code}` | {name} | `{side}` | {qty:.0f} | {amount:.2f} | `{status}` | `{decision}` | `{gate}` |".format(
                    code=str(row.get("code", "")),
                    name=str(row.get("name", "")),
                    side=str(row.get("side", "")),
                    qty=_as_float(row.get("order_quantity")),
                    amount=_as_float(row.get("estimated_amount")),
                    status=str(row.get("review_status", "")) or "N/A",
                    decision=str(row.get("review_decision", "")),
                    gate=str(row.get("review_gate", "")),
                )
            )
        rows = "\n".join(lines)
    return f"""# Live Shadow Review Decisions

Generated at: `{snapshot.get("generated_at")}`

This is a research-only review audit. It does not connect to brokers, place orders, cancel orders, or create executable order instructions.

## Summary

| 项目 | 值 |
| --- | ---: |
| 状态 | `{snapshot.get("status")}` |
| 日期 | {snapshot.get("as_of_date")} |
| 复核行数 | {snapshot.get("review_row_count")} |
| 已填复核状态 | {snapshot.get("filled_review_status_count")} |
| 空状态行数 | {snapshot.get("blank_status_count")} |
| 非法状态行数 | {snapshot.get("invalid_status_count")} |
| 大市值/跟踪规则拦截 | {snapshot.get("tracking_blocked_count")} |
| 已人工考虑 | {snapshot.get("manual_considered_count")} |
| 继续观察 | {snapshot.get("watch_only_count")} |
| 跳过 | {snapshot.get("skipped_count")} |
| 需要补数据 | {snapshot.get("needs_data_count")} |
| 券商动作 | `{snapshot.get("broker_action")}` |

## Decisions

| 股票代码 | 股票名称 | 方向 | 数量 | 金额 | 复核状态 | 复核结果 | 复核门禁 |
| --- | --- | --- | ---: | ---: | --- | --- | --- |
{rows}

## Files

- Decisions CSV: `{snapshot.get("decisions_path")}`
- Snapshot JSON: `{snapshot.get("snapshot_path")}`
"""


def run_live_shadow_import_template(
    project_root: str | Path = Path("."),
    targets_file: str | Path | None = DEFAULT_TARGETS_FILE,
    output_dir: str | Path | None = DEFAULT_IMPORT_TEMPLATE_OUTPUT_DIR,
    as_of_date: str | date | None = None,
    blank_rows: int = 10,
) -> LiveShadowImportTemplateResult:
    root = Path(project_root).resolve()
    if output_dir is None:
        output_dir = DEFAULT_IMPORT_TEMPLATE_OUTPUT_DIR
    resolved_targets = _resolve(root, targets_file, DEFAULT_TARGETS_FILE)
    resolved_output = _resolve(root, output_dir, DEFAULT_IMPORT_TEMPLATE_OUTPUT_DIR)
    as_of = _parse_date(as_of_date)
    targets = _normalize_targets(resolved_targets)
    holdings, cash = _build_live_shadow_import_templates(targets, as_of, blank_rows)
    resolved_output.mkdir(parents=True, exist_ok=True)
    holdings_path = resolved_output / "live_holdings_import_template.csv"
    cash_path = resolved_output / "live_cash_import_template.csv"
    workbook_path = resolved_output / "live_shadow_import_template.xlsx"
    snapshot_path = resolved_output / "live_shadow_import_template_snapshot.json"
    report_path = resolved_output / "live_shadow_import_template.md"
    holdings.to_csv(holdings_path, index=False, encoding="utf-8-sig")
    cash.to_csv(cash_path, index=False, encoding="utf-8-sig")
    _write_live_shadow_import_template_xlsx(holdings, cash, workbook_path)
    snapshot = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "status": "completed",
        "as_of_date": as_of,
        "target_row_count": int(len(targets)),
        "blank_row_count": int(blank_rows),
        "holdings_template_path": str(holdings_path),
        "cash_template_path": str(cash_path),
        "workbook_path": str(workbook_path),
        "snapshot_path": str(snapshot_path),
        "report_path": str(report_path),
        "targets_file": str(resolved_targets),
        "broker_action": "none",
        "research_only": True,
    }
    snapshot_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    report_path.write_text(_render_live_shadow_import_template_report(snapshot), encoding="utf-8")
    return LiveShadowImportTemplateResult(
        output_dir=resolved_output,
        holdings_template_path=holdings_path,
        cash_template_path=cash_path,
        workbook_path=workbook_path,
        snapshot_path=snapshot_path,
        report_path=report_path,
        snapshot=snapshot,
        holdings_template=holdings,
        cash_template=cash,
    )


def run_live_shadow_review_queue(
    project_root: str | Path = Path("."),
    orders_file: str | Path | None = None,
    targets_file: str | Path | None = DEFAULT_TARGETS_FILE,
    output_dir: str | Path | None = DEFAULT_REVIEW_OUTPUT_DIR,
    as_of_date: str | date | None = None,
) -> LiveShadowReviewQueueResult:
    root = Path(project_root).resolve()
    if output_dir is None:
        output_dir = DEFAULT_REVIEW_OUTPUT_DIR
    resolved_orders = _resolve(root, orders_file, None)
    resolved_targets = _resolve(root, targets_file, DEFAULT_TARGETS_FILE)
    resolved_output = _resolve(root, output_dir, DEFAULT_REVIEW_OUTPUT_DIR)
    as_of = _parse_date(as_of_date)
    orders = _read_csv(resolved_orders, ("code", "side", "action", "order_quantity", "price", "estimated_amount"))
    target_context = _load_target_review_context(resolved_targets)
    review = _build_live_shadow_review_queue(orders, target_context, as_of)
    resolved_output.mkdir(parents=True, exist_ok=True)
    review_path = resolved_output / "live_shadow_review_queue.csv"
    workbook_path = resolved_output / "live_shadow_review_queue.xlsx"
    snapshot_path = resolved_output / "live_shadow_review_queue_snapshot.json"
    report_path = resolved_output / "live_shadow_review_queue.md"
    review.to_csv(review_path, index=False, encoding="utf-8-sig")
    _write_live_shadow_review_queue_xlsx(review, workbook_path)
    blocked = review[review["review_gate"] == "blocked_by_tracking_rule"] if not review.empty else pd.DataFrame()
    snapshot = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "status": "blocked_tracking_rule_review" if len(blocked) else "ready_for_manual_review",
        "as_of_date": as_of,
        "review_row_count": int(len(review)),
        "tracking_blocked_count": int(len(blocked)),
        "manual_review_count": int((review["review_gate"] == "manual_review_required").sum()) if not review.empty else 0,
        "tracking_blocked_codes": blocked["code"].astype(str).tolist() if not blocked.empty else [],
        "orders_file": str(resolved_orders),
        "targets_file": str(resolved_targets),
        "review_queue_path": str(review_path),
        "workbook_path": str(workbook_path),
        "snapshot_path": str(snapshot_path),
        "report_path": str(report_path),
        "broker_action": "none",
        "research_only": True,
    }
    snapshot_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    report_path.write_text(_render_live_shadow_review_queue_report(snapshot, review), encoding="utf-8")
    return LiveShadowReviewQueueResult(
        output_dir=resolved_output,
        review_queue_path=review_path,
        workbook_path=workbook_path,
        snapshot_path=snapshot_path,
        report_path=report_path,
        snapshot=snapshot,
        review_queue=review,
    )


def run_live_shadow_review_decisions(
    project_root: str | Path = Path("."),
    review_file: str | Path | None = None,
    output_dir: str | Path | None = DEFAULT_REVIEW_DECISION_OUTPUT_DIR,
    as_of_date: str | date | None = None,
) -> LiveShadowReviewDecisionResult:
    root = Path(project_root).resolve()
    if output_dir is None:
        output_dir = DEFAULT_REVIEW_DECISION_OUTPUT_DIR
    resolved_review = _resolve(root, review_file, None)
    resolved_output = _resolve(root, output_dir, DEFAULT_REVIEW_DECISION_OUTPUT_DIR)
    as_of = _parse_date(as_of_date)
    review = _read_csv(resolved_review, ("code", "review_gate", "review_status_to_fill"), sheet_name="review_queue")
    decisions = _build_live_shadow_review_decisions(review, as_of)
    resolved_output.mkdir(parents=True, exist_ok=True)
    decisions_path = resolved_output / "live_shadow_review_decisions.csv"
    snapshot_path = resolved_output / "live_shadow_review_decisions_snapshot.json"
    report_path = resolved_output / "live_shadow_review_decisions.md"
    decisions.to_csv(decisions_path, index=False, encoding="utf-8-sig")
    review_status = decisions["review_status"].fillna("").astype(str).str.strip() if not decisions.empty else pd.Series(dtype=str)
    decision = decisions["review_decision"].fillna("").astype(str).str.strip() if not decisions.empty else pd.Series(dtype=str)
    snapshot = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "status": _live_shadow_review_decision_status(decisions),
        "as_of_date": as_of,
        "review_row_count": int(len(decisions)),
        "filled_review_status_count": int((review_status != "").sum()),
        "blank_status_count": int((decision == "incomplete_review").sum()),
        "invalid_status_count": int((decision == "invalid_review_status").sum()),
        "tracking_blocked_count": int((decision == "blocked_by_tracking_rule").sum()),
        "manual_considered_count": int((decision == "manual_considered").sum()),
        "watch_only_count": int((decision == "watch_only").sum()),
        "skipped_count": int((decision == "skipped").sum()),
        "needs_data_count": int((decision == "needs_data").sum()),
        "review_file": str(resolved_review),
        "decisions_path": str(decisions_path),
        "snapshot_path": str(snapshot_path),
        "report_path": str(report_path),
        "broker_action": "none",
        "research_only": True,
    }
    snapshot_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    report_path.write_text(_render_live_shadow_review_decisions_report(snapshot, decisions), encoding="utf-8")
    return LiveShadowReviewDecisionResult(
        output_dir=resolved_output,
        decisions_path=decisions_path,
        snapshot_path=snapshot_path,
        report_path=report_path,
        snapshot=snapshot,
        decisions=decisions,
    )


def run_live_shadow_preflight(
    project_root: str | Path = Path("."),
    holdings_file: str | Path | None = None,
    targets_file: str | Path | None = DEFAULT_TARGETS_FILE,
    prices_file: str | Path | None = None,
    cash: float = 0.0,
    output_dir: str | Path | None = DEFAULT_PRE_FLIGHT_OUTPUT_DIR,
    as_of_date: str | date | None = None,
    lot_size: int = 100,
    min_trade_value: float = 1000.0,
    max_position_weight: float = 0.20,
    max_gross_exposure: float = 1.0,
) -> LiveShadowPreflightResult:
    root = Path(project_root).resolve()
    if output_dir is None:
        output_dir = DEFAULT_PRE_FLIGHT_OUTPUT_DIR
    resolved_holdings = _resolve(root, holdings_file, None)
    resolved_targets = _resolve(root, targets_file, DEFAULT_TARGETS_FILE)
    resolved_prices = _resolve(root, prices_file, None) if prices_file is not None else None
    resolved_output = _resolve(root, output_dir, DEFAULT_PRE_FLIGHT_OUTPUT_DIR)
    if cash is None:
        cash = 0.0

    as_of = _parse_date(as_of_date)
    snapshot_path = resolved_output / "live_shadow_preflight_snapshot.json"
    report_path = resolved_output / "live_shadow_preflight_report.md"

    blockers: list[str] = []
    warnings: list[str] = []
    status = "passed"
    plan_snapshot: dict[str, Any] = {
        "status": status,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "as_of_date": as_of,
        "cash": float(cash),
        "holdings_file": str(resolved_holdings),
        "targets_file": str(resolved_targets),
        "prices_file": str(resolved_prices) if resolved_prices is not None else None,
        "snapshot_path": str(snapshot_path),
        "output_dir": str(resolved_output),
        "trade_plan_status": "manual_review_only",
        "broker_action": "none",
        "research_only": True,
    }
    try:
        holdings = _normalize_holdings(resolved_holdings)
        targets = _normalize_targets(resolved_targets)
        prices = _normalize_prices(resolved_prices)
        reconcile, orders, context, computed_warnings = _build_shadow_plan(
            holdings=holdings,
            targets=targets,
            prices=prices,
            cash=float(cash),
            as_of=as_of,
            lot_size=lot_size,
            min_trade_value=min_trade_value,
            max_position_weight=max_position_weight,
            max_gross_exposure=max_gross_exposure,
        )
        warnings.extend(computed_warnings)
        plan_snapshot.update(context)
    except Exception as error:  # Preflight should always be explicit and consumable by scripts.
        status = "blocked"
        blockers.append(str(error))
        warnings.append(str(error))
        plan_snapshot["error"] = str(error)
        plan_snapshot["order_count"] = 0
        plan_snapshot["buy_count"] = 0
        plan_snapshot["sell_count"] = 0
        plan_snapshot["planned_buy_amount"] = 0.0
        plan_snapshot["planned_sell_amount"] = 0.0
        plan_snapshot["estimated_cash_after_orders"] = float(cash)
        plan_snapshot["current_equity"] = float(cash)
        plan_snapshot["warning_count"] = len(warnings)
    else:
        plan_snapshot["warning_count"] = len(warnings)
    plan_snapshot["status"] = status
    plan_snapshot["blockers"] = blockers
    plan_snapshot["warnings"] = warnings
    resolved_output.mkdir(parents=True, exist_ok=True)
    snapshot_path.write_text(json.dumps(plan_snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    report_path.write_text(_render_preflight_report(plan_snapshot), encoding="utf-8")
    return LiveShadowPreflightResult(
        output_dir=resolved_output,
        snapshot_path=snapshot_path,
        report_path=report_path,
        holdings_path=resolved_holdings,
        targets_path=resolved_targets,
        prices_path=resolved_prices,
        snapshot=plan_snapshot,
    )


def run_live_shadow(
    project_root: str | Path = Path("."),
    holdings_file: str | Path | None = None,
    targets_file: str | Path | None = DEFAULT_TARGETS_FILE,
    prices_file: str | Path | None = None,
    output_dir: str | Path | None = DEFAULT_OUTPUT_DIR,
    cash: float = 0.0,
    as_of_date: str | date | None = None,
    lot_size: int = 100,
    min_trade_value: float = 1000.0,
    max_position_weight: float = 0.20,
    max_gross_exposure: float = 1.0,
) -> LiveShadowResult:
    root = Path(project_root).resolve()
    resolved_holdings = _resolve(root, holdings_file, None)
    resolved_targets = _resolve(root, targets_file, DEFAULT_TARGETS_FILE)
    resolved_prices = _resolve(root, prices_file, None) if prices_file is not None else None
    resolved_output = _resolve(root, output_dir, DEFAULT_OUTPUT_DIR)
    if cash is None:
        cash = 0.0
    if min_trade_value is None:
        min_trade_value = 0.0

    holdings = _normalize_holdings(resolved_holdings)
    targets = _normalize_targets(resolved_targets)
    prices = _normalize_prices(resolved_prices)
    as_of = _parse_date(as_of_date)
    reconcile, orders, payload_snapshot, warnings = _build_shadow_plan(
        holdings=holdings,
        targets=targets,
        prices=prices,
        cash=float(cash),
        as_of=as_of,
        lot_size=lot_size,
        min_trade_value=min_trade_value,
        max_position_weight=max_position_weight,
        max_gross_exposure=max_gross_exposure,
    )
    snapshot = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "as_of_date": as_of,
        "trade_plan_status": "manual_review_only",
        "broker_action": "none",
        "research_only": True,
        "holdings_file": str(resolved_holdings),
        "targets_file": str(resolved_targets),
        "prices_file": str(resolved_prices) if resolved_prices is not None else None,
        "output_dir": str(resolved_output),
        "orders_path": str(resolved_output / "live_shadow_orders.csv"),
        "reconcile_path": str(resolved_output / "live_shadow_reconcile.csv"),
        "snapshot_path": str(resolved_output / "live_shadow_snapshot.json"),
        "report_path": str(resolved_output / "live_shadow_plan.md"),
        "cash": float(cash),
        **payload_snapshot,
        "warnings": warnings,
    }
    snapshot["warning_count"] = len(warnings)

    resolved_output.mkdir(parents=True, exist_ok=True)
    orders_path = resolved_output / "live_shadow_orders.csv"
    reconcile_path = resolved_output / "live_shadow_reconcile.csv"
    snapshot_path = resolved_output / "live_shadow_snapshot.json"
    report_path = resolved_output / "live_shadow_plan.md"

    snapshot["note"] = "Shadow-account plan for manual review only; no broker API is used."

    orders.to_csv(orders_path, index=False, encoding="utf-8")
    reconcile.to_csv(reconcile_path, index=False, encoding="utf-8")
    snapshot_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    report_path.write_text(_render_report(snapshot, orders), encoding="utf-8")

    return LiveShadowResult(
        output_dir=resolved_output,
        snapshot_path=snapshot_path,
        report_path=report_path,
        orders_path=orders_path,
        reconcile_path=reconcile_path,
        holdings_path=resolved_holdings,
        targets_path=resolved_targets,
        prices_path=resolved_prices,
        snapshot=snapshot,
        orders=orders,
        reconcile=reconcile,
    )
