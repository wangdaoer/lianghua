"""Paper-account ledger built from portfolio walk-forward allocator outputs."""

from __future__ import annotations

import json
import hashlib
import io
import re
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import yaml

from ._compat import read_text
from .artifact_io import publish_json_if_semantically_changed, write_bytes_if_changed, write_text_if_changed
from .portfolio import (
    CurveConfig,
    PortfolioConfig,
    _parse_satellite_filter,
    _validate_weights,
    load_portfolio_config,
    run_portfolio_combine,
)
from .market_cap import DEFAULT_STOCK_MARKET_CAP_PATH, load_stock_market_cap_cache
from .market_data_source import MarketSnapshotLoadResult
from .paper_formatting import (
    _as_float,
    _format_date,
    _format_yyyymmdd,
    _maybe_pct,
    _md_text,
    _negative_threshold,
    _num,
    _optional_float,
    _parse_yyyymmdd,
    _pct,
    _safe_text,
    _signed_pct,
)


DEFAULT_PAPER_CONFIG = Path("configs/portfolio_core_satellite_quality_v2_guarded.yaml")
DEFAULT_ALLOCATOR_DIR = Path(
    "outputs/portfolio_source_selection/main_chinext_source_selection_highgain_pos8_dd50_cap30_activation_dd50_20260624"
)
DEFAULT_TRIGGER_SIGNAL_PATH = Path("D:/codex/outputs/signal_history/signals_latest.csv")
DEFAULT_STOCK_REVIEW_NOTES_PATH = Path("outputs/research/stock_target_review_notes.csv")
DEFAULT_STOCK_REVIEW_OUTCOMES_HISTORY_PATH = Path("outputs/research/stock_target_review_outcomes_history.csv")
DEFAULT_STOCK_TRACKING_MAX_MARKET_CAP_YI = 0.0
DEFAULT_STOCK_REVIEW_OUTCOME_MIN_EVALUABLE = 20
DEFAULT_STOCK_REVIEW_OUTCOME_MIN_GROUP_EVALUABLE = 5

STOCK_TARGET_REVIEW_NOTE_COLUMNS = [
    "layer",
    "code",
    "name",
    "first_seen_date",
    "last_seen_date",
    "last_review_bucket",
    "last_review_stage",
    "last_review_priority_score",
    "manual_status",
    "manual_note",
    "reviewed_at",
    "reviewed_by",
    "next_review_date",
    "last_model_reason",
    "broker_action",
    "research_only",
]
STOCK_TARGET_REVIEW_MANUAL_COLUMNS = [
    "manual_status",
    "manual_note",
    "reviewed_at",
    "reviewed_by",
    "next_review_date",
]
MANUAL_STATUS_ALIASES = {
    "": "unreviewed",
    "unreviewed": "unreviewed",
    "pending": "unreviewed",
    "pending_review": "unreviewed",
    "待复核": "unreviewed",
    "未复核": "unreviewed",
    "reviewed": "reviewed",
    "done": "reviewed",
    "checked": "reviewed",
    "已复核": "reviewed",
    "已确认": "reviewed",
    "watch": "watch",
    "monitor": "watch",
    "watching": "watch",
    "观察": "watch",
    "继续观察": "watch",
    "resolved": "resolved",
    "closed": "resolved",
    "cleared": "resolved",
    "已解决": "resolved",
    "关闭": "resolved",
    "exclude": "exclude_candidate",
    "exclude_candidate": "exclude_candidate",
    "avoid": "exclude_candidate",
    "剔除候选": "exclude_candidate",
    "排除候选": "exclude_candidate",
}
MANUAL_STATUS_EXPLANATIONS = {
    "unreviewed": "No human review note has been recorded yet.",
    "reviewed": "Human review has been recorded; keep the model row visible for audit.",
    "watch": "Human review asks to keep watching this row.",
    "resolved": "Human review marks the item as resolved for the current audit loop.",
    "exclude_candidate": "Human review marks the row as an exclusion candidate; model targets are not changed automatically.",
    "other": "Manual status is not one of the built-in values; keep it visible for review.",
}
STOCK_TARGET_REVIEW_ACTION_COLUMNS = [
    "date",
    "action_rank",
    "action_code",
    "action_stage",
    "action_priority_score",
    "source_review_rank",
    "review_bucket",
    "review_stage",
    "layer",
    "code",
    "name",
    "portfolio_target_weight",
    "unrealized_return",
    "manual_status",
    "manual_status_normalized",
    "manual_review_state",
    "next_review_date",
    "action_reason",
    "recommended_review",
    "broker_action",
    "research_only",
]
STOCK_TARGET_REVIEW_ASSISTANT_COLUMNS = [
    "date",
    "assistant_rank",
    "action_code",
    "review_bucket",
    "review_stage",
    "layer",
    "code",
    "name",
    "portfolio_target_weight",
    "unrealized_return",
    "holding_days",
    "review_priority_score",
    "latest_price_date",
    "latest_close",
    "latest_pct_change",
    "support_price",
    "pressure_price",
    "stop_loss_reference",
    "trigger_condition",
    "invalidation_condition",
    "suggested_review_posture",
    "local_price_history_status",
    "local_price_history_path",
    "price_staleness_days",
    "return_5d",
    "return_20d",
    "drawdown_from_20d_high",
    "manual_status_options",
    "manual_note_prompt",
    "evidence_checklist",
    "review_reason",
    "recommended_review",
    "broker_action",
    "research_only",
]
STOCK_TARGET_REVIEW_DECISION_TEMPLATE_COLUMNS = [
    "date",
    "decision_rank",
    "source_assistant_rank",
    "layer",
    "code",
    "name",
    "review_posture",
    "status_hint",
    "manual_status_to_fill",
    "manual_note_to_fill",
    "next_review_date_to_fill",
    "reviewed_by_to_fill",
    "allowed_manual_statuses",
    "portfolio_target_weight",
    "unrealized_return",
    "latest_close",
    "latest_pct_change",
    "support_price",
    "pressure_price",
    "stop_loss_reference",
    "trigger_condition",
    "invalidation_condition",
    "evidence_checklist",
    "review_reason",
    "recommended_review",
    "notes_path",
    "broker_action",
    "research_only",
]
STOCK_TARGET_REVIEW_DECISION_TEMPLATE_USER_FILL_COLUMNS = [
    "manual_status_to_fill",
    "manual_note_to_fill",
    "next_review_date_to_fill",
    "reviewed_by_to_fill",
]
STOCK_TARGET_REVIEW_DECISION_APPLY_AUDIT_COLUMNS = [
    "date",
    "apply_rank",
    "apply_status",
    "apply_message",
    "layer",
    "code",
    "name",
    "previous_manual_status",
    "new_manual_status",
    "manual_note",
    "reviewed_at",
    "reviewed_by",
    "next_review_date",
    "review_posture",
    "review_reason",
    "broker_action",
    "research_only",
]
STOCK_TARGET_COLUMNS = [
    "date",
    "layer",
    "code",
    "name",
    "layer_target_weight",
    "layer_internal_weight",
    "portfolio_target_weight",
    "portfolio_target_value",
    "model_quantity",
    "model_market_value",
    "close_price",
    "price_date",
    "price_source",
    "avg_cost",
    "unrealized_return",
    "holding_days",
    "last_trade_price",
    "last_trade_date",
    "target_action",
    "target_explanation",
    "risk_filter_status",
    "selection_explanation",
    "source_filter_explanation",
    "source_risk_explanation",
    "source_strategy_profile",
    "source_strategy_name",
    "source_project_name",
    "factor_weights",
    "factor_weights_text",
    "source_config_path",
    "trigger_monitor_status",
    "trigger_summary",
    "trigger_run_time",
    "trigger_signal_age_days",
    "trigger_signal_validity_status",
    "trigger_signal_type",
    "trigger_action",
    "trigger_reason",
    "trigger_score",
    "trigger_score_level",
    "trigger_pct",
    "trigger_last",
    "source_backtest_dir",
    "source_trades_path",
    "execution_gate_action",
]
STOCK_TARGET_REVIEW_OUTCOME_HORIZONS = (1, 5, 10, 20)
STOCK_TARGET_REVIEW_OUTCOME_COLUMNS = [
    "date",
    "layer",
    "code",
    "name",
    "review_bucket",
    "review_stage",
    "review_priority_score",
    "action_code",
    "manual_status",
    "manual_status_normalized",
    "manual_review_state",
    "entry_date",
    "entry_close",
    "price_source",
    "outcome_status",
    "broker_action",
    "research_only",
    "price_source_size",
    "price_source_mtime_ns",
    "price_source_fingerprint",
]
STOCK_TARGET_REVIEW_OUTCOME_SOURCE_METADATA_COLUMNS = (
    "price_source_size",
    "price_source_mtime_ns",
    "price_source_fingerprint",
)
for _horizon in STOCK_TARGET_REVIEW_OUTCOME_HORIZONS:
    STOCK_TARGET_REVIEW_OUTCOME_COLUMNS.extend(
        [
            f"future_date_{_horizon}d",
            f"future_close_{_horizon}d",
            f"return_{_horizon}d",
            f"outcome_status_{_horizon}d",
        ]
    )


STOCK_TARGET_REVIEW_OUTCOME_ANALYSIS_DIMENSIONS = (
    "overall",
    "action_code",
    "review_bucket",
    "review_stage",
    "manual_status_normalized",
    "layer",
)
STOCK_TARGET_REVIEW_OUTCOME_ANALYSIS_COLUMNS = [
    "dimension",
    "group_value",
    "history_row_count",
    "complete_count",
    "partial_count",
    "pending_count",
    "missing_entry_price_count",
]
for _horizon in STOCK_TARGET_REVIEW_OUTCOME_HORIZONS:
    STOCK_TARGET_REVIEW_OUTCOME_ANALYSIS_COLUMNS.extend(
        [
            f"evaluable_{_horizon}d",
            f"coverage_{_horizon}d",
            f"avg_return_{_horizon}d",
            f"median_return_{_horizon}d",
            f"win_rate_{_horizon}d",
            f"avg_win_{_horizon}d",
            f"avg_loss_{_horizon}d",
        ]
    )


STOCK_TARGET_REVIEW_OUTCOME_CALENDAR_COLUMNS = [
    "as_of_date",
    "horizon",
    "analysis_status",
    "readiness_status",
    "maturity_state",
    "pending_count",
    "due_pending_count",
    "evaluable_count",
    "remaining_to_min_evaluable",
    "estimated_next_evaluable_date",
    "days_until_next_evaluable",
    "estimated_all_pending_mature_by",
    "estimated_total_after_pending_mature",
    "min_evaluable",
    "qualified_group_count",
    "min_group_evaluable",
    "basis",
    "research_only",
    "broker_action",
]
STOCK_TARGET_REVIEW_OUTCOME_DUE_COLUMNS = ["due_rank"] + STOCK_TARGET_REVIEW_OUTCOME_CALENDAR_COLUMNS


@dataclass(frozen=True)
class PaperAccountResult:
    output_dir: Path
    ledger_path: Path
    audit_path: Path
    monthly_returns_path: Path
    target_holdings_path: Path
    target_holdings_json_path: Path
    target_holdings_report_path: Path
    stock_targets_path: Path
    stock_targets_json_path: Path
    stock_targets_report_path: Path
    stock_target_review_path: Path
    stock_target_review_json_path: Path
    stock_target_review_report_path: Path
    stock_target_review_notes_path: Path
    stock_target_review_notes_snapshot_path: Path
    stock_target_review_actions_path: Path
    stock_target_review_actions_json_path: Path
    stock_target_review_actions_report_path: Path
    stock_target_review_assistant_path: Path
    stock_target_review_assistant_json_path: Path
    stock_target_review_assistant_report_path: Path
    stock_target_review_decision_template_path: Path
    stock_target_review_decision_template_json_path: Path
    stock_target_review_decision_template_report_path: Path
    stock_target_review_decision_template_xlsx_path: Path
    stock_target_review_outcomes_path: Path
    stock_target_review_outcomes_json_path: Path
    stock_target_review_outcomes_report_path: Path
    stock_target_review_outcomes_history_path: Path
    stock_target_review_outcomes_history_snapshot_path: Path
    stock_target_review_outcomes_history_json_path: Path
    stock_target_review_outcomes_history_report_path: Path
    stock_target_review_outcome_analysis_path: Path
    stock_target_review_outcome_analysis_json_path: Path
    stock_target_review_outcome_analysis_report_path: Path
    stock_target_review_outcome_calendar_path: Path
    stock_target_review_outcome_calendar_json_path: Path
    stock_target_review_outcome_calendar_report_path: Path
    stock_target_review_outcome_due_path: Path
    stock_target_review_outcome_due_json_path: Path
    stock_target_review_outcome_due_report_path: Path
    metrics_path: Path
    report_path: Path
    ledger: pd.DataFrame
    audit: pd.DataFrame
    target_holdings: pd.DataFrame
    target_holdings_payload: dict[str, Any]
    stock_targets: pd.DataFrame
    stock_targets_payload: dict[str, Any]
    stock_target_review: pd.DataFrame
    stock_target_review_payload: dict[str, Any]
    stock_target_review_actions: pd.DataFrame
    stock_target_review_actions_payload: dict[str, Any]
    stock_target_review_assistant: pd.DataFrame
    stock_target_review_assistant_payload: dict[str, Any]
    stock_target_review_decision_template: pd.DataFrame
    stock_target_review_decision_template_payload: dict[str, Any]
    stock_target_review_outcomes: pd.DataFrame
    stock_target_review_outcomes_payload: dict[str, Any]
    stock_target_review_outcomes_history: pd.DataFrame
    stock_target_review_outcomes_history_payload: dict[str, Any]
    stock_target_review_outcome_analysis: pd.DataFrame
    stock_target_review_outcome_analysis_payload: dict[str, Any]
    stock_target_review_outcome_calendar: pd.DataFrame
    stock_target_review_outcome_calendar_payload: dict[str, Any]
    stock_target_review_outcome_due: pd.DataFrame
    stock_target_review_outcome_due_payload: dict[str, Any]
    monthly_returns: pd.DataFrame
    metrics: dict[str, Any]


@dataclass(frozen=True)
class StockTargetReviewDecisionApplyResult:
    output_dir: Path
    template_path: Path
    notes_path: Path
    notes_snapshot_path: Path
    audit_path: Path
    payload_path: Path
    report_path: Path
    notes: pd.DataFrame
    audit: pd.DataFrame
    payload: dict[str, Any]


def _resolve(project_root: Path, path: str | Path | None, default: Path) -> Path:
    raw = Path(path) if path is not None else default
    return raw if raw.is_absolute() else project_root / raw


def _read_json(path: Path) -> dict[str, Any]:
    try:
        return json.loads(read_text(path))
    except (OSError, json.JSONDecodeError):
        return {}


def _read_yaml(path: Path) -> dict[str, Any]:
    try:
        data = yaml.safe_load(read_text(path))
    except (OSError, yaml.YAMLError):
        return {}
    return data if isinstance(data, dict) else {}


def _json_records(frame: pd.DataFrame) -> list[dict[str, Any]]:
    if frame.empty:
        return []
    records: list[dict[str, Any]] = []
    for row in frame.to_dict(orient="records"):
        clean: dict[str, Any] = {}
        for key, value in row.items():
            if isinstance(value, pd.Timestamp):
                clean[key] = value.strftime("%Y-%m-%d")
            elif pd.isna(value):
                clean[key] = None
            else:
                clean[key] = value
        records.append(clean)
    return records


def _as_dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _selected_params_path(project_root: Path, allocator_dir: Path, raw_path: Any) -> Path:
    raw = Path(str(raw_path))
    if raw.is_absolute():
        return raw
    if str(raw).replace("\\", "/").startswith("outputs/"):
        return project_root / raw
    return allocator_dir / raw


def _resolve_equity_column(path: Path, preferred: str) -> str:
    try:
        columns = pd.read_csv(path, nrows=0).columns
    except (OSError, pd.errors.EmptyDataError, pd.errors.ParserError):
        return preferred
    options = [preferred, "equity", "stitched_equity", "equity_curve"]
    for option in options:
        if option in set(columns):
            return option
    return preferred


def _candidate_config_from_payload(
    base: PortfolioConfig,
    payload: dict[str, Any],
    start_date: str,
    end_date: str,
) -> PortfolioConfig:
    allocation = _as_dict(payload.get("allocation")) or payload
    regime = _as_dict(allocation.get("regime_overrides"))
    source_path = payload.get("source_path")
    satellite = base.satellite
    if source_path:
        resolved_source = Path(str(source_path))
        if not resolved_source.is_absolute():
            resolved_source = base.project_root / resolved_source
        satellite = CurveConfig(
            name=base.satellite.name,
            path=resolved_source,
            equity_column=_resolve_equity_column(resolved_source, base.satellite.equity_column),
        )
    return PortfolioConfig(
        project_root=base.project_root,
        name=base.name,
        initial_cash=base.initial_cash,
        output_dir=base.output_dir,
        core=base.core,
        satellite=satellite,
        benchmark_path=base.benchmark_path,
        benchmark_close_column=base.benchmark_close_column,
        ma_window=int(regime.get("ma_window", base.ma_window)),
        drop_window=int(regime.get("drop_window", base.drop_window)),
        risk_on_drop_threshold=float(regime.get("risk_on_drop_threshold", base.risk_on_drop_threshold)),
        crash_drop_threshold=float(regime.get("crash_drop_threshold", base.crash_drop_threshold)),
        default_regime=str(regime.get("default_regime", base.default_regime)),
        weights=_validate_weights(allocation.get("weights", base.weights)),
        satellite_filter=_parse_satellite_filter(allocation.get("satellite_filter", {})),
        start_date=start_date,
        end_date=end_date,
        notes=base.notes,
    )


def _load_allocator_windows(project_root: Path, allocator_dir: Path) -> pd.DataFrame:
    summary_path = allocator_dir / "portfolio_walk_forward_summary.csv"
    if not summary_path.exists():
        raise FileNotFoundError(f"Portfolio walk-forward summary not found: {summary_path}")
    summary = pd.read_csv(summary_path)
    required = {"window", "test_start", "test_end", "selected_candidate", "selected_params_path"}
    missing = required - set(summary.columns)
    if missing:
        raise ValueError(f"Portfolio walk-forward summary missing columns: {sorted(missing)}")
    data = summary.copy()
    data["_test_start_sort"] = data["test_start"].map(_parse_yyyymmdd)
    data = data.dropna(subset=["_test_start_sort"]).sort_values("_test_start_sort").reset_index(drop=True)
    if data.empty:
        raise ValueError(f"No usable windows in {summary_path}")
    data["selected_params_resolved"] = [
        str(_selected_params_path(project_root, allocator_dir, raw_path)) for raw_path in data["selected_params_path"]
    ]
    return data


PAPER_EQUITY_CACHE_VERSION = 3


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _window_equity_fingerprint(
    base_config: PortfolioConfig,
    allocator_dir: Path,
    config_path: Path,
) -> tuple[str, pd.DataFrame]:
    windows = _load_allocator_windows(base_config.project_root, allocator_dir)
    dependencies = [
        config_path,
        allocator_dir / "portfolio_walk_forward_summary.csv",
        Path(base_config.core.path),
        Path(base_config.satellite.path),
        Path(base_config.benchmark_path),
    ]
    for raw_path in windows["selected_params_resolved"].astype(str):
        params_path = Path(raw_path)
        dependencies.append(params_path)
        payload = _read_json(params_path)
        source_path = payload.get("source_path") if payload else None
        if source_path:
            resolved_source = Path(str(source_path))
            if not resolved_source.is_absolute():
                resolved_source = base_config.project_root / resolved_source
            dependencies.append(resolved_source)

    records = []
    for path in sorted({item.resolve() for item in dependencies}, key=str):
        records.append(
            {
                "path": str(path),
                "sha256": _file_sha256(path) if path.is_file() else "missing",
            }
        )
    encoded = json.dumps(
        {"version": PAPER_EQUITY_CACHE_VERSION, "dependencies": records},
        ensure_ascii=False,
        sort_keys=True,
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest(), windows


def _window_equity_frame(
    base_config: PortfolioConfig,
    allocator_dir: Path,
    *,
    config_path: Path | None = None,
    cache_dir: Path | None = None,
) -> pd.DataFrame:
    windows: pd.DataFrame | None = None
    fingerprint: str | None = None
    cache_path = cache_dir / "window_equity.csv" if cache_dir is not None else None
    metadata_path = cache_dir / "window_equity_cache.json" if cache_dir is not None else None
    if config_path is not None and cache_path is not None and metadata_path is not None:
        fingerprint, windows = _window_equity_fingerprint(base_config, allocator_dir, config_path)
        if cache_path.exists() and metadata_path.exists():
            metadata = _read_json(metadata_path)
            if metadata.get("fingerprint") == fingerprint:
                try:
                    cached = pd.read_csv(cache_path, float_precision="round_trip")
                except (OSError, pd.errors.EmptyDataError, pd.errors.ParserError):
                    cached = pd.DataFrame()
                if not cached.empty and "date" in cached.columns:
                    cached["date"] = pd.to_datetime(cached["date"], errors="coerce")
                    cached = cached.dropna(subset=["date"])
                    if not cached.empty:
                        return cached.reset_index(drop=True)

    if windows is None:
        windows = _load_allocator_windows(base_config.project_root, allocator_dir)
    frames: list[pd.DataFrame] = []
    for row in windows.to_dict("records"):
        params_path = Path(str(row["selected_params_resolved"]))
        payload = _read_json(params_path)
        if not payload:
            raise ValueError(f"Selected params are missing or invalid: {params_path}")
        start_date = _format_yyyymmdd(row["test_start"])
        end_date = _format_yyyymmdd(row["test_end"])
        config = _candidate_config_from_payload(base_config, payload, start_date, end_date)
        run_id = f"{row['window']}_{row['selected_candidate']}_paper"
        result = run_portfolio_combine(config, run_id=run_id, write_outputs=False)
        equity = result.equity.copy()
        if equity.empty:
            continue
        equity["window"] = str(row["window"])
        equity["selected_candidate"] = str(row["selected_candidate"])
        equity["selected_score"] = _as_float(row.get("selected_score"), default=np.nan)
        equity["train_start"] = _format_date(row.get("train_start", ""))
        equity["train_end"] = _format_date(row.get("train_end", ""))
        equity["test_start"] = _format_date(row.get("test_start", ""))
        equity["test_end"] = _format_date(row.get("test_end", ""))
        equity["selected_params_path"] = str(params_path)
        frames.append(equity)
    if not frames:
        raise ValueError("No paper-account equity frames generated.")
    combined = pd.concat(frames, ignore_index=True)
    combined["date"] = pd.to_datetime(combined["date"], errors="coerce")
    combined = combined.dropna(subset=["date"]).sort_values(["date", "window"]).drop_duplicates(subset=["date"], keep="last")
    combined = combined.reset_index(drop=True)
    if cache_path is not None and metadata_path is not None and fingerprint is not None:
        write_text_if_changed(cache_path, combined.to_csv(index=False, float_format="%.17g"))
        write_text_if_changed(
            metadata_path,
            json.dumps(
                {"version": PAPER_EQUITY_CACHE_VERSION, "fingerprint": fingerprint},
                ensure_ascii=False,
                indent=2,
            ),
        )
    return combined


def _target_change_turnover(previous: tuple[float, float, float] | None, current: tuple[float, float, float]) -> float:
    if previous is None:
        previous = (0.0, 0.0, 1.0)
    return 0.5 * sum(abs(curr - prev) for curr, prev in zip(current, previous))


def build_paper_account_ledger(
    base_equity: pd.DataFrame,
    initial_cash: float,
    rebalance_cost_rate: float = 0.0,
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any], pd.DataFrame]:
    if base_equity.empty:
        raise ValueError("base_equity cannot be empty.")
    if rebalance_cost_rate < 0:
        raise ValueError("rebalance_cost_rate cannot be negative.")

    ledger_rows: list[dict[str, Any]] = []
    audit_rows: list[dict[str, Any]] = []
    previous_equity = float(initial_cash)
    previous_target: tuple[float, float, float] | None = None
    previous_state: dict[str, Any] | None = None
    running_max = float(initial_cash)

    for row in base_equity.itertuples(index=False):
        current_target = (
            _as_float(getattr(row, "core_weight", 0.0)),
            _as_float(getattr(row, "satellite_weight", 0.0)),
            _as_float(getattr(row, "cash_weight", 0.0)),
        )
        turnover = _target_change_turnover(previous_target, current_target)
        estimated_fee = previous_equity * turnover * rebalance_cost_rate
        gross_return = _as_float(getattr(row, "portfolio_return", 0.0))
        equity_after_fee = max(previous_equity - estimated_fee, 0.0)
        ending_equity = equity_after_fee * (1.0 + gross_return)
        net_return = ending_equity / previous_equity - 1.0 if previous_equity > 0 else 0.0
        running_max = max(running_max, ending_equity)
        current_drawdown = ending_equity / running_max - 1.0 if running_max > 0 else 0.0

        state = {
            "window": str(getattr(row, "window", "")),
            "selected_candidate": str(getattr(row, "selected_candidate", "")),
            "effective_regime": str(getattr(row, "effective_regime", "")),
            "effective_reason": str(getattr(row, "effective_reason", "")),
            "satellite_effective_reason": str(getattr(row, "satellite_effective_reason", "")),
            "target": current_target,
        }
        event_reasons: list[str] = []
        if previous_state is None:
            event_reasons.append("initial_allocation")
        else:
            if state["window"] != previous_state["window"]:
                event_reasons.append("window_switch")
            if state["selected_candidate"] != previous_state["selected_candidate"]:
                event_reasons.append("candidate_change")
            if state["effective_regime"] != previous_state["effective_regime"]:
                event_reasons.append("risk_regime_change")
            if state["satellite_effective_reason"] != previous_state["satellite_effective_reason"]:
                event_reasons.append("satellite_filter_change")
            if tuple(round(value, 10) for value in current_target) != tuple(
                round(value, 10) for value in previous_state["target"]
            ):
                event_reasons.append("target_weight_change")
        if event_reasons:
            previous_target_values = previous_state["target"] if previous_state else (0.0, 0.0, 1.0)
            audit_rows.append(
                {
                    "date": getattr(row, "date"),
                    "event_type": ";".join(event_reasons),
                    "window": state["window"],
                    "selected_candidate": state["selected_candidate"],
                    "from_regime": previous_state["effective_regime"] if previous_state else "",
                    "to_regime": state["effective_regime"],
                    "from_core_weight": previous_target_values[0],
                    "from_satellite_weight": previous_target_values[1],
                    "from_cash_weight": previous_target_values[2],
                    "to_core_weight": current_target[0],
                    "to_satellite_weight": current_target[1],
                    "to_cash_weight": current_target[2],
                    "target_change_turnover": turnover,
                    "estimated_fee": estimated_fee,
                    "effective_reason": state["effective_reason"],
                    "satellite_effective_reason": state["satellite_effective_reason"],
                    "selected_params_path": str(getattr(row, "selected_params_path", "")),
                }
            )

        ledger_rows.append(
            {
                "date": getattr(row, "date"),
                "window": state["window"],
                "selected_candidate": state["selected_candidate"],
                "effective_regime": state["effective_regime"],
                "effective_reason": state["effective_reason"],
                "satellite_effective_scale": _as_float(getattr(row, "satellite_effective_scale", 1.0), 1.0),
                "satellite_effective_reason": state["satellite_effective_reason"],
                "core_target_weight": current_target[0],
                "satellite_target_weight": current_target[1],
                "cash_target_weight": current_target[2],
                "core_return": _as_float(getattr(row, "core_return", 0.0)),
                "satellite_return": _as_float(getattr(row, "satellite_return", 0.0)),
                "gross_return": gross_return,
                "net_return": net_return,
                "starting_equity": previous_equity,
                "target_change_turnover": turnover,
                "rebalance_cost_rate": rebalance_cost_rate,
                "estimated_fee": estimated_fee,
                "ending_equity": ending_equity,
                "target_core_value": ending_equity * current_target[0],
                "target_satellite_value": ending_equity * current_target[1],
                "target_cash_value": ending_equity * current_target[2],
                "current_drawdown": current_drawdown,
                "signal_regime": str(getattr(row, "signal_regime", "")),
                "signal_reason": str(getattr(row, "signal_reason", "")),
                "benchmark_close": _as_float(getattr(row, "benchmark_close", np.nan), default=np.nan),
                "benchmark_ma": _as_float(getattr(row, "benchmark_ma", np.nan), default=np.nan),
                "benchmark_drop": _as_float(getattr(row, "benchmark_drop", np.nan), default=np.nan),
            }
        )
        previous_equity = ending_equity
        previous_target = current_target
        previous_state = state

    ledger = pd.DataFrame(ledger_rows)
    audit = pd.DataFrame(audit_rows)
    metrics = _paper_metrics(ledger, audit, float(initial_cash), rebalance_cost_rate)
    monthly = _paper_monthly_returns(ledger)
    return ledger, audit, metrics, monthly


def _paper_monthly_returns(ledger: pd.DataFrame) -> pd.DataFrame:
    if ledger.empty:
        return pd.DataFrame(columns=["date", "monthly_return"])
    series = ledger.set_index("date")["ending_equity"].sort_index()
    try:
        monthly_equity = series.resample("ME").last()
    except ValueError:
        monthly_equity = series.resample("M").last()
    monthly = monthly_equity.pct_change().dropna().reset_index()
    monthly.columns = ["date", "monthly_return"]
    return monthly


def _paper_metrics(
    ledger: pd.DataFrame,
    audit: pd.DataFrame,
    initial_cash: float,
    rebalance_cost_rate: float,
) -> dict[str, Any]:
    if ledger.empty:
        return {
            "initial_cash": initial_cash,
            "final_equity": initial_cash,
            "total_return": 0.0,
            "cagr": 0.0,
            "max_drawdown": 0.0,
            "sharpe": 0.0,
            "rebalance_cost_rate": rebalance_cost_rate,
        }
    data = ledger.sort_values("date").copy()
    final_equity = _as_float(data["ending_equity"].iloc[-1], initial_cash)
    total_return = final_equity / initial_cash - 1.0
    first_date = pd.Timestamp(data["date"].iloc[0])
    last_date = pd.Timestamp(data["date"].iloc[-1])
    years = max((last_date - first_date).days, 1) / 365.25
    cagr = (final_equity / initial_cash) ** (1.0 / years) - 1.0 if final_equity > 0 else 0.0
    returns = pd.to_numeric(data["net_return"], errors="coerce").dropna()
    sharpe = 0.0
    if len(returns) > 1 and float(returns.std()) != 0.0:
        sharpe = float(np.sqrt(252.0) * returns.mean() / returns.std())
    return {
        "initial_cash": float(initial_cash),
        "final_equity": float(final_equity),
        "total_return": float(total_return),
        "cagr": float(cagr),
        "max_drawdown": float(pd.to_numeric(data["current_drawdown"], errors="coerce").min()),
        "sharpe": float(sharpe),
        "observation_days": int(len(data)),
        "latest_date": _format_date(data["date"].iloc[-1]),
        "latest_window": str(data["window"].iloc[-1]),
        "latest_candidate": str(data["selected_candidate"].iloc[-1]),
        "latest_regime": str(data["effective_regime"].iloc[-1]),
        "latest_core_weight": _as_float(data["core_target_weight"].iloc[-1]),
        "latest_satellite_weight": _as_float(data["satellite_target_weight"].iloc[-1]),
        "latest_cash_weight": _as_float(data["cash_target_weight"].iloc[-1]),
        "current_drawdown": _as_float(data["current_drawdown"].iloc[-1]),
        "average_core_weight": float(pd.to_numeric(data["core_target_weight"], errors="coerce").mean()),
        "average_satellite_weight": float(pd.to_numeric(data["satellite_target_weight"], errors="coerce").mean()),
        "average_cash_weight": float(pd.to_numeric(data["cash_target_weight"], errors="coerce").mean()),
        "satellite_active_day_ratio": float((pd.to_numeric(data["satellite_target_weight"], errors="coerce") > 0).mean()),
        "total_target_change_turnover": float(pd.to_numeric(data["target_change_turnover"], errors="coerce").sum()),
        "total_estimated_fee": float(pd.to_numeric(data["estimated_fee"], errors="coerce").sum()),
        "rebalance_cost_rate": float(rebalance_cost_rate),
        "audit_event_count": int(len(audit)),
    }


def _target_action(previous_weight: float, target_weight: float) -> str:
    if abs(target_weight - previous_weight) < 1e-12:
        return "hold"
    if previous_weight <= 1e-12 and target_weight > 1e-12:
        return "open"
    if previous_weight > 1e-12 and target_weight <= 1e-12:
        return "close"
    if target_weight > previous_weight:
        return "increase"
    return "reduce"


def build_target_holdings(
    ledger: pd.DataFrame,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Build a layer-level target holdings sheet from the latest paper ledger row."""

    if ledger.empty:
        empty = pd.DataFrame(
            columns=[
                "date",
                "layer",
                "target_weight",
                "previous_weight",
                "weight_change",
                "target_value",
                "target_action",
                "effective_regime",
                "effective_reason",
                "satellite_effective_reason",
                "selected_candidate",
                "window",
                "signal_regime",
                "signal_reason",
                "note",
            ]
        )
        return empty, {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "status": "empty_ledger",
            "target_holding_count": 0,
            "research_only": True,
            "broker_action": "none",
        }

    data = ledger.sort_values("date").reset_index(drop=True)
    latest = data.iloc[-1]
    previous = data.iloc[-2] if len(data) > 1 else None
    latest_equity = _as_float(latest.get("ending_equity"))
    layer_specs = [
        ("core", "core_target_weight", "target_core_value", "Defensive core model curve target."),
        ("satellite", "satellite_target_weight", "target_satellite_value", "Allocator-gated satellite model curve target."),
        ("cash", "cash_target_weight", "target_cash_value", "Unallocated cash buffer in the simulation."),
    ]
    rows: list[dict[str, Any]] = []
    for layer, weight_column, value_column, note in layer_specs:
        target_weight = _as_float(latest.get(weight_column))
        previous_weight = _as_float(previous.get(weight_column)) if previous is not None else (1.0 if layer == "cash" else 0.0)
        target_value = _as_float(latest.get(value_column), latest_equity * target_weight)
        rows.append(
            {
                "date": latest.get("date"),
                "layer": layer,
                "target_weight": target_weight,
                "previous_weight": previous_weight,
                "weight_change": target_weight - previous_weight,
                "target_value": target_value,
                "target_action": _target_action(previous_weight, target_weight),
                "effective_regime": str(latest.get("effective_regime", "")),
                "effective_reason": str(latest.get("effective_reason", "")),
                "satellite_effective_reason": str(latest.get("satellite_effective_reason", "")),
                "selected_candidate": str(latest.get("selected_candidate", "")),
                "window": str(latest.get("window", "")),
                "signal_regime": str(latest.get("signal_regime", "")),
                "signal_reason": str(latest.get("signal_reason", "")),
                "note": note,
            }
        )
    holdings = pd.DataFrame(rows)
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "status": "ok",
        "latest_date": _format_date(latest.get("date")),
        "latest_equity": latest_equity,
        "latest_candidate": str(latest.get("selected_candidate", "")),
        "latest_regime": str(latest.get("effective_regime", "")),
        "effective_reason": str(latest.get("effective_reason", "")),
        "satellite_effective_reason": str(latest.get("satellite_effective_reason", "")),
        "signal_regime": str(latest.get("signal_regime", "")),
        "signal_reason": str(latest.get("signal_reason", "")),
        "target_holding_count": int(len(holdings)),
        "active_target_count": int((pd.to_numeric(holdings["target_weight"], errors="coerce") > 0).sum()),
        "total_target_weight": float(pd.to_numeric(holdings["target_weight"], errors="coerce").sum()),
        "total_target_value": float(pd.to_numeric(holdings["target_value"], errors="coerce").sum()),
        "target_actions": holdings[["layer", "target_action", "target_weight", "weight_change", "target_value"]].to_dict(
            orient="records"
        ),
        "research_only": True,
        "broker_action": "none",
        "note": "Layer-level simulation target only; this is not a broker order list.",
    }
    return holdings, payload


def _render_target_holdings_report(
    holdings: pd.DataFrame,
    payload: dict[str, Any],
    output_dir: Path,
) -> str:
    if holdings.empty:
        rows = "| `N/A` | `N/A` | N/A | N/A | N/A | Empty paper ledger |"
    else:
        rows = "\n".join(
            "| `{layer}` | `{action}` | {weight} | {change} | {value} | {note} |".format(
                layer=str(row["layer"]),
                action=str(row["target_action"]),
                weight=_pct(row["target_weight"]),
                change=_signed_pct(row["weight_change"]),
                value=_num(row["target_value"], 2),
                note=str(row["note"]),
            )
            for _, row in holdings.iterrows()
        )
    return f"""# Paper Target Holdings

Generated at: `{payload.get("generated_at")}`

This is a research-only simulation target sheet. It does not connect to brokers, place orders, or provide investment advice.

## Summary

| Item | Value |
| --- | ---: |
| Status | `{payload.get("status")}` |
| Latest date | {payload.get("latest_date", "N/A")} |
| Latest candidate | `{payload.get("latest_candidate", "N/A")}` |
| Latest regime | `{payload.get("latest_regime", "N/A")}` |
| Latest equity | {_num(payload.get("latest_equity"), 2)} |
| Total target weight | {_pct(payload.get("total_target_weight"))} |
| Active target count | {payload.get("active_target_count", 0)} |
| Broker action | `{payload.get("broker_action")}` |

## Target Holdings

| Layer | Action | Target weight | Weight change | Target value | Note |
| --- | --- | ---: | ---: | ---: | --- |
{rows}

## Signal Context

| Item | Value |
| --- | ---: |
| Effective reason | `{payload.get("effective_reason", "N/A")}` |
| Satellite reason | `{payload.get("satellite_effective_reason", "N/A")}` |
| Signal regime | `{payload.get("signal_regime", "N/A")}` |
| Signal reason | `{payload.get("signal_reason", "N/A")}` |

## Files

- Target holdings CSV: `{output_dir / "target_holdings.csv"}`
- Target holdings JSON: `{output_dir / "target_holdings.json"}`
"""


def _latest_curve_window(curve_path: Path) -> tuple[str | None, str | None]:
    if not curve_path.exists():
        return None, None
    try:
        curve = pd.read_csv(curve_path)
    except (OSError, pd.errors.EmptyDataError):
        return None, None
    if curve.empty or "date" not in curve.columns:
        return None, None
    data = curve.copy()
    data["date"] = pd.to_datetime(data["date"], errors="coerce")
    data = data.dropna(subset=["date"]).sort_values("date")
    if data.empty:
        return None, None
    latest = data.iloc[-1]
    latest_date = pd.Timestamp(latest["date"]).strftime("%Y-%m-%d")
    window = latest.get("window")
    if window in (None, "") or pd.isna(window):
        return None, latest_date
    return str(window), latest_date


def _resolve_backtest_dir(project_root: Path, curve_path: Path) -> tuple[Path | None, str | None, str]:
    window, latest_date = _latest_curve_window(curve_path)
    if window:
        raw = Path(window)
        candidates = [raw] if raw.is_absolute() else [project_root / "outputs" / "backtests" / raw.name, curve_path.parent / raw]
        for candidate in candidates:
            if candidate.exists():
                return candidate, latest_date, "curve_window"
    summary_path = curve_path.parent / "walk_forward_summary.csv"
    if summary_path.exists():
        try:
            summary = pd.read_csv(summary_path)
        except (OSError, pd.errors.EmptyDataError):
            summary = pd.DataFrame()
        if not summary.empty and "test_run_dir" in summary.columns:
            sort_column = "test_end" if "test_end" in summary.columns else None
            data = summary.copy()
            if sort_column:
                data["_sort"] = data[sort_column].map(_parse_yyyymmdd)
                data = data.dropna(subset=["_sort"]).sort_values("_sort")
            latest = data.iloc[-1]
            raw = Path(str(latest["test_run_dir"]))
            candidates = [raw] if raw.is_absolute() else [project_root / raw, project_root / "outputs" / "backtests" / raw.name]
            for candidate in candidates:
                if candidate.exists():
                    return candidate, latest_date, "walk_forward_summary"
    return None, latest_date, "missing_backtest_dir"


def _latest_stock_close(project_root: Path, code: str, as_of_date: Any) -> tuple[float | None, str | None, str]:
    target_date = pd.to_datetime(as_of_date, errors="coerce")
    candidates = [
        project_root / "data" / "processed" / "stocks" / f"{code}.csv",
        project_root / "data" / "processed" / f"{code}.csv",
    ]
    for path in candidates:
        if not path.exists():
            continue
        try:
            frame = pd.read_csv(path)
        except (OSError, pd.errors.EmptyDataError):
            continue
        if frame.empty or not {"date", "close"}.issubset(frame.columns):
            continue
        data = frame[["date", "close"]].copy()
        data["date"] = pd.to_datetime(data["date"], errors="coerce")
        data["close"] = pd.to_numeric(data["close"], errors="coerce")
        data = data.dropna(subset=["date", "close"]).sort_values("date")
        if pd.notna(target_date):
            data = data[data["date"] <= pd.Timestamp(target_date)]
        if data.empty:
            continue
        latest = data.iloc[-1]
        return float(latest["close"]), pd.Timestamp(latest["date"]).strftime("%Y-%m-%d"), str(path)
    return None, None, "missing_local_close"


def _market_snapshot_close_lookup(
    market_snapshot: MarketSnapshotLoadResult | None,
    as_of_date: Any,
) -> dict[str, tuple[float, str, str]]:
    if market_snapshot is None or not market_snapshot.rows or not market_snapshot.trade_date:
        return {}
    snapshot_date = pd.to_datetime(market_snapshot.trade_date, errors="coerce")
    target_date = pd.to_datetime(as_of_date, errors="coerce")
    if pd.isna(snapshot_date) or pd.isna(target_date) or pd.Timestamp(snapshot_date).normalize() != pd.Timestamp(target_date).normalize():
        return {}
    date_text = pd.Timestamp(snapshot_date).strftime("%Y-%m-%d")
    source = f"market_snapshot:{market_snapshot.source_kind}"
    lookup: dict[str, tuple[float, str, str]] = {}
    for row in market_snapshot.rows:
        code = _clean_code(row.get("security_code") or row.get("code") or row.get("symbol"))
        close = None
        for field in ("close_price", "close", "last_price"):
            close = _optional_float(row.get(field))
            if close is not None:
                break
        if code and close is not None and close > 0:
            lookup[code] = (float(close), date_text, source)
    return lookup


def _clean_code(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    if not digits:
        return ""
    return digits[-6:].zfill(6)


def _first_column(columns: list[str], candidates: list[str]) -> str | None:
    column_set = set(columns)
    for candidate in candidates:
        if candidate in column_set:
            return candidate
    return None


def _format_factor_weights(factor_weights: Any) -> str:
    if not isinstance(factor_weights, dict) or not factor_weights:
        return "N/A"
    items = sorted(factor_weights.items(), key=lambda item: _as_float(item[1]), reverse=True)
    return ", ".join(f"{key}={_as_float(value) * 100:.0f}%" for key, value in items)


def _top_factor_text(factor_weights: Any) -> str:
    if not isinstance(factor_weights, dict) or not factor_weights:
        return "no_factor_weights"
    items = sorted(factor_weights.items(), key=lambda item: _as_float(item[1]), reverse=True)
    return "+".join(str(key) for key, _ in items[:3])


def _strategy_explanation(backtest_dir: Path | None) -> dict[str, Any]:
    config_path = backtest_dir / "config_used.yaml" if backtest_dir is not None else None
    config = _read_yaml(config_path) if config_path is not None else {}
    strategy = config.get("strategy", {}) if isinstance(config.get("strategy"), dict) else {}
    risk = config.get("risk", {}) if isinstance(config.get("risk"), dict) else {}
    project = config.get("project", {}) if isinstance(config.get("project"), dict) else {}
    strategy_name = _safe_text(strategy.get("name"), "unknown")
    max_positions = int(_as_float(strategy.get("max_positions"), 0))
    rebalance_interval = int(_as_float(strategy.get("factor_rebalance_interval"), 0))
    min_score = _as_float(strategy.get("min_score"), 0.0)
    allocation_buffer = _as_float(strategy.get("allocation_buffer"), 0.0)
    factor_weights = strategy.get("factor_weights") if isinstance(strategy.get("factor_weights"), dict) else {}
    factor_text = _format_factor_weights(factor_weights)
    factor_json = json.dumps(factor_weights, ensure_ascii=False, sort_keys=True) if factor_weights else "{}"

    if strategy_name == "multi_factor":
        strategy_profile = f"multi_factor:{_top_factor_text(factor_weights)}"
        selection_explanation = (
            f"Local walk-forward multi-factor source; max_positions={max_positions}, "
            f"rebalance={rebalance_interval}d, min_score={min_score:.1f}, weights: {factor_text}."
        )
    else:
        strategy_profile = strategy_name
        selection_explanation = (
            f"Local walk-forward {strategy_name} source; max_positions={max_positions}, "
            f"allocation_buffer={allocation_buffer:.2f}."
        )

    entry_filter = bool(strategy.get("factor_entry_filter_enabled", False))
    breadth_filter = bool(strategy.get("market_breadth_enabled", False))
    breadth_exposure = bool(strategy.get("market_breadth_exposure_enabled", False))
    filter_parts = [
        "entry_filter=on" if entry_filter else "entry_filter=off",
        "breadth_filter=on" if breadth_filter else "breadth_filter=off",
        "breadth_exposure=on" if breadth_exposure else "breadth_exposure=off",
    ]
    risk_enabled = bool(risk.get("enabled", False))
    if risk_enabled:
        risk_explanation = (
            f"Risk model on: benchmark={_safe_text(risk.get('benchmark_code'), 'N/A')} "
            f"MA{int(_as_float(risk.get('benchmark_ma_window'), 0))}, "
            f"{int(_as_float(risk.get('benchmark_drop_window'), 0))}d drop threshold={_pct(risk.get('benchmark_drop_threshold'))}, "
            f"off_exposure={_pct(risk.get('benchmark_off_exposure'))}, "
            f"drawdown_guard={_pct(risk.get('protection_drawdown'))}."
        )
    else:
        risk_explanation = "Risk model not configured in source backtest."

    return {
        "project_name": _safe_text(project.get("name"), "unknown"),
        "strategy_name": strategy_name,
        "strategy_profile": strategy_profile,
        "selection_explanation": selection_explanation,
        "source_filter_explanation": "; ".join(filter_parts),
        "source_risk_explanation": risk_explanation,
        "source_risk_enabled": risk_enabled,
        "factor_weights_text": factor_text,
        "factor_weights_json": factor_json,
        "config_path": str(config_path) if config_path is not None and config_path.exists() else None,
        "config_status": "ok" if config else "missing_or_unreadable",
    }


def _read_latest_trigger_signals(signal_path: str | Path | None) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    resolved_path = Path(signal_path) if signal_path is not None else DEFAULT_TRIGGER_SIGNAL_PATH
    meta: dict[str, Any] = {
        "latest_trigger_signal_path": str(resolved_path),
        "trigger_signal_status": "missing",
        "trigger_signal_count": 0,
        "latest_trigger_run_time": None,
    }
    if not resolved_path.exists():
        return {}, meta
    try:
        signals = pd.read_csv(resolved_path, dtype={"code": str}, encoding="utf-8-sig")
    except UnicodeDecodeError:
        try:
            signals = pd.read_csv(resolved_path, dtype={"code": str}, encoding="gbk")
        except (OSError, pd.errors.ParserError, UnicodeDecodeError):
            meta["trigger_signal_status"] = "unreadable"
            return {}, meta
    except (OSError, pd.errors.ParserError):
        meta["trigger_signal_status"] = "unreadable"
        return {}, meta
    if signals.empty:
        meta["trigger_signal_status"] = "empty"
        return {}, meta

    columns = list(signals.columns)
    code_col = _first_column(columns, ["code", "代码", "股票代码", "symbol"])
    if code_col is None:
        meta["trigger_signal_status"] = "missing_code_column"
        return {}, meta
    run_time_col = _first_column(columns, ["run_time", "datetime", "time", "日期"])
    signal_type_col = _first_column(columns, ["signal_type", "type", "类型", "信号类型"])
    action_col = _first_column(columns, ["action", "动作", "建议动作"])
    reason_col = _first_column(columns, ["reason", "理由", "trigger_reason"])
    score_col = _first_column(columns, ["score", "评分"])
    score_level_col = _first_column(columns, ["score_level", "评级", "level"])
    pct_col = _first_column(columns, ["pct", "涨跌幅"])
    last_col = _first_column(columns, ["last", "price", "latest_price", "最新价"])

    latest: dict[str, dict[str, Any]] = {}
    run_times: list[str] = []
    for row in signals.to_dict(orient="records"):
        code = _clean_code(row.get(code_col))
        if not code:
            continue
        run_time = _safe_text(row.get(run_time_col)) if run_time_col else ""
        if run_time:
            run_times.append(run_time)
        latest[code] = {
            "trigger_name": _safe_text(row.get("name"), code),
            "trigger_run_time": run_time or None,
            "trigger_signal_type": _safe_text(row.get(signal_type_col), "N/A") if signal_type_col else "N/A",
            "trigger_action": _safe_text(row.get(action_col), "N/A") if action_col else "N/A",
            "trigger_reason": _safe_text(row.get(reason_col), "N/A") if reason_col else "N/A",
            "trigger_score": _optional_float(row.get(score_col)) if score_col else None,
            "trigger_score_level": _safe_text(row.get(score_level_col), "N/A") if score_level_col else "N/A",
            "trigger_pct": _optional_float(row.get(pct_col)) if pct_col else None,
            "trigger_last": _optional_float(row.get(last_col)) if last_col else None,
        }
    meta["trigger_signal_status"] = "ok"
    meta["trigger_signal_count"] = len(latest)
    meta["latest_trigger_run_time"] = max(run_times) if run_times else None
    return latest, meta


def _trigger_summary(trigger: dict[str, Any] | None) -> str:
    if not trigger:
        return "not_in_latest_trigger_report"
    score = trigger.get("trigger_score")
    score_text = f"{score:.0f}" if score is not None else "N/A"
    return f"{trigger.get('trigger_signal_type', 'N/A')}/{trigger.get('trigger_action', 'N/A')}/score={score_text}"


def _trigger_signal_validity(trigger: dict[str, Any] | None, latest_date: Any, valid_days: int = 3) -> dict[str, Any]:
    if not trigger:
        return {"trigger_signal_age_days": pd.NA, "trigger_signal_validity_status": "no_current_trigger"}
    run_time = trigger.get("trigger_run_time")
    signal_date = pd.to_datetime(str(run_time).split("_")[0], errors="coerce") if run_time else pd.NaT
    as_of = pd.to_datetime(latest_date, errors="coerce")
    if pd.isna(signal_date) or pd.isna(as_of):
        return {"trigger_signal_age_days": pd.NA, "trigger_signal_validity_status": "trigger_date_unknown"}
    age_days = int((pd.Timestamp(as_of).normalize() - pd.Timestamp(signal_date).normalize()).days)
    return {
        "trigger_signal_age_days": age_days,
        "trigger_signal_validity_status": "fresh_trigger_signal" if age_days <= valid_days else "stale_trigger_signal",
    }


def _open_positions_from_trades(trades_path: Path, as_of_date: Any) -> pd.DataFrame:
    columns = ["code", "name", "quantity", "avg_cost", "last_trade_price", "last_trade_date"]
    if not trades_path.exists():
        return pd.DataFrame(columns=columns)
    try:
        trades = pd.read_csv(trades_path, dtype={"code": str})
    except (OSError, pd.errors.EmptyDataError):
        return pd.DataFrame(columns=columns)
    required = {"date", "code", "name", "side", "price", "quantity"}
    if trades.empty or not required.issubset(trades.columns):
        return pd.DataFrame(columns=columns)
    data = trades.copy()
    data["date"] = pd.to_datetime(data["date"], errors="coerce")
    target_date = pd.to_datetime(as_of_date, errors="coerce")
    if pd.notna(target_date):
        data = data[data["date"] <= pd.Timestamp(target_date)]
    data["code"] = data["code"].astype(str).str.zfill(6)
    data["price"] = pd.to_numeric(data["price"], errors="coerce")
    data["quantity"] = pd.to_numeric(data["quantity"], errors="coerce")
    data = data.dropna(subset=["date", "code", "price", "quantity"]).sort_values("date")

    state: dict[str, dict[str, Any]] = {}
    for row in data.itertuples(index=False):
        code = str(row.code).zfill(6)
        item = state.setdefault(
            code,
            {"code": code, "name": str(row.name), "quantity": 0.0, "avg_cost": 0.0, "last_trade_price": 0.0, "last_trade_date": None},
        )
        side = str(row.side).upper()
        price = float(row.price)
        quantity = max(float(row.quantity), 0.0)
        item["name"] = str(row.name)
        item["last_trade_price"] = price
        item["last_trade_date"] = pd.Timestamp(row.date).strftime("%Y-%m-%d")
        if side == "BUY" and quantity > 0:
            old_quantity = float(item["quantity"])
            new_quantity = old_quantity + quantity
            item["avg_cost"] = ((float(item["avg_cost"]) * old_quantity) + (price * quantity)) / new_quantity
            item["quantity"] = new_quantity
        elif side == "SELL" and quantity > 0:
            item["quantity"] = max(float(item["quantity"]) - quantity, 0.0)
            if item["quantity"] <= 1e-8:
                item["quantity"] = 0.0
                item["avg_cost"] = 0.0

    rows = [item for item in state.values() if float(item["quantity"]) > 1e-8]
    return pd.DataFrame(rows, columns=columns)


def _empty_market_cap_cache_payload(path: Path, status: str, error: str | None = None) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "stock_market_cap_path": str(path),
        "stock_market_cap_cache_status": status,
        "stock_market_cap_cache_row_count": 0,
        "stock_market_cap_cache_latest_snapshot_date": None,
        "stock_market_cap_cache_updated_at": None,
        "stock_market_cap_cache_source": None,
    }
    if error:
        payload["stock_market_cap_cache_error"] = error
    return payload


def _load_market_cap_cache_payload(path: Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    try:
        frame = load_stock_market_cap_cache(path)
    except FileNotFoundError:
        return pd.DataFrame(), _empty_market_cap_cache_payload(path, "missing")
    except Exception as exc:
        return pd.DataFrame(), _empty_market_cap_cache_payload(path, "invalid", str(exc))

    payload = {
        "stock_market_cap_path": str(path),
        "stock_market_cap_cache_status": "ok",
        "stock_market_cap_cache_row_count": int(len(frame)),
        "stock_market_cap_cache_latest_snapshot_date": _safe_text(frame["snapshot_date"].max()) if "snapshot_date" in frame.columns and not frame.empty else None,
        "stock_market_cap_cache_updated_at": _safe_text(frame["updated_at"].max()) if "updated_at" in frame.columns and not frame.empty else None,
        "stock_market_cap_cache_source": _safe_text(frame["source"].iloc[0]) if "source" in frame.columns and not frame.empty else None,
    }
    return frame, payload


def _boolish(value: Any) -> bool:
    if isinstance(value, (bool, np.bool_)):
        return bool(value)
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def apply_stock_market_cap_tracking_rule(
    stock_targets: pd.DataFrame,
    market_cap_frame: pd.DataFrame,
    max_market_cap_yi: float = DEFAULT_STOCK_TRACKING_MAX_MARKET_CAP_YI,
    market_cap_cache_payload: dict[str, Any] | None = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Mark stock targets that should not enter the temporary tracking queue."""
    threshold = None if max_market_cap_yi is None else float(max_market_cap_yi)
    limit_enabled = threshold is not None and threshold > 0
    market_cap_lookup: dict[str, dict[str, Any]] = {}
    if not market_cap_frame.empty and {"code", "market_cap_yi"}.issubset(market_cap_frame.columns):
        clean_caps = market_cap_frame.copy()
        clean_caps["code"] = clean_caps["code"].map(_clean_code)
        market_cap_lookup = clean_caps.set_index("code").to_dict(orient="index")

    frame = stock_targets.copy()
    added_columns = [
        "market_cap_yi",
        "market_cap_snapshot_date",
        "market_cap_source",
        "market_cap_status",
        "tracking_excluded",
        "tracking_rule_status",
        "tracking_exclusion_reason",
    ]
    if frame.empty:
        for column in added_columns:
            frame[column] = []
    else:
        market_caps: list[float] = []
        snapshot_dates: list[str] = []
        sources: list[str] = []
        cap_statuses: list[str] = []
        tracking_excluded: list[bool] = []
        tracking_statuses: list[str] = []
        tracking_reasons: list[str] = []
        for row in frame.to_dict(orient="records"):
            code = _clean_code(row.get("code"))
            record = market_cap_lookup.get(code)
            market_cap_yi = _optional_float(record.get("market_cap_yi")) if record else None
            if market_cap_yi is None:
                market_caps.append(np.nan)
                snapshot_dates.append(_safe_text(record.get("snapshot_date")) if record else "")
                sources.append(_safe_text(record.get("source")) if record else "")
                cap_statuses.append("missing")
                tracking_excluded.append(False)
                tracking_statuses.append("tracking_allowed_market_cap_missing")
                tracking_reasons.append("market cap snapshot is missing; keep row visible and do not guess")
                continue

            excluded = bool(limit_enabled and market_cap_yi > threshold)
            market_caps.append(float(market_cap_yi))
            snapshot_dates.append(_safe_text(record.get("snapshot_date")))
            sources.append(_safe_text(record.get("source")))
            cap_statuses.append("ok")
            tracking_excluded.append(excluded)
            if excluded:
                tracking_statuses.append("excluded_large_market_cap")
                tracking_reasons.append(f"market_cap_yi {market_cap_yi:.2f} > max_tracking_market_cap_yi {threshold:.2f}")
            else:
                tracking_statuses.append("tracking_allowed")
                tracking_reasons.append("")

        frame["market_cap_yi"] = market_caps
        frame["market_cap_snapshot_date"] = snapshot_dates
        frame["market_cap_source"] = sources
        frame["market_cap_status"] = cap_statuses
        frame["tracking_excluded"] = tracking_excluded
        frame["tracking_rule_status"] = tracking_statuses
        frame["tracking_exclusion_reason"] = tracking_reasons

    excluded_count = int(frame["tracking_excluded"].sum()) if "tracking_excluded" in frame.columns and not frame.empty else 0
    missing_count = int((frame.get("market_cap_status") == "missing").sum()) if not frame.empty else 0
    payload = dict(market_cap_cache_payload or {})
    payload.update(
        {
            "stock_tracking_max_market_cap_yi": threshold if limit_enabled else None,
            "stock_tracking_requested_max_market_cap_yi": threshold,
            "stock_tracking_excluded_large_market_cap_count": excluded_count,
            "stock_tracking_allowed_count": int(len(frame) - excluded_count),
            "stock_tracking_market_cap_missing_count": missing_count,
            "stock_tracking_rule_note": "Rows above max market cap stay in stock_targets for audit but are skipped by stock_target_review tracking.",
        }
    )
    return frame, payload


def build_stock_targets(
    project_root: Path,
    portfolio_config: PortfolioConfig,
    target_holdings: pd.DataFrame,
    latest_date: Any,
    trigger_signal_path: str | Path | None = None,
    stock_market_cap_path: str | Path | None = None,
    stock_tracking_max_market_cap_yi: float = DEFAULT_STOCK_TRACKING_MAX_MARKET_CAP_YI,
    market_snapshot: MarketSnapshotLoadResult | None = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    source_layers: list[dict[str, Any]] = []
    target_by_layer = target_holdings.set_index("layer").to_dict(orient="index") if not target_holdings.empty else {}
    curve_by_layer = {"core": portfolio_config.core.path, "satellite": portfolio_config.satellite.path}
    trigger_signals, trigger_meta = _read_latest_trigger_signals(trigger_signal_path)
    resolved_market_cap_path = _resolve(project_root, stock_market_cap_path, DEFAULT_STOCK_MARKET_CAP_PATH)
    market_cap_frame, market_cap_payload = _load_market_cap_cache_payload(resolved_market_cap_path)
    snapshot_closes = _market_snapshot_close_lookup(market_snapshot, latest_date)

    for layer, curve_path in curve_by_layer.items():
        layer_target = target_by_layer.get(layer, {})
        layer_target_weight = _as_float(layer_target.get("target_weight"), 0.0)
        layer_target_value = _as_float(layer_target.get("target_value"), 0.0)
        backtest_dir, curve_latest_date, source_method = _resolve_backtest_dir(project_root, curve_path)
        trades_path = backtest_dir / "trades.csv" if backtest_dir is not None else None
        positions = _open_positions_from_trades(trades_path, latest_date) if trades_path is not None else pd.DataFrame()
        strategy_info = _strategy_explanation(backtest_dir)
        source_layers.append(
            {
                "layer": layer,
                "curve_path": str(curve_path),
                "curve_latest_date": curve_latest_date,
                "source_method": source_method,
                "backtest_dir": str(backtest_dir) if backtest_dir is not None else None,
                "trades_path": str(trades_path) if trades_path is not None else None,
                "open_position_count": int(len(positions)),
                "layer_target_weight": layer_target_weight,
                "source_project_name": strategy_info["project_name"],
                "source_strategy_name": strategy_info["strategy_name"],
                "source_strategy_profile": strategy_info["strategy_profile"],
                "source_selection_explanation": strategy_info["selection_explanation"],
                "source_filter_explanation": strategy_info["source_filter_explanation"],
                "source_risk_explanation": strategy_info["source_risk_explanation"],
                "source_config_path": strategy_info["config_path"],
                "source_config_status": strategy_info["config_status"],
            }
        )
        if positions.empty:
            continue
        valued_rows: list[dict[str, Any]] = []
        for position in positions.to_dict(orient="records"):
            code = _clean_code(position.get("code"))
            close_price, close_date, price_source = snapshot_closes.get(code, (None, None, ""))
            if close_price is None:
                close_price, close_date, price_source = _latest_stock_close(project_root, code, latest_date)
            if close_price is None:
                close_price = _as_float(position.get("last_trade_price"), 0.0)
                close_date = position.get("last_trade_date")
                price_source = "latest_trade_price"
            model_quantity = _as_float(position.get("quantity"), 0.0)
            model_market_value = model_quantity * close_price
            valued_rows.append(
                {
                    **position,
                    "close_price": close_price,
                    "price_date": close_date,
                    "price_source": price_source,
                    "model_market_value": model_market_value,
                }
            )
        layer_model_value = sum(_as_float(item.get("model_market_value"), 0.0) for item in valued_rows)
        for item in valued_rows:
            internal_weight = _as_float(item.get("model_market_value"), 0.0) / layer_model_value if layer_model_value > 0 else 0.0
            portfolio_weight = layer_target_weight * internal_weight
            portfolio_value = layer_target_value * internal_weight
            code = str(item["code"]).zfill(6)
            avg_cost = _as_float(item.get("avg_cost"), 0.0)
            close_price = _as_float(item.get("close_price"), 0.0)
            unrealized_return = close_price / avg_cost - 1.0 if avg_cost > 0 and close_price > 0 else np.nan
            last_trade_date = pd.to_datetime(item.get("last_trade_date"), errors="coerce")
            latest_ts = pd.to_datetime(latest_date, errors="coerce")
            holding_days = int((pd.Timestamp(latest_ts) - pd.Timestamp(last_trade_date)).days) if pd.notna(latest_ts) and pd.notna(last_trade_date) else np.nan
            target_action = "target_hold" if layer_target_weight > 1e-12 else "suppressed_by_layer_weight"
            if target_action == "suppressed_by_layer_weight":
                risk_filter_status = "blocked_by_layer_weight"
                target_explanation = "Layer target weight is 0%, so the source position is listed but receives no portfolio allocation."
            elif strategy_info["source_risk_enabled"]:
                risk_filter_status = "active_with_source_risk_model"
                target_explanation = "Portfolio target equals current layer weight multiplied by source-model internal position weight."
            else:
                risk_filter_status = "active_without_source_risk_model"
                target_explanation = "Portfolio target equals current layer weight multiplied by source-model internal position weight."
            trigger = trigger_signals.get(code)
            trigger_validity = _trigger_signal_validity(trigger, latest_date)
            rows.append(
                {
                    "date": _format_date(latest_date),
                    "layer": layer,
                    "code": code,
                    "name": str(item["name"]),
                    "layer_target_weight": layer_target_weight,
                    "layer_internal_weight": internal_weight,
                    "portfolio_target_weight": portfolio_weight,
                    "portfolio_target_value": portfolio_value,
                    "model_quantity": _as_float(item.get("quantity"), 0.0),
                    "model_market_value": _as_float(item.get("model_market_value"), 0.0),
                    "close_price": _as_float(item.get("close_price"), 0.0),
                    "price_date": item.get("price_date"),
                    "price_source": item.get("price_source"),
                    "avg_cost": avg_cost,
                    "unrealized_return": unrealized_return,
                    "holding_days": holding_days,
                    "last_trade_price": _as_float(item.get("last_trade_price"), 0.0),
                    "last_trade_date": item.get("last_trade_date"),
                    "target_action": target_action,
                    "target_explanation": target_explanation,
                    "risk_filter_status": risk_filter_status,
                    "selection_explanation": strategy_info["selection_explanation"],
                    "source_filter_explanation": strategy_info["source_filter_explanation"],
                    "source_risk_explanation": strategy_info["source_risk_explanation"],
                    "source_strategy_profile": strategy_info["strategy_profile"],
                    "source_strategy_name": strategy_info["strategy_name"],
                    "source_project_name": strategy_info["project_name"],
                    "factor_weights": strategy_info["factor_weights_json"],
                    "factor_weights_text": strategy_info["factor_weights_text"],
                    "source_config_path": strategy_info["config_path"],
                    "trigger_monitor_status": "matched_latest_trigger" if trigger else "not_in_latest_trigger_report",
                    "trigger_summary": _trigger_summary(trigger),
                    "trigger_run_time": trigger.get("trigger_run_time") if trigger else None,
                    "trigger_signal_age_days": trigger_validity["trigger_signal_age_days"],
                    "trigger_signal_validity_status": trigger_validity["trigger_signal_validity_status"],
                    "trigger_signal_type": trigger.get("trigger_signal_type") if trigger else None,
                    "trigger_action": trigger.get("trigger_action") if trigger else None,
                    "trigger_reason": trigger.get("trigger_reason") if trigger else None,
                    "trigger_score": trigger.get("trigger_score") if trigger else np.nan,
                    "trigger_score_level": trigger.get("trigger_score_level") if trigger else None,
                    "trigger_pct": trigger.get("trigger_pct") if trigger else np.nan,
                    "trigger_last": trigger.get("trigger_last") if trigger else np.nan,
                    "source_backtest_dir": str(backtest_dir) if backtest_dir is not None else None,
                    "source_trades_path": str(trades_path) if trades_path is not None else None,
                    "execution_gate_action": "manual_review_only",
                }
            )

    external_fresh_trigger_count = 0
    existing_codes = {_clean_code(row.get("code")) for row in rows}
    for code, trigger in sorted(trigger_signals.items()):
        clean_code = _clean_code(code)
        if not clean_code or clean_code in existing_codes:
            continue
        trigger_validity = _trigger_signal_validity(trigger, latest_date)
        if trigger_validity["trigger_signal_validity_status"] != "fresh_trigger_signal":
            continue
        external_fresh_trigger_count += 1
        rows.append(
            {
                "date": _format_date(latest_date),
                "layer": "trigger",
                "code": clean_code,
                "name": trigger.get("trigger_name") or clean_code,
                "layer_target_weight": 0.0,
                "layer_internal_weight": 0.0,
                "portfolio_target_weight": 0.0,
                "portfolio_target_value": 0.0,
                "model_quantity": 0.0,
                "model_market_value": 0.0,
                "close_price": trigger.get("trigger_last") if trigger.get("trigger_last") is not None else np.nan,
                "price_date": None,
                "price_source": "latest_trigger_signal",
                "avg_cost": np.nan,
                "unrealized_return": np.nan,
                "holding_days": np.nan,
                "last_trade_price": np.nan,
                "last_trade_date": None,
                "target_action": "trigger_watch_candidate",
                "target_explanation": "Fresh trigger-monitor candidate is not in the current model portfolio; listed for research review only.",
                "risk_filter_status": "not_in_current_portfolio_target",
                "selection_explanation": "External trigger-monitor candidate added with zero portfolio weight.",
                "source_filter_explanation": "Not selected by current portfolio source.",
                "source_risk_explanation": "No broker action; candidate-review row only.",
                "source_strategy_profile": "trigger_monitor_candidate",
                "source_strategy_name": "trigger_monitor",
                "source_project_name": "trigger_monitor",
                "factor_weights": "{}",
                "factor_weights_text": "",
                "source_config_path": None,
                "trigger_monitor_status": "matched_latest_trigger",
                "trigger_summary": _trigger_summary(trigger),
                "trigger_run_time": trigger.get("trigger_run_time"),
                "trigger_signal_age_days": trigger_validity["trigger_signal_age_days"],
                "trigger_signal_validity_status": trigger_validity["trigger_signal_validity_status"],
                "trigger_signal_type": trigger.get("trigger_signal_type"),
                "trigger_action": trigger.get("trigger_action"),
                "trigger_reason": trigger.get("trigger_reason"),
                "trigger_score": trigger.get("trigger_score"),
                "trigger_score_level": trigger.get("trigger_score_level"),
                "trigger_pct": trigger.get("trigger_pct"),
                "trigger_last": trigger.get("trigger_last"),
                "source_backtest_dir": None,
                "source_trades_path": None,
                "execution_gate_action": "usable_for_candidate_review",
            }
        )

    frame = pd.DataFrame(rows, columns=STOCK_TARGET_COLUMNS)
    if not frame.empty:
        frame = frame.sort_values(["layer", "portfolio_target_weight", "code"], ascending=[True, False, True]).reset_index(drop=True)
    frame, tracking_payload = apply_stock_market_cap_tracking_rule(
        frame,
        market_cap_frame,
        max_market_cap_yi=stock_tracking_max_market_cap_yi,
        market_cap_cache_payload=market_cap_payload,
    )
    trigger_match_count = (
        int((frame.get("trigger_monitor_status") == "matched_latest_trigger").sum()) if not frame.empty else 0
    )
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "status": "ok" if source_layers else "missing_sources",
        "latest_date": _format_date(latest_date),
        "stock_target_count": int(len(frame)),
        "active_stock_target_count": int((pd.to_numeric(frame.get("portfolio_target_weight", 0), errors="coerce") > 0).sum()) if not frame.empty else 0,
        "suppressed_stock_count": int((frame.get("target_action") == "suppressed_by_layer_weight").sum()) if not frame.empty else 0,
        "total_portfolio_target_weight": float(pd.to_numeric(frame.get("portfolio_target_weight", 0), errors="coerce").sum()) if not frame.empty else 0.0,
        "source_layers": source_layers,
        "latest_trigger_signal_path": trigger_meta["latest_trigger_signal_path"],
        "trigger_signal_status": trigger_meta["trigger_signal_status"],
        "trigger_signal_count": trigger_meta["trigger_signal_count"],
        "latest_trigger_run_time": trigger_meta["latest_trigger_run_time"],
        "trigger_match_count": trigger_match_count,
        "external_fresh_trigger_candidate_count": external_fresh_trigger_count,
        **tracking_payload,
        "stock_targets": _json_records(frame),
        "research_only": True,
        "broker_action": "none",
        "note": "Stock-level target decomposition from local backtest trades; this is not a broker order list.",
    }
    return frame, payload


def _stock_review_assessment(
    row: dict[str, Any],
    drawdown_threshold: float = -0.10,
    watch_drawdown_threshold: float = -0.07,
    loss_attention_threshold: float = -0.05,
    gain_attention_threshold: float = 0.10,
    watch_score_threshold: float = 30.0,
) -> dict[str, Any]:
    drawdown_threshold = _negative_threshold(drawdown_threshold, -0.10)
    watch_drawdown_threshold = _negative_threshold(watch_drawdown_threshold, -0.07)
    loss_attention_threshold = _negative_threshold(loss_attention_threshold, -0.05)
    gain_attention_threshold = abs(_as_float(gain_attention_threshold, 0.10))
    score = 0.0
    reasons: list[str] = []
    portfolio_weight = _as_float(row.get("portfolio_target_weight"), 0.0)
    unrealized_return = _optional_float(row.get("unrealized_return"))
    holding_days = _optional_float(row.get("holding_days"))
    target_action = _safe_text(row.get("target_action"))
    trigger_status = _safe_text(row.get("trigger_monitor_status"))
    risk_status = _safe_text(row.get("risk_filter_status"))

    if trigger_status == "matched_latest_trigger":
        score += 45.0
        reasons.append("matched latest trigger-monitor signal")

    if target_action == "suppressed_by_layer_weight" or risk_status == "blocked_by_layer_weight":
        score += 35.0
        reasons.append("source position is blocked by current layer weight")

    if unrealized_return is not None:
        if unrealized_return <= drawdown_threshold:
            score += 35.0
            reasons.append(f"unrealized loss <= {_pct(drawdown_threshold)}")
        elif unrealized_return <= watch_drawdown_threshold:
            score += 25.0
            reasons.append(f"unrealized loss <= {_pct(watch_drawdown_threshold)}")
        elif unrealized_return <= loss_attention_threshold:
            score += 15.0
            reasons.append(f"unrealized loss <= {_pct(loss_attention_threshold)}")
        elif unrealized_return >= gain_attention_threshold:
            score += 8.0
            reasons.append(f"unrealized gain >= {_pct(gain_attention_threshold)}, review trailing risk")

    if portfolio_weight >= 0.07:
        score += 15.0
        reasons.append("larger target weight >= 7%")
    elif portfolio_weight >= 0.065:
        score += 8.0
        reasons.append("target weight near equal-weight top sleeve")

    if holding_days is not None and holding_days <= 1:
        score += 5.0
        reasons.append("recent source trade")

    score = min(score, 100.0)
    if trigger_status == "matched_latest_trigger":
        bucket = "trigger_review"
        stage = "review_required"
        recommended = "Review latest trigger context before relying on this target; no broker order is generated."
    elif unrealized_return is not None and unrealized_return <= drawdown_threshold:
        bucket = "drawdown_review"
        stage = "review_required"
        recommended = "Review loss source, source risk gate, and next-session gap risk; no broker order is generated."
    elif target_action == "suppressed_by_layer_weight" or risk_status == "blocked_by_layer_weight":
        bucket = "suppressed_layer_review"
        stage = "monitor"
        recommended = "Keep visible for audit: source position exists, but current portfolio layer target is 0%."
    elif score >= watch_score_threshold:
        bucket = "watch_review"
        stage = "monitor"
        recommended = "Routine watch with extra attention to weight, loss state, or recent source trade."
    else:
        bucket = "routine_review"
        stage = "routine"
        recommended = "Routine model-target review only."

    return {
        "review_priority_score": round(score, 2),
        "review_bucket": bucket,
        "review_stage": stage,
        "review_reason": "; ".join(reasons) if reasons else "no elevated review signal",
        "recommended_review": recommended,
    }


def build_stock_target_review(
    stock_targets: pd.DataFrame,
    stock_targets_payload: dict[str, Any],
    drawdown_threshold: float = -0.10,
    watch_drawdown_threshold: float = -0.07,
    loss_attention_threshold: float = -0.05,
    gain_attention_threshold: float = 0.10,
    watch_score_threshold: float = 30.0,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    review_columns = [
        "date",
        "review_rank",
        "review_bucket",
        "review_stage",
        "review_priority_score",
        "layer",
        "code",
        "name",
        "portfolio_target_weight",
        "portfolio_target_value",
        "market_cap_yi",
        "tracking_rule_status",
        "unrealized_return",
        "holding_days",
        "target_action",
        "risk_filter_status",
        "trigger_monitor_status",
        "trigger_summary",
        "source_strategy_profile",
        "review_reason",
        "recommended_review",
        "broker_action",
        "research_only",
    ]
    drawdown_threshold = _negative_threshold(drawdown_threshold, -0.10)
    watch_drawdown_threshold = _negative_threshold(watch_drawdown_threshold, -0.07)
    loss_attention_threshold = _negative_threshold(loss_attention_threshold, -0.05)
    gain_attention_threshold = abs(_as_float(gain_attention_threshold, 0.10))
    watch_score_threshold = _as_float(watch_score_threshold, 30.0)
    rows: list[dict[str, Any]] = []
    for raw in stock_targets.to_dict(orient="records"):
        if _boolish(raw.get("tracking_excluded")):
            continue
        assessment = _stock_review_assessment(
            raw,
            drawdown_threshold=drawdown_threshold,
            watch_drawdown_threshold=watch_drawdown_threshold,
            loss_attention_threshold=loss_attention_threshold,
            gain_attention_threshold=gain_attention_threshold,
            watch_score_threshold=watch_score_threshold,
        )
        rows.append(
            {
                "date": raw.get("date"),
                "review_rank": 0,
                **assessment,
                "layer": raw.get("layer"),
                "code": _clean_code(raw.get("code")),
                "name": raw.get("name"),
                "portfolio_target_weight": _as_float(raw.get("portfolio_target_weight"), 0.0),
                "portfolio_target_value": _as_float(raw.get("portfolio_target_value"), 0.0),
                "market_cap_yi": _optional_float(raw.get("market_cap_yi")),
                "tracking_rule_status": raw.get("tracking_rule_status"),
                "unrealized_return": _optional_float(raw.get("unrealized_return")),
                "holding_days": _optional_float(raw.get("holding_days")),
                "target_action": raw.get("target_action"),
                "risk_filter_status": raw.get("risk_filter_status"),
                "trigger_monitor_status": raw.get("trigger_monitor_status"),
                "trigger_summary": raw.get("trigger_summary"),
                "source_strategy_profile": raw.get("source_strategy_profile"),
                "review_reason": assessment["review_reason"],
                "recommended_review": assessment["recommended_review"],
                "broker_action": "none",
                "research_only": True,
            }
        )
    frame = pd.DataFrame(rows, columns=review_columns)
    if not frame.empty:
        frame = frame.sort_values(
            ["review_priority_score", "portfolio_target_weight", "code"],
            ascending=[False, False, True],
        ).reset_index(drop=True)
        frame["review_rank"] = range(1, len(frame) + 1)

    stage_counts = frame["review_stage"].value_counts().to_dict() if not frame.empty else {}
    bucket_counts = frame["review_bucket"].value_counts().to_dict() if not frame.empty else {}
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "status": "ok" if stock_targets_payload.get("status") == "ok" else stock_targets_payload.get("status", "missing_sources"),
        "latest_date": stock_targets_payload.get("latest_date"),
        "review_row_count": int(len(frame)),
        "review_required_count": int(stage_counts.get("review_required", 0)),
        "monitor_count": int(stage_counts.get("monitor", 0)),
        "routine_count": int(stage_counts.get("routine", 0)),
        "trigger_review_count": int(bucket_counts.get("trigger_review", 0)),
        "drawdown_review_count": int(bucket_counts.get("drawdown_review", 0)),
        "suppressed_layer_review_count": int(bucket_counts.get("suppressed_layer_review", 0)),
        "watch_review_count": int(bucket_counts.get("watch_review", 0)),
        "routine_review_count": int(bucket_counts.get("routine_review", 0)),
        "drawdown_threshold": drawdown_threshold,
        "watch_drawdown_threshold": watch_drawdown_threshold,
        "loss_attention_threshold": loss_attention_threshold,
        "gain_attention_threshold": gain_attention_threshold,
        "watch_score_threshold": watch_score_threshold,
        "top_review_items": _json_records(frame.head(5)),
        "source_stock_target_count": stock_targets_payload.get("stock_target_count", 0),
        "review_skipped_tracking_excluded_count": stock_targets_payload.get("stock_tracking_excluded_large_market_cap_count", 0),
        "stock_tracking_excluded_large_market_cap_count": stock_targets_payload.get("stock_tracking_excluded_large_market_cap_count", 0),
        "stock_tracking_max_market_cap_yi": stock_targets_payload.get("stock_tracking_max_market_cap_yi"),
        "stock_tracking_market_cap_missing_count": stock_targets_payload.get("stock_tracking_market_cap_missing_count", 0),
        "stock_market_cap_cache_status": stock_targets_payload.get("stock_market_cap_cache_status"),
        "stock_market_cap_path": stock_targets_payload.get("stock_market_cap_path"),
        "trigger_match_count": stock_targets_payload.get("trigger_match_count", 0),
        "latest_trigger_run_time": stock_targets_payload.get("latest_trigger_run_time"),
        "research_only": True,
        "broker_action": "none",
        "note": "Review priority panel derived from stock_targets; it does not change model targets or generate broker orders.",
    }
    return frame, payload


def apply_review_required_observation_exclusion(
    stock_targets: pd.DataFrame,
    stock_targets_payload: dict[str, Any],
    review: pd.DataFrame,
    review_payload: dict[str, Any],
) -> tuple[pd.DataFrame, dict[str, Any], pd.DataFrame, dict[str, Any]]:
    """Remove review-required rows from observation targets while retaining audit evidence."""
    targets = stock_targets.copy()
    reviews = review.copy()
    target_payload = dict(stock_targets_payload)
    updated_review_payload = dict(review_payload)

    if reviews.empty or "review_stage" not in reviews.columns:
        target_payload.update(
            {
                "source_stock_target_count": int(len(targets)),
                "review_required_excluded_count": 0,
                "review_required_excluded_codes": [],
                "stock_targets": _json_records(targets),
            }
        )
        updated_review_payload["excluded_review_required_count"] = 0
        return targets, target_payload, reviews, updated_review_payload

    reviews["original_review_stage"] = reviews["review_stage"]
    reviews["observation_excluded"] = False
    reviews["observation_exclusion_reason"] = ""
    required_mask = reviews["review_stage"].astype(str).eq("review_required")
    excluded_keys = {
        (_safe_text(row.get("layer")), _clean_code(row.get("code")))
        for row in reviews.loc[required_mask].to_dict(orient="records")
    }
    excluded_codes = sorted({code for _, code in excluded_keys if code})

    if targets.empty:
        target_mask = pd.Series(False, index=targets.index, dtype=bool)
    else:
        target_mask = pd.Series(
            [
                (_safe_text(row.get("layer")), _clean_code(row.get("code"))) in excluded_keys
                for row in targets.to_dict(orient="records")
            ],
            index=targets.index,
            dtype=bool,
        )
    observation_targets = targets.loc[~target_mask].copy().reset_index(drop=True)

    reviews.loc[required_mask, "review_stage"] = "excluded"
    reviews.loc[required_mask, "observation_excluded"] = True
    reviews.loc[required_mask, "observation_exclusion_reason"] = (
        "Hard rule: review_required rows are excluded from all observation targets."
    )
    reviews.loc[required_mask, "recommended_review"] = (
        "Automatically excluded from observation targets; retained only for audit history."
    )

    active_count = (
        int((pd.to_numeric(observation_targets.get("portfolio_target_weight", 0), errors="coerce") > 0).sum())
        if not observation_targets.empty
        else 0
    )
    suppressed_count = (
        int((observation_targets.get("target_action") == "suppressed_by_layer_weight").sum())
        if not observation_targets.empty
        else 0
    )
    trigger_match_count = (
        int((observation_targets.get("trigger_monitor_status") == "matched_latest_trigger").sum())
        if not observation_targets.empty
        else 0
    )
    tracking_allowed_count = (
        int((~observation_targets.get("tracking_excluded", pd.Series(False, index=observation_targets.index)).map(_boolish)).sum())
        if not observation_targets.empty
        else 0
    )
    target_payload.update(
        {
            "source_stock_target_count": int(len(targets)),
            "stock_target_count": int(len(observation_targets)),
            "active_stock_target_count": active_count,
            "suppressed_stock_count": suppressed_count,
            "total_portfolio_target_weight": (
                float(pd.to_numeric(observation_targets.get("portfolio_target_weight", 0), errors="coerce").sum())
                if not observation_targets.empty
                else 0.0
            ),
            "trigger_match_count": trigger_match_count,
            "external_fresh_trigger_candidate_count": int(
                (observation_targets.get("layer") == "trigger").sum()
            )
            if not observation_targets.empty
            else 0,
            "stock_tracking_allowed_count": tracking_allowed_count,
            "review_required_excluded_count": int(required_mask.sum()),
            "review_required_excluded_codes": excluded_codes,
            "review_required_exclusion_rule": "exclude_all_review_required_from_observation_targets",
            "stock_targets": _json_records(observation_targets),
            "note": "Review-required rows are removed from observation targets and retained only in the review audit.",
        }
    )

    stage_counts = reviews["review_stage"].value_counts().to_dict()
    active_reviews = reviews.loc[~reviews["observation_excluded"].map(_boolish)]
    active_bucket_counts = active_reviews["review_bucket"].value_counts().to_dict() if "review_bucket" in active_reviews else {}
    updated_review_payload.update(
        {
            "review_required_count": int(stage_counts.get("review_required", 0)),
            "monitor_count": int(stage_counts.get("monitor", 0)),
            "routine_count": int(stage_counts.get("routine", 0)),
            "excluded_count": int(stage_counts.get("excluded", 0)),
            "original_trigger_review_count": int(updated_review_payload.get("trigger_review_count", 0)),
            "original_drawdown_review_count": int(updated_review_payload.get("drawdown_review_count", 0)),
            "trigger_review_count": int(active_bucket_counts.get("trigger_review", 0)),
            "drawdown_review_count": int(active_bucket_counts.get("drawdown_review", 0)),
            "original_review_required_count": int(required_mask.sum()),
            "excluded_review_required_count": int(required_mask.sum()),
            "excluded_review_required_codes": excluded_codes,
            "top_review_items": _json_records(reviews.head(5)),
            "note": "Review-required rows are automatically excluded from observation targets; audit rows remain visible.",
        }
    )
    return observation_targets, target_payload, reviews, updated_review_payload


def _stock_review_note_columns(frame: pd.DataFrame) -> list[str]:
    extras = [column for column in frame.columns if column not in STOCK_TARGET_REVIEW_NOTE_COLUMNS]
    return STOCK_TARGET_REVIEW_NOTE_COLUMNS + extras


def _load_stock_target_review_notes(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=STOCK_TARGET_REVIEW_NOTE_COLUMNS)
    try:
        frame = pd.read_csv(path, dtype={"code": str})
    except (OSError, pd.errors.EmptyDataError):
        return pd.DataFrame(columns=STOCK_TARGET_REVIEW_NOTE_COLUMNS)
    for column in STOCK_TARGET_REVIEW_NOTE_COLUMNS:
        if column not in frame.columns:
            frame[column] = ""
    frame["code"] = frame["code"].map(_clean_code)
    frame["layer"] = frame["layer"].map(lambda value: _safe_text(value))
    return frame[_stock_review_note_columns(frame)]


def _stock_review_note_key(row: dict[str, Any]) -> tuple[str, str]:
    return (_safe_text(row.get("layer")), _clean_code(row.get("code")))


def _has_manual_review(record: dict[str, Any]) -> bool:
    manual_status = _normalize_manual_status(record.get("manual_status"))
    manual_note = _safe_text(record.get("manual_note"))
    return bool(manual_note or (manual_status and manual_status != "unreviewed"))


def _normalize_manual_status(value: Any) -> str:
    text = _safe_text(value).strip()
    key = text.lower()
    return MANUAL_STATUS_ALIASES.get(key, MANUAL_STATUS_ALIASES.get(text, "other"))


def _manual_review_state(normalized_status: str) -> str:
    if normalized_status == "unreviewed":
        return "pending_review"
    if normalized_status == "watch":
        return "reviewed_watch"
    if normalized_status == "resolved":
        return "reviewed_resolved"
    if normalized_status == "exclude_candidate":
        return "reviewed_exclude_candidate"
    if normalized_status == "reviewed":
        return "reviewed_closed"
    return "reviewed_other"


def _manual_followup_required(normalized_status: str) -> bool:
    return normalized_status in {"unreviewed", "watch", "exclude_candidate", "other"}


def sync_stock_target_review_notes(
    review: pd.DataFrame,
    notes_path: str | Path,
    notes_snapshot_path: str | Path,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Merge current model review rows into a persistent manual-review note sheet."""
    resolved_notes = Path(notes_path)
    resolved_snapshot = Path(notes_snapshot_path)
    notes = _load_stock_target_review_notes(resolved_notes)
    records = notes.to_dict(orient="records")
    index_by_key: dict[tuple[str, str], int] = {}
    for index, record in enumerate(records):
        key = _stock_review_note_key(record)
        if key[1]:
            index_by_key[key] = index

    review_with_notes = review.copy()
    note_statuses: list[str] = []
    manual_statuses: list[str] = []
    manual_review_states: list[str] = []
    manual_followups: list[bool] = []
    note_records_for_review: list[dict[str, Any]] = []
    for row in review.to_dict(orient="records"):
        key = _stock_review_note_key(row)
        if key in index_by_key:
            record = records[index_by_key[key]]
        else:
            record = {column: "" for column in STOCK_TARGET_REVIEW_NOTE_COLUMNS}
            records.append(record)
            index_by_key[key] = len(records) - 1

        current_date = _format_date(row.get("date"))
        record["layer"] = key[0]
        record["code"] = key[1]
        record["name"] = _safe_text(row.get("name"))
        record["first_seen_date"] = _safe_text(record.get("first_seen_date"), current_date)
        record["last_seen_date"] = current_date
        record["last_review_bucket"] = _safe_text(row.get("review_bucket"))
        record["last_review_stage"] = _safe_text(row.get("review_stage"))
        record["last_review_priority_score"] = _as_float(row.get("review_priority_score"), 0.0)
        record["last_model_reason"] = _safe_text(row.get("review_reason"))
        record["broker_action"] = "none"
        record["research_only"] = True
        if not _safe_text(record.get("manual_status")):
            record["manual_status"] = "unreviewed"
        for column in STOCK_TARGET_REVIEW_MANUAL_COLUMNS:
            record[column] = _safe_text(record.get(column))

        normalized_status = _normalize_manual_status(record.get("manual_status"))
        manual_state = _manual_review_state(normalized_status)
        notes_status = "manual_reviewed" if _has_manual_review(record) else "unreviewed"
        note_statuses.append(notes_status)
        manual_statuses.append(normalized_status)
        manual_review_states.append(manual_state)
        manual_followups.append(_manual_followup_required(normalized_status))
        note_records_for_review.append(record.copy())

    if records:
        updated_notes = pd.DataFrame(records)
    else:
        updated_notes = pd.DataFrame(columns=STOCK_TARGET_REVIEW_NOTE_COLUMNS)
    for column in STOCK_TARGET_REVIEW_NOTE_COLUMNS:
        if column not in updated_notes.columns:
            updated_notes[column] = ""
    updated_notes = updated_notes[_stock_review_note_columns(updated_notes)]

    if not review_with_notes.empty:
        for column in STOCK_TARGET_REVIEW_MANUAL_COLUMNS:
            review_with_notes[column] = [_safe_text(record.get(column)) for record in note_records_for_review]
        review_with_notes["notes_first_seen_date"] = [
            _safe_text(record.get("first_seen_date")) for record in note_records_for_review
        ]
        review_with_notes["notes_last_seen_date"] = [
            _safe_text(record.get("last_seen_date")) for record in note_records_for_review
        ]
        review_with_notes["notes_status"] = note_statuses
        review_with_notes["manual_status_normalized"] = manual_statuses
        review_with_notes["manual_review_state"] = manual_review_states
        review_with_notes["manual_status_explanation"] = [
            MANUAL_STATUS_EXPLANATIONS.get(status, MANUAL_STATUS_EXPLANATIONS["other"]) for status in manual_statuses
        ]
        review_with_notes["manual_followup_required"] = manual_followups
    else:
        for column in STOCK_TARGET_REVIEW_MANUAL_COLUMNS + [
            "notes_first_seen_date",
            "notes_last_seen_date",
            "notes_status",
            "manual_status_normalized",
            "manual_review_state",
            "manual_status_explanation",
            "manual_followup_required",
        ]:
            review_with_notes[column] = []

    resolved_notes.parent.mkdir(parents=True, exist_ok=True)
    resolved_snapshot.parent.mkdir(parents=True, exist_ok=True)
    updated_notes.to_csv(resolved_notes, index=False, encoding="utf-8")
    updated_notes.to_csv(resolved_snapshot, index=False, encoding="utf-8")

    manual_note_count = int(sum(status == "manual_reviewed" for status in note_statuses))
    unreviewed_count = int(sum(status == "unreviewed" for status in note_statuses))
    manual_status_counts = pd.Series(manual_statuses).value_counts().to_dict() if manual_statuses else {}
    review_required_unreviewed_count = (
        int(((review_with_notes.get("review_stage") == "review_required") & (review_with_notes.get("notes_status") == "unreviewed")).sum())
        if not review_with_notes.empty
        else 0
    )
    review_required_pending_count = (
        int(((review_with_notes.get("review_stage") == "review_required") & (review_with_notes.get("manual_status_normalized") == "unreviewed")).sum())
        if not review_with_notes.empty
        else 0
    )
    review_required_reviewed_count = (
        int(((review_with_notes.get("review_stage") == "review_required") & (review_with_notes.get("manual_status_normalized") != "unreviewed")).sum())
        if not review_with_notes.empty
        else 0
    )
    payload = {
        "notes_path": str(resolved_notes),
        "notes_snapshot_path": str(resolved_snapshot),
        "notes_row_count": int(len(updated_notes)),
        "manual_note_count": manual_note_count,
        "unreviewed_count": unreviewed_count,
        "review_required_unreviewed_count": review_required_unreviewed_count,
        "review_required_pending_count": review_required_pending_count,
        "review_required_reviewed_count": review_required_reviewed_count,
        "manual_status_counts": manual_status_counts,
        "manual_pending_count": int(manual_status_counts.get("unreviewed", 0)),
        "manual_reviewed_count": int(manual_status_counts.get("reviewed", 0)),
        "manual_watch_count": int(manual_status_counts.get("watch", 0)),
        "manual_resolved_count": int(manual_status_counts.get("resolved", 0)),
        "manual_exclude_candidate_count": int(manual_status_counts.get("exclude_candidate", 0)),
        "manual_other_status_count": int(manual_status_counts.get("other", 0)),
        "notes_status_counts": review_with_notes["notes_status"].value_counts().to_dict()
        if not review_with_notes.empty
        else {},
        "notes_instruction": "Edit manual_status/manual_note/reviewed_at in the persistent notes CSV after human review; reruns preserve those fields.",
    }
    return review_with_notes, payload


def _action_from_review_row(row: dict[str, Any], latest_date: Any) -> dict[str, Any] | None:
    normalized_status = _safe_text(row.get("manual_status_normalized"), "unreviewed")
    review_stage = _safe_text(row.get("review_stage"))
    next_review_date = _safe_text(row.get("next_review_date"))
    latest_ts = pd.to_datetime(latest_date, errors="coerce")
    next_ts = pd.to_datetime(next_review_date, errors="coerce") if next_review_date else pd.NaT
    next_review_due = bool(pd.notna(next_ts) and pd.notna(latest_ts) and pd.Timestamp(next_ts).date() <= pd.Timestamp(latest_ts).date())

    if review_stage == "review_required" and normalized_status == "unreviewed":
        return {
            "action_code": "review_required_pending",
            "action_stage": "review_required",
            "action_priority_score": 100.0 + _as_float(row.get("review_priority_score"), 0.0) / 100.0,
            "action_reason": "Model row is review_required and has not been manually reviewed.",
            "recommended_review": "Open the review row, inspect trigger/drawdown context, then update manual_status/manual_note.",
        }
    if normalized_status == "exclude_candidate":
        return {
            "action_code": "manual_exclusion_candidate",
            "action_stage": "monitor",
            "action_priority_score": 80.0 + _as_float(row.get("review_priority_score"), 0.0) / 100.0,
            "action_reason": "Manual status marks this row as an exclusion candidate.",
            "recommended_review": "Review the exclusion rationale separately; this label does not change model targets.",
        }
    if normalized_status == "other":
        return {
            "action_code": "manual_status_normalization",
            "action_stage": "monitor",
            "action_priority_score": 70.0 + _as_float(row.get("review_priority_score"), 0.0) / 100.0,
            "action_reason": "Manual status is not recognized by the built-in review taxonomy.",
            "recommended_review": "Normalize manual_status to unreviewed, reviewed, watch, resolved, or exclude_candidate.",
        }
    if normalized_status == "watch":
        return {
            "action_code": "manual_watch_followup",
            "action_stage": "monitor",
            "action_priority_score": 60.0 + _as_float(row.get("review_priority_score"), 0.0) / 100.0,
            "action_reason": "Manual status asks to keep this row on watch.",
            "recommended_review": "Keep this row in routine follow-up and update next_review_date/manual_note as needed.",
        }
    if next_review_due and normalized_status not in {"resolved", "unreviewed"}:
        return {
            "action_code": "manual_next_review_due",
            "action_stage": "monitor",
            "action_priority_score": 55.0 + _as_float(row.get("review_priority_score"), 0.0) / 100.0,
            "action_reason": f"Manual next_review_date {next_review_date} is due by latest model date {_format_date(latest_date)}.",
            "recommended_review": "Refresh the manual note and move next_review_date forward if the follow-up remains open.",
        }
    return None


def build_stock_target_review_actions(
    review: pd.DataFrame,
    review_payload: dict[str, Any],
) -> tuple[pd.DataFrame, dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    latest_date = review_payload.get("latest_date")
    for raw in review.to_dict(orient="records"):
        action = _action_from_review_row(raw, latest_date)
        if action is None:
            continue
        rows.append(
            {
                "date": raw.get("date"),
                "action_rank": 0,
                **action,
                "source_review_rank": raw.get("review_rank"),
                "review_bucket": raw.get("review_bucket"),
                "review_stage": raw.get("review_stage"),
                "layer": raw.get("layer"),
                "code": _clean_code(raw.get("code")),
                "name": raw.get("name"),
                "portfolio_target_weight": _as_float(raw.get("portfolio_target_weight"), 0.0),
                "unrealized_return": _optional_float(raw.get("unrealized_return")),
                "manual_status": raw.get("manual_status"),
                "manual_status_normalized": raw.get("manual_status_normalized"),
                "manual_review_state": raw.get("manual_review_state"),
                "next_review_date": raw.get("next_review_date"),
                "broker_action": "none",
                "research_only": True,
            }
        )
    frame = pd.DataFrame(rows, columns=STOCK_TARGET_REVIEW_ACTION_COLUMNS)
    if not frame.empty:
        frame = frame.sort_values(
            ["action_priority_score", "source_review_rank", "code"],
            ascending=[False, True, True],
        ).reset_index(drop=True)
        frame["action_rank"] = range(1, len(frame) + 1)

    action_counts = frame["action_code"].value_counts().to_dict() if not frame.empty else {}
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "status": review_payload.get("status", "missing_sources"),
        "latest_date": latest_date,
        "action_count": int(len(frame)),
        "review_required_pending_count": int(action_counts.get("review_required_pending", 0)),
        "manual_watch_followup_count": int(action_counts.get("manual_watch_followup", 0)),
        "manual_exclusion_candidate_count": int(action_counts.get("manual_exclusion_candidate", 0)),
        "manual_status_normalization_count": int(action_counts.get("manual_status_normalization", 0)),
        "manual_next_review_due_count": int(action_counts.get("manual_next_review_due", 0)),
        "action_counts": action_counts,
        "top_actions": _json_records(frame.head(10)),
        "source_review_row_count": review_payload.get("review_row_count", 0),
        "notes_path": review_payload.get("notes_path"),
        "research_only": True,
        "broker_action": "none",
        "note": "Action queue is a manual review checklist derived from stock_target_review; it does not change model targets or place orders.",
    }
    return frame, payload


def _render_stock_target_review_actions_report(actions: pd.DataFrame, payload: dict[str, Any], output_dir: Path) -> str:
    if actions.empty:
        rows = "| N/A | `N/A` | `N/A` | `N/A` | N/A | N/A | `N/A` | No open stock-target review actions |"
    else:
        rows = "\n".join(
            "| {rank} | `{code}` | {name} | `{action}` | `{stage}` | {weight} | {ret} | {reason} |".format(
                rank=int(row["action_rank"]),
                code=str(row.get("code")).zfill(6),
                name=_md_text(row.get("name"), 40),
                action=_md_text(row.get("action_code"), 40),
                stage=_md_text(row.get("action_stage"), 30),
                weight=_pct(row.get("portfolio_target_weight")),
                ret=_pct(row.get("unrealized_return")),
                reason=_md_text(row.get("action_reason"), 120),
            )
            for _, row in actions.iterrows()
        )
    return f"""# Paper Stock Target Review Actions

Generated at: `{payload.get("generated_at")}`

This is a research-only manual action queue. It does not connect to brokers, place orders, change target weights, or provide investment advice.

## Summary

| Item | Value |
| --- | ---: |
| Status | `{payload.get("status")}` |
| Latest date | {payload.get("latest_date", "N/A")} |
| Open actions | {payload.get("action_count", 0)} |
| Pending model reviews | {payload.get("review_required_pending_count", 0)} |
| Manual watch follow-ups | {payload.get("manual_watch_followup_count", 0)} |
| Manual exclusion candidates | {payload.get("manual_exclusion_candidate_count", 0)} |
| Manual status normalization | {payload.get("manual_status_normalization_count", 0)} |
| Manual next-review due | {payload.get("manual_next_review_due_count", 0)} |
| Notes CSV | `{payload.get("notes_path", "N/A")}` |
| Broker action | `{payload.get("broker_action")}` |

## Action Queue

| Rank | Code | Name | Action | Stage | Target weight | Unrealized return | Reason |
| ---: | --- | --- | --- | --- | ---: | ---: | --- |
{rows}

## Files

- Actions CSV: `{output_dir / "stock_target_review_actions.csv"}`
- Actions JSON: `{output_dir / "stock_target_review_actions.json"}`
- Review CSV: `{output_dir / "stock_target_review.csv"}`
- Review notes CSV: `{payload.get("notes_path", "N/A")}`
"""


def _stock_price_history(
    project_root: Path,
    code: str,
    cache: dict[str, tuple[pd.DataFrame, str]] | None = None,
    persistent_cache_dir: Path | None = None,
) -> tuple[pd.DataFrame, str]:
    clean_code = _clean_code(code)
    if cache is not None and clean_code in cache:
        return cache[clean_code]
    candidates = [
        project_root / "data" / "processed" / "stocks" / f"{clean_code}.csv",
        project_root / "data" / "processed" / f"{clean_code}.csv",
    ]
    for path in candidates:
        if not path.exists():
            continue
        source_stat = path.stat()
        source_fingerprint = f"1:{path.resolve()}:{source_stat.st_size}:{source_stat.st_mtime_ns}"
        persistent_path = persistent_cache_dir / f"{clean_code}.npz" if persistent_cache_dir is not None else None
        if persistent_path is not None and persistent_path.exists():
            try:
                with np.load(persistent_path, allow_pickle=False) as stored:
                    if str(stored["source_fingerprint"].item()) == source_fingerprint:
                        data = pd.DataFrame(
                            {
                                "date": pd.to_datetime(stored["date_ns"], unit="ns"),
                                "close": stored["close"],
                                "high": stored["high"],
                                "low": stored["low"],
                            }
                        )
                    else:
                        data = pd.DataFrame()
            except (OSError, ValueError, KeyError):
                data = pd.DataFrame()
            if not data.empty:
                result = (data, str(path))
                if cache is not None:
                    cache[clean_code] = result
                return result
        try:
            frame = pd.read_csv(path)
        except (OSError, pd.errors.EmptyDataError):
            continue
        if frame.empty or not {"date", "close"}.issubset(frame.columns):
            continue
        data = frame[["date", "close"]].copy()
        data["high"] = frame["high"] if "high" in frame.columns else frame["close"]
        data["low"] = frame["low"] if "low" in frame.columns else frame["close"]
        data["date"] = pd.to_datetime(data["date"], errors="coerce")
        data["close"] = pd.to_numeric(data["close"], errors="coerce")
        data["high"] = pd.to_numeric(data["high"], errors="coerce").fillna(data["close"])
        data["low"] = pd.to_numeric(data["low"], errors="coerce").fillna(data["close"])
        data = data.dropna(subset=["date", "close"]).sort_values("date").drop_duplicates(subset=["date"], keep="last")
        if data.empty:
            continue
        result = (data.reset_index(drop=True), str(path))
        if persistent_path is not None:
            buffer = io.BytesIO()
            np.savez(
                buffer,
                source_fingerprint=np.asarray(source_fingerprint),
                date_ns=result[0]["date"].astype("int64").to_numpy(),
                close=result[0]["close"].to_numpy(dtype=float),
                high=result[0]["high"].to_numpy(dtype=float),
                low=result[0]["low"].to_numpy(dtype=float),
            )
            write_bytes_if_changed(persistent_path, buffer.getvalue())
        if cache is not None:
            cache[clean_code] = result
        return result
    result = (pd.DataFrame(columns=["date", "close", "high", "low"]), "missing_local_price_history")
    if cache is not None:
        cache[clean_code] = result
    return result


def _review_condition_text(
    latest_close: float | None,
    pressure_price: float | None,
    stop_loss_reference: float | None,
    review_bucket: Any,
    unrealized_return: Any,
) -> tuple[str, str, str]:
    bucket = _safe_text(review_bucket)
    unrealized = _optional_float(unrealized_return)
    pressure_text = f"{pressure_price:.2f}" if pressure_price is not None and pd.notna(pressure_price) else "N/A"
    stop_text = f"{stop_loss_reference:.2f}" if stop_loss_reference is not None and pd.notna(stop_loss_reference) else "N/A"
    close_text = f"{latest_close:.2f}" if latest_close is not None and pd.notna(latest_close) else "N/A"

    trigger = f"Review strength only if close reclaims local pressure {pressure_text}; last close {close_text}."
    invalidation = f"Review risk if close breaks stop-loss reference {stop_text}; no broker action is generated."
    posture = "observe"
    if bucket in {"drawdown_review", "watch_review"} or (unrealized is not None and unrealized <= -0.05):
        posture = "risk_review"
    elif unrealized is not None and unrealized >= 0.10:
        posture = "profit_protection_review"
    return trigger, invalidation, posture


def _stock_review_price_context(
    project_root: Path,
    code: str,
    review_date: Any,
    review_bucket: Any = None,
    unrealized_return: Any = None,
    price_history_cache: dict[str, tuple[pd.DataFrame, str]] | None = None,
    price_history_cache_dir: Path | None = None,
) -> dict[str, Any]:
    history, path = _stock_price_history(project_root, code, price_history_cache, price_history_cache_dir)
    fields: dict[str, Any] = {
        "latest_price_date": None,
        "latest_close": None,
        "latest_pct_change": None,
        "support_price": None,
        "pressure_price": None,
        "stop_loss_reference": None,
        "trigger_condition": "missing local price history",
        "invalidation_condition": "missing local price history",
        "suggested_review_posture": "data_review",
        "local_price_history_status": path if path == "missing_local_price_history" else "missing_local_price_history",
        "local_price_history_path": "" if path == "missing_local_price_history" else path,
        "price_staleness_days": None,
        "return_5d": None,
        "return_20d": None,
        "drawdown_from_20d_high": None,
    }
    if history.empty:
        return fields

    target_date = pd.to_datetime(review_date, errors="coerce")
    eligible = history[history["date"] <= pd.Timestamp(target_date)] if pd.notna(target_date) else history
    if eligible.empty:
        fields["local_price_history_status"] = "no_price_on_or_before_review_date"
        return fields

    latest_index = int(eligible.index[-1])
    latest = eligible.loc[latest_index]
    latest_date = pd.Timestamp(latest["date"])
    latest_close = float(latest["close"])
    fields["latest_price_date"] = latest_date.strftime("%Y-%m-%d")
    fields["latest_close"] = latest_close
    if latest_index >= 1:
        prev_close = float(history.iloc[latest_index - 1]["close"])
        fields["latest_pct_change"] = latest_close / prev_close - 1.0 if prev_close else None
    if pd.notna(target_date):
        fields["price_staleness_days"] = int((pd.Timestamp(target_date).normalize() - latest_date.normalize()).days)
    fields["local_price_history_status"] = "ok"
    if fields["price_staleness_days"] is not None and fields["price_staleness_days"] > 5:
        fields["local_price_history_status"] = "stale_local_price_history"

    if latest_index >= 5:
        prev_5 = float(history.iloc[latest_index - 5]["close"])
        fields["return_5d"] = latest_close / prev_5 - 1.0 if prev_5 else None
    if latest_index >= 20:
        prev_20 = float(history.iloc[latest_index - 20]["close"])
        fields["return_20d"] = latest_close / prev_20 - 1.0 if prev_20 else None
    window_20 = history.iloc[max(0, latest_index - 19) : latest_index + 1]
    high_20 = float(window_20["close"].max()) if not window_20.empty else 0.0
    fields["drawdown_from_20d_high"] = latest_close / high_20 - 1.0 if high_20 else None
    support_price = float(window_20["low"].min()) if not window_20.empty and "low" in window_20.columns else None
    pressure_price = float(window_20["high"].max()) if not window_20.empty and "high" in window_20.columns else None
    stop_loss_reference = support_price * 0.98 if support_price is not None and pd.notna(support_price) else None
    fields["support_price"] = support_price
    fields["pressure_price"] = pressure_price
    fields["stop_loss_reference"] = stop_loss_reference
    trigger, invalidation, posture = _review_condition_text(
        latest_close,
        pressure_price,
        stop_loss_reference,
        review_bucket=review_bucket,
        unrealized_return=unrealized_return,
    )
    fields["trigger_condition"] = trigger
    fields["invalidation_condition"] = invalidation
    fields["suggested_review_posture"] = posture
    return fields


def _assistant_manual_note_prompt(row: dict[str, Any]) -> str:
    action_code = _safe_text(row.get("action_code"))
    review_bucket = _safe_text(row.get("review_bucket"))
    if action_code == "review_required_pending" or review_bucket == "drawdown_review":
        return "Record whether the row is data issue, source-model drawdown, broken trend, or watch-only; add next_review_date if still open."
    if action_code == "manual_watch_followup":
        return "Refresh the watch reason and next_review_date, or mark resolved if the follow-up is no longer needed."
    if action_code == "manual_exclusion_candidate":
        return "Record the exclusion reason separately; model targets are not changed automatically."
    return "Record the human review conclusion, evidence checked, and whether another review date is needed."


def build_stock_target_review_assistant(
    project_root: str | Path,
    review: pd.DataFrame,
    actions: pd.DataFrame,
    review_payload: dict[str, Any],
    price_history_cache: dict[str, tuple[pd.DataFrame, str]] | None = None,
    price_history_cache_dir: Path | None = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Build a research-only assistant sheet for rows that need human review."""
    root = Path(project_root)
    review_lookup = {
        (_safe_text(row.get("layer")), _clean_code(row.get("code"))): row for row in review.to_dict(orient="records")
    }
    rows: list[dict[str, Any]] = []
    included_keys: set[tuple[str, str]] = set()

    def append_assistant_row(source: dict[str, Any]) -> None:
        key = (_safe_text(source.get("layer")), _clean_code(source.get("code")))
        included_keys.add(key)
        price_context = _stock_review_price_context(
            root,
            source.get("code"),
            source.get("date"),
            review_bucket=source.get("review_bucket"),
            unrealized_return=source.get("unrealized_return"),
            price_history_cache=price_history_cache,
            price_history_cache_dir=price_history_cache_dir,
        )
        evidence = [
            "review model reason",
            "check local price history freshness",
            "check support, pressure, and stop-loss reference",
            "check source cost and holding-days context",
            "check source risk and layer-weight gate",
            "write manual_status/manual_note after human review",
        ]
        if _safe_text(source.get("trigger_monitor_status")) == "matched_latest_trigger":
            evidence.insert(1, "check latest trigger-monitor context")
        assistant_rank = _as_float(source.get("action_rank"), 0.0)
        if assistant_rank <= 0:
            assistant_rank = _as_float(source.get("review_rank"), len(rows) + 1)
        rows.append(
            {
                "date": source.get("date"),
                "assistant_rank": int(assistant_rank),
                "action_code": source.get("action_code") or "manual_followup_pending",
                "review_bucket": source.get("review_bucket"),
                "review_stage": source.get("review_stage"),
                "layer": source.get("layer"),
                "code": _clean_code(source.get("code")),
                "name": source.get("name"),
                "portfolio_target_weight": _as_float(source.get("portfolio_target_weight"), 0.0),
                "unrealized_return": _optional_float(source.get("unrealized_return")),
                "holding_days": _optional_float(source.get("holding_days")),
                "review_priority_score": _as_float(source.get("review_priority_score"), 0.0),
                **price_context,
                "manual_status_options": "reviewed|watch|resolved|exclude_candidate|unreviewed",
                "manual_note_prompt": _assistant_manual_note_prompt(source),
                "evidence_checklist": "; ".join(evidence),
                "review_reason": source.get("review_reason") or source.get("action_reason"),
                "recommended_review": source.get("recommended_review"),
                "broker_action": "none",
                "research_only": True,
            }
        )

    for action in actions.to_dict(orient="records"):
        key = (_safe_text(action.get("layer")), _clean_code(action.get("code")))
        source = {**review_lookup.get(key, {}), **action}
        append_assistant_row(source)

    for review_row in review.to_dict(orient="records"):
        key = (_safe_text(review_row.get("layer")), _clean_code(review_row.get("code")))
        if key in included_keys or not _boolish(review_row.get("manual_followup_required")):
            continue
        source = {
            **review_row,
            "action_code": "manual_followup_pending",
            "action_reason": review_row.get("review_reason"),
        }
        append_assistant_row(source)

    frame = pd.DataFrame(rows, columns=STOCK_TARGET_REVIEW_ASSISTANT_COLUMNS)
    if not frame.empty:
        frame = frame.sort_values(["assistant_rank", "code"], ascending=[True, True]).reset_index(drop=True)
        frame["assistant_rank"] = range(1, len(frame) + 1)

    status_counts = frame["local_price_history_status"].value_counts().to_dict() if not frame.empty else {}
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "status": review_payload.get("status", "missing_sources"),
        "latest_date": review_payload.get("latest_date"),
        "assistant_row_count": int(len(frame)),
        "review_required_pending_count": int((frame.get("action_code") == "review_required_pending").sum()) if not frame.empty else 0,
        "price_history_ok_count": int(status_counts.get("ok", 0)),
        "price_history_stale_count": int(status_counts.get("stale_local_price_history", 0)),
        "price_history_missing_count": int(status_counts.get("missing_local_price_history", 0)),
        "price_history_status_counts": status_counts,
        "manual_status_options": ["reviewed", "watch", "resolved", "exclude_candidate", "unreviewed"],
        "notes_path": review_payload.get("notes_path"),
        "top_assistant_rows": _json_records(frame.head(10)),
        "research_only": True,
        "broker_action": "none",
        "note": "Assistant rows summarize local evidence for human review; they do not set manual_status, change model targets, or place orders.",
    }
    return frame, payload


def _render_stock_target_review_assistant_report(assistant: pd.DataFrame, payload: dict[str, Any], output_dir: Path) -> str:
    if assistant.empty:
        rows = "| N/A | `N/A` | `N/A` | `N/A` | N/A | N/A | N/A | N/A | N/A | N/A | N/A | N/A | `N/A` | No open assistant rows |"
    else:
        rows = "\n".join(
            "| {rank} | `{code}` | {name} | `{action}` | {weight} | {ret} | {close} | {pct} | {support} | {pressure} | {stop} | `{posture}` | `{price_status}` | {prompt} |".format(
                rank=int(row["assistant_rank"]),
                code=str(row.get("code")).zfill(6),
                name=_md_text(row.get("name"), 40),
                action=_md_text(row.get("action_code"), 40),
                weight=_pct(row.get("portfolio_target_weight")),
                ret=_maybe_pct(row.get("unrealized_return")),
                close=_num(row.get("latest_close")),
                pct=_maybe_pct(row.get("latest_pct_change")),
                support=_num(row.get("support_price")),
                pressure=_num(row.get("pressure_price")),
                stop=_num(row.get("stop_loss_reference")),
                posture=_md_text(row.get("suggested_review_posture"), 40),
                price_status=_md_text(row.get("local_price_history_status"), 40),
                prompt=_md_text(
                    f"{row.get('trigger_condition', '')} {row.get('invalidation_condition', '')} {row.get('manual_note_prompt', '')}",
                    220,
                ),
            )
            for _, row in assistant.iterrows()
        )
    return f"""# Paper Stock Target Review Assistant

Generated at: `{payload.get("generated_at")}`

This is a research-only evidence checklist for manual stock-target review. It does not connect to brokers, place orders, change target weights, set manual_status, or provide investment advice.

## Summary

| Item | Value |
| --- | ---: |
| Status | `{payload.get("status")}` |
| Latest date | {payload.get("latest_date", "N/A")} |
| Assistant rows | {payload.get("assistant_row_count", 0)} |
| Pending model reviews | {payload.get("review_required_pending_count", 0)} |
| Price history ok / stale / missing | {payload.get("price_history_ok_count", 0)} / {payload.get("price_history_stale_count", 0)} / {payload.get("price_history_missing_count", 0)} |
| Notes CSV | `{payload.get("notes_path", "N/A")}` |
| Broker action | `{payload.get("broker_action")}` |

## Review Assistant Queue

| Rank | Code | Name | Action | Target weight | Unrealized return | Last close | Day change | Support | Pressure | Stop-loss ref | Review posture | Price status | Conditions and note prompt |
| ---: | --- | --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | --- | --- | --- |
{rows}

## Manual Status Guide

- `reviewed`: human review has been recorded; keep row visible for audit.
- `watch`: keep the row in routine follow-up.
- `resolved`: current review item is closed for this audit loop.
- `exclude_candidate`: mark as an exclusion candidate; model targets are not changed automatically.
- `unreviewed`: no human review conclusion has been recorded yet.

## Files

- Assistant CSV: `{output_dir / "stock_target_review_assistant.csv"}`
- Assistant JSON: `{output_dir / "stock_target_review_assistant.json"}`
- Review actions: `{output_dir / "stock_target_review_actions.csv"}`
- Review notes CSV: `{payload.get("notes_path", "N/A")}`
"""


def _decision_status_hint(review_posture: Any) -> str:
    posture = _safe_text(review_posture)
    if posture == "risk_review":
        return "risk_review: choose watch/resolved/reviewed after human evidence check"
    if posture == "profit_protection_review":
        return "profit_protection_review: choose watch/reviewed/resolved after human evidence check"
    if posture == "data_review":
        return "data_review: verify source data before choosing status"
    return "observe: choose reviewed/watch/resolved after human evidence check"


def build_stock_target_review_decision_template(
    assistant: pd.DataFrame,
    assistant_payload: dict[str, Any],
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Build a fill-in template for human stock-target review decisions."""
    rows: list[dict[str, Any]] = []
    notes_path = assistant_payload.get("notes_path")
    for raw in assistant.to_dict(orient="records"):
        posture = raw.get("suggested_review_posture")
        rows.append(
            {
                "date": raw.get("date"),
                "decision_rank": 0,
                "source_assistant_rank": int(_as_float(raw.get("assistant_rank"), len(rows) + 1)),
                "layer": raw.get("layer"),
                "code": _clean_code(raw.get("code")),
                "name": raw.get("name"),
                "review_posture": posture,
                "status_hint": _decision_status_hint(posture),
                "manual_status_to_fill": "",
                "manual_note_to_fill": "",
                "next_review_date_to_fill": "",
                "reviewed_by_to_fill": "",
                "allowed_manual_statuses": "reviewed|watch|resolved|exclude_candidate|unreviewed",
                "portfolio_target_weight": _as_float(raw.get("portfolio_target_weight"), 0.0),
                "unrealized_return": _optional_float(raw.get("unrealized_return")),
                "latest_close": _optional_float(raw.get("latest_close")),
                "latest_pct_change": _optional_float(raw.get("latest_pct_change")),
                "support_price": _optional_float(raw.get("support_price")),
                "pressure_price": _optional_float(raw.get("pressure_price")),
                "stop_loss_reference": _optional_float(raw.get("stop_loss_reference")),
                "trigger_condition": raw.get("trigger_condition"),
                "invalidation_condition": raw.get("invalidation_condition"),
                "evidence_checklist": raw.get("evidence_checklist"),
                "review_reason": raw.get("review_reason"),
                "recommended_review": raw.get("recommended_review"),
                "notes_path": notes_path,
                "broker_action": "none",
                "research_only": True,
            }
        )
    frame = pd.DataFrame(rows, columns=STOCK_TARGET_REVIEW_DECISION_TEMPLATE_COLUMNS)
    if not frame.empty:
        frame = frame.sort_values(["source_assistant_rank", "code"], ascending=[True, True]).reset_index(drop=True)
        frame["decision_rank"] = range(1, len(frame) + 1)
    blank_status_count = int((frame["manual_status_to_fill"].map(_safe_text) == "").sum()) if not frame.empty else 0
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "status": assistant_payload.get("status", "missing_sources"),
        "latest_date": assistant_payload.get("latest_date"),
        "decision_template_row_count": int(len(frame)),
        "blank_manual_status_count": blank_status_count,
        "source_assistant_row_count": assistant_payload.get("assistant_row_count", len(assistant)),
        "allowed_manual_statuses": ["reviewed", "watch", "resolved", "exclude_candidate", "unreviewed"],
        "notes_path": notes_path,
        "top_template_rows": _json_records(frame.head(10)),
        "research_only": True,
        "broker_action": "none",
        "note": "Decision template rows are intentionally blank for human review; they do not set manual_status, change model targets, or place orders.",
    }
    return frame, payload


def _render_stock_target_review_decision_template_report(
    template: pd.DataFrame,
    payload: dict[str, Any],
    output_dir: Path,
) -> str:
    if template.empty:
        rows = "| N/A | `N/A` | `N/A` | `N/A` | `N/A` | N/A | N/A | N/A | `N/A` | No rows require a decision template |"
    else:
        rows = "\n".join(
            "| {rank} | `{code}` | {name} | `{posture}` | `{status}` | {ret} | {support} | {pressure} | {hint} | {reason} |".format(
                rank=int(row["decision_rank"]),
                code=str(row.get("code")).zfill(6),
                name=_md_text(row.get("name"), 36),
                posture=_md_text(row.get("review_posture"), 36),
                status=_md_text(row.get("manual_status_to_fill"), 20),
                ret=_maybe_pct(row.get("unrealized_return")),
                support=_num(row.get("support_price")),
                pressure=_num(row.get("pressure_price")),
                hint=_md_text(row.get("status_hint"), 90),
                reason=_md_text(row.get("review_reason"), 90),
            )
            for _, row in template.iterrows()
        )
    return f"""# Paper Stock Target Review Decision Template

Generated at: `{payload.get("generated_at")}`

This is a research-only fill-in template for human stock-target review. It does not set manual_status, connect to brokers, place orders, change target weights, or provide investment advice.

## Summary

| Item | Value |
| --- | ---: |
| Status | `{payload.get("status")}` |
| Latest date | {payload.get("latest_date", "N/A")} |
| Template rows | {payload.get("decision_template_row_count", 0)} |
| Blank manual status rows | {payload.get("blank_manual_status_count", 0)} |
| Allowed statuses | `reviewed|watch|resolved|exclude_candidate|unreviewed` |
| Notes CSV to update | `{payload.get("notes_path", "N/A")}` |
| Broker action | `{payload.get("broker_action")}` |

## Fill-In Queue

| Rank | Code | Name | Review posture | manual_status_to_fill | Unrealized | Support | Pressure | Status hint | Reason |
| ---: | --- | --- | --- | --- | ---: | ---: | ---: | --- | --- |
{rows}

## How To Use

Fill `manual_status_to_fill`, `manual_note_to_fill`, and optionally `next_review_date_to_fill` / `reviewed_by_to_fill`, then use the filled rows to update the persistent notes CSV.

## Files

- Decision template CSV: `{output_dir / "stock_target_review_decision_template.csv"}`
- Highlighted fill-in workbook: `{payload.get("decision_template_xlsx_path", output_dir / "stock_target_review_decision_template.xlsx")}`
- Decision template JSON: `{output_dir / "stock_target_review_decision_template.json"}`
- Decision template report: `{output_dir / "stock_target_review_decision_template.md"}`
- Persistent notes CSV: `{payload.get("notes_path", "N/A")}`
"""


def _xlsx_scalar(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def write_stock_target_review_decision_template_xlsx(
    template: pd.DataFrame,
    payload: dict[str, Any],
    output_path: str | Path,
) -> Path:
    """Write a color-highlighted Excel template for human fill-in fields."""
    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as error:  # pragma: no cover - dependency is covered by requirements and CI env.
        raise RuntimeError("openpyxl is required to write the highlighted decision-template workbook.") from error

    resolved = Path(output_path)
    resolved.parent.mkdir(parents=True, exist_ok=True)
    columns = list(template.columns) if not template.empty else list(STOCK_TARGET_REVIEW_DECISION_TEMPLATE_COLUMNS)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "decision_template"
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = True

    header_fill = PatternFill("solid", fgColor="FF1F4E78")
    header_font = Font(color="FFFFFFFF", bold=True)
    fill_header_fill = PatternFill("solid", fgColor="FFFFC000")
    required_fill = PatternFill("solid", fgColor="FFFFE699")
    optional_fill = PatternFill("solid", fgColor="FFFFF2CC")
    neutral_fill = PatternFill("solid", fgColor="FFFFFFFF")
    thin = Side(style="thin", color="FFD9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.append(columns)
    for record in template.to_dict(orient="records"):
        sheet.append([_xlsx_scalar(record.get(column)) for column in columns])

    max_row = max(sheet.max_row, 2)
    max_col = len(columns)
    sheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    fill_column_indexes = {
        column: columns.index(column) + 1
        for column in STOCK_TARGET_REVIEW_DECISION_TEMPLATE_USER_FILL_COLUMNS
        if column in columns
    }

    for col_idx, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = fill_header_fill if column in fill_column_indexes else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        if column in fill_column_indexes:
            cell.comment = Comment("Need human input here.", "quant_etf_lab")
        width = max(12, min(42, max(len(str(column)) + 2, 14)))
        if column in {"manual_note_to_fill", "trigger_condition", "invalidation_condition", "evidence_checklist", "review_reason", "recommended_review"}:
            width = 36
        if column in {"manual_status_to_fill", "next_review_date_to_fill", "reviewed_by_to_fill"}:
            width = 24
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx in range(2, max_row + 1):
        for col_idx, column in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=column in fill_column_indexes or column in {"trigger_condition", "invalidation_condition", "review_reason", "recommended_review"})
            if column == "manual_status_to_fill":
                cell.fill = required_fill
            elif column in fill_column_indexes:
                cell.fill = optional_fill
            else:
                cell.fill = neutral_fill
            if column in {"portfolio_target_weight", "unrealized_return", "latest_pct_change"}:
                cell.number_format = "0.00%"
            elif column in {"latest_close", "support_price", "pressure_price", "stop_loss_reference"}:
                cell.number_format = "0.00"
        sheet.row_dimensions[row_idx].height = 32

    if "manual_status_to_fill" in fill_column_indexes:
        status_col = get_column_letter(fill_column_indexes["manual_status_to_fill"])
        allowed = ",".join(["reviewed", "watch", "resolved", "exclude_candidate", "unreviewed"])
        validation = DataValidation(type="list", formula1=f'"{allowed}"', allow_blank=True)
        validation.error = "Use one of: reviewed, watch, resolved, exclude_candidate, unreviewed."
        validation.errorTitle = "Invalid manual status"
        validation.prompt = "Choose a manual review status."
        validation.promptTitle = "Manual status"
        sheet.add_data_validation(validation)
        validation.add(f"{status_col}2:{status_col}{max_row}")

    sheet["A1"].comment = Comment(
        f"Fill highlighted columns only, then apply with stock-review-apply-template. Latest date: {_safe_text(payload.get('latest_date'), 'N/A')}",
        "quant_etf_lab",
    )
    workbook.save(resolved)
    return resolved


def _publish_stock_target_review_decision_template(
    template: pd.DataFrame,
    payload: dict[str, Any],
    *,
    output_dir: Path,
    csv_path: Path,
    json_path: Path,
    report_path: Path,
    xlsx_path: Path,
) -> bool:
    """Publish the template and rebuild Excel only when its source changed."""

    csv_changed = write_text_if_changed(csv_path, template.to_csv(index=False))
    published_payload, payload_changed = publish_json_if_semantically_changed(json_path, payload)
    sources_changed = csv_changed or payload_changed
    if sources_changed or not report_path.exists():
        write_text_if_changed(
            report_path,
            _render_stock_target_review_decision_template_report(template, published_payload, output_dir),
        )
    if sources_changed or not xlsx_path.exists():
        write_stock_target_review_decision_template_xlsx(template, published_payload, xlsx_path)
        return True
    return False


def _load_stock_target_review_decision_template(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Decision template CSV does not exist: {path}")
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        try:
            frame = pd.read_excel(path, sheet_name="decision_template", dtype={"code": str})
        except ValueError:
            frame = pd.read_excel(path, sheet_name=0, dtype={"code": str})
    else:
        frame = pd.read_csv(path, dtype={"code": str})
    for column in STOCK_TARGET_REVIEW_DECISION_TEMPLATE_COLUMNS:
        if column not in frame.columns:
            frame[column] = ""
    frame["code"] = frame["code"].map(_clean_code)
    frame["layer"] = frame["layer"].map(_safe_text)
    return frame


def _template_apply_allowed_statuses() -> set[str]:
    return {"reviewed", "watch", "resolved", "exclude_candidate", "unreviewed"}


def _build_decision_apply_audit_row(
    raw: dict[str, Any],
    apply_rank: int,
    apply_status: str,
    apply_message: str,
    previous_manual_status: Any = "",
    new_manual_status: Any = "",
    reviewed_at: Any = "",
) -> dict[str, Any]:
    return {
        "date": _format_date(raw.get("date")),
        "apply_rank": apply_rank,
        "apply_status": apply_status,
        "apply_message": apply_message,
        "layer": _safe_text(raw.get("layer")),
        "code": _clean_code(raw.get("code")),
        "name": _safe_text(raw.get("name")),
        "previous_manual_status": _safe_text(previous_manual_status),
        "new_manual_status": _safe_text(new_manual_status),
        "manual_note": _safe_text(raw.get("manual_note_to_fill")),
        "reviewed_at": _safe_text(reviewed_at),
        "reviewed_by": _safe_text(raw.get("reviewed_by_to_fill")),
        "next_review_date": _safe_text(raw.get("next_review_date_to_fill")),
        "review_posture": _safe_text(raw.get("review_posture")),
        "review_reason": _safe_text(raw.get("review_reason")),
        "broker_action": "none",
        "research_only": True,
    }


def _render_stock_target_review_decision_apply_report(
    audit: pd.DataFrame,
    payload: dict[str, Any],
) -> str:
    if audit.empty:
        rows = "| N/A | `N/A` | `N/A` | `N/A` | `N/A` | N/A |"
    else:
        rows = "\n".join(
            "| {rank} | `{code}` | {name} | `{status}` | `{new_status}` | {message} |".format(
                rank=int(row.get("apply_rank", 0)),
                code=str(row.get("code", "")).zfill(6),
                name=_md_text(row.get("name"), 36),
                status=_md_text(row.get("apply_status"), 32),
                new_status=_md_text(row.get("new_manual_status"), 32),
                message=_md_text(row.get("apply_message"), 100),
            )
            for _, row in audit.iterrows()
        )
    return f"""# Stock Target Review Decision Apply

Generated at: `{payload.get("generated_at")}`

This research-only report records local manual-review template rows applied to the persistent stock-target notes CSV. It does not connect to brokers, place orders, change model targets, or provide investment advice.

## Summary

| Item | Value |
| --- | ---: |
| Template rows | {payload.get("template_row_count", 0)} |
| Applied rows | {payload.get("applied_count", 0)} |
| Blank ignored rows | {payload.get("blank_ignored_count", 0)} |
| Invalid status rows | {payload.get("invalid_status_count", 0)} |
| Notes rows after apply | {payload.get("notes_row_count", 0)} |
| Broker action | `{payload.get("broker_action")}` |

## Audit

| Rank | Code | Name | Apply status | New manual status | Message |
| ---: | --- | --- | --- | --- | --- |
{rows}

## Files

- Decision template CSV: `{payload.get("template_path")}`
- Persistent notes CSV: `{payload.get("notes_path")}`
- Notes snapshot CSV: `{payload.get("notes_snapshot_path")}`
- Apply audit CSV: `{payload.get("audit_path")}`
"""


def apply_stock_target_review_decision_template(
    template_path: str | Path,
    notes_path: str | Path,
    output_dir: str | Path,
    *,
    reviewed_at: str | None = None,
) -> StockTargetReviewDecisionApplyResult:
    """Apply filled human-review template rows to the persistent notes CSV."""
    resolved_template = Path(template_path)
    resolved_notes = Path(notes_path)
    resolved_output = Path(output_dir)
    resolved_output.mkdir(parents=True, exist_ok=True)
    notes_snapshot_path = resolved_output / "stock_target_review_notes_after_apply.csv"
    audit_path = resolved_output / "stock_target_review_decision_apply_audit.csv"
    payload_path = resolved_output / "stock_target_review_decision_apply.json"
    report_path = resolved_output / "stock_target_review_decision_apply.md"

    template = _load_stock_target_review_decision_template(resolved_template)
    notes = _load_stock_target_review_notes(resolved_notes)
    records = notes.to_dict(orient="records")
    index_by_key: dict[tuple[str, str], int] = {}
    for index, record in enumerate(records):
        key = _stock_review_note_key(record)
        if key[1]:
            index_by_key[key] = index

    allowed_statuses = _template_apply_allowed_statuses()
    audit_rows: list[dict[str, Any]] = []
    applied_count = 0
    blank_ignored_count = 0
    invalid_status_count = 0
    apply_time = reviewed_at or datetime.now().isoformat(timespec="seconds")

    for apply_rank, raw in enumerate(template.to_dict(orient="records"), start=1):
        status_text = _safe_text(raw.get("manual_status_to_fill")).strip()
        normalized_status = _normalize_manual_status(status_text)
        key = _stock_review_note_key(raw)
        previous_status = ""
        if key in index_by_key:
            previous_status = records[index_by_key[key]].get("manual_status", "")

        if not status_text:
            blank_ignored_count += 1
            audit_rows.append(
                _build_decision_apply_audit_row(
                    raw,
                    apply_rank,
                    "blank_ignored",
                    "manual_status_to_fill is blank; notes were not changed.",
                    previous_manual_status=previous_status,
                )
            )
            continue
        if normalized_status not in allowed_statuses:
            invalid_status_count += 1
            audit_rows.append(
                _build_decision_apply_audit_row(
                    raw,
                    apply_rank,
                    "invalid_status",
                    f"manual_status_to_fill={status_text!r} is not allowed; notes were not changed.",
                    previous_manual_status=previous_status,
                    new_manual_status=status_text,
                )
            )
            continue
        if not key[1]:
            invalid_status_count += 1
            audit_rows.append(
                _build_decision_apply_audit_row(
                    raw,
                    apply_rank,
                    "invalid_status",
                    "code is blank or invalid; notes were not changed.",
                    previous_manual_status=previous_status,
                    new_manual_status=normalized_status,
                )
            )
            continue

        if key in index_by_key:
            record = records[index_by_key[key]]
        else:
            record = {column: "" for column in STOCK_TARGET_REVIEW_NOTE_COLUMNS}
            records.append(record)
            index_by_key[key] = len(records) - 1

        current_date = _format_date(raw.get("date"))
        record["layer"] = key[0]
        record["code"] = key[1]
        record["name"] = _safe_text(raw.get("name"))
        record["first_seen_date"] = _safe_text(record.get("first_seen_date"), current_date)
        record["last_seen_date"] = current_date
        record["last_review_bucket"] = _safe_text(record.get("last_review_bucket"), _safe_text(raw.get("review_posture")))
        record["last_review_stage"] = _safe_text(record.get("last_review_stage"), "manual_template_apply")
        record["manual_status"] = normalized_status
        record["manual_note"] = _safe_text(raw.get("manual_note_to_fill"))
        record["reviewed_at"] = apply_time
        record["reviewed_by"] = _safe_text(raw.get("reviewed_by_to_fill"))
        record["next_review_date"] = _safe_text(raw.get("next_review_date_to_fill"))
        record["last_model_reason"] = _safe_text(raw.get("review_reason"), _safe_text(record.get("last_model_reason")))
        record["broker_action"] = "none"
        record["research_only"] = True
        applied_count += 1
        audit_rows.append(
            _build_decision_apply_audit_row(
                raw,
                apply_rank,
                "applied",
                "manual review row applied to persistent notes.",
                previous_manual_status=previous_status,
                new_manual_status=normalized_status,
                reviewed_at=apply_time,
            )
        )

    if records:
        updated_notes = pd.DataFrame(records)
    else:
        updated_notes = pd.DataFrame(columns=STOCK_TARGET_REVIEW_NOTE_COLUMNS)
    for column in STOCK_TARGET_REVIEW_NOTE_COLUMNS:
        if column not in updated_notes.columns:
            updated_notes[column] = ""
    updated_notes = updated_notes[_stock_review_note_columns(updated_notes)]

    audit = pd.DataFrame(audit_rows, columns=STOCK_TARGET_REVIEW_DECISION_APPLY_AUDIT_COLUMNS)
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "template_path": str(resolved_template),
        "notes_path": str(resolved_notes),
        "notes_snapshot_path": str(notes_snapshot_path),
        "audit_path": str(audit_path),
        "report_path": str(report_path),
        "template_row_count": int(len(template)),
        "notes_row_count": int(len(updated_notes)),
        "applied_count": int(applied_count),
        "blank_ignored_count": int(blank_ignored_count),
        "invalid_status_count": int(invalid_status_count),
        "allowed_manual_statuses": sorted(allowed_statuses),
        "top_audit_rows": _json_records(audit.head(10)),
        "research_only": True,
        "broker_action": "none",
        "note": "Only valid nonblank manual_status_to_fill rows are written to the local persistent notes CSV; no broker action is generated.",
    }

    if applied_count > 0:
        resolved_notes.parent.mkdir(parents=True, exist_ok=True)
        updated_notes.to_csv(resolved_notes, index=False, encoding="utf-8")
    updated_notes.to_csv(notes_snapshot_path, index=False, encoding="utf-8")
    audit.to_csv(audit_path, index=False, encoding="utf-8")
    payload_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    report_path.write_text(_render_stock_target_review_decision_apply_report(audit, payload), encoding="utf-8")

    return StockTargetReviewDecisionApplyResult(
        output_dir=resolved_output,
        template_path=resolved_template,
        notes_path=resolved_notes,
        notes_snapshot_path=notes_snapshot_path,
        audit_path=audit_path,
        payload_path=payload_path,
        report_path=report_path,
        notes=updated_notes,
        audit=audit,
        payload=payload,
    )


def _future_return_fields(
    history: pd.DataFrame,
    review_date: Any,
    horizons: tuple[int, ...] = STOCK_TARGET_REVIEW_OUTCOME_HORIZONS,
) -> dict[str, Any]:
    fields: dict[str, Any] = {"entry_date": None, "entry_close": None, "outcome_status": "missing_entry_price"}
    for horizon in horizons:
        fields[f"future_date_{horizon}d"] = None
        fields[f"future_close_{horizon}d"] = None
        fields[f"return_{horizon}d"] = None
        fields[f"outcome_status_{horizon}d"] = "missing_entry_price"
    if history.empty:
        return fields

    target_date = pd.to_datetime(review_date, errors="coerce")
    if pd.isna(target_date):
        target_date = history["date"].max()
    eligible = history[history["date"] <= pd.Timestamp(target_date)]
    if eligible.empty:
        return fields
    entry = eligible.iloc[-1]
    entry_index = int(eligible.index[-1])
    entry_close = float(entry["close"])
    fields["entry_date"] = pd.Timestamp(entry["date"]).strftime("%Y-%m-%d")
    fields["entry_close"] = entry_close
    statuses: list[str] = []
    for horizon in horizons:
        future_index = entry_index + int(horizon)
        if future_index >= len(history):
            status = "pending"
            future_date = None
            future_close = None
            future_return = None
        else:
            future = history.iloc[future_index]
            status = "available"
            future_date = pd.Timestamp(future["date"]).strftime("%Y-%m-%d")
            future_close = float(future["close"])
            future_return = (future_close / entry_close - 1.0) if entry_close else None
        fields[f"future_date_{horizon}d"] = future_date
        fields[f"future_close_{horizon}d"] = future_close
        fields[f"return_{horizon}d"] = future_return
        fields[f"outcome_status_{horizon}d"] = status
        statuses.append(status)
    if all(status == "available" for status in statuses):
        fields["outcome_status"] = "complete"
    elif any(status == "available" for status in statuses):
        fields["outcome_status"] = "partial"
    else:
        fields["outcome_status"] = "pending"
    return fields


def _outcome_group_summary(frame: pd.DataFrame, group_column: str) -> list[dict[str, Any]]:
    if frame.empty or group_column not in frame.columns:
        return []
    rows: list[dict[str, Any]] = []
    for group_value, group in frame.groupby(group_column, dropna=False):
        row: dict[str, Any] = {group_column: group_value, "count": int(len(group))}
        for horizon in STOCK_TARGET_REVIEW_OUTCOME_HORIZONS:
            returns = pd.to_numeric(group.get(f"return_{horizon}d"), errors="coerce").dropna()
            row[f"evaluable_{horizon}d"] = int(len(returns))
            row[f"avg_return_{horizon}d"] = float(returns.mean()) if not returns.empty else None
            row[f"win_rate_{horizon}d"] = float((returns > 0).mean()) if not returns.empty else None
        rows.append(row)
    return rows


def _has_nonempty_text(value: Any) -> bool:
    return bool(_safe_text(value))


def _outcome_horizon_summary(frame: pd.DataFrame) -> dict[str, dict[str, Any]]:
    horizon_summary: dict[str, dict[str, Any]] = {}
    for horizon in STOCK_TARGET_REVIEW_OUTCOME_HORIZONS:
        returns = pd.to_numeric(frame.get(f"return_{horizon}d"), errors="coerce").dropna() if not frame.empty else pd.Series(dtype=float)
        horizon_summary[f"{horizon}d"] = {
            "evaluable_count": int(len(returns)),
            "pending_count": int((frame.get(f"outcome_status_{horizon}d") == "pending").sum()) if not frame.empty else 0,
            "avg_return": float(returns.mean()) if not returns.empty else None,
            "win_rate": float((returns > 0).mean()) if not returns.empty else None,
        }
    return horizon_summary


def build_stock_target_review_outcomes(
    project_root: str | Path,
    review: pd.DataFrame,
    actions: pd.DataFrame,
    review_payload: dict[str, Any],
    price_history_cache: dict[str, tuple[pd.DataFrame, str]] | None = None,
    price_history_cache_dir: Path | None = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    root = Path(project_root)
    action_lookup: dict[tuple[str, str], str] = {}
    if not actions.empty:
        for item in actions.to_dict(orient="records"):
            action_lookup[(_safe_text(item.get("layer")), _clean_code(item.get("code")))] = _safe_text(item.get("action_code"))

    rows: list[dict[str, Any]] = []
    for raw in review.to_dict(orient="records"):
        code = _clean_code(raw.get("code"))
        layer = _safe_text(raw.get("layer"))
        history, price_source = _stock_price_history(root, code, price_history_cache, price_history_cache_dir)
        fields = _future_return_fields(history, raw.get("date"))
        rows.append(
            {
                "date": raw.get("date"),
                "layer": layer,
                "code": code,
                "name": raw.get("name"),
                "review_bucket": raw.get("review_bucket"),
                "review_stage": raw.get("review_stage"),
                "review_priority_score": _as_float(raw.get("review_priority_score"), 0.0),
                "action_code": action_lookup.get((layer, code), ""),
                "manual_status": raw.get("manual_status"),
                "manual_status_normalized": raw.get("manual_status_normalized"),
                "manual_review_state": raw.get("manual_review_state"),
                **fields,
                "price_source": price_source,
                "broker_action": "none",
                "research_only": True,
            }
        )
    frame = pd.DataFrame(rows, columns=STOCK_TARGET_REVIEW_OUTCOME_COLUMNS)
    if not frame.empty:
        frame = frame.sort_values(["date", "review_priority_score", "code"], ascending=[True, False, True]).reset_index(drop=True)

    horizon_summary = _outcome_horizon_summary(frame)
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "status": review_payload.get("status", "missing_sources"),
        "latest_date": review_payload.get("latest_date"),
        "outcome_row_count": int(len(frame)),
        "complete_count": int((frame.get("outcome_status") == "complete").sum()) if not frame.empty else 0,
        "partial_count": int((frame.get("outcome_status") == "partial").sum()) if not frame.empty else 0,
        "pending_count": int((frame.get("outcome_status") == "pending").sum()) if not frame.empty else 0,
        "missing_entry_price_count": int((frame.get("outcome_status") == "missing_entry_price").sum()) if not frame.empty else 0,
        "horizon_summary": horizon_summary,
        "manual_status_summary": _outcome_group_summary(frame, "manual_status_normalized"),
        "action_code_summary": _outcome_group_summary(frame[frame["action_code"].map(_has_nonempty_text)] if not frame.empty else frame, "action_code"),
        "notes_path": review_payload.get("notes_path"),
        "research_only": True,
        "broker_action": "none",
        "note": "Outcome tracking uses local OHLCV after the review date when available; pending horizons are not guessed.",
    }
    return frame, payload


def _stock_review_outcome_columns(frame: pd.DataFrame) -> list[str]:
    extras = [column for column in frame.columns if column not in STOCK_TARGET_REVIEW_OUTCOME_COLUMNS]
    return STOCK_TARGET_REVIEW_OUTCOME_COLUMNS + extras


def _review_date_text(value: Any) -> str:
    if isinstance(value, str):
        text = value.strip()
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}(?:[ T].*)?", text):
            return text[:10]
        if re.fullmatch(r"\d{8}", text):
            return f"{text[:4]}-{text[4:6]}-{text[6:8]}"
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return _safe_text(value)
    return pd.Timestamp(parsed).strftime("%Y-%m-%d")


def _stock_review_outcome_key(row: dict[str, Any]) -> tuple[str, str, str]:
    return (_review_date_text(row.get("date")), _safe_text(row.get("layer")), _clean_code(row.get("code")))


def _load_stock_target_review_outcomes_history(path: Path) -> pd.DataFrame:
    if not path.exists() or path.stat().st_size == 0:
        return pd.DataFrame(columns=STOCK_TARGET_REVIEW_OUTCOME_COLUMNS)
    try:
        frame = pd.read_csv(path, dtype={"code": str, "price_source_fingerprint": str})
    except (OSError, pd.errors.EmptyDataError):
        return pd.DataFrame(columns=STOCK_TARGET_REVIEW_OUTCOME_COLUMNS)
    for column in STOCK_TARGET_REVIEW_OUTCOME_COLUMNS:
        if column not in frame.columns:
            frame[column] = pd.NA
    if not frame.empty:
        frame["date"] = frame["date"].map(_review_date_text)
        frame["layer"] = frame["layer"].map(_safe_text)
        frame["code"] = frame["code"].map(_clean_code)
    return frame[_stock_review_outcome_columns(frame)]


def _load_stock_review_outcome_price_source(path: Path) -> pd.DataFrame:
    try:
        frame = pd.read_csv(path)
    except (OSError, pd.errors.EmptyDataError):
        return pd.DataFrame(columns=["date", "close", "high", "low"])
    if frame.empty or not {"date", "close"}.issubset(frame.columns):
        return pd.DataFrame(columns=["date", "close", "high", "low"])
    data = frame[["date", "close"]].copy()
    data["high"] = frame["high"] if "high" in frame.columns else frame["close"]
    data["low"] = frame["low"] if "low" in frame.columns else frame["close"]
    data["date"] = pd.to_datetime(data["date"], errors="coerce")
    data["close"] = pd.to_numeric(data["close"], errors="coerce")
    data["high"] = pd.to_numeric(data["high"], errors="coerce").fillna(data["close"])
    data["low"] = pd.to_numeric(data["low"], errors="coerce").fillna(data["close"])
    return data.dropna(subset=["date", "close"]).sort_values("date").drop_duplicates(subset=["date"], keep="last").reset_index(drop=True)


def _price_source_fingerprint(path: Path) -> tuple[str, Any]:
    source_stat = path.stat()
    return f"{source_stat.st_size}:{source_stat.st_mtime_ns}", source_stat


def _stock_review_outcome_has_open_horizon(record: dict[str, Any]) -> bool:
    if _safe_text(record.get("outcome_status")) == "complete":
        return False
    return any(_safe_text(record.get(f"outcome_status_{horizon}d")) != "available" for horizon in STOCK_TARGET_REVIEW_OUTCOME_HORIZONS)


def _refresh_stock_review_outcome_record_from_price_source(
    record: dict[str, Any],
    price_history_cache: dict[Path, pd.DataFrame] | None = None,
) -> bool:
    if not _stock_review_outcome_has_open_horizon(record):
        return False
    source = _safe_text(record.get("price_source"))
    if not source or source == "missing_local_price_history":
        return False
    source_path = Path(source)
    if not source_path.exists() or not source_path.is_file():
        return False
    source_fingerprint, source_stat = _price_source_fingerprint(source_path)
    if _safe_text(record.get("price_source_fingerprint")) == source_fingerprint:
        return False

    before_statuses = {
        horizon: _safe_text(record.get(f"outcome_status_{horizon}d"))
        for horizon in STOCK_TARGET_REVIEW_OUTCOME_HORIZONS
    }
    cache_key = source_path.resolve()
    if price_history_cache is not None and cache_key in price_history_cache:
        history = price_history_cache[cache_key]
    else:
        history = _load_stock_review_outcome_price_source(source_path)
        if price_history_cache is not None:
            price_history_cache[cache_key] = history
    fields = _future_return_fields(history, record.get("date"))
    for key, value in fields.items():
        record[key] = value
    record["price_source_size"] = source_stat.st_size
    record["price_source_mtime_ns"] = source_stat.st_mtime_ns
    record["price_source_fingerprint"] = source_fingerprint

    return any(
        before_statuses[horizon] != "available"
        and _safe_text(record.get(f"outcome_status_{horizon}d")) == "available"
        for horizon in STOCK_TARGET_REVIEW_OUTCOME_HORIZONS
    )


def _normalize_stock_review_outcome_record(record: dict[str, Any], columns: list[str]) -> dict[str, Any]:
    normalized = {column: record.get(column, pd.NA) for column in columns}
    normalized["date"] = _review_date_text(normalized.get("date"))
    normalized["layer"] = _safe_text(normalized.get("layer"))
    normalized["code"] = _clean_code(normalized.get("code"))
    if not _safe_text(normalized.get("broker_action")):
        normalized["broker_action"] = "none"
    if not _safe_text(normalized.get("research_only")):
        normalized["research_only"] = True
    return normalized


def sync_stock_target_review_outcomes_history(
    outcomes: pd.DataFrame,
    history_path: str | Path,
    history_snapshot_path: str | Path,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Append or update current outcome rows in the persistent outcome history."""
    resolved_history = Path(history_path)
    resolved_snapshot = Path(history_snapshot_path)
    existing = _load_stock_target_review_outcomes_history(resolved_history)

    combined_columns = STOCK_TARGET_REVIEW_OUTCOME_COLUMNS.copy()
    for frame in (existing, outcomes):
        for column in frame.columns:
            if column not in combined_columns:
                combined_columns.append(column)

    records = [
        _normalize_stock_review_outcome_record(record, combined_columns)
        for record in existing.to_dict(orient="records")
    ]
    index_by_key: dict[tuple[str, str, str], int] = {}
    for index, record in enumerate(records):
        key = _stock_review_outcome_key(record)
        if key[0] and key[2]:
            index_by_key[key] = index

    updated_count = 0
    inserted_count = 0
    for raw in outcomes.to_dict(orient="records"):
        record = _normalize_stock_review_outcome_record(raw, combined_columns)
        key = _stock_review_outcome_key(record)
        if key[0] and key[2] and key in index_by_key:
            existing_record = records[index_by_key[key]]
            for column, value in record.items():
                if column in STOCK_TARGET_REVIEW_OUTCOME_SOURCE_METADATA_COLUMNS and pd.isna(value):
                    continue
                existing_record[column] = value
            updated_count += 1
        else:
            records.append(record)
            if key[0] and key[2]:
                index_by_key[key] = len(records) - 1
            inserted_count += 1

    matured_count = 0
    source_paths_to_load: set[Path] = set()
    for record in records:
        if not _stock_review_outcome_has_open_horizon(record):
            continue
        source = _safe_text(record.get("price_source"))
        if not source or source == "missing_local_price_history":
            continue
        source_path = Path(source)
        if not source_path.exists() or not source_path.is_file():
            continue
        source_fingerprint, _ = _price_source_fingerprint(source_path)
        if _safe_text(record.get("price_source_fingerprint")) != source_fingerprint:
            source_paths_to_load.add(source_path.resolve())
    source_paths = sorted(source_paths_to_load, key=str)
    price_history_cache: dict[Path, pd.DataFrame] = {}
    if source_paths:
        worker_count = min(8, len(source_paths))
        with ThreadPoolExecutor(max_workers=worker_count, thread_name_prefix="outcome-price") as executor:
            loaded_histories = executor.map(_load_stock_review_outcome_price_source, source_paths)
            price_history_cache.update(zip(source_paths, loaded_histories))
    for record in records:
        if _refresh_stock_review_outcome_record_from_price_source(record, price_history_cache):
            matured_count += 1

    history = pd.DataFrame(records, columns=combined_columns)
    if not history.empty:
        history["_date_sort"] = pd.to_datetime(history["date"], errors="coerce")
        sort_columns = ["_date_sort", "layer", "code"]
        ascending = [True, True, True]
        if "review_priority_score" in history.columns:
            history["_priority_sort"] = pd.to_numeric(history["review_priority_score"], errors="coerce")
            sort_columns.insert(1, "_priority_sort")
            ascending.insert(1, False)
        history = history.sort_values(sort_columns, ascending=ascending, na_position="last")
        history = history.drop(columns=["_date_sort", "_priority_sort"], errors="ignore").reset_index(drop=True)

    resolved_history.parent.mkdir(parents=True, exist_ok=True)
    resolved_snapshot.parent.mkdir(parents=True, exist_ok=True)
    history.to_csv(resolved_history, index=False, encoding="utf-8")
    history.to_csv(resolved_snapshot, index=False, encoding="utf-8")

    status_counts = history["outcome_status"].value_counts(dropna=False).to_dict() if not history.empty else {}
    parsed_dates = pd.to_datetime(history["date"], errors="coerce") if not history.empty else pd.Series(dtype="datetime64[ns]")
    latest_review_date = pd.Timestamp(parsed_dates.max()).strftime("%Y-%m-%d") if not parsed_dates.empty and pd.notna(parsed_dates.max()) else None
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "history_path": str(resolved_history),
        "history_snapshot_path": str(resolved_snapshot),
        "history_row_count": int(len(history)),
        "history_current_row_count": int(len(outcomes)),
        "history_updated_row_count": int(len(outcomes) + matured_count),
        "history_matured_row_count": int(matured_count),
        "history_replaced_row_count": int(updated_count),
        "history_inserted_row_count": int(inserted_count),
        "history_price_source_count": int(len(source_paths)),
        "history_price_source_workers": int(min(8, len(source_paths))) if source_paths else 0,
        "history_complete_count": int(status_counts.get("complete", 0)),
        "history_partial_count": int(status_counts.get("partial", 0)),
        "history_pending_count": int(status_counts.get("pending", 0)),
        "history_missing_entry_price_count": int(status_counts.get("missing_entry_price", 0)),
        "history_latest_review_date": latest_review_date,
        "history_horizon_summary": _outcome_horizon_summary(history),
        "history_manual_status_summary": _outcome_group_summary(history, "manual_status_normalized"),
        "history_action_code_summary": _outcome_group_summary(history[history["action_code"].map(_has_nonempty_text)] if not history.empty else history, "action_code"),
        "research_only": True,
        "broker_action": "none",
        "note": "Persistent research-only outcome history keyed by review date, layer, and code; existing rows are refreshed from local price_source files as future OHLCV becomes available.",
    }
    return history, payload


def _render_stock_target_review_outcomes_report(outcomes: pd.DataFrame, payload: dict[str, Any], output_dir: Path) -> str:
    if outcomes.empty:
        rows = "| `N/A` | `N/A` | N/A | `N/A` | N/A | N/A | N/A | N/A | No stock target outcomes to track |"
    else:
        rows = "\n".join(
            "| `{code}` | {name} | {entry} | `{status}` | {r1} | {r5} | {r10} | {r20} | {action} |".format(
                code=str(row.get("code")).zfill(6),
                name=_md_text(row.get("name"), 40),
                entry=row.get("entry_date") or "N/A",
                status=_md_text(row.get("manual_status_normalized"), 30),
                r1=_maybe_pct(row.get("return_1d")),
                r5=_maybe_pct(row.get("return_5d")),
                r10=_maybe_pct(row.get("return_10d")),
                r20=_maybe_pct(row.get("return_20d")),
                action=_md_text(row.get("action_code"), 40),
            )
            for _, row in outcomes.iterrows()
        )
    horizon_rows = "\n".join(
        "| `{horizon}` | {evaluable} | {pending} | {avg_return} | {win_rate} |".format(
            horizon=horizon,
            evaluable=item.get("evaluable_count", 0),
            pending=item.get("pending_count", 0),
            avg_return=_maybe_pct(item.get("avg_return")),
            win_rate=_maybe_pct(item.get("win_rate")),
        )
        for horizon, item in (payload.get("horizon_summary") or {}).items()
    )
    if not horizon_rows:
        horizon_rows = "| `N/A` | 0 | 0 | N/A | N/A |"
    return f"""# Paper Stock Target Review Outcomes

Generated at: `{payload.get("generated_at")}`

This is a research-only outcome tracker for manual review rows. It uses local OHLCV data after the review date when available. Pending horizons are not guessed.

## Summary

| Item | Value |
| --- | ---: |
| Status | `{payload.get("status")}` |
| Latest date | {payload.get("latest_date", "N/A")} |
| Outcome rows | {payload.get("outcome_row_count", 0)} |
| Complete / partial / pending | {payload.get("complete_count", 0)} / {payload.get("partial_count", 0)} / {payload.get("pending_count", 0)} |
| Missing entry price | {payload.get("missing_entry_price_count", 0)} |
| Notes CSV | `{payload.get("notes_path", "N/A")}` |
| Broker action | `{payload.get("broker_action")}` |

## Horizon Summary

| Horizon | Evaluable | Pending | Avg return | Win rate |
| --- | ---: | ---: | ---: | ---: |
{horizon_rows}

## Outcome Rows

| Code | Name | Entry date | Manual status | 1D | 5D | 10D | 20D | Action |
| --- | --- | --- | --- | ---: | ---: | ---: | ---: | --- |
{rows}

## Files

- Outcomes CSV: `{output_dir / "stock_target_review_outcomes.csv"}`
- Outcomes JSON: `{output_dir / "stock_target_review_outcomes.json"}`
- Review CSV: `{output_dir / "stock_target_review.csv"}`
- Actions CSV: `{output_dir / "stock_target_review_actions.csv"}`
"""


def _render_stock_target_review_outcomes_history_report(history: pd.DataFrame, payload: dict[str, Any], output_dir: Path) -> str:
    recent = history.copy()
    if not recent.empty:
        recent["_date_sort"] = pd.to_datetime(recent["date"], errors="coerce")
        recent["_priority_sort"] = pd.to_numeric(recent.get("review_priority_score"), errors="coerce")
        recent = recent.sort_values(["_date_sort", "_priority_sort", "code"], ascending=[False, False, True], na_position="last")
        recent = recent.drop(columns=["_date_sort", "_priority_sort"], errors="ignore").head(30)

    if recent.empty:
        rows = "| `N/A` | `N/A` | `N/A` | N/A | `N/A` | N/A | N/A | N/A | N/A | No outcome history rows |"
    else:
        rows = "\n".join(
            "| {date} | `{layer}` | `{code}` | {name} | `{status}` | {r1} | {r5} | {r10} | {r20} | {action} |".format(
                date=row.get("date") or "N/A",
                layer=_md_text(row.get("layer"), 30),
                code=str(row.get("code")).zfill(6),
                name=_md_text(row.get("name"), 40),
                status=_md_text(row.get("outcome_status"), 30),
                r1=_maybe_pct(row.get("return_1d")),
                r5=_maybe_pct(row.get("return_5d")),
                r10=_maybe_pct(row.get("return_10d")),
                r20=_maybe_pct(row.get("return_20d")),
                action=_md_text(row.get("action_code"), 40),
            )
            for _, row in recent.iterrows()
        )

    horizon_rows = "\n".join(
        "| `{horizon}` | {evaluable} | {pending} | {avg_return} | {win_rate} |".format(
            horizon=horizon,
            evaluable=item.get("evaluable_count", 0),
            pending=item.get("pending_count", 0),
            avg_return=_maybe_pct(item.get("avg_return")),
            win_rate=_maybe_pct(item.get("win_rate")),
        )
        for horizon, item in (payload.get("history_horizon_summary") or {}).items()
    )
    if not horizon_rows:
        horizon_rows = "| `N/A` | 0 | 0 | N/A | N/A |"

    return f"""# Paper Stock Target Review Outcome History

Generated at: `{payload.get("generated_at")}`

This is a cumulative research-only outcome tracker. It updates existing rows by review date, layer, and code as future local OHLCV becomes available. It does not connect to brokers, place orders, or provide investment advice.

## Summary

| Item | Value |
| --- | ---: |
| History rows | {payload.get("history_row_count", 0)} |
| Current-run outcome rows | {payload.get("history_current_row_count", payload.get("history_updated_row_count", 0))} |
| Matured history rows refreshed | {payload.get("history_matured_row_count", 0)} |
| Total rows touched | {payload.get("history_updated_row_count", 0)} |
| Updated / inserted rows | {payload.get("history_replaced_row_count", 0)} / {payload.get("history_inserted_row_count", 0)} |
| Complete / partial / pending | {payload.get("history_complete_count", 0)} / {payload.get("history_partial_count", 0)} / {payload.get("history_pending_count", 0)} |
| Missing entry price | {payload.get("history_missing_entry_price_count", 0)} |
| Latest review date | {payload.get("history_latest_review_date", "N/A")} |
| Persistent history CSV | `{payload.get("history_path", "N/A")}` |
| Run snapshot CSV | `{payload.get("history_snapshot_path", "N/A")}` |
| Broker action | `{payload.get("broker_action")}` |

## Horizon Summary

| Horizon | Evaluable | Pending | Avg return | Win rate |
| --- | ---: | ---: | ---: | ---: |
{horizon_rows}

## Recent Outcome Rows

| Review date | Layer | Code | Name | Outcome status | 1D | 5D | 10D | 20D | Action |
| --- | --- | --- | --- | --- | ---: | ---: | ---: | ---: | --- |
{rows}

## Files

- Persistent history CSV: `{payload.get("history_path", "N/A")}`
- Run history snapshot CSV: `{output_dir / "stock_target_review_outcomes_history.csv"}`
- Run history JSON: `{output_dir / "stock_target_review_outcomes_history.json"}`
- Current-run outcomes CSV: `{output_dir / "stock_target_review_outcomes.csv"}`
"""


def _analysis_group_value(value: Any) -> str:
    text = _safe_text(value)
    return text if text else "N/A"


def _stock_review_outcome_analysis_row(group: pd.DataFrame, dimension: str, group_value: str) -> dict[str, Any]:
    row: dict[str, Any] = {
        "dimension": dimension,
        "group_value": group_value,
        "history_row_count": int(len(group)),
    }
    status_counts = group["outcome_status"].value_counts(dropna=False).to_dict() if "outcome_status" in group.columns else {}
    row.update(
        {
            "complete_count": int(status_counts.get("complete", 0)),
            "partial_count": int(status_counts.get("partial", 0)),
            "pending_count": int(status_counts.get("pending", 0)),
            "missing_entry_price_count": int(status_counts.get("missing_entry_price", 0)),
        }
    )
    for horizon in STOCK_TARGET_REVIEW_OUTCOME_HORIZONS:
        return_column = f"return_{horizon}d"
        source_returns = group[return_column] if return_column in group.columns else pd.Series(dtype=float)
        returns = pd.to_numeric(source_returns, errors="coerce").dropna()
        wins = returns[returns > 0]
        losses = returns[returns < 0]
        row[f"evaluable_{horizon}d"] = int(len(returns))
        row[f"coverage_{horizon}d"] = float(len(returns) / len(group)) if len(group) else None
        row[f"avg_return_{horizon}d"] = float(returns.mean()) if not returns.empty else None
        row[f"median_return_{horizon}d"] = float(returns.median()) if not returns.empty else None
        row[f"win_rate_{horizon}d"] = float((returns > 0).mean()) if not returns.empty else None
        row[f"avg_win_{horizon}d"] = float(wins.mean()) if not wins.empty else None
        row[f"avg_loss_{horizon}d"] = float(losses.mean()) if not losses.empty else None
    return row


def _best_outcome_analysis_group(analysis: pd.DataFrame, horizon: int, direction: str) -> dict[str, Any] | None:
    if analysis.empty:
        return None
    avg_column = f"avg_return_{horizon}d"
    eval_column = f"evaluable_{horizon}d"
    if avg_column not in analysis.columns or eval_column not in analysis.columns:
        return None
    evaluable = analysis[pd.to_numeric(analysis[eval_column], errors="coerce") > 0].copy()
    evaluable = evaluable[evaluable["dimension"] != "overall"]
    if evaluable.empty:
        return None
    evaluable["_avg_sort"] = pd.to_numeric(evaluable[avg_column], errors="coerce")
    evaluable = evaluable.dropna(subset=["_avg_sort"])
    if evaluable.empty:
        return None
    ascending = direction == "worst"
    best = evaluable.sort_values(["_avg_sort", eval_column], ascending=[ascending, False]).iloc[0]
    return {
        "dimension": best.get("dimension"),
        "group_value": best.get("group_value"),
        "evaluable_count": int(best.get(eval_column, 0)),
        "avg_return": _optional_float(best.get(avg_column)),
        "win_rate": _optional_float(best.get(f"win_rate_{horizon}d")),
    }


def _estimated_weekday_horizon_date(value: Any, horizon: int) -> str | None:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    current = pd.Timestamp(parsed).normalize()
    steps = 0
    while steps < int(horizon):
        current += pd.Timedelta(days=1)
        if current.weekday() < 5:
            steps += 1
    return current.strftime("%Y-%m-%d")


def _resolve_outcome_maturity_date(row: dict[str, Any], horizon: int) -> str | None:
    future_date = row.get(f"future_date_{horizon}d")
    parsed_future = pd.to_datetime(future_date, errors="coerce")
    if not pd.isna(parsed_future):
        return pd.Timestamp(parsed_future).strftime("%Y-%m-%d")
    return _estimated_weekday_horizon_date(row.get("entry_date") or row.get("date"), horizon)


def _outcome_maturity_forecast(history: pd.DataFrame, min_evaluable: int) -> dict[str, dict[str, Any]]:
    forecast: dict[str, dict[str, Any]] = {}
    for horizon in STOCK_TARGET_REVIEW_OUTCOME_HORIZONS:
        key = f"{horizon}d"
        return_column = f"return_{horizon}d"
        status_column = f"outcome_status_{horizon}d"
        returns = (
            pd.to_numeric(history[return_column], errors="coerce").dropna()
            if not history.empty and return_column in history.columns
            else pd.Series(dtype=float)
        )
        evaluable_count = int(len(returns))
        if history.empty or status_column not in history.columns:
            pending = pd.DataFrame()
        else:
            pending = history[history[status_column].astype(str) == "pending"].copy()
        pending_records = pending.to_dict(orient="records")
        pending_dates = [_resolve_outcome_maturity_date(row, horizon) for row in pending_records]
        estimated_dates = [item for item in pending_dates if item]
        forecast[key] = {
            "pending_count": int(len(pending)),
            "due_pending_count": 0,
            "pending_maturity_dates": pending_dates,
            "evaluable_count": evaluable_count,
            "remaining_to_min_evaluable": max(int(min_evaluable) - evaluable_count, 0),
            "estimated_next_evaluable_date": min(estimated_dates) if estimated_dates else None,
            "estimated_all_pending_mature_by": max(estimated_dates) if estimated_dates else None,
            "estimated_total_after_pending_mature": evaluable_count + int(len(pending)),
            "basis": "weekday_estimate_without_holiday_calendar",
        }
    return forecast


def build_stock_target_review_outcome_analysis(
    history: pd.DataFrame,
    history_payload: dict[str, Any] | None = None,
    min_evaluable: int = DEFAULT_STOCK_REVIEW_OUTCOME_MIN_EVALUABLE,
    min_group_evaluable: int = DEFAULT_STOCK_REVIEW_OUTCOME_MIN_GROUP_EVALUABLE,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Summarize cumulative review outcomes by source tags for model audit."""
    payload_source = history_payload or {}
    min_evaluable = max(1, int(min_evaluable))
    min_group_evaluable = max(1, int(min_group_evaluable))
    rows: list[dict[str, Any]] = []
    if not history.empty:
        rows.append(_stock_review_outcome_analysis_row(history, "overall", "all"))
        for dimension in STOCK_TARGET_REVIEW_OUTCOME_ANALYSIS_DIMENSIONS:
            if dimension == "overall" or dimension not in history.columns:
                continue
            grouped = history.copy()
            grouped["_group_value"] = grouped[dimension].map(_analysis_group_value)
            for group_value, group in grouped.groupby("_group_value", dropna=False):
                rows.append(_stock_review_outcome_analysis_row(group.drop(columns=["_group_value"], errors="ignore"), dimension, str(group_value)))

    analysis = pd.DataFrame(rows, columns=STOCK_TARGET_REVIEW_OUTCOME_ANALYSIS_COLUMNS)
    if not analysis.empty:
        analysis["_dimension_order"] = analysis["dimension"].map(
            {dimension: index for index, dimension in enumerate(STOCK_TARGET_REVIEW_OUTCOME_ANALYSIS_DIMENSIONS)}
        ).fillna(len(STOCK_TARGET_REVIEW_OUTCOME_ANALYSIS_DIMENSIONS))
        analysis = analysis.sort_values(["_dimension_order", "history_row_count", "group_value"], ascending=[True, False, True])
        analysis = analysis.drop(columns=["_dimension_order"], errors="ignore").reset_index(drop=True)

    total_evaluable: dict[str, int] = {}
    horizon_readiness: dict[str, dict[str, Any]] = {}
    ready_horizons: list[str] = []
    overall_rows = analysis.loc[analysis["dimension"] == "overall"] if not analysis.empty else pd.DataFrame()
    non_overall = analysis[analysis["dimension"] != "overall"].copy() if not analysis.empty else pd.DataFrame()
    for horizon in STOCK_TARGET_REVIEW_OUTCOME_HORIZONS:
        key = f"{horizon}d"
        eval_column = f"evaluable_{horizon}d"
        total = int(overall_rows[eval_column].iloc[0]) if not overall_rows.empty and eval_column in overall_rows.columns else 0
        qualified_group_count = (
            int((pd.to_numeric(non_overall.get(eval_column), errors="coerce") >= min_group_evaluable).sum())
            if not non_overall.empty and eval_column in non_overall.columns
            else 0
        )
        if total <= 0:
            status = "waiting_for_evaluable_returns"
        elif total < min_evaluable:
            status = "insufficient_overall_sample"
        elif qualified_group_count <= 0:
            status = "insufficient_group_sample"
        else:
            status = "ready_for_group_review"
            ready_horizons.append(key)
        total_evaluable[key] = total
        horizon_readiness[key] = {
            "status": status,
            "evaluable_count": total,
            "min_evaluable": min_evaluable,
            "qualified_group_count": qualified_group_count,
            "min_group_evaluable": min_group_evaluable,
        }
    if ready_horizons:
        analysis_status = "ready_for_review"
    elif any(value > 0 for value in total_evaluable.values()):
        analysis_status = "sample_insufficient"
    else:
        analysis_status = "waiting_for_evaluable_returns"
    sample_warning = (
        "No horizons have evaluable returns yet."
        if analysis_status == "waiting_for_evaluable_returns"
        else "Sample size is below the configured anti-overfit readiness gates."
        if analysis_status == "sample_insufficient"
        else "At least one horizon meets the configured readiness gates."
    )
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "analysis_status": analysis_status,
        "analysis_row_count": int(len(analysis)),
        "min_evaluable": min_evaluable,
        "min_group_evaluable": min_group_evaluable,
        "source_history_path": payload_source.get("history_path"),
        "source_history_row_count": payload_source.get("history_row_count", int(len(history))),
        "source_history_latest_review_date": payload_source.get("history_latest_review_date"),
        "total_evaluable_by_horizon": total_evaluable,
        "horizon_readiness": horizon_readiness,
        "maturity_forecast": _outcome_maturity_forecast(history, min_evaluable),
        "ready_horizons": ready_horizons,
        "ready_horizon_count": len(ready_horizons),
        "sample_warning": sample_warning,
        "best_groups": {
            f"{horizon}d": _best_outcome_analysis_group(analysis, horizon, "best")
            for horizon in STOCK_TARGET_REVIEW_OUTCOME_HORIZONS
        },
        "worst_groups": {
            f"{horizon}d": _best_outcome_analysis_group(analysis, horizon, "worst")
            for horizon in STOCK_TARGET_REVIEW_OUTCOME_HORIZONS
        },
        "top_analysis_rows": _json_records(analysis.head(20)),
        "research_only": True,
        "broker_action": "none",
        "note": "Outcome analysis summarizes historical review outcomes only; it does not change model selection, target weights, or broker actions.",
    }
    return analysis, payload


def _cached_stock_target_review_outcome_analysis(
    history: pd.DataFrame,
    history_payload: dict[str, Any],
    *,
    min_evaluable: int,
    min_group_evaluable: int,
    cache_dir: Path,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    history_text = history.to_csv(index=False, float_format="%.17g")
    fingerprint_payload = {
        "version": 1,
        "history_sha256": hashlib.sha256(history_text.encode("utf-8")).hexdigest(),
        "min_evaluable": int(min_evaluable),
        "min_group_evaluable": int(min_group_evaluable),
    }
    fingerprint = hashlib.sha256(
        json.dumps(fingerprint_payload, sort_keys=True).encode("utf-8")
    ).hexdigest()
    analysis_path = cache_dir / "outcome_analysis.csv"
    payload_path = cache_dir / "outcome_analysis.json"
    metadata_path = cache_dir / "outcome_analysis_cache.json"
    if analysis_path.exists() and payload_path.exists() and metadata_path.exists():
        metadata = _read_json(metadata_path)
        cached_payload = _read_json(payload_path)
        if metadata.get("fingerprint") == fingerprint and cached_payload:
            try:
                cached_analysis = pd.read_csv(analysis_path, float_precision="round_trip")
            except (OSError, pd.errors.EmptyDataError, pd.errors.ParserError):
                cached_analysis = None
            if cached_analysis is not None:
                return cached_analysis, cached_payload

    analysis, payload = build_stock_target_review_outcome_analysis(
        history,
        history_payload,
        min_evaluable=min_evaluable,
        min_group_evaluable=min_group_evaluable,
    )
    write_text_if_changed(analysis_path, analysis.to_csv(index=False, float_format="%.17g"))
    write_text_if_changed(payload_path, json.dumps(payload, ensure_ascii=False, indent=2))
    write_text_if_changed(
        metadata_path,
        json.dumps({"version": 1, "fingerprint": fingerprint}, ensure_ascii=False, indent=2),
    )
    return analysis, payload


def _calendar_reference_date(value: Any) -> str | None:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return pd.Timestamp(parsed).strftime("%Y-%m-%d")


def build_stock_target_review_outcome_calendar(
    analysis_payload: dict[str, Any],
    as_of_date: Any | None = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Build a standalone outcome maturity calendar from outcome-analysis metadata."""
    forecast = analysis_payload.get("maturity_forecast") or {}
    readiness = analysis_payload.get("horizon_readiness") or {}
    analysis_status = str(analysis_payload.get("analysis_status") or "unknown")
    reference_date = _calendar_reference_date(as_of_date)
    reference_ts = pd.to_datetime(reference_date, errors="coerce") if reference_date else pd.NaT
    rows: list[dict[str, Any]] = []
    next_candidates: list[tuple[pd.Timestamp, str]] = []
    due_candidates: list[tuple[pd.Timestamp, str]] = []

    for horizon in STOCK_TARGET_REVIEW_OUTCOME_HORIZONS:
        key = f"{horizon}d"
        forecast_item = forecast.get(key, {}) or {}
        readiness_item = readiness.get(key, {}) or {}
        pending_count = int(_as_float(forecast_item.get("pending_count"), 0.0))
        due_pending_count = int(_as_float(forecast_item.get("due_pending_count"), 0.0))
        evaluable_count = int(_as_float(forecast_item.get("evaluable_count"), 0.0))
        remaining = int(_as_float(forecast_item.get("remaining_to_min_evaluable"), 0.0))
        next_date = forecast_item.get("estimated_next_evaluable_date")
        all_by = forecast_item.get("estimated_all_pending_mature_by")
        total_after_pending = int(_as_float(forecast_item.get("estimated_total_after_pending_mature"), evaluable_count))
        readiness_status = str(readiness_item.get("status") or "unknown")
        maturity_dates = [
            pd.to_datetime(item, errors="coerce") for item in (forecast_item.get("pending_maturity_dates") or [])
        ]
        mature_dates = [
            item
            for item in maturity_dates
            if pd.notna(item) and pd.Timestamp(item).normalize() <= pd.Timestamp(reference_ts).normalize()
        ] if pd.notna(reference_ts) else []
        if due_pending_count <= 0 and mature_dates:
            due_pending_count = int(len(mature_dates))

        next_ts = pd.to_datetime(next_date, errors="coerce")
        days_until_next = None
        if pd.notna(reference_ts) and pd.notna(next_ts):
            days_until_next = int((pd.Timestamp(next_ts).normalize() - pd.Timestamp(reference_ts).normalize()).days)

        maturity_state = "pending_maturity"
        if due_pending_count > 0:
            maturity_state = "due_for_recheck"
            if mature_dates:
                earliest_due = pd.Timestamp(min(mature_dates))
                days_until_next = int((earliest_due.normalize() - pd.Timestamp(reference_ts).normalize()).days)
                next_ts = earliest_due
        if pending_count <= 0:
            maturity_state = "no_pending"
        elif days_until_next is None:
            maturity_state = "pending_date_unknown"
        elif due_pending_count > 0:
            maturity_state = "due_for_recheck"
        elif days_until_next <= 0:
            maturity_state = "due_for_recheck"
        else:
            maturity_state = "pending_maturity"
        row = {
            "as_of_date": reference_date,
            "horizon": key,
            "analysis_status": analysis_status,
            "readiness_status": readiness_status,
            "maturity_state": maturity_state,
            "pending_count": pending_count,
            "due_pending_count": due_pending_count,
            "evaluable_count": evaluable_count,
            "remaining_to_min_evaluable": remaining,
            "estimated_next_evaluable_date": next_date,
            "days_until_next_evaluable": days_until_next,
            "estimated_all_pending_mature_by": all_by,
            "estimated_total_after_pending_mature": total_after_pending,
            "min_evaluable": int(_as_float(readiness_item.get("min_evaluable"), _as_float(analysis_payload.get("min_evaluable"), 0.0))),
            "qualified_group_count": int(_as_float(readiness_item.get("qualified_group_count"), 0.0)),
            "min_group_evaluable": int(
                _as_float(readiness_item.get("min_group_evaluable"), _as_float(analysis_payload.get("min_group_evaluable"), 0.0))
            ),
            "basis": forecast_item.get("basis") or "N/A",
            "research_only": True,
            "broker_action": "none",
        }
        rows.append(row)

        if pending_count > 0 and next_date:
            parsed = pd.to_datetime(next_date, errors="coerce")
            if pd.notna(parsed):
                next_candidates.append((pd.Timestamp(parsed), key))
                if maturity_state == "due_for_recheck":
                    due_candidates.append((pd.Timestamp(parsed), key))

    calendar = pd.DataFrame(rows, columns=STOCK_TARGET_REVIEW_OUTCOME_CALENDAR_COLUMNS)
    next_action_date = None
    next_action_horizon = None
    if next_candidates:
        next_timestamp, next_action_horizon = min(
            next_candidates,
            key=lambda item: (item[0], int(str(item[1]).rstrip("d"))),
        )
        next_action_date = next_timestamp.strftime("%Y-%m-%d")

    next_due_date = None
    next_due_horizon = None
    if due_candidates:
        due_timestamp, next_due_horizon = min(
            due_candidates,
            key=lambda item: (item[0], int(str(item[1]).rstrip("d"))),
        )
        next_due_date = due_timestamp.strftime("%Y-%m-%d")

    calendar_ready_count = int((calendar["readiness_status"] == "ready_for_group_review").sum()) if not calendar.empty else 0
    calendar_pending_count = int(calendar["pending_count"].sum()) if not calendar.empty else 0
    calendar_due_count = int((calendar["maturity_state"] == "due_for_recheck").sum()) if not calendar.empty else 0
    calendar_due_pending_count = int(
        calendar.loc[calendar["maturity_state"] == "due_for_recheck", "due_pending_count"].sum()
    ) if calendar_due_count else 0
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "as_of_date": reference_date,
        "analysis_status": analysis_status,
        "calendar_status": (
            "due_for_recheck"
            if calendar_due_count > 0
            else "ready_for_review"
            if calendar_ready_count > 0
            else "pending_maturity"
            if calendar_pending_count > 0
            else analysis_status
        ),
        "ready_horizon_count": int(analysis_payload.get("ready_horizon_count") or calendar_ready_count),
        "calendar_row_count": int(len(calendar)),
        "calendar_ready_count": calendar_ready_count,
        "calendar_pending_count": calendar_pending_count,
        "calendar_due_count": calendar_due_count,
        "calendar_due_pending_count": calendar_due_pending_count,
        "next_action_date": next_action_date,
        "next_action_horizon": next_action_horizon,
        "next_due_date": next_due_date,
        "next_due_horizon": next_due_horizon,
        "source_history_path": analysis_payload.get("source_history_path"),
        "source_history_row_count": analysis_payload.get("source_history_row_count"),
        "source_history_latest_review_date": analysis_payload.get("source_history_latest_review_date"),
        "min_evaluable": analysis_payload.get("min_evaluable"),
        "min_group_evaluable": analysis_payload.get("min_group_evaluable"),
        "calendar_rows": _json_records(calendar),
        "research_only": True,
        "broker_action": "none",
        "note": "Outcome calendar is a research-only maturity schedule for later review; it does not change model selection, target weights, or broker actions.",
    }
    return calendar, payload


def build_stock_target_review_outcome_due_queue(
    calendar: pd.DataFrame,
    calendar_payload: dict[str, Any],
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Filter the outcome calendar to horizons that are due but still pending."""
    if calendar.empty or "maturity_state" not in calendar.columns:
        due = pd.DataFrame(columns=STOCK_TARGET_REVIEW_OUTCOME_DUE_COLUMNS)
    else:
        due = calendar[calendar["maturity_state"].astype(str) == "due_for_recheck"].copy()
        if not due.empty:
            due["_next_sort"] = pd.to_datetime(due["estimated_next_evaluable_date"], errors="coerce")
            due["_horizon_sort"] = pd.to_numeric(due["horizon"].astype(str).str.rstrip("d"), errors="coerce")
            due = due.sort_values(["_next_sort", "_horizon_sort"], ascending=[True, True]).reset_index(drop=True)
            due.insert(0, "due_rank", range(1, len(due) + 1))
            due = due.drop(columns=["_next_sort", "_horizon_sort"], errors="ignore")
        else:
            due = pd.DataFrame(columns=STOCK_TARGET_REVIEW_OUTCOME_DUE_COLUMNS)
    if not due.empty:
        due = due.reindex(columns=STOCK_TARGET_REVIEW_OUTCOME_DUE_COLUMNS)

    next_due_date = None
    next_due_horizon = None
    if not due.empty:
        next_due_date = due["estimated_next_evaluable_date"].iloc[0]
        next_due_horizon = due["horizon"].iloc[0]

    due_pending_count = int(due["due_pending_count"].sum()) if not due.empty else 0
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "as_of_date": calendar_payload.get("as_of_date"),
        "due_status": "due_for_recheck" if not due.empty else "no_due_horizons",
        "due_row_count": int(len(due)),
        "due_pending_count": due_pending_count,
        "next_due_date": next_due_date,
        "next_due_horizon": next_due_horizon,
        "calendar_status": calendar_payload.get("calendar_status"),
        "source_history_path": calendar_payload.get("source_history_path"),
        "source_history_row_count": calendar_payload.get("source_history_row_count"),
        "source_history_latest_review_date": calendar_payload.get("source_history_latest_review_date"),
        "due_rows": _json_records(due),
        "research_only": True,
        "broker_action": "none",
        "note": "Outcome due queue lists maturity horizons that should be rechecked from local OHLCV; it does not change model selection, target weights, or broker actions.",
    }
    return due, payload


def _render_outcome_readiness_rows(payload: dict[str, Any]) -> str:
    rows: list[str] = []
    for horizon, item in (payload.get("horizon_readiness") or {}).items():
        rows.append(
            "| `{horizon}` | `{status}` | {evaluable} | {min_evaluable} | {qualified} | {min_group} |".format(
                horizon=horizon,
                status=item.get("status", "N/A"),
                evaluable=item.get("evaluable_count", 0),
                min_evaluable=item.get("min_evaluable", "N/A"),
                qualified=item.get("qualified_group_count", 0),
                min_group=item.get("min_group_evaluable", "N/A"),
            )
        )
    return "\n".join(rows) if rows else "| `N/A` | `N/A` | 0 | N/A | 0 | N/A |"


def _render_outcome_maturity_forecast_rows(payload: dict[str, Any]) -> str:
    rows: list[str] = []
    for horizon, item in (payload.get("maturity_forecast") or {}).items():
        rows.append(
            "| `{horizon}` | {pending} | {evaluable} | {remaining} | {next_date} | {all_by} | `{basis}` |".format(
                horizon=horizon,
                pending=item.get("pending_count", 0),
                evaluable=item.get("evaluable_count", 0),
                remaining=item.get("remaining_to_min_evaluable", 0),
                next_date=item.get("estimated_next_evaluable_date") or "N/A",
                all_by=item.get("estimated_all_pending_mature_by") or "N/A",
                basis=item.get("basis", "N/A"),
            )
        )
    return "\n".join(rows) if rows else "| `N/A` | 0 | 0 | 0 | N/A | N/A | `N/A` |"


def _render_stock_target_review_outcome_calendar_report(calendar: pd.DataFrame, payload: dict[str, Any], output_dir: Path) -> str:
    if calendar.empty:
        rows = "| `N/A` | `N/A` | `N/A` | 0 | 0 | 0 | N/A | N/A | N/A | `N/A` |"
    else:
        rows = "\n".join(
            "| `{horizon}` | `{status}` | `{maturity}` | {pending} | {evaluable} | {remaining} | {next_date} | {days} | {all_by} | `{basis}` |".format(
                horizon=_md_text(row.get("horizon"), 20),
                status=_md_text(row.get("readiness_status"), 40),
                maturity=_md_text(row.get("maturity_state"), 40),
                pending=int(_as_float(row.get("pending_count"), 0.0)),
                evaluable=int(_as_float(row.get("evaluable_count"), 0.0)),
                remaining=int(_as_float(row.get("remaining_to_min_evaluable"), 0.0)),
                next_date=row.get("estimated_next_evaluable_date") or "N/A",
                days=row.get("days_until_next_evaluable") if pd.notna(row.get("days_until_next_evaluable")) else "N/A",
                all_by=row.get("estimated_all_pending_mature_by") or "N/A",
                basis=_md_text(row.get("basis"), 60),
            )
            for _, row in calendar.iterrows()
        )

    return f"""# Paper Stock Target Review Outcome Calendar

Generated at: `{payload.get("generated_at")}`

This is a research-only outcome maturity calendar. It estimates when pending 1/5/10/20 trading-day review outcomes can be rechecked from local OHLCV data; it does not connect to brokers, place orders, or change model targets.

## Summary

| Item | Value |
| --- | ---: |
| Calendar status | `{payload.get("calendar_status")}` |
| Analysis status | `{payload.get("analysis_status")}` |
| As-of date | {payload.get("as_of_date") or "N/A"} |
| Calendar rows | {payload.get("calendar_row_count", 0)} |
| Ready horizons | {payload.get("calendar_ready_count", 0)} |
| Pending horizon rows | {payload.get("calendar_pending_count", 0)} |
| Due horizons | {payload.get("calendar_due_count", 0)} |
| Due pending rows | {payload.get("calendar_due_pending_count", 0)} |
| Next action date | {payload.get("next_action_date") or "N/A"} |
| Next action horizon | `{payload.get("next_action_horizon") or "N/A"}` |
| Next due date | {payload.get("next_due_date") or "N/A"} |
| Next due horizon | `{payload.get("next_due_horizon") or "N/A"}` |
| Source history rows | {payload.get("source_history_row_count", 0)} |
| Source latest review date | {payload.get("source_history_latest_review_date", "N/A")} |
| Min evaluable rows | {payload.get("min_evaluable", "N/A")} |
| Min group evaluable rows | {payload.get("min_group_evaluable", "N/A")} |
| Broker action | `{payload.get("broker_action")}` |

## Calendar

| Horizon | Readiness | Maturity state | Pending rows | Evaluable now | Remaining to min | Estimated next evaluable date | Days until next | Estimated all pending mature by | Basis |
| --- | --- | --- | ---: | ---: | ---: | --- | ---: | --- | --- |
{rows}

## Files

- Calendar CSV: `{output_dir / "stock_target_review_outcome_calendar.csv"}`
- Calendar JSON: `{output_dir / "stock_target_review_outcome_calendar.json"}`
- Source analysis CSV: `{output_dir / "stock_target_review_outcome_analysis.csv"}`
- Source history CSV: `{payload.get("source_history_path", "N/A")}`
"""


def _render_stock_target_review_outcome_due_report(due: pd.DataFrame, payload: dict[str, Any], output_dir: Path) -> str:
    if due.empty:
        rows = "| N/A | `N/A` | 0 | N/A | N/A | No due outcome horizons |"
    else:
        rows = "\n".join(
            "| {rank} | `{horizon}` | {pending} | {next_date} | {days} | {action} |".format(
                rank=int(_as_float(row.get("due_rank"), 0.0)),
                horizon=_md_text(row.get("horizon"), 20),
                pending=int(_as_float(row.get("due_pending_count"), _as_float(row.get("pending_count"), 0.0))),
                next_date=row.get("estimated_next_evaluable_date") or "N/A",
                days=row.get("days_until_next_evaluable") if pd.notna(row.get("days_until_next_evaluable")) else "N/A",
                action="refresh_local_ohlcv_and_rerun_outcomes",
            )
            for _, row in due.iterrows()
        )

    return f"""# Paper Stock Target Review Outcome Due Queue

Generated at: `{payload.get("generated_at")}`

This is a research-only due queue for outcome maturity checks. It flags pending horizons whose estimated maturity date is on or before the local paper-account as-of date; it does not connect to brokers, place orders, or change model targets.

## Summary

| Item | Value |
| --- | ---: |
| Due status | `{payload.get("due_status")}` |
| As-of date | {payload.get("as_of_date") or "N/A"} |
| Due horizons | {payload.get("due_row_count", 0)} |
| Due pending rows | {payload.get("due_pending_count", 0)} |
| Next due date | {payload.get("next_due_date") or "N/A"} |
| Next due horizon | `{payload.get("next_due_horizon") or "N/A"}` |
| Calendar status | `{payload.get("calendar_status")}` |
| Source history rows | {payload.get("source_history_row_count", 0)} |
| Source latest review date | {payload.get("source_history_latest_review_date", "N/A")} |
| Broker action | `{payload.get("broker_action")}` |

## Due Queue

| Rank | Horizon | Pending rows | Estimated maturity date | Days until next | Research action |
| ---: | --- | ---: | --- | ---: | --- |
{rows}

## Files

- Due CSV: `{output_dir / "stock_target_review_outcome_due.csv"}`
- Due JSON: `{output_dir / "stock_target_review_outcome_due.json"}`
- Calendar CSV: `{output_dir / "stock_target_review_outcome_calendar.csv"}`
- Source history CSV: `{payload.get("source_history_path", "N/A")}`
"""


def _render_stock_target_review_outcome_analysis_report(analysis: pd.DataFrame, payload: dict[str, Any], output_dir: Path) -> str:
    if analysis.empty:
        rows = "| `N/A` | `N/A` | 0 | 0 | N/A | N/A | N/A | N/A | No outcome analysis rows |"
    else:
        display = analysis.head(80)
        rows = "\n".join(
            "| `{dimension}` | `{group}` | {count} | {eval1} | {avg1} | {win1} | {eval5} | {avg5} | {win5} |".format(
                dimension=_md_text(row.get("dimension"), 40),
                group=_md_text(row.get("group_value"), 60),
                count=int(_as_float(row.get("history_row_count"), 0.0)),
                eval1=int(_as_float(row.get("evaluable_1d"), 0.0)),
                avg1=_maybe_pct(row.get("avg_return_1d")),
                win1=_maybe_pct(row.get("win_rate_1d")),
                eval5=int(_as_float(row.get("evaluable_5d"), 0.0)),
                avg5=_maybe_pct(row.get("avg_return_5d")),
                win5=_maybe_pct(row.get("win_rate_5d")),
            )
            for _, row in display.iterrows()
        )

    horizon_rows = "\n".join(
        f"| `{horizon}` | {count} |"
        for horizon, count in (payload.get("total_evaluable_by_horizon") or {}).items()
    )
    if not horizon_rows:
        horizon_rows = "| `N/A` | 0 |"

    best_rows: list[str] = []
    for horizon, best in (payload.get("best_groups") or {}).items():
        worst = (payload.get("worst_groups") or {}).get(horizon)
        if best is None and worst is None:
            best_rows.append(f"| `{horizon}` | N/A | N/A |")
            continue
        best_text = (
            "{dimension}/{group} ({avg}, win {win}, n={count})".format(
                dimension=_md_text(best.get("dimension"), 30),
                group=_md_text(best.get("group_value"), 40),
                avg=_maybe_pct(best.get("avg_return")),
                win=_maybe_pct(best.get("win_rate")),
                count=best.get("evaluable_count", 0),
            )
            if best
            else "N/A"
        )
        worst_text = (
            "{dimension}/{group} ({avg}, win {win}, n={count})".format(
                dimension=_md_text(worst.get("dimension"), 30),
                group=_md_text(worst.get("group_value"), 40),
                avg=_maybe_pct(worst.get("avg_return")),
                win=_maybe_pct(worst.get("win_rate")),
                count=worst.get("evaluable_count", 0),
            )
            if worst
            else "N/A"
        )
        best_rows.append(f"| `{horizon}` | {best_text} | {worst_text} |")
    best_table = "\n".join(best_rows) if best_rows else "| `N/A` | N/A | N/A |"

    return f"""# Paper Stock Target Review Outcome Analysis

Generated at: `{payload.get("generated_at")}`

This is a research-only historical effectiveness view. It groups cumulative review outcomes by source tags and does not connect to brokers, place orders, or change model targets.

## Summary

| Item | Value |
| --- | ---: |
| Analysis status | `{payload.get("analysis_status")}` |
| Analysis rows | {payload.get("analysis_row_count", 0)} |
| Source history rows | {payload.get("source_history_row_count", 0)} |
| Source latest review date | {payload.get("source_history_latest_review_date", "N/A")} |
| Min evaluable rows | {payload.get("min_evaluable", "N/A")} |
| Min group evaluable rows | {payload.get("min_group_evaluable", "N/A")} |
| Ready horizons | {payload.get("ready_horizon_count", 0)} |
| Sample warning | {payload.get("sample_warning", "N/A")} |
| Source history CSV | `{payload.get("source_history_path", "N/A")}` |
| Broker action | `{payload.get("broker_action")}` |

## Evaluable Samples

| Horizon | Evaluable rows |
| --- | ---: |
{horizon_rows}

## Readiness Gates

| Horizon | Status | Evaluable | Min evaluable | Qualified groups | Min group rows |
| --- | --- | ---: | ---: | ---: | ---: |
{_render_outcome_readiness_rows(payload)}

## Maturity Forecast

| Horizon | Pending rows | Evaluable now | Remaining to min | Estimated next evaluable date | Estimated all pending mature by | Basis |
| --- | ---: | ---: | ---: | --- | --- | --- |
{_render_outcome_maturity_forecast_rows(payload)}

## Best And Worst Groups

| Horizon | Highest avg return group | Lowest avg return group |
| --- | --- | --- |
{best_table}

## Group Analysis

| Dimension | Group | Rows | 1D n | 1D avg | 1D win | 5D n | 5D avg | 5D win |
| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: |
{rows}

## Files

- Analysis CSV: `{output_dir / "stock_target_review_outcome_analysis.csv"}`
- Analysis JSON: `{output_dir / "stock_target_review_outcome_analysis.json"}`
- Outcome history CSV: `{payload.get("source_history_path", "N/A")}`
"""


def _render_stock_target_review_report(review: pd.DataFrame, payload: dict[str, Any], output_dir: Path) -> str:
    if review.empty:
        rows = "| N/A | `N/A` | `N/A` | `N/A` | N/A | N/A | N/A | N/A | N/A | `N/A` | `N/A` | N/A | No stock target review rows found |"
    else:
        rows = "\n".join(
            "| {rank} | `{bucket}` | `{stage}` | `{code}` | {name} | {weight} | {cap} | {ret} | {score} | `{manual_status}` | `{manual_state}` | {manual_note} | {reason} |".format(
                rank=int(row["review_rank"]),
                bucket=_md_text(row.get("review_bucket"), 40),
                stage=_md_text(row.get("review_stage"), 30),
                code=str(row.get("code")).zfill(6),
                name=_md_text(row.get("name"), 40),
                weight=_pct(row.get("portfolio_target_weight")),
                cap=_num(row.get("market_cap_yi"), 2) if _optional_float(row.get("market_cap_yi")) is not None else "N/A",
                ret=_pct(row.get("unrealized_return")),
                score=_num(row.get("review_priority_score"), 2),
                manual_status=_md_text(row.get("manual_status"), 40),
                manual_state=_md_text(row.get("manual_review_state"), 40),
                manual_note=_md_text(row.get("manual_note"), 80),
                reason=_md_text(row.get("review_reason"), 140),
            )
            for _, row in review.iterrows()
        )
    return f"""# Paper Stock Target Review

Generated at: `{payload.get("generated_at")}`

This is a research-only target-stock review panel. It ranks items for human review and does not connect to brokers, place orders, or provide investment advice.

## Summary

| Item | Value |
| --- | ---: |
| Status | `{payload.get("status")}` |
| Latest date | {payload.get("latest_date", "N/A")} |
| Review rows | {payload.get("review_row_count", 0)} |
| Large-cap rows skipped from tracking | {payload.get("review_skipped_tracking_excluded_count", 0)} |
| Max tracking market cap | {payload.get("stock_tracking_max_market_cap_yi", "N/A")} Yi Yuan |
| Market-cap cache status | `{payload.get("stock_market_cap_cache_status", "N/A")}` |
| Review required | {payload.get("review_required_count", 0)} |
| Monitor | {payload.get("monitor_count", 0)} |
| Routine | {payload.get("routine_count", 0)} |
| Trigger reviews | {payload.get("trigger_review_count", 0)} |
| Drawdown reviews | {payload.get("drawdown_review_count", 0)} |
| Suppressed-layer reviews | {payload.get("suppressed_layer_review_count", 0)} |
| Watch reviews | {payload.get("watch_review_count", 0)} |
| Manual reviewed current rows | {payload.get("manual_note_count", 0)} |
| Unreviewed current rows | {payload.get("unreviewed_count", 0)} |
| Review-required unreviewed rows | {payload.get("review_required_unreviewed_count", 0)} |
| Manual watch rows | {payload.get("manual_watch_count", 0)} |
| Manual resolved rows | {payload.get("manual_resolved_count", 0)} |
| Manual exclude-candidate rows | {payload.get("manual_exclude_candidate_count", 0)} |
| Manual other-status rows | {payload.get("manual_other_status_count", 0)} |
| Persistent notes file | `{payload.get("notes_path", "N/A")}` |
| Drawdown threshold | {_pct(payload.get("drawdown_threshold"))} |
| Watch drawdown threshold | {_pct(payload.get("watch_drawdown_threshold"))} |
| Loss attention threshold | {_pct(payload.get("loss_attention_threshold"))} |
| Trigger matches | {payload.get("trigger_match_count", 0)} |
| Broker action | `{payload.get("broker_action")}` |

## Review Queue

| Rank | Bucket | Stage | Code | Name | Target weight | Market cap Yi | Unrealized return | Score | Manual status | Manual state | Manual note | Reason |
| ---: | --- | --- | --- | --- | ---: | ---: | ---: | ---: | --- | --- | --- | --- |
{rows}

## Files

- Review CSV: `{output_dir / "stock_target_review.csv"}`
- Review JSON: `{output_dir / "stock_target_review.json"}`
- Persistent notes CSV: `{payload.get("notes_path", "N/A")}`
- Run notes snapshot CSV: `{payload.get("notes_snapshot_path", "N/A")}`
- Source stock targets: `{output_dir / "stock_targets.csv"}`
"""


def _render_stock_targets_report(stock_targets: pd.DataFrame, payload: dict[str, Any], output_dir: Path) -> str:
    if stock_targets.empty:
        rows = "| `N/A` | `N/A` | `N/A` | N/A | N/A | N/A | `N/A` | `N/A` | `N/A` | `N/A` | No stock-level source positions found |"
    else:
        rows = "\n".join(
            "| `{layer}` | `{code}` | {name} | {weight} | {value} | {cap} | {ret} | `{action}` | `{tracking}` | `{risk}` | {trigger} | {why} |".format(
                layer=str(row["layer"]),
                code=str(row["code"]).zfill(6),
                name=_md_text(row["name"], 40),
                weight=_pct(row["portfolio_target_weight"]),
                value=_num(row["portfolio_target_value"], 2),
                cap=_num(row.get("market_cap_yi"), 2) if _optional_float(row.get("market_cap_yi")) is not None else "N/A",
                ret=_pct(row.get("unrealized_return")),
                action=str(row["target_action"]),
                tracking=str(row.get("tracking_rule_status", "N/A")),
                risk=str(row.get("risk_filter_status", "N/A")),
                trigger=_md_text(row.get("trigger_summary"), 80),
                why=_md_text(row.get("selection_explanation"), 140),
            )
            for _, row in stock_targets.iterrows()
        )
    source_rows = "\n".join(
        "| `{layer}` | {count} | {weight} | `{profile}` | `{config}` | {risk} | `{path}` |".format(
            layer=str(item.get("layer")),
            count=item.get("open_position_count", 0),
            weight=_pct(item.get("layer_target_weight")),
            profile=_md_text(item.get("source_strategy_profile"), 60),
            config=str(item.get("source_config_status")),
            risk=_md_text(item.get("source_risk_explanation"), 120),
            path=str(item.get("trades_path")),
        )
        for item in payload.get("source_layers", [])
    )
    if not source_rows:
        source_rows = "| `N/A` | 0 | N/A | `N/A` | `missing_sources` | N/A | `N/A` |"
    return f"""# Paper Stock Targets

Generated at: `{payload.get("generated_at")}`

This is a research-only stock-level target decomposition from local backtest trades. It does not connect to brokers, place orders, or provide investment advice.

## Summary

| Item | Value |
| --- | ---: |
| Status | `{payload.get("status")}` |
| Latest date | {payload.get("latest_date", "N/A")} |
| Stock target rows | {payload.get("stock_target_count", 0)} |
| Source rows before review exclusion | {payload.get("source_stock_target_count", payload.get("stock_target_count", 0))} |
| Active stock targets | {payload.get("active_stock_target_count", 0)} |
| Suppressed stock rows | {payload.get("suppressed_stock_count", 0)} |
| Review-required rows excluded | {payload.get("review_required_excluded_count", 0)} |
| Large-cap rows excluded from tracking | {payload.get("stock_tracking_excluded_large_market_cap_count", 0)} |
| Max tracking market cap | {payload.get("stock_tracking_max_market_cap_yi", "N/A")} Yi Yuan |
| Market-cap missing rows | {payload.get("stock_tracking_market_cap_missing_count", 0)} |
| Market-cap cache status | `{payload.get("stock_market_cap_cache_status", "N/A")}` |
| Market-cap cache path | `{payload.get("stock_market_cap_path", "N/A")}` |
| Total portfolio target weight | {_pct(payload.get("total_portfolio_target_weight"))} |
| Trigger signal status | `{payload.get("trigger_signal_status")}` |
| Latest trigger run time | `{payload.get("latest_trigger_run_time")}` |
| Trigger signal rows | {payload.get("trigger_signal_count", 0)} |
| Matched target stocks | {payload.get("trigger_match_count", 0)} |
| Broker action | `{payload.get("broker_action")}` |

## Stock Targets

| Layer | Code | Name | Target weight | Target value | Market cap Yi | Unrealized return | Action | Tracking | Risk/filter | Trigger monitor | Why selected |
| --- | --- | --- | ---: | ---: | ---: | ---: | --- | --- | --- | --- | --- |
{rows}

## Source Layers

| Layer | Open positions | Layer target weight | Strategy profile | Config | Source risk | Trades path |
| --- | ---: | ---: | --- | --- | --- | --- |
{source_rows}

## Files

- Stock targets CSV: `{output_dir / "stock_targets.csv"}`
- Stock targets JSON: `{output_dir / "stock_targets.json"}`
- Latest trigger signals: `{payload.get("latest_trigger_signal_path")}`
"""


def _render_report(metrics: dict[str, Any], output_dir: Path, allocator_dir: Path, config_path: Path) -> str:
    return f"""# Paper Account Ledger

Generated at: `{datetime.now().isoformat(timespec="seconds")}`

This is a research-only simulation ledger. It does not connect to brokers, place orders, or provide investment advice.

## Current State

| Item | Value |
| --- | ---: |
| Latest date | {metrics.get("latest_date", "N/A")} |
| Latest window | {metrics.get("latest_window", "N/A")} |
| Latest allocator candidate | {metrics.get("latest_candidate", "N/A")} |
| Latest effective regime | {metrics.get("latest_regime", "N/A")} |
| Core target weight | {_pct(metrics.get("latest_core_weight"))} |
| Satellite target weight | {_pct(metrics.get("latest_satellite_weight"))} |
| Cash target weight | {_pct(metrics.get("latest_cash_weight"))} |
| Current drawdown | {_pct(metrics.get("current_drawdown"))} |

## Simulation Metrics

| Metric | Value |
| --- | ---: |
| Initial cash | {_num(metrics.get("initial_cash"), 2)} |
| Final equity | {_num(metrics.get("final_equity"), 2)} |
| Total return | {_pct(metrics.get("total_return"))} |
| CAGR | {_pct(metrics.get("cagr"))} |
| Max drawdown | {_pct(metrics.get("max_drawdown"))} |
| Sharpe | {_num(metrics.get("sharpe"), 3)} |
| Average core weight | {_pct(metrics.get("average_core_weight"))} |
| Average satellite weight | {_pct(metrics.get("average_satellite_weight"))} |
| Satellite active day ratio | {_pct(metrics.get("satellite_active_day_ratio"))} |
| Target-change turnover | {_num(metrics.get("total_target_change_turnover"), 3)} |
| Estimated allocation fee | {_num(metrics.get("total_estimated_fee"), 2)} |
| Rebalance cost rate | {_pct(metrics.get("rebalance_cost_rate"))} |
| Audit events | {metrics.get("audit_event_count", 0)} |

## Notes

- The ledger uses the allocator's selected rolling windows and target weights.
- Underlying core and satellite curves are treated as already cost-adjusted by their own backtests.
- `estimated_fee` only applies the optional portfolio allocation-level `rebalance_cost_rate` to target-weight changes.
- Events are logged when the window, selected candidate, regime, satellite filter state, or target weight changes.

## Files

- Ledger: `{output_dir / "ledger.csv"}`
- Rebalance audit: `{output_dir / "rebalance_audit.csv"}`
- Monthly returns: `{output_dir / "monthly_returns.csv"}`
- Target holdings: `{output_dir / "target_holdings.csv"}`
- Target holdings report: `{output_dir / "target_holdings.md"}`
- Stock targets: `{output_dir / "stock_targets.csv"}`
- Stock targets report: `{output_dir / "stock_targets.md"}`
- Stock target review: `{output_dir / "stock_target_review.csv"}`
- Stock target review report: `{output_dir / "stock_target_review.md"}`
- Stock target review notes: `{metrics.get("stock_target_review_notes_path")}`
- Stock target review notes snapshot: `{metrics.get("stock_target_review_notes_snapshot_path")}`
- Stock target review actions: `{metrics.get("stock_target_review_actions_path")}`
- Stock target review actions report: `{metrics.get("stock_target_review_actions_report_path")}`
- Stock target review assistant: `{metrics.get("stock_target_review_assistant_path")}`
- Stock target review assistant report: `{metrics.get("stock_target_review_assistant_report_path")}`
- Stock target review decision template: `{metrics.get("stock_target_review_decision_template_path")}`
- Stock target review decision template report: `{metrics.get("stock_target_review_decision_template_report_path")}`
- Stock target review decision template XLSX: `{metrics.get("stock_target_review_decision_template_xlsx_path")}`
- Stock target review outcomes: `{metrics.get("stock_target_review_outcomes_path")}`
- Stock target review outcomes report: `{metrics.get("stock_target_review_outcomes_report_path")}`
- Stock target review outcome history: `{metrics.get("stock_target_review_outcomes_history_path")}`
- Stock target review outcome history report: `{metrics.get("stock_target_review_outcomes_history_report_path")}`
- Stock target review outcome analysis: `{metrics.get("stock_target_review_outcome_analysis_path")}`
- Stock target review outcome analysis report: `{metrics.get("stock_target_review_outcome_analysis_report_path")}`
- Stock target review outcome calendar: `{metrics.get("stock_target_review_outcome_calendar_path")}`
- Stock target review outcome calendar report: `{metrics.get("stock_target_review_outcome_calendar_report_path")}`
- Stock target review outcome due queue: `{metrics.get("stock_target_review_outcome_due_path")}`
- Stock target review outcome due report: `{metrics.get("stock_target_review_outcome_due_report_path")}`
- Metrics: `{output_dir / "metrics.json"}`
- Portfolio config: `{config_path}`
- Allocator directory: `{allocator_dir}`
"""


def run_paper_account(
    project_root: str | Path = Path("."),
    config_path: str | Path | None = None,
    allocator_dir: str | Path | None = None,
    output_dir: str | Path | None = None,
    rebalance_cost_rate: float = 0.0,
    trigger_signal_path: str | Path | None = None,
    stock_market_cap_path: str | Path | None = None,
    stock_tracking_max_market_cap_yi: float = DEFAULT_STOCK_TRACKING_MAX_MARKET_CAP_YI,
    stock_review_notes_path: str | Path | None = None,
    stock_review_outcomes_history_path: str | Path | None = None,
    stock_review_drawdown_threshold: float = -0.10,
    stock_review_watch_drawdown_threshold: float = -0.07,
    stock_review_loss_attention_threshold: float = -0.05,
    stock_review_gain_attention_threshold: float = 0.10,
    stock_review_watch_score_threshold: float = 30.0,
    stock_review_outcome_min_evaluable: int = DEFAULT_STOCK_REVIEW_OUTCOME_MIN_EVALUABLE,
    stock_review_outcome_min_group_evaluable: int = DEFAULT_STOCK_REVIEW_OUTCOME_MIN_GROUP_EVALUABLE,
    market_snapshot: MarketSnapshotLoadResult | None = None,
) -> PaperAccountResult:
    root = Path(project_root).resolve()
    resolved_config = _resolve(root, config_path, DEFAULT_PAPER_CONFIG)
    resolved_allocator = _resolve(root, allocator_dir, DEFAULT_ALLOCATOR_DIR)
    resolved_output = _resolve(root, output_dir, Path("outputs/research/paper_account_latest"))
    resolved_stock_review_notes = _resolve(root, stock_review_notes_path, DEFAULT_STOCK_REVIEW_NOTES_PATH)
    resolved_stock_review_outcomes_history = _resolve(
        root,
        stock_review_outcomes_history_path,
        DEFAULT_STOCK_REVIEW_OUTCOMES_HISTORY_PATH,
    )

    base_config = load_portfolio_config(resolved_config)
    base_equity = _window_equity_frame(
        base_config,
        resolved_allocator,
        config_path=resolved_config,
        cache_dir=resolved_output / ".cache",
    )
    ledger, audit, metrics, monthly = build_paper_account_ledger(
        base_equity,
        initial_cash=base_config.initial_cash,
        rebalance_cost_rate=rebalance_cost_rate,
    )
    target_holdings, target_payload = build_target_holdings(ledger)
    stock_targets, stock_targets_payload = build_stock_targets(
        root,
        base_config,
        target_holdings,
        metrics.get("latest_date"),
        trigger_signal_path=trigger_signal_path,
        stock_market_cap_path=stock_market_cap_path,
        stock_tracking_max_market_cap_yi=stock_tracking_max_market_cap_yi,
        market_snapshot=market_snapshot,
    )
    stock_target_review, stock_target_review_payload = build_stock_target_review(
        stock_targets,
        stock_targets_payload,
        drawdown_threshold=stock_review_drawdown_threshold,
        watch_drawdown_threshold=stock_review_watch_drawdown_threshold,
        loss_attention_threshold=stock_review_loss_attention_threshold,
        gain_attention_threshold=stock_review_gain_attention_threshold,
        watch_score_threshold=stock_review_watch_score_threshold,
    )
    stock_targets, stock_targets_payload, stock_target_review, stock_target_review_payload = (
        apply_review_required_observation_exclusion(
            stock_targets,
            stock_targets_payload,
            stock_target_review,
            stock_target_review_payload,
        )
    )
    stock_target_review_notes_snapshot_path = resolved_output / "stock_target_review_notes.csv"
    stock_target_review, notes_payload = sync_stock_target_review_notes(
        stock_target_review,
        resolved_stock_review_notes,
        stock_target_review_notes_snapshot_path,
    )
    stock_target_review_payload.update(notes_payload)
    stock_target_review_actions, stock_target_review_actions_payload = build_stock_target_review_actions(
        stock_target_review,
        stock_target_review_payload,
    )
    price_history_cache: dict[str, tuple[pd.DataFrame, str]] = {}
    price_history_cache_dir = resolved_output / ".cache" / "stock_price_history"
    stock_target_review_assistant, stock_target_review_assistant_payload = build_stock_target_review_assistant(
        root,
        stock_target_review,
        stock_target_review_actions,
        stock_target_review_payload,
        price_history_cache=price_history_cache,
        price_history_cache_dir=price_history_cache_dir,
    )
    stock_target_review_decision_template, stock_target_review_decision_template_payload = build_stock_target_review_decision_template(
        stock_target_review_assistant,
        stock_target_review_assistant_payload,
    )
    stock_target_review_outcomes, stock_target_review_outcomes_payload = build_stock_target_review_outcomes(
        root,
        stock_target_review,
        stock_target_review_actions,
        stock_target_review_payload,
        price_history_cache=price_history_cache,
        price_history_cache_dir=price_history_cache_dir,
    )
    stock_target_review_outcomes_history_snapshot_path = resolved_output / "stock_target_review_outcomes_history.csv"
    stock_target_review_outcomes_history, stock_target_review_outcomes_history_payload = sync_stock_target_review_outcomes_history(
        stock_target_review_outcomes,
        resolved_stock_review_outcomes_history,
        stock_target_review_outcomes_history_snapshot_path,
    )
    stock_target_review_outcome_analysis, stock_target_review_outcome_analysis_payload = _cached_stock_target_review_outcome_analysis(
        stock_target_review_outcomes_history,
        stock_target_review_outcomes_history_payload,
        min_evaluable=stock_review_outcome_min_evaluable,
        min_group_evaluable=stock_review_outcome_min_group_evaluable,
        cache_dir=resolved_output / ".cache" / "outcome_analysis",
    )
    stock_target_review_outcome_calendar, stock_target_review_outcome_calendar_payload = build_stock_target_review_outcome_calendar(
        stock_target_review_outcome_analysis_payload,
        as_of_date=metrics.get("latest_date"),
    )
    stock_target_review_outcome_due, stock_target_review_outcome_due_payload = build_stock_target_review_outcome_due_queue(
        stock_target_review_outcome_calendar,
        stock_target_review_outcome_calendar_payload,
    )
    metrics.update(
        {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "config_path": str(resolved_config),
            "allocator_dir": str(resolved_allocator),
            "target_holdings_path": str(resolved_output / "target_holdings.csv"),
            "target_holdings_json_path": str(resolved_output / "target_holdings.json"),
            "target_holdings_report_path": str(resolved_output / "target_holdings.md"),
            "stock_targets_path": str(resolved_output / "stock_targets.csv"),
            "stock_targets_json_path": str(resolved_output / "stock_targets.json"),
            "stock_targets_report_path": str(resolved_output / "stock_targets.md"),
            "stock_target_review_path": str(resolved_output / "stock_target_review.csv"),
            "stock_target_review_json_path": str(resolved_output / "stock_target_review.json"),
            "stock_target_review_report_path": str(resolved_output / "stock_target_review.md"),
            "stock_target_review_notes_path": str(resolved_stock_review_notes),
            "stock_target_review_notes_snapshot_path": str(stock_target_review_notes_snapshot_path),
            "stock_target_review_actions_path": str(resolved_output / "stock_target_review_actions.csv"),
            "stock_target_review_actions_json_path": str(resolved_output / "stock_target_review_actions.json"),
            "stock_target_review_actions_report_path": str(resolved_output / "stock_target_review_actions.md"),
            "stock_target_review_assistant_path": str(resolved_output / "stock_target_review_assistant.csv"),
            "stock_target_review_assistant_json_path": str(resolved_output / "stock_target_review_assistant.json"),
            "stock_target_review_assistant_report_path": str(resolved_output / "stock_target_review_assistant.md"),
            "stock_target_review_decision_template_path": str(resolved_output / "stock_target_review_decision_template.csv"),
            "stock_target_review_decision_template_json_path": str(resolved_output / "stock_target_review_decision_template.json"),
            "stock_target_review_decision_template_report_path": str(resolved_output / "stock_target_review_decision_template.md"),
            "stock_target_review_decision_template_xlsx_path": str(resolved_output / "stock_target_review_decision_template.xlsx"),
            "stock_target_review_outcomes_path": str(resolved_output / "stock_target_review_outcomes.csv"),
            "stock_target_review_outcomes_json_path": str(resolved_output / "stock_target_review_outcomes.json"),
            "stock_target_review_outcomes_report_path": str(resolved_output / "stock_target_review_outcomes.md"),
            "stock_target_review_outcomes_history_path": str(resolved_stock_review_outcomes_history),
            "stock_target_review_outcomes_history_snapshot_path": str(stock_target_review_outcomes_history_snapshot_path),
            "stock_target_review_outcomes_history_json_path": str(resolved_output / "stock_target_review_outcomes_history.json"),
            "stock_target_review_outcomes_history_report_path": str(resolved_output / "stock_target_review_outcomes_history.md"),
            "stock_target_review_outcome_analysis_path": str(resolved_output / "stock_target_review_outcome_analysis.csv"),
            "stock_target_review_outcome_analysis_json_path": str(resolved_output / "stock_target_review_outcome_analysis.json"),
            "stock_target_review_outcome_analysis_report_path": str(resolved_output / "stock_target_review_outcome_analysis.md"),
            "stock_target_review_outcome_calendar_path": str(resolved_output / "stock_target_review_outcome_calendar.csv"),
            "stock_target_review_outcome_calendar_json_path": str(resolved_output / "stock_target_review_outcome_calendar.json"),
            "stock_target_review_outcome_calendar_report_path": str(resolved_output / "stock_target_review_outcome_calendar.md"),
            "stock_target_review_outcome_due_path": str(resolved_output / "stock_target_review_outcome_due.csv"),
            "stock_target_review_outcome_due_json_path": str(resolved_output / "stock_target_review_outcome_due.json"),
            "stock_target_review_outcome_due_report_path": str(resolved_output / "stock_target_review_outcome_due.md"),
            "stock_target_trigger_signal_status": stock_targets_payload.get("trigger_signal_status"),
            "stock_target_trigger_match_count": stock_targets_payload.get("trigger_match_count"),
            "stock_target_trigger_signal_path": stock_targets_payload.get("latest_trigger_signal_path"),
            "stock_tracking_max_market_cap_yi": stock_targets_payload.get("stock_tracking_max_market_cap_yi"),
            "stock_tracking_excluded_large_market_cap_count": stock_targets_payload.get("stock_tracking_excluded_large_market_cap_count"),
            "stock_tracking_allowed_count": stock_targets_payload.get("stock_tracking_allowed_count"),
            "stock_tracking_market_cap_missing_count": stock_targets_payload.get("stock_tracking_market_cap_missing_count"),
            "stock_market_cap_cache_status": stock_targets_payload.get("stock_market_cap_cache_status"),
            "stock_market_cap_path": stock_targets_payload.get("stock_market_cap_path"),
            "stock_market_cap_cache_row_count": stock_targets_payload.get("stock_market_cap_cache_row_count"),
            "stock_market_cap_cache_latest_snapshot_date": stock_targets_payload.get("stock_market_cap_cache_latest_snapshot_date"),
            "stock_market_cap_cache_updated_at": stock_targets_payload.get("stock_market_cap_cache_updated_at"),
            "stock_target_review_required_count": stock_target_review_payload.get("review_required_count"),
            "stock_target_review_excluded_count": stock_target_review_payload.get("excluded_review_required_count"),
            "stock_target_review_monitor_count": stock_target_review_payload.get("monitor_count"),
            "stock_target_review_drawdown_count": stock_target_review_payload.get("drawdown_review_count"),
            "stock_target_review_manual_note_count": stock_target_review_payload.get("manual_note_count"),
            "stock_target_review_unreviewed_count": stock_target_review_payload.get("unreviewed_count"),
            "stock_target_review_required_unreviewed_count": stock_target_review_payload.get("review_required_unreviewed_count"),
            "stock_target_review_manual_pending_count": stock_target_review_payload.get("manual_pending_count"),
            "stock_target_review_manual_reviewed_count": stock_target_review_payload.get("manual_reviewed_count"),
            "stock_target_review_manual_watch_count": stock_target_review_payload.get("manual_watch_count"),
            "stock_target_review_manual_resolved_count": stock_target_review_payload.get("manual_resolved_count"),
            "stock_target_review_manual_exclude_candidate_count": stock_target_review_payload.get("manual_exclude_candidate_count"),
            "stock_target_review_manual_other_status_count": stock_target_review_payload.get("manual_other_status_count"),
            "stock_target_review_action_count": stock_target_review_actions_payload.get("action_count"),
            "stock_target_review_action_pending_model_count": stock_target_review_actions_payload.get("review_required_pending_count"),
            "stock_target_review_action_manual_watch_count": stock_target_review_actions_payload.get("manual_watch_followup_count"),
            "stock_target_review_action_manual_exclusion_count": stock_target_review_actions_payload.get("manual_exclusion_candidate_count"),
            "stock_target_review_action_status_normalization_count": stock_target_review_actions_payload.get("manual_status_normalization_count"),
            "stock_target_review_action_next_review_due_count": stock_target_review_actions_payload.get("manual_next_review_due_count"),
            "stock_target_review_assistant_count": stock_target_review_assistant_payload.get("assistant_row_count"),
            "stock_target_review_assistant_pending_count": stock_target_review_assistant_payload.get("review_required_pending_count"),
            "stock_target_review_assistant_price_ok_count": stock_target_review_assistant_payload.get("price_history_ok_count"),
            "stock_target_review_assistant_price_stale_count": stock_target_review_assistant_payload.get("price_history_stale_count"),
            "stock_target_review_assistant_price_missing_count": stock_target_review_assistant_payload.get("price_history_missing_count"),
            "stock_target_review_decision_template_count": stock_target_review_decision_template_payload.get("decision_template_row_count"),
            "stock_target_review_decision_template_blank_status_count": stock_target_review_decision_template_payload.get("blank_manual_status_count"),
            "stock_target_review_outcome_row_count": stock_target_review_outcomes_payload.get("outcome_row_count"),
            "stock_target_review_outcome_complete_count": stock_target_review_outcomes_payload.get("complete_count"),
            "stock_target_review_outcome_partial_count": stock_target_review_outcomes_payload.get("partial_count"),
            "stock_target_review_outcome_pending_count": stock_target_review_outcomes_payload.get("pending_count"),
            "stock_target_review_outcome_missing_entry_price_count": stock_target_review_outcomes_payload.get("missing_entry_price_count"),
            "stock_target_review_outcomes_history_row_count": stock_target_review_outcomes_history_payload.get("history_row_count"),
            "stock_target_review_outcomes_history_updated_row_count": stock_target_review_outcomes_history_payload.get("history_updated_row_count"),
            "stock_target_review_outcomes_history_complete_count": stock_target_review_outcomes_history_payload.get("history_complete_count"),
            "stock_target_review_outcomes_history_partial_count": stock_target_review_outcomes_history_payload.get("history_partial_count"),
            "stock_target_review_outcomes_history_pending_count": stock_target_review_outcomes_history_payload.get("history_pending_count"),
            "stock_target_review_outcomes_history_missing_entry_price_count": stock_target_review_outcomes_history_payload.get("history_missing_entry_price_count"),
            "stock_target_review_outcomes_history_latest_review_date": stock_target_review_outcomes_history_payload.get("history_latest_review_date"),
            "stock_target_review_outcome_analysis_status": stock_target_review_outcome_analysis_payload.get("analysis_status"),
            "stock_target_review_outcome_analysis_row_count": stock_target_review_outcome_analysis_payload.get("analysis_row_count"),
            "stock_target_review_outcome_analysis_min_evaluable": stock_target_review_outcome_analysis_payload.get("min_evaluable"),
            "stock_target_review_outcome_analysis_min_group_evaluable": stock_target_review_outcome_analysis_payload.get("min_group_evaluable"),
            "stock_target_review_outcome_analysis_ready_horizon_count": stock_target_review_outcome_analysis_payload.get("ready_horizon_count"),
            "stock_target_review_outcome_analysis_sample_warning": stock_target_review_outcome_analysis_payload.get("sample_warning"),
            "stock_target_review_outcome_calendar_row_count": stock_target_review_outcome_calendar_payload.get("calendar_row_count"),
            "stock_target_review_outcome_calendar_ready_count": stock_target_review_outcome_calendar_payload.get("calendar_ready_count"),
            "stock_target_review_outcome_calendar_pending_count": stock_target_review_outcome_calendar_payload.get("calendar_pending_count"),
            "stock_target_review_outcome_calendar_due_count": stock_target_review_outcome_calendar_payload.get("calendar_due_count"),
            "stock_target_review_outcome_calendar_due_pending_count": stock_target_review_outcome_calendar_payload.get("calendar_due_pending_count"),
            "stock_target_review_outcome_calendar_next_action_date": stock_target_review_outcome_calendar_payload.get("next_action_date"),
            "stock_target_review_outcome_calendar_next_action_horizon": stock_target_review_outcome_calendar_payload.get("next_action_horizon"),
            "stock_target_review_outcome_calendar_next_due_date": stock_target_review_outcome_calendar_payload.get("next_due_date"),
            "stock_target_review_outcome_calendar_next_due_horizon": stock_target_review_outcome_calendar_payload.get("next_due_horizon"),
            "stock_target_review_outcome_due_status": stock_target_review_outcome_due_payload.get("due_status"),
            "stock_target_review_outcome_due_row_count": stock_target_review_outcome_due_payload.get("due_row_count"),
            "stock_target_review_outcome_due_pending_count": stock_target_review_outcome_due_payload.get("due_pending_count"),
            "stock_target_review_outcome_due_next_date": stock_target_review_outcome_due_payload.get("next_due_date"),
            "stock_target_review_outcome_due_next_horizon": stock_target_review_outcome_due_payload.get("next_due_horizon"),
            "stock_target_review_outcome_maturity_next_1d_date": (
                stock_target_review_outcome_analysis_payload.get("maturity_forecast") or {}
            ).get("1d", {}).get("estimated_next_evaluable_date"),
            "stock_target_review_outcome_maturity_next_5d_date": (
                stock_target_review_outcome_analysis_payload.get("maturity_forecast") or {}
            ).get("5d", {}).get("estimated_next_evaluable_date"),
            "stock_target_review_outcome_maturity_next_10d_date": (
                stock_target_review_outcome_analysis_payload.get("maturity_forecast") or {}
            ).get("10d", {}).get("estimated_next_evaluable_date"),
            "stock_target_review_outcome_maturity_next_20d_date": (
                stock_target_review_outcome_analysis_payload.get("maturity_forecast") or {}
            ).get("20d", {}).get("estimated_next_evaluable_date"),
            "stock_target_review_outcome_analysis_evaluable_1d_count": (
                stock_target_review_outcome_analysis_payload.get("total_evaluable_by_horizon") or {}
            ).get("1d"),
            "stock_target_review_outcome_analysis_evaluable_5d_count": (
                stock_target_review_outcome_analysis_payload.get("total_evaluable_by_horizon") or {}
            ).get("5d"),
            "stock_target_review_outcome_analysis_evaluable_10d_count": (
                stock_target_review_outcome_analysis_payload.get("total_evaluable_by_horizon") or {}
            ).get("10d"),
            "stock_target_review_outcome_analysis_evaluable_20d_count": (
                stock_target_review_outcome_analysis_payload.get("total_evaluable_by_horizon") or {}
            ).get("20d"),
            "stock_target_review_drawdown_threshold": stock_target_review_payload.get("drawdown_threshold"),
            "stock_target_review_watch_drawdown_threshold": stock_target_review_payload.get("watch_drawdown_threshold"),
            "stock_target_review_loss_attention_threshold": stock_target_review_payload.get("loss_attention_threshold"),
        }
    )
    target_payload.update(
        {
            "config_path": str(resolved_config),
            "allocator_dir": str(resolved_allocator),
        }
    )
    stock_targets_payload.update(
        {
            "config_path": str(resolved_config),
            "allocator_dir": str(resolved_allocator),
        }
    )
    stock_target_review_payload.update(
        {
            "config_path": str(resolved_config),
            "allocator_dir": str(resolved_allocator),
        }
    )
    stock_target_review_actions_payload.update(
        {
            "config_path": str(resolved_config),
            "allocator_dir": str(resolved_allocator),
        }
    )
    stock_target_review_assistant_payload.update(
        {
            "config_path": str(resolved_config),
            "allocator_dir": str(resolved_allocator),
            "assistant_path": str(resolved_output / "stock_target_review_assistant.csv"),
            "assistant_json_path": str(resolved_output / "stock_target_review_assistant.json"),
            "assistant_report_path": str(resolved_output / "stock_target_review_assistant.md"),
        }
    )
    stock_target_review_decision_template_payload.update(
        {
            "config_path": str(resolved_config),
            "allocator_dir": str(resolved_allocator),
            "decision_template_path": str(resolved_output / "stock_target_review_decision_template.csv"),
            "decision_template_json_path": str(resolved_output / "stock_target_review_decision_template.json"),
            "decision_template_report_path": str(resolved_output / "stock_target_review_decision_template.md"),
            "decision_template_xlsx_path": str(resolved_output / "stock_target_review_decision_template.xlsx"),
        }
    )
    stock_target_review_outcomes_payload.update(
        {
            "config_path": str(resolved_config),
            "allocator_dir": str(resolved_allocator),
        }
    )
    stock_target_review_outcomes_history_payload.update(
        {
            "config_path": str(resolved_config),
            "allocator_dir": str(resolved_allocator),
            "history_json_path": str(resolved_output / "stock_target_review_outcomes_history.json"),
            "history_report_path": str(resolved_output / "stock_target_review_outcomes_history.md"),
        }
    )
    stock_target_review_outcome_analysis_payload.update(
        {
            "config_path": str(resolved_config),
            "allocator_dir": str(resolved_allocator),
            "analysis_path": str(resolved_output / "stock_target_review_outcome_analysis.csv"),
            "analysis_json_path": str(resolved_output / "stock_target_review_outcome_analysis.json"),
            "analysis_report_path": str(resolved_output / "stock_target_review_outcome_analysis.md"),
        }
    )
    stock_target_review_outcome_calendar_payload.update(
        {
            "config_path": str(resolved_config),
            "allocator_dir": str(resolved_allocator),
            "calendar_path": str(resolved_output / "stock_target_review_outcome_calendar.csv"),
            "calendar_json_path": str(resolved_output / "stock_target_review_outcome_calendar.json"),
            "calendar_report_path": str(resolved_output / "stock_target_review_outcome_calendar.md"),
            "due_path": str(resolved_output / "stock_target_review_outcome_due.csv"),
            "due_json_path": str(resolved_output / "stock_target_review_outcome_due.json"),
            "due_report_path": str(resolved_output / "stock_target_review_outcome_due.md"),
        }
    )
    stock_target_review_outcome_due_payload.update(
        {
            "config_path": str(resolved_config),
            "allocator_dir": str(resolved_allocator),
            "due_path": str(resolved_output / "stock_target_review_outcome_due.csv"),
            "due_json_path": str(resolved_output / "stock_target_review_outcome_due.json"),
            "due_report_path": str(resolved_output / "stock_target_review_outcome_due.md"),
            "calendar_path": str(resolved_output / "stock_target_review_outcome_calendar.csv"),
            "calendar_report_path": str(resolved_output / "stock_target_review_outcome_calendar.md"),
        }
    )

    resolved_output.mkdir(parents=True, exist_ok=True)
    ledger_path = resolved_output / "ledger.csv"
    audit_path = resolved_output / "rebalance_audit.csv"
    monthly_path = resolved_output / "monthly_returns.csv"
    target_holdings_path = resolved_output / "target_holdings.csv"
    target_holdings_json_path = resolved_output / "target_holdings.json"
    target_holdings_report_path = resolved_output / "target_holdings.md"
    stock_targets_path = resolved_output / "stock_targets.csv"
    stock_targets_json_path = resolved_output / "stock_targets.json"
    stock_targets_report_path = resolved_output / "stock_targets.md"
    stock_target_review_path = resolved_output / "stock_target_review.csv"
    stock_target_review_json_path = resolved_output / "stock_target_review.json"
    stock_target_review_report_path = resolved_output / "stock_target_review.md"
    stock_target_review_actions_path = resolved_output / "stock_target_review_actions.csv"
    stock_target_review_actions_json_path = resolved_output / "stock_target_review_actions.json"
    stock_target_review_actions_report_path = resolved_output / "stock_target_review_actions.md"
    stock_target_review_assistant_path = resolved_output / "stock_target_review_assistant.csv"
    stock_target_review_assistant_json_path = resolved_output / "stock_target_review_assistant.json"
    stock_target_review_assistant_report_path = resolved_output / "stock_target_review_assistant.md"
    stock_target_review_decision_template_path = resolved_output / "stock_target_review_decision_template.csv"
    stock_target_review_decision_template_json_path = resolved_output / "stock_target_review_decision_template.json"
    stock_target_review_decision_template_report_path = resolved_output / "stock_target_review_decision_template.md"
    stock_target_review_decision_template_xlsx_path = resolved_output / "stock_target_review_decision_template.xlsx"
    stock_target_review_outcomes_path = resolved_output / "stock_target_review_outcomes.csv"
    stock_target_review_outcomes_json_path = resolved_output / "stock_target_review_outcomes.json"
    stock_target_review_outcomes_report_path = resolved_output / "stock_target_review_outcomes.md"
    stock_target_review_outcomes_history_snapshot_path = resolved_output / "stock_target_review_outcomes_history.csv"
    stock_target_review_outcomes_history_json_path = resolved_output / "stock_target_review_outcomes_history.json"
    stock_target_review_outcomes_history_report_path = resolved_output / "stock_target_review_outcomes_history.md"
    stock_target_review_outcome_analysis_path = resolved_output / "stock_target_review_outcome_analysis.csv"
    stock_target_review_outcome_analysis_json_path = resolved_output / "stock_target_review_outcome_analysis.json"
    stock_target_review_outcome_analysis_report_path = resolved_output / "stock_target_review_outcome_analysis.md"
    stock_target_review_outcome_calendar_path = resolved_output / "stock_target_review_outcome_calendar.csv"
    stock_target_review_outcome_calendar_json_path = resolved_output / "stock_target_review_outcome_calendar.json"
    stock_target_review_outcome_calendar_report_path = resolved_output / "stock_target_review_outcome_calendar.md"
    stock_target_review_outcome_due_path = resolved_output / "stock_target_review_outcome_due.csv"
    stock_target_review_outcome_due_json_path = resolved_output / "stock_target_review_outcome_due.json"
    stock_target_review_outcome_due_report_path = resolved_output / "stock_target_review_outcome_due.md"
    metrics_path = resolved_output / "metrics.json"
    report_path = resolved_output / "paper_account.md"
    ledger.to_csv(ledger_path, index=False, encoding="utf-8")
    audit.to_csv(audit_path, index=False, encoding="utf-8")
    monthly.to_csv(monthly_path, index=False, encoding="utf-8")
    target_holdings.to_csv(target_holdings_path, index=False, encoding="utf-8")
    target_holdings_json_path.write_text(json.dumps(target_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    target_holdings_report_path.write_text(
        _render_target_holdings_report(target_holdings, target_payload, resolved_output),
        encoding="utf-8",
    )
    stock_targets.to_csv(stock_targets_path, index=False, encoding="utf-8")
    stock_targets_json_path.write_text(json.dumps(stock_targets_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    stock_targets_report_path.write_text(
        _render_stock_targets_report(stock_targets, stock_targets_payload, resolved_output),
        encoding="utf-8",
    )
    stock_target_review.to_csv(stock_target_review_path, index=False, encoding="utf-8")
    stock_target_review_json_path.write_text(
        json.dumps(stock_target_review_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    stock_target_review_report_path.write_text(
        _render_stock_target_review_report(stock_target_review, stock_target_review_payload, resolved_output),
        encoding="utf-8",
    )
    stock_target_review_actions.to_csv(stock_target_review_actions_path, index=False, encoding="utf-8")
    stock_target_review_actions_json_path.write_text(
        json.dumps(stock_target_review_actions_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    stock_target_review_actions_report_path.write_text(
        _render_stock_target_review_actions_report(stock_target_review_actions, stock_target_review_actions_payload, resolved_output),
        encoding="utf-8",
    )
    stock_target_review_assistant.to_csv(stock_target_review_assistant_path, index=False, encoding="utf-8")
    stock_target_review_assistant_json_path.write_text(
        json.dumps(stock_target_review_assistant_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    stock_target_review_assistant_report_path.write_text(
        _render_stock_target_review_assistant_report(
            stock_target_review_assistant,
            stock_target_review_assistant_payload,
            resolved_output,
        ),
        encoding="utf-8",
    )
    _publish_stock_target_review_decision_template(
        stock_target_review_decision_template,
        stock_target_review_decision_template_payload,
        output_dir=resolved_output,
        csv_path=stock_target_review_decision_template_path,
        json_path=stock_target_review_decision_template_json_path,
        report_path=stock_target_review_decision_template_report_path,
        xlsx_path=stock_target_review_decision_template_xlsx_path,
    )
    stock_target_review_outcomes.to_csv(stock_target_review_outcomes_path, index=False, encoding="utf-8")
    stock_target_review_outcomes_json_path.write_text(
        json.dumps(stock_target_review_outcomes_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    stock_target_review_outcomes_report_path.write_text(
        _render_stock_target_review_outcomes_report(stock_target_review_outcomes, stock_target_review_outcomes_payload, resolved_output),
        encoding="utf-8",
    )
    stock_target_review_outcomes_history_json_path.write_text(
        json.dumps(stock_target_review_outcomes_history_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    stock_target_review_outcomes_history_report_path.write_text(
        _render_stock_target_review_outcomes_history_report(
            stock_target_review_outcomes_history,
            stock_target_review_outcomes_history_payload,
            resolved_output,
        ),
        encoding="utf-8",
    )
    stock_target_review_outcome_analysis.to_csv(stock_target_review_outcome_analysis_path, index=False, encoding="utf-8")
    stock_target_review_outcome_analysis_json_path.write_text(
        json.dumps(stock_target_review_outcome_analysis_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    stock_target_review_outcome_analysis_report_path.write_text(
        _render_stock_target_review_outcome_analysis_report(
            stock_target_review_outcome_analysis,
            stock_target_review_outcome_analysis_payload,
            resolved_output,
        ),
        encoding="utf-8",
    )
    stock_target_review_outcome_calendar.to_csv(stock_target_review_outcome_calendar_path, index=False, encoding="utf-8")
    stock_target_review_outcome_calendar_json_path.write_text(
        json.dumps(stock_target_review_outcome_calendar_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    stock_target_review_outcome_calendar_report_path.write_text(
        _render_stock_target_review_outcome_calendar_report(
            stock_target_review_outcome_calendar,
            stock_target_review_outcome_calendar_payload,
            resolved_output,
        ),
        encoding="utf-8",
    )
    stock_target_review_outcome_due.to_csv(stock_target_review_outcome_due_path, index=False, encoding="utf-8")
    stock_target_review_outcome_due_json_path.write_text(
        json.dumps(stock_target_review_outcome_due_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    stock_target_review_outcome_due_report_path.write_text(
        _render_stock_target_review_outcome_due_report(
            stock_target_review_outcome_due,
            stock_target_review_outcome_due_payload,
            resolved_output,
        ),
        encoding="utf-8",
    )
    metrics_path.write_text(json.dumps(metrics, ensure_ascii=False, indent=2), encoding="utf-8")
    report_path.write_text(
        _render_report(metrics, resolved_output, resolved_allocator, resolved_config),
        encoding="utf-8",
    )

    return PaperAccountResult(
        output_dir=resolved_output,
        ledger_path=ledger_path,
        audit_path=audit_path,
        monthly_returns_path=monthly_path,
        target_holdings_path=target_holdings_path,
        target_holdings_json_path=target_holdings_json_path,
        target_holdings_report_path=target_holdings_report_path,
        stock_targets_path=stock_targets_path,
        stock_targets_json_path=stock_targets_json_path,
        stock_targets_report_path=stock_targets_report_path,
        stock_target_review_path=stock_target_review_path,
        stock_target_review_json_path=stock_target_review_json_path,
        stock_target_review_report_path=stock_target_review_report_path,
        stock_target_review_notes_path=resolved_stock_review_notes,
        stock_target_review_notes_snapshot_path=stock_target_review_notes_snapshot_path,
        stock_target_review_actions_path=stock_target_review_actions_path,
        stock_target_review_actions_json_path=stock_target_review_actions_json_path,
        stock_target_review_actions_report_path=stock_target_review_actions_report_path,
        stock_target_review_assistant_path=stock_target_review_assistant_path,
        stock_target_review_assistant_json_path=stock_target_review_assistant_json_path,
        stock_target_review_assistant_report_path=stock_target_review_assistant_report_path,
        stock_target_review_decision_template_path=stock_target_review_decision_template_path,
        stock_target_review_decision_template_json_path=stock_target_review_decision_template_json_path,
        stock_target_review_decision_template_report_path=stock_target_review_decision_template_report_path,
        stock_target_review_decision_template_xlsx_path=stock_target_review_decision_template_xlsx_path,
        stock_target_review_outcomes_path=stock_target_review_outcomes_path,
        stock_target_review_outcomes_json_path=stock_target_review_outcomes_json_path,
        stock_target_review_outcomes_report_path=stock_target_review_outcomes_report_path,
        stock_target_review_outcomes_history_path=resolved_stock_review_outcomes_history,
        stock_target_review_outcomes_history_snapshot_path=stock_target_review_outcomes_history_snapshot_path,
        stock_target_review_outcomes_history_json_path=stock_target_review_outcomes_history_json_path,
        stock_target_review_outcomes_history_report_path=stock_target_review_outcomes_history_report_path,
        stock_target_review_outcome_analysis_path=stock_target_review_outcome_analysis_path,
        stock_target_review_outcome_analysis_json_path=stock_target_review_outcome_analysis_json_path,
        stock_target_review_outcome_analysis_report_path=stock_target_review_outcome_analysis_report_path,
        stock_target_review_outcome_calendar_path=stock_target_review_outcome_calendar_path,
        stock_target_review_outcome_calendar_json_path=stock_target_review_outcome_calendar_json_path,
        stock_target_review_outcome_calendar_report_path=stock_target_review_outcome_calendar_report_path,
        stock_target_review_outcome_due_path=stock_target_review_outcome_due_path,
        stock_target_review_outcome_due_json_path=stock_target_review_outcome_due_json_path,
        stock_target_review_outcome_due_report_path=stock_target_review_outcome_due_report_path,
        metrics_path=metrics_path,
        report_path=report_path,
        ledger=ledger,
        audit=audit,
        target_holdings=target_holdings,
        target_holdings_payload=target_payload,
        stock_targets=stock_targets,
        stock_targets_payload=stock_targets_payload,
        stock_target_review=stock_target_review,
        stock_target_review_payload=stock_target_review_payload,
        stock_target_review_actions=stock_target_review_actions,
        stock_target_review_actions_payload=stock_target_review_actions_payload,
        stock_target_review_assistant=stock_target_review_assistant,
        stock_target_review_assistant_payload=stock_target_review_assistant_payload,
        stock_target_review_decision_template=stock_target_review_decision_template,
        stock_target_review_decision_template_payload=stock_target_review_decision_template_payload,
        stock_target_review_outcomes=stock_target_review_outcomes,
        stock_target_review_outcomes_payload=stock_target_review_outcomes_payload,
        stock_target_review_outcomes_history=stock_target_review_outcomes_history,
        stock_target_review_outcomes_history_payload=stock_target_review_outcomes_history_payload,
        stock_target_review_outcome_analysis=stock_target_review_outcome_analysis,
        stock_target_review_outcome_analysis_payload=stock_target_review_outcome_analysis_payload,
        stock_target_review_outcome_calendar=stock_target_review_outcome_calendar,
        stock_target_review_outcome_calendar_payload=stock_target_review_outcome_calendar_payload,
        stock_target_review_outcome_due=stock_target_review_outcome_due,
        stock_target_review_outcome_due_payload=stock_target_review_outcome_due_payload,
        monthly_returns=monthly,
        metrics=metrics,
    )
