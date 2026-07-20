from __future__ import annotations

import io
import tempfile
import unittest
import json
import os
import sys
from contextlib import redirect_stdout
from dataclasses import replace
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import numpy as np
import pandas as pd

from quant_etf_lab.attribution import find_drawdown_periods, load_equity_drawdown, run_drawdown_attribution
from quant_etf_lab.backtest import BacktestResult, _benchmark_curve, _scaled_buy_orders, _score_weighted_position_weights, run_backtest
from quant_etf_lab.cli import _alert_gate_failed, _history_gate_failed, _walk_forward_lock_key, build_parser, main, resolve_walk_forward_options
from quant_etf_lab.config import CostsConfig, DataConfig, ETFSpec, LabConfig, ProjectConfig, RiskConfig, StrategyConfig, UniverseSourceConfig, load_config, load_config_mapping, parse_config
from quant_etf_lab.data import (
    AkShareDataError,
    UniverseSourceDataError,
    cache_paths,
    fetch_akshare_history,
    filter_history_by_date,
    is_a_share_chinext_code,
    is_a_share_main_board_code,
    is_a_share_main_or_chinext_code,
    load_universe_history,
    load_universe_file,
    normalize_a_share_main_board_frame,
    normalize_akshare_frame,
    normalize_constituents_frame,
    resolve_universe,
    update_data,
    universe_cache_path,
)
from quant_etf_lab.allocator_promotion import run_allocator_promotion_review
from quant_etf_lab.dashboard import run_daily_dashboard
from quant_etf_lab.daily_alerts import build_daily_alert_payload, write_daily_alerts
from quant_etf_lab.daily_check import run_daily_model_check
from quant_etf_lab.daily_pipeline import _live_preflight_context, _next_step_context, run_daily_pipeline
from quant_etf_lab.diagnostics import build_diagnostics, run_batch_diagnostics, run_diagnostics
from quant_etf_lab.factor_lab import build_symbol_factor_frame, run_factor_lab
from quant_etf_lab.indicators import moving_average, rsi
from quant_etf_lab.live_shadow import (
    run_live_shadow,
    run_live_shadow_import_template,
    run_live_shadow_review_decisions,
    run_live_shadow_review_queue,
)
from quant_etf_lab.live_preflight import run_live_preflight
from quant_etf_lab.market_sentiment import _load_sampled_config, build_market_sentiment_frame, classify_sentiment_state, reference_exposure
from quant_etf_lab.market_cap import normalize_stock_market_cap_frame
from quant_etf_lab.network_lab import estimate_mutual_information, run_network_lab
import quant_etf_lab.paper_account as paper_account_module
from quant_etf_lab.paper_account import (
    apply_stock_market_cap_tracking_rule,
    build_stock_target_review,
    build_stock_target_review_actions,
    build_stock_target_review_assistant,
    build_stock_target_review_decision_template,
    build_stock_target_review_outcome_analysis,
    build_stock_target_review_outcome_calendar,
    build_stock_target_review_outcome_due_queue,
    build_stock_target_review_outcomes,
    _render_stock_target_review_assistant_report,
    _render_stock_target_review_decision_template_report,
    run_paper_account,
    sync_stock_target_review_outcomes_history,
    sync_stock_target_review_notes,
)
from quant_etf_lab.phase2 import run_phase2_status
from quant_etf_lab.phase2_review import run_phase2_review
from quant_etf_lab.pipeline_history import run_pipeline_history_review
import quant_etf_lab.portfolio as portfolio_module
from quant_etf_lab.portfolio import (
    CurveConfig,
    PortfolioCandidate,
    PortfolioResult,
    PortfolioSourceCandidate,
    SatelliteFilterConfig,
    _apply_source_group_preselection,
    _select_source_candidate,
    build_portfolio_source_candidate_grid,
    build_portfolio_candidate_grid,
    generate_portfolio_windows,
    load_portfolio_config,
    load_portfolio_source_selection_config,
    run_portfolio_combine,
    run_portfolio_source_selection_walk_forward,
    score_portfolio_metrics,
)
from quant_etf_lab.process_lock import ProcessLockError, _tasklist_contains_pid, process_lock, seed_process_lock
from quant_etf_lab.risk import build_benchmark_risk_frame, decide_next_risk
from quant_etf_lab.report import write_report
from quant_etf_lab.strategies import build_signals
from quant_etf_lab.strategies import build_signal_frames
from quant_etf_lab.trading_gate import build_a_share_trading_gate
from quant_etf_lab.walk_forward import (
    ParameterCandidate,
    _candidate_selection_gate_reason,
    _walk_forward_validation_dates,
    _signal_warmup_start,
    _trim_signal_frames,
    build_parameter_grid,
    generate_rolling_windows,
    run_walk_forward,
    score_training_metrics,
)


def make_config(root: Path) -> LabConfig:
    return LabConfig(
        project_root=root,
        project=ProjectConfig(
            name="test",
            initial_cash=100000.0,
            data_dir=root / "data",
            output_dir=root / "outputs" / "backtests",
        ),
        data=DataConfig(start_date="20200101", end_date=None, period="daily", adjust="qfq"),
        universe=(ETFSpec(code="TEST", name="Test ETF"),),
        strategy=StrategyConfig(
            name="trend",
            fast_window=2,
            slow_window=3,
            trend_window=4,
            rsi_window=3,
            rsi_oversold=35,
            rsi_recover=40,
            rsi_exit=65,
            max_positions=1,
            allocation_buffer=0.995,
        ),
        costs=CostsConfig(commission_rate=0.001, slippage_rate=0.01),
    )


def make_price_frame(code: str, name: str, closes: list[float]) -> pd.DataFrame:
    dates = pd.date_range("2020-01-01", periods=len(closes), freq="D")
    return pd.DataFrame(
        {
            "date": dates,
            "code": code,
            "name": name,
            "open": closes,
            "high": [value * 1.01 for value in closes],
            "low": [value * 0.99 for value in closes],
            "close": closes,
            "volume": [1000] * len(closes),
            "amount": [100000] * len(closes),
        }
    )


def make_ohlcv_frame(
    code: str,
    name: str,
    opens: list[float],
    highs: list[float],
    lows: list[float],
    closes: list[float],
    volumes: list[float],
) -> pd.DataFrame:
    dates = pd.date_range("2020-01-01", periods=len(closes), freq="D")
    return pd.DataFrame(
        {
            "date": dates,
            "code": code,
            "name": name,
            "open": opens,
            "high": highs,
            "low": lows,
            "close": closes,
            "volume": volumes,
            "amount": [close * volume for close, volume in zip(closes, volumes)],
        }
    )


class CoreTests(unittest.TestCase):
    def test_normalize_akshare_frame_sorts_and_renames(self) -> None:
        raw = pd.DataFrame(
            {
                "日期": ["2020-01-03", "2020-01-02"],
                "开盘": [2.0, 1.0],
                "最高": [2.2, 1.2],
                "最低": [1.8, 0.9],
                "收盘": [2.1, 1.1],
                "成交量": [200, 100],
                "成交额": [420, 110],
            }
        )
        frame = normalize_akshare_frame(raw, ETFSpec("TEST", "Test ETF"))
        self.assertEqual(list(frame["date"].dt.strftime("%Y-%m-%d")), ["2020-01-02", "2020-01-03"])
        self.assertEqual(list(frame.columns[:9]), ["date", "code", "name", "open", "high", "low", "close", "volume", "amount"])
        self.assertEqual(frame.loc[0, "code"], "TEST")

    def test_normalize_akshare_frame_drops_nonpositive_prices(self) -> None:
        raw = pd.DataFrame(
            {
                "日期": ["2020-01-01", "2020-01-02"],
                "开盘": [-1.0, 2.0],
                "最高": [1.0, 2.2],
                "最低": [0.5, 1.8],
                "收盘": [0.8, 2.1],
                "成交量": [100, 200],
                "成交额": [100, 420],
            }
        )
        frame = normalize_akshare_frame(raw, ETFSpec("TEST", "Test ETF"))
        self.assertEqual(len(frame), 1)
        self.assertEqual(frame.loc[0, "open"], 2.0)

    def test_stock_history_fallback_records_daily_source(self) -> None:
        calls: list[str] = []

        class FakeAkShare:
            @staticmethod
            def stock_zh_a_hist(**_: object) -> pd.DataFrame:
                calls.append("hist")
                raise RuntimeError("primary outage")

            @staticmethod
            def stock_zh_a_daily(**_: object) -> pd.DataFrame:
                calls.append("daily")
                return pd.DataFrame(
                    {
                        "date": ["2020-01-02"],
                        "open": [10.0],
                        "high": [10.5],
                        "low": [9.5],
                        "close": [10.2],
                        "volume": [1000],
                        "amount": [10200],
                    }
                )

        with patch.dict(sys.modules, {"akshare": FakeAkShare}):
            frame = fetch_akshare_history(
                ETFSpec("000001", "Ping An", asset_type="stock"),
                start_date="20200101",
                end_date="20200131",
                period="daily",
                adjust="qfq",
            )

        self.assertEqual(calls, ["hist", "daily"])
        self.assertEqual(frame.attrs["source"], "akshare.stock_zh_a_daily")
        self.assertEqual(list(frame["code"]), ["000001"])

    def test_stock_history_fallback_error_keeps_primary_context(self) -> None:
        class FakeAkShare:
            @staticmethod
            def stock_zh_a_hist(**_: object) -> pd.DataFrame:
                raise RuntimeError("primary outage")

            @staticmethod
            def stock_zh_a_daily(**_: object) -> pd.DataFrame:
                raise ValueError("daily outage")

        with patch.dict(sys.modules, {"akshare": FakeAkShare}):
            with self.assertRaises(AkShareDataError) as caught:
                fetch_akshare_history(
                    ETFSpec("000001", "Ping An", asset_type="stock"),
                    start_date="20200101",
                    end_date="20200131",
                    period="daily",
                    adjust="qfq",
                )

        message = str(caught.exception)
        self.assertIn("akshare.stock_zh_a_hist", message)
        self.assertIn("primary outage", message)
        self.assertIn("akshare.stock_zh_a_daily", message)
        self.assertIn("daily outage", message)

    def test_update_data_continue_on_error_records_structured_failure(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            with patch(
                "quant_etf_lab.data.fetch_akshare_history_with_retries",
                side_effect=AkShareDataError("network outage"),
            ):
                written = update_data(config, continue_on_error=True, retry_count=1, pause_seconds=0)

            failures = pd.read_csv(Path(tmp) / "data" / "meta" / "update_failures.csv")

        self.assertEqual(written, [])
        self.assertEqual(failures.loc[0, "stage"], "fetch")
        self.assertEqual(failures.loc[0, "error_type"], "AkShareDataError")
        self.assertIn("network outage", failures.loc[0, "error"])
        self.assertIn("recorded_at", failures.columns)

    def test_load_universe_history_skip_missing_does_not_swallow_code_errors(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            with patch("quant_etf_lab.data.load_cached_history", side_effect=KeyError("programming bug")):
                with self.assertRaises(KeyError):
                    load_universe_history(config, allow_fetch=False, skip_missing=True)

    def test_resolve_universe_uses_cached_source_after_fetch_failure(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = make_config(Path(tmp))
            source = UniverseSourceConfig(type="csindex", symbol="000300", asset_type="stock", limit=None)
            config = base.__class__(
                project_root=base.project_root,
                project=base.project,
                data=base.data,
                universe=(),
                strategy=base.strategy,
                costs=base.costs,
                universe_file=None,
                universe_source=source,
                risk=base.risk,
            )
            cached = pd.DataFrame({"code": ["000001"], "name": ["Ping An"], "asset_type": ["stock"]})
            cache_path = universe_cache_path(config, source)
            cache_path.parent.mkdir(parents=True, exist_ok=True)
            cached.to_csv(cache_path, index=False)

            with patch(
                "quant_etf_lab.data.fetch_universe_source",
                side_effect=UniverseSourceDataError("source outage"),
            ):
                instruments = resolve_universe(config)

            failures = pd.read_csv(Path(tmp) / "data" / "meta" / "universe_source_failures.csv")

        self.assertEqual([instrument.code for instrument in instruments], ["000001"])
        self.assertEqual(failures.loc[0, "stage"], "source_fetch")
        self.assertEqual(failures.loc[0, "error_type"], "UniverseSourceDataError")
        self.assertIn("source outage", failures.loc[0, "error"])

    def test_resolve_universe_uses_main_chinext_fallback_after_source_failure(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = make_config(Path(tmp))
            source = UniverseSourceConfig(type="ashare_main_chinext", symbol="all", asset_type="stock", limit=2)
            config = base.__class__(
                project_root=base.project_root,
                project=base.project,
                data=base.data,
                universe=(),
                strategy=base.strategy,
                costs=base.costs,
                universe_file=None,
                universe_source=source,
                risk=base.risk,
            )
            fallback = pd.DataFrame(
                {
                    "code": ["000001", "300001"],
                    "name": ["Ping An", "Chinext A"],
                    "asset_type": ["stock", "stock"],
                }
            )
            with patch(
                "quant_etf_lab.data.fetch_universe_source",
                side_effect=UniverseSourceDataError("source outage"),
            ), patch(
                "quant_etf_lab.data.build_main_chinext_universe_from_fallbacks",
                return_value=fallback,
            ):
                instruments = resolve_universe(config)

            failures = pd.read_csv(Path(tmp) / "data" / "meta" / "universe_source_failures.csv")
            cached = pd.read_csv(universe_cache_path(config, source), dtype={"code": str})

        self.assertEqual([instrument.code for instrument in instruments], ["000001", "300001"])
        self.assertEqual(list(failures["stage"]), ["source_fetch"])
        self.assertEqual(list(cached["code"].str.zfill(6)), ["000001", "300001"])

    def test_resolve_universe_does_not_swallow_source_code_errors(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = make_config(Path(tmp))
            source = UniverseSourceConfig(type="csindex", symbol="000300", asset_type="stock", limit=None)
            config = base.__class__(
                project_root=base.project_root,
                project=base.project,
                data=base.data,
                universe=(),
                strategy=base.strategy,
                costs=base.costs,
                universe_file=None,
                universe_source=source,
                risk=base.risk,
            )
            with patch("quant_etf_lab.data.fetch_universe_source", side_effect=KeyError("programming bug")):
                with self.assertRaises(KeyError):
                    resolve_universe(config)

    def test_indicators(self) -> None:
        close = pd.Series([1, 2, 3, 4, 5], dtype=float)
        ma = moving_average(close, 3)
        self.assertTrue(pd.isna(ma.iloc[1]))
        self.assertAlmostEqual(ma.iloc[-1], 4.0)
        values = rsi(close, 3)
        self.assertTrue(values.dropna().between(0, 100).all())

    def test_build_signals_adds_shifted_volatility_stop_loss_pct(self) -> None:
        closes = [100.0, 102.0, 101.0, 103.0, 102.0, 104.0]
        frame = make_price_frame("TEST", "Test ETF", closes)
        with tempfile.TemporaryDirectory() as tmp:
            base = make_config(Path(tmp))
            strategy = base.strategy.__class__(
                **{
                    **base.strategy.__dict__,
                    "volatility_stop_loss_enabled": True,
                    "volatility_stop_window": 3,
                    "volatility_stop_multiplier": 2.0,
                    "volatility_stop_min_pct": 0.03,
                    "volatility_stop_max_pct": 0.10,
                }
            )
            signals = build_signals(frame, strategy)

        returns = pd.Series(closes, dtype="float64").pct_change()
        expected = (returns.rolling(3, min_periods=3).std() * 2.0).clip(lower=0.03, upper=0.10).shift(1)
        self.assertIn("trade_stop_loss_pct", signals.columns)
        self.assertTrue(pd.isna(signals.loc[3, "trade_stop_loss_pct"]))
        self.assertAlmostEqual(float(signals.loc[4, "trade_stop_loss_pct"]), float(expected.iloc[4]))

    def test_scaled_buy_orders_prorate_cash_shortage(self) -> None:
        orders = _scaled_buy_orders(
            {1: 100.0, 0: 100.0},
            np.array([10.0, 20.0]),
            np.array([True, True]),
            cash=100.0,
            commission_rate=0.0,
            slippage_rate=0.0,
            lot_size=1,
        )
        by_idx = {order.idx: order for order in orders}

        self.assertEqual(set(by_idx), {0, 1})
        self.assertAlmostEqual(sum(order.required_cash for order in orders), 100.0)
        self.assertAlmostEqual(by_idx[0].required_cash, 50.0)
        self.assertAlmostEqual(by_idx[1].required_cash, 50.0)

    def test_score_weighted_position_weights_can_use_bounded_tilt(self) -> None:
        raw = _score_weighted_position_weights(
            [0, 1],
            np.array([100.0, 1.0]),
            target_exposure=1.0,
            method="proportional",
        )
        tilted = _score_weighted_position_weights(
            [0, 1],
            np.array([100.0, 1.0]),
            target_exposure=1.0,
            method="tilt",
            tilt_strength=0.5,
            min_multiplier=0.5,
            max_multiplier=1.5,
        )

        self.assertAlmostEqual(float(raw.sum()), 1.0)
        self.assertGreater(float(raw[0] / raw[1]), 50.0)
        self.assertAlmostEqual(float(tilted.sum()), 1.0)
        self.assertAlmostEqual(float(tilted[0]), 0.75)
        self.assertAlmostEqual(float(tilted[1]), 0.25)

    def test_benchmark_curve_handles_empty_inputs(self) -> None:
        empty = _benchmark_curve({}, pd.Series(dtype="datetime64[ns]"), 100000.0)
        self.assertTrue(empty.empty)
        self.assertEqual(list(empty.columns), ["date", "benchmark_equity"])

        dates = pd.Series(pd.date_range("2020-01-01", periods=2, freq="D"))
        flat = _benchmark_curve({}, dates, 100000.0)
        self.assertEqual(list(flat["benchmark_equity"]), [100000.0, 100000.0])

    def test_market_sentiment_uses_prior_limit_up_premium(self) -> None:
        histories = {
            "000001": make_price_frame("000001", "A", [10.0, 11.0, 11.55, 11.0]),
            "000002": make_price_frame("000002", "B", [10.0, 10.0, 10.0, 10.0]),
        }
        sentiment = build_market_sentiment_frame(histories, window=20)
        row = sentiment[sentiment["date"] == pd.Timestamp("2020-01-03")].iloc[0]

        self.assertEqual(int(row["prior_limit_up_count"]), 1)
        self.assertAlmostEqual(float(row["prior_limit_up_premium"]), 5.0)
        self.assertEqual(int(row["coverage_count"]), 2)

    def test_market_sentiment_state_maps_to_reference_exposure(self) -> None:
        self.assertEqual(classify_sentiment_state(0.80), "hot")
        self.assertEqual(classify_sentiment_state(0.25), "warm")
        self.assertEqual(classify_sentiment_state(0.00), "neutral")
        self.assertEqual(classify_sentiment_state(-0.50), "weak")
        self.assertEqual(classify_sentiment_state(-0.90), "cold")
        self.assertAlmostEqual(reference_exposure("hot"), 1.0)
        self.assertAlmostEqual(reference_exposure("cold"), 0.3)

    def test_market_sentiment_sampled_config_falls_back_for_cache_errors(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            config = config.__class__(
                project_root=config.project_root,
                project=config.project,
                data=config.data,
                universe=config.universe,
                strategy=config.strategy,
                costs=config.costs,
                universe_file=config.universe_file,
                universe_source=UniverseSourceConfig(type="ashare_main_chinext", symbol="all", asset_type="stock"),
                risk=config.risk,
            )

            with patch("quant_etf_lab.market_sentiment.load_cached_universe", side_effect=FileNotFoundError("missing cache")), patch(
                "quant_etf_lab.market_sentiment.resolve_universe",
                return_value=config.universe,
            ):
                sampled = _load_sampled_config(config, max_symbols=1)

            self.assertEqual(len(sampled.universe), 1)
            self.assertIsNone(sampled.universe_source)

    def test_market_sentiment_sampled_config_does_not_swallow_unexpected_errors(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            config = config.__class__(
                project_root=config.project_root,
                project=config.project,
                data=config.data,
                universe=config.universe,
                strategy=config.strategy,
                costs=config.costs,
                universe_file=config.universe_file,
                universe_source=UniverseSourceConfig(type="ashare_main_chinext", symbol="all", asset_type="stock"),
                risk=config.risk,
            )

            with patch("quant_etf_lab.market_sentiment.load_cached_universe", side_effect=RuntimeError("programming bug")), patch(
                "quant_etf_lab.market_sentiment.resolve_universe",
                return_value=config.universe,
            ):
                with self.assertRaisesRegex(RuntimeError, "programming bug"):
                    _load_sampled_config(config, max_symbols=1)

    def test_symbol_factor_frame_uses_past_data_for_factors(self) -> None:
        frame = make_price_frame("000001", "A", [float(value) for value in range(1, 35)])
        base = build_symbol_factor_frame(frame, horizons=[1])
        mutated_frame = frame.copy()
        mutated_frame.loc[25:, "close"] = mutated_frame.loc[25:, "close"] * 10
        mutated = build_symbol_factor_frame(mutated_frame, horizons=[1])

        self.assertAlmostEqual(base.loc[20, "momentum_20d"], mutated.loc[20, "momentum_20d"])
        self.assertAlmostEqual(base.loc[20, "ma20_gap"], mutated.loc[20, "ma20_gap"])
        self.assertNotAlmostEqual(base.loc[24, "fwd_return_1d"], mutated.loc[24, "fwd_return_1d"])

    def test_factor_lab_writes_ic_and_quantile_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            data_dir = root / "data"
            data_dir.mkdir()
            dates = pd.date_range("2020-01-01", periods=90, freq="D")
            for idx, code in enumerate(["000001", "000002", "000003", "000004"]):
                drift = 0.001 * (idx + 1)
                closes = [10.0 * (1.0 + drift) ** day for day in range(len(dates))]
                frame = pd.DataFrame(
                    {
                        "date": dates,
                        "code": code,
                        "name": f"S{idx}",
                        "open": closes,
                        "high": [value * 1.01 for value in closes],
                        "low": [value * 0.99 for value in closes],
                        "close": closes,
                        "volume": [1000 * (idx + 1)] * len(dates),
                        "amount": [close * 1000 * (idx + 1) for close in closes],
                    }
                )
                frame.to_csv(data_dir / f"{code}.csv", index=False)

            result = run_factor_lab(
                project_root=root,
                data_dir=data_dir,
                output_dir=root / "factor_lab",
                factors=["momentum_20d", "reversal_5d"],
                horizons=[1, 5],
                quantiles=2,
                min_obs=3,
                save_panel=False,
            )

            self.assertTrue(result.report_path.exists())
            self.assertTrue(result.ic_summary_path.exists())
            self.assertTrue(result.quantile_summary_path.exists())
            self.assertFalse(result.factor_panel_path.exists())
            self.assertFalse(result.panel.empty)
            self.assertFalse(result.ic_summary.empty)
            self.assertFalse(result.quantile_summary.empty)
            self.assertEqual(result.snapshot["loaded_symbol_count"], 4)
            self.assertFalse(result.snapshot["save_panel"])
            self.assertIn("Factor Lab Report", result.report_path.read_text(encoding="utf-8"))

    def test_mutual_information_detects_nonlinear_dependency_when_correlation_is_small(self) -> None:
        x = np.array([-2.0, -1.0, 0.0, 1.0, 2.0, -2.0, -1.0, 0.0, 1.0, 2.0])
        y = x * x

        corr = float(np.corrcoef(x, y)[0, 1])
        mi = estimate_mutual_information(x, y, bins=4)

        self.assertAlmostEqual(corr, 0.0, places=8)
        self.assertGreater(mi, 0.1)

    def test_network_lab_writes_research_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            data_dir = root / "data" / "processed" / "stocks"
            data_dir.mkdir(parents=True)
            dates = pd.date_range("2024-01-01", periods=11, freq="D")
            returns_by_code = {
                "000001": [0.010, 0.020, -0.010, 0.030, 0.000, 0.010, 0.020, -0.020, 0.010, 0.030],
                "000002": [0.012, 0.018, -0.011, 0.028, 0.001, 0.009, 0.019, -0.018, 0.012, 0.029],
                "000003": [-0.010, 0.005, 0.015, -0.020, 0.010, -0.005, 0.020, 0.000, -0.015, 0.005],
            }
            for code, returns in returns_by_code.items():
                closes = [10.0]
                for item in returns:
                    closes.append(closes[-1] * (1.0 + item))
                frame = pd.DataFrame(
                    {
                        "date": dates,
                        "code": code,
                        "name": f"S{code[-1]}",
                        "open": closes,
                        "high": [value * 1.01 for value in closes],
                        "low": [value * 0.99 for value in closes],
                        "close": closes,
                        "volume": [1000] * len(closes),
                        "amount": [value * 1000 for value in closes],
                    }
                )
                frame.to_csv(data_dir / f"{code}.csv", index=False)

            result = run_network_lab(
                project_root=root,
                data_dir=data_dir,
                output_dir=root / "network_lab",
                symbols=["000001", "000002", "000003"],
                min_obs=6,
                top_edges=3,
                bins=4,
            )

            self.assertTrue(result.report_path.exists())
            self.assertTrue(result.edge_path.exists())
            self.assertTrue(result.mst_path.exists())
            self.assertEqual(result.snapshot["research_only"], True)
            self.assertEqual(result.snapshot["broker_action"], "none")
            self.assertEqual(result.snapshot["loaded_symbol_count"], 3)
            self.assertEqual(result.snapshot["mst_edge_count"], 2)
            self.assertGreaterEqual(result.snapshot["top_residual_mutual_information"], 0.0)
            pair = result.edges[
                ((result.edges["source"] == "000001") & (result.edges["target"] == "000002"))
                | ((result.edges["source"] == "000002") & (result.edges["target"] == "000001"))
            ].iloc[0]
            self.assertGreater(float(pair["correlation"]), 0.95)
            self.assertLess(float(pair["correlation_distance"]), 0.35)
            self.assertIn("Network Lab Report", result.report_path.read_text(encoding="utf-8"))

    def test_filter_history_by_date_is_inclusive(self) -> None:
        frame = make_price_frame("TEST", "Test ETF", [10, 11, 12, 13, 14])
        filtered = filter_history_by_date(frame, "20200102", "20200104")
        self.assertEqual(
            list(filtered["date"].dt.strftime("%Y-%m-%d")),
            ["2020-01-02", "2020-01-03", "2020-01-04"],
        )

    def test_backtest_uses_next_open_after_signal(self) -> None:
        dates = pd.date_range("2020-01-01", periods=6, freq="D")
        frame = pd.DataFrame(
            {
                "date": dates,
                "code": "TEST",
                "name": "Test ETF",
                "open": [10, 10, 10, 12, 13, 14],
                "high": [10, 10, 11, 13, 14, 15],
                "low": [9, 9, 9, 11, 12, 13],
                "close": [10, 10, 10, 12, 13, 14],
                "volume": [1000] * 6,
                "amount": [10000] * 6,
            }
        )
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            result = run_backtest(config, histories={"TEST": frame}, run_id="test", write_outputs=False)
        self.assertFalse(result.trades.empty)
        first_trade = result.trades.iloc[0]
        self.assertEqual(first_trade["side"], "BUY")
        self.assertEqual(pd.Timestamp(first_trade["date"]), pd.Timestamp("2020-01-05"))
        self.assertAlmostEqual(first_trade["price"], 13 * 1.01)
        self.assertGreater(first_trade["fee"], 0)

    def test_backtest_accepts_precomputed_signal_frames(self) -> None:
        frame = make_price_frame("TEST", "Test ETF", [10, 10, 11, 12, 13])
        signals = frame.copy()
        signals["trade_signal"] = [0, 1, 1, 0, 0]
        signals["trade_score"] = [0, 1, 1, 0, 0]
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            with patch("quant_etf_lab.backtest.build_signal_frames", side_effect=AssertionError("should not rebuild")):
                result = run_backtest(
                    config,
                    histories={"TEST": frame},
                    signal_frames={"TEST": signals},
                    run_id="test",
                    write_outputs=False,
                )
        self.assertFalse(result.equity.empty)
        self.assertFalse(result.trades.empty)

    def test_backtest_score_weighted_allocation_prefers_higher_score(self) -> None:
        dates = pd.date_range("2020-01-01", periods=3, freq="D")
        high_score_frame = make_ohlcv_frame(
            "AAA",
            "Alpha",
            opens=[10.0, 10.0, 10.0],
            highs=[11.0, 11.0, 11.0],
            lows=[9.0, 9.0, 9.0],
            closes=[10.0, 10.0, 10.0],
            volumes=[1000, 1000, 1000],
        )
        low_score_frame = make_ohlcv_frame(
            "BBB",
            "Beta",
            opens=[10.0, 10.0, 10.0],
            highs=[11.0, 11.0, 11.0],
            lows=[9.0, 9.0, 9.0],
            closes=[10.0, 10.0, 10.0],
            volumes=[1000, 1000, 1000],
        )
        signal_frames = {
            "AAA": high_score_frame.assign(
                trade_signal=[1, 1, 1],
                trade_score=[10.0, 10.0, 10.0],
                trade_exit_risk=[0.0, 0.0, 0.0],
                trade_entry_allowed=[1.0, 1.0, 1.0],
            ),
            "BBB": low_score_frame.assign(
                trade_signal=[1, 1, 1],
                trade_score=[1.0, 1.0, 1.0],
                trade_exit_risk=[0.0, 0.0, 0.0],
                trade_entry_allowed=[1.0, 1.0, 1.0],
            ),
        }
        with tempfile.TemporaryDirectory() as tmp:
            base_config = make_config(Path(tmp))
            strategy_weighted = base_config.strategy.__class__(
                **{
                    **base_config.strategy.__dict__,
                    "score_weighted_allocation": True,
                    "allocation_buffer": 1.0,
                    "lot_size": 1,
                    "max_positions": 2,
                }
            )
            strategy_equal = base_config.strategy.__class__(
                **{
                    **base_config.strategy.__dict__,
                    "score_weighted_allocation": False,
                    "allocation_buffer": 1.0,
                    "lot_size": 1,
                    "max_positions": 2,
                }
            )
            strategy_tilt = base_config.strategy.__class__(
                **{
                    **base_config.strategy.__dict__,
                    "score_weighted_allocation": True,
                    "score_allocation_method": "tilt",
                    "score_allocation_tilt_strength": 0.5,
                    "score_allocation_min_multiplier": 0.5,
                    "score_allocation_max_multiplier": 1.5,
                    "allocation_buffer": 1.0,
                    "lot_size": 1,
                    "max_positions": 2,
                }
            )
            config_weighted = base_config.__class__(
                project_root=base_config.project_root,
                project=base_config.project,
                data=base_config.data,
                universe=(ETFSpec(code="AAA", name="Alpha"), ETFSpec(code="BBB", name="Beta")),
                strategy=strategy_weighted,
                costs=base_config.costs,
                universe_file=base_config.universe_file,
                universe_source=base_config.universe_source,
                risk=base_config.risk,
            )
            config_equal = base_config.__class__(
                project_root=base_config.project_root,
                project=base_config.project,
                data=base_config.data,
                universe=(ETFSpec(code="AAA", name="Alpha"), ETFSpec(code="BBB", name="Beta")),
                strategy=strategy_equal,
                costs=base_config.costs,
                universe_file=base_config.universe_file,
                universe_source=base_config.universe_source,
                risk=base_config.risk,
            )
            config_tilt = base_config.__class__(
                project_root=base_config.project_root,
                project=base_config.project,
                data=base_config.data,
                universe=(ETFSpec(code="AAA", name="Alpha"), ETFSpec(code="BBB", name="Beta")),
                strategy=strategy_tilt,
                costs=base_config.costs,
                universe_file=base_config.universe_file,
                universe_source=base_config.universe_source,
                risk=base_config.risk,
            )
            with patch(
                "quant_etf_lab.backtest.build_signal_frames",
                side_effect=AssertionError("should not rebuild"),
            ):
                weighted = run_backtest(
                    config_weighted,
                    histories={"AAA": high_score_frame, "BBB": low_score_frame},
                    signal_frames=signal_frames,
                    run_id="weighted",
                    write_outputs=False,
                )
                equal = run_backtest(
                    config_equal,
                    histories={"AAA": high_score_frame, "BBB": low_score_frame},
                    signal_frames=signal_frames,
                    run_id="equal",
                    write_outputs=False,
                )
                tilted = run_backtest(
                    config_tilt,
                    histories={"AAA": high_score_frame, "BBB": low_score_frame},
                    signal_frames=signal_frames,
                    run_id="tilted",
                    write_outputs=False,
                )

        weighted_buys = weighted.trades[weighted.trades["side"] == "BUY"]
        equal_buys = equal.trades[equal.trades["side"] == "BUY"]
        tilted_buys = tilted.trades[tilted.trades["side"] == "BUY"]
        weighted_qty = dict(zip(weighted_buys["code"], weighted_buys["quantity"]))
        equal_qty = dict(zip(equal_buys["code"], equal_buys["quantity"]))
        tilted_qty = dict(zip(tilted_buys["code"], tilted_buys["quantity"]))
        self.assertGreater(weighted_qty["AAA"], weighted_qty["BBB"] * 4)
        self.assertTrue(0.9 <= float(equal_qty["AAA"] / equal_qty["BBB"]) <= 1.2)
        self.assertTrue(2.5 <= float(tilted_qty["AAA"] / tilted_qty["BBB"]) <= 3.5)

    def test_backtest_stop_loss_exit_reason(self) -> None:
        dates = pd.date_range("2020-01-01", periods=8, freq="D")
        frame = pd.DataFrame(
            {
                "date": dates,
                "code": "TEST",
                "name": "Test ETF",
                "open": [10, 10, 10, 12, 13, 11, 11, 11],
                "high": [10, 10, 11, 13, 14, 12, 12, 12],
                "low": [9, 9, 9, 11, 12, 10, 10, 10],
                "close": [10, 10, 10, 12, 13, 11, 11, 11],
                "volume": [1000] * 8,
                "amount": [10000] * 8,
            }
        )
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            strategy = config.strategy.__class__(
                **{**config.strategy.__dict__, "stop_loss_pct": 0.10, "min_hold_days": 20}
            )
            config = config.__class__(
                project_root=config.project_root,
                project=config.project,
                data=config.data,
                universe=config.universe,
                strategy=strategy,
                costs=config.costs,
                universe_file=config.universe_file,
                universe_source=config.universe_source,
                risk=config.risk,
            )
            result = run_backtest(config, histories={"TEST": frame}, run_id="test", write_outputs=False)
        sells = result.trades[result.trades["side"] == "SELL"]
        self.assertFalse(sells.empty)
        self.assertIn("stop_loss", set(sells["exit_reason"]))

    def test_backtest_uses_signal_frame_dynamic_stop_loss(self) -> None:
        dates = pd.date_range("2020-01-01", periods=4, freq="D")
        frame = pd.DataFrame(
            {
                "date": dates,
                "code": "TEST",
                "name": "Test ETF",
                "open": [10.0, 10.0, 9.4, 9.4],
                "high": [10.0, 10.0, 9.4, 9.4],
                "low": [10.0, 10.0, 9.4, 9.4],
                "close": [10.0, 10.0, 9.4, 9.4],
                "volume": [1000] * 4,
                "amount": [10000] * 4,
                "trade_signal": [1, 1, 1, 1],
                "trade_score": [1.0, 1.0, 1.0, 1.0],
                "trade_stop_loss_pct": [0.05, 0.05, 0.05, 0.05],
            }
        )
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            config = config.__class__(
                project_root=config.project_root,
                project=config.project,
                data=config.data,
                universe=config.universe,
                strategy=config.strategy.__class__(
                    **{
                        **config.strategy.__dict__,
                        "lot_size": 1,
                    }
                ),
                costs=CostsConfig(commission_rate=0.0, slippage_rate=0.0),
                universe_file=config.universe_file,
                universe_source=config.universe_source,
                risk=config.risk,
            )
            result = run_backtest(
                config,
                histories={"TEST": frame},
                signal_frames={"TEST": frame},
                run_id="test",
                write_outputs=False,
            )

        sells = result.trades[result.trades["side"] == "SELL"]
        self.assertFalse(sells.empty)
        self.assertEqual(set(sells["exit_reason"]), {"stop_loss"})
        self.assertEqual(pd.Timestamp(sells.iloc[0]["date"]), pd.Timestamp("2020-01-03"))

    def test_backtest_time_stop_exits_stagnant_position_after_min_hold(self) -> None:
        dates = pd.date_range("2020-01-01", periods=6, freq="D")
        frame = pd.DataFrame(
            {
                "date": dates,
                "code": "TEST",
                "name": "Test ETF",
                "open": [10.0, 10.0, 10.1, 10.1, 10.1, 10.1],
                "high": [10.0, 10.1, 10.1, 10.1, 10.1, 10.1],
                "low": [10.0, 10.0, 10.0, 10.0, 10.0, 10.0],
                "close": [10.0, 10.1, 10.1, 10.1, 10.1, 10.1],
                "volume": [1000] * 6,
                "amount": [10000] * 6,
                "trade_signal": [1, 1, 1, 1, 1, 1],
                "trade_score": [1.0] * 6,
            }
        )
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            strategy = config.strategy.__class__(
                **{
                    **config.strategy.__dict__,
                    "lot_size": 1,
                    "time_stop_enabled": True,
                    "time_stop_min_hold_days": 3,
                    "time_stop_return_threshold_pct": 0.02,
                }
            )
            config = config.__class__(
                project_root=config.project_root,
                project=config.project,
                data=config.data,
                universe=config.universe,
                strategy=strategy,
                costs=CostsConfig(commission_rate=0.0, slippage_rate=0.0),
                universe_file=config.universe_file,
                universe_source=config.universe_source,
                risk=config.risk,
            )
            result = run_backtest(
                config,
                histories={"TEST": frame},
                signal_frames={"TEST": frame},
                run_id="test",
                write_outputs=False,
            )

        sells = result.trades[result.trades["side"] == "SELL"]
        self.assertFalse(sells.empty)
        time_stop_sells = sells[sells["exit_reason"] == "time_stop"]
        self.assertFalse(time_stop_sells.empty)
        self.assertEqual(pd.Timestamp(time_stop_sells.iloc[0]["date"]), pd.Timestamp("2020-01-04"))

    def test_loss_cooldown_blocks_immediate_reentry(self) -> None:
        dates = pd.date_range("2020-01-01", periods=6, freq="D")
        frame = pd.DataFrame(
            {
                "date": dates,
                "code": "TEST",
                "name": "Test ETF",
                "open": [10, 8, 8, 8, 10, 10],
                "high": [10, 8, 8, 8, 10, 10],
                "low": [10, 8, 8, 8, 10, 10],
                "close": [10, 8, 8, 8, 10, 10],
                "volume": [1000] * 6,
                "amount": [10000] * 6,
                "trade_signal": [1] * 6,
                "trade_score": [1.0] * 6,
            }
        )
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            strategy = config.strategy.__class__(
                **{
                    **config.strategy.__dict__,
                    "stop_loss_pct": 0.10,
                    "loss_cooldown_days": 2,
                    "loss_cooldown_min_losses": 1,
                }
            )
            config = config.__class__(
                project_root=config.project_root,
                project=config.project,
                data=config.data,
                universe=config.universe,
                strategy=strategy,
                costs=config.costs,
                universe_file=config.universe_file,
                universe_source=config.universe_source,
                risk=config.risk,
            )
            with patch("quant_etf_lab.backtest.build_signal_frames", return_value={"TEST": frame}):
                result = run_backtest(config, histories={"TEST": frame}, run_id="test", write_outputs=False)
        buys = result.trades[result.trades["side"] == "BUY"]
        sells = result.trades[result.trades["side"] == "SELL"]
        self.assertEqual(
            list(buys["date"].dt.strftime("%Y-%m-%d")),
            ["2020-01-01", "2020-01-05"],
        )
        self.assertEqual(list(sells["exit_reason"]), ["stop_loss"])
        self.assertEqual(int(result.metrics["cooldown_event_count"]), 1)
        self.assertEqual(result.cooldown_events.loc[0, "cooldown_until"], pd.Timestamp("2020-01-04"))

    def test_loss_cooldown_records_beyond_backtest_end(self) -> None:
        dates = pd.date_range("2020-01-01", periods=3, freq="D")
        frame = pd.DataFrame(
            {
                "date": dates,
                "code": "TEST",
                "name": "Test ETF",
                "open": [10, 8, 8],
                "high": [10, 8, 8],
                "low": [10, 8, 8],
                "close": [10, 8, 8],
                "volume": [1000] * 3,
                "amount": [10000] * 3,
                "trade_signal": [1] * 3,
                "trade_score": [1.0] * 3,
            }
        )
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            strategy = config.strategy.__class__(
                **{
                    **config.strategy.__dict__,
                    "stop_loss_pct": 0.10,
                    "loss_cooldown_days": 10,
                    "loss_cooldown_min_losses": 1,
                }
            )
            config = config.__class__(
                project_root=config.project_root,
                project=config.project,
                data=config.data,
                universe=config.universe,
                strategy=strategy,
                costs=config.costs,
                universe_file=config.universe_file,
                universe_source=config.universe_source,
                risk=config.risk,
            )
            with patch("quant_etf_lab.backtest.build_signal_frames", return_value={"TEST": frame}):
                result = run_backtest(config, histories={"TEST": frame}, run_id="test", write_outputs=False)

        event = result.cooldown_events.iloc[0]
        self.assertEqual(event["cooldown_until"], pd.Timestamp("2020-01-03"))
        self.assertEqual(int(event["cooldown_until_index"]), 11)
        self.assertTrue(bool(event["cooldown_beyond_backtest_end"]))

    def test_backtest_forces_exit_when_stock_exceeds_market_cap_limit(self) -> None:
        dates = pd.date_range("2020-01-01", periods=3, freq="D")
        signal_frames = {
            "000001": pd.DataFrame(
                {
                    "date": dates,
                    "code": "000001",
                    "name": "Allowed",
                    "open": [10.0, 10.0, 10.0],
                    "high": [10.0, 10.0, 10.0],
                    "low": [10.0, 10.0, 10.0],
                    "close": [10.0, 10.0, 10.0],
                    "volume": [1000, 1000, 1000],
                    "amount": [10000, 10000, 10000],
                    "trade_signal": [1, 1, 1],
                    "trade_score": [1.0, 1.0, 1.0],
                }
            ),
            "600001": pd.DataFrame(
                {
                    "date": dates,
                    "code": "600001",
                    "name": "Blocked",
                    "open": [10.0, 10.0, 10.0],
                    "high": [10.0, 10.0, 10.0],
                    "low": [10.0, 10.0, 10.0],
                    "close": [10.0, 10.0, 10.0],
                    "volume": [1000, 1000, 1000],
                    "amount": [10000, 10000, 10000],
                    "trade_signal": [1, 1, 1],
                    "trade_score": [1.0, 1.0, 1.0],
                }
            ),
        }
        histories = {code: frame[["date", "code", "name", "open", "high", "low", "close", "volume", "amount"]]
                     for code, frame in signal_frames.items()}
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            strategy = config.strategy.__class__(
                **{
                    **config.strategy.__dict__,
                    "max_positions": 2,
                    "lot_size": 1,
                    "allocation_buffer": 1.0,
                }
            )
            cap_path = Path(tmp) / "data" / "processed" / "stock_market_cap_yi.csv"
            cap_path.parent.mkdir(parents=True, exist_ok=True)
            pd.DataFrame(
                {
                    "code": ["000001", "600001"],
                    "snapshot_date": ["2020-01-02", "2020-01-02"],
                    "market_cap_yi": [1000.0, 2000.0],
                }
            ).to_csv(cap_path, index=False)
            config = config.__class__(
                project_root=config.project_root,
                project=config.project,
                data=config.data,
                universe=(
                    ETFSpec(code="000001", name="Allowed"),
                    ETFSpec(code="600001", name="Blocked"),
                ),
                strategy=strategy,
                costs=CostsConfig(commission_rate=0.0, slippage_rate=0.0),
                universe_file=config.universe_file,
                universe_source=config.universe_source,
                stock_market_cap_path=cap_path,
                stock_tracking_max_market_cap_yi=1500.0,
                risk=config.risk,
            )
            result = run_backtest(
                config,
                histories=histories,
                signal_frames=signal_frames,
                run_id="stock-cap-blocked-exit-test",
                write_outputs=False,
            )

        sells = result.trades[result.trades["side"] == "SELL"]
        blocked_sells = sells[sells["code"] == "600001"]
        self.assertEqual(len(blocked_sells), 1)
        self.assertEqual(blocked_sells.iloc[0]["exit_reason"], "market_cap_blocked")
        self.assertEqual(pd.Timestamp(blocked_sells.iloc[0]["date"]), pd.Timestamp("2020-01-02"))
        self.assertEqual(len(sells), 1)
        self.assertEqual(result.equity["stock_cap_blocked_count"].iloc[1], 1)
        self.assertEqual(int(result.equity["active_positions"].iloc[-1]), 1)
        self.assertEqual(int(result.metrics["stock_cap_blocked_days"]), 2)

    def test_backtest_blocks_new_entry_after_cap_threshold(self) -> None:
        dates = pd.date_range("2020-01-01", periods=3, freq="D")
        signal_frames = {
            "000001": pd.DataFrame(
                {
                    "date": dates,
                    "code": "000001",
                    "name": "Allowed",
                    "open": [10.0, 10.0, 10.0],
                    "high": [10.0, 10.0, 10.0],
                    "low": [10.0, 10.0, 10.0],
                    "close": [10.0, 10.0, 10.0],
                    "volume": [1000, 1000, 1000],
                    "amount": [10000, 10000, 10000],
                    "trade_signal": [1, 1, 1],
                    "trade_score": [1.0, 1.0, 1.0],
                }
            ),
            "600001": pd.DataFrame(
                {
                    "date": dates,
                    "code": "600001",
                    "name": "Blocked",
                    "open": [10.0, 10.0, 10.0],
                    "high": [10.0, 10.0, 10.0],
                    "low": [10.0, 10.0, 10.0],
                    "close": [10.0, 10.0, 10.0],
                    "volume": [1000, 1000, 1000],
                    "amount": [10000, 10000, 10000],
                    "trade_signal": [1, 1, 1],
                    "trade_score": [1.0, 1.0, 1.0],
                }
            ),
        }
        histories = {code: frame[["date", "code", "name", "open", "high", "low", "close", "volume", "amount"]]
                     for code, frame in signal_frames.items()}
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            cap_path = Path(tmp) / "data" / "processed" / "stock_market_cap_yi.csv"
            cap_path.parent.mkdir(parents=True, exist_ok=True)
            pd.DataFrame(
                {
                    "code": ["000001", "600001"],
                    "snapshot_date": ["2020-01-01", "2020-01-01"],
                    "market_cap_yi": [1000.0, 2000.0],
                }
            ).to_csv(cap_path, index=False)
            config = config.__class__(
                project_root=config.project_root,
                project=config.project,
                data=config.data,
                universe=(
                    ETFSpec(code="000001", name="Allowed"),
                    ETFSpec(code="600001", name="Blocked"),
                ),
                strategy=config.strategy,
                costs=config.costs,
                universe_file=config.universe_file,
                universe_source=config.universe_source,
                stock_market_cap_path=cap_path,
                stock_tracking_max_market_cap_yi=1500.0,
                risk=config.risk,
            )
            result = run_backtest(
                config,
                histories=histories,
                signal_frames=signal_frames,
                run_id="stock-cap-blocked-entry-test",
                write_outputs=False,
            )

        buys = result.trades[result.trades["side"] == "BUY"]
        self.assertEqual(set(buys["code"]), {"000001"})
        self.assertFalse(result.equity["stock_cap_blocked_count"].empty)
        self.assertTrue((result.equity["stock_cap_blocked_count"] == 1).all())

    def test_rebalance_loss_guard_defers_small_loss_rebalance_exit(self) -> None:
        dates = pd.date_range("2020-01-01", periods=4, freq="D")
        frame = pd.DataFrame(
            {
                "date": dates,
                "code": "TEST",
                "name": "Test ETF",
                "open": [10, 9.5, 9.5, 9.5],
                "high": [10, 9.5, 9.5, 9.5],
                "low": [10, 9.5, 9.5, 9.5],
                "close": [10, 9.5, 9.5, 9.5],
                "volume": [1000] * 4,
                "amount": [10000] * 4,
                "trade_signal": [1, 0, 0, 0],
                "trade_score": [1.0, 0.0, 0.0, 0.0],
            }
        )
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            strategy = config.strategy.__class__(
                **{
                    **config.strategy.__dict__,
                    "rebalance_loss_guard_pct": 0.10,
                    "rebalance_loss_guard_max_days": 3,
                }
            )
            config = config.__class__(
                project_root=config.project_root,
                project=config.project,
                data=config.data,
                universe=config.universe,
                strategy=strategy,
                costs=config.costs,
                universe_file=config.universe_file,
                universe_source=config.universe_source,
                risk=config.risk,
            )
            with patch("quant_etf_lab.backtest.build_signal_frames", return_value={"TEST": frame}):
                result = run_backtest(config, histories={"TEST": frame}, run_id="test", write_outputs=False)
        self.assertEqual(len(result.trades[result.trades["side"] == "SELL"]), 0)
        self.assertGreater(float(result.metrics["rebalance_loss_guarded_day_ratio"]), 0.0)
        self.assertGreater(int(result.equity["active_positions"].iloc[-1]), 0)

    def test_rebalance_loss_guard_allows_large_loss_exit(self) -> None:
        dates = pd.date_range("2020-01-01", periods=4, freq="D")
        frame = pd.DataFrame(
            {
                "date": dates,
                "code": "TEST",
                "name": "Test ETF",
                "open": [10, 8.5, 8.5, 8.5],
                "high": [10, 8.5, 8.5, 8.5],
                "low": [10, 8.5, 8.5, 8.5],
                "close": [10, 8.5, 8.5, 8.5],
                "volume": [1000] * 4,
                "amount": [10000] * 4,
                "trade_signal": [1, 0, 0, 0],
                "trade_score": [1.0, 0.0, 0.0, 0.0],
            }
        )
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            strategy = config.strategy.__class__(**{**config.strategy.__dict__, "rebalance_loss_guard_pct": 0.10})
            config = config.__class__(
                project_root=config.project_root,
                project=config.project,
                data=config.data,
                universe=config.universe,
                strategy=strategy,
                costs=config.costs,
                universe_file=config.universe_file,
                universe_source=config.universe_source,
                risk=config.risk,
            )
            with patch("quant_etf_lab.backtest.build_signal_frames", return_value={"TEST": frame}):
                result = run_backtest(config, histories={"TEST": frame}, run_id="test", write_outputs=False)
        sells = result.trades[result.trades["side"] == "SELL"]
        self.assertFalse(sells.empty)
        self.assertEqual(set(sells["exit_reason"]), {"rebalance"})

    def test_constituents_normalization_and_stock_cache_paths(self) -> None:
        raw = pd.DataFrame(
            {
                "成分券代码": ["1", "000002"],
                "成分券名称": ["平安银行", "万科A"],
                "日期": ["2026-06-11", "2026-06-11"],
            }
        )
        frame = normalize_constituents_frame(raw, "stock")
        self.assertEqual(list(frame["code"]), ["000001", "000002"])
        self.assertEqual(frame.loc[0, "asset_type"], "stock")

        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            stock = ETFSpec(code="000001", name="平安银行", asset_type="stock")
            self.assertIn("stocks", str(cache_paths(config, stock)["processed"]))

    def test_a_share_main_board_filter(self) -> None:
        raw = pd.DataFrame(
            {
                "证券代码": ["600000", "688001", "000001", "300001", "002001", "430001"],
                "证券简称": ["浦发银行", "科创样本", "平安银行", "创业样本", "新和成", "北交样本"],
                "板块": ["主板", "科创板", "主板", "创业板", "主板", "北交所"],
                "上市日期": ["1999-11-10"] * 6,
            }
        )
        frame = normalize_a_share_main_board_frame(raw)
        self.assertEqual(list(frame["code"]), ["000001", "002001", "600000"])
        self.assertTrue(is_a_share_main_board_code("603000"))
        self.assertFalse(is_a_share_main_board_code("688001"))
        self.assertFalse(is_a_share_main_board_code("300001"))
        self.assertTrue(is_a_share_chinext_code("301001"))
        self.assertTrue(is_a_share_main_or_chinext_code("300001"))

    def test_a_share_main_board_filter_coalesces_sh_and_sz_columns(self) -> None:
        raw = pd.DataFrame(
            {
                "证券代码": ["600000", None],
                "证券简称": ["浦发银行", None],
                "A股代码": [None, "000001"],
                "A股简称": [None, "平安银行"],
                "板块": [None, "主板"],
            }
        )
        frame = normalize_a_share_main_board_frame(raw)
        self.assertEqual(list(frame["code"]), ["000001", "600000"])

    def test_a_share_main_chinext_filter_includes_chinext(self) -> None:
        raw = pd.DataFrame(
            {
                "证券代码": ["600000", "688001", "000001", "300001", "301001", "430001"],
                "证券简称": ["浦发银行", "科创样本", "平安银行", "创业样本", "创业样本2", "北交样本"],
                "板块": ["主板", "科创板", "主板", "创业板", "创业板", "北交所"],
            }
        )
        frame = normalize_a_share_main_board_frame(raw, include_chinext=True)
        self.assertEqual(list(frame["code"]), ["000001", "300001", "301001", "600000"])

    def test_parse_config_with_csi300_universe_source(self) -> None:
        raw = {
            "project": {"data_dir": "data", "output_dir": "outputs/backtests"},
            "universe_source": {"type": "csindex", "symbol": "000300", "asset_type": "stock", "limit": 5},
            "strategy": {"name": "trend", "max_positions": 2, "lot_size": 100},
            "costs": {"commission_rate": 0.0003, "slippage_rate": 0.0005, "stamp_tax_rate": 0.0005},
        }
        config = parse_config(raw, Path("D:/codex/量化"))
        self.assertEqual(config.universe_source.symbol, "000300")
        self.assertEqual(config.universe_source.limit, 5)
        self.assertEqual(config.costs.stamp_tax_rate, 0.0005)

    def test_parse_config_with_main_board_universe_source(self) -> None:
        raw = {
            "project": {"data_dir": "data", "output_dir": "outputs/backtests"},
            "universe_source": {"type": "ashare_main_board", "symbol": "all", "asset_type": "stock", "limit": 80},
            "strategy": {"name": "thsdk_monitor", "max_positions": 5},
        }
        config = parse_config(raw, Path("D:/codex/量化"))
        self.assertEqual(config.universe_source.type, "ashare_main_board")
        self.assertEqual(config.universe_source.symbol, "all")
        self.assertEqual(config.universe_source.limit, 80)

    def test_parse_config_with_main_chinext_universe_source(self) -> None:
        raw = {
            "project": {"data_dir": "data", "output_dir": "outputs/backtests"},
            "universe_source": {"type": "ashare_main_chinext", "symbol": "all", "asset_type": "stock", "limit": 120},
            "strategy": {"name": "thsdk_monitor", "max_positions": 5},
        }
        config = parse_config(raw, Path("D:/codex/量化"))
        self.assertEqual(config.universe_source.type, "ashare_main_chinext")
        self.assertEqual(config.universe_source.symbol, "all")
        self.assertEqual(config.universe_source.limit, 120)

    def test_load_custom_universe_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "custom.csv"
            path.write_text("code,name,asset_type\n70,特发信息,stock\n000070,重复,stock\n300285,国瓷材料,stock\n", encoding="utf-8")
            frame = load_universe_file(path)
        self.assertEqual(list(frame["code"]), ["000070", "300285"])
        self.assertEqual(frame.loc[0, "name"], "特发信息")

    def test_parse_config_with_custom_universe_file(self) -> None:
        raw = {
            "project": {"data_dir": "data", "output_dir": "outputs/backtests"},
            "universe_file": "configs/custom_stocks.csv",
            "strategy": {"name": "trend", "max_positions": 5, "min_hold_days": 5},
        }
        config = parse_config(raw, Path("D:/codex/量化"))
        self.assertEqual(config.universe_file, Path("D:/codex/量化/configs/custom_stocks.csv"))
        self.assertEqual(config.strategy.min_hold_days, 5)

    def test_parse_config_score_weighted_allocation_default_false(self) -> None:
        raw = {
            "project": {"data_dir": "data", "output_dir": "outputs/backtests"},
            "universe": [{"code": "TEST", "name": "Test ETF"}],
            "strategy": {"name": "trend", "max_positions": 1},
        }
        config = parse_config(raw, Path("D:/codex/量化"))
        self.assertFalse(config.strategy.score_weighted_allocation)

    def test_parse_config_score_weighted_allocation_enabled(self) -> None:
        raw = {
            "project": {"data_dir": "data", "output_dir": "outputs/backtests"},
            "universe": [{"code": "TEST", "name": "Test ETF"}],
            "strategy": {
                "name": "trend",
                "max_positions": 1,
                "score_weighted_allocation": True,
                "score_allocation_method": "tilt",
                "score_allocation_tilt_strength": 0.25,
                "score_allocation_min_multiplier": 0.75,
                "score_allocation_max_multiplier": 1.25,
            },
        }
        config = parse_config(raw, Path("D:/codex/量化"))
        self.assertTrue(config.strategy.score_weighted_allocation)
        self.assertEqual(config.strategy.score_allocation_method, "tilt")
        self.assertAlmostEqual(config.strategy.score_allocation_tilt_strength, 0.25)
        self.assertAlmostEqual(config.strategy.score_allocation_min_multiplier, 0.75)
        self.assertAlmostEqual(config.strategy.score_allocation_max_multiplier, 1.25)

    def test_parse_config_volatility_stop_loss(self) -> None:
        raw = {
            "project": {"data_dir": "data", "output_dir": "outputs/backtests"},
            "universe": [{"code": "TEST", "name": "Test ETF"}],
            "strategy": {
                "name": "trend",
                "max_positions": 1,
                "volatility_stop_loss_enabled": True,
                "volatility_stop_window": 10,
                "volatility_stop_multiplier": 2.5,
                "volatility_stop_min_pct": 0.04,
                "volatility_stop_max_pct": 0.15,
            },
        }
        config = parse_config(raw, Path("D:/codex/量化"))
        self.assertTrue(config.strategy.volatility_stop_loss_enabled)
        self.assertEqual(config.strategy.volatility_stop_window, 10)
        self.assertAlmostEqual(config.strategy.volatility_stop_multiplier, 2.5)
        self.assertAlmostEqual(config.strategy.volatility_stop_min_pct, 0.04)
        self.assertAlmostEqual(config.strategy.volatility_stop_max_pct, 0.15)

    def test_parse_config_time_stop(self) -> None:
        raw = {
            "project": {"data_dir": "data", "output_dir": "outputs/backtests"},
            "universe": [{"code": "TEST", "name": "Test ETF"}],
            "strategy": {
                "name": "trend",
                "max_positions": 1,
                "time_stop_enabled": True,
                "time_stop_min_hold_days": 3,
                "time_stop_return_threshold_pct": 0.02,
            },
        }
        config = parse_config(raw, Path("D:/codex/量化"))
        self.assertTrue(config.strategy.time_stop_enabled)
        self.assertEqual(config.strategy.time_stop_min_hold_days, 3)
        self.assertAlmostEqual(config.strategy.time_stop_return_threshold_pct, 0.02)

    def test_load_config_extends_base_with_deep_merge(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            configs = root / "configs"
            variants = configs / "variants"
            variants.mkdir(parents=True)
            (configs / "base.yaml").write_text(
                """
project:
  name: base_model
  data_dir: data
  output_dir: outputs/backtests
universe:
  - code: "000001"
    name: Ping An
    asset_type: stock
strategy:
  name: multi_factor
  max_positions: 5
  factor_weights:
    momentum: 0.35
    trend: 0.25
    reversal: 0.15
costs:
  commission_rate: 0.0003
  slippage_rate: 0.0005
""",
                encoding="utf-8",
            )
            child = variants / "child.yaml"
            child.write_text(
                """
extends: ../base.yaml
project:
  name: child_model
strategy:
  max_positions: 3
  factor_weights:
    momentum: 0.50
""",
                encoding="utf-8",
            )

            config = load_config(child)

        self.assertEqual(config.project_root, root.resolve())
        self.assertEqual(config.project.name, "child_model")
        self.assertEqual(config.project.data_dir, root.resolve() / "data")
        self.assertEqual(config.strategy.max_positions, 3)
        self.assertEqual(config.strategy.factor_weights["momentum"], 0.50)
        self.assertEqual(config.strategy.factor_weights["trend"], 0.25)
        self.assertEqual(config.strategy.factor_weights["reversal"], 0.15)

    def test_load_config_extends_detects_cycles(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            configs = root / "configs"
            configs.mkdir()
            a = configs / "a.yaml"
            b = configs / "b.yaml"
            a.write_text("extends: b.yaml\n", encoding="utf-8")
            b.write_text("extends: a.yaml\n", encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "Circular config extends"):
                load_config_mapping(a)

    def test_migrated_pos3_extends_configs_keep_variant_overrides(self) -> None:
        root = Path.cwd()

        base = load_config(root / "configs" / "ashare_main_chinext_thsdk_hold_prefer_warm120_dd_light_pos3.yaml")
        self.assertEqual(base.strategy.max_positions, 3)
        self.assertEqual(base.strategy.min_hold_days, 20)
        self.assertAlmostEqual(base.strategy.take_profit_pct or 0.0, 0.30)
        self.assertAlmostEqual(base.risk.protection_drawdown, 0.40)

        breadth = load_config(
            root / "configs" / "ashare_main_chinext_thsdk_hold_prefer_warm120_dd_light_pos3_breadth_soft30.yaml"
        )
        self.assertTrue(breadth.strategy.market_breadth_exposure_enabled)
        self.assertEqual(breadth.strategy.market_breadth_min_count, 500)
        self.assertAlmostEqual(breadth.strategy.market_breadth_weak_exposure, 0.50)
        self.assertAlmostEqual(breadth.risk.benchmark_off_exposure, 0.10)

        lossguard = load_config(
            root / "configs" / "ashare_main_chinext_thsdk_hold_prefer_warm120_dd_light_pos3_lossguard5d40.yaml"
        )
        self.assertAlmostEqual(lossguard.strategy.rebalance_loss_guard_pct or 0.0, 0.05)
        self.assertEqual(lossguard.strategy.rebalance_loss_guard_max_days, 40)
        self.assertTrue(lossguard.strategy.prefer_existing_positions)

        rsrs = load_config(
            root
            / "configs"
            / "ashare_main_chinext_thsdk_hold_prefer_warm120_dd_light_pos3_lossguard5d40_rsrs_z240.yaml"
        )
        self.assertTrue(rsrs.risk.benchmark_rsrs_enabled)
        self.assertEqual(rsrs.risk.benchmark_rsrs_zscore_window, 240)
        self.assertAlmostEqual(rsrs.risk.benchmark_rsrs_threshold, -0.5)
        self.assertEqual(rsrs.strategy.rebalance_loss_guard_max_days, 40)

        xscore = load_config(
            root / "configs" / "ashare_main_chinext_thsdk_hold_prefer_warm120_dd_light_pos3_xscore.yaml"
        )
        self.assertTrue(xscore.strategy.cross_sectional_score_enabled)
        self.assertAlmostEqual(xscore.strategy.cross_sectional_score_weights["signal"], 0.45)
        self.assertEqual(xscore.strategy.factor_min_history, 120)
        self.assertEqual(xscore.risk.drawdown_levels[-1], (0.40, 0.0))

    def test_migrated_light_extends_configs_keep_position_and_breadth_overrides(self) -> None:
        root = Path.cwd()

        light = load_config(root / "configs" / "ashare_main_chinext_thsdk_hold_prefer_warm120_dd_light.yaml")
        self.assertEqual(light.strategy.max_positions, 5)
        self.assertEqual(light.strategy.min_hold_days, 20)
        self.assertTrue(light.strategy.prefer_existing_positions)
        self.assertAlmostEqual(light.risk.benchmark_off_exposure, 0.10)
        self.assertEqual(light.risk.drawdown_levels[-1], (0.40, 0.0))

        breadth30 = load_config(
            root / "configs" / "ashare_main_chinext_thsdk_hold_prefer_warm120_dd_light_breadth30.yaml"
        )
        self.assertEqual(breadth30.strategy.max_positions, 5)
        self.assertTrue(breadth30.strategy.market_breadth_enabled)
        self.assertAlmostEqual(breadth30.strategy.market_breadth_min_ratio, 0.30)
        self.assertEqual(breadth30.strategy.market_breadth_min_count, 500)

        breadth40 = load_config(
            root / "configs" / "ashare_main_chinext_thsdk_hold_prefer_warm120_dd_light_breadth40.yaml"
        )
        self.assertEqual(breadth40.strategy.max_positions, 5)
        self.assertTrue(breadth40.strategy.market_breadth_enabled)
        self.assertAlmostEqual(breadth40.strategy.market_breadth_min_ratio, 0.40)
        self.assertEqual(breadth40.strategy.market_breadth_min_count, 500)

        pos3 = load_config(root / "configs" / "ashare_main_chinext_thsdk_hold_prefer_warm120_dd_light_pos3.yaml")
        self.assertEqual(pos3.strategy.max_positions, 3)
        self.assertFalse(pos3.strategy.market_breadth_enabled)

    def test_migrated_hold_prefer_extends_configs_keep_neutral_drawdown(self) -> None:
        root = Path.cwd()

        plain = load_config(root / "configs" / "ashare_main_chinext_thsdk_hold_prefer.yaml")
        self.assertEqual(plain.strategy.max_positions, 5)
        self.assertEqual(plain.strategy.signal_min_history, 0)
        self.assertEqual(plain.risk.drawdown_levels, ((0.99, 1.0),))
        self.assertAlmostEqual(plain.risk.protection_drawdown, 0.99)
        self.assertAlmostEqual(plain.risk.recovery_drawdown, 0.15)

        warm120 = load_config(root / "configs" / "ashare_main_chinext_thsdk_hold_prefer_warm120.yaml")
        self.assertEqual(warm120.strategy.max_positions, 5)
        self.assertEqual(warm120.strategy.signal_min_history, 120)
        self.assertEqual(warm120.risk.drawdown_levels, plain.risk.drawdown_levels)
        self.assertAlmostEqual(warm120.risk.protection_drawdown, 0.99)

    def test_migrated_thsdk_robust_extends_configs_keep_universe_overrides(self) -> None:
        root = Path.cwd()

        chinext = load_config(root / "configs" / "ashare_main_chinext_thsdk_robust.yaml")
        self.assertEqual(chinext.universe_source.type, "ashare_main_chinext")
        self.assertEqual(chinext.universe_source.symbol, "all")
        self.assertIsNone(chinext.universe_source.limit)
        self.assertFalse(chinext.strategy.prefer_existing_positions)
        self.assertEqual(chinext.risk.drawdown_levels, ((0.99, 1.0),))

        chinext_sample = load_config(root / "configs" / "ashare_main_chinext_thsdk_robust_sample120.yaml")
        self.assertEqual(chinext_sample.data.start_date, "20220101")
        self.assertEqual(chinext_sample.universe_source.type, "ashare_main_chinext")
        self.assertEqual(chinext_sample.universe_source.symbol, "sample120")
        self.assertEqual(chinext_sample.universe_source.limit, 120)

        stop_loss = load_config(root / "configs" / "ashare_main_chinext_thsdk_robust_sl8_tp30.yaml")
        self.assertAlmostEqual(stop_loss.strategy.stop_loss_pct or 0.0, 0.08)
        self.assertAlmostEqual(stop_loss.strategy.take_profit_pct or 0.0, 0.30)

        main_board = load_config(root / "configs" / "ashare_main_board_thsdk_robust.yaml")
        self.assertEqual(main_board.universe_source.type, "ashare_main_board")
        self.assertEqual(main_board.universe_source.symbol, "all")

        main_board_sample = load_config(root / "configs" / "ashare_main_board_thsdk_robust_sample80.yaml")
        self.assertEqual(main_board_sample.data.start_date, "20220101")
        self.assertEqual(main_board_sample.universe_source.type, "ashare_main_board")
        self.assertEqual(main_board_sample.universe_source.symbol, "sample80")
        self.assertEqual(main_board_sample.universe_source.limit, 80)

    def test_migrated_main_chinext_multifactor_extends_configs_keep_variant_weights(self) -> None:
        root = Path.cwd()

        stable = load_config(root / "configs" / "ashare_main_chinext_multifactor_stable.yaml")
        self.assertEqual(stable.strategy.name, "multi_factor")
        self.assertEqual(stable.strategy.max_positions, 10)
        self.assertEqual(stable.strategy.factor_rebalance_interval, 20)
        self.assertAlmostEqual(stable.strategy.factor_weights["momentum"], 0.20)
        self.assertAlmostEqual(stable.strategy.factor_weights["reversal"], 0.25)
        self.assertEqual(stable.risk.drawdown_levels[-1], (0.30, 0.0))
        self.assertAlmostEqual(stable.risk.benchmark_off_exposure, 0.50)

        stable_sample = load_config(root / "configs" / "ashare_main_chinext_multifactor_stable_sample120.yaml")
        self.assertEqual(stable_sample.universe_source.symbol, "sample120")
        self.assertEqual(stable_sample.universe_source.limit, 120)
        self.assertEqual(stable_sample.strategy.factor_weights, stable.strategy.factor_weights)

        stable_sl8_sample = load_config(root / "configs" / "ashare_main_chinext_multifactor_stable_sl8_sample120.yaml")
        self.assertEqual(stable_sl8_sample.universe_source.symbol, "sample120")
        self.assertEqual(stable_sl8_sample.universe_source.limit, 120)
        self.assertEqual(stable_sl8_sample.strategy.factor_weights, stable.strategy.factor_weights)
        self.assertAlmostEqual(stable_sl8_sample.strategy.stop_loss_pct or 0.0, 0.08)

        stable_volstop = load_config(root / "configs" / "ashare_main_chinext_multifactor_stable_volstop_w20_m25_sample120.yaml")
        self.assertEqual(stable_volstop.universe_source.symbol, "sample120")
        self.assertEqual(stable_volstop.universe_source.limit, 120)
        self.assertEqual(stable_volstop.strategy.factor_weights, stable.strategy.factor_weights)
        self.assertTrue(stable_volstop.strategy.volatility_stop_loss_enabled)
        self.assertEqual(stable_volstop.strategy.volatility_stop_window, 20)
        self.assertAlmostEqual(stable_volstop.strategy.volatility_stop_multiplier, 2.5)
        self.assertAlmostEqual(stable_volstop.strategy.volatility_stop_min_pct, 0.04)
        self.assertAlmostEqual(stable_volstop.strategy.volatility_stop_max_pct, 0.14)

        stable_timestop = load_config(root / "configs" / "ashare_main_chinext_multifactor_stable_timestop40_sample120.yaml")
        self.assertEqual(stable_timestop.universe_source.symbol, "sample120")
        self.assertEqual(stable_timestop.universe_source.limit, 120)
        self.assertEqual(stable_timestop.strategy.factor_weights, stable.strategy.factor_weights)
        self.assertTrue(stable_timestop.strategy.time_stop_enabled)
        self.assertEqual(stable_timestop.strategy.time_stop_min_hold_days, 40)
        self.assertAlmostEqual(stable_timestop.strategy.time_stop_return_threshold_pct, 0.0)

        stable_combo = load_config(root / "configs" / "ashare_main_chinext_multifactor_stable_volstop_timestop_sample120.yaml")
        self.assertEqual(stable_combo.universe_source.symbol, "sample120")
        self.assertEqual(stable_combo.universe_source.limit, 120)
        self.assertEqual(stable_combo.strategy.factor_weights, stable.strategy.factor_weights)
        self.assertTrue(stable_combo.strategy.volatility_stop_loss_enabled)
        self.assertTrue(stable_combo.strategy.time_stop_enabled)
        self.assertEqual(stable_combo.strategy.time_stop_min_hold_days, 40)
        self.assertAlmostEqual(stable_combo.strategy.time_stop_return_threshold_pct, 0.02)

        reversal = load_config(root / "configs" / "ashare_main_chinext_multifactor_reversal.yaml")
        self.assertAlmostEqual(reversal.strategy.factor_weights["momentum"], 0.05)
        self.assertAlmostEqual(reversal.strategy.factor_weights["reversal"], 0.55)
        self.assertAlmostEqual(reversal.risk.benchmark_off_exposure, 0.50)

        reversal_sample = load_config(root / "configs" / "ashare_main_chinext_multifactor_reversal_sample120.yaml")
        self.assertEqual(reversal_sample.universe_source.symbol, "sample120")
        self.assertEqual(reversal_sample.universe_source.limit, 120)
        self.assertEqual(reversal_sample.strategy.factor_weights, reversal.strategy.factor_weights)

        satellite = load_config(root / "configs" / "ashare_main_chinext_multifactor_satellite.yaml")
        self.assertEqual(satellite.strategy.factor_rebalance_interval, 10)
        self.assertAlmostEqual(satellite.strategy.factor_weights["momentum"], 0.45)
        self.assertAlmostEqual(satellite.strategy.factor_weights["trend"], 0.30)
        self.assertAlmostEqual(satellite.risk.benchmark_off_exposure, 0.40)
        self.assertEqual(satellite.risk.drawdown_levels, ((0.10, 0.75), (0.20, 0.45), (0.30, 0.0)))

        quality = load_config(root / "configs" / "ashare_main_chinext_multifactor_satellite_quality.yaml")
        self.assertTrue(quality.strategy.factor_entry_filter_enabled)
        self.assertTrue(quality.strategy.market_breadth_enabled)
        self.assertEqual(quality.strategy.market_breadth_min_count, 500)
        self.assertEqual(quality.strategy.factor_weights, satellite.strategy.factor_weights)

        quality_sample = load_config(
            root / "configs" / "ashare_main_chinext_multifactor_satellite_quality_sample120.yaml"
        )
        self.assertEqual(quality_sample.universe_source.symbol, "sample120")
        self.assertEqual(quality_sample.universe_source.limit, 120)
        self.assertEqual(quality_sample.strategy.market_breadth_min_count, 60)
        self.assertTrue(quality_sample.strategy.market_breadth_exposure_enabled)

    def test_migrated_custom_watchlist_thsdk_extends_configs_keep_risk_variants(self) -> None:
        root = Path.cwd()

        monitor = load_config(root / "configs" / "custom_watchlist_thsdk.yaml")
        self.assertEqual(monitor.universe_file, root / "configs" / "custom_stocks.csv")
        self.assertEqual(monitor.strategy.name, "thsdk_monitor")
        self.assertEqual(monitor.strategy.min_score, 60)
        self.assertEqual(monitor.strategy.min_hold_days, 0)
        self.assertFalse(monitor.risk.enabled)

        hold20 = load_config(root / "configs" / "custom_watchlist_thsdk_hold20.yaml")
        self.assertEqual(hold20.strategy.min_score, 60)
        self.assertEqual(hold20.strategy.min_hold_days, 20)
        self.assertFalse(hold20.risk.enabled)

        market40 = load_config(root / "configs" / "custom_watchlist_thsdk_hold20_market40.yaml")
        self.assertEqual(market40.strategy.min_hold_days, 20)
        self.assertAlmostEqual(market40.risk.benchmark_off_exposure, 0.40)
        self.assertEqual(market40.risk.drawdown_levels, ((0.99, 1.0),))
        self.assertAlmostEqual(market40.risk.protection_drawdown, 0.99)

        tp30 = load_config(root / "configs" / "custom_watchlist_thsdk_hold20_market40_tp30.yaml")
        self.assertAlmostEqual(tp30.strategy.take_profit_pct or 0.0, 0.30)
        self.assertEqual(tp30.strategy.min_score, 60)
        self.assertAlmostEqual(tp30.risk.benchmark_off_exposure, 0.40)

        score65 = load_config(root / "configs" / "custom_watchlist_thsdk_hold20_market40_tp30_score65.yaml")
        self.assertEqual(score65.strategy.min_score, 65)
        self.assertAlmostEqual(score65.strategy.take_profit_pct or 0.0, 0.30)

        score70 = load_config(root / "configs" / "custom_watchlist_thsdk_hold20_market40_tp30_score70_lowdd.yaml")
        self.assertEqual(score70.strategy.min_score, 70)
        self.assertAlmostEqual(score70.risk.benchmark_off_exposure, 0.40)

        stop_loss = load_config(root / "configs" / "custom_watchlist_thsdk_hold20_market40_sl8_tp30.yaml")
        self.assertAlmostEqual(stop_loss.strategy.stop_loss_pct or 0.0, 0.08)
        self.assertAlmostEqual(stop_loss.strategy.take_profit_pct or 0.0, 0.30)

        robust = load_config(root / "configs" / "custom_watchlist_thsdk_hold20_market10_tp30_score65_robust.yaml")
        self.assertEqual(robust.strategy.min_score, 65)
        self.assertAlmostEqual(robust.strategy.take_profit_pct or 0.0, 0.30)
        self.assertAlmostEqual(robust.risk.benchmark_off_exposure, 0.10)
        self.assertEqual(robust.risk.drawdown_levels, ((0.99, 1.0),))

        risk = load_config(root / "configs" / "custom_watchlist_thsdk_hold20_risk.yaml")
        self.assertAlmostEqual(risk.risk.benchmark_off_exposure, 0.40)
        self.assertEqual(risk.risk.drawdown_levels, ((0.10, 0.70), (0.20, 0.40), (0.30, 0.0)))
        self.assertAlmostEqual(risk.risk.protection_drawdown, 0.30)

    def test_migrated_custom_watchlist_multifactor_extends_configs_keep_variants(self) -> None:
        root = Path.cwd()

        plain = load_config(root / "configs" / "custom_watchlist_multifactor.yaml")
        self.assertEqual(plain.universe_file, root / "configs" / "custom_stocks.csv")
        self.assertEqual(plain.strategy.name, "multi_factor")
        self.assertEqual(plain.strategy.max_positions, 5)
        self.assertEqual(plain.strategy.factor_rebalance_interval, 5)
        self.assertAlmostEqual(plain.strategy.factor_weights["momentum"], 0.35)
        self.assertAlmostEqual(plain.strategy.factor_weights["trend"], 0.25)
        self.assertFalse(plain.risk.enabled)

        risk = load_config(root / "configs" / "custom_watchlist_multifactor_risk.yaml")
        risk_extends = load_config(root / "configs" / "custom_watchlist_multifactor_risk_extends.yaml")
        self.assertEqual(risk.strategy.factor_weights, plain.strategy.factor_weights)
        self.assertEqual(risk.strategy.factor_rebalance_interval, 5)
        self.assertTrue(risk.risk.enabled)
        self.assertAlmostEqual(risk.risk.benchmark_off_exposure, 0.50)
        self.assertEqual(risk.risk.drawdown_levels, ((0.10, 0.70), (0.20, 0.40), (0.30, 0.0)))
        self.assertEqual(risk_extends.strategy.factor_weights, risk.strategy.factor_weights)
        self.assertEqual(risk_extends.risk.drawdown_levels, risk.risk.drawdown_levels)

        tuned = load_config(root / "configs" / "custom_watchlist_multifactor_tuned.yaml")
        self.assertEqual(tuned.strategy.factor_rebalance_interval, 20)
        self.assertAlmostEqual(tuned.strategy.factor_weights["momentum"], 0.20)
        self.assertAlmostEqual(tuned.strategy.factor_weights["volatility"], 0.25)
        self.assertAlmostEqual(tuned.risk.benchmark_off_exposure, 0.80)
        self.assertAlmostEqual(tuned.risk.benchmark_crash_exposure, 0.20)
        self.assertEqual(tuned.risk.drawdown_levels, ((0.10, 0.80), (0.20, 0.60), (0.35, 0.0)))

        momentum_tuned = load_config(root / "configs" / "custom_watchlist_multifactor_momentum_tuned.yaml")
        self.assertTrue(momentum_tuned.strategy.momentum_focus_enabled)
        self.assertTrue(momentum_tuned.strategy.momentum_focus_only)
        self.assertAlmostEqual(momentum_tuned.strategy.momentum_focus_threshold_pct, 5.0)
        self.assertAlmostEqual(momentum_tuned.strategy.momentum_focus_limit_up_boost, 0.20)
        self.assertAlmostEqual(momentum_tuned.strategy.momentum_focus_strong_gain_boost, 0.05)
        self.assertEqual(momentum_tuned.strategy.score_allocation_method, "tilt")
        self.assertAlmostEqual(momentum_tuned.strategy.score_allocation_tilt_strength, 0.60)

        bear = load_config(root / "configs" / "custom_watchlist_multifactor_bear_v2.yaml")
        self.assertEqual(bear.strategy.factor_rebalance_interval, 20)
        self.assertAlmostEqual(bear.strategy.factor_weights["momentum"], 0.15)
        self.assertAlmostEqual(bear.strategy.factor_weights["volatility"], 0.35)
        self.assertAlmostEqual(bear.risk.benchmark_drop_threshold, -0.05)
        self.assertAlmostEqual(bear.risk.benchmark_off_exposure, 0.35)
        self.assertEqual(bear.risk.drawdown_levels, ((0.08, 0.55), (0.15, 0.30), (0.25, 0.0)))

        opportunity = load_config(root / "configs" / "custom_watchlist_multifactor_opportunity_v2.yaml")
        self.assertEqual(opportunity.strategy.factor_weights, tuned.strategy.factor_weights)
        self.assertAlmostEqual(opportunity.risk.benchmark_drop_threshold, -0.06)
        self.assertAlmostEqual(opportunity.risk.benchmark_off_exposure, 0.50)
        self.assertEqual(
            opportunity.risk.drawdown_levels,
            ((0.08, 0.65), (0.15, 0.45), (0.25, 0.15), (0.30, 0.0)),
        )

    def test_migrated_trend_extends_configs_keep_asset_specific_costs(self) -> None:
        root = Path.cwd()

        etf = load_config(root / "configs" / "etf_trend.yaml")
        self.assertEqual(etf.strategy.name, "trend")
        self.assertEqual(etf.strategy.max_positions, 3)
        self.assertEqual([item.code for item in etf.universe], ["510300", "510500", "159915", "588000", "512100", "518880", "511010"])
        self.assertAlmostEqual(etf.costs.commission_rate, 0.0003)
        self.assertAlmostEqual(etf.costs.slippage_rate, 0.0005)
        self.assertAlmostEqual(etf.costs.stamp_tax_rate, 0.0)

        csi300 = load_config(root / "configs" / "csi300_trend.yaml")
        self.assertEqual(csi300.universe_source.type, "csindex")
        self.assertEqual(csi300.universe_source.symbol, "000300")
        self.assertEqual(csi300.strategy.max_positions, 10)
        self.assertEqual(csi300.strategy.lot_size, 100)
        self.assertAlmostEqual(csi300.costs.stamp_tax_rate, 0.0005)

        watchlist = load_config(root / "configs" / "custom_watchlist_trend.yaml")
        self.assertEqual(watchlist.universe_file, root / "configs" / "custom_stocks.csv")
        self.assertEqual(watchlist.strategy.max_positions, 5)
        self.assertEqual(watchlist.strategy.lot_size, 100)
        self.assertAlmostEqual(watchlist.costs.stamp_tax_rate, 0.0005)

    def test_migrated_portfolio_extends_configs_keep_curve_and_filter_variants(self) -> None:
        root = Path.cwd()

        proxy = load_portfolio_config(root / "configs" / "portfolio_core_satellite_v1.yaml")
        self.assertEqual(proxy.name, "main_chinext_core_satellite_v1")
        self.assertEqual(proxy.core.path, root / "outputs" / "walk_forward" / "main_chinext_stable_v2_full_20260613_011322" / "oos_equity_stitched.csv")
        self.assertEqual(proxy.satellite.path, root / "outputs" / "backtests" / "ashare_main_chinext_hold_prefer_warm120_dd_light" / "equity_curve.csv")
        self.assertEqual(proxy.satellite.equity_column, "equity")
        self.assertAlmostEqual(proxy.weights["risk_on"]["satellite"], 0.25)
        self.assertFalse(proxy.satellite_filter.enabled)

        rolling = load_portfolio_config(root / "configs" / "portfolio_core_satellite_rolling_v1.yaml")
        self.assertEqual(rolling.satellite.path, root / "outputs" / "walk_forward" / "main_chinext_satellite_v1_full_20260613_074155" / "oos_equity_stitched.csv")
        self.assertEqual(rolling.satellite.equity_column, "stitched_equity")
        self.assertAlmostEqual(rolling.weights["risk_on"]["satellite"], 0.25)

        sat30 = load_portfolio_config(root / "configs" / "portfolio_core_satellite_rolling_sat30.yaml")
        self.assertAlmostEqual(sat30.weights["risk_on"]["core"], 0.70)
        self.assertAlmostEqual(sat30.weights["risk_on"]["satellite"], 0.30)
        self.assertAlmostEqual(sat30.weights["risk_off"]["satellite"], 0.0)
        self.assertFalse(sat30.satellite_filter.enabled)

        guarded = load_portfolio_config(root / "configs" / "portfolio_core_satellite_rolling_guarded.yaml")
        self.assertTrue(guarded.satellite_filter.enabled)
        self.assertEqual(guarded.satellite_filter.momentum_window, 10)
        self.assertAlmostEqual(guarded.satellite_filter.max_drawdown, 0.20)
        self.assertAlmostEqual(guarded.satellite_filter.reduced_drawdown, 0.12)
        self.assertAlmostEqual(guarded.satellite_filter.reduced_scale, 0.75)
        self.assertFalse(guarded.satellite_filter.require_above_ma)

        quality = load_portfolio_config(root / "configs" / "portfolio_core_satellite_quality_v2_guarded.yaml")
        self.assertEqual(quality.satellite.path, root / "outputs" / "walk_forward" / "main_chinext_satellite_quality_v2_full" / "oos_equity_stitched.csv")
        self.assertEqual(quality.satellite.equity_column, "stitched_equity")
        self.assertTrue(quality.satellite_filter.enabled)
        self.assertEqual(quality.satellite_filter, guarded.satellite_filter)

    def test_migrated_drawdown_guard_configs_keep_risk_profiles(self) -> None:
        root = Path.cwd()

        guard = load_config(root / "configs" / "ashare_main_chinext_thsdk_hold_prefer_dd_guard.yaml")
        self.assertEqual(guard.strategy.max_positions, 5)
        self.assertEqual(guard.strategy.signal_min_history, 0)
        self.assertEqual(guard.risk.drawdown_levels, ((0.10, 0.50), (0.20, 0.20), (0.30, 0.0)))
        self.assertAlmostEqual(guard.risk.protection_drawdown, 0.30)
        self.assertAlmostEqual(guard.risk.recovery_drawdown, 0.10)

        moderate = load_config(root / "configs" / "ashare_main_chinext_thsdk_hold_prefer_dd_moderate.yaml")
        self.assertEqual(moderate.strategy.max_positions, 5)
        self.assertEqual(moderate.strategy.signal_min_history, 0)
        self.assertEqual(moderate.risk.drawdown_levels, ((0.15, 0.70), (0.25, 0.35), (0.35, 0.0)))
        self.assertAlmostEqual(moderate.risk.protection_drawdown, 0.35)
        self.assertAlmostEqual(moderate.risk.recovery_drawdown, 0.12)

        warm_guard = load_config(root / "configs" / "ashare_main_chinext_thsdk_hold_prefer_warm120_dd_guard.yaml")
        self.assertEqual(warm_guard.strategy.max_positions, 5)
        self.assertEqual(warm_guard.strategy.signal_min_history, 120)
        self.assertEqual(warm_guard.risk.drawdown_levels, guard.risk.drawdown_levels)
        self.assertAlmostEqual(warm_guard.risk.recovery_drawdown, 0.10)

        light = load_config(root / "configs" / "ashare_main_chinext_thsdk_hold_prefer_warm120_dd_light.yaml")
        self.assertEqual(light.risk.drawdown_levels, ((0.20, 0.70), (0.30, 0.35), (0.40, 0.0)))
        self.assertEqual(light.strategy.signal_min_history, 120)

    def test_parse_config_with_risk_block(self) -> None:
        raw = {
            "project": {"data_dir": "data", "output_dir": "outputs/backtests"},
            "universe_file": "configs/custom_stocks.csv",
            "strategy": {"name": "multi_factor", "max_positions": 5},
            "risk": {
                "enabled": True,
                "benchmark_code": "510300",
                "benchmark_ma_window": 120,
                "benchmark_rsrs_enabled": True,
                "benchmark_rsrs_window": 18,
                "benchmark_rsrs_zscore_window": 120,
                "benchmark_rsrs_threshold": -0.5,
                "benchmark_rsrs_off_exposure": 0.25,
                "drawdown_levels": [
                    {"drawdown": 0.10, "exposure": 0.70},
                    {"drawdown": 0.30, "exposure": 0.0},
                ],
            },
        }
        config = parse_config(raw, Path("D:/codex/量化"))
        self.assertTrue(config.risk.enabled)
        self.assertEqual(config.risk.benchmark_code, "510300")
        self.assertTrue(config.risk.benchmark_rsrs_enabled)
        self.assertEqual(config.risk.benchmark_rsrs_off_exposure, 0.25)
        self.assertEqual(config.risk.drawdown_levels[-1], (0.30, 0.0))

    def test_thsdk_monitor_breakout_signal(self) -> None:
        dates = pd.date_range("2020-01-01", periods=7, freq="D")
        frame = pd.DataFrame(
            {
                "date": dates,
                "code": "TEST",
                "name": "Test ETF",
                "open": [10, 10, 10, 10, 10, 10, 11],
                "high": [10.2, 10.2, 10.2, 10.2, 10.2, 10.2, 12.5],
                "low": [9.8, 9.8, 9.8, 9.8, 9.8, 9.8, 10.8],
                "close": [10, 10, 10, 10, 10, 10, 12],
                "volume": [1000] * 7,
                "amount": [10000, 10000, 10000, 10000, 10000, 10000, 18000],
            }
        )
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            strategy = config.strategy.__class__(**{**config.strategy.__dict__, "name": "thsdk_monitor", "min_score": 60})
        signals = build_signals(frame, strategy)
        self.assertEqual(int(signals.iloc[-1]["breakout"]), 1)
        self.assertEqual(int(signals.iloc[-1]["signal"]), 1)

    def test_market_breadth_filter_blocks_new_thsdk_entries(self) -> None:
        strong = make_price_frame("AAA", "Strong", [10, 10, 10, 10, 10, 10, 12])
        strong["high"] = [10.2, 10.2, 10.2, 10.2, 10.2, 10.2, 12.5]
        strong["low"] = [9.8, 9.8, 9.8, 9.8, 9.8, 9.8, 10.8]
        strong["amount"] = [10000, 10000, 10000, 10000, 10000, 10000, 18000]
        weak = make_price_frame("BBB", "Weak", [10, 9, 8, 7, 6, 5, 4])
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            strategy = config.strategy.__class__(
                **{
                    **config.strategy.__dict__,
                    "name": "thsdk_monitor",
                    "min_score": 60,
                    "market_breadth_enabled": True,
                    "market_breadth_window": 2,
                    "market_breadth_min_ratio": 0.75,
                    "market_breadth_min_count": 2,
                }
            )
        frames = build_signal_frames({"AAA": strong, "BBB": weak}, strategy)
        strong_signals = frames["AAA"]
        self.assertAlmostEqual(float(strong_signals.iloc[-1]["market_breadth_ratio"]), 0.5)
        self.assertFalse(bool(strong_signals.iloc[-1]["market_breadth_ok"]))
        self.assertEqual(int(strong_signals.iloc[-1]["entry_candidate"]), 0)
        self.assertEqual(int(strong_signals.iloc[-1]["signal"]), 0)

    def test_market_breadth_filter_keeps_existing_positions_eligible(self) -> None:
        dates = pd.date_range("2020-01-01", periods=4, freq="D")
        aaa = pd.DataFrame(
            {
                "date": dates,
                "code": "AAA",
                "name": "Held",
                "open": [10, 10, 10, 10],
                "high": [10, 10, 10, 10],
                "low": [10, 10, 10, 10],
                "close": [10, 10, 10, 10],
                "volume": [1000] * 4,
                "amount": [10000] * 4,
                "trade_signal": [1, 1, 1, 1],
                "trade_score": [2.0, 2.0, 2.0, 2.0],
                "trade_entry_allowed": [1, 0, 0, 0],
            }
        )
        bbb = aaa.copy()
        bbb["code"] = "BBB"
        bbb["name"] = "Blocked"
        bbb["trade_score"] = [1.0, 1.0, 1.0, 1.0]
        bbb["trade_entry_allowed"] = [0, 0, 0, 0]
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            strategy = config.strategy.__class__(
                **{**config.strategy.__dict__, "max_positions": 2, "prefer_existing_positions": True}
            )
            config = config.__class__(
                project_root=config.project_root,
                project=config.project,
                data=config.data,
                universe=config.universe,
                strategy=strategy,
                costs=config.costs,
                universe_file=config.universe_file,
                universe_source=config.universe_source,
                risk=config.risk,
            )
            result = run_backtest(
                config,
                histories={"AAA": aaa, "BBB": bbb},
                signal_frames={"AAA": aaa, "BBB": bbb},
                run_id="test",
                write_outputs=False,
            )
        buys = result.trades[result.trades["side"] == "BUY"]
        self.assertEqual(set(buys["code"]), {"AAA"})

    def test_market_breadth_soft_exposure_reduces_position_size(self) -> None:
        dates = pd.date_range("2020-01-01", periods=3, freq="D")
        frame = pd.DataFrame(
            {
                "date": dates,
                "code": "TEST",
                "name": "Test ETF",
                "open": [10, 10, 10],
                "high": [10, 10, 10],
                "low": [10, 10, 10],
                "close": [10, 10, 10],
                "volume": [1000] * 3,
                "amount": [10000] * 3,
                "trade_signal": [1, 1, 1],
                "trade_score": [1.0, 1.0, 1.0],
                "trade_entry_allowed": [1, 1, 1],
                "market_breadth_ratio": [0.20, 0.20, 0.20],
                "market_breadth_count": [1000, 1000, 1000],
            }
        )
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            strategy = config.strategy.__class__(
                **{
                    **config.strategy.__dict__,
                    "market_breadth_exposure_enabled": True,
                    "market_breadth_min_count": 500,
                    "market_breadth_weak_ratio": 0.30,
                    "market_breadth_weak_exposure": 0.50,
                }
            )
            config = config.__class__(
                project_root=config.project_root,
                project=config.project,
                data=config.data,
                universe=config.universe,
                strategy=strategy,
                costs=config.costs,
                universe_file=config.universe_file,
                universe_source=config.universe_source,
                risk=config.risk,
            )
            result = run_backtest(
                config,
                histories={"TEST": frame},
                signal_frames={"TEST": frame},
                run_id="test",
                write_outputs=False,
            )
        buy = result.trades[result.trades["side"] == "BUY"].iloc[0]
        self.assertLess(float(buy["gross_amount"]), config.project.initial_cash * 0.55)
        self.assertAlmostEqual(float(result.equity.iloc[0]["market_breadth_exposure"]), 0.5)

    def test_cross_sectional_score_overlay_uses_next_day_score(self) -> None:
        config = StrategyConfig(
            name="thsdk_monitor",
            fast_window=2,
            slow_window=3,
            trend_window=4,
            rsi_window=3,
            rsi_oversold=35,
            rsi_recover=40,
            rsi_exit=65,
            max_positions=1,
            allocation_buffer=0.995,
            lot_size=100,
            min_score=60,
            cross_sectional_score_enabled=True,
            cross_sectional_score_weights={"signal": 0.0, "momentum": 1.0},
            factor_momentum_window=1,
            factor_reversal_window=1,
            factor_volatility_window=2,
            factor_liquidity_window=2,
            factor_trend_window=2,
            factor_min_history=2,
        )
        histories = {
            "AAA": make_price_frame("AAA", "Strong", [10, 10, 11, 12, 13]),
            "BBB": make_price_frame("BBB", "Flat", [10, 10, 10, 10, 10]),
        }
        frames = build_signal_frames(histories, config)

        strong = frames["AAA"]
        flat = frames["BBB"]
        self.assertGreater(float(strong.loc[2, "cross_sectional_score"]), float(flat.loc[2, "cross_sectional_score"]))
        self.assertGreater(float(strong.loc[3, "trade_score"]), float(flat.loc[3, "trade_score"]))
        self.assertAlmostEqual(float(strong.loc[3, "trade_score"]), float(strong.loc[2, "cross_sectional_score"]))

    def test_cross_sectional_score_overlay_can_rank_ma_convergence(self) -> None:
        config = StrategyConfig(
            name="thsdk_monitor",
            fast_window=2,
            slow_window=3,
            trend_window=4,
            rsi_window=3,
            rsi_oversold=35,
            rsi_recover=40,
            rsi_exit=65,
            max_positions=1,
            allocation_buffer=0.995,
            lot_size=100,
            min_score=60,
            cross_sectional_score_enabled=True,
            cross_sectional_score_weights={"signal": 0.0, "convergence": 1.0},
            factor_convergence_windows=(2, 3),
            factor_min_history=3,
        )
        histories = {
            "AAA": make_price_frame("AAA", "Tight", [10, 10, 10, 10, 10, 10]),
            "BBB": make_price_frame("BBB", "Loose", [10, 13, 8, 14, 7, 15]),
        }
        frames = build_signal_frames(histories, config)

        tight = frames["AAA"]
        loose = frames["BBB"]
        self.assertGreater(float(tight.loc[3, "cross_sectional_score"]), float(loose.loc[3, "cross_sectional_score"]))
        self.assertGreater(float(tight.loc[4, "trade_score"]), float(loose.loc[4, "trade_score"]))

    def test_cross_sectional_score_overlay_can_rank_volume_price_corr(self) -> None:
        config = StrategyConfig(
            name="thsdk_monitor",
            fast_window=2,
            slow_window=3,
            trend_window=4,
            rsi_window=3,
            rsi_oversold=35,
            rsi_recover=40,
            rsi_exit=65,
            max_positions=1,
            allocation_buffer=0.995,
            lot_size=100,
            min_score=60,
            cross_sectional_score_enabled=True,
            cross_sectional_score_weights={"signal": 0.0, "volume_price_corr": 1.0},
            factor_volume_price_window=3,
            factor_min_history=3,
        )
        histories = {
            "AAA": make_ohlcv_frame("AAA", "Diverging", [10, 11, 12, 13], [10.2, 11.2, 12.2, 13.2], [9.8, 10.8, 11.8, 12.8], [10, 11, 12, 13], [4000, 3000, 2000, 1000]),
            "BBB": make_ohlcv_frame("BBB", "Crowded", [10, 11, 12, 13], [10.2, 11.2, 12.2, 13.2], [9.8, 10.8, 11.8, 12.8], [10, 11, 12, 13], [1000, 2000, 3000, 4000]),
        }
        frames = build_signal_frames(histories, config)

        diverging = frames["AAA"]
        crowded = frames["BBB"]
        self.assertGreater(float(diverging.loc[2, "cross_sectional_score"]), float(crowded.loc[2, "cross_sectional_score"]))
        self.assertGreater(float(diverging.loc[3, "trade_score"]), float(crowded.loc[3, "trade_score"]))

    def test_cross_sectional_score_overlay_can_rank_shadow_support(self) -> None:
        config = StrategyConfig(
            name="thsdk_monitor",
            fast_window=2,
            slow_window=3,
            trend_window=4,
            rsi_window=3,
            rsi_oversold=35,
            rsi_recover=40,
            rsi_exit=65,
            max_positions=1,
            allocation_buffer=0.995,
            lot_size=100,
            min_score=60,
            cross_sectional_score_enabled=True,
            cross_sectional_score_weights={"signal": 0.0, "shadow_support": 1.0},
            factor_shadow_window=3,
            factor_min_history=3,
        )
        histories = {
            "AAA": make_ohlcv_frame("AAA", "Lower Shadow", [10, 10, 10, 10], [10.2, 10.2, 10.2, 10.2], [9.0, 9.0, 9.0, 9.0], [10, 10, 10, 10], [1000, 1000, 1000, 1000]),
            "BBB": make_ohlcv_frame("BBB", "Upper Shadow", [10, 10, 10, 10], [11.0, 11.0, 11.0, 11.0], [9.8, 9.8, 9.8, 9.8], [10, 10, 10, 10], [1000, 1000, 1000, 1000]),
        }
        frames = build_signal_frames(histories, config)

        lower_shadow = frames["AAA"]
        upper_shadow = frames["BBB"]
        self.assertGreater(float(lower_shadow.loc[2, "cross_sectional_score"]), float(upper_shadow.loc[2, "cross_sectional_score"]))
        self.assertGreater(float(lower_shadow.loc[3, "trade_score"]), float(upper_shadow.loc[3, "trade_score"]))

    def test_multifactor_signal_frames_rank_cross_sectionally(self) -> None:
        config = StrategyConfig(
            name="multi_factor",
            fast_window=2,
            slow_window=3,
            trend_window=4,
            rsi_window=3,
            rsi_oversold=35,
            rsi_recover=40,
            rsi_exit=65,
            max_positions=1,
            allocation_buffer=0.995,
            lot_size=100,
            factor_momentum_window=2,
            factor_reversal_window=1,
            factor_volatility_window=2,
            factor_liquidity_window=2,
            factor_trend_window=2,
            factor_min_history=3,
            factor_rebalance_interval=1,
            factor_weights={"momentum": 1.0},
        )
        histories = {
            "AAA": make_price_frame("AAA", "Strong", [10, 11, 12, 13, 14, 15]),
            "BBB": make_price_frame("BBB", "Weak", [10, 10, 10, 10, 10, 10]),
        }
        frames = build_signal_frames(histories, config)
        strong = frames["AAA"]
        weak = frames["BBB"]
        self.assertEqual(int(strong.loc[2, "signal"]), 1)
        self.assertEqual(int(weak.loc[2, "signal"]), 0)
        self.assertEqual(int(strong.loc[2, "trade_signal"]), 0)
        self.assertEqual(int(strong.loc[3, "trade_signal"]), 1)

    def test_multifactor_momentum_focus_boost_marks_candidate_rows(self) -> None:
        config = StrategyConfig(
            name="multi_factor",
            fast_window=2,
            slow_window=3,
            trend_window=4,
            rsi_window=3,
            rsi_oversold=35,
            rsi_recover=40,
            rsi_exit=65,
            max_positions=1,
            allocation_buffer=0.995,
            lot_size=100,
            factor_momentum_window=1,
            factor_reversal_window=1,
            factor_volatility_window=2,
            factor_liquidity_window=2,
            factor_trend_window=2,
            factor_min_history=1,
            factor_rebalance_interval=1,
            factor_weights={"momentum": 1.0},
            momentum_focus_enabled=True,
            momentum_focus_limit_up_boost=0.2,
            momentum_focus_strong_gain_boost=0.05,
            momentum_focus_threshold_pct=7.0,
        )
        histories = {
            "300001": make_price_frame("300001", "Strong", [10, 10, 10, 12, 12]),
            "600001": make_price_frame("600001", "Weak", [10, 10, 9.9, 9.8, 9.7]),
        }
        frames = build_signal_frames(histories, config)
        strong = frames["300001"]
        weak = frames["600001"]

        self.assertEqual(strong.loc[3, "momentum_focus_signal"], "strong_gain")
        self.assertAlmostEqual(float(strong.loc[3, "momentum_focus_boost"]), 0.05)
        self.assertFalse(bool(strong.loc[0, "momentum_focus_flag"]))
        self.assertTrue(bool(strong.loc[3, "momentum_focus_flag"]))
        self.assertFalse(bool(weak.loc[2, "momentum_focus_flag"]))
        self.assertAlmostEqual(float(weak.loc[2, "momentum_focus_boost"]), 0.0)
        self.assertEqual(int(strong.loc[3, "signal"]), 1)
        self.assertEqual(int(weak.loc[3, "signal"]), 0)

    def test_multifactor_chip_reversal_factor_is_default_off_and_can_rank_events(self) -> None:
        config = StrategyConfig(
            name="multi_factor",
            fast_window=2,
            slow_window=3,
            trend_window=4,
            rsi_window=3,
            rsi_oversold=35,
            rsi_recover=40,
            rsi_exit=65,
            max_positions=1,
            allocation_buffer=0.995,
            lot_size=100,
            factor_momentum_window=1,
            factor_reversal_window=1,
            factor_volatility_window=2,
            factor_liquidity_window=2,
            factor_trend_window=2,
            factor_min_history=3,
            factor_rebalance_interval=1,
            factor_weights={"chip_reversal": 1.0},
            factor_chip_reversal_drawdown_window=3,
            factor_chip_reversal_cost_window=3,
            factor_chip_reversal_min_drawdown_pct=10.0,
            factor_chip_reversal_min_score=0.05,
        )
        histories = {
            "600001": make_price_frame("600001", "Deep Pullback", [10, 12, 11, 8, 8.2]),
            "600002": make_price_frame("600002", "Flat", [10, 10, 10, 10, 10]),
        }
        frames = build_signal_frames(histories, config)
        deep = frames["600001"]
        flat = frames["600002"]

        self.assertGreater(float(deep.loc[3, "factor_chip_reversal"]), 0.05)
        self.assertTrue(pd.isna(flat.loc[3, "factor_chip_reversal"]))
        self.assertEqual(int(deep.loc[3, "signal"]), 1)
        self.assertEqual(int(flat.loc[3, "signal"]), 0)
        self.assertEqual(int(deep.loc[4, "trade_signal"]), 1)

    def test_parse_config_accepts_chip_reversal_factor_settings(self) -> None:
        raw = {
            "project": {"name": "chip-factor", "initial_cash": 100000},
            "data": {"start_date": "20200101", "period": "daily", "adjust": "qfq"},
            "universe": [{"code": "600001", "name": "A"}],
            "strategy": {
                "name": "multi_factor",
                "max_positions": 1,
                "factor_weights": {"chip_reversal": 1.0},
                "factor_chip_reversal_drawdown_window": 10,
                "factor_chip_reversal_cost_window": 15,
                "factor_chip_reversal_min_drawdown_pct": 8.5,
                "factor_chip_reversal_min_score": 0.04,
            },
        }
        config = parse_config(raw, Path("D:/tmp/project"))

        self.assertEqual(config.strategy.factor_weights["chip_reversal"], 1.0)
        self.assertEqual(config.strategy.factor_chip_reversal_drawdown_window, 10)
        self.assertEqual(config.strategy.factor_chip_reversal_cost_window, 15)
        self.assertAlmostEqual(config.strategy.factor_chip_reversal_min_drawdown_pct, 8.5)
        self.assertAlmostEqual(config.strategy.factor_chip_reversal_min_score, 0.04)

    def test_multifactor_momentum_focus_only_filters_out_non_focus_stocks(self) -> None:
        config = StrategyConfig(
            name="multi_factor",
            fast_window=2,
            slow_window=3,
            trend_window=4,
            rsi_window=3,
            rsi_oversold=35,
            rsi_recover=40,
            rsi_exit=65,
            max_positions=1,
            allocation_buffer=0.995,
            lot_size=100,
            factor_momentum_window=1,
            factor_reversal_window=1,
            factor_volatility_window=2,
            factor_liquidity_window=2,
            factor_trend_window=2,
            factor_min_history=1,
            factor_rebalance_interval=1,
            factor_weights={"momentum": 1.0},
            momentum_focus_enabled=True,
            momentum_focus_only=True,
            momentum_focus_threshold_pct=7.0,
            momentum_focus_limit_up_boost=0.0,
            momentum_focus_strong_gain_boost=0.0,
        )
        histories = {
            "300001": make_price_frame("300001", "Focus", [10, 10, 10, 10, 20]),
            "600001": make_price_frame("600001", "Distractor", [10, 10.5, 10.9, 11.2, 11.5]),
        }
        frames = build_signal_frames(histories, config)
        focus = frames["300001"]
        distractor = frames["600001"]

        self.assertTrue(bool(focus.loc[4, "momentum_focus_flag"]))
        self.assertFalse(bool(distractor.loc[4, "momentum_focus_flag"]))
        self.assertEqual(int(focus.loc[4, "signal"]), 1)
        self.assertTrue(pd.isna(distractor.loc[4, "factor_score"]))
        self.assertTrue(pd.isna(distractor.loc[4, "score"]))

    def test_multifactor_entry_filter_blocks_weak_trend_candidate(self) -> None:
        config = StrategyConfig(
            name="multi_factor",
            fast_window=2,
            slow_window=3,
            trend_window=4,
            rsi_window=3,
            rsi_oversold=35,
            rsi_recover=40,
            rsi_exit=65,
            max_positions=1,
            allocation_buffer=0.995,
            lot_size=100,
            factor_momentum_window=1,
            factor_reversal_window=1,
            factor_volatility_window=2,
            factor_liquidity_window=2,
            factor_trend_window=2,
            factor_min_history=3,
            factor_rebalance_interval=1,
            factor_weights={"momentum": 1.0},
            factor_entry_filter_enabled=True,
            factor_entry_trend_window=3,
            factor_entry_min_trend=0.02,
            factor_entry_momentum_window=1,
            factor_entry_min_momentum=-0.10,
        )
        histories = {
            "AAA": make_price_frame("AAA", "Weak Trend", [10, 9, 8, 7, 6, 6.5]),
            "BBB": make_price_frame("BBB", "Healthy Trend", [5, 5, 5, 5.2, 5.4, 5.6]),
        }
        frames = build_signal_frames(histories, config)

        weak_trend = frames["AAA"]
        healthy = frames["BBB"]
        self.assertFalse(bool(weak_trend.loc[5, "factor_entry_ok"]))
        self.assertEqual(int(weak_trend.loc[5, "signal"]), 0)
        self.assertEqual(int(healthy.loc[5, "signal"]), 1)

    def test_multifactor_market_breadth_blocks_next_day_new_entries(self) -> None:
        config = StrategyConfig(
            name="multi_factor",
            fast_window=2,
            slow_window=3,
            trend_window=4,
            rsi_window=3,
            rsi_oversold=35,
            rsi_recover=40,
            rsi_exit=65,
            max_positions=1,
            allocation_buffer=0.995,
            lot_size=100,
            factor_momentum_window=1,
            factor_reversal_window=1,
            factor_volatility_window=2,
            factor_liquidity_window=2,
            factor_trend_window=2,
            factor_min_history=2,
            factor_rebalance_interval=1,
            factor_weights={"momentum": 1.0},
            market_breadth_enabled=True,
            market_breadth_window=2,
            market_breadth_min_ratio=0.75,
            market_breadth_min_count=2,
        )
        histories = {
            "AAA": make_price_frame("AAA", "Strong", [10, 11, 12, 13]),
            "BBB": make_price_frame("BBB", "Weak", [10, 9, 8, 7]),
        }
        frames = build_signal_frames(histories, config)

        strong = frames["AAA"]
        self.assertAlmostEqual(float(strong.loc[2, "market_breadth_ratio"]), 0.5)
        self.assertFalse(bool(strong.loc[2, "market_breadth_ok"]))
        self.assertEqual(int(strong.loc[3, "trade_entry_allowed"]), 0)

    def test_multifactor_entry_filter_affects_next_trading_day(self) -> None:
        config = StrategyConfig(
            name="multi_factor",
            fast_window=2,
            slow_window=3,
            trend_window=4,
            rsi_window=3,
            rsi_oversold=35,
            rsi_recover=40,
            rsi_exit=65,
            max_positions=1,
            allocation_buffer=0.995,
            lot_size=100,
            factor_momentum_window=1,
            factor_reversal_window=1,
            factor_volatility_window=2,
            factor_liquidity_window=2,
            factor_trend_window=2,
            factor_min_history=3,
            factor_rebalance_interval=1,
            factor_weights={"momentum": 1.0},
            factor_entry_filter_enabled=True,
            factor_entry_trend_window=3,
            factor_entry_min_trend=0.0,
            factor_entry_momentum_window=1,
            factor_entry_min_momentum=-0.10,
        )
        frames = build_signal_frames({"AAA": make_price_frame("AAA", "Test", [10, 10, 11, 12, 8])}, config)
        frame = frames["AAA"]

        self.assertEqual(int(frame.loc[3, "signal"]), 1)
        self.assertFalse(bool(frame.loc[4, "factor_entry_ok"]))
        self.assertEqual(int(frame.loc[4, "signal"]), 0)
        self.assertEqual(int(frame.loc[4, "trade_signal"]), 1)

    def test_risk_benchmark_below_ma_reduces_exposure(self) -> None:
        risk = RiskConfig(enabled=True, benchmark_off_exposure=0.5)
        row = pd.Series(
            {
                "benchmark_exposure": 0.5,
                "benchmark_reason": "benchmark_below_ma",
                "benchmark_risk_on": False,
            }
        )
        decision = decide_next_risk(risk, equity=100.0, peak_equity=100.0, protected=False, benchmark_row=row)
        self.assertEqual(decision.exposure, 0.5)

    def test_rsrs_benchmark_filter_reduces_next_exposure(self) -> None:
        dates = pd.date_range("2020-01-01", periods=8, freq="D")
        lows = [9.0, 10.0, 11.0, 12.0, 9.0, 10.0, 11.0, 12.0]
        highs = [18.0, 20.0, 22.0, 24.0, 13.5, 15.0, 16.5, 18.0]
        benchmark = pd.DataFrame(
            {
                "date": dates,
                "code": "510300",
                "name": "CSI300 ETF",
                "open": [10.0] * len(dates),
                "high": highs,
                "low": lows,
                "close": [10.0] * len(dates),
                "volume": [1000] * len(dates),
                "amount": [10000] * len(dates),
            }
        )
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            risk = RiskConfig(
                enabled=True,
                benchmark_ma_window=2,
                benchmark_rsrs_enabled=True,
                benchmark_rsrs_window=4,
                benchmark_rsrs_zscore_window=3,
                benchmark_rsrs_threshold=-0.5,
                benchmark_rsrs_off_exposure=0.25,
            )
            config = LabConfig(
                project_root=config.project_root,
                project=config.project,
                data=config.data,
                universe=config.universe,
                strategy=config.strategy,
                costs=config.costs,
                risk=risk,
            )
            benchmark_path = Path(tmp) / "data" / "processed" / "510300.csv"
            benchmark_path.parent.mkdir(parents=True, exist_ok=True)
            benchmark.to_csv(benchmark_path, index=False)
            risk_frame = build_benchmark_risk_frame(config, list(dates))

        weak_row = risk_frame.iloc[-1]
        self.assertFalse(bool(weak_row["benchmark_rsrs_risk_on"]))
        self.assertEqual(float(weak_row["benchmark_exposure"]), 0.25)
        self.assertIn("benchmark_rsrs_off", str(weak_row["benchmark_reason"]))

    def test_risk_drawdown_protection_goes_flat(self) -> None:
        risk = RiskConfig(enabled=True, protection_drawdown=0.30)
        row = pd.Series({"benchmark_exposure": 1.0, "benchmark_reason": "benchmark_on", "benchmark_risk_on": True})
        decision = decide_next_risk(risk, equity=70.0, peak_equity=100.0, protected=False, benchmark_row=row)
        self.assertEqual(decision.exposure, 0.0)
        self.assertTrue(decision.protected)

    def test_risk_protection_requires_recovery(self) -> None:
        risk = RiskConfig(enabled=True, recovery_drawdown=0.15)
        row = pd.Series({"benchmark_exposure": 1.0, "benchmark_reason": "benchmark_on", "benchmark_risk_on": True})
        decision = decide_next_risk(risk, equity=80.0, peak_equity=100.0, protected=True, benchmark_row=row)
        self.assertEqual(decision.exposure, 0.0)
        self.assertTrue(decision.protected)

    def test_risk_protection_recovers_at_threshold(self) -> None:
        risk = RiskConfig(enabled=True, recovery_drawdown=0.15)
        row = pd.Series({"benchmark_exposure": 1.0, "benchmark_reason": "benchmark_on", "benchmark_risk_on": True})
        decision = decide_next_risk(risk, equity=85.0, peak_equity=100.0, protected=True, benchmark_row=row)
        self.assertFalse(decision.protected)
        self.assertEqual(decision.exposure, 0.7)

    def test_risk_signal_affects_next_trading_day(self) -> None:
        dates = pd.date_range("2020-01-01", periods=5, freq="D")
        frame = pd.DataFrame(
            {
                "date": dates,
                "code": "TEST",
                "name": "Test ETF",
                "open": [10, 10, 10, 10, 10],
                "high": [10.5] * 5,
                "low": [9.5] * 5,
                "close": [10, 11, 12, 13, 14],
                "volume": [1000] * 5,
                "amount": [10000] * 5,
            }
        )
        benchmark = pd.DataFrame(
            {
                "date": dates,
                "code": "510300",
                "name": "CSI300 ETF",
                "open": [10, 9, 8, 7, 6],
                "high": [10, 9, 8, 7, 6],
                "low": [10, 9, 8, 7, 6],
                "close": [10, 9, 8, 7, 6],
                "volume": [1000] * 5,
                "amount": [10000] * 5,
            }
        )
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            risk = RiskConfig(
                enabled=True,
                benchmark_ma_window=2,
                benchmark_drop_window=1,
                benchmark_drop_threshold=-0.50,
                benchmark_off_exposure=0.5,
            )
            config = LabConfig(
                project_root=config.project_root,
                project=config.project,
                data=config.data,
                universe=config.universe,
                strategy=config.strategy,
                costs=config.costs,
                risk=risk,
            )
            benchmark_path = Path(tmp) / "data" / "processed" / "510300.csv"
            benchmark_path.parent.mkdir(parents=True, exist_ok=True)
            benchmark.to_csv(benchmark_path, index=False)
            result = run_backtest(config, histories={"TEST": frame}, run_id="risk-test", write_outputs=False)
        risk_curve = result.risk_curve
        self.assertEqual(float(risk_curve.loc[1, "risk_exposure"]), 1.0)
        self.assertEqual(float(risk_curve.loc[1, "next_risk_exposure"]), 0.5)
        self.assertEqual(float(risk_curve.loc[2, "risk_exposure"]), 0.5)

    def test_generate_rolling_windows(self) -> None:
        windows = generate_rolling_windows("20180101", "20240630", train_years=3, test_months=12, step_months=12)
        self.assertEqual(windows[0].train_start, "20180101")
        self.assertEqual(windows[0].train_end, "20201231")
        self.assertEqual(windows[0].test_start, "20210101")
        self.assertEqual(windows[0].test_end, "20211231")
        self.assertEqual(windows[-1].test_end, "20240630")

    def test_walk_forward_validation_dates_hold_out_recent_training_months(self) -> None:
        window = generate_rolling_windows("20180101", "20220630", train_years=4, test_months=6, step_months=6)[0]
        validation = _walk_forward_validation_dates(window, validation_months=6)
        self.assertIsNotNone(validation)
        assert validation is not None
        self.assertEqual(validation.train_start, "20180101")
        self.assertEqual(validation.train_end, "20210630")
        self.assertEqual(validation.validation_start, "20210701")
        self.assertEqual(validation.validation_end, "20211231")

    def test_walk_forward_validation_dates_reject_full_training_window_holdout(self) -> None:
        window = generate_rolling_windows("20180101", "20220630", train_years=4, test_months=6, step_months=6)[0]
        with self.assertRaisesRegex(ValueError, "selection_validation_months"):
            _walk_forward_validation_dates(window, validation_months=48)

    def test_walk_forward_signal_warmup_uses_prior_history(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            strategy = config.strategy.__class__(
                **{
                    **config.strategy.__dict__,
                    "factor_min_history": 120,
                    "factor_trend_window": 120,
                }
            )
            config = config.__class__(
                project_root=config.project_root,
                project=config.project,
                data=config.data,
                universe=config.universe,
                strategy=strategy,
                costs=config.costs,
                universe_file=config.universe_file,
                universe_source=config.universe_source,
                risk=config.risk,
            )
        self.assertEqual(_signal_warmup_start(config, "20220101"), "20201107")

    def test_trim_signal_frames_keeps_precomputed_first_trade_signal(self) -> None:
        dates = pd.date_range("2021-12-30", periods=4, freq="D")
        frame = pd.DataFrame(
            {
                "date": dates,
                "open": [10, 10, 11, 12],
                "close": [10, 11, 12, 13],
                "trade_signal": [0, 1, 1, 0],
                "trade_score": [0.0, 0.8, 0.7, 0.0],
            }
        )
        trimmed = _trim_signal_frames({"TEST": frame}, "20220101", "20220102")["TEST"]
        self.assertEqual(list(trimmed["date"].dt.strftime("%Y-%m-%d")), ["2022-01-01", "2022-01-02"])
        self.assertEqual(int(trimmed.iloc[0]["trade_signal"]), 1)

    def test_bear_grid_includes_strict_risk_candidates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            candidates = build_parameter_grid(make_config(Path(tmp)), grid="bear")
        self.assertEqual(len(candidates), 9)
        self.assertTrue(any("risk_bear_strict" in candidate.name for candidate in candidates))
        strict = next(candidate for candidate in candidates if "risk_bear_strict" in candidate.name)
        self.assertEqual(strict.risk_overrides["protection_drawdown"], 0.25)

    def test_walk_forward_bear_v2_preset(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--preset", "bear-v2"])
        options = resolve_walk_forward_options(args)
        self.assertEqual(options["grid"], "bear")
        self.assertEqual(options["objective"], "robust")
        self.assertEqual(options["train_years"], 4)
        self.assertEqual(options["max_train_drawdown"], 0.30)

    def test_walk_forward_opportunity_v2_preset(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--preset", "opportunity-v2"])
        options = resolve_walk_forward_options(args)
        self.assertEqual(options["grid"], "opportunity")
        self.assertEqual(options["objective"], "robust")
        self.assertEqual(options["train_years"], 4)
        self.assertEqual(options["max_train_drawdown"], 0.25)

    def test_walk_forward_opportunity_v2_growthfull_preset(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--preset", "opportunity-v2-growthfull"])
        options = resolve_walk_forward_options(args)
        self.assertEqual(options["grid"], "stable_v2")
        self.assertEqual(options["objective"], "opportunity_growth")
        self.assertEqual(options["train_years"], 4)
        self.assertEqual(options["max_train_drawdown"], 0.28)
        self.assertEqual(options["test_months"], 12)
        self.assertEqual(options["step_months"], 6)

    def test_walk_forward_highgain_dd40_preset(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--preset", "highgain-dd40"])
        options = resolve_walk_forward_options(args)
        self.assertEqual(options["grid"], "opportunity_v2")
        self.assertEqual(options["objective"], "capital")
        self.assertEqual(options["train_years"], 4)
        self.assertEqual(options["test_months"], 6)
        self.assertEqual(options["step_months"], 6)
        self.assertEqual(options["max_train_drawdown"], 0.40)
        self.assertEqual(options["selection_validation_months"], 6)

    def test_walk_forward_opportunity_v2_q_full_preset(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--preset", "opportunity-v2-q-full"])
        options = resolve_walk_forward_options(args)
        self.assertEqual(options["grid"], "stable_v2")
        self.assertEqual(options["objective"], "opportunity_q_full")
        self.assertEqual(options["train_years"], 4)
        self.assertEqual(options["max_train_drawdown"], 0.25)
        self.assertEqual(options["test_months"], 12)
        self.assertEqual(options["step_months"], 6)
        self.assertEqual(options["selection_validation_months"], 6)

    def test_walk_forward_opportunity_v2_q_full_aggressive_preset(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--preset", "opportunity-v2-q-full-aggressive"])
        options = resolve_walk_forward_options(args)
        self.assertEqual(options["grid"], "stable_v2")
        self.assertEqual(options["objective"], "opportunity_q_full")
        self.assertEqual(options["train_years"], 4)
        self.assertEqual(options["max_train_drawdown"], 0.28)
        self.assertEqual(options["test_months"], 12)
        self.assertEqual(options["step_months"], 6)
        self.assertEqual(options["selection_validation_months"], 0)

    def test_walk_forward_main_chinext_stable_preset(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--preset", "main-chinext-stable"])
        options = resolve_walk_forward_options(args)
        self.assertEqual(options["grid"], "stable")
        self.assertEqual(options["objective"], "stable")
        self.assertEqual(options["train_years"], 4)
        self.assertEqual(options["test_months"], 6)
        self.assertEqual(options["step_months"], 6)
        self.assertEqual(options["max_train_drawdown"], 0.25)

    def test_walk_forward_main_chinext_stable_v2_preset(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--preset", "main-chinext-stable-v2"])
        options = resolve_walk_forward_options(args)
        self.assertEqual(options["grid"], "stable_v2")
        self.assertEqual(options["objective"], "stable")
        self.assertEqual(options["train_years"], 4)
        self.assertEqual(options["test_months"], 6)
        self.assertEqual(options["step_months"], 6)
        self.assertEqual(options["max_train_drawdown"], 0.20)

    def test_walk_forward_main_chinext_reversal_preset(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--preset", "main-chinext-reversal-v1"])
        options = resolve_walk_forward_options(args)
        self.assertEqual(options["grid"], "reversal")
        self.assertEqual(options["objective"], "stable")
        self.assertEqual(options["train_years"], 4)
        self.assertEqual(options["test_months"], 6)
        self.assertEqual(options["step_months"], 6)
        self.assertEqual(options["max_train_drawdown"], 0.25)

    def test_walk_forward_main_chinext_satellite_preset(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--preset", "main-chinext-satellite-v1"])
        options = resolve_walk_forward_options(args)
        self.assertEqual(options["grid"], "satellite")
        self.assertEqual(options["objective"], "satellite")
        self.assertEqual(options["train_years"], 4)
        self.assertEqual(options["test_months"], 6)
        self.assertEqual(options["step_months"], 6)
        self.assertEqual(options["max_train_drawdown"], 0.35)

    def test_walk_forward_main_chinext_satellite_v2_preset(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--preset", "main-chinext-satellite-v2"])
        options = resolve_walk_forward_options(args)
        self.assertEqual(options["grid"], "satellite_v2")
        self.assertEqual(options["objective"], "satellite")
        self.assertEqual(options["train_years"], 4)
        self.assertEqual(options["test_months"], 6)
        self.assertEqual(options["step_months"], 6)
        self.assertEqual(options["max_train_drawdown"], 0.35)

    def test_walk_forward_preset_can_be_overridden(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--preset", "bear-v2", "--train-years", "3"])
        options = resolve_walk_forward_options(args)
        self.assertEqual(options["grid"], "bear")
        self.assertEqual(options["train_years"], 3)

    def test_walk_forward_resume_flag(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--resume"])
        self.assertTrue(args.resume)

    def test_walk_forward_window_limit_flag(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--window-limit", "1"])
        self.assertEqual(args.window_limit, 1)

    def test_walk_forward_selection_validation_months_flag(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--selection-validation-months", "6"])
        options = resolve_walk_forward_options(args)
        self.assertEqual(args.selection_validation_months, 6)
        self.assertEqual(options["selection_validation_months"], 6)

    def test_walk_forward_opportunity_quality_objective_flag(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--objective", "opportunity_quality"])
        options = resolve_walk_forward_options(args)
        self.assertEqual(args.objective, "opportunity_quality")
        self.assertEqual(options["objective"], "opportunity_quality")

    def test_walk_forward_opportunity_growth_objective_flag(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--objective", "opportunity_growth"])
        options = resolve_walk_forward_options(args)
        self.assertEqual(args.objective, "opportunity_growth")
        self.assertEqual(options["objective"], "opportunity_growth")

    def test_walk_forward_opportunity_q_full_objective_flag(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--objective", "opportunity_q_full"])
        options = resolve_walk_forward_options(args)
        self.assertEqual(args.objective, "opportunity_q_full")
        self.assertEqual(options["objective"], "opportunity_q_full")

    def test_walk_forward_selection_validation_chooses_validation_winner(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            config = make_config(root)
            config = config.__class__(
                project_root=config.project_root,
                project=config.project,
                data=config.data.__class__(start_date="20180101", end_date="20220630", period="daily", adjust="qfq"),
                universe=config.universe,
                strategy=config.strategy,
                costs=config.costs,
                universe_file=config.universe_file,
                universe_source=config.universe_source,
                risk=config.risk,
            )
            dates = pd.date_range("2018-01-01", "2022-06-30", freq="D")
            histories = {
                "TEST": pd.DataFrame(
                    {
                        "date": dates,
                        "open": np.full(len(dates), 10.0),
                        "high": np.full(len(dates), 10.5),
                        "low": np.full(len(dates), 9.5),
                        "close": np.full(len(dates), 10.0),
                        "volume": np.full(len(dates), 1000.0),
                    }
                )
            }
            candidates = (
                ParameterCandidate("train_star", {"momentum": 1.0}, 10, 1, {}),
                ParameterCandidate("validation_star", {"momentum": 1.0}, 20, 1, {}),
            )

            def fake_result(run_id: str, run_dir: Path, start_date: str, candidate_rebalance: int) -> BacktestResult:
                is_fit = start_date == "20180101"
                is_validation = start_date == "20210701"
                if is_fit and candidate_rebalance == 10:
                    total_return, cagr, sharpe = 0.60, 0.16, 1.2
                elif is_validation and candidate_rebalance == 20:
                    total_return, cagr, sharpe = 0.20, 0.18, 1.1
                else:
                    total_return, cagr, sharpe = -0.10, -0.08, -0.4
                equity = pd.DataFrame(
                    {
                        "date": pd.to_datetime([start_date, pd.Timestamp(start_date) + pd.Timedelta(days=1)]),
                        "equity": [100000.0, 100000.0 * (1.0 + total_return)],
                    }
                )
                metrics = {
                    "final_equity": float(equity["equity"].iloc[-1]),
                    "total_return": total_return,
                    "benchmark_return": 0.0,
                    "excess_return": total_return,
                    "cagr": cagr,
                    "max_drawdown": -0.05,
                    "sharpe": sharpe,
                    "trade_count": 10,
                    "win_rate": 0.6,
                    "payoff_ratio": 1.2,
                    "profit_factor": 1.3,
                    "average_position_exposure": 0.5,
                    "average_risk_exposure": 0.5,
                }
                return BacktestResult(
                    run_id=run_id,
                    run_dir=run_dir,
                    equity=equity,
                    trades=pd.DataFrame(),
                    benchmark=pd.DataFrame(),
                    monthly_returns=pd.DataFrame(),
                    metrics=metrics,
                    risk_curve=pd.DataFrame(),
                    risk_events=pd.DataFrame(),
                    cooldown_events=pd.DataFrame(),
                )

            def fake_run_backtest(backtest_config, histories, signal_frames, run_id, write_outputs):
                return fake_result(
                    run_id,
                    root / "outputs" / "backtests" / run_id,
                    backtest_config.data.start_date,
                    int(backtest_config.strategy.factor_rebalance_interval),
                )

            with (
                patch("quant_etf_lab.walk_forward.load_universe_history", return_value=histories),
                patch("quant_etf_lab.walk_forward.build_parameter_grid", return_value=candidates),
                patch("quant_etf_lab.walk_forward.build_signal_frames", return_value={}),
                patch("quant_etf_lab.walk_forward.run_backtest", side_effect=fake_run_backtest),
                patch("quant_etf_lab.walk_forward.write_report", return_value=root / "report.html"),
            ):
                _, summary = run_walk_forward(
                    config,
                    train_years=4,
                    test_months=6,
                    step_months=6,
                    selection_validation_months=6,
                    output_dir=root / "wf",
                    run_id_prefix="selection_validation",
                    window_limit=1,
                )

            self.assertEqual(summary.iloc[0]["selected_candidate"], "validation_star")
            self.assertEqual(int(summary.iloc[0]["selection_validation_months"]), 6)
            self.assertEqual(str(summary.iloc[0]["validation_start"]), "20210701")
            self.assertIn("selected_validation_score", summary.columns)

    def test_walk_forward_opportunity_v2_gate_keeps_current_when_new_candidate_validation_loses(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            config = make_config(root)
            config = config.__class__(
                project_root=config.project_root,
                project=config.project,
                data=config.data.__class__(start_date="20180101", end_date="20220630", period="daily", adjust="qfq"),
                universe=config.universe,
                strategy=config.strategy,
                costs=config.costs,
                universe_file=config.universe_file,
                universe_source=config.universe_source,
                risk=config.risk,
            )
            dates = pd.date_range("2018-01-01", "2022-06-30", freq="D")
            histories = {
                "TEST": pd.DataFrame(
                    {
                        "date": dates,
                        "open": np.full(len(dates), 10.0),
                        "high": np.full(len(dates), 10.5),
                        "low": np.full(len(dates), 9.5),
                        "close": np.full(len(dates), 10.0),
                        "volume": np.full(len(dates), 1000.0),
                    }
                )
            }
            candidates = (
                ParameterCandidate("current_reb10_risk_loose", {"momentum": 1.0}, 10, 10, {}),
                ParameterCandidate("strong_gain_quality_reb15_risk_growth_dd40", {"momentum": 0.8, "trend": 0.2}, 15, 12, {}),
            )

            def fake_result(run_id: str, run_dir: Path, start_date: str, rebalance: int) -> BacktestResult:
                is_validation = start_date == "20210701"
                if is_validation and rebalance == 15:
                    total_return, cagr, sharpe = -0.01, -0.02, 0.8
                elif is_validation:
                    total_return, cagr, sharpe = -0.15, -0.20, -0.8
                elif rebalance == 15:
                    total_return, cagr, sharpe = 0.40, 0.12, 0.9
                else:
                    total_return, cagr, sharpe = -0.05, -0.02, -0.2
                equity = pd.DataFrame(
                    {
                        "date": pd.to_datetime([start_date, pd.Timestamp(start_date) + pd.Timedelta(days=1)]),
                        "equity": [100000.0, 100000.0 * (1.0 + total_return)],
                    }
                )
                metrics = {
                    "final_equity": float(equity["equity"].iloc[-1]),
                    "total_return": total_return,
                    "benchmark_return": 0.0,
                    "excess_return": total_return,
                    "cagr": cagr,
                    "max_drawdown": -0.05,
                    "sharpe": sharpe,
                    "trade_count": 10,
                    "win_rate": 0.6,
                    "payoff_ratio": 1.2,
                    "profit_factor": 1.3,
                    "average_position_exposure": 0.5,
                    "average_risk_exposure": 0.5,
                }
                return BacktestResult(
                    run_id=run_id,
                    run_dir=run_dir,
                    equity=equity,
                    trades=pd.DataFrame(),
                    benchmark=pd.DataFrame(),
                    monthly_returns=pd.DataFrame(),
                    metrics=metrics,
                    risk_curve=pd.DataFrame(),
                    risk_events=pd.DataFrame(),
                    cooldown_events=pd.DataFrame(),
                )

            def fake_run_backtest(backtest_config, histories, signal_frames, run_id, write_outputs):
                return fake_result(
                    run_id,
                    root / "outputs" / "backtests" / run_id,
                    backtest_config.data.start_date,
                    int(backtest_config.strategy.factor_rebalance_interval),
                )

            with (
                patch("quant_etf_lab.walk_forward.load_universe_history", return_value=histories),
                patch("quant_etf_lab.walk_forward.build_parameter_grid", return_value=candidates),
                patch("quant_etf_lab.walk_forward.build_signal_frames", return_value={}),
                patch("quant_etf_lab.walk_forward.run_backtest", side_effect=fake_run_backtest),
                patch("quant_etf_lab.walk_forward.write_report", return_value=root / "report.html"),
            ):
                _, summary = run_walk_forward(
                    config,
                    train_years=4,
                    test_months=6,
                    step_months=6,
                    grid="opportunity_v2",
                    objective="capital",
                    selection_validation_months=6,
                    output_dir=root / "wf",
                    run_id_prefix="selection_gate",
                    window_limit=1,
                )

            self.assertEqual(summary.iloc[0]["selected_candidate"], "current_reb10_risk_loose")

    def test_data_update_continue_on_error_flag(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["data", "update", "--continue-on-error"])
        self.assertTrue(args.continue_on_error)

    def test_data_update_market_cap_cli_options(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "data",
                "update-market-cap",
                "--output",
                "data/processed/stock_market_cap_yi.csv",
                "--retry-count",
                "2",
                "--pause-seconds",
                "0.2",
                "--symbols-file",
                "outputs/research/paper/stock_targets.csv",
                "--as-of-date",
                "2024-04-02",
                "--lookback-days",
                "5",
                "--daily-data-dir",
                "D:/codex/daily-market-data",
                "--ingest-project-dir",
                "D:/codex/2026-06-15-exchange-data-ingest",
            ]
        )
        self.assertEqual(args.data_command, "update-market-cap")
        self.assertEqual(args.output, "data/processed/stock_market_cap_yi.csv")
        self.assertEqual(args.retry_count, 2)
        self.assertAlmostEqual(args.pause_seconds, 0.2)
        self.assertEqual(args.symbols_file, "outputs/research/paper/stock_targets.csv")
        self.assertEqual(args.as_of_date, "2024-04-02")
        self.assertEqual(args.lookback_days, 5)
        self.assertEqual(args.daily_data_dir, "D:/codex/daily-market-data")
        self.assertEqual(args.ingest_project_dir, "D:/codex/2026-06-15-exchange-data-ingest")
        self.assertFalse(args.no_local_daily)

    def test_backtest_skip_missing_flag(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["backtest", "--skip-missing"])
        self.assertTrue(args.skip_missing)

    def test_walk_forward_skip_missing_flag(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["walk-forward", "--skip-missing"])
        self.assertTrue(args.skip_missing)

    def test_sentiment_cli_reference_options(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "sentiment",
                "--config",
                "configs/ashare.yaml",
                "--output-dir",
                "outputs/research/sentiment",
                "--max-symbols",
                "80",
                "--window",
                "60",
                "--strict",
            ]
        )
        self.assertEqual(args.command, "sentiment")
        self.assertEqual(args.config, "configs/ashare.yaml")
        self.assertEqual(args.output_dir, "outputs/research/sentiment")
        self.assertEqual(args.max_symbols, 80)
        self.assertEqual(args.window, 60)
        self.assertTrue(args.strict)

    def test_factor_lab_cli_options(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "factor-lab",
                "--data-dir",
                "data/processed/stocks",
                "--output-dir",
                "outputs/research/factors",
                "--factors",
                "momentum_20d,reversal_5d",
                "--horizons",
                "1,5",
                "--quantiles",
                "3",
                "--min-obs",
                "20",
                "--max-symbols",
                "100",
                "--start-date",
                "2024-01-01",
                "--end-date",
                "2024-12-31",
                "--warmup-days",
                "120",
                "--no-save-panel",
                "--recursive",
            ]
        )
        self.assertEqual(args.command, "factor-lab")
        self.assertEqual(args.data_dir, "data/processed/stocks")
        self.assertEqual(args.output_dir, "outputs/research/factors")
        self.assertEqual(args.factors, "momentum_20d,reversal_5d")
        self.assertEqual(args.horizons, "1,5")
        self.assertEqual(args.quantiles, 3)
        self.assertEqual(args.min_obs, 20)
        self.assertEqual(args.max_symbols, 100)
        self.assertEqual(args.start_date, "2024-01-01")
        self.assertEqual(args.end_date, "2024-12-31")
        self.assertEqual(args.warmup_days, 120)
        self.assertFalse(args.save_panel)
        self.assertTrue(args.recursive)

    def test_network_lab_cli_options(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "network-lab",
                "--data-dir",
                "data/processed/stocks",
                "--output-dir",
                "outputs/research/network",
                "--symbols",
                "000001,000002,000003",
                "--start-date",
                "2024-01-01",
                "--end-date",
                "2024-12-31",
                "--lookback-days",
                "120",
                "--max-symbols",
                "30",
                "--top-edges",
                "5",
                "--bins",
                "4",
                "--min-obs",
                "5",
                "--recursive",
            ]
        )
        self.assertEqual(args.command, "network-lab")
        self.assertEqual(args.data_dir, "data/processed/stocks")
        self.assertEqual(args.output_dir, "outputs/research/network")
        self.assertEqual(args.symbols, "000001,000002,000003")
        self.assertEqual(args.start_date, "2024-01-01")
        self.assertEqual(args.end_date, "2024-12-31")
        self.assertEqual(args.lookback_days, 120)
        self.assertEqual(args.max_symbols, 30)
        self.assertEqual(args.top_edges, 5)
        self.assertEqual(args.bins, 4)
        self.assertEqual(args.min_obs, 5)
        self.assertTrue(args.recursive)

    def test_phase2_cli_options(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "phase2",
                "--output-dir",
                "outputs/research/phase2",
                "--baseline-run-dir",
                "outputs/backtests/base",
                "--core-wf-dir",
                "outputs/walk_forward/core",
                "--satellite-wf-dir",
                "outputs/walk_forward/satellite",
                "--portfolio-run-dir",
                "outputs/portfolios/pf",
                "--portfolio-wf-dir",
                "outputs/portfolio_walk_forward/pfwf",
            ]
        )
        self.assertEqual(args.command, "phase2")
        self.assertEqual(args.output_dir, "outputs/research/phase2")
        self.assertEqual(args.baseline_run_dir, "outputs/backtests/base")
        self.assertEqual(args.core_wf_dir, "outputs/walk_forward/core")
        self.assertEqual(args.satellite_wf_dir, "outputs/walk_forward/satellite")
        self.assertEqual(args.portfolio_run_dir, "outputs/portfolios/pf")
        self.assertEqual(args.portfolio_wf_dir, "outputs/portfolio_walk_forward/pfwf")

    def test_phase2_review_cli_options(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "phase2-review",
                "--allocator-dir",
                "outputs/portfolio_walk_forward/allocator",
                "--components-path",
                "outputs/research/phase2/phase2_components.csv",
                "--promotion-review-dir",
                "outputs/research/promotion",
                "--output-dir",
                "outputs/research/review",
            ]
        )
        self.assertEqual(args.command, "phase2-review")
        self.assertEqual(args.allocator_dir, "outputs/portfolio_walk_forward/allocator")
        self.assertEqual(args.components_path, "outputs/research/phase2/phase2_components.csv")
        self.assertEqual(args.promotion_review_dir, "outputs/research/promotion")
        self.assertEqual(args.output_dir, "outputs/research/review")

    def test_research_allocator_cli_defaults_use_promoted_candidate(self) -> None:
        parser = build_parser()
        expected = (
            "outputs/portfolio_source_selection/main_chinext_source_selection_highgain_pos8_dd50_cap30_activation_dd50_20260624"
        )
        expected_components = "outputs/research/phase2_model_status_source_selection_default_20260614/phase2_components.csv"
        expected_market_data = "D:/codex/daily-market-data"

        phase2_review = parser.parse_args(["phase2-review"])
        daily_check = parser.parse_args(["daily-check"])
        paper_account = parser.parse_args(["paper-account"])
        dashboard = parser.parse_args(["dashboard"])
        daily_pipeline = parser.parse_args(["daily-pipeline"])

        self.assertEqual(phase2_review.allocator_dir, expected)
        self.assertEqual(daily_check.allocator_dir, expected)
        self.assertEqual(paper_account.allocator_dir, expected)
        self.assertEqual(dashboard.allocator_dir, expected)
        self.assertEqual(daily_pipeline.allocator_dir, expected)
        self.assertEqual(phase2_review.components_path, expected_components)
        self.assertEqual(daily_check.components_path, expected_components)
        self.assertEqual(daily_pipeline.components_path, expected_components)
        self.assertEqual(dashboard.data_cache_dir, expected_market_data)
        self.assertEqual(daily_pipeline.data_cache_dir, expected_market_data)

    def test_daily_check_cli_options(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "daily-check",
                "--output-dir",
                "outputs/research/daily",
                "--phase2-review-dir",
                "outputs/research/daily/phase2",
                "--allocator-dir",
                "outputs/portfolio_walk_forward/allocator",
                "--components-path",
                "outputs/research/phase2/phase2_components.csv",
                "--promotion-review-dir",
                "outputs/research/promotion",
                "--max-staleness-days",
                "5",
                "--as-of-date",
                "2024-04-02",
                "--date-stamp",
            ]
        )
        self.assertEqual(args.command, "daily-check")
        self.assertEqual(args.output_dir, "outputs/research/daily")
        self.assertEqual(args.phase2_review_dir, "outputs/research/daily/phase2")
        self.assertEqual(args.allocator_dir, "outputs/portfolio_walk_forward/allocator")
        self.assertEqual(args.components_path, "outputs/research/phase2/phase2_components.csv")
        self.assertEqual(args.promotion_review_dir, "outputs/research/promotion")
        self.assertEqual(args.max_staleness_days, 5)
        self.assertEqual(args.as_of_date, "2024-04-02")
        self.assertTrue(args.date_stamp)

    def test_dashboard_cli_options(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "dashboard",
                "--output-dir",
                "outputs/research/dashboard",
                "--daily-check-dir",
                "outputs/research/daily",
                "--paper-account-dir",
                "outputs/research/paper",
                "--sentiment-dir",
                "outputs/research/sentiment",
                "--data-cache-dir",
                "data/processed",
                "--allocator-wf-dir",
                "outputs/portfolio_walk_forward/allocator",
                "--trigger-report",
                "D:/codex/outputs/trigger_reports/latest_trigger.md",
                "--daily-run-status-dir",
                "outputs/research/daily_run_status",
                "--allocator-observation-dir",
                "outputs/research/allocator_observation",
                "--max-staleness-days",
                "4",
                "--min-cache-fresh-ratio",
                "0.8",
                "--as-of-date",
                "2024-04-02",
                "--date-stamp",
            ]
        )
        self.assertEqual(args.command, "dashboard")
        self.assertEqual(args.output_dir, "outputs/research/dashboard")
        self.assertEqual(args.daily_check_dir, "outputs/research/daily")
        self.assertEqual(args.paper_account_dir, "outputs/research/paper")
        self.assertEqual(args.sentiment_dir, "outputs/research/sentiment")
        self.assertEqual(args.data_cache_dir, "data/processed")
        self.assertEqual(args.allocator_dir, "outputs/portfolio_walk_forward/allocator")
        self.assertEqual(args.trigger_report, "D:/codex/outputs/trigger_reports/latest_trigger.md")
        self.assertEqual(args.daily_run_status_dir, "outputs/research/daily_run_status")
        self.assertEqual(args.allocator_observation_dir, "outputs/research/allocator_observation")
        self.assertEqual(args.max_staleness_days, 4)
        self.assertAlmostEqual(args.min_cache_fresh_ratio, 0.8)
        self.assertEqual(args.as_of_date, "2024-04-02")
        self.assertTrue(args.date_stamp)

    def test_dashboard_cli_prints_live_preflight_lines(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            output_dir = root / "dashboard"
            result = SimpleNamespace(
                output_dir=output_dir,
                snapshot_path=output_dir / "latest_dashboard_snapshot.json",
                report_path=output_dir / "latest_dashboard.md",
                snapshot={
                    "dashboard_posture": "core_base_watch_allocator_gate",
                    "action_posture": "review_core_base_allocator_gate",
                    "sentiment_state": "warm",
                    "market_cache_status": "fresh_enough",
                    "allocator_input_status": "fresh_enough",
                    "trigger_freshness_status": "fresh_enough",
                    "model_audit_status": "ok",
                    "pipeline_history_health_state": "ok",
                    "allocator_observation_decision_status": "outcomes_ready_for_review",
                    "live_preflight_status": "live_preflight_completed",
                    "live_preflight_decision": "ready_for_manual_broker_reconciliation",
                    "live_preflight_blocking_items_count": 1,
                    "live_preflight_monitor_items_count": 0,
                    "live_preflight_status_path": str(root / "outputs" / "logs" / "live_preflight_20260615.status.json"),
                    "live_preflight_report_path": str(root / "outputs" / "research" / "live_preflight" / "live_preflight.md"),
                    "live_preflight_snapshot_path": str(root / "outputs" / "research" / "live_preflight" / "live_preflight_snapshot.json"),
                    "daily_preflight_skipped": True,
                },
            )
            for file_path in (
                result.snapshot["live_preflight_status_path"],
                result.snapshot["live_preflight_report_path"],
                result.snapshot["live_preflight_snapshot_path"],
            ):
                Path(file_path).parent.mkdir(parents=True, exist_ok=True)
                Path(file_path).write_text("ok", encoding="utf-8")
            output = io.StringIO()
            with patch("quant_etf_lab.cli.run_daily_dashboard", return_value=result):
                with redirect_stdout(output):
                    self.assertEqual(
                        main(
                            [
                                "dashboard",
                                "--output-dir",
                                str(output_dir),
                            ]
                        ),
                        0,
                    )
            output_text = output.getvalue()
            self.assertIn("Dashboard completed:", output_text)
            self.assertIn("Dashboard posture: core_base_watch_allocator_gate", output_text)
            self.assertIn("Live preflight status: live_preflight_completed", output_text)
            self.assertIn("Live preflight decision: ready_for_manual_broker_reconciliation", output_text)
            self.assertIn("Live preflight blockers/monitors: 1 / 0", output_text)
            self.assertIn("Daily preflight skipped: `True`", output_text)
            self.assertIn(f"Live preflight status path: {result.snapshot['live_preflight_status_path']}", output_text)
            self.assertIn(f"Live preflight report path: {result.snapshot['live_preflight_report_path']}", output_text)
            self.assertIn(f"Live preflight snapshot path: {result.snapshot['live_preflight_snapshot_path']}", output_text)
            self.assertTrue(Path(result.snapshot["live_preflight_status_path"]).exists())
            self.assertTrue(Path(result.snapshot["live_preflight_report_path"]).exists())
            self.assertTrue(Path(result.snapshot["live_preflight_snapshot_path"]).exists())

    def test_daily_pipeline_cli_prints_live_preflight_and_alert_lines(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            output_dir = root / "pipeline"
            result = SimpleNamespace(
                output_dir=output_dir,
                snapshot_path=output_dir / "daily_pipeline_snapshot.json",
                report_path=output_dir / "daily_pipeline.md",
                history_path=root / "history.csv",
                daily_check_result=SimpleNamespace(report_path=root / "daily" / "daily_model_check.md"),
                paper_account_result=SimpleNamespace(
                    report_path=root / "paper" / "paper_account.md",
                    target_holdings_report_path=root / "paper" / "paper_account_target_holdings.md",
                    stock_targets_report_path=root / "paper" / "stock_targets.md",
                    stock_target_review_report_path=root / "paper" / "stock_target_review.md",
                    stock_target_review_notes_path=root / "paper" / "stock_target_review_notes.md",
                    stock_target_review_actions_report_path=root / "paper" / "stock_target_review_actions.md",
                    stock_target_review_assistant_report_path=root / "paper" / "stock_target_review_assistant.md",
                    stock_target_review_decision_template_report_path=root / "paper" / "stock_target_review_decision_template.md",
                    stock_target_review_decision_template_xlsx_path=root / "paper" / "stock_target_review_decision_template.xlsx",
                    stock_target_review_outcomes_report_path=root / "paper" / "stock_target_review_outcomes.md",
                    stock_target_review_outcomes_history_report_path=root / "paper" / "stock_target_review_outcomes_history.md",
                    stock_target_review_outcome_analysis_report_path=root / "paper" / "stock_target_review_outcome_analysis.md",
                    stock_target_review_outcome_calendar_report_path=root / "paper" / "stock_target_review_outcome_calendar.md",
                    stock_target_review_outcome_due_report_path=root / "paper" / "stock_target_review_outcome_due.md",
                ),
                dashboard_result=SimpleNamespace(report_path=root / "dashboard" / "latest_dashboard.md"),
                alerts_result=SimpleNamespace(payload={"dashboard_alerts": []}, report_path=root / "alerts" / "alerts.md"),
                snapshot={
                    "dashboard_posture": "core_base_watch_allocator_gate",
                    "trading_day_gate_status": "trading_day_data_ready",
                    "after_close_data_status": "ready",
                    "paper_latest_regime": "risk_on",
                    "history_path": str(root / "history.csv"),
                    "history_review_report_path": str(root / "history" / "history_review.md"),
                    "history_review_health_state": "ok",
                    "history_review_alert_count": 0,
                    "satellite_risk_budget_report_path": str(root / "satellite" / "satellite_risk_budget_review.md"),
                    "satellite_risk_budget_decision": "eligible_for_small_satellite_trial",
                    "live_shadow_preflight_report_path": str(root / "preflight" / "live_shadow_preflight.md"),
                    "live_shadow_preflight_decision": "blocked",
                    "live_shadow_report_path": str(root / "shadow" / "live_shadow.md"),
                    "live_preflight_report_path": str(root / "live_preflight" / "live_preflight.md"),
                    "live_preflight_decision": "ready_for_manual_broker_reconciliation",
                    "daily_preflight_skipped": True,
                    "live_shadow_status": "ready",
                    "live_shadow_order_count": 0,
                    "model_audit_status": "ok",
                    "model_audit_walk_forward_action_items": 0,
                    "promotion_decision": "promote_candidate",
                    "promotion_report_path": str(root / "promotion" / "promotion_review.md"),
                    "alert_level": "info",
                    "alert_action_stage": "monitor",
                    "alert_count": 0,
                    "alerts_report_path": str(root / "alerts" / "alerts.md"),
                    "live_preflight_status_path": str(root / "outputs" / "logs" / "live_preflight_20260615.status.json"),
                    "live_preflight_blocking_items_count": 0,
                    "live_preflight_monitor_items_count": 1,
                },
            )
            for file_path in (
                result.snapshot["live_preflight_status_path"],
                result.snapshot["live_preflight_report_path"],
                result.snapshot["history_path"],
                result.snapshot["history_review_report_path"],
                result.snapshot["satellite_risk_budget_report_path"],
                result.snapshot["live_shadow_preflight_report_path"],
                result.snapshot["live_shadow_report_path"],
                result.snapshot["live_preflight_report_path"],
                result.snapshot["promotion_report_path"],
                result.snapshot["alerts_report_path"],
            ):
                path = Path(file_path)
                path.parent.mkdir(parents=True, exist_ok=True)
                path.write_text("ok", encoding="utf-8")
            output = io.StringIO()
            with patch("quant_etf_lab.cli.run_daily_pipeline", return_value=result):
                with redirect_stdout(output):
                    self.assertEqual(
                        main(
                            [
                                "daily-pipeline",
                                "--output-dir",
                                str(output_dir),
                                "--components-path",
                                str(root / "components.csv"),
                                "--allocator-wf-dir",
                                str(root / "allocator"),
                                "--config",
                                str(root / "configs" / "portfolio.yaml"),
                                "--history-file",
                                str(root / "history.csv"),
                                "--network-lab-snapshot",
                                str(root / "network_lab_snapshot.json"),
                                "--network-max-cluster-count-warning",
                                "2",
                                "--network-residual-mi-warning",
                                "0.25",
                            ]
                        ),
                        0,
                    )
            output_text = output.getvalue()
            self.assertIn(f"Dashboard posture: core_base_watch_allocator_gate", output_text)
            self.assertIn("Trading-day gate: trading_day_data_ready", output_text)
            self.assertIn(f"Live preflight: {result.snapshot['live_preflight_report_path']}", output_text)
            self.assertIn(f"Live preflight decision: {result.snapshot['live_preflight_decision']}", output_text)
            self.assertIn(f"Daily preflight skipped: {result.snapshot['daily_preflight_skipped']}", output_text)
            self.assertIn(f"Alert level: {result.snapshot['alert_level']}", output_text)
            self.assertIn(f"Alert count: {result.snapshot['alert_count']}", output_text)
            self.assertIn(f"History: {result.snapshot['history_path']}", output_text)
            self.assertIn(
                f"Paper stock target review decision template: "
                f"{result.paper_account_result.stock_target_review_decision_template_report_path}",
                output_text,
            )
            self.assertIn(
                f"Paper stock target review decision template XLSX: "
                f"{result.paper_account_result.stock_target_review_decision_template_xlsx_path}",
                output_text,
            )
            self.assertTrue(Path(result.snapshot["history_path"]).exists())
            self.assertTrue(Path(result.snapshot["live_preflight_report_path"]).exists())

    def test_daily_pipeline_main_passes_network_args_to_run(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            output_dir = root / "pipeline"
            snapshot_path = output_dir / "daily_pipeline_snapshot.json"
            report_path = output_dir / "daily_pipeline.md"
            result = SimpleNamespace(
                output_dir=output_dir,
                snapshot_path=snapshot_path,
                report_path=report_path,
                history_path=root / "history.csv",
                daily_check_result=SimpleNamespace(report_path=root / "daily" / "daily_model_check.md"),
                paper_account_result=SimpleNamespace(
                    report_path=root / "paper" / "paper_account.md",
                    target_holdings_report_path=root / "paper" / "paper_account_target_holdings.md",
                    stock_targets_report_path=root / "paper" / "stock_targets.md",
                    stock_target_review_report_path=root / "paper" / "stock_target_review.md",
                    stock_target_review_notes_path=root / "paper" / "stock_target_review_notes.md",
                    stock_target_review_actions_report_path=root / "paper" / "stock_target_review_actions.md",
                    stock_target_review_assistant_report_path=root / "paper" / "stock_target_review_assistant.md",
                    stock_target_review_decision_template_report_path=root / "paper" / "stock_target_review_decision_template.md",
                    stock_target_review_decision_template_xlsx_path=root
                    / "paper"
                    / "stock_target_review_decision_template.xlsx",
                    stock_target_review_outcomes_report_path=root / "paper" / "stock_target_review_outcomes.md",
                    stock_target_review_outcomes_history_report_path=root / "paper" / "stock_target_review_outcomes_history.md",
                    stock_target_review_outcome_analysis_report_path=root / "paper" / "stock_target_review_outcome_analysis.md",
                    stock_target_review_outcome_calendar_report_path=root / "paper" / "stock_target_review_outcome_calendar.md",
                    stock_target_review_outcome_due_report_path=root / "paper" / "stock_target_review_outcome_due.md",
                ),
                dashboard_result=SimpleNamespace(report_path=root / "dashboard" / "latest_dashboard.md"),
                snapshot={
                    "dashboard_posture": "core_base_watch_allocator_gate",
                    "trading_day_gate_status": "trading_day_data_ready",
                    "after_close_data_status": "ready",
                    "paper_latest_regime": "risk_on",
                    "history_path": str(root / "history.csv"),
                    "history_review_report_path": str(root / "history" / "history_review.md"),
                    "history_review_health_state": "ok",
                    "history_review_alert_count": 0,
                    "satellite_risk_budget_report_path": str(root / "satellite" / "satellite_risk_budget_review.md"),
                    "satellite_risk_budget_decision": "eligible_for_small_satellite_trial",
                    "live_shadow_preflight_report_path": str(root / "preflight" / "live_shadow_preflight.md"),
                    "live_shadow_preflight_decision": "blocked",
                    "live_shadow_report_path": str(root / "shadow" / "live_shadow.md"),
                    "live_preflight_report_path": str(root / "live_preflight" / "live_preflight.md"),
                    "live_preflight_decision": "ready_for_manual_broker_reconciliation",
                    "daily_preflight_skipped": True,
                    "model_audit_status": "ok",
                    "model_audit_walk_forward_action_items": 0,
                    "promotion_decision": "promote_candidate",
                    "promotion_report_path": str(root / "promotion" / "promotion_review.md"),
                    "alert_level": "info",
                    "alert_action_stage": "monitor",
                    "alert_count": 0,
                    "alerts_report_path": str(root / "alerts" / "alerts.md"),
                    "live_preflight_status_path": str(root / "outputs" / "logs" / "live_preflight_20260615.status.json"),
                    "live_preflight_blocking_items_count": 0,
                    "live_preflight_monitor_items_count": 1,
                },
            )
            with patch("quant_etf_lab.cli.run_daily_pipeline") as mock_run:
                mock_run.return_value = result
                self.assertEqual(
                    main(
                        [
                            "daily-pipeline",
                            "--output-dir",
                            str(output_dir),
                            "--components-path",
                            str(root / "components.csv"),
                            "--allocator-wf-dir",
                            str(root / "allocator"),
                            "--config",
                            str(root / "configs" / "portfolio.yaml"),
                            "--history-file",
                            str(root / "history.csv"),
                            "--network-lab-snapshot",
                            str(root / "network_lab_snapshot.json"),
                            "--network-max-cluster-count-warning",
                            "2",
                            "--network-residual-mi-warning",
                            "0.25",
                        ]
                    ),
                    0,
                )
            called_kwargs = mock_run.call_args.kwargs
            self.assertEqual(str(called_kwargs["network_lab_snapshot"]), str(root / "network_lab_snapshot.json"))
            self.assertEqual(called_kwargs["network_max_cluster_count_warning"], 2)
            self.assertAlmostEqual(called_kwargs["network_residual_mi_warning"], 0.25)

    def test_daily_pipeline_cli_options(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "daily-pipeline",
                "--output-dir",
                "outputs/research/pipeline",
                "--daily-check-output-dir",
                "outputs/research/daily",
                "--paper-account-output-dir",
                "outputs/research/paper",
                "--dashboard-output-dir",
                "outputs/research/dashboard",
                "--phase2-status-dir",
                "outputs/research/phase2",
                "--allocator-wf-dir",
                "outputs/portfolio_walk_forward/allocator",
                "--promotion-review-dir",
                "outputs/research/promotion",
                "--config",
                "configs/portfolio.yaml",
                "--sentiment-dir",
                "outputs/research/sentiment",
                "--data-cache-dir",
                "data/processed",
                "--history-file",
                "outputs/research/history.csv",
                "--history-review-output-dir",
                "outputs/research/history_review",
                "--satellite-risk-budget-output-dir",
                "outputs/research/satellite_risk_budget",
                "--satellite-trial-replay-output-dir",
                "outputs/research/satellite_trial_replay",
                "--satellite-trial-replay-horizons",
                "1d,5d,10d",
                "--live-shadow-output-dir",
                "outputs/research/live_shadow",
                "--live-shadow-preflight",
                "--live-shadow-preflight-fail-on-blocked",
                "--live-shadow-holdings-file",
                "inputs/live_holdings.csv",
                "--live-shadow-prices-file",
                "inputs/live_prices.csv",
                "--live-shadow-cash",
                "50000",
                "--live-shadow-lot-size",
                "100",
                "--live-shadow-min-trade-value",
                "800",
                "--live-shadow-max-position-weight",
                "0.2",
                "--live-shadow-max-gross-exposure",
                "0.9",
                "--network-lab-snapshot",
                "inputs/network_lab_snapshot.json",
                "--network-max-cluster-count-warning",
                "2",
                "--network-residual-mi-warning",
                "0.25",
                "--trigger-report",
                "D:/codex/outputs/trigger_reports/latest_trigger.md",
                "--max-staleness-days",
                "4",
                "--min-cache-fresh-ratio",
                "0.8",
                "--rebalance-cost-rate",
                "0.0003",
                "--stock-market-cap-path",
                "data/processed/stock_market_cap_yi.csv",
                "--stock-tracking-max-market-cap-yi",
                "1500",
                "--stock-review-notes-path",
                "outputs/research/stock_target_review_notes.csv",
                "--stock-review-outcomes-history-path",
                "outputs/research/stock_target_review_outcomes_history.csv",
                "--stock-review-drawdown-threshold",
                "-0.08",
                "--stock-review-watch-drawdown-threshold",
                "-0.06",
                "--stock-review-loss-attention-threshold",
                "-0.04",
                "--stock-review-gain-attention-threshold",
                "0.12",
                "--stock-review-watch-score-threshold",
                "25",
                "--stock-review-outcome-min-evaluable",
                "12",
                "--stock-review-outcome-min-group-evaluable",
                "4",
                "--stock-review-warning-only-after-close",
                "--chip-reversal-candidate-outcomes",
                "--chip-reversal-candidates-path",
                "outputs/research/chip_reversal_daily_candidates_20260616/chip_reversal_daily_candidates.csv",
                "--chip-reversal-outcomes-output-dir",
                "outputs/research/chip_reversal_candidate_outcomes_20260616",
                "--chip-reversal-outcome-horizons",
                "1,2",
                "--chip-reversal-outcome-success-threshold",
                "0.03",
                "--chip-reversal-outcome-min-ready-per-horizon",
                "30",
                "--chip-reversal-outcome-min-success-rate",
                "0.55",
                "--history-lookback-runs",
                "7",
                "--history-drawdown-watch-threshold",
                "-0.07",
                "--history-min-sharpe-watch",
                "0.55",
                "--fail-on-alert-level",
                "warning",
                "--fail-on-history-health",
                "watch",
                "--as-of-date",
                "2024-04-02",
                "--no-date-stamp",
            ]
        )
        self.assertEqual(args.command, "daily-pipeline")
        self.assertEqual(args.output_dir, "outputs/research/pipeline")
        self.assertEqual(args.daily_check_output_dir, "outputs/research/daily")
        self.assertEqual(args.paper_account_output_dir, "outputs/research/paper")
        self.assertEqual(args.dashboard_output_dir, "outputs/research/dashboard")
        self.assertEqual(args.components_path, "outputs/research/phase2")
        self.assertEqual(args.allocator_dir, "outputs/portfolio_walk_forward/allocator")
        self.assertEqual(args.promotion_review_dir, "outputs/research/promotion")
        self.assertEqual(args.config, "configs/portfolio.yaml")
        self.assertEqual(args.sentiment_dir, "outputs/research/sentiment")
        self.assertEqual(args.data_cache_dir, "data/processed")
        self.assertEqual(args.history_file, "outputs/research/history.csv")
        self.assertEqual(args.history_review_output_dir, "outputs/research/history_review")
        self.assertEqual(args.satellite_risk_budget_output_dir, "outputs/research/satellite_risk_budget")
        self.assertEqual(args.satellite_trial_replay_output_dir, "outputs/research/satellite_trial_replay")
        self.assertEqual(args.satellite_trial_replay_horizons, "1d,5d,10d")
        self.assertEqual(args.live_shadow_output_dir, "outputs/research/live_shadow")
        self.assertTrue(args.live_shadow_preflight)
        self.assertTrue(args.live_shadow_preflight_fail_on_blocked)
        self.assertEqual(args.live_shadow_holdings_file, "inputs/live_holdings.csv")
        self.assertEqual(args.live_shadow_prices_file, "inputs/live_prices.csv")
        self.assertEqual(args.live_shadow_cash, 50000.0)
        self.assertEqual(args.live_shadow_lot_size, 100)
        self.assertAlmostEqual(args.live_shadow_min_trade_value, 800.0)
        self.assertAlmostEqual(args.live_shadow_max_position_weight, 0.2)
        self.assertAlmostEqual(args.live_shadow_max_gross_exposure, 0.9)
        self.assertEqual(args.network_lab_snapshot, "inputs/network_lab_snapshot.json")
        self.assertEqual(args.network_max_cluster_count_warning, 2)
        self.assertAlmostEqual(args.network_residual_mi_warning, 0.25)
        self.assertEqual(args.max_staleness_days, 4)
        self.assertAlmostEqual(args.min_cache_fresh_ratio, 0.8)
        self.assertAlmostEqual(args.rebalance_cost_rate, 0.0003)
        self.assertEqual(args.stock_market_cap_path, "data/processed/stock_market_cap_yi.csv")
        self.assertAlmostEqual(args.stock_tracking_max_market_cap_yi, 1500.0)
        self.assertEqual(args.stock_review_notes_path, "outputs/research/stock_target_review_notes.csv")
        self.assertEqual(args.stock_review_outcomes_history_path, "outputs/research/stock_target_review_outcomes_history.csv")
        self.assertAlmostEqual(args.stock_review_drawdown_threshold, -0.08)
        self.assertAlmostEqual(args.stock_review_watch_drawdown_threshold, -0.06)
        self.assertAlmostEqual(args.stock_review_loss_attention_threshold, -0.04)
        self.assertAlmostEqual(args.stock_review_gain_attention_threshold, 0.12)
        self.assertAlmostEqual(args.stock_review_watch_score_threshold, 25)
        self.assertEqual(args.stock_review_outcome_min_evaluable, 12)
        self.assertEqual(args.stock_review_outcome_min_group_evaluable, 4)
        self.assertTrue(args.stock_review_warning_only_after_close)
        self.assertTrue(args.chip_reversal_candidate_outcomes)
        self.assertEqual(
            args.chip_reversal_candidates_path,
            "outputs/research/chip_reversal_daily_candidates_20260616/chip_reversal_daily_candidates.csv",
        )
        self.assertEqual(args.chip_reversal_outcomes_output_dir, "outputs/research/chip_reversal_candidate_outcomes_20260616")
        self.assertEqual(args.chip_reversal_outcome_horizons, "1,2")
        self.assertAlmostEqual(args.chip_reversal_outcome_success_threshold, 0.03)
        self.assertEqual(args.chip_reversal_outcome_min_ready_per_horizon, 30)
        self.assertAlmostEqual(args.chip_reversal_outcome_min_success_rate, 0.55)
        self.assertTrue(args.run_history_review)
        self.assertEqual(args.history_lookback_runs, 7)
        self.assertAlmostEqual(args.history_drawdown_watch_threshold, -0.07)
        self.assertAlmostEqual(args.history_min_sharpe_watch, 0.55)
        self.assertEqual(args.fail_on_alert_level, "warning")
        self.assertEqual(args.fail_on_history_health, "watch")
        self.assertEqual(args.as_of_date, "2024-04-02")
        self.assertFalse(args.date_stamp)

    def test_daily_pipeline_runs_chip_reversal_candidate_outcomes_when_enabled(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            daily_result = SimpleNamespace(
                output_dir=root / "daily",
                report_path=root / "daily" / "daily_model_check.md",
                snapshot_path=root / "daily" / "daily_model_check_snapshot.json",
                phase2_result=None,
                snapshot={
                    "data_freshness_status": "fresh_enough",
                    "action_posture": "review_core_base_allocator_gate",
                    "phase2_posture": "core_base_allocator_gated_satellite",
                    "latest_date": "2026-06-16",
                },
            )

            class FakePaperResult:
                output_dir = root / "paper"
                ledger = pd.DataFrame()
                audit = pd.DataFrame()
                target_holdings = pd.DataFrame()
                stock_targets = pd.DataFrame()
                stock_target_review = pd.DataFrame()
                stock_target_review_actions = pd.DataFrame()
                stock_target_review_assistant = pd.DataFrame()
                stock_target_review_outcomes = pd.DataFrame()
                stock_target_review_outcomes_history = pd.DataFrame()
                stock_target_review_outcome_analysis = pd.DataFrame()
                stock_target_review_outcome_calendar = pd.DataFrame()
                stock_target_review_outcome_due = pd.DataFrame()
                monthly_returns = pd.DataFrame()
                metrics = {
                    "latest_date": "2026-06-16",
                    "latest_regime": "risk_off",
                    "latest_candidate": "core_only",
                    "latest_core_weight": 1.0,
                    "latest_satellite_weight": 0.0,
                    "latest_cash_weight": 0.0,
                    "final_equity": 1000000.0,
                    "total_return": 0.0,
                    "max_drawdown": -0.02,
                    "current_drawdown": -0.01,
                    "sharpe": 1.0,
                    "audit_event_count": 0,
                }

                def __getattr__(self, name: str) -> object:
                    if name.endswith("_payload"):
                        return {}
                    if name.endswith("_path") or name == "report_path":
                        return self.output_dir / f"{name}.csv"
                    raise AttributeError(name)

            dashboard_result = SimpleNamespace(
                output_dir=root / "dashboard",
                report_path=root / "dashboard" / "latest_dashboard.md",
                snapshot_path=root / "dashboard" / "latest_dashboard_snapshot.json",
                snapshot={
                    "dashboard_posture": "core_base_watch_allocator_gate",
                    "sentiment_state": "neutral",
                    "sentiment_freshness_status": "fresh_enough",
                    "market_cache_status": "fresh_enough",
                    "allocator_input_status": "fresh_enough",
                    "paper_freshness_status": "fresh_enough",
                    "trigger_freshness_status": "fresh_enough",
                    "model_audit_status": "ok",
                    "model_audit_walk_forward_action_items": 0,
                    "model_audit_walk_forward_resume_candidates": 0,
                    "model_audit_walk_forward_archive_review_candidates": 0,
                    "promotion_status": "ok",
                    "promotion_decision": "watch_candidate",
                    "pipeline_history_status": "ok",
                    "pipeline_history_health_state": "ok",
                    "pipeline_history_alert_count": 0,
                    "pipeline_history_latest_as_of_date": "2026-06-16",
                },
            )
            chip_result = SimpleNamespace(
                output_dir=root / "chip_outcomes",
                outcomes_path=root / "chip_outcomes" / "chip_reversal_candidate_outcomes.csv",
                summary_path=root / "chip_outcomes" / "chip_reversal_candidate_outcome_summary.csv",
                group_summary_path=root / "chip_outcomes" / "chip_reversal_candidate_outcome_group_summary.csv",
                snapshot_path=root / "chip_outcomes" / "chip_reversal_candidate_outcomes_snapshot.json",
                report_path=root / "chip_outcomes" / "chip_reversal_candidate_outcomes.md",
                outcomes=pd.DataFrame(),
                summary=pd.DataFrame(),
                group_summary=pd.DataFrame(),
                snapshot={
                    "status": "ok",
                    "candidate_count": 54,
                    "ready_outcome_count": 0,
                    "pending_outcome_count": 108,
                    "outcome_readiness_status": "waiting_for_future_bar",
                    "outcome_analysis_status": "waiting_for_future_bar",
                    "outcome_ready_horizons": [],
                    "outcome_pending_horizons": ["1d", "2d"],
                    "next_outcome_review_horizon": "1d",
                    "next_outcome_review_reason": "future_bar_not_available",
                    "outcome_sample_warning": "waiting_for_future_bar",
                    "latest_available_market_trade_date": "2026-06-16",
                    "promotion_gate_status": "blocked",
                    "promotion_gate_reasons": ["1d:ready_count_below_minimum"],
                    "market_source_kind": "daily_market_data_csv",
                    "market_trade_date": "2026-06-16",
                    "broker_action": "none",
                    "research_only": True,
                },
            )
            satellite_result = SimpleNamespace(
                output_dir=root / "risk",
                report_path=root / "risk" / "satellite_risk_budget_review.md",
                snapshot_path=root / "risk" / "satellite_risk_budget_snapshot.json",
                checklist_path=root / "risk" / "satellite_risk_budget_checklist.csv",
                snapshot={
                    "risk_budget_decision": "wait_for_outcome_samples",
                    "risk_budget_reason": "Waiting for outcome samples.",
                    "next_action_stage": "wait",
                    "recommended_satellite_budget": 0.0,
                    "selected_horizon": "",
                    "outcome_ready_horizon_count": 0,
                },
            )
            alerts_result = SimpleNamespace(
                output_dir=root / "alerts",
                report_path=root / "alerts" / "daily_alerts.md",
                json_path=root / "alerts" / "daily_alerts.json",
                latest_json_path=root / "alerts" / "latest.json",
                latest_report_path=root / "alerts" / "latest.md",
                payload={"alerts": [], "alert_level": "info", "action_stage": "monitor", "alert_count": 0},
            )
            candidates_path = root / "chip_candidates.csv"

            with patch("quant_etf_lab.daily_pipeline.run_daily_model_check", return_value=daily_result), patch(
                "quant_etf_lab.daily_pipeline.run_paper_account", return_value=FakePaperResult()
            ), patch(
                "quant_etf_lab.daily_pipeline.run_daily_dashboard", return_value=dashboard_result
            ), patch(
                "quant_etf_lab.daily_pipeline.run_chip_reversal_candidate_outcomes", return_value=chip_result
            ) as chip_mock, patch(
                "quant_etf_lab.daily_pipeline.run_satellite_risk_budget_review", return_value=satellite_result
            ), patch(
                "quant_etf_lab.daily_pipeline.write_daily_alerts", return_value=alerts_result
            ):
                result = run_daily_pipeline(
                    project_root=root,
                    output_dir=root / "pipeline",
                    as_of_date="2026-06-16",
                    date_stamp=False,
                    history_path=None,
                    run_history_review=False,
                    chip_reversal_candidate_outcomes=True,
                    chip_reversal_candidates_path=candidates_path,
                    chip_reversal_outcomes_output_dir=root / "chip_outcomes",
                )

            chip_mock.assert_called_once()
            self.assertEqual(chip_mock.call_args.kwargs["candidates_path"], candidates_path)
            self.assertEqual(result.chip_reversal_candidate_outcomes_result, chip_result)
            self.assertEqual(result.snapshot["chip_reversal_candidate_outcomes_status"], "ok")
            self.assertEqual(result.snapshot["chip_reversal_candidate_outcomes_candidate_count"], 54)
            self.assertEqual(result.snapshot["chip_reversal_candidate_outcomes_ready_count"], 0)
            self.assertEqual(result.snapshot["chip_reversal_candidate_outcomes_pending_count"], 108)
            self.assertEqual(result.snapshot["chip_reversal_candidate_outcomes_readiness_status"], "waiting_for_future_bar")
            self.assertEqual(result.snapshot["chip_reversal_candidate_outcomes_analysis_status"], "waiting_for_future_bar")
            self.assertEqual(result.snapshot["chip_reversal_candidate_outcomes_ready_horizons"], [])
            self.assertEqual(result.snapshot["chip_reversal_candidate_outcomes_pending_horizons"], ["1d", "2d"])
            self.assertEqual(result.snapshot["chip_reversal_candidate_outcomes_next_review_horizon"], "1d")
            self.assertEqual(result.snapshot["chip_reversal_candidate_outcomes_next_review_reason"], "future_bar_not_available")
            self.assertEqual(result.snapshot["chip_reversal_candidate_outcomes_sample_warning"], "waiting_for_future_bar")
            self.assertEqual(result.snapshot["chip_reversal_candidate_outcomes_latest_market_trade_date"], "2026-06-16")
            self.assertEqual(result.snapshot["chip_reversal_candidate_outcomes_promotion_gate_status"], "blocked")
            self.assertEqual(result.snapshot["chip_reversal_candidate_outcomes_broker_action"], "none")

    def test_daily_pipeline_live_shadow_preflight_flag_defaults_to_fail(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "daily-pipeline",
                "--live-shadow-preflight",
            ]
        )

        self.assertTrue(args.live_shadow_preflight)
        self.assertTrue(args.live_shadow_preflight_fail_on_blocked)

    def test_daily_pipeline_live_shadow_preflight_flag_can_disable_fail(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "daily-pipeline",
                "--live-shadow-preflight",
                "--no-live-shadow-preflight-fail-on-blocked",
            ]
        )

        self.assertTrue(args.live_shadow_preflight)
        self.assertFalse(args.live_shadow_preflight_fail_on_blocked)

    def test_pipeline_history_cli_options(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "pipeline-history",
                "--history-file",
                "outputs/research/history.csv",
                "--output-dir",
                "outputs/research/history_review",
                "--as-of-date",
                "2024-04-03",
                "--lookback-runs",
                "5",
                "--max-staleness-days",
                "2",
                "--drawdown-watch-threshold",
                "-0.08",
                "--min-sharpe-watch",
                "0.6",
                "--fail-on-health",
                "watch",
            ]
        )
        self.assertEqual(args.command, "pipeline-history")
        self.assertEqual(args.history_file, "outputs/research/history.csv")
        self.assertEqual(args.output_dir, "outputs/research/history_review")
        self.assertEqual(args.as_of_date, "2024-04-03")
        self.assertEqual(args.lookback_runs, 5)
        self.assertEqual(args.max_staleness_days, 2)
        self.assertAlmostEqual(args.drawdown_watch_threshold, -0.08)
        self.assertAlmostEqual(args.min_sharpe_watch, 0.6)
        self.assertEqual(args.fail_on_health, "watch")

    def test_live_shadow_cli_options(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "live-shadow",
                "--holdings-file",
                "inputs/live_holdings.csv",
                "--targets-file",
                "outputs/research/paper/stock_targets.csv",
                "--prices-file",
                "inputs/live_prices.csv",
                "--cash",
                "50000",
                "--output-dir",
                "outputs/research/live_shadow",
                "--as-of-date",
                "2026-06-12",
                "--lot-size",
                "100",
                "--min-trade-value",
                "800",
                "--max-position-weight",
                "0.2",
                "--max-gross-exposure",
                "0.9",
            ]
        )

        self.assertEqual(args.command, "live-shadow")
        self.assertEqual(args.holdings_file, "inputs/live_holdings.csv")
        self.assertEqual(args.targets_file, "outputs/research/paper/stock_targets.csv")
        self.assertEqual(args.prices_file, "inputs/live_prices.csv")
        self.assertEqual(args.cash, 50000.0)
        self.assertEqual(args.output_dir, "outputs/research/live_shadow")
        self.assertEqual(args.as_of_date, "2026-06-12")
        self.assertEqual(args.lot_size, 100)
        self.assertAlmostEqual(args.min_trade_value, 800.0)
        self.assertAlmostEqual(args.max_position_weight, 0.2)
        self.assertAlmostEqual(args.max_gross_exposure, 0.9)

    def test_live_shadow_preflight_cli_options(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "live-shadow-preflight",
                "--holdings-file",
                "inputs/live_holdings.csv",
                "--targets-file",
                "outputs/research/paper/stock_targets.csv",
                "--prices-file",
                "inputs/live_prices.csv",
                "--cash",
                "50000",
                "--output-dir",
                "outputs/research/live_shadow_preflight",
                "--as-of-date",
                "2026-06-12",
                "--lot-size",
                "100",
                "--min-trade-value",
                "800",
                "--max-position-weight",
                "0.2",
                "--max-gross-exposure",
                "0.9",
                "--fail-on-blocked",
            ]
        )

        self.assertEqual(args.command, "live-shadow-preflight")
        self.assertEqual(args.holdings_file, "inputs/live_holdings.csv")
        self.assertEqual(args.targets_file, "outputs/research/paper/stock_targets.csv")
        self.assertEqual(args.prices_file, "inputs/live_prices.csv")
        self.assertEqual(args.cash, 50000.0)
        self.assertEqual(args.output_dir, "outputs/research/live_shadow_preflight")
        self.assertEqual(args.as_of_date, "2026-06-12")
        self.assertEqual(args.lot_size, 100)
        self.assertAlmostEqual(args.min_trade_value, 800.0)
        self.assertAlmostEqual(args.max_position_weight, 0.2)
        self.assertAlmostEqual(args.max_gross_exposure, 0.9)
        self.assertTrue(args.fail_on_blocked)

    def test_live_shadow_template_cli_options(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "live-shadow-template",
                "--targets-file",
                "outputs/research/paper/stock_targets.csv",
                "--output-dir",
                "outputs/research/live_shadow_import",
                "--as-of-date",
                "2026-06-15",
                "--blank-rows",
                "3",
            ]
        )

        self.assertEqual(args.command, "live-shadow-template")
        self.assertEqual(args.targets_file, "outputs/research/paper/stock_targets.csv")
        self.assertEqual(args.output_dir, "outputs/research/live_shadow_import")
        self.assertEqual(args.as_of_date, "2026-06-15")
        self.assertEqual(args.blank_rows, 3)

    def test_live_shadow_review_cli_options(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "live-shadow-review",
                "--orders-file",
                "outputs/research/live_shadow/live_shadow_orders.csv",
                "--targets-file",
                "outputs/research/paper/stock_targets.csv",
                "--output-dir",
                "outputs/research/live_shadow_review",
                "--as-of-date",
                "2026-06-15",
                "--fail-on-blocked",
            ]
        )

        self.assertEqual(args.command, "live-shadow-review")
        self.assertEqual(args.orders_file, "outputs/research/live_shadow/live_shadow_orders.csv")
        self.assertEqual(args.targets_file, "outputs/research/paper/stock_targets.csv")
        self.assertEqual(args.output_dir, "outputs/research/live_shadow_review")
        self.assertEqual(args.as_of_date, "2026-06-15")
        self.assertTrue(args.fail_on_blocked)

    def test_cli_health_gates(self) -> None:
        self.assertFalse(_alert_gate_failed("warning", "none"))
        self.assertFalse(_alert_gate_failed("info", "warning"))
        self.assertTrue(_alert_gate_failed("warning", "warning"))
        self.assertTrue(_alert_gate_failed("critical", "warning"))
        self.assertFalse(_history_gate_failed("ok", "watch"))
        self.assertTrue(_history_gate_failed("watch", "watch"))
        self.assertTrue(_history_gate_failed("blocked", "watch"))

    def test_paper_account_cli_options(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "paper-account",
                "--config",
                "configs/portfolio.yaml",
                "--allocator-wf-dir",
                "outputs/portfolio_walk_forward/allocator",
                "--output-dir",
                "outputs/research/paper",
                "--rebalance-cost-rate",
                "0.001",
                "--trigger-signal-path",
                "outputs/signal_history/signals_latest.csv",
                "--stock-market-cap-path",
                "data/processed/stock_market_cap_yi.csv",
                "--stock-tracking-max-market-cap-yi",
                "1500",
                "--stock-review-notes-path",
                "outputs/research/stock_target_review_notes.csv",
                "--stock-review-outcomes-history-path",
                "outputs/research/stock_target_review_outcomes_history.csv",
                "--stock-review-drawdown-threshold",
                "-0.08",
                "--stock-review-watch-drawdown-threshold",
                "-0.06",
                "--stock-review-loss-attention-threshold",
                "-0.04",
                "--stock-review-gain-attention-threshold",
                "0.12",
                "--stock-review-watch-score-threshold",
                "25",
                "--stock-review-outcome-min-evaluable",
                "12",
                "--stock-review-outcome-min-group-evaluable",
                "4",
            ]
        )
        self.assertEqual(args.command, "paper-account")
        self.assertEqual(args.config, "configs/portfolio.yaml")
        self.assertEqual(args.allocator_dir, "outputs/portfolio_walk_forward/allocator")
        self.assertEqual(args.output_dir, "outputs/research/paper")
        self.assertAlmostEqual(args.rebalance_cost_rate, 0.001)
        self.assertEqual(args.trigger_signal_path, "outputs/signal_history/signals_latest.csv")
        self.assertEqual(args.stock_market_cap_path, "data/processed/stock_market_cap_yi.csv")
        self.assertAlmostEqual(args.stock_tracking_max_market_cap_yi, 1500.0)
        self.assertEqual(args.stock_review_notes_path, "outputs/research/stock_target_review_notes.csv")
        self.assertEqual(args.stock_review_outcomes_history_path, "outputs/research/stock_target_review_outcomes_history.csv")
        self.assertAlmostEqual(args.stock_review_drawdown_threshold, -0.08)
        self.assertAlmostEqual(args.stock_review_watch_drawdown_threshold, -0.06)
        self.assertAlmostEqual(args.stock_review_loss_attention_threshold, -0.04)
        self.assertAlmostEqual(args.stock_review_gain_attention_threshold, 0.12)
        self.assertAlmostEqual(args.stock_review_watch_score_threshold, 25)
        self.assertEqual(args.stock_review_outcome_min_evaluable, 12)
        self.assertEqual(args.stock_review_outcome_min_group_evaluable, 4)

    def test_stock_review_apply_template_cli_options(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "stock-review-apply-template",
                "--template-path",
                "outputs/research/paper_account_latest/stock_target_review_decision_template.csv",
                "--notes-path",
                "outputs/research/stock_target_review_notes.csv",
                "--output-dir",
                "outputs/research/stock_review_template_apply_latest",
                "--reviewed-at",
                "2026-06-15T16:10:00",
            ]
        )

        self.assertEqual(args.command, "stock-review-apply-template")
        self.assertEqual(args.template_path, "outputs/research/paper_account_latest/stock_target_review_decision_template.csv")
        self.assertEqual(args.notes_path, "outputs/research/stock_target_review_notes.csv")
        self.assertEqual(args.output_dir, "outputs/research/stock_review_template_apply_latest")
        self.assertEqual(args.reviewed_at, "2026-06-15T16:10:00")

    def test_stock_review_apply_template_cli_prints_summary(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            result = SimpleNamespace(
                output_dir=root / "apply",
                template_path=root / "template.csv",
                notes_path=root / "notes.csv",
                notes_snapshot_path=root / "apply" / "stock_target_review_notes_after_apply.csv",
                audit_path=root / "apply" / "stock_target_review_decision_apply_audit.csv",
                report_path=root / "apply" / "stock_target_review_decision_apply.md",
                payload={
                    "applied_count": 1,
                    "blank_ignored_count": 2,
                    "invalid_status_count": 0,
                    "broker_action": "none",
                },
            )
            output = io.StringIO()
            with patch("quant_etf_lab.cli.apply_stock_target_review_decision_template", return_value=result) as mocked_apply:
                with redirect_stdout(output):
                    self.assertEqual(
                        main(
                            [
                                "stock-review-apply-template",
                                "--template-path",
                                str(result.template_path),
                                "--notes-path",
                                str(result.notes_path),
                                "--output-dir",
                                str(result.output_dir),
                                "--reviewed-at",
                                "2026-06-15T16:10:00",
                            ]
                        ),
                        0,
                    )
            mocked_apply.assert_called_once()
            output_text = output.getvalue()
            self.assertIn("Stock review template apply completed:", output_text)
            self.assertIn("Applied rows: 1", output_text)
            self.assertIn("Blank ignored rows: 2", output_text)
            self.assertIn("Invalid status rows: 0", output_text)
            self.assertIn("Broker action: none", output_text)

    def test_phase2_review_writes_monitoring_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            allocator = root / "allocator"
            phase2 = root / "phase2"
            promotion = root / "promotion"
            allocator.mkdir()
            phase2.mkdir()
            promotion.mkdir()
            core_params = allocator / "w1_selected_params.json"
            satellite_params = allocator / "w2_selected_params.json"
            core_params.write_text(
                json.dumps(
                    {
                        "weights": {
                            "risk_on": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                            "risk_off": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                            "crash": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                        },
                        "satellite_filter": {"enabled": False},
                        "regime_overrides": {},
                    }
                ),
                encoding="utf-8",
            )
            satellite_params.write_text(
                json.dumps(
                    {
                        "weights": {
                            "risk_on": {"core": 0.8, "satellite": 0.2, "cash": 0.0},
                            "risk_off": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                            "crash": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                        },
                        "satellite_filter": {"enabled": False},
                        "regime_overrides": {
                            "ma_window": 60,
                            "risk_on_drop_threshold": -0.03,
                            "crash_drop_threshold": -0.07,
                        },
                    }
                ),
                encoding="utf-8",
            )
            pd.DataFrame(
                {
                    "window": ["w1", "w2"],
                    "train_start": ["20220101", "20220701"],
                    "train_end": ["20220630", "20221231"],
                    "test_start": ["20220701", "20230101"],
                    "test_end": ["20221231", "20230630"],
                    "selected_candidate": ["core_only", "sat20_ma60_drop03_unfiltered"],
                    "selected_score": [0.1, 0.2],
                    "selected_params_path": [str(core_params), str(satellite_params)],
                    "train_total_return": [0.05, 0.10],
                    "test_total_return": [0.02, 0.06],
                    "train_max_drawdown": [-0.08, -0.09],
                    "test_max_drawdown": [-0.04, -0.05],
                    "train_sharpe": [0.5, 0.8],
                    "test_sharpe": [0.4, 0.9],
                    "test_average_satellite_weight": [0.0, 0.08],
                    "test_satellite_active_day_ratio": [0.0, 0.4],
                    "test_risk_on_day_ratio": [0.0, 0.5],
                }
            ).to_csv(allocator / "portfolio_walk_forward_summary.csv", index=False)
            pd.DataFrame(
                {
                    "date": pd.to_datetime(["2024-01-31", "2024-02-29", "2024-03-31"]),
                    "window": ["w1", "w1", "w2"],
                    "stitched_equity": [100.0, 105.0, 102.0],
                }
            ).to_csv(allocator / "oos_equity_stitched.csv", index=False)
            pd.DataFrame(
                {
                    "layer": ["core", "portfolio", "allocator"],
                    "status": ["complete", "complete", "complete"],
                    "total_return": [0.01, 0.00, 0.02],
                    "max_drawdown": [-0.06, -0.08, -0.04],
                    "sharpe": [0.3, 0.1, 1.2],
                }
            ).to_csv(phase2 / "phase2_components.csv", index=False)
            (promotion / "allocator_promotion_snapshot.json").write_text(
                json.dumps(
                    {
                        "decision": "promote_candidate",
                        "candidate_passes_headline_gate": True,
                        "sensitivity_run_support_count": 2,
                        "sensitivity_group_support_count": 1,
                        "evidence_support_count": 1,
                        "support_group_count": 2,
                        "support_groups": ["execution_cost_stress", "guardrail"],
                        "sensitivity_support_count": 2,
                        "min_sensitivity_support": 2,
                        "return_edge": 0.062,
                        "sharpe_edge": 0.119,
                        "drawdown_change": 0.019,
                    }
                ),
                encoding="utf-8",
            )
            (promotion / "allocator_promotion_review.md").write_text("# Allocator Promotion Review\n", encoding="utf-8")

            result = run_phase2_review(
                project_root=root,
                output_dir=root / "review",
                components_path=phase2 / "phase2_components.csv",
                allocator_dir=allocator,
                promotion_review_dir=promotion,
            )

            self.assertTrue(result.report_path.exists())
            self.assertTrue(result.snapshot_path.exists())
            self.assertEqual(result.snapshot["selected_candidate"], "sat20_ma60_drop03_unfiltered")
            self.assertEqual(int(result.snapshot["regime_ma_window"]), 60)
            self.assertAlmostEqual(float(result.snapshot["risk_on_satellite_weight"]), 0.2)
            self.assertEqual(result.snapshot["posture"], "core_base_allocator_gated_satellite")
            self.assertEqual(result.snapshot["test_period"], "2023-01-01 to 2023-06-30")
            self.assertEqual(result.snapshot["promotion_decision"], "promote_candidate")
            self.assertEqual(int(result.snapshot["promotion_sensitivity_run_support_count"]), 2)
            self.assertEqual(int(result.snapshot["promotion_sensitivity_group_support_count"]), 1)
            self.assertEqual(int(result.snapshot["promotion_evidence_support_count"]), 1)
            self.assertEqual(int(result.snapshot["promotion_support_group_count"]), 2)
            self.assertEqual(int(result.snapshot["promotion_sensitivity_support_count"]), 2)
            self.assertAlmostEqual(float(result.snapshot["promotion_return_edge"]), 0.062)
            report_text = result.report_path.read_text(encoding="utf-8")
            self.assertIn("Defensive core is the base", report_text)
            self.assertIn("Activation allocator Sharpe | 1.200", report_text)
            self.assertIn("Allocator Promotion Watch", report_text)
            self.assertIn("promote_candidate", report_text)
            self.assertIn("Independent support groups | 2 / 2", report_text)
            self.assertIn("Sensitivity group support | 1", report_text)
            self.assertIn("Evidence group support | 1", report_text)
            self.assertIn("Sensitivity run support | 2", report_text)

    def test_phase2_review_reads_source_selection_nested_allocation_params(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            allocator = root / "allocator"
            phase2 = root / "phase2"
            promotion = root / "promotion"
            allocator.mkdir()
            phase2.mkdir()
            promotion.mkdir()
            nested_params = allocator / "w1_selected_params.json"
            nested_params.write_text(
                json.dumps(
                    {
                        "allocation": {
                            "weights": {
                                "risk_on": {"core": 0.75, "satellite": 0.25, "cash": 0.0},
                                "risk_off": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                                "crash": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                            },
                            "satellite_filter": {"enabled": False},
                            "regime_overrides": {
                                "ma_window": 60,
                                "risk_on_drop_threshold": -0.03,
                                "crash_drop_threshold": -0.07,
                            },
                        },
                        "source_name": "reversal",
                        "source_path": str(root / "reversal.csv"),
                    }
                ),
                encoding="utf-8",
            )
            pd.DataFrame(
                {
                    "window": ["w1"],
                    "train_start": ["20220101"],
                    "train_end": ["20220630"],
                    "test_start": ["20220701"],
                    "test_end": ["20221231"],
                    "selected_candidate": ["reversal__sat25_ma60_drop03_unfiltered"],
                    "selected_score": [0.2],
                    "selected_params_path": [str(nested_params)],
                    "test_total_return": [0.06],
                    "test_max_drawdown": [-0.05],
                    "test_sharpe": [0.9],
                    "test_average_satellite_weight": [0.08],
                    "test_risk_on_day_ratio": [0.5],
                }
            ).to_csv(allocator / "portfolio_walk_forward_summary.csv", index=False)
            pd.DataFrame(
                {
                    "date": pd.to_datetime(["2024-01-31", "2024-02-29"]),
                    "window": ["w1", "w1"],
                    "stitched_equity": [100.0, 106.0],
                }
            ).to_csv(allocator / "oos_equity_stitched.csv", index=False)
            pd.DataFrame(
                {
                    "layer": ["core", "portfolio", "allocator"],
                    "status": ["complete", "complete", "complete"],
                    "total_return": [0.03, 0.01, 0.06],
                    "max_drawdown": [-0.06, -0.08, -0.04],
                    "sharpe": [0.3, 0.1, 1.2],
                }
            ).to_csv(phase2 / "phase2_components.csv", index=False)
            (promotion / "allocator_promotion_snapshot.json").write_text(
                json.dumps(
                    {
                        "decision": "promote_candidate",
                        "candidate_passes_headline_gate": True,
                        "sensitivity_support_count": 2,
                        "min_sensitivity_support": 2,
                        "return_edge": 0.062,
                        "sharpe_edge": 0.119,
                        "drawdown_change": 0.019,
                    }
                ),
                encoding="utf-8",
            )
            (promotion / "allocator_promotion_review.md").write_text("# Allocator Promotion Review\n", encoding="utf-8")

            result = run_phase2_review(
                project_root=root,
                output_dir=root / "review",
                components_path=phase2 / "phase2_components.csv",
                allocator_dir=allocator,
                promotion_review_dir=promotion,
            )

            self.assertEqual(result.snapshot["selected_candidate"], "reversal__sat25_ma60_drop03_unfiltered")
            self.assertAlmostEqual(float(result.snapshot["risk_on_satellite_weight"]), 0.25)
            self.assertAlmostEqual(float(result.snapshot["risk_off_satellite_weight"]), 0.0)
            self.assertEqual(int(result.snapshot["regime_ma_window"]), 60)
            self.assertEqual(result.snapshot["posture"], "core_base_allocator_gated_satellite")
            self.assertEqual(result.snapshot["promotion_decision"], "promote_candidate")

    def test_phase2_review_auto_resolves_latest_promotion_dir(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            allocator = root / "allocator"
            phase2 = root / "phase2"
            research_root = root / "outputs" / "research"
            allocator.mkdir(parents=True, exist_ok=True)
            phase2.mkdir(parents=True, exist_ok=True)
            research_root.mkdir(parents=True, exist_ok=True)

            params_path = allocator / "w1_selected_params.json"
            params_path.write_text(
                json.dumps(
                    {
                        "weights": {
                            "risk_on": {"core": 0.7, "satellite": 0.3, "cash": 0.0},
                            "risk_off": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                            "crash": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                        },
                        "satellite_filter": {"enabled": False},
                        "regime_overrides": {
                            "ma_window": 60,
                            "risk_on_drop_threshold": -0.03,
                            "crash_drop_threshold": -0.07,
                        },
                    }
                ),
                encoding="utf-8",
            )
            pd.DataFrame(
                {
                    "window": ["w1"],
                    "train_start": ["20220101"],
                    "train_end": ["20220630"],
                    "test_start": ["20220701"],
                    "test_end": ["20221231"],
                    "selected_candidate": ["sat20_ma60_drop03_unfiltered"],
                    "selected_score": [0.2],
                    "selected_params_path": [str(params_path)],
                    "test_total_return": [0.06],
                    "test_max_drawdown": [-0.05],
                    "test_sharpe": [0.9],
                    "test_average_satellite_weight": [0.3],
                    "test_risk_on_day_ratio": [0.5],
                }
            ).to_csv(allocator / "portfolio_walk_forward_summary.csv", index=False)
            pd.DataFrame(
                {
                    "date": pd.to_datetime(["2024-01-31", "2024-02-29"]),
                    "window": ["w1", "w1"],
                    "stitched_equity": [100.0, 106.0],
                }
            ).to_csv(allocator / "oos_equity_stitched.csv", index=False)
            pd.DataFrame(
                {
                    "layer": ["core", "portfolio", "allocator"],
                    "status": ["complete", "complete", "complete"],
                    "total_return": [0.03, 0.01, 0.06],
                    "max_drawdown": [-0.06, -0.08, -0.04],
                    "sharpe": [0.3, 0.1, 1.2],
                }
            ).to_csv(phase2 / "phase2_components.csv", index=False)

            older = research_root / "allocator_promotion_test_old_20260614"
            latest = research_root / "allocator_promotion_test_latest_20260616"
            for dir_path, decision in [(older, "watch"), (latest, "promote_candidate")]:
                dir_path.mkdir(parents=True, exist_ok=True)
                (dir_path / "allocator_promotion_snapshot.json").write_text(
                    json.dumps(
                        {
                            "decision": decision,
                            "candidate_passes_headline_gate": True,
                            "sensitivity_support_count": 2,
                            "min_sensitivity_support": 2,
                            "return_edge": 0.045,
                            "sharpe_edge": 0.055,
                            "drawdown_change": -0.005,
                        }
                    ),
                    encoding="utf-8",
                )
                (dir_path / "allocator_promotion_review.md").write_text("# Allocator Promotion Review\n", encoding="utf-8")

            result = run_phase2_review(
                project_root=root,
                output_dir=root / "review",
                components_path=phase2 / "phase2_components.csv",
                allocator_dir=allocator,
            )

            self.assertEqual(result.snapshot["promotion_status"], "ok")
            self.assertEqual(result.snapshot["promotion_decision"], "promote_candidate")
            self.assertEqual(result.snapshot["promotion_review_dir"], str(latest))

    def test_allocator_promotion_promotes_candidate_with_support(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            baseline = root / "baseline"
            candidate = root / "candidate"
            sensitivity_a = root / "sensitivity_a"
            evidence = root / "evidence.json"
            output = root / "promotion"

            for run_dir, equity, window_returns in [
                (baseline, [100.0, 108.0, 111.0, 114.0], [0.02, 0.01, 0.02]),
                (candidate, [100.0, 120.0, 129.0, 145.0], [0.12, 0.08, 0.04]),
                (sensitivity_a, [100.0, 114.0, 124.0, 130.0], [0.09, 0.06, 0.02]),
            ]:
                run_dir.mkdir(parents=True, exist_ok=True)
                pd.DataFrame(
                    {
                        "date": pd.to_datetime(["2024-01-31", "2024-02-29", "2024-03-31", "2024-04-30"]),
                        "window": ["w1", "w1", "w1", "w1"],
                        "stitched_equity": equity,
                    }
                ).to_csv(run_dir / "oos_equity_stitched.csv", index=False)
                pd.DataFrame(
                    {
                        "window": ["w1", "w1", "w1"],
                        "test_total_return": window_returns,
                    }
                ).to_csv(run_dir / "portfolio_walk_forward_summary.csv", index=False)

            evidence.write_text(
                json.dumps(
                    {
                        "status": "supports_candidate",
                        "return_edge": 0.08,
                        "sharpe_edge": 0.05,
                        "drawdown_change": -0.02,
                    }
                ),
                encoding="utf-8",
            )

            result = run_allocator_promotion_review(
                baseline_dir=baseline,
                candidate_dir=candidate,
                sensitivity_dirs=[sensitivity_a],
                sensitivity_groups=["execution_cost_stress"],
                evidence_snapshots=[evidence],
                evidence_groups=["evidence_guardrail"],
                output_dir=output,
                min_return_edge=0.03,
                min_sharpe_edge=0.05,
                max_drawdown_worsening=0.02,
                min_positive_window_ratio=0.80,
                min_sensitivity_support=2,
            )

            self.assertEqual(result.decision, "promote_candidate")
            self.assertTrue(result.snapshot["candidate_passes_headline_gate"])
            self.assertEqual(int(result.snapshot["sensitivity_run_support_count"]), 1)
            self.assertEqual(int(result.snapshot["support_group_count"]), 2)
            self.assertEqual(int(result.snapshot["sensitivity_support_count"]), 2)
            self.assertEqual(result.snapshot["evidence_snapshots"], [str(evidence)])
            self.assertTrue(result.snapshot_path.exists())
            self.assertTrue(result.report_path.exists())
            self.assertIn("promote_candidate", result.snapshot["decision"])
            self.assertIn("Decision", result.report_path.read_text(encoding="utf-8"))

    def test_allocator_promotion_watches_when_support_is_insufficient(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            baseline = root / "baseline"
            candidate = root / "candidate"
            sensitivity_a = root / "sensitivity_a"
            output = root / "promotion"

            for run_dir, equity, window_returns in [
                (baseline, [100.0, 108.0, 111.0, 114.0], [0.02, 0.01, 0.02]),
                (candidate, [100.0, 112.0, 121.0, 130.0], [0.01, 0.03, 0.01]),
                (sensitivity_a, [100.0, 104.0, 105.0, 106.0], [-0.01, 0.01, -0.02]),
            ]:
                run_dir.mkdir(parents=True, exist_ok=True)
                pd.DataFrame(
                    {
                        "date": pd.to_datetime(["2024-01-31", "2024-02-29", "2024-03-31", "2024-04-30"]),
                        "window": ["w1", "w1", "w1", "w1"],
                        "stitched_equity": equity,
                    }
                ).to_csv(run_dir / "oos_equity_stitched.csv", index=False)
                pd.DataFrame(
                    {
                        "window": ["w1", "w1", "w1"],
                        "test_total_return": window_returns,
                    }
                ).to_csv(run_dir / "portfolio_walk_forward_summary.csv", index=False)

            result = run_allocator_promotion_review(
                baseline_dir=baseline,
                candidate_dir=candidate,
                sensitivity_dirs=[sensitivity_a],
                sensitivity_groups=["execution_cost_stress"],
                evidence_snapshots=(),
                output_dir=output,
                min_return_edge=0.01,
                min_sharpe_edge=0.05,
                max_drawdown_worsening=0.00,
                min_positive_window_ratio=0.80,
                min_sensitivity_support=1,
            )

            self.assertEqual(result.decision, "watch_candidate")
            self.assertTrue(result.snapshot["candidate_passes_headline_gate"])
            self.assertEqual(int(result.snapshot["support_group_count"]), 0)

    def test_allocator_promotion_rejects_candidate_if_headline_gate_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            baseline = root / "baseline"
            candidate = root / "candidate"
            sensitivity_a = root / "sensitivity_a"
            output = root / "promotion"

            for run_dir, equity, window_returns in [
                (baseline, [100.0, 118.0, 124.0, 130.0], [0.05, 0.05, 0.04]),
                (candidate, [100.0, 107.0, 112.0, 114.0], [0.03, 0.02, 0.02]),
                (sensitivity_a, [100.0, 115.0, 123.0, 132.0], [0.06, 0.07, 0.07]),
            ]:
                run_dir.mkdir(parents=True, exist_ok=True)
                pd.DataFrame(
                    {
                        "date": pd.to_datetime(["2024-01-31", "2024-02-29", "2024-03-31", "2024-04-30"]),
                        "window": ["w1", "w1", "w1", "w1"],
                        "stitched_equity": equity,
                    }
                ).to_csv(run_dir / "oos_equity_stitched.csv", index=False)
                pd.DataFrame(
                    {
                        "window": ["w1", "w1", "w1"],
                        "test_total_return": window_returns,
                    }
                ).to_csv(run_dir / "portfolio_walk_forward_summary.csv", index=False)

            result = run_allocator_promotion_review(
                baseline_dir=baseline,
                candidate_dir=candidate,
                sensitivity_dirs=[sensitivity_a],
                sensitivity_groups=["execution_cost_stress"],
                output_dir=output,
                min_return_edge=0.03,
                min_sharpe_edge=0.02,
                max_drawdown_worsening=0.01,
                min_positive_window_ratio=0.70,
                min_sensitivity_support=1,
            )

            self.assertEqual(result.decision, "reject_candidate")
            self.assertFalse(result.snapshot["candidate_passes_headline_gate"])
            self.assertEqual(int(result.snapshot["support_group_count"]), 0)
            self.assertIn("Decision", result.report_path.read_text(encoding="utf-8"))

    def test_allocator_promotion_rejects_mismatched_sensitivity_groups(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            baseline = root / "baseline"
            candidate = root / "candidate"
            sensitivity_a = root / "sensitivity_a"

            for run_dir, equity in [
                (baseline, [100.0, 118.0, 124.0, 130.0]),
                (candidate, [100.0, 132.0, 147.0, 153.0]),
                (sensitivity_a, [100.0, 118.0, 126.0, 129.0]),
            ]:
                run_dir.mkdir(parents=True, exist_ok=True)
                pd.DataFrame(
                    {
                        "date": pd.to_datetime(["2024-01-31", "2024-02-29", "2024-03-31", "2024-04-30"]),
                        "window": ["w1", "w1", "w1", "w1"],
                        "stitched_equity": equity,
                    }
                ).to_csv(run_dir / "oos_equity_stitched.csv", index=False)
                pd.DataFrame({"window": ["w1"], "test_total_return": [0.02]}).to_csv(
                    run_dir / "portfolio_walk_forward_summary.csv", index=False
                )

            with self.assertRaisesRegex(ValueError, "sensitivity_groups must have the same length"):
                run_allocator_promotion_review(
                    baseline_dir=baseline,
                    candidate_dir=candidate,
                    sensitivity_dirs=[sensitivity_a],
                    sensitivity_groups=["g1", "g2"],
                    output_dir=root / "promotion",
                )

    def test_allocator_promotion_rejects_mismatched_evidence_groups(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            baseline = root / "baseline"
            candidate = root / "candidate"
            sensitivity_a = root / "sensitivity_a"
            evidence_a = root / "evidence_a.json"
            evidence_b = root / "evidence_b.json"

            for run_dir, equity in [
                (baseline, [100.0, 118.0, 124.0, 130.0]),
                (candidate, [100.0, 132.0, 147.0, 153.0]),
                (sensitivity_a, [100.0, 118.0, 126.0, 129.0]),
            ]:
                run_dir.mkdir(parents=True, exist_ok=True)
                pd.DataFrame(
                    {
                        "date": pd.to_datetime(["2024-01-31", "2024-02-29", "2024-03-31", "2024-04-30"]),
                        "window": ["w1", "w1", "w1", "w1"],
                        "stitched_equity": equity,
                    }
                ).to_csv(run_dir / "oos_equity_stitched.csv", index=False)
                pd.DataFrame({"window": ["w1"], "test_total_return": [0.02]}).to_csv(
                    run_dir / "portfolio_walk_forward_summary.csv", index=False
                )
            evidence_a.write_text(json.dumps({"status": "supports_candidate"}), encoding="utf-8")
            evidence_b.write_text(json.dumps({"status": "supports_candidate"}), encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "evidence_groups must have the same length"):
                run_allocator_promotion_review(
                    baseline_dir=baseline,
                    candidate_dir=candidate,
                    sensitivity_dirs=[sensitivity_a],
                    sensitivity_groups=["g1"],
                    evidence_snapshots=[evidence_a, evidence_b],
                    evidence_groups=["e1"],
                    output_dir=root / "promotion",
                )

    def test_allocator_promotion_evidence_support_flag_from_status(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            baseline = root / "baseline"
            candidate = root / "candidate"
            sensitivity_a = root / "sensitivity_a"
            evidence = root / "evidence_support.json"
            output = root / "promotion"

            for run_dir, equity, window_returns in [
                (baseline, [100.0, 120.0, 128.0, 135.0], [0.03, 0.02, 0.01]),
                (candidate, [100.0, 180.0, 265.0, 360.0], [0.25, 0.20, 0.15]),
                (sensitivity_a, [100.0, 160.0, 230.0, 310.0], [0.18, 0.16, 0.12]),
            ]:
                run_dir.mkdir(parents=True, exist_ok=True)
                pd.DataFrame(
                    {
                        "date": pd.to_datetime(["2024-01-31", "2024-02-29", "2024-03-31", "2024-04-30"]),
                        "window": ["w1", "w1", "w1", "w1"],
                        "stitched_equity": equity,
                    }
                ).to_csv(run_dir / "oos_equity_stitched.csv", index=False)
                pd.DataFrame(
                    {
                        "window": ["w1", "w1", "w1"],
                        "test_total_return": window_returns,
                    }
                ).to_csv(run_dir / "portfolio_walk_forward_summary.csv", index=False)

            evidence.write_text(
                json.dumps({"status": "supports_candidate", "min_return_edge": 0.04, "min_sharpe_edge": 0.03}),
                encoding="utf-8",
            )

            result = run_allocator_promotion_review(
                baseline_dir=baseline,
                candidate_dir=candidate,
                sensitivity_dirs=[sensitivity_a],
                sensitivity_groups=["execution_cost_stress"],
                evidence_snapshots=[evidence],
                evidence_groups=["evidence_guardrail"],
                output_dir=output,
                min_return_edge=0.05,
                min_sharpe_edge=0.0,
                min_positive_window_ratio=0.80,
                min_sensitivity_support=2,
            )

            self.assertTrue(result.snapshot["candidate_passes_headline_gate"])
            self.assertEqual(int(result.snapshot["support_group_count"]), 2)
            self.assertIn("evidence_guardrail", result.snapshot["support_groups"])

    def test_allocator_promotion_evidence_support_flag_from_boolean_fields(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            baseline = root / "baseline"
            candidate = root / "candidate"
            sensitivity_a = root / "sensitivity_a"
            evidence = root / "evidence_support_bool.json"
            output = root / "promotion"

            for run_dir, equity, window_returns in [
                (baseline, [100.0, 120.0, 128.0, 135.0], [0.03, 0.02, 0.01]),
                (candidate, [100.0, 180.0, 265.0, 360.0], [0.25, 0.20, 0.15]),
                (sensitivity_a, [100.0, 90.0, 85.0, 80.0], [-0.02, -0.05, -0.02]),
            ]:
                run_dir.mkdir(parents=True, exist_ok=True)
                pd.DataFrame(
                    {
                        "date": pd.to_datetime(["2024-01-31", "2024-02-29", "2024-03-31", "2024-04-30"]),
                        "window": ["w1", "w1", "w1", "w1"],
                        "stitched_equity": equity,
                    }
                ).to_csv(run_dir / "oos_equity_stitched.csv", index=False)
                pd.DataFrame(
                    {
                        "window": ["w1", "w1", "w1"],
                        "test_total_return": window_returns,
                    }
                ).to_csv(run_dir / "portfolio_walk_forward_summary.csv", index=False)

            evidence.write_text(
                json.dumps(
                    {
                        "supports_candidate": True,
                        "all_cost_rates_support_candidate": True,
                        "min_return_edge": 0.04,
                        "min_sharpe_edge": 0.03,
                    }
                ),
                encoding="utf-8",
            )

            result = run_allocator_promotion_review(
                baseline_dir=baseline,
                candidate_dir=candidate,
                sensitivity_dirs=[sensitivity_a],
                sensitivity_groups=["execution_cost_stress"],
                evidence_snapshots=[evidence],
                evidence_groups=["evidence_guardrail"],
                output_dir=output,
                min_return_edge=0.05,
                min_sharpe_edge=0.0,
                min_positive_window_ratio=0.80,
                min_sensitivity_support=1,
            )

            self.assertEqual(result.decision, "promote_candidate")
            self.assertEqual(int(result.snapshot["support_group_count"]), 1)
            self.assertIn("evidence_guardrail", result.snapshot["support_groups"])

    def test_daily_check_writes_after_close_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            allocator = root / "allocator"
            phase2 = root / "phase2"
            promotion = root / "promotion"
            allocator.mkdir()
            phase2.mkdir()
            promotion.mkdir()
            selected_params = allocator / "selected_params.json"
            selected_params.write_text(
                json.dumps(
                    {
                        "weights": {
                            "risk_on": {"core": 0.8, "satellite": 0.2, "cash": 0.0},
                            "risk_off": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                            "crash": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                        },
                        "satellite_filter": {"enabled": False},
                        "regime_overrides": {
                            "ma_window": 60,
                            "risk_on_drop_threshold": -0.03,
                            "crash_drop_threshold": -0.07,
                        },
                    }
                ),
                encoding="utf-8",
            )
            pd.DataFrame(
                {
                    "window": ["w1"],
                    "train_start": ["20220101"],
                    "train_end": ["20221231"],
                    "test_start": ["20230101"],
                    "test_end": ["20230630"],
                    "selected_candidate": ["sat20_ma60_drop03_unfiltered"],
                    "selected_score": [0.2],
                    "selected_params_path": [str(selected_params)],
                    "test_total_return": [0.06],
                    "test_max_drawdown": [-0.05],
                    "test_sharpe": [0.9],
                    "test_average_core_weight": [0.92],
                    "test_average_satellite_weight": [0.08],
                    "test_risk_on_day_ratio": [0.5],
                    "test_risk_off_day_ratio": [0.5],
                    "test_crash_day_ratio": [0.0],
                }
            ).to_csv(allocator / "portfolio_walk_forward_summary.csv", index=False)
            pd.DataFrame(
                {
                    "date": pd.to_datetime(["2024-03-27", "2024-03-28", "2024-03-31"]),
                    "window": ["w1", "w1", "w1"],
                    "stitched_equity": [100.0, 104.0, 103.0],
                }
            ).to_csv(allocator / "oos_equity_stitched.csv", index=False)
            pd.DataFrame(
                {
                    "layer": ["core", "portfolio", "allocator"],
                    "status": ["complete", "complete", "complete"],
                    "total_return": [0.03, 0.02, 0.04],
                    "max_drawdown": [-0.06, -0.08, -0.04],
                    "sharpe": [0.4, 0.2, 1.1],
                }
            ).to_csv(phase2 / "phase2_components.csv", index=False)
            (promotion / "allocator_promotion_snapshot.json").write_text(
                json.dumps(
                    {
                        "decision": "promote_candidate",
                        "candidate_passes_headline_gate": True,
                        "sensitivity_run_support_count": 2,
                        "sensitivity_group_support_count": 1,
                        "evidence_support_count": 1,
                        "support_group_count": 2,
                        "sensitivity_support_count": 2,
                        "min_sensitivity_support": 2,
                        "return_edge": 0.062,
                        "sharpe_edge": 0.119,
                        "drawdown_change": 0.019,
                    }
                ),
                encoding="utf-8",
            )

            result = run_daily_model_check(
                project_root=root,
                output_dir=root / "daily",
                components_path=phase2 / "phase2_components.csv",
                allocator_dir=allocator,
                promotion_review_dir=promotion,
                as_of_date="2024-04-02",
                max_staleness_days=3,
                date_stamp=True,
            )

            self.assertTrue(result.report_path.exists())
            self.assertTrue(result.snapshot_path.exists())
            self.assertTrue(result.phase2_result.report_path.exists())
            self.assertEqual(result.output_dir.name, "daily_20240402")
            self.assertEqual(result.snapshot["data_freshness_status"], "fresh_enough")
            self.assertEqual(result.snapshot["days_since_latest"], 2)
            self.assertEqual(result.snapshot["phase2_posture"], "core_base_allocator_gated_satellite")
            self.assertEqual(result.snapshot["action_posture"], "review_core_base_allocator_gate")
            self.assertEqual(result.snapshot["promotion_decision"], "promote_candidate")
            self.assertEqual(int(result.snapshot["promotion_support_group_count"]), 2)
            self.assertEqual(int(result.snapshot["promotion_evidence_support_count"]), 1)
            self.assertEqual(int(result.snapshot["promotion_sensitivity_run_support_count"]), 2)
            self.assertEqual(int(result.snapshot["promotion_sensitivity_group_support_count"]), 1)
            self.assertAlmostEqual(float(result.snapshot["promotion_return_edge"]), 0.062)
            report_text = result.report_path.read_text(encoding="utf-8")
            self.assertIn("Daily Model Check", report_text)
            self.assertIn("Promotion decision", report_text)
            self.assertIn("Promotion independent support groups | 2 / 2", report_text)
            self.assertIn("Promotion evidence group support | 1", report_text)
            self.assertIn("Promotion sensitivity group support | 1", report_text)
            self.assertIn("Promotion sensitivity run support | 2", report_text)

    def test_paper_account_writes_ledger_and_audit(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            configs_dir = root / "configs"
            curves_dir = root / "curves"
            allocator = root / "allocator"
            backtests = root / "outputs" / "backtests"
            stock_cache = root / "data" / "processed" / "stocks"
            configs_dir.mkdir()
            curves_dir.mkdir()
            allocator.mkdir()
            (backtests / "core_model_run").mkdir(parents=True)
            (backtests / "satellite_model_run").mkdir(parents=True)
            stock_cache.mkdir(parents=True)
            dates = pd.date_range("2020-01-01", periods=8, freq="D")
            pd.DataFrame({"date": dates, "equity": [100, 101, 102, 104, 105, 106, 107, 108], "window": ["core_model_run"] * len(dates)}).to_csv(
                curves_dir / "core.csv",
                index=False,
            )
            pd.DataFrame({"date": dates, "equity": [100, 102, 104, 103, 107, 109, 108, 112], "window": ["satellite_model_run"] * len(dates)}).to_csv(
                curves_dir / "satellite.csv",
                index=False,
            )
            pd.DataFrame({"date": dates, "close": [10, 11, 12, 13, 14, 15, 16, 17]}).to_csv(
                curves_dir / "benchmark.csv",
                index=False,
            )
            pd.DataFrame(
                {
                    "date": ["2020-01-06", "2020-01-08"],
                    "code": ["000001", "000001"],
                    "name": ["One", "One"],
                    "open": [10.0, 12.0],
                    "high": [10.0, 12.0],
                    "low": [10.0, 12.0],
                    "close": [10.0, 12.0],
                    "volume": [1000, 1000],
                    "amount": [10000, 12000],
                }
            ).to_csv(stock_cache / "000001.csv", index=False)
            pd.DataFrame(
                {
                    "date": ["2020-01-06", "2020-01-08"],
                    "code": ["000002", "000002"],
                    "name": ["Two", "Two"],
                    "open": [20.0, 22.0],
                    "high": [20.0, 22.0],
                    "low": [20.0, 22.0],
                    "close": [20.0, 22.0],
                    "volume": [1000, 1000],
                    "amount": [20000, 22000],
                }
            ).to_csv(stock_cache / "000002.csv", index=False)
            pd.DataFrame(
                {
                    "date": ["2020-01-06", "2020-01-08"],
                    "code": ["000003", "000003"],
                    "name": ["Three", "Three"],
                    "open": [5.0, 6.0],
                    "high": [5.0, 6.0],
                    "low": [5.0, 6.0],
                    "close": [5.0, 6.0],
                    "volume": [1000, 1000],
                    "amount": [5000, 6000],
                }
            ).to_csv(stock_cache / "000003.csv", index=False)
            pd.DataFrame(
                {
                    "date": ["2020-01-06", "2020-01-06", "2020-01-07"],
                    "code": ["000001", "000002", "000002"],
                    "name": ["One", "Two", "Two"],
                    "side": ["BUY", "BUY", "SELL"],
                    "price": [10.0, 20.0, 21.0],
                    "quantity": [1000.0, 500.0, 100.0],
                    "gross_amount": [10000.0, 10000.0, 2100.0],
                    "fee": [3.0, 3.0, 2.0],
                    "realized_pnl": [None, None, 98.0],
                    "cash_after": [990000.0, 980000.0, 982098.0],
                    "exit_reason": ["", "", "rebalance"],
                }
            ).to_csv(backtests / "core_model_run" / "trades.csv", index=False)
            (backtests / "core_model_run" / "config_used.yaml").write_text(
                """
project:
  name: core_test_model
strategy:
  name: multi_factor
  max_positions: 15
  factor_rebalance_interval: 40
  min_score: 60
  allocation_buffer: 0.95
  factor_weights:
    momentum: 0.2
    trend: 0.2
    reversal: 0.25
    volatility: 0.25
    liquidity: 0.1
risk:
  enabled: true
  benchmark_code: '510300'
  benchmark_ma_window: 120
  benchmark_drop_window: 20
  benchmark_drop_threshold: -0.04
  benchmark_off_exposure: 0.2
  protection_drawdown: 0.2
""",
                encoding="utf-8",
            )
            pd.DataFrame(
                {
                    "date": ["2020-01-06"],
                    "code": ["000003"],
                    "name": ["Three"],
                    "side": ["BUY"],
                    "price": [5.0],
                    "quantity": [1000.0],
                    "gross_amount": [5000.0],
                    "fee": [1.5],
                    "realized_pnl": [None],
                    "cash_after": [995000.0],
                    "exit_reason": [""],
                }
            ).to_csv(backtests / "satellite_model_run" / "trades.csv", index=False)
            (backtests / "satellite_model_run" / "config_used.yaml").write_text(
                """
project:
  name: satellite_test_model
strategy:
  name: multi_factor
  max_positions: 5
  factor_rebalance_interval: 20
  min_score: 55
  allocation_buffer: 0.95
  factor_entry_filter_enabled: true
  market_breadth_enabled: true
  market_breadth_exposure_enabled: true
  factor_weights:
    momentum: 0.4
    trend: 0.3
    reversal: 0.1
    volatility: 0.1
    liquidity: 0.1
risk:
  enabled: true
  benchmark_code: '510300'
  benchmark_ma_window: 120
  benchmark_drop_window: 20
  benchmark_drop_threshold: -0.06
  benchmark_off_exposure: 0.0
  protection_drawdown: 0.3
""",
                encoding="utf-8",
            )
            trigger_signal_path = root / "signals_latest.csv"
            pd.DataFrame(
                {
                    "run_time": ["2020-01-08_160000"],
                    "code": ["000001"],
                    "name": ["One"],
                    "last": [12.0],
                    "pct": [1.5],
                    "signal_type": ["观察"],
                    "action": ["不操作"],
                    "reason": ["unit test trigger"],
                    "score": [68],
                    "score_level": ["B"],
                }
            ).to_csv(trigger_signal_path, index=False, encoding="utf-8")
            config_path = configs_dir / "portfolio.yaml"
            config_path.write_text(
                f"""
project:
  name: paper_test
  initial_cash: 1000000
  output_dir: {str(root / "outputs" / "portfolios")}
curves:
  core:
    path: {str(curves_dir / "core.csv")}
    equity_column: equity
  satellite:
    path: {str(curves_dir / "satellite.csv")}
    equity_column: equity
regime:
  benchmark_path: {str(curves_dir / "benchmark.csv")}
  benchmark_close_column: close
  ma_window: 2
  drop_window: 1
  risk_on_drop_threshold: -0.20
  crash_drop_threshold: -0.50
  default_regime: risk_off
weights:
  risk_on:
    core: 0.8
    satellite: 0.2
  risk_off:
    core: 1.0
    satellite: 0.0
  crash:
    core: 1.0
    satellite: 0.0
""",
                encoding="utf-8",
            )
            params1 = allocator / "params1.json"
            params1.write_text(
                json.dumps(
                    {
                        "weights": {
                            "risk_on": {"core": 0.8, "satellite": 0.2, "cash": 0.0},
                            "risk_off": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                            "crash": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                        },
                        "satellite_filter": {"enabled": False},
                        "regime_overrides": {
                            "ma_window": 2,
                            "risk_on_drop_threshold": -0.2,
                            "crash_drop_threshold": -0.5,
                        },
                    }
                ),
                encoding="utf-8",
            )
            params2 = allocator / "params2.json"
            params2.write_text(
                json.dumps(
                    {
                        "weights": {
                            "risk_on": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                            "risk_off": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                            "crash": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                        },
                        "satellite_filter": {"enabled": False},
                        "regime_overrides": {
                            "ma_window": 2,
                            "risk_on_drop_threshold": -0.2,
                            "crash_drop_threshold": -0.5,
                        },
                    }
                ),
                encoding="utf-8",
            )
            pd.DataFrame(
                {
                    "window": ["pf_01", "pf_02"],
                    "train_start": [20190101, 20190105],
                    "train_end": [20191231, 20200104],
                    "test_start": [20200101, 20200105],
                    "test_end": [20200104, 20200108],
                    "selected_candidate": ["sat20_test", "core_only"],
                    "selected_score": [0.4, 0.2],
                    "selected_params_path": ["params1.json", "params2.json"],
                }
            ).to_csv(allocator / "portfolio_walk_forward_summary.csv", index=False)

            result = run_paper_account(
                project_root=root,
                config_path=config_path,
                allocator_dir=allocator,
                output_dir=root / "paper",
                rebalance_cost_rate=0.001,
                trigger_signal_path=trigger_signal_path,
            )

            self.assertTrue(result.ledger_path.exists())
            self.assertTrue(result.audit_path.exists())
            self.assertTrue(result.target_holdings_path.exists())
            self.assertTrue(result.target_holdings_json_path.exists())
            self.assertTrue(result.target_holdings_report_path.exists())
            self.assertTrue(result.stock_targets_path.exists())
            self.assertTrue(result.stock_targets_json_path.exists())
            self.assertTrue(result.stock_targets_report_path.exists())
            self.assertTrue(result.stock_target_review_path.exists())
            self.assertTrue(result.stock_target_review_json_path.exists())
            self.assertTrue(result.stock_target_review_report_path.exists())
            self.assertTrue(result.stock_target_review_notes_path.exists())
            self.assertTrue(result.stock_target_review_notes_snapshot_path.exists())
            self.assertTrue(result.stock_target_review_actions_path.exists())
            self.assertTrue(result.stock_target_review_actions_json_path.exists())
            self.assertTrue(result.stock_target_review_actions_report_path.exists())
            self.assertTrue(result.stock_target_review_assistant_path.exists())
            self.assertTrue(result.stock_target_review_assistant_json_path.exists())
            self.assertTrue(result.stock_target_review_assistant_report_path.exists())
            self.assertTrue(result.stock_target_review_decision_template_path.exists())
            self.assertTrue(result.stock_target_review_decision_template_json_path.exists())
            self.assertTrue(result.stock_target_review_decision_template_report_path.exists())
            self.assertTrue(result.stock_target_review_decision_template_xlsx_path.exists())
            self.assertTrue(result.stock_target_review_outcomes_path.exists())
            self.assertTrue(result.stock_target_review_outcomes_json_path.exists())
            self.assertTrue(result.stock_target_review_outcomes_report_path.exists())
            self.assertTrue(result.stock_target_review_outcomes_history_path.exists())
            self.assertTrue(result.stock_target_review_outcomes_history_snapshot_path.exists())
            self.assertTrue(result.stock_target_review_outcomes_history_json_path.exists())
            self.assertTrue(result.stock_target_review_outcomes_history_report_path.exists())
            self.assertTrue(result.stock_target_review_outcome_analysis_path.exists())
            self.assertTrue(result.stock_target_review_outcome_analysis_json_path.exists())
            self.assertTrue(result.stock_target_review_outcome_analysis_report_path.exists())
            self.assertTrue(result.stock_target_review_outcome_calendar_path.exists())
            self.assertTrue(result.stock_target_review_outcome_calendar_json_path.exists())
            self.assertTrue(result.stock_target_review_outcome_calendar_report_path.exists())
            self.assertTrue(result.stock_target_review_outcome_due_path.exists())
            self.assertTrue(result.stock_target_review_outcome_due_json_path.exists())
            self.assertTrue(result.stock_target_review_outcome_due_report_path.exists())
            self.assertTrue(result.report_path.exists())
            self.assertEqual(len(result.ledger), 8)
            self.assertEqual(len(result.target_holdings), 3)
            self.assertEqual(len(result.stock_targets), 2)
            self.assertEqual(len(result.stock_target_review), 3)
            self.assertEqual(set(result.target_holdings["layer"]), {"core", "satellite", "cash"})
            self.assertAlmostEqual(float(result.target_holdings["target_weight"].sum()), 1.0)
            self.assertLess(float(result.stock_targets["portfolio_target_weight"].sum()), 1.0)
            self.assertEqual(int(result.stock_targets_payload["active_stock_target_count"]), 1)
            self.assertEqual(int(result.stock_targets_payload["suppressed_stock_count"]), 1)
            self.assertEqual(result.stock_targets_payload["trigger_signal_status"], "ok")
            self.assertEqual(int(result.stock_targets_payload["trigger_match_count"]), 0)
            self.assertEqual(int(result.stock_targets_payload["review_required_excluded_count"]), 1)
            self.assertEqual(result.stock_targets_payload["review_required_excluded_codes"], ["000001"])
            self.assertEqual(int(result.stock_target_review_payload["review_required_count"]), 0)
            self.assertEqual(int(result.stock_target_review_payload["excluded_review_required_count"]), 1)
            self.assertGreaterEqual(int(result.stock_target_review_payload["monitor_count"]), 1)
            self.assertEqual(int(result.stock_target_review_payload["manual_note_count"]), 0)
            self.assertEqual(int(result.stock_target_review_payload["review_required_unreviewed_count"]), 0)
            self.assertEqual(int(result.stock_target_review_payload["manual_pending_count"]), 3)
            self.assertEqual(int(result.stock_target_review_payload["manual_watch_count"]), 0)
            self.assertEqual(int(result.stock_target_review_actions_payload["action_count"]), 0)
            self.assertEqual(int(result.stock_target_review_actions_payload["review_required_pending_count"]), 0)
            self.assertEqual(int(result.stock_target_review_assistant_payload["assistant_row_count"]), 3)
            self.assertEqual(int(result.stock_target_review_assistant_payload["review_required_pending_count"]), 0)
            self.assertEqual(int(result.stock_target_review_decision_template_payload["decision_template_row_count"]), 3)
            self.assertEqual(int(result.stock_target_review_decision_template_payload["blank_manual_status_count"]), 3)
            self.assertEqual(int(result.stock_target_review_outcomes_payload["outcome_row_count"]), 3)
            self.assertEqual(int(result.stock_target_review_outcomes_history_payload["history_row_count"]), 3)
            self.assertEqual(int(result.stock_target_review_outcomes_history_payload["history_updated_row_count"]), 3)
            self.assertGreaterEqual(int(result.stock_target_review_outcome_analysis_payload["analysis_row_count"]), 1)
            self.assertEqual(int(result.stock_target_review_outcome_analysis_payload["min_evaluable"]), 20)
            self.assertEqual(int(result.stock_target_review_outcome_analysis_payload["min_group_evaluable"]), 5)
            self.assertEqual(int(result.stock_target_review_outcome_calendar_payload["calendar_row_count"]), 4)
            self.assertGreaterEqual(int(result.stock_target_review_outcome_calendar_payload["calendar_pending_count"]), 0)
            self.assertGreaterEqual(int(result.stock_target_review_outcome_due_payload["due_row_count"]), 0)
            self.assertGreaterEqual(int(result.stock_target_review_outcome_due_payload["due_pending_count"]), 0)
            self.assertIn(
                result.stock_target_review_outcome_analysis_payload["analysis_status"],
                {"waiting_for_evaluable_returns", "sample_insufficient", "ready_for_review"},
            )
            self.assertEqual(result.stock_target_review.iloc[0]["code"], "000001")
            self.assertEqual(result.stock_target_review.iloc[0]["review_bucket"], "trigger_review")
            self.assertIn("manual_status", result.stock_target_review.columns)
            self.assertIn("manual_status_normalized", result.stock_target_review.columns)
            self.assertIn("manual_review_state", result.stock_target_review.columns)
            self.assertIn("notes_status", result.stock_target_review.columns)
            self.assertIn("suppressed_by_layer_weight", set(result.stock_targets["target_action"]))
            self.assertIn("selection_explanation", result.stock_targets.columns)
            self.assertIn("risk_filter_status", result.stock_targets.columns)
            self.assertIn("trigger_monitor_status", result.stock_targets.columns)
            self.assertIn("unrealized_return", result.stock_targets.columns)
            self.assertIn("multi_factor", ";".join(result.stock_targets["source_strategy_profile"].astype(str)))
            self.assertNotIn("000001", set(result.stock_targets["code"]))
            matched_review = result.stock_target_review[result.stock_target_review["code"] == "000001"].iloc[0]
            self.assertEqual(matched_review["review_stage"], "excluded")
            self.assertTrue(bool(matched_review["observation_excluded"]))
            suppressed_stock = result.stock_targets[result.stock_targets["target_action"] == "suppressed_by_layer_weight"].iloc[0]
            self.assertEqual(suppressed_stock["risk_filter_status"], "blocked_by_layer_weight")
            self.assertEqual(result.target_holdings_payload["broker_action"], "none")
            self.assertEqual(result.stock_targets_payload["broker_action"], "none")
            self.assertEqual(result.target_holdings_payload["target_holding_count"], 3)
            self.assertEqual(result.metrics["latest_candidate"], "core_only")
            self.assertEqual(result.metrics["target_holdings_path"], str(root / "paper" / "target_holdings.csv"))
            self.assertEqual(result.metrics["stock_targets_path"], str(root / "paper" / "stock_targets.csv"))
            self.assertEqual(result.metrics["stock_target_trigger_match_count"], 0)
            self.assertEqual(result.metrics["stock_target_review_required_count"], 0)
            self.assertEqual(result.metrics["stock_target_review_excluded_count"], 1)
            self.assertEqual(result.metrics["stock_target_review_required_unreviewed_count"], 0)
            self.assertEqual(result.metrics["stock_target_review_manual_pending_count"], 3)
            self.assertEqual(result.metrics["stock_target_review_action_count"], 0)
            self.assertEqual(result.metrics["stock_target_review_assistant_count"], 3)
            self.assertEqual(result.metrics["stock_target_review_decision_template_count"], 3)
            self.assertEqual(result.metrics["stock_target_review_decision_template_blank_status_count"], 3)
            self.assertEqual(
                result.metrics["stock_target_review_decision_template_xlsx_path"],
                str(root / "paper" / "stock_target_review_decision_template.xlsx"),
            )
            self.assertEqual(result.metrics["stock_target_review_outcome_row_count"], 3)
            self.assertEqual(result.metrics["stock_target_review_outcomes_history_row_count"], 3)
            self.assertGreaterEqual(result.metrics["stock_target_review_outcome_analysis_row_count"], 1)
            self.assertEqual(result.metrics["stock_target_review_outcome_analysis_min_evaluable"], 20)
            self.assertEqual(result.metrics["stock_target_review_outcome_analysis_min_group_evaluable"], 5)
            self.assertEqual(result.metrics["stock_target_review_outcome_calendar_row_count"], 4)
            self.assertIn("stock_target_review_outcome_calendar_next_action_date", result.metrics)
            self.assertIn("stock_target_review_outcome_due_row_count", result.metrics)
            self.assertIn("stock_target_review_outcome_due_next_date", result.metrics)
            self.assertIn("stock_target_review_outcome_maturity_next_1d_date", result.metrics)
            self.assertEqual(result.metrics["stock_target_review_notes_path"], str(root / "outputs" / "research" / "stock_target_review_notes.csv"))
            self.assertEqual(result.metrics["stock_target_review_outcomes_history_path"], str(root / "outputs" / "research" / "stock_target_review_outcomes_history.csv"))
            self.assertGreater(result.metrics["total_estimated_fee"], 0.0)
            event_text = ";".join(result.audit["event_type"].astype(str))
            self.assertIn("window_switch", event_text)
            self.assertIn("candidate_change", event_text)
            self.assertIn("estimated_fee", result.ledger.columns)
            self.assertIn("Paper Account Ledger", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Paper Target Holdings", result.target_holdings_report_path.read_text(encoding="utf-8"))
            stock_report_text = result.stock_targets_report_path.read_text(encoding="utf-8")
            self.assertIn("Paper Stock Targets", stock_report_text)
            self.assertIn("Trigger monitor", stock_report_text)
            self.assertIn("Why selected", stock_report_text)
            review_report_text = result.stock_target_review_report_path.read_text(encoding="utf-8")
            self.assertIn("Paper Stock Target Review", review_report_text)
            self.assertIn("trigger_review", review_report_text)
            self.assertIn("Persistent notes CSV", review_report_text)
            action_report_text = result.stock_target_review_actions_report_path.read_text(encoding="utf-8")
            self.assertIn("Paper Stock Target Review Actions", action_report_text)
            self.assertIn("No open stock-target review actions", action_report_text)
            assistant_report_text = result.stock_target_review_assistant_report_path.read_text(encoding="utf-8")
            self.assertIn("Paper Stock Target Review Assistant", assistant_report_text)
            self.assertIn("Manual Status Guide", assistant_report_text)
            self.assertIn("does not connect to brokers", assistant_report_text)
            decision_template_text = result.stock_target_review_decision_template_report_path.read_text(encoding="utf-8")
            self.assertIn("Paper Stock Target Review Decision Template", decision_template_text)
            self.assertIn("manual_status_to_fill", decision_template_text)
            self.assertIn("Highlighted fill-in workbook", decision_template_text)
            self.assertIn("does not set manual_status", decision_template_text)
            outcome_report_text = result.stock_target_review_outcomes_report_path.read_text(encoding="utf-8")
            self.assertIn("Paper Stock Target Review Outcomes", outcome_report_text)
            self.assertIn("Pending horizons are not guessed", outcome_report_text)
            outcome_history_report_text = result.stock_target_review_outcomes_history_report_path.read_text(encoding="utf-8")
            self.assertIn("Paper Stock Target Review Outcome History", outcome_history_report_text)
            self.assertIn("cumulative research-only outcome tracker", outcome_history_report_text)
            outcome_analysis_report_text = result.stock_target_review_outcome_analysis_report_path.read_text(encoding="utf-8")
            self.assertIn("Paper Stock Target Review Outcome Analysis", outcome_analysis_report_text)
            self.assertIn("historical effectiveness view", outcome_analysis_report_text)
            self.assertIn("Readiness Gates", outcome_analysis_report_text)
            self.assertIn("Maturity Forecast", outcome_analysis_report_text)
            outcome_calendar_report_text = result.stock_target_review_outcome_calendar_report_path.read_text(encoding="utf-8")
            self.assertIn("Paper Stock Target Review Outcome Calendar", outcome_calendar_report_text)
            self.assertIn("research-only outcome maturity calendar", outcome_calendar_report_text)
            outcome_due_report_text = result.stock_target_review_outcome_due_report_path.read_text(encoding="utf-8")
            self.assertIn("Paper Stock Target Review Outcome Due Queue", outcome_due_report_text)
            self.assertIn("research-only due queue", outcome_due_report_text)

    def test_live_shadow_writes_manual_review_plan(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            holdings = root / "holdings.csv"
            targets = root / "stock_targets.csv"
            prices = root / "prices.csv"

            pd.DataFrame(
                [
                    {"code": "000001", "name": "Alpha", "quantity": 100, "current_price": 10.0},
                    {"code": "000002", "name": "Beta", "quantity": 200, "current_price": 20.0},
                ]
            ).to_csv(holdings, index=False)
            pd.DataFrame(
                [
                    {"code": "000001", "name": "Alpha", "portfolio_target_weight": 0.20, "close_price": 10.0},
                    {"code": "000003", "name": "Gamma", "portfolio_target_weight": 0.30, "close_price": 5.0},
                ]
            ).to_csv(targets, index=False)
            pd.DataFrame(
                [
                    {"code": "000001", "price": 10.0},
                    {"code": "000002", "price": 20.0},
                    {"code": "000003", "price": 5.0},
                ]
            ).to_csv(prices, index=False)

            result = run_live_shadow(
                project_root=root,
                holdings_file=holdings,
                targets_file=targets,
                prices_file=prices,
                output_dir=root / "live_shadow",
                cash=5000.0,
                as_of_date="2026-06-12",
                lot_size=100,
                min_trade_value=0.0,
                max_position_weight=1.0,
            )

            orders = pd.read_csv(result.orders_path, dtype={"code": str})
            reconcile = pd.read_csv(result.reconcile_path, dtype={"code": str})
            sides = dict(zip(orders["code"], orders["side"]))
            quantities = dict(zip(orders["code"], orders["order_quantity"]))

            self.assertEqual(result.snapshot["trade_plan_status"], "manual_review_only")
            self.assertTrue(result.snapshot["research_only"])
            self.assertEqual(result.snapshot["broker_action"], "none")
            self.assertEqual(result.snapshot["order_count"], 3)
            self.assertEqual(sides["000001"], "BUY")
            self.assertEqual(int(quantities["000001"]), 100)
            self.assertEqual(sides["000002"], "SELL")
            self.assertEqual(int(quantities["000002"]), 200)
            self.assertEqual(sides["000003"], "BUY")
            self.assertEqual(int(quantities["000003"]), 600)
            self.assertIn("manual_review_only", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("000002", set(reconcile["code"]))

    def test_live_shadow_caps_available_quantity_and_recomputes_inconsistent_market_value(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            holdings = root / "holdings.xlsx"
            targets = root / "stock_targets.csv"

            with pd.ExcelWriter(holdings) as writer:
                pd.DataFrame(
                    [
                        {
                            "code": "000895",
                            "name": "Alpha",
                            "quantity": 1000,
                            "available_quantity": 10000,
                            "current_price": 24.0,
                            "market_value": 240000.0,
                        }
                    ]
                ).to_excel(writer, sheet_name="holdings_to_fill", index=False)
            pd.DataFrame(columns=["code", "name", "portfolio_target_weight", "close_price"]).to_csv(targets, index=False)

            result = run_live_shadow(
                project_root=root,
                holdings_file=holdings,
                targets_file=targets,
                output_dir=root / "live_shadow",
                cash=0.0,
                as_of_date="2026-06-15",
                lot_size=100,
                min_trade_value=0.0,
                max_position_weight=1.0,
            )

            reconcile = pd.read_csv(result.reconcile_path, dtype={"code": str})
            row = reconcile.iloc[0]
            self.assertEqual(result.snapshot["current_equity"], 24000.0)
            self.assertEqual(row["code"], "000895")
            self.assertEqual(row["available_quantity"], 1000.0)
            self.assertEqual(row["current_value"], 24000.0)
            self.assertEqual(row["price"], 24.0)

    def test_live_shadow_prefers_holding_price_over_target_price_without_prices_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            holdings = root / "holdings.csv"
            targets = root / "stock_targets.csv"

            pd.DataFrame(
                [
                    {"code": "000895", "name": "Alpha", "quantity": 1000, "current_price": 24.0},
                ]
            ).to_csv(holdings, index=False)
            pd.DataFrame(
                [
                    {"code": "000895", "name": "Alpha", "portfolio_target_weight": 0.50, "close_price": 30.0},
                ]
            ).to_csv(targets, index=False)

            result = run_live_shadow(
                project_root=root,
                holdings_file=holdings,
                targets_file=targets,
                output_dir=root / "live_shadow",
                cash=24000.0,
                as_of_date="2026-06-15",
                lot_size=100,
                min_trade_value=0.0,
                max_position_weight=1.0,
            )

            reconcile = pd.read_csv(result.reconcile_path, dtype={"code": str})
            row = reconcile.iloc[0]
            self.assertEqual(row["price"], 24.0)
            self.assertEqual(row["target_quantity"], 1000)
            self.assertEqual(row["order_quantity"], 0)

    def test_live_shadow_import_template_writes_highlighted_fill_workbook(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            targets = root / "stock_targets.csv"
            pd.DataFrame(
                [
                    {"code": "000001", "name": "Alpha", "portfolio_target_weight": 0.20, "close_price": 10.0},
                    {"code": "000003", "name": "Gamma", "portfolio_target_weight": 0.30, "close_price": 5.0},
                ]
            ).to_csv(targets, index=False)

            result = run_live_shadow_import_template(
                project_root=root,
                targets_file=targets,
                output_dir=root / "live_shadow_import",
                as_of_date="2026-06-15",
                blank_rows=2,
            )

            holdings = pd.read_csv(result.holdings_template_path, dtype={"code": str})
            cash = pd.read_csv(result.cash_template_path)
            workbook = load_workbook(result.workbook_path)
            sheet = workbook["holdings_to_fill"]
            headers = [sheet.cell(row=1, column=column).value for column in range(1, sheet.max_column + 1)]
            quantity_column = headers.index("持仓数量") + 1
            price_column = headers.index("当前价格") + 1

            self.assertEqual(result.snapshot["broker_action"], "none")
            self.assertTrue(result.snapshot["research_only"])
            self.assertEqual(result.snapshot["target_row_count"], 2)
            self.assertEqual(result.snapshot["blank_row_count"], 2)
            self.assertTrue(result.workbook_path.exists())
            self.assertTrue(result.report_path.exists())
            self.assertEqual(list(holdings["code"].head(2)), ["000001", "000003"])
            self.assertEqual(len(holdings), 4)
            self.assertIn("quantity", holdings.columns)
            self.assertIn("current_price", holdings.columns)
            self.assertIn("cash", cash.columns)
            self.assertIn("cash_to_fill", workbook.sheetnames)
            self.assertEqual(sheet.cell(row=1, column=quantity_column).fill.fgColor.rgb, "FFFFC000")
            self.assertEqual(sheet.cell(row=2, column=quantity_column).fill.fgColor.rgb, "FFFFE699")
            self.assertEqual(sheet.cell(row=2, column=price_column).fill.fgColor.rgb, "FFFFE699")
            report_text = result.report_path.read_text(encoding="utf-8")
            self.assertIn("| 项目 | 值 |", report_text)
            self.assertIn("`股票代码`, `持仓数量`, and `当前价格`", report_text)
            self.assertIn("does not connect to brokers", report_text)

    def test_live_shadow_reads_chinese_import_template_headers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            holdings = root / "holdings.xlsx"
            targets = root / "stock_targets.csv"

            with pd.ExcelWriter(holdings) as writer:
                pd.DataFrame(
                    [
                        {
                            "股票代码": "000895",
                            "股票名称": "Alpha",
                            "持仓数量": 1000,
                            "可卖数量": 800,
                            "当前价格": 24.0,
                            "市值": 999999.0,
                        }
                    ]
                ).to_excel(writer, sheet_name="holdings_to_fill", index=False)
            pd.DataFrame(columns=["code", "name", "portfolio_target_weight", "close_price"]).to_csv(targets, index=False)

            result = run_live_shadow(
                project_root=root,
                holdings_file=holdings,
                targets_file=targets,
                output_dir=root / "live_shadow",
                cash=0.0,
                as_of_date="2026-06-15",
                lot_size=100,
                min_trade_value=0.0,
                max_position_weight=1.0,
            )

            reconcile = pd.read_csv(result.reconcile_path, dtype={"code": str})
            row = reconcile.iloc[0]
            self.assertEqual(result.snapshot["current_equity"], 24000.0)
            self.assertEqual(row["code"], "000895")
            self.assertEqual(row["available_quantity"], 800.0)
            self.assertEqual(row["current_value"], 24000.0)
            self.assertEqual(row["price"], 24.0)

    def test_live_shadow_review_queue_flags_tracking_excluded_orders(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            orders = root / "live_shadow_orders.csv"
            targets = root / "stock_targets.csv"
            pd.DataFrame(
                [
                    {
                        "code": "000001",
                        "name": "Alpha",
                        "side": "BUY",
                        "action": "buy_to_target",
                        "current_quantity": 100,
                        "target_quantity": 200,
                        "order_quantity": 100,
                        "price": 10.0,
                        "estimated_amount": 1000.0,
                    },
                    {
                        "code": "600919",
                        "name": "LargeBank",
                        "side": "SELL",
                        "action": "sell_to_target",
                        "current_quantity": 1000,
                        "target_quantity": 500,
                        "order_quantity": 500,
                        "price": 11.0,
                        "estimated_amount": 5500.0,
                    },
                ]
            ).to_csv(orders, index=False)
            pd.DataFrame(
                [
                    {
                        "code": "000001",
                        "name": "Alpha",
                        "tracking_rule_status": "tracking_allowed",
                        "tracking_excluded": False,
                        "market_cap_yi": 120.0,
                    },
                    {
                        "code": "600919",
                        "name": "LargeBank",
                        "tracking_rule_status": "excluded_large_market_cap",
                        "tracking_excluded": True,
                        "tracking_exclusion_reason": "market_cap_yi 2181.97 > max_tracking_market_cap_yi 1500.00",
                        "market_cap_yi": 2181.97,
                    },
                ]
            ).to_csv(targets, index=False)

            result = run_live_shadow_review_queue(
                project_root=root,
                orders_file=orders,
                targets_file=targets,
                output_dir=root / "live_shadow_review",
                as_of_date="2026-06-15",
            )

            review = pd.read_csv(result.review_queue_path, dtype={"code": str})
            gates = dict(zip(review["code"], review["review_gate"]))
            workbook = load_workbook(result.workbook_path)
            sheet = workbook["review_queue"]
            headers = [sheet.cell(row=1, column=column).value for column in range(1, sheet.max_column + 1)]
            status_column = headers.index("复核状态（请填写）") + 1

            self.assertEqual(result.snapshot["status"], "blocked_tracking_rule_review")
            self.assertEqual(result.snapshot["review_row_count"], 2)
            self.assertEqual(result.snapshot["tracking_blocked_count"], 1)
            self.assertIn("review_status_to_fill", review.columns)
            self.assertEqual(gates["000001"], "manual_review_required")
            self.assertEqual(gates["600919"], "blocked_by_tracking_rule")
            self.assertTrue(result.workbook_path.exists())
            self.assertEqual(sheet.cell(row=1, column=status_column).fill.fgColor.rgb, "FFFFC000")
            self.assertEqual(sheet.cell(row=2, column=status_column).fill.fgColor.rgb, "FFFFE699")
            report_text = result.report_path.read_text(encoding="utf-8")
            self.assertIn("| 股票代码 | 股票名称 | 方向 | 数量 | 金额 | 复核门禁 | 跟踪原因 |", report_text)
            self.assertIn("does not connect to brokers", report_text)

    def test_live_shadow_review_decisions_apply_filled_chinese_queue(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            review_file = root / "review_queue.xlsx"
            with pd.ExcelWriter(review_file) as writer:
                pd.DataFrame(
                    [
                        {
                            "日期": "2026-06-15",
                            "股票代码": "000001",
                            "股票名称": "Alpha",
                            "方向": "BUY",
                            "差异数量": 100,
                            "估算金额": 1000.0,
                            "复核门禁": "manual_review_required",
                            "复核状态（请填写）": "reviewed_for_manual_consideration",
                            "备注（请填写）": "ok",
                        },
                        {
                            "日期": "2026-06-15",
                            "股票代码": "600919",
                            "股票名称": "LargeBank",
                            "方向": "BUY",
                            "差异数量": 300,
                            "估算金额": 3543.0,
                            "复核门禁": "blocked_by_tracking_rule",
                            "复核状态（请填写）": "reviewed_for_manual_consideration",
                            "暂停跟踪原因": "market_cap_yi 2181.97 > max_tracking_market_cap_yi 1500.00",
                        },
                        {
                            "日期": "2026-06-15",
                            "股票代码": "000003",
                            "股票名称": "Gamma",
                            "方向": "SELL",
                            "差异数量": 200,
                            "估算金额": 2000.0,
                            "复核门禁": "manual_review_required",
                            "复核状态（请填写）": "watch_only",
                        },
                        {
                            "日期": "2026-06-15",
                            "股票代码": "000004",
                            "股票名称": "Delta",
                            "方向": "BUY",
                            "差异数量": 500,
                            "估算金额": 5000.0,
                            "复核门禁": "manual_review_required",
                            "复核状态（请填写）": "needs_data",
                        },
                    ]
                ).to_excel(writer, sheet_name="review_queue", index=False)

            result = run_live_shadow_review_decisions(
                project_root=root,
                review_file=review_file,
                output_dir=root / "review_decisions",
                as_of_date="2026-06-15",
            )

            decisions = pd.read_csv(result.decisions_path, dtype={"code": str})
            decision_by_code = dict(zip(decisions["code"], decisions["review_decision"]))
            self.assertEqual(result.snapshot["status"], "review_completed_with_tracking_blockers")
            self.assertEqual(result.snapshot["review_row_count"], 4)
            self.assertEqual(result.snapshot["filled_review_status_count"], 4)
            self.assertEqual(result.snapshot["invalid_status_count"], 0)
            self.assertEqual(result.snapshot["blank_status_count"], 0)
            self.assertEqual(result.snapshot["tracking_blocked_count"], 1)
            self.assertEqual(result.snapshot["manual_considered_count"], 1)
            self.assertEqual(result.snapshot["watch_only_count"], 1)
            self.assertEqual(result.snapshot["needs_data_count"], 1)
            self.assertEqual(decision_by_code["000001"], "manual_considered")
            self.assertEqual(decision_by_code["600919"], "blocked_by_tracking_rule")
            self.assertEqual(decision_by_code["000003"], "watch_only")
            self.assertEqual(decision_by_code["000004"], "needs_data")
            self.assertTrue(result.report_path.exists())
            self.assertIn("does not connect to brokers", result.report_path.read_text(encoding="utf-8"))

    def test_live_shadow_review_decisions_flags_blank_and_invalid_statuses(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            review_file = root / "review_queue.csv"
            pd.DataFrame(
                [
                    {
                        "code": "000001",
                        "name": "Alpha",
                        "side": "BUY",
                        "order_quantity": 100,
                        "estimated_amount": 1000.0,
                        "review_gate": "manual_review_required",
                        "review_status_to_fill": "",
                    },
                    {
                        "code": "000002",
                        "name": "Beta",
                        "side": "SELL",
                        "order_quantity": 200,
                        "estimated_amount": 2000.0,
                        "review_gate": "manual_review_required",
                        "review_status_to_fill": "approve",
                    },
                ]
            ).to_csv(review_file, index=False)

            result = run_live_shadow_review_decisions(
                project_root=root,
                review_file=review_file,
                output_dir=root / "review_decisions",
                as_of_date="2026-06-15",
            )

            decisions = pd.read_csv(result.decisions_path, dtype={"code": str})
            decision_by_code = dict(zip(decisions["code"], decisions["review_decision"]))
            self.assertEqual(result.snapshot["status"], "invalid_review_status")
            self.assertEqual(result.snapshot["blank_status_count"], 1)
            self.assertEqual(result.snapshot["invalid_status_count"], 1)
            self.assertEqual(decision_by_code["000001"], "incomplete_review")
            self.assertEqual(decision_by_code["000002"], "invalid_review_status")

    def test_paper_account_applies_source_selection_nested_allocation_and_source_path(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            configs_dir = root / "configs"
            curves_dir = root / "curves"
            allocator = root / "allocator"
            configs_dir.mkdir()
            curves_dir.mkdir()
            allocator.mkdir()
            dates = pd.date_range("2020-01-01", periods=4, freq="D")
            pd.DataFrame({"date": dates, "equity": [100, 100, 100, 100]}).to_csv(curves_dir / "core.csv", index=False)
            pd.DataFrame({"date": dates, "equity": [100, 100, 100, 100]}).to_csv(
                curves_dir / "default_satellite.csv",
                index=False,
            )
            pd.DataFrame({"date": dates, "equity": [100, 120, 150, 200]}).to_csv(
                curves_dir / "selected_source.csv",
                index=False,
            )
            pd.DataFrame({"date": dates, "close": [10, 11, 12, 13]}).to_csv(curves_dir / "benchmark.csv", index=False)
            config_path = configs_dir / "portfolio.yaml"
            config_path.write_text(
                f"""
project:
  name: source_selection_paper_test
  initial_cash: 1000000
  output_dir: {str(root / "outputs" / "portfolios")}
curves:
  core:
    path: {str(curves_dir / "core.csv")}
    equity_column: equity
  satellite:
    path: {str(curves_dir / "default_satellite.csv")}
    equity_column: equity
regime:
  benchmark_path: {str(curves_dir / "benchmark.csv")}
  benchmark_close_column: close
  ma_window: 2
  drop_window: 1
  risk_on_drop_threshold: -0.20
  crash_drop_threshold: -0.50
  default_regime: risk_on
weights:
  risk_on:
    core: 1.0
    satellite: 0.0
  risk_off:
    core: 1.0
    satellite: 0.0
  crash:
    core: 1.0
    satellite: 0.0
""",
                encoding="utf-8",
            )
            params = allocator / "params_source_selection.json"
            params.write_text(
                json.dumps(
                    {
                        "source_name": "selected_source",
                        "source_path": str(curves_dir / "selected_source.csv"),
                        "allocation": {
                            "weights": {
                                "risk_on": {"core": 0.0, "satellite": 1.0, "cash": 0.0},
                                "risk_off": {"core": 0.0, "satellite": 1.0, "cash": 0.0},
                                "crash": {"core": 0.0, "satellite": 1.0, "cash": 0.0},
                            },
                            "satellite_filter": {"enabled": False},
                            "regime_overrides": {
                                "ma_window": 2,
                                "risk_on_drop_threshold": -0.2,
                                "crash_drop_threshold": -0.5,
                            },
                        },
                    }
                ),
                encoding="utf-8",
            )
            pd.DataFrame(
                {
                    "window": ["pf_01"],
                    "train_start": [20190101],
                    "train_end": [20191231],
                    "test_start": [20200101],
                    "test_end": [20200104],
                    "selected_candidate": ["selected_source__sat100"],
                    "selected_score": [1.0],
                    "selected_params_path": ["params_source_selection.json"],
                }
            ).to_csv(allocator / "portfolio_walk_forward_summary.csv", index=False)

            result = run_paper_account(
                project_root=root,
                config_path=config_path,
                allocator_dir=allocator,
                output_dir=root / "paper",
            )

            self.assertGreater(float(result.metrics["total_return"]), 0.90)
            self.assertEqual(result.metrics["latest_candidate"], "selected_source__sat100")
            self.assertAlmostEqual(float(result.metrics["average_satellite_weight"]), 1.0)
            self.assertAlmostEqual(float(result.metrics["average_core_weight"]), 0.0)

    def test_paper_account_auto_falls_back_to_stitched_equity_source_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            configs_dir = root / "configs"
            curves_dir = root / "curves"
            allocator = root / "allocator"
            configs_dir.mkdir()
            curves_dir.mkdir()
            allocator.mkdir()
            dates = pd.date_range("2020-01-01", periods=4, freq="D")
            pd.DataFrame({"date": dates, "equity": [100, 100, 100, 100]}).to_csv(curves_dir / "core.csv", index=False)
            pd.DataFrame({"date": dates, "equity": [100, 100, 100, 100]}).to_csv(
                curves_dir / "default_satellite.csv",
                index=False,
            )
            pd.DataFrame({"date": dates, "stitched_equity": [100, 150, 180, 220]}).to_csv(
                curves_dir / "stitched_source.csv",
                index=False,
            )
            pd.DataFrame({"date": dates, "close": [10, 11, 12, 13]}).to_csv(curves_dir / "benchmark.csv", index=False)
            config_path = configs_dir / "portfolio.yaml"
            config_path.write_text(
                f"""
project:
  name: source_selection_stitched_column_test
  initial_cash: 1000000
  output_dir: {str(root / "outputs" / "portfolios")}
curves:
  core:
    path: {str(curves_dir / "core.csv")}
    equity_column: equity
  satellite:
    path: {str(curves_dir / "default_satellite.csv")}
    equity_column: equity
regime:
  benchmark_path: {str(curves_dir / "benchmark.csv")}
  benchmark_close_column: close
  ma_window: 2
  drop_window: 1
  risk_on_drop_threshold: -0.20
  crash_drop_threshold: -0.50
  default_regime: risk_on
weights:
  risk_on:
    core: 1.0
    satellite: 0.0
  risk_off:
    core: 1.0
    satellite: 0.0
  crash:
    core: 1.0
    satellite: 0.0
""",
                encoding="utf-8",
            )
            params = allocator / "params_source_selection_stitched.csv"
            params.write_text(
                json.dumps(
                    {
                        "source_name": "stitched_source",
                        "source_path": str(curves_dir / "stitched_source.csv"),
                        "allocation": {
                            "weights": {
                                "risk_on": {"core": 0.0, "satellite": 1.0, "cash": 0.0},
                                "risk_off": {"core": 0.0, "satellite": 1.0, "cash": 0.0},
                                "crash": {"core": 0.0, "satellite": 1.0, "cash": 0.0},
                            },
                            "satellite_filter": {"enabled": False},
                            "regime_overrides": {
                                "ma_window": 2,
                                "risk_on_drop_threshold": -0.2,
                                "crash_drop_threshold": -0.5,
                            },
                        },
                    }
                ),
                encoding="utf-8",
            )
            pd.DataFrame(
                {
                    "window": ["pf_01"],
                    "train_start": [20190101],
                    "train_end": [20191231],
                    "test_start": [20200101],
                    "test_end": [20200104],
                    "selected_candidate": ["selected_stitched__sat100"],
                    "selected_score": [1.0],
                    "selected_params_path": ["params_source_selection_stitched.csv"],
                }
            ).to_csv(allocator / "portfolio_walk_forward_summary.csv", index=False)

            result = run_paper_account(
                project_root=root,
                config_path=config_path,
                allocator_dir=allocator,
                output_dir=root / "paper",
            )

            self.assertEqual(result.metrics["latest_candidate"], "selected_stitched__sat100")
            self.assertGreater(float(result.metrics["total_return"]), 0.90)
            self.assertEqual(result.target_holdings_payload.get("status"), "ok")

    def test_market_cap_snapshot_normalizes_yuan_to_yi(self) -> None:
        raw = pd.DataFrame(
            {
                "代码": ["000001", "600000"],
                "名称": ["One", "Two"],
                "总市值": [1600 * 100_000_000, 900 * 100_000_000],
                "流通市值": [1400 * 100_000_000, 800 * 100_000_000],
                "最新价": [12.3, 8.9],
                "涨跌幅": [1.2, -0.5],
            }
        )
        frame = normalize_stock_market_cap_frame(raw, snapshot_date="2024-04-02", updated_at="2024-04-02T16:00:00")

        self.assertEqual(list(frame["code"]), ["000001", "600000"])
        self.assertAlmostEqual(float(frame.loc[frame["code"] == "000001", "market_cap_yi"].iloc[0]), 1600.0)
        self.assertEqual(frame.loc[frame["code"] == "000001", "snapshot_date"].iloc[0], "2024-04-02")

    def test_stock_market_cap_tracking_rule_excludes_large_caps_from_review(self) -> None:
        targets = pd.DataFrame(
            [
                {
                    "date": "2024-04-02",
                    "layer": "core",
                    "code": "000001",
                    "name": "Large",
                    "portfolio_target_weight": 0.08,
                    "portfolio_target_value": 80000.0,
                    "unrealized_return": 0.02,
                    "holding_days": 3,
                    "target_action": "target_hold",
                    "risk_filter_status": "active_with_source_risk_model",
                    "trigger_monitor_status": "matched_latest_trigger",
                    "trigger_summary": "matched",
                    "source_strategy_profile": "multi_factor:test",
                },
                {
                    "date": "2024-04-02",
                    "layer": "core",
                    "code": "000002",
                    "name": "Smaller",
                    "portfolio_target_weight": 0.07,
                    "portfolio_target_value": 70000.0,
                    "unrealized_return": -0.03,
                    "holding_days": 5,
                    "target_action": "target_hold",
                    "risk_filter_status": "active_with_source_risk_model",
                    "trigger_monitor_status": "not_in_latest_trigger_report",
                    "trigger_summary": "not_in_latest_trigger_report",
                    "source_strategy_profile": "multi_factor:test",
                },
            ]
        )
        market_caps = pd.DataFrame(
            {
                "code": ["000001", "000002"],
                "name": ["Large", "Smaller"],
                "market_cap_yi": [1600.0, 900.0],
                "snapshot_date": ["2024-04-02", "2024-04-02"],
                "updated_at": ["2024-04-02T16:00:00", "2024-04-02T16:00:00"],
                "source": ["unit", "unit"],
            }
        )
        enriched, tracking_payload = apply_stock_market_cap_tracking_rule(
            targets,
            market_caps,
            max_market_cap_yi=1500.0,
            market_cap_cache_payload={"stock_market_cap_cache_status": "ok"},
        )

        excluded = enriched[enriched["code"] == "000001"].iloc[0]
        self.assertTrue(bool(excluded["tracking_excluded"]))
        self.assertEqual(excluded["tracking_rule_status"], "excluded_large_market_cap")
        self.assertEqual(int(tracking_payload["stock_tracking_excluded_large_market_cap_count"]), 1)

        review, review_payload = build_stock_target_review(
            enriched,
            {
                "status": "ok",
                "latest_date": "2024-04-02",
                "stock_target_count": 2,
                "trigger_match_count": 1,
                "latest_trigger_run_time": "2024-04-02T15:10:00",
                **tracking_payload,
            },
        )
        self.assertEqual(list(review["code"]), ["000002"])
        self.assertEqual(int(review_payload["review_skipped_tracking_excluded_count"]), 1)

    def test_stock_target_review_thresholds_are_configurable(self) -> None:
        targets = pd.DataFrame(
            [
                {
                    "date": "2024-04-02",
                    "layer": "core",
                    "code": "000001",
                    "name": "One",
                    "portfolio_target_weight": 0.066,
                    "portfolio_target_value": 50000.0,
                    "unrealized_return": -0.09,
                    "holding_days": 3,
                    "target_action": "target_hold",
                    "risk_filter_status": "active_with_source_risk_model",
                    "trigger_monitor_status": "not_in_latest_trigger_report",
                    "trigger_summary": "not_in_latest_trigger_report",
                    "source_strategy_profile": "multi_factor:test",
                }
            ]
        )

        default_review, default_payload = build_stock_target_review(targets, {"status": "ok", "latest_date": "2024-04-02", "stock_target_count": 1})
        strict_review, strict_payload = build_stock_target_review(
            targets,
            {"status": "ok", "latest_date": "2024-04-02", "stock_target_count": 1},
            drawdown_threshold=-0.08,
            watch_drawdown_threshold=-0.06,
            loss_attention_threshold=-0.04,
        )

        self.assertEqual(default_review.iloc[0]["review_bucket"], "watch_review")
        self.assertEqual(int(default_payload["drawdown_review_count"]), 0)
        self.assertEqual(strict_review.iloc[0]["review_bucket"], "drawdown_review")
        self.assertEqual(int(strict_payload["drawdown_review_count"]), 1)
        self.assertAlmostEqual(float(strict_payload["drawdown_threshold"]), -0.08)

    def test_stock_target_review_notes_preserve_manual_fields(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            notes_path = root / "stock_target_review_notes.csv"
            snapshot_path = root / "run" / "stock_target_review_notes.csv"
            pd.DataFrame(
                [
                    {
                        "layer": "core",
                        "code": "000001",
                        "name": "One",
                        "first_seen_date": "2024-04-01",
                        "last_seen_date": "2024-04-01",
                        "last_review_bucket": "trigger_review",
                        "last_review_stage": "review_required",
                        "last_review_priority_score": 55.0,
                        "manual_status": "reviewed",
                        "manual_note": "confirmed local context",
                        "reviewed_at": "2024-04-01T16:30:00",
                        "reviewed_by": "tester",
                        "next_review_date": "2024-04-03",
                        "last_model_reason": "previous model reason",
                        "broker_action": "none",
                        "research_only": True,
                    }
                ]
            ).to_csv(notes_path, index=False)
            targets = pd.DataFrame(
                [
                    {
                        "date": "2024-04-02",
                        "layer": "core",
                        "code": "000001",
                        "name": "One",
                        "portfolio_target_weight": 0.07,
                        "portfolio_target_value": 70000.0,
                        "unrealized_return": 0.02,
                        "holding_days": 3,
                        "target_action": "target_hold",
                        "risk_filter_status": "active_with_source_risk_model",
                        "trigger_monitor_status": "matched_latest_trigger",
                        "trigger_summary": "matched",
                        "source_strategy_profile": "multi_factor:test",
                    },
                    {
                        "date": "2024-04-02",
                        "layer": "satellite",
                        "code": "000002",
                        "name": "Two",
                        "portfolio_target_weight": 0.06,
                        "portfolio_target_value": 60000.0,
                        "unrealized_return": -0.12,
                        "holding_days": 4,
                        "target_action": "target_hold",
                        "risk_filter_status": "active_with_source_risk_model",
                        "trigger_monitor_status": "not_in_latest_trigger_report",
                        "trigger_summary": "none",
                        "source_strategy_profile": "multi_factor:test",
                    },
                ]
            )

            review, _ = build_stock_target_review(
                targets,
                {"status": "ok", "latest_date": "2024-04-02", "stock_target_count": 2},
            )
            enriched, notes_payload = sync_stock_target_review_notes(review, notes_path, snapshot_path)

            one = enriched[enriched["code"] == "000001"].iloc[0]
            two = enriched[enriched["code"] == "000002"].iloc[0]
            self.assertEqual(one["manual_status"], "reviewed")
            self.assertEqual(one["manual_note"], "confirmed local context")
            self.assertEqual(one["manual_status_normalized"], "reviewed")
            self.assertEqual(one["manual_review_state"], "reviewed_closed")
            self.assertEqual(two["manual_status"], "unreviewed")
            self.assertEqual(two["manual_status_normalized"], "unreviewed")
            self.assertEqual(two["notes_status"], "unreviewed")
            self.assertEqual(int(notes_payload["manual_note_count"]), 1)
            self.assertEqual(int(notes_payload["unreviewed_count"]), 1)
            self.assertEqual(int(notes_payload["review_required_unreviewed_count"]), 1)
            self.assertEqual(int(notes_payload["manual_pending_count"]), 1)
            self.assertEqual(int(notes_payload["manual_reviewed_count"]), 1)
            self.assertEqual(int(notes_payload["manual_watch_count"]), 0)
            self.assertTrue(snapshot_path.exists())
            persisted = pd.read_csv(notes_path, dtype={"code": str})
            self.assertEqual(set(persisted["code"]), {"000001", "000002"})
            persisted_one = persisted[persisted["code"] == "000001"].iloc[0]
            self.assertEqual(persisted_one["manual_note"], "confirmed local context")
            self.assertEqual(persisted_one["last_seen_date"], "2024-04-02")

    def test_stock_target_review_notes_classify_manual_statuses(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            notes_path = root / "stock_target_review_notes.csv"
            snapshot_path = root / "run" / "stock_target_review_notes.csv"
            pd.DataFrame(
                [
                    {"layer": "core", "code": "000001", "manual_status": "reviewed", "manual_note": ""},
                    {"layer": "core", "code": "000002", "manual_status": "watch", "manual_note": ""},
                    {"layer": "core", "code": "000003", "manual_status": "resolved", "manual_note": ""},
                    {"layer": "core", "code": "000004", "manual_status": "exclude_candidate", "manual_note": ""},
                    {"layer": "core", "code": "000005", "manual_status": "custom_pause", "manual_note": ""},
                ]
            ).to_csv(notes_path, index=False)
            targets = pd.DataFrame(
                [
                    {
                        "date": "2024-04-02",
                        "layer": "core",
                        "code": f"{code:06d}",
                        "name": f"Stock{code}",
                        "portfolio_target_weight": 0.05,
                        "portfolio_target_value": 50000.0,
                        "unrealized_return": -0.12 if code == 6 else 0.01,
                        "holding_days": 3,
                        "target_action": "target_hold",
                        "risk_filter_status": "active_with_source_risk_model",
                        "trigger_monitor_status": "not_in_latest_trigger_report",
                        "trigger_summary": "none",
                        "source_strategy_profile": "multi_factor:test",
                    }
                    for code in range(1, 7)
                ]
            )

            review, _ = build_stock_target_review(
                targets,
                {"status": "ok", "latest_date": "2024-04-02", "stock_target_count": 6},
            )
            enriched, notes_payload = sync_stock_target_review_notes(review, notes_path, snapshot_path)

            statuses = dict(zip(enriched["code"], enriched["manual_status_normalized"]))
            states = dict(zip(enriched["code"], enriched["manual_review_state"]))
            followups = dict(zip(enriched["code"], enriched["manual_followup_required"]))
            self.assertEqual(statuses["000001"], "reviewed")
            self.assertEqual(statuses["000002"], "watch")
            self.assertEqual(statuses["000003"], "resolved")
            self.assertEqual(statuses["000004"], "exclude_candidate")
            self.assertEqual(statuses["000005"], "other")
            self.assertEqual(statuses["000006"], "unreviewed")
            self.assertEqual(states["000002"], "reviewed_watch")
            self.assertEqual(states["000004"], "reviewed_exclude_candidate")
            self.assertFalse(bool(followups["000001"]))
            self.assertFalse(bool(followups["000003"]))
            self.assertTrue(bool(followups["000002"]))
            self.assertTrue(bool(followups["000004"]))
            self.assertTrue(bool(followups["000005"]))
            self.assertTrue(bool(followups["000006"]))
            self.assertEqual(int(notes_payload["manual_pending_count"]), 1)
            self.assertEqual(int(notes_payload["manual_reviewed_count"]), 1)
            self.assertEqual(int(notes_payload["manual_watch_count"]), 1)
            self.assertEqual(int(notes_payload["manual_resolved_count"]), 1)
            self.assertEqual(int(notes_payload["manual_exclude_candidate_count"]), 1)
            self.assertEqual(int(notes_payload["manual_other_status_count"]), 1)
            self.assertEqual(int(notes_payload["review_required_pending_count"]), 1)
            self.assertEqual(int(notes_payload["review_required_reviewed_count"]), 0)

    def test_stock_target_review_actions_build_manual_queue(self) -> None:
        review = pd.DataFrame(
            [
                {
                    "date": "2024-04-02",
                    "review_rank": 1,
                    "review_bucket": "drawdown_review",
                    "review_stage": "review_required",
                    "review_priority_score": 48.0,
                    "layer": "core",
                    "code": "000001",
                    "name": "One",
                    "portfolio_target_weight": 0.07,
                    "unrealized_return": -0.12,
                    "manual_status": "unreviewed",
                    "manual_status_normalized": "unreviewed",
                    "manual_review_state": "pending_review",
                    "next_review_date": "",
                },
                {
                    "date": "2024-04-02",
                    "review_rank": 2,
                    "review_bucket": "watch_review",
                    "review_stage": "monitor",
                    "review_priority_score": 33.0,
                    "layer": "core",
                    "code": "000002",
                    "name": "Two",
                    "portfolio_target_weight": 0.05,
                    "unrealized_return": 0.02,
                    "manual_status": "watch",
                    "manual_status_normalized": "watch",
                    "manual_review_state": "reviewed_watch",
                    "next_review_date": "",
                },
                {
                    "date": "2024-04-02",
                    "review_rank": 3,
                    "review_bucket": "routine_review",
                    "review_stage": "routine",
                    "review_priority_score": 10.0,
                    "layer": "core",
                    "code": "000003",
                    "name": "Three",
                    "portfolio_target_weight": 0.04,
                    "unrealized_return": 0.03,
                    "manual_status": "exclude_candidate",
                    "manual_status_normalized": "exclude_candidate",
                    "manual_review_state": "reviewed_exclude_candidate",
                    "next_review_date": "",
                },
                {
                    "date": "2024-04-02",
                    "review_rank": 4,
                    "review_bucket": "routine_review",
                    "review_stage": "routine",
                    "review_priority_score": 8.0,
                    "layer": "core",
                    "code": "000004",
                    "name": "Four",
                    "portfolio_target_weight": 0.04,
                    "unrealized_return": 0.01,
                    "manual_status": "custom_pause",
                    "manual_status_normalized": "other",
                    "manual_review_state": "reviewed_other",
                    "next_review_date": "",
                },
                {
                    "date": "2024-04-02",
                    "review_rank": 5,
                    "review_bucket": "routine_review",
                    "review_stage": "routine",
                    "review_priority_score": 5.0,
                    "layer": "core",
                    "code": "000005",
                    "name": "Five",
                    "portfolio_target_weight": 0.04,
                    "unrealized_return": 0.01,
                    "manual_status": "reviewed",
                    "manual_status_normalized": "reviewed",
                    "manual_review_state": "reviewed_closed",
                    "next_review_date": "",
                },
            ]
        )

        actions, payload = build_stock_target_review_actions(
            review,
            {"status": "ok", "latest_date": "2024-04-02", "review_row_count": 5, "notes_path": "notes.csv"},
        )

        self.assertEqual(int(payload["action_count"]), 4)
        self.assertEqual(int(payload["review_required_pending_count"]), 1)
        self.assertEqual(int(payload["manual_watch_followup_count"]), 1)
        self.assertEqual(int(payload["manual_exclusion_candidate_count"]), 1)
        self.assertEqual(int(payload["manual_status_normalization_count"]), 1)
        self.assertEqual(actions.iloc[0]["action_code"], "review_required_pending")
        self.assertEqual(actions.iloc[-1]["action_code"], "manual_watch_followup")
        self.assertEqual(set(actions["broker_action"]), {"none"})

    def test_stock_target_review_assistant_uses_local_context_without_future_data(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            data_dir = root / "data" / "processed" / "stocks"
            data_dir.mkdir(parents=True)
            pd.DataFrame(
                {
                    "date": pd.date_range("2024-04-01", periods=25, freq="D"),
                    "close": [float(value) for value in range(10, 35)],
                }
            ).to_csv(data_dir / "000001.csv", index=False)
            review = pd.DataFrame(
                [
                    {
                        "date": "2024-04-21",
                        "review_rank": 1,
                        "review_bucket": "drawdown_review",
                        "review_stage": "review_required",
                        "review_priority_score": 48.0,
                        "layer": "core",
                        "code": "000001",
                        "name": "One",
                        "portfolio_target_weight": 0.07,
                        "unrealized_return": -0.12,
                        "holding_days": 3,
                        "trigger_monitor_status": "not_in_latest_trigger_report",
                        "review_reason": "unrealized loss <= -10.00%",
                        "recommended_review": "Review loss source; no broker order is generated.",
                        "manual_status": "unreviewed",
                        "manual_status_normalized": "unreviewed",
                        "manual_review_state": "pending_review",
                        "next_review_date": "",
                    }
                ]
            )
            actions, _ = build_stock_target_review_actions(
                review,
                {"status": "ok", "latest_date": "2024-04-21", "review_row_count": 1, "notes_path": "notes.csv"},
            )

            assistant, payload = build_stock_target_review_assistant(
                root,
                review,
                actions,
                {"status": "ok", "latest_date": "2024-04-21", "notes_path": "notes.csv"},
            )

            self.assertEqual(int(payload["assistant_row_count"]), 1)
            self.assertEqual(int(payload["review_required_pending_count"]), 1)
            self.assertEqual(int(payload["price_history_ok_count"]), 1)
            row = assistant.iloc[0]
            self.assertEqual(row["latest_price_date"], "2024-04-21")
            self.assertAlmostEqual(float(row["return_20d"]), 2.0)
            self.assertIn("manual_status/manual_note", row["evidence_checklist"])
            self.assertIn("reviewed|watch|resolved", row["manual_status_options"])
            self.assertEqual(row["broker_action"], "none")
            self.assertTrue(bool(row["research_only"]))

    def test_stock_target_review_assistant_includes_manual_followup_rows_without_actions(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            data_dir = root / "data" / "processed" / "stocks"
            data_dir.mkdir(parents=True)
            pd.DataFrame(
                {
                    "date": pd.date_range("2024-04-01", periods=22, freq="D"),
                    "close": [20.0 + float(value) for value in range(22)],
                }
            ).to_csv(data_dir / "000002.csv", index=False)
            review = pd.DataFrame(
                [
                    {
                        "date": "2024-04-22",
                        "review_rank": 6,
                        "review_bucket": "routine_review",
                        "review_stage": "routine",
                        "review_priority_score": 18.0,
                        "layer": "core",
                        "code": "000002",
                        "name": "Two",
                        "portfolio_target_weight": 0.05,
                        "unrealized_return": -0.06,
                        "holding_days": 3,
                        "trigger_monitor_status": "not_in_latest_trigger_report",
                        "review_reason": "unrealized loss <= -5.00%",
                        "recommended_review": "Review routine loss context; no broker order is generated.",
                        "manual_status": "unreviewed",
                        "manual_status_normalized": "unreviewed",
                        "manual_review_state": "pending_review",
                        "manual_followup_required": True,
                        "next_review_date": "",
                    }
                ]
            )
            actions = pd.DataFrame(columns=["layer", "code", "action_code"])

            assistant, payload = build_stock_target_review_assistant(
                root,
                review,
                actions,
                {"status": "ok", "latest_date": "2024-04-22", "notes_path": "notes.csv"},
            )

            self.assertEqual(int(payload["assistant_row_count"]), 1)
            row = assistant.iloc[0]
            self.assertEqual(row["action_code"], "manual_followup_pending")
            self.assertEqual(row["review_bucket"], "routine_review")
            self.assertEqual(row["latest_price_date"], "2024-04-22")
            self.assertIn("human review conclusion", row["manual_note_prompt"])
            self.assertEqual(row["broker_action"], "none")

    def test_stock_target_review_assistant_adds_local_review_levels_and_conditions(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            data_dir = root / "data" / "processed" / "stocks"
            data_dir.mkdir(parents=True)
            pd.DataFrame(
                {
                    "date": pd.date_range("2024-04-01", periods=22, freq="D"),
                    "high": [
                        10.5,
                        11.5,
                        12.5,
                        13.5,
                        14.5,
                        15.5,
                        16.5,
                        17.5,
                        18.5,
                        19.5,
                        20.5,
                        21.5,
                        22.5,
                        23.5,
                        24.5,
                        25.5,
                        26.5,
                        27.5,
                        28.5,
                        29.5,
                        30.5,
                        31.5,
                    ],
                    "low": [
                        9.5,
                        10.5,
                        11.5,
                        12.5,
                        13.5,
                        14.5,
                        15.5,
                        16.5,
                        17.5,
                        18.5,
                        19.5,
                        20.5,
                        21.5,
                        22.5,
                        23.5,
                        24.5,
                        25.5,
                        26.5,
                        27.5,
                        28.5,
                        29.5,
                        30.5,
                    ],
                    "close": [10.0 + float(value) for value in range(22)],
                }
            ).to_csv(data_dir / "000003.csv", index=False)
            review = pd.DataFrame(
                [
                    {
                        "date": "2024-04-22",
                        "review_rank": 1,
                        "review_bucket": "watch_review",
                        "review_stage": "monitor",
                        "review_priority_score": 30.0,
                        "layer": "core",
                        "code": "000003",
                        "name": "Three",
                        "portfolio_target_weight": 0.05,
                        "unrealized_return": -0.08,
                        "holding_days": 0,
                        "trigger_monitor_status": "not_in_latest_trigger_report",
                        "review_reason": "unrealized loss <= -7.00%",
                        "recommended_review": "Routine watch with extra attention to loss state.",
                        "manual_status": "unreviewed",
                        "manual_status_normalized": "unreviewed",
                        "manual_review_state": "pending_review",
                        "manual_followup_required": True,
                    }
                ]
            )

            assistant, _ = build_stock_target_review_assistant(
                root,
                review,
                pd.DataFrame(columns=["layer", "code", "action_code"]),
                {"status": "ok", "latest_date": "2024-04-22", "notes_path": "notes.csv"},
            )

            row = assistant.iloc[0]
            self.assertAlmostEqual(float(row["latest_close"]), 31.0)
            self.assertAlmostEqual(float(row["latest_pct_change"]), 1 / 30)
            self.assertAlmostEqual(float(row["support_price"]), 11.5)
            self.assertAlmostEqual(float(row["pressure_price"]), 31.5)
            self.assertAlmostEqual(float(row["stop_loss_reference"]), 11.27)
            self.assertEqual(row["suggested_review_posture"], "risk_review")
            self.assertIn("31.50", row["trigger_condition"])
            self.assertIn("11.27", row["invalidation_condition"])
            self.assertEqual(row["broker_action"], "none")

            report = _render_stock_target_review_assistant_report(
                assistant,
                {"generated_at": "2024-04-22T16:00:00", "status": "ok", "latest_date": "2024-04-22", "broker_action": "none"},
                root / "paper",
            )
            self.assertIn("Support", report)
            self.assertIn("Pressure", report)
            self.assertIn("Stop-loss ref", report)
            self.assertIn("Review posture", report)
            self.assertIn("risk_review", report)
            self.assertIn("31.50", report)
            self.assertIn("11.27", report)

    def test_stock_target_review_decision_template_keeps_status_blank_for_human_review(self) -> None:
        assistant = pd.DataFrame(
            [
                {
                    "date": "2026-06-15",
                    "assistant_rank": 1,
                    "action_code": "manual_followup_pending",
                    "review_bucket": "watch_review",
                    "review_stage": "monitor",
                    "layer": "core",
                    "code": "002216",
                    "name": "三全食品",
                    "portfolio_target_weight": 0.050271,
                    "unrealized_return": -0.073567,
                    "latest_close": 12.44,
                    "latest_pct_change": -0.0048,
                    "support_price": 11.90,
                    "pressure_price": 13.35,
                    "stop_loss_reference": 11.66,
                    "suggested_review_posture": "risk_review",
                    "trigger_condition": "Review strength only if close reclaims local pressure 13.35.",
                    "invalidation_condition": "Review risk if close breaks stop-loss reference 11.66.",
                    "evidence_checklist": "review model reason; check support, pressure, and stop-loss reference",
                    "review_reason": "unrealized loss <= -7.00%",
                    "recommended_review": "Routine watch with extra attention to loss state.",
                    "broker_action": "none",
                    "research_only": True,
                }
            ]
        )

        template, payload = build_stock_target_review_decision_template(
            assistant,
            {"status": "ok", "latest_date": "2026-06-15", "notes_path": "outputs/research/stock_target_review_notes.csv"},
        )

        self.assertEqual(int(payload["decision_template_row_count"]), 1)
        self.assertEqual(int(payload["blank_manual_status_count"]), 1)
        row = template.iloc[0]
        self.assertEqual(row["manual_status_to_fill"], "")
        self.assertEqual(row["allowed_manual_statuses"], "reviewed|watch|resolved|exclude_candidate|unreviewed")
        self.assertEqual(row["status_hint"], "risk_review: choose watch/resolved/reviewed after human evidence check")
        self.assertEqual(row["broker_action"], "none")
        self.assertTrue(bool(row["research_only"]))
        report = _render_stock_target_review_decision_template_report(template, payload, Path("paper"))
        self.assertIn("Paper Stock Target Review Decision Template", report)
        self.assertIn("manual_status_to_fill", report)
        self.assertIn("does not set manual_status", report)
        self.assertIn("002216", report)

    def test_stock_target_review_decision_template_xlsx_highlights_user_fill_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            template = pd.DataFrame(
                [
                    {
                        "date": "2026-06-15",
                        "decision_rank": 1,
                        "source_assistant_rank": 1,
                        "layer": "core",
                        "code": "002216",
                        "name": "Sanquan",
                        "review_posture": "risk_review",
                        "status_hint": "choose watch/resolved/reviewed",
                        "manual_status_to_fill": "",
                        "manual_note_to_fill": "",
                        "next_review_date_to_fill": "",
                        "reviewed_by_to_fill": "",
                        "allowed_manual_statuses": "reviewed|watch|resolved|exclude_candidate|unreviewed",
                        "portfolio_target_weight": 0.05,
                        "unrealized_return": -0.07,
                        "latest_close": 12.44,
                        "latest_pct_change": -0.004,
                        "support_price": 11.90,
                        "pressure_price": 13.35,
                        "stop_loss_reference": 11.66,
                        "trigger_condition": "Review strength only if close reclaims local pressure.",
                        "invalidation_condition": "Review risk if close breaks stop-loss reference.",
                        "evidence_checklist": "check support, pressure, stop-loss reference",
                        "review_reason": "unrealized loss <= -7.00%",
                        "recommended_review": "Routine watch.",
                        "notes_path": "outputs/research/stock_target_review_notes.csv",
                        "broker_action": "none",
                        "research_only": True,
                    }
                ]
            )
            xlsx_path = root / "stock_target_review_decision_template.xlsx"

            self.assertTrue(hasattr(paper_account_module, "write_stock_target_review_decision_template_xlsx"))
            paper_account_module.write_stock_target_review_decision_template_xlsx(
                template,
                {"latest_date": "2026-06-15"},
                xlsx_path,
            )

            from openpyxl import load_workbook

            workbook = load_workbook(xlsx_path)
            sheet = workbook["decision_template"]
            header_map = {cell.value: cell.column for cell in sheet[1]}
            fill_columns = [
                "manual_status_to_fill",
                "manual_note_to_fill",
                "next_review_date_to_fill",
                "reviewed_by_to_fill",
            ]
            for column in fill_columns:
                cell = sheet.cell(row=2, column=header_map[column])
                self.assertIn(cell.fill.fill_type, {"solid"})
                self.assertNotIn(cell.fill.fgColor.rgb, {None, "00000000", "00FFFFFF", "FFFFFFFF"})
            code_cell = sheet.cell(row=2, column=header_map["code"])
            self.assertNotEqual(code_cell.fill.fgColor.rgb, sheet.cell(row=2, column=header_map["manual_status_to_fill"]).fill.fgColor.rgb)
            self.assertGreaterEqual(len(sheet.data_validations.dataValidation), 1)
            self.assertEqual(sheet.freeze_panes, "A2")

    def test_apply_stock_target_review_decision_template_updates_only_valid_filled_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            template_path = root / "stock_target_review_decision_template.csv"
            notes_path = root / "stock_target_review_notes.csv"
            output_dir = root / "apply"
            pd.DataFrame(
                [
                    {
                        "date": "2026-06-15",
                        "layer": "core",
                        "code": "002216",
                        "name": "Sanquan",
                        "review_posture": "risk_review",
                        "manual_status_to_fill": "watch",
                        "manual_note_to_fill": "support intact; keep watching",
                        "next_review_date_to_fill": "2026-06-18",
                        "reviewed_by_to_fill": "tester",
                        "review_reason": "loss review",
                    },
                    {
                        "date": "2026-06-15",
                        "layer": "core",
                        "code": "003018",
                        "name": "Jinfu",
                        "review_posture": "profit_protection_review",
                        "manual_status_to_fill": "",
                        "manual_note_to_fill": "blank status should be ignored",
                        "next_review_date_to_fill": "2026-06-20",
                        "reviewed_by_to_fill": "tester",
                        "review_reason": "profit review",
                    },
                    {
                        "date": "2026-06-15",
                        "layer": "satellite",
                        "code": "601077",
                        "name": "CQRCB",
                        "review_posture": "risk_review",
                        "manual_status_to_fill": "sell_now",
                        "manual_note_to_fill": "invalid status should not write",
                        "next_review_date_to_fill": "2026-06-19",
                        "reviewed_by_to_fill": "tester",
                        "review_reason": "invalid review",
                    },
                ]
            ).to_csv(template_path, index=False)
            pd.DataFrame(
                [
                    {
                        "layer": "core",
                        "code": "002216",
                        "name": "OldName",
                        "first_seen_date": "2026-06-10",
                        "last_seen_date": "2026-06-14",
                        "last_review_bucket": "watch_review",
                        "last_review_stage": "monitor",
                        "last_review_priority_score": 25.0,
                        "manual_status": "unreviewed",
                        "manual_note": "",
                        "reviewed_at": "",
                        "reviewed_by": "",
                        "next_review_date": "",
                        "last_model_reason": "previous reason",
                        "broker_action": "none",
                        "research_only": True,
                    }
                ]
            ).to_csv(notes_path, index=False)

            self.assertTrue(hasattr(paper_account_module, "apply_stock_target_review_decision_template"))
            result = paper_account_module.apply_stock_target_review_decision_template(
                template_path,
                notes_path,
                output_dir,
                reviewed_at="2026-06-15T16:10:00",
            )

            self.assertEqual(int(result.payload["applied_count"]), 1)
            self.assertEqual(int(result.payload["blank_ignored_count"]), 1)
            self.assertEqual(int(result.payload["invalid_status_count"]), 1)
            self.assertEqual(result.payload["broker_action"], "none")
            self.assertTrue(result.audit_path.exists())
            self.assertTrue(result.report_path.exists())
            persisted = pd.read_csv(notes_path, dtype={"code": str})
            self.assertEqual(set(persisted["code"]), {"002216"})
            row = persisted.iloc[0]
            self.assertEqual(row["manual_status"], "watch")
            self.assertEqual(row["manual_note"], "support intact; keep watching")
            self.assertEqual(row["reviewed_at"], "2026-06-15T16:10:00")
            self.assertEqual(row["reviewed_by"], "tester")
            self.assertEqual(row["next_review_date"], "2026-06-18")
            self.assertEqual(row["last_model_reason"], "loss review")
            self.assertEqual(row["broker_action"], "none")
            audit = pd.read_csv(result.audit_path, dtype={"code": str})
            statuses = dict(zip(audit["code"], audit["apply_status"]))
            self.assertEqual(statuses["002216"], "applied")
            self.assertEqual(statuses["003018"], "blank_ignored")
            self.assertEqual(statuses["601077"], "invalid_status")
            report = result.report_path.read_text(encoding="utf-8")
            self.assertIn("Stock Target Review Decision Apply", report)
            self.assertIn("research-only", report)

    def test_apply_stock_target_review_decision_template_reads_filled_xlsx_template(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            template = pd.DataFrame(
                [
                    {
                        "date": "2026-06-15",
                        "decision_rank": 1,
                        "source_assistant_rank": 1,
                        "layer": "core",
                        "code": "002216",
                        "name": "Sanquan",
                        "review_posture": "risk_review",
                        "status_hint": "choose watch/resolved/reviewed",
                        "manual_status_to_fill": "",
                        "manual_note_to_fill": "",
                        "next_review_date_to_fill": "",
                        "reviewed_by_to_fill": "",
                        "allowed_manual_statuses": "reviewed|watch|resolved|exclude_candidate|unreviewed",
                        "portfolio_target_weight": 0.05,
                        "unrealized_return": -0.07,
                        "latest_close": 12.44,
                        "latest_pct_change": -0.004,
                        "support_price": 11.90,
                        "pressure_price": 13.35,
                        "stop_loss_reference": 11.66,
                        "trigger_condition": "Review strength only if close reclaims local pressure.",
                        "invalidation_condition": "Review risk if close breaks stop-loss reference.",
                        "evidence_checklist": "check support, pressure, stop-loss reference",
                        "review_reason": "unrealized loss <= -7.00%",
                        "recommended_review": "Routine watch.",
                        "notes_path": "outputs/research/stock_target_review_notes.csv",
                        "broker_action": "none",
                        "research_only": True,
                    }
                ]
            )
            xlsx_path = root / "stock_target_review_decision_template.xlsx"
            notes_path = root / "stock_target_review_notes.csv"
            output_dir = root / "apply"
            paper_account_module.write_stock_target_review_decision_template_xlsx(
                template,
                {"latest_date": "2026-06-15"},
                xlsx_path,
            )

            from openpyxl import load_workbook

            workbook = load_workbook(xlsx_path)
            sheet = workbook["decision_template"]
            header_map = {cell.value: cell.column for cell in sheet[1]}
            sheet.cell(row=2, column=header_map["manual_status_to_fill"]).value = "watch"
            sheet.cell(row=2, column=header_map["manual_note_to_fill"]).value = "filled in Excel"
            sheet.cell(row=2, column=header_map["reviewed_by_to_fill"]).value = "tester"
            workbook.save(xlsx_path)

            result = paper_account_module.apply_stock_target_review_decision_template(
                xlsx_path,
                notes_path,
                output_dir,
                reviewed_at="2026-06-15T16:30:00",
            )

            self.assertEqual(int(result.payload["applied_count"]), 1)
            self.assertEqual(int(result.payload["blank_ignored_count"]), 0)
            persisted = pd.read_csv(notes_path, dtype={"code": str})
            self.assertEqual(persisted.iloc[0]["code"], "002216")
            self.assertEqual(persisted.iloc[0]["manual_status"], "watch")
            self.assertEqual(persisted.iloc[0]["manual_note"], "filled in Excel")
            self.assertEqual(persisted.iloc[0]["reviewed_by"], "tester")

    def test_stock_target_review_outcomes_track_future_returns_and_pending(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            data_dir = root / "data" / "processed" / "stocks"
            data_dir.mkdir(parents=True)
            pd.DataFrame(
                {
                    "date": pd.date_range("2024-04-01", periods=8, freq="D"),
                    "close": [10.0, 11.0, 12.0, 13.0, 14.0, 15.0, 16.0, 17.0],
                }
            ).to_csv(data_dir / "000001.csv", index=False)
            review = pd.DataFrame(
                [
                    {
                        "date": "2024-04-01",
                        "layer": "core",
                        "code": "000001",
                        "name": "One",
                        "review_bucket": "drawdown_review",
                        "review_stage": "review_required",
                        "review_priority_score": 48.0,
                        "manual_status": "unreviewed",
                        "manual_status_normalized": "unreviewed",
                        "manual_review_state": "pending_review",
                    },
                    {
                        "date": "2024-04-01",
                        "layer": "core",
                        "code": "000002",
                        "name": "Two",
                        "review_bucket": "watch_review",
                        "review_stage": "monitor",
                        "review_priority_score": 20.0,
                        "manual_status": "watch",
                        "manual_status_normalized": "watch",
                        "manual_review_state": "reviewed_watch",
                    },
                ]
            )
            actions = pd.DataFrame(
                [
                    {
                        "layer": "core",
                        "code": "000001",
                        "action_code": "review_required_pending",
                    }
                ]
            )

            outcomes, payload = build_stock_target_review_outcomes(
                root,
                review,
                actions,
                {"status": "ok", "latest_date": "2024-04-01", "notes_path": "notes.csv"},
            )

            one = outcomes[outcomes["code"] == "000001"].iloc[0]
            two = outcomes[outcomes["code"] == "000002"].iloc[0]
            self.assertEqual(one["entry_date"], "2024-04-01")
            self.assertAlmostEqual(float(one["return_1d"]), 0.10)
            self.assertAlmostEqual(float(one["return_5d"]), 0.50)
            self.assertEqual(one["outcome_status_10d"], "pending")
            self.assertEqual(one["outcome_status"], "partial")
            self.assertEqual(one["action_code"], "review_required_pending")
            self.assertEqual(two["outcome_status"], "missing_entry_price")
            self.assertEqual(int(payload["outcome_row_count"]), 2)
            self.assertEqual(int(payload["partial_count"]), 1)
            self.assertEqual(int(payload["missing_entry_price_count"]), 1)
            self.assertEqual(payload["horizon_summary"]["1d"]["evaluable_count"], 1)
            self.assertAlmostEqual(float(payload["horizon_summary"]["1d"]["avg_return"]), 0.10)

    def test_stock_target_review_outcomes_history_updates_existing_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            history_path = root / "outputs" / "research" / "stock_target_review_outcomes_history.csv"
            snapshot_path = root / "paper" / "stock_target_review_outcomes_history.csv"
            history_path.parent.mkdir(parents=True)
            pd.DataFrame(
                [
                    {
                        "date": "2024-04-01",
                        "layer": "core",
                        "code": "000001",
                        "name": "One",
                        "outcome_status": "pending",
                        "return_1d": pd.NA,
                        "broker_action": "none",
                        "research_only": True,
                    }
                ]
            ).to_csv(history_path, index=False)

            current = pd.DataFrame(
                [
                    {
                        "date": "2024-04-01",
                        "layer": "core",
                        "code": "000001",
                        "name": "One",
                        "review_bucket": "drawdown_review",
                        "review_stage": "review_required",
                        "review_priority_score": 50.0,
                        "action_code": "review_required_pending",
                        "manual_status": "unreviewed",
                        "manual_status_normalized": "unreviewed",
                        "manual_review_state": "pending_review",
                        "entry_date": "2024-04-01",
                        "entry_close": 10.0,
                        "price_source": "local",
                        "outcome_status": "partial",
                        "return_1d": 0.10,
                        "outcome_status_1d": "available",
                        "broker_action": "none",
                        "research_only": True,
                    },
                    {
                        "date": "2024-04-01",
                        "layer": "satellite",
                        "code": "000002",
                        "name": "Two",
                        "review_bucket": "watch_review",
                        "review_stage": "monitor",
                        "review_priority_score": 20.0,
                        "action_code": "",
                        "manual_status": "watch",
                        "manual_status_normalized": "watch",
                        "manual_review_state": "reviewed_watch",
                        "entry_date": "2024-04-01",
                        "entry_close": 20.0,
                        "price_source": "local",
                        "outcome_status": "pending",
                        "return_1d": pd.NA,
                        "outcome_status_1d": "pending",
                        "broker_action": "none",
                        "research_only": True,
                    },
                ]
            )

            history, payload = sync_stock_target_review_outcomes_history(current, history_path, snapshot_path)

            self.assertTrue(history_path.exists())
            self.assertTrue(snapshot_path.exists())
            self.assertEqual(len(history), 2)
            one = history[history["code"] == "000001"].iloc[0]
            self.assertEqual(one["outcome_status"], "partial")
            self.assertAlmostEqual(float(one["return_1d"]), 0.10)
            self.assertEqual(int(payload["history_row_count"]), 2)
            self.assertEqual(int(payload["history_updated_row_count"]), 2)
            self.assertEqual(int(payload["history_replaced_row_count"]), 1)
            self.assertEqual(int(payload["history_inserted_row_count"]), 1)
            self.assertEqual(int(payload["history_partial_count"]), 1)
            self.assertEqual(int(payload["history_pending_count"]), 1)
            self.assertEqual(payload["history_latest_review_date"], "2024-04-01")

    def test_stock_target_review_outcomes_history_refreshes_matured_pending_rows_from_price_source(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            data_dir = root / "data" / "processed" / "stocks"
            data_dir.mkdir(parents=True)
            price_path = data_dir / "000001.csv"
            pd.DataFrame(
                {
                    "date": ["2024-04-01", "2024-04-02"],
                    "close": [10.0, 11.0],
                }
            ).to_csv(price_path, index=False)

            history_path = root / "outputs" / "research" / "stock_target_review_outcomes_history.csv"
            snapshot_path = root / "paper" / "stock_target_review_outcomes_history.csv"
            history_path.parent.mkdir(parents=True)
            pd.DataFrame(
                [
                    {
                        "date": "2024-04-01",
                        "layer": "core",
                        "code": "000001",
                        "name": "One",
                        "review_bucket": "drawdown_review",
                        "review_stage": "review_required",
                        "entry_date": "2024-04-01",
                        "entry_close": 10.0,
                        "price_source": str(price_path),
                        "outcome_status": "pending",
                        "future_date_1d": pd.NA,
                        "future_close_1d": pd.NA,
                        "return_1d": pd.NA,
                        "outcome_status_1d": "pending",
                        "outcome_status_5d": "pending",
                        "outcome_status_10d": "pending",
                        "outcome_status_20d": "pending",
                        "broker_action": "none",
                        "research_only": True,
                    }
                ]
            ).to_csv(history_path, index=False)

            current = pd.DataFrame(columns=paper_account_module.STOCK_TARGET_REVIEW_OUTCOME_COLUMNS)

            history, payload = sync_stock_target_review_outcomes_history(current, history_path, snapshot_path)

            row = history.iloc[0]
            self.assertEqual(row["future_date_1d"], "2024-04-02")
            self.assertAlmostEqual(float(row["future_close_1d"]), 11.0)
            self.assertAlmostEqual(float(row["return_1d"]), 0.10)
            self.assertEqual(row["outcome_status_1d"], "available")
            self.assertEqual(row["outcome_status"], "partial")
            self.assertEqual(int(payload["history_matured_row_count"]), 1)
            self.assertEqual(int(payload["history_partial_count"]), 1)
            self.assertEqual(int(payload["history_pending_count"]), 0)

    def test_stock_target_review_outcomes_history_report_separates_current_and_matured_counts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            output_dir = Path(tmp)
            payload = {
                "generated_at": "2024-04-03T15:30:00",
                "history_row_count": 40,
                "history_current_row_count": 9,
                "history_matured_row_count": 22,
                "history_updated_row_count": 31,
                "history_replaced_row_count": 9,
                "history_inserted_row_count": 0,
                "history_complete_count": 0,
                "history_partial_count": 31,
                "history_pending_count": 9,
                "history_missing_entry_price_count": 0,
                "history_latest_review_date": "2026-06-15",
                "history_horizon_summary": {},
                "broker_action": "none",
            }

            report = paper_account_module._render_stock_target_review_outcomes_history_report(
                pd.DataFrame(),
                payload,
                output_dir,
            )

            self.assertIn("| Current-run outcome rows | 9 |", report)
            self.assertIn("| Matured history rows refreshed | 22 |", report)
            self.assertIn("| Total rows touched | 31 |", report)

    def test_stock_target_review_outcome_analysis_groups_history_returns(self) -> None:
        history = pd.DataFrame(
            [
                {
                    "date": "2024-04-01",
                    "layer": "core",
                    "code": "000001",
                    "name": "One",
                    "review_bucket": "trigger_review",
                    "review_stage": "review_required",
                    "action_code": "review_required_pending",
                    "manual_status_normalized": "unreviewed",
                    "outcome_status": "complete",
                    "entry_date": "2024-04-01",
                    "return_1d": 0.10,
                    "outcome_status_1d": "available",
                    "return_5d": 0.20,
                    "outcome_status_5d": "available",
                    "outcome_status_10d": "pending",
                    "outcome_status_20d": "pending",
                    "broker_action": "none",
                    "research_only": True,
                },
                {
                    "date": "2024-04-01",
                    "layer": "core",
                    "code": "000002",
                    "name": "Two",
                    "review_bucket": "trigger_review",
                    "review_stage": "review_required",
                    "action_code": "review_required_pending",
                    "manual_status_normalized": "unreviewed",
                    "outcome_status": "complete",
                    "entry_date": "2024-04-01",
                    "return_1d": -0.05,
                    "outcome_status_1d": "available",
                    "return_5d": -0.10,
                    "outcome_status_5d": "available",
                    "outcome_status_10d": "pending",
                    "outcome_status_20d": "pending",
                    "broker_action": "none",
                    "research_only": True,
                },
                {
                    "date": "2024-04-01",
                    "layer": "satellite",
                    "code": "000003",
                    "name": "Three",
                    "review_bucket": "watch_review",
                    "review_stage": "monitor",
                    "action_code": "",
                    "manual_status_normalized": "watch",
                    "outcome_status": "pending",
                    "entry_date": "2024-04-01",
                    "return_1d": pd.NA,
                    "outcome_status_1d": "pending",
                    "return_5d": pd.NA,
                    "outcome_status_5d": "pending",
                    "outcome_status_10d": "pending",
                    "outcome_status_20d": "pending",
                    "broker_action": "none",
                    "research_only": True,
                },
            ]
        )

        analysis, payload = build_stock_target_review_outcome_analysis(
            history,
            {
                "history_path": "history.csv",
                "history_row_count": 3,
                "history_latest_review_date": "2024-04-01",
            },
            min_evaluable=2,
            min_group_evaluable=2,
        )

        overall = analysis[analysis["dimension"] == "overall"].iloc[0]
        action = analysis[
            (analysis["dimension"] == "action_code") & (analysis["group_value"] == "review_required_pending")
        ].iloc[0]
        self.assertEqual(payload["analysis_status"], "ready_for_review")
        self.assertEqual(int(payload["analysis_row_count"]), len(analysis))
        self.assertEqual(payload["min_evaluable"], 2)
        self.assertEqual(payload["min_group_evaluable"], 2)
        self.assertEqual(payload["ready_horizon_count"], 2)
        self.assertIn("1d", payload["ready_horizons"])
        self.assertEqual(payload["horizon_readiness"]["1d"]["status"], "ready_for_group_review")
        self.assertEqual(payload["horizon_readiness"]["10d"]["status"], "waiting_for_evaluable_returns")
        self.assertEqual(payload["total_evaluable_by_horizon"]["1d"], 2)
        self.assertEqual(payload["maturity_forecast"]["1d"]["pending_count"], 1)
        self.assertEqual(payload["maturity_forecast"]["1d"]["estimated_next_evaluable_date"], "2024-04-02")
        self.assertEqual(payload["maturity_forecast"]["5d"]["estimated_next_evaluable_date"], "2024-04-08")
        self.assertEqual(payload["maturity_forecast"]["20d"]["basis"], "weekday_estimate_without_holiday_calendar")
        self.assertEqual(int(overall["history_row_count"]), 3)
        self.assertEqual(int(overall["evaluable_1d"]), 2)
        self.assertAlmostEqual(float(overall["avg_return_1d"]), 0.025)
        self.assertAlmostEqual(float(overall["win_rate_1d"]), 0.5)
        self.assertEqual(int(action["history_row_count"]), 2)
        self.assertEqual(int(action["evaluable_5d"]), 2)
        self.assertAlmostEqual(float(action["avg_return_5d"]), 0.05)
        self.assertEqual(payload["best_groups"]["1d"]["group_value"], "review_required_pending")

        calendar, calendar_payload = build_stock_target_review_outcome_calendar(payload, as_of_date="2024-04-02")
        self.assertEqual(list(calendar["horizon"]), ["1d", "5d", "10d", "20d"])
        self.assertEqual(int(calendar_payload["calendar_row_count"]), 4)
        self.assertEqual(int(calendar_payload["calendar_ready_count"]), 2)
        self.assertEqual(int(calendar_payload["calendar_pending_count"]), 8)
        self.assertEqual(int(calendar_payload["calendar_due_count"]), 1)
        self.assertEqual(int(calendar_payload["calendar_due_pending_count"]), 1)
        self.assertEqual(calendar_payload["next_action_date"], "2024-04-02")
        self.assertEqual(calendar_payload["next_action_horizon"], "1d")
        self.assertEqual(calendar_payload["next_due_date"], "2024-04-02")
        self.assertEqual(calendar_payload["next_due_horizon"], "1d")
        one_day = calendar[calendar["horizon"] == "1d"].iloc[0]
        self.assertEqual(one_day["readiness_status"], "ready_for_group_review")
        self.assertEqual(one_day["maturity_state"], "due_for_recheck")
        self.assertEqual(one_day["estimated_next_evaluable_date"], "2024-04-02")
        self.assertEqual(int(one_day["days_until_next_evaluable"]), 0)
        self.assertTrue(bool(one_day["research_only"]))
        self.assertEqual(one_day["broker_action"], "none")
        due, due_payload = build_stock_target_review_outcome_due_queue(calendar, calendar_payload)
        self.assertEqual(int(due_payload["due_row_count"]), 1)
        self.assertEqual(int(due_payload["due_pending_count"]), 1)
        self.assertEqual(due_payload["next_due_date"], "2024-04-02")
        self.assertEqual(due_payload["next_due_horizon"], "1d")
        self.assertEqual(due.iloc[0]["horizon"], "1d")

        _, strict_payload = build_stock_target_review_outcome_analysis(
            history,
            {"history_path": "history.csv", "history_row_count": 3, "history_latest_review_date": "2024-04-01"},
            min_evaluable=20,
            min_group_evaluable=5,
        )
        self.assertEqual(strict_payload["analysis_status"], "sample_insufficient")
        self.assertEqual(strict_payload["horizon_readiness"]["1d"]["status"], "insufficient_overall_sample")

    def test_stock_target_review_outcome_calendar_due_counts_only_due_rows(self) -> None:
        history = pd.DataFrame(
            [
                {
                    "date": "2024-04-01",
                    "layer": "core",
                    "code": "000001",
                    "name": "One",
                    "review_bucket": "trigger_review",
                    "review_stage": "review_required",
                    "action_code": "review_required_pending",
                    "manual_status_normalized": "unreviewed",
                    "outcome_status": "pending",
                    "entry_date": "2024-04-01",
                    "return_1d": 0.01,
                    "outcome_status_1d": "available",
                    "return_5d": pd.NA,
                    "outcome_status_5d": "pending",
                    "return_10d": pd.NA,
                    "outcome_status_10d": "pending",
                    "return_20d": pd.NA,
                    "outcome_status_20d": "pending",
                    "broker_action": "none",
                    "research_only": True,
                },
                {
                    "date": "2024-04-04",
                    "layer": "core",
                    "code": "000002",
                    "name": "Two",
                    "review_bucket": "trigger_review",
                    "review_stage": "review_required",
                    "action_code": "review_required_pending",
                    "manual_status_normalized": "unreviewed",
                    "outcome_status": "pending",
                    "entry_date": "2024-04-04",
                    "return_1d": -0.01,
                    "outcome_status_1d": "available",
                    "return_5d": pd.NA,
                    "outcome_status_5d": "pending",
                    "return_10d": pd.NA,
                    "outcome_status_10d": "pending",
                    "return_20d": pd.NA,
                    "outcome_status_20d": "pending",
                    "broker_action": "none",
                    "research_only": True,
                },
            ]
        )

        analysis, payload = build_stock_target_review_outcome_analysis(
            history,
            {"history_path": "history.csv", "history_row_count": 2, "history_latest_review_date": "2024-04-04"},
            min_evaluable=1,
            min_group_evaluable=1,
        )
        calendar, calendar_payload = build_stock_target_review_outcome_calendar(payload, as_of_date="2024-04-09")
        five_day_row = calendar[calendar["horizon"] == "5d"].iloc[0]
        self.assertEqual(int(five_day_row["pending_count"]), 2)
        self.assertEqual(int(five_day_row["due_pending_count"]), 1)
        self.assertEqual(calendar_payload["calendar_pending_count"], 6)
        self.assertEqual(calendar_payload["calendar_due_count"], 1)
        self.assertEqual(calendar_payload["calendar_due_pending_count"], 1)
        due, due_payload = build_stock_target_review_outcome_due_queue(calendar, calendar_payload)
        self.assertEqual(int(due_payload["due_row_count"]), 1)
        self.assertEqual(int(due_payload["due_pending_count"]), 1)

    def test_dashboard_writes_unified_local_overview(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            daily = root / "daily_model_check_20240402"
            paper = root / "paper_account_20240402"
            sentiment = root / "market_sentiment_state_20240402"
            cache = root / "data" / "processed"
            allocator = root / "allocator"
            trigger = root / "latest_trigger.md"
            pipeline_history = root / "pipeline_history"
            daily_run_status = root / "daily_run_status"
            allocator_observation = root / "allocator_observation"
            daily.mkdir()
            paper.mkdir()
            sentiment.mkdir()
            cache.mkdir(parents=True)
            allocator.mkdir()
            pipeline_history.mkdir()
            daily_run_status.mkdir()
            allocator_observation.mkdir()
            (daily / "daily_model_check_snapshot.json").write_text(
                json.dumps(
                    {
                        "generated_at": "2024-04-02T16:00:00",
                        "as_of_date": "2024-04-02",
                        "latest_date": "2024-04-02",
                        "data_freshness_status": "fresh_enough",
                        "phase2_posture": "core_base_allocator_gated_satellite",
                        "action_posture": "review_core_base_allocator_gate",
                        "selected_candidate": "sat20_ma60_drop03_unfiltered",
                        "risk_on_satellite_weight": 0.2,
                        "current_drawdown": -0.03,
                        "return_20d": 0.01,
                        "allocator_sharpe": 1.2,
                        "promotion_status": "ok",
                        "promotion_decision": "promote_candidate",
                        "promotion_return_edge": 0.062,
                        "promotion_sharpe_edge": 0.119,
                        "promotion_sensitivity_support_count": 2,
                        "promotion_sensitivity_group_support_count": 1,
                        "promotion_support_group_count": 2,
                        "promotion_evidence_support_count": 1,
                        "promotion_sensitivity_run_support_count": 2,
                        "promotion_min_sensitivity_support": 2,
                        "phase2_report_path": "phase2_review.md",
                        "phase2_snapshot_path": "phase2_review_snapshot.json",
                    }
                ),
                encoding="utf-8",
            )
            (daily / "daily_model_check.md").write_text("# Daily Model Check\n", encoding="utf-8")
            (paper / "metrics.json").write_text(
                json.dumps(
                    {
                        "generated_at": "2024-04-02T16:02:00",
                        "latest_date": "2024-04-02",
                        "latest_window": "pf_01",
                        "latest_candidate": "sat20_ma60_drop03_unfiltered",
                        "latest_regime": "risk_on",
                        "latest_core_weight": 0.8,
                        "latest_satellite_weight": 0.2,
                        "latest_cash_weight": 0.0,
                        "final_equity": 1100000.0,
                        "total_return": 0.1,
                        "max_drawdown": -0.05,
                        "current_drawdown": -0.02,
                        "sharpe": 1.1,
                        "audit_event_count": 3,
                    }
                ),
                encoding="utf-8",
            )
            (paper / "paper_account.md").write_text("# Paper Account Ledger\n", encoding="utf-8")
            pd.DataFrame({"date": pd.to_datetime(["2024-04-01", "2024-04-02"]), "close": [10.0, 10.2]}).to_csv(
                cache / "000001.csv",
                index=False,
            )
            (allocator / "selected_params.json").write_text(
                json.dumps(
                    {
                        "weights": {
                            "risk_on": {"core": 0.8, "satellite": 0.2, "cash": 0.0},
                            "risk_off": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                            "crash": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                        },
                        "satellite_filter": {"enabled": False},
                        "regime_overrides": {},
                    }
                ),
                encoding="utf-8",
            )
            pd.DataFrame(
                {"date": pd.to_datetime(["2024-04-01", "2024-04-02"]), "window": ["pf_01", "pf_01"], "stitched_equity": [100, 101]}
            ).to_csv(allocator / "oos_equity_stitched.csv", index=False)
            pd.DataFrame(
                {
                    "window": ["pf_01"],
                    "test_end": [20240402],
                    "selected_params_path": ["selected_params.json"],
                }
            ).to_csv(allocator / "portfolio_walk_forward_summary.csv", index=False)
            (sentiment / "latest_market_sentiment.json").write_text(
                json.dumps(
                    {
                        "date": "2024-04-02",
                        "sentiment_state": "warm",
                        "sentiment_score": 0.5,
                        "reference_exposure": 1.0,
                        "coverage_count": 100,
                        "advance_ratio": 0.7,
                        "limit_up_count": 10,
                        "limit_down_count": 1,
                        "market_return": 2.0,
                    }
                ),
                encoding="utf-8",
            )
            (sentiment / "latest_market_sentiment.md").write_text("# Market Sentiment State\n", encoding="utf-8")
            trigger.write_text(
                "# THSDK strategy monitor 2024-04-02_154500\n\n- Trigger count: 3\n",
                encoding="utf-8",
            )
            model_audit = root / "model_audit"
            model_audit.mkdir()
            (model_audit / "model_build_audit_snapshot.json").write_text(
                json.dumps(
                    {
                        "generated_at": "2024-04-02T16:00:00",
                        "duplicate_config_groups": 0,
                        "root_configs_without_extends": 0,
                        "walk_forward_action_items": 2,
                        "walk_forward_resume_candidates": 1,
                        "walk_forward_archive_review_candidates": 1,
                    }
                ),
                encoding="utf-8",
            )
            (model_audit / "model_build_audit.md").write_text("# Model Build Audit\n", encoding="utf-8")
            pd.DataFrame(
                {
                    "run_id": ["wf_partial"],
                    "issue_type": ["partial_window_outputs"],
                    "recommended_action": ["resume_or_finalize"],
                    "priority": ["high"],
                    "requires_manual_confirmation": [False],
                }
            ).to_csv(model_audit / "walk_forward_run_actions.csv", index=False)
            (pipeline_history / "pipeline_history_review_snapshot.json").write_text(
                json.dumps(
                    {
                        "generated_at": "2024-04-02T16:10:00",
                        "health_state": "watch",
                        "alert_count": 2,
                        "latest_as_of_date": "2024-04-02",
                        "latest_satellite_risk_budget_status": "completed",
                        "latest_satellite_risk_budget_decision": "eligible_for_small_satellite_trial",
                        "previous_satellite_risk_budget_decision": "wait_for_outcome_samples",
                        "satellite_risk_budget_decision_changed": True,
                        "latest_satellite_risk_budget_recommended_satellite_weight": 0.05,
                        "satellite_risk_budget_recommended_weight_change": 0.05,
                        "latest_satellite_risk_budget_selected_horizon": "1d",
                        "latest_satellite_risk_budget_report_path": "risk/satellite_risk_budget_review.md",
                    }
                ),
                encoding="utf-8",
            )
            (pipeline_history / "pipeline_history_review.md").write_text("# Pipeline History Review\n", encoding="utf-8")
            (allocator_observation / "allocator_observation_snapshot.json").write_text(
                json.dumps(
                    {
                        "generated_at": "2024-04-02T16:20:00",
                        "status": "allocator_observation_completed",
                        "as_of_date": "2024-04-02",
                        "observation_status": "outcomes_ready_for_review",
                        "next_action_stage": "review_outcome_analysis",
                        "risk_budget_decision": "eligible_for_budget_review",
                        "next_review_date": "2024-04-03",
                        "next_review_horizon": "5d",
                        "outcome_analysis_status": "ready",
                        "outcome_ready_horizon_count": 2,
                        "outcome_due_count": 1,
                        "blocking_items": [],
                        "monitor_items": ["manual_pending_rows=1"],
                        "pipeline_snapshot": "daily_pipeline_snapshot.json",
                    }
                ),
                encoding="utf-8",
            )
            (allocator_observation / "allocator_observation.md").write_text(
                "# Post-Switch Allocator Observation\n",
                encoding="utf-8",
            )
            (daily_run_status / "daily_run_status_snapshot.json").write_text(
                json.dumps(
                    {
                        "run_state": "outcome_ready",
                        "run_state_severity": "ok",
                        "problem_state": False,
                        "latest_live_preflight_status": "live_preflight_completed",
                        "latest_live_preflight_decision": "ready_for_manual_broker_reconciliation",
                        "latest_live_preflight_status_path": str(root / "outputs" / "logs" / "live_preflight_20240402_163000.status.json"),
                        "latest_live_preflight_snapshot_path": str(root / "outputs" / "research" / "live_preflight_20240402" / "live_preflight_snapshot.json"),
                        "latest_live_preflight_report_path": str(root / "outputs" / "research" / "live_preflight_20240402" / "live_preflight.md"),
                        "latest_live_preflight_blocking_items_count": 1,
                        "latest_live_preflight_monitor_items_count": 0,
                        "latest_live_preflight_live_shadow_review_decisions_path": str(
                            root / "outputs" / "research" / "live_shadow_review_decisions_20240402" / "live_shadow_review_decisions.csv"
                        ),
                        "latest_live_preflight_live_shadow_review_decision_count": 4,
                        "latest_live_preflight_live_shadow_review_blocking_decision_count": 1,
                        "latest_live_preflight_live_shadow_review_monitor_decision_count": 2,
                        "latest_live_preflight_live_shadow_review_unknown_decision_count": 1,
                        "latest_observation_snapshot_path": str(
                            allocator_observation / "allocator_observation_snapshot.json"
                        ),
                        "latest_observation_status_path": str(
                            root / "outputs" / "logs" / "allocator_observation_20240402_162000.status.json"
                        ),
                    }
                ),
                encoding="utf-8",
            )

            result = run_daily_dashboard(
                project_root=root,
                output_dir=root / "dashboard",
                daily_check_dir=daily,
                paper_account_dir=paper,
                sentiment_dir=sentiment,
                data_cache_dir=cache,
                allocator_dir=allocator,
                trigger_report=trigger,
                model_audit_dir=model_audit,
                pipeline_history_dir=pipeline_history,
                daily_run_status_dir=daily_run_status,
                as_of_date="2024-04-02",
                date_stamp=True,
            )

            self.assertTrue(result.report_path.exists())
            self.assertTrue(result.snapshot_path.exists())
            self.assertEqual(result.output_dir.name, "dashboard_20240402")
            self.assertEqual(result.snapshot["dashboard_posture"], "core_base_watch_allocator_gate")
            self.assertEqual(result.snapshot["trigger_freshness_status"], "fresh_enough")
            self.assertEqual(result.snapshot["sentiment_state"], "warm")
            self.assertEqual(result.snapshot["paper_account_status"], "ok")
            self.assertEqual(result.snapshot["paper_freshness_status"], "fresh_enough")
            self.assertEqual(result.snapshot["market_cache_status"], "fresh_enough")
            self.assertEqual(result.snapshot["allocator_input_status"], "fresh_enough")
            self.assertEqual(result.snapshot["paper_latest_regime"], "risk_on")
            self.assertEqual(result.snapshot["model_audit_status"], "needs_attention")
            self.assertEqual(result.snapshot["model_audit_walk_forward_action_items"], 2)
            self.assertEqual(result.snapshot["model_audit_top_action"], "resume_or_finalize")
            self.assertEqual(result.snapshot["promotion_decision"], "promote_candidate")
            self.assertEqual(int(result.snapshot["promotion_support_group_count"]), 2)
            self.assertEqual(int(result.snapshot["promotion_evidence_support_count"]), 1)
            self.assertEqual(int(result.snapshot["promotion_sensitivity_group_support_count"]), 1)
            self.assertEqual(int(result.snapshot["promotion_sensitivity_run_support_count"]), 2)
            self.assertEqual(result.snapshot["pipeline_history_status"], "ok")
            self.assertEqual(result.snapshot["pipeline_history_health_state"], "watch")
            self.assertEqual(result.snapshot["pipeline_history_latest_satellite_risk_budget_decision"], "eligible_for_small_satellite_trial")
            self.assertTrue(result.snapshot["pipeline_history_satellite_risk_budget_decision_changed"])
            self.assertAlmostEqual(result.snapshot["pipeline_history_latest_satellite_risk_budget_recommended_satellite_weight"], 0.05)
            self.assertEqual(result.snapshot["daily_run_state"], "outcome_ready")
            self.assertEqual(result.snapshot["live_preflight_status"], "live_preflight_completed")
            self.assertEqual(result.snapshot["live_preflight_decision"], "ready_for_manual_broker_reconciliation")
            self.assertEqual(result.snapshot["live_preflight_blocking_items_count"], 1)
            self.assertEqual(result.snapshot["live_preflight_monitor_items_count"], 0)
            self.assertEqual(result.snapshot["live_preflight_live_shadow_review_decision_count"], 4)
            self.assertEqual(result.snapshot["live_preflight_live_shadow_review_blocking_decision_count"], 1)
            self.assertEqual(result.snapshot["live_preflight_live_shadow_review_monitor_decision_count"], 2)
            self.assertEqual(result.snapshot["live_preflight_live_shadow_review_unknown_decision_count"], 1)
            self.assertEqual(result.snapshot["allocator_observation_status"], "ok")
            self.assertEqual(result.snapshot["allocator_observation_decision_status"], "outcomes_ready_for_review")
            self.assertEqual(result.snapshot["allocator_observation_next_action_stage"], "review_outcome_analysis")
            self.assertEqual(result.snapshot["allocator_observation_risk_budget_decision"], "eligible_for_budget_review")
            self.assertEqual(result.snapshot["allocator_observation_outcome_ready_horizon_count"], 2)
            self.assertEqual(result.snapshot["allocator_observation_next_review_date"], "2024-04-03")
            report_text = result.report_path.read_text(encoding="utf-8")
            self.assertIn("Latest Dashboard", report_text)
            self.assertIn("Freshness Gates", report_text)
            self.assertIn("Model Build Hygiene", report_text)
            self.assertIn("Allocator Observation", report_text)
            self.assertIn("Live preflight", report_text)
            self.assertIn("Live-shadow review decisions", report_text)
            self.assertIn("outcomes_ready_for_review", report_text)
            self.assertIn("eligible_for_budget_review", report_text)
            self.assertIn("Pipeline History", report_text)
            self.assertIn("eligible_for_small_satellite_trial", report_text)
            self.assertIn("Recommended satellite budget | 5.00%", report_text)
            self.assertIn("Decision changed | True", report_text)
            self.assertIn("Allocator Promotion Watch", report_text)
            self.assertIn("Independent support groups | 2 / 2", report_text)
            self.assertIn("Sensitivity group support | 1", report_text)
            self.assertIn("Evidence group support | 1", report_text)
            self.assertIn("Sensitivity run support | 2", report_text)
            self.assertIn("Paper Account", report_text)
            self.assertIn("Market average return | 2.00%", report_text)

    def test_daily_pipeline_runs_daily_check_paper_account_and_dashboard(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            configs_dir = root / "configs"
            curves_dir = root / "curves"
            allocator = root / "allocator"
            phase2 = root / "phase2"
            promotion = root / "promotion"
            sentiment = root / "sentiment"
            model_audit = root / "model_audit"
            cache = root / "data" / "processed"
            trigger = root / "latest_trigger.md"
            history = root / "history.csv"
            live_holdings = root / "live_holdings.csv"
            live_prices = root / "live_prices.csv"
            configs_dir.mkdir()
            curves_dir.mkdir()
            allocator.mkdir()
            phase2.mkdir()
            promotion.mkdir()
            sentiment.mkdir()
            model_audit.mkdir()
            cache.mkdir(parents=True)
            dates = pd.date_range("2020-01-01", periods=8, freq="D")
            pd.DataFrame({"date": dates, "equity": [100, 101, 102, 104, 105, 106, 107, 108]}).to_csv(
                curves_dir / "core.csv",
                index=False,
            )
            pd.DataFrame({"date": dates, "equity": [100, 102, 104, 103, 107, 109, 108, 112]}).to_csv(
                curves_dir / "satellite.csv",
                index=False,
            )
            pd.DataFrame({"date": dates, "close": [10, 11, 12, 13, 14, 15, 16, 17]}).to_csv(
                curves_dir / "benchmark.csv",
                index=False,
            )
            pd.DataFrame({"date": dates, "close": [10, 11, 12, 13, 14, 15, 16, 17]}).to_csv(
                cache / "510300.csv",
                index=False,
            )
            config_path = configs_dir / "portfolio.yaml"
            config_path.write_text(
                f"""
project:
  name: pipeline_test
  initial_cash: 1000000
  output_dir: {str(root / "outputs" / "portfolios")}
curves:
  core:
    path: {str(curves_dir / "core.csv")}
    equity_column: equity
  satellite:
    path: {str(curves_dir / "satellite.csv")}
    equity_column: equity
regime:
  benchmark_path: {str(curves_dir / "benchmark.csv")}
  benchmark_close_column: close
  ma_window: 2
  drop_window: 1
  risk_on_drop_threshold: -0.20
  crash_drop_threshold: -0.50
  default_regime: risk_off
weights:
  risk_on:
    core: 0.8
    satellite: 0.2
  risk_off:
    core: 1.0
    satellite: 0.0
  crash:
    core: 1.0
    satellite: 0.0
""",
                encoding="utf-8",
            )
            params = allocator / "params.json"
            params.write_text(
                json.dumps(
                    {
                        "weights": {
                            "risk_on": {"core": 0.8, "satellite": 0.2, "cash": 0.0},
                            "risk_off": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                            "crash": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                        },
                        "satellite_filter": {"enabled": False},
                        "regime_overrides": {
                            "ma_window": 2,
                            "risk_on_drop_threshold": -0.2,
                            "crash_drop_threshold": -0.5,
                        },
                    }
                ),
                encoding="utf-8",
            )
            pd.DataFrame(
                {
                    "window": ["pf_01"],
                    "train_start": [20190101],
                    "train_end": [20191231],
                    "test_start": [20200101],
                    "test_end": [20200108],
                    "selected_candidate": ["sat20_test"],
                    "selected_score": [0.4],
                    "selected_params_path": ["params.json"],
                    "test_total_return": [0.08],
                    "test_max_drawdown": [-0.03],
                    "test_sharpe": [1.2],
                    "test_average_core_weight": [0.85],
                    "test_average_satellite_weight": [0.15],
                    "test_risk_on_day_ratio": [0.75],
                    "test_risk_off_day_ratio": [0.25],
                    "test_crash_day_ratio": [0.0],
                }
            ).to_csv(allocator / "portfolio_walk_forward_summary.csv", index=False)
            pd.DataFrame(
                {"date": dates, "window": ["pf_01"] * len(dates), "stitched_equity": [100, 101, 102, 104, 105, 106, 107, 108]}
            ).to_csv(allocator / "oos_equity_stitched.csv", index=False)
            pd.DataFrame(
                {
                    "layer": ["core", "portfolio", "allocator"],
                    "status": ["complete", "complete", "complete"],
                    "total_return": [0.08, 0.05, 0.08],
                    "max_drawdown": [-0.03, -0.04, -0.03],
                    "sharpe": [1.0, 0.8, 1.2],
                }
            ).to_csv(phase2 / "phase2_components.csv", index=False)
            (promotion / "allocator_promotion_snapshot.json").write_text(
                json.dumps(
                    {
                        "status": "ok",
                        "decision": "promote_candidate",
                        "candidate_passes_headline_gate": True,
                        "sensitivity_run_support_count": 2,
                        "sensitivity_group_support_count": 1,
                        "support_group_count": 2,
                        "evidence_support_count": 1,
                        "sensitivity_support_count": 2,
                        "min_sensitivity_support": 2,
                        "return_edge": 0.062,
                        "sharpe_edge": 0.119,
                        "drawdown_change": 0.019,
                    }
                ),
                encoding="utf-8",
            )
            (promotion / "allocator_promotion_review.md").write_text("# Allocator Promotion Review\n", encoding="utf-8")
            (sentiment / "latest_market_sentiment.json").write_text(
                json.dumps(
                    {
                        "date": "2020-01-08",
                        "sentiment_state": "warm",
                        "sentiment_score": 0.6,
                        "reference_exposure": 1.0,
                        "coverage_count": 50,
                        "advance_ratio": 0.6,
                        "limit_up_count": 3,
                        "limit_down_count": 0,
                        "market_return": 1.2,
                    }
                ),
                encoding="utf-8",
            )
            (sentiment / "latest_market_sentiment.md").write_text("# Sentiment\n", encoding="utf-8")
            trigger.write_text("# THSDK strategy monitor 2020-01-08_154500\n\n- Trigger count: 1\n", encoding="utf-8")
            (model_audit / "model_build_audit_snapshot.json").write_text(
                json.dumps(
                    {
                        "generated_at": "2020-01-08T16:00:00",
                        "duplicate_config_groups": 0,
                        "root_configs_without_extends": 0,
                        "walk_forward_action_items": 2,
                        "walk_forward_resume_candidates": 1,
                        "walk_forward_archive_review_candidates": 1,
                    }
                ),
                encoding="utf-8",
            )
            (model_audit / "model_build_audit.md").write_text("# Model Build Audit\n", encoding="utf-8")
            pd.DataFrame(
                {
                    "run_id": ["wf_partial"],
                    "issue_type": ["partial_window_outputs"],
                    "recommended_action": ["resume_or_finalize"],
                    "priority": ["high"],
                    "requires_manual_confirmation": [False],
                }
            ).to_csv(model_audit / "walk_forward_run_actions.csv", index=False)
            pd.DataFrame(
                {
                    "generated_at": ["2020-01-07T16:00:00"],
                    "as_of_date": ["2020-01-07"],
                    "dashboard_posture": ["stale_model_data"],
                    "action_posture": ["wait_for_fresh_data"],
                    "market_cache_status": ["stale"],
                    "allocator_input_status": ["fresh_enough"],
                    "paper_latest_regime": ["risk_off"],
                    "paper_latest_candidate": ["core_only"],
                    "paper_latest_satellite_weight": [0.0],
                    "paper_final_equity": [950000.0],
                    "paper_total_return": [-0.05],
                }
            ).to_csv(history, index=False)
            pd.DataFrame(
                [
                    {"code": "000001", "name": "One", "quantity": 100, "current_price": 17.0},
                    {"code": "000002", "name": "Two", "quantity": 0, "current_price": 22.0},
                ]
            ).to_csv(live_holdings, index=False)
            pd.DataFrame(
                [
                    {"code": "000001", "price": 17.0},
                    {"code": "000002", "price": 22.0},
                    {"code": "000003", "price": 6.0},
                ]
            ).to_csv(live_prices, index=False)

            result = run_daily_pipeline(
                project_root=root,
                output_dir=root / "pipeline",
                components_path=phase2,
                allocator_dir=allocator,
                promotion_review_dir=promotion,
                portfolio_config_path=config_path,
                sentiment_dir=sentiment,
                data_cache_dir=curves_dir,
                trigger_report=trigger,
                model_audit_dir=model_audit,
                history_path=history,
                live_shadow_holdings_file=live_holdings,
                live_shadow_prices_file=live_prices,
                live_shadow_cash=20000.0,
                live_shadow_output_dir=root / "live_shadow",
                as_of_date="2020-01-08",
                date_stamp=True,
            )

            self.assertTrue(result.report_path.exists())
            self.assertTrue(result.snapshot_path.exists())
            self.assertTrue(result.daily_check_result.report_path.exists())
            self.assertTrue(result.paper_account_result.report_path.exists())
            self.assertTrue(result.paper_account_result.target_holdings_report_path.exists())
            self.assertTrue(result.paper_account_result.stock_targets_report_path.exists())
            self.assertTrue(result.paper_account_result.stock_target_review_report_path.exists())
            self.assertTrue(result.paper_account_result.stock_target_review_notes_path.exists())
            self.assertTrue(result.paper_account_result.stock_target_review_notes_snapshot_path.exists())
            self.assertTrue(result.paper_account_result.stock_target_review_actions_report_path.exists())
            self.assertTrue(result.paper_account_result.stock_target_review_assistant_report_path.exists())
            self.assertTrue(result.paper_account_result.stock_target_review_outcomes_report_path.exists())
            self.assertTrue(result.paper_account_result.stock_target_review_outcomes_history_report_path.exists())
            self.assertTrue(result.paper_account_result.stock_target_review_outcome_analysis_report_path.exists())
            self.assertTrue(result.paper_account_result.stock_target_review_outcome_calendar_report_path.exists())
            self.assertTrue(result.paper_account_result.stock_target_review_outcome_due_report_path.exists())
            self.assertTrue(result.dashboard_result.report_path.exists())
            self.assertIsNotNone(result.live_shadow_result)
            self.assertTrue(result.live_shadow_result.report_path.exists())
            self.assertTrue(result.live_shadow_result.orders_path.exists())
            self.assertTrue(result.live_shadow_result.reconcile_path.exists())
            self.assertIsNotNone(result.history_review_result)
            self.assertTrue(result.history_review_result.report_path.exists())
            self.assertTrue(result.alerts_result.report_path.exists())
            self.assertTrue(result.alerts_result.json_path.exists())
            self.assertEqual(result.history_path, history)
            self.assertEqual(result.output_dir.name, "pipeline_20200108")
            self.assertEqual(result.snapshot["data_freshness_status"], "fresh_enough")
            self.assertEqual(result.daily_check_result.snapshot["promotion_decision"], "promote_candidate")
            self.assertEqual(result.dashboard_result.snapshot["promotion_decision"], "promote_candidate")
            self.assertEqual(result.snapshot["promotion_decision"], "promote_candidate")
            self.assertEqual(int(result.daily_check_result.snapshot["promotion_support_group_count"]), 2)
            self.assertEqual(int(result.dashboard_result.snapshot["promotion_support_group_count"]), 2)
            self.assertEqual(int(result.snapshot["promotion_support_group_count"]), 2)
            self.assertEqual(int(result.snapshot["promotion_evidence_support_count"]), 1)
            self.assertEqual(int(result.daily_check_result.snapshot["promotion_sensitivity_group_support_count"]), 1)
            self.assertEqual(int(result.dashboard_result.snapshot["promotion_sensitivity_group_support_count"]), 1)
            self.assertEqual(int(result.snapshot["promotion_sensitivity_group_support_count"]), 1)
            self.assertEqual(int(result.snapshot["promotion_sensitivity_run_support_count"]), 2)
            self.assertAlmostEqual(float(result.snapshot["promotion_return_edge"]), 0.062)
            self.assertEqual(result.snapshot["trading_day_gate_status"], "trading_day_data_ready")
            self.assertEqual(result.snapshot["after_close_data_status"], "ready")
            self.assertEqual(result.snapshot["paper_target_holding_count"], 3)
            self.assertEqual(result.snapshot["paper_target_broker_action"], "none")
            self.assertTrue(str(result.snapshot["paper_target_holdings_report_path"]).endswith("target_holdings.md"))
            self.assertTrue(str(result.snapshot["paper_stock_targets_report_path"]).endswith("stock_targets.md"))
            self.assertTrue(str(result.snapshot["paper_stock_target_review_report_path"]).endswith("stock_target_review.md"))
            self.assertTrue(str(result.snapshot["paper_stock_target_review_actions_report_path"]).endswith("stock_target_review_actions.md"))
            self.assertTrue(str(result.snapshot["paper_stock_target_review_assistant_report_path"]).endswith("stock_target_review_assistant.md"))
            self.assertTrue(str(result.snapshot["paper_stock_target_review_assistant_path"]).endswith("stock_target_review_assistant.csv"))
            self.assertTrue(
                str(result.snapshot["paper_stock_target_review_decision_template_report_path"]).endswith(
                    "stock_target_review_decision_template.md"
                )
            )
            self.assertTrue(
                str(result.snapshot["paper_stock_target_review_decision_template_path"]).endswith(
                    "stock_target_review_decision_template.csv"
                )
            )
            self.assertTrue(
                str(result.snapshot["paper_stock_target_review_decision_template_xlsx_path"]).endswith(
                    "stock_target_review_decision_template.xlsx"
                )
            )
            self.assertTrue(str(result.snapshot["paper_stock_target_review_outcomes_report_path"]).endswith("stock_target_review_outcomes.md"))
            self.assertTrue(str(result.snapshot["paper_stock_target_review_outcomes_history_report_path"]).endswith("stock_target_review_outcomes_history.md"))
            self.assertTrue(str(result.snapshot["paper_stock_target_review_outcomes_history_path"]).endswith("stock_target_review_outcomes_history.csv"))
            self.assertTrue(str(result.snapshot["paper_stock_target_review_outcomes_history_snapshot_path"]).endswith("stock_target_review_outcomes_history.csv"))
            self.assertTrue(str(result.snapshot["paper_stock_target_review_outcome_analysis_report_path"]).endswith("stock_target_review_outcome_analysis.md"))
            self.assertTrue(str(result.snapshot["paper_stock_target_review_outcome_analysis_path"]).endswith("stock_target_review_outcome_analysis.csv"))
            self.assertTrue(str(result.snapshot["paper_stock_target_review_outcome_calendar_report_path"]).endswith("stock_target_review_outcome_calendar.md"))
            self.assertTrue(str(result.snapshot["paper_stock_target_review_outcome_calendar_path"]).endswith("stock_target_review_outcome_calendar.csv"))
            self.assertTrue(str(result.snapshot["paper_stock_target_review_outcome_due_report_path"]).endswith("stock_target_review_outcome_due.md"))
            self.assertTrue(str(result.snapshot["paper_stock_target_review_outcome_due_path"]).endswith("stock_target_review_outcome_due.csv"))
            self.assertTrue(str(result.snapshot["paper_stock_target_review_notes_path"]).endswith("stock_target_review_notes.csv"))
            self.assertTrue(str(result.snapshot["paper_stock_target_review_notes_snapshot_path"]).endswith("stock_target_review_notes.csv"))
            self.assertEqual(result.snapshot["live_shadow_status"], "completed")
            self.assertEqual(result.snapshot["live_shadow_trade_plan_status"], "manual_review_only")
            self.assertEqual(result.snapshot["live_shadow_broker_action"], "none")
            self.assertTrue(result.snapshot["live_shadow_research_only"])
            self.assertGreaterEqual(result.snapshot["live_shadow_order_count"], 1)
            self.assertTrue(str(result.snapshot["live_shadow_report_path"]).endswith("live_shadow_plan.md"))
            self.assertTrue(str(result.snapshot["live_shadow_orders_path"]).endswith("live_shadow_orders.csv"))
            self.assertTrue(str(result.snapshot["live_shadow_reconcile_path"]).endswith("live_shadow_reconcile.csv"))
            self.assertGreaterEqual(result.snapshot["paper_stock_target_count"], 0)
            self.assertGreaterEqual(result.snapshot["paper_stock_target_review_required_count"], 0)
            self.assertGreaterEqual(result.snapshot["paper_stock_target_review_unreviewed_count"], 0)
            self.assertGreaterEqual(result.snapshot["paper_stock_target_review_required_unreviewed_count"], 0)
            self.assertGreaterEqual(result.snapshot["paper_stock_target_review_manual_pending_count"], 0)
            self.assertGreaterEqual(result.snapshot["paper_stock_target_review_manual_watch_count"], 0)
            self.assertGreaterEqual(result.snapshot["paper_stock_target_review_manual_exclude_candidate_count"], 0)
            self.assertGreaterEqual(result.snapshot["paper_stock_target_review_action_count"], 0)
            self.assertGreaterEqual(result.snapshot["paper_stock_target_review_assistant_count"], 0)
            self.assertGreaterEqual(result.snapshot["paper_stock_target_review_assistant_pending_count"], 0)
            self.assertGreaterEqual(result.snapshot["paper_stock_target_review_decision_template_count"], 0)
            self.assertGreaterEqual(result.snapshot["paper_stock_target_review_decision_template_blank_status_count"], 0)
            self.assertGreaterEqual(result.snapshot["paper_stock_target_review_outcome_row_count"], 0)
            self.assertGreaterEqual(result.snapshot["paper_stock_target_review_outcomes_history_row_count"], 0)
            self.assertGreaterEqual(result.snapshot["paper_stock_target_review_outcome_analysis_row_count"], 0)
            self.assertEqual(result.snapshot["paper_stock_target_review_outcome_analysis_min_evaluable"], 20)
            self.assertEqual(result.snapshot["paper_stock_target_review_outcome_analysis_min_group_evaluable"], 5)
            self.assertGreaterEqual(result.snapshot["paper_stock_target_review_outcome_analysis_ready_horizon_count"], 0)
            self.assertEqual(result.snapshot["paper_stock_target_review_outcome_calendar_row_count"], 4)
            self.assertIn("paper_stock_target_review_outcome_calendar_next_action_date", result.snapshot)
            self.assertIn("paper_stock_target_review_outcome_due_row_count", result.snapshot)
            self.assertIn("paper_stock_target_review_outcome_due_next_date", result.snapshot)
            self.assertIn("paper_stock_target_review_outcome_maturity_next_1d_date", result.snapshot)
            self.assertIn("paper_stock_target_review_outcome_maturity_next_20d_date", result.snapshot)
            self.assertEqual(result.snapshot["satellite_risk_budget_status"], "completed")
            self.assertEqual(result.snapshot["satellite_risk_budget_decision"], "blocked_by_pipeline_gates")
            self.assertEqual(
                result.snapshot["satellite_risk_budget_reason"],
                "Resolve research pipeline blockers before changing satellite risk budget.",
            )
            self.assertTrue(str(result.snapshot["satellite_risk_budget_report_path"]).endswith("satellite_risk_budget_review.md"))
            self.assertTrue(str(result.snapshot["satellite_risk_budget_snapshot_path"]).endswith("satellite_risk_budget_snapshot.json"))
            self.assertTrue(str(result.snapshot["satellite_risk_budget_checklist_path"]).endswith("satellite_risk_budget_checklist.csv"))
            self.assertTrue(Path(result.snapshot["satellite_risk_budget_report_path"]).exists())
            self.assertEqual(result.snapshot["dashboard_posture"], "core_base_watch_allocator_gate")
            self.assertEqual(result.snapshot["model_audit_status"], "needs_attention")
            self.assertEqual(result.snapshot["model_audit_walk_forward_action_items"], 2)
            self.assertEqual(result.snapshot["model_audit_walk_forward_resume_candidates"], 1)
            self.assertEqual(result.snapshot["model_audit_walk_forward_archive_review_candidates"], 1)
            self.assertEqual(result.snapshot["alert_level"], "warning")
            self.assertEqual(result.snapshot["alert_action_stage"], "review_required")
            self.assertGreaterEqual(result.snapshot["alert_action_item_count"], 1)
            self.assertGreaterEqual(result.snapshot["alert_count"], 1)
            self.assertIn(
                result.snapshot["pipeline_next_step_stage"],
                {
                    "update_manual_review_notes",
                    "review_stock_target_actions",
                    "inspect_warning_alerts",
                    "accumulate_outcome_samples",
                    "refresh_outcome_due_queue",
                    "review_outcome_analysis",
                    "routine_monitor",
                },
            )
            self.assertTrue(str(result.snapshot["pipeline_next_step_action"]))
            self.assertIn("pipeline_user_intervention_required", result.snapshot)
            self.assertIn("satellite_weight_opened", result.alerts_result.report_path.read_text(encoding="utf-8"))
            self.assertIn("model_audit_needs_attention", result.alerts_result.report_path.read_text(encoding="utf-8"))
            self.assertIn("review_model_build_audit", result.alerts_result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Model build audit:", result.alerts_result.report_path.read_text(encoding="utf-8"))
            self.assertEqual(result.snapshot["history_status"], "appended")
            self.assertEqual(result.snapshot["history_review_status"], "completed")
            self.assertEqual(result.snapshot["history_review_alert_refresh_status"], "completed")
            self.assertEqual(result.alerts_result.payload["history_review_alert_refresh_status"], "completed")
            self.assertEqual(result.alerts_result.payload["history_review_latest_alert_level"], "warning")
            self.assertEqual(result.alerts_result.payload["history_review_latest_alert_action_stage"], "review_required")
            self.assertEqual(result.snapshot["dashboard_alert_refresh_status"], "completed")
            self.assertEqual(result.alerts_result.payload["dashboard_alert_refresh_status"], "completed")
            self.assertEqual(result.snapshot["history_review_health_state"], "watch")
            self.assertGreaterEqual(int(result.snapshot["history_review_alert_count"]), 1)
            self.assertEqual(result.history_review_result.snapshot["latest_alert_level"], "warning")
            self.assertEqual(result.history_review_result.snapshot["latest_alert_action_stage"], "review_required")
            self.assertTrue(
                any("Daily alert level is warning" in item for item in result.history_review_result.snapshot["alerts"])
            )
            self.assertEqual(result.snapshot["history_previous_as_of_date"], "2020-01-07")
            self.assertGreater(result.snapshot["history_change_count"], 0)
            self.assertGreater(result.snapshot["history_paper_final_equity_change"], 0.0)
            self.assertEqual(result.snapshot["dashboard_history_refresh_status"], "completed")
            self.assertEqual(result.dashboard_result.snapshot["pipeline_history_status"], "ok")
            self.assertEqual(
                result.dashboard_result.snapshot["pipeline_history_health_state"],
                result.snapshot["history_review_health_state"],
            )
            self.assertEqual(
                result.dashboard_result.snapshot["pipeline_history_latest_as_of_date"],
                result.snapshot["history_review_latest_as_of_date"],
            )
            self.assertEqual(result.dashboard_result.snapshot["market_cache_status"], "fresh_enough")
            self.assertEqual(result.dashboard_result.snapshot["allocator_input_status"], "fresh_enough")
            self.assertEqual(result.dashboard_result.snapshot["paper_account_status"], "ok")
            history_frame = pd.read_csv(history)
            self.assertEqual(len(history_frame), 2)
            self.assertEqual(history_frame.iloc[-1]["as_of_date"], "2020-01-08")
            self.assertEqual(history_frame.iloc[-1]["pipeline_next_step_stage"], result.snapshot["pipeline_next_step_stage"])
            self.assertEqual(history_frame.iloc[-1]["model_audit_status"], "needs_attention")
            self.assertEqual(history_frame.iloc[-1]["model_audit_walk_forward_action_items"], 2)
            self.assertEqual(int(history_frame.iloc[-1]["promotion_support_group_count"]), 2)
            self.assertEqual(int(history_frame.iloc[-1]["promotion_evidence_support_count"]), 1)
            self.assertEqual(history_frame.iloc[-1]["dashboard_history_refresh_status"], "completed")
            self.assertEqual(history_frame.iloc[-1]["dashboard_alert_refresh_status"], "completed")
            self.assertEqual(history_frame.iloc[-1]["dashboard_pipeline_history_status"], "ok")
            self.assertEqual(history_frame.iloc[-1]["dashboard_pipeline_history_health_state"], "watch")
            self.assertEqual(history_frame.iloc[-1]["live_shadow_status"], "completed")
            self.assertEqual(history_frame.iloc[-1]["live_shadow_trade_plan_status"], "manual_review_only")
            self.assertEqual(history_frame.iloc[-1]["live_shadow_broker_action"], "none")
            self.assertGreaterEqual(int(history_frame.iloc[-1]["live_shadow_order_count"]), 1)
            self.assertEqual(history_frame.iloc[-1]["history_review_alert_refresh_status"], "completed")
            self.assertEqual(history_frame.iloc[-1]["history_review_latest_alert_level"], "warning")
            self.assertEqual(history_frame.iloc[-1]["history_review_latest_alert_action_stage"], "review_required")
            self.assertEqual(int(history_frame.iloc[-1]["history_review_latest_alert_count"]), result.snapshot["alert_count"])
            self.assertEqual(history_frame.iloc[-1]["dashboard_pipeline_history_latest_as_of_date"], "2020-01-08")
            self.assertEqual(history_frame.iloc[-1]["alert_level"], result.snapshot["alert_level"])
            self.assertEqual(history_frame.iloc[-1]["alert_action_stage"], result.snapshot["alert_action_stage"])
            self.assertEqual(int(history_frame.iloc[-1]["alert_count"]), result.snapshot["alert_count"])
            self.assertTrue(str(history_frame.iloc[-1]["alerts_report_path"]).endswith("alerts.md"))
            self.assertIn("Daily Research Pipeline", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Next Step", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Model Build Hygiene", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Independent support groups | 2 / 2", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Evidence group support | 1", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Alerts", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Paper Target Holdings", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Paper stock targets", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Paper stock target review", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Stock target action report", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Stock target assistant report", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Stock target outcome report", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Stock target outcome history report", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Stock target outcome analysis report", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Review notes CSV", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Live Shadow", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("manual_review_only", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Pipeline History", result.dashboard_result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Pipeline history | `ok` / `watch`", result.dashboard_result.report_path.read_text(encoding="utf-8"))

    def test_daily_pipeline_retryable_when_paper_account_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)

            fake_daily_result = SimpleNamespace(
                output_dir=root / "daily",
                report_path=root / "daily" / "daily_model_check.md",
                snapshot_path=root / "daily" / "daily_model_check_snapshot.json",
                snapshot={
                    "data_freshness_status": "fresh_enough",
                    "action_posture": "monitor",
                    "phase2_posture": "core_base_allocator_gate",
                    "latest_date": "2024-01-02",
                },
            )

            fake_dashboard_result = SimpleNamespace(
                output_dir=root / "dashboard",
                snapshot_path=root / "dashboard" / "latest_dashboard_snapshot.json",
                report_path=root / "dashboard" / "latest_dashboard.md",
                snapshot={
                    "dashboard_posture": "core_base_watch_allocator_gate",
                    "sentiment_state": "neutral",
                    "sentiment_freshness_status": "fresh_enough",
                    "market_cache_status": "fresh_enough",
                    "allocator_input_status": "fresh_enough",
                    "paper_freshness_status": "missing",
                    "trigger_freshness_status": "missing",
                    "model_audit_status": "ok",
                    "model_audit_walk_forward_action_items": 0,
                    "model_audit_walk_forward_resume_candidates": 0,
                    "model_audit_walk_forward_archive_review_candidates": 0,
                    "promotion_status": "ok",
                    "promotion_decision": "wait",
                    "promotion_sensitivity_support_count": 0,
                    "promotion_sensitivity_group_support_count": 0,
                    "promotion_support_group_count": 0,
                    "promotion_evidence_support_count": 0,
                    "promotion_sensitivity_run_support_count": 0,
                    "promotion_min_sensitivity_support": 0,
                    "promotion_return_edge": 0.0,
                    "promotion_sharpe_edge": 0.0,
                    "promotion_drawdown_change": 0.0,
                    "promotion_report_path": str(root / "promotion" / "allocator_promotion_review.md"),
                    "pipeline_history_status": "ok",
                    "pipeline_history_health_state": "ok",
                    "pipeline_history_alert_count": 0,
                    "pipeline_history_latest_as_of_date": "2024-01-02",
                },
            )

            fake_risk_budget_result = SimpleNamespace(
                output_dir=root / "risk",
                snapshot_path=root / "risk" / "satellite_risk_budget_snapshot.json",
                report_path=root / "risk" / "satellite_risk_budget_review.md",
                checklist_path=root / "risk" / "satellite_risk_budget_checklist.csv",
                snapshot={
                    "risk_budget_decision": "wait_for_outcome_samples",
                    "risk_budget_reason": "No outcome samples yet.",
                    "next_action_stage": "monitor",
                    "recommended_satellite_budget": 0.0,
                    "selected_horizon": "",
                    "outcome_ready_horizon_count": 0,
                    "satellite_trial_rule_count": 2,
                    "satellite_trial_rules_path": str(root / "risk" / "satellite_trial_rules.csv"),
                    "satellite_trial_rules_json_path": str(root / "risk" / "satellite_trial_rules.json"),
                },
            )
            fake_trial_replay_result = SimpleNamespace(
                output_dir=root / "trial_replay",
                summary_path=root / "trial_replay" / "satellite_trial_replay_summary.csv",
                matches_path=root / "trial_replay" / "satellite_trial_replay_matches.csv",
                snapshot_path=root / "trial_replay" / "satellite_trial_replay_snapshot.json",
                report_path=root / "trial_replay" / "satellite_trial_replay.md",
                snapshot={
                    "status": "completed",
                    "matched_event_count": 8,
                    "union_best_horizon": "5d",
                    "union_best_avg_return_edge": 0.04,
                    "union_best_win_rate_edge": 0.20,
                },
            )

            fake_live_preflight = SimpleNamespace(
                output_dir=root / "live_preflight",
                snapshot_path=root / "live_preflight" / "live_preflight_snapshot.json",
                report_path=root / "live_preflight" / "live_preflight.md",
                checklist_path=root / "live_preflight" / "live_preflight_checklist.csv",
                snapshot={
                    "decision": "hold",
                    "broker_connection_status": "not_connected",
                    "blocking_items": [],
                    "monitor_items": [],
                    "pipeline_snapshot_status": "present",
                    "pipeline_snapshot_path": str(root / "pipeline_snapshot.json"),
                    "active_target_count": 0,
                    "stock_target_review_action_count": 0,
                },
            )

            fake_alerts = SimpleNamespace(
                output_dir=root / "alerts",
                report_path=root / "alerts" / "alerts.md",
                json_path=root / "alerts" / "alerts.json",
                latest_report_path=root / "alerts" / "alerts_latest.md",
                payload={
                    "alert_level": "warning",
                    "alert_count": 1,
                    "action_stage": "review_required",
                    "action_summary": "Paper account failed.",
                    "action_item_count": 1,
                    "action_required": True,
                    "critical_count": 0,
                    "warning_count": 1,
                    "info_count": 0,
                    "alerts": [
                        {"code": "paper_account_failed"},
                    ],
                },
            )

            with patch(
                "quant_etf_lab.daily_pipeline.run_daily_model_check", return_value=fake_daily_result
            ), patch(
                "quant_etf_lab.daily_pipeline.run_paper_account",
                side_effect=RuntimeError("simulated paper account outage"),
            ), patch(
                "quant_etf_lab.daily_pipeline.run_daily_dashboard", return_value=fake_dashboard_result
            ), patch(
                "quant_etf_lab.daily_pipeline.run_satellite_risk_budget_review",
                return_value=fake_risk_budget_result,
            ), patch(
                "quant_etf_lab.daily_pipeline.run_satellite_trial_replay",
                return_value=fake_trial_replay_result,
            ), patch(
                "quant_etf_lab.daily_pipeline.run_live_preflight", return_value=fake_live_preflight
            ), patch(
                "quant_etf_lab.daily_pipeline.write_daily_alerts", return_value=fake_alerts
            ):
                result = run_daily_pipeline(
                    project_root=root,
                    output_dir=root / "pipeline",
                    daily_check_output_dir=root / "daily",
                    as_of_date="2024-01-02",
                    date_stamp=False,
                    run_history_review=False,
                    history_path=None,
                )

            self.assertEqual(result.snapshot["paper_account_status"], "failed")
            self.assertIn("simulated paper account outage", result.snapshot["paper_account_error"])
            self.assertEqual(result.snapshot["pipeline_next_step_stage"], "retry_paper_account")
            self.assertEqual(result.snapshot["pipeline_blocker"], "paper_account_step_failed")
            self.assertTrue(result.snapshot["pipeline_user_intervention_required"])
            self.assertEqual(result.snapshot["satellite_trial_rule_count"], 2)
            self.assertTrue(str(result.snapshot["satellite_trial_rules_path"]).endswith("satellite_trial_rules.csv"))
            self.assertTrue(str(result.snapshot["satellite_trial_rules_json_path"]).endswith("satellite_trial_rules.json"))
            self.assertEqual(result.snapshot["satellite_trial_replay_status"], "completed")
            self.assertEqual(result.snapshot["satellite_trial_replay_matched_event_count"], 8)
            self.assertEqual(result.snapshot["satellite_trial_replay_best_horizon"], "5d")
            self.assertAlmostEqual(result.snapshot["satellite_trial_replay_best_avg_return_edge"], 0.04)
            self.assertTrue(str(result.snapshot["satellite_trial_replay_report_path"]).endswith("satellite_trial_replay.md"))
            self.assertIn("Satellite trial rules", result.report_path.read_text(encoding="utf-8"))
            self.assertIn("Trial replay status", result.report_path.read_text(encoding="utf-8"))
            self.assertTrue(result.paper_account_result.report_path.exists())
            self.assertTrue(result.paper_account_result.stock_target_review_decision_template_path.exists())
            self.assertTrue(result.paper_account_result.stock_target_review_decision_template_json_path.exists())
            self.assertTrue(result.paper_account_result.stock_target_review_decision_template_report_path.exists())
            self.assertTrue(result.paper_account_result.stock_target_review_decision_template_xlsx_path.exists())
            self.assertEqual(result.snapshot["paper_stock_target_review_decision_template_count"], 0)
            self.assertEqual(result.snapshot["paper_stock_target_review_decision_template_blank_status_count"], 0)
            self.assertTrue(result.paper_account_result.output_dir.exists())
            self.assertTrue(result.snapshot_path.exists())
            self.assertTrue(result.report_path.exists())

    def test_live_preflight_treats_manual_watch_followups_as_monitor_not_blocker(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            dashboard_path = root / "dashboard" / "latest_dashboard_snapshot.json"
            paper_dir = root / "paper"
            output_dir = root / "live_preflight"
            dashboard_path.parent.mkdir(parents=True)
            paper_dir.mkdir(parents=True)
            dashboard = {
                "dashboard_posture": "core_base_watch_allocator_gate",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_account_status": "ok",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "trading_day_gate_status": "trading_day_data_ready",
                "after_close_data_status": "ready",
            }
            dashboard_path.write_text(json.dumps(dashboard), encoding="utf-8")
            pd.DataFrame(
                [
                    {
                        "code": "002216",
                        "portfolio_target_weight": 0.05,
                    }
                ]
            ).to_csv(paper_dir / "target_holdings.csv", index=False)
            pd.DataFrame(
                [
                    {
                        "action_code": "manual_watch_followup",
                        "action_stage": "monitor",
                        "code": "002216",
                        "broker_action": "none",
                    }
                ]
            ).to_csv(paper_dir / "stock_target_review_actions.csv", index=False)
            (paper_dir / "metrics.json").write_text(
                json.dumps(
                    {
                        "latest_regime": "risk_on",
                        "latest_core_weight": 0.75,
                        "latest_satellite_weight": 0.25,
                        "latest_cash_weight": 0.0,
                        "latest_date": "2026-06-15",
                        "target_holdings_path": str(paper_dir / "target_holdings.csv"),
                        "stock_target_review_actions_path": str(paper_dir / "stock_target_review_actions.csv"),
                        "stock_target_review_action_count": 1,
                        "stock_target_review_action_manual_watch_count": 1,
                    }
                ),
                encoding="utf-8",
            )

            result = run_live_preflight(
                dashboard_snapshot=dashboard_path,
                paper_account_dir=paper_dir,
                output_dir=output_dir,
            )

            self.assertEqual(result.snapshot["decision"], "ready_for_manual_review_pre_stage")
            self.assertEqual(result.snapshot["stock_target_review_action_count"], 1)
            self.assertEqual(result.snapshot["stock_target_review_blocking_action_count"], 0)
            self.assertEqual(result.snapshot["stock_target_review_monitor_action_count"], 1)
            self.assertNotIn("stock_target_review_action_count=1", result.snapshot["blocking_items"])
            self.assertIn("stock_target_review_monitor_action_count=1", result.snapshot["monitor_items"])

    def test_live_preflight_treats_unknown_stock_review_actions_as_blockers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            dashboard_path = root / "dashboard" / "latest_dashboard_snapshot.json"
            paper_dir = root / "paper"
            output_dir = root / "live_preflight"
            dashboard_path.parent.mkdir(parents=True)
            paper_dir.mkdir(parents=True)
            dashboard = {
                "dashboard_posture": "core_base_watch_allocator_gate",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_account_status": "ok",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "trading_day_gate_status": "trading_day_data_ready",
                "after_close_data_status": "ready",
            }
            dashboard_path.write_text(json.dumps(dashboard), encoding="utf-8")
            pd.DataFrame([{"code": "002216", "portfolio_target_weight": 0.05}]).to_csv(
                paper_dir / "target_holdings.csv",
                index=False,
            )
            pd.DataFrame(
                [
                    {
                        "action_code": "unexpected_manual_review_action",
                        "action_stage": "review_required",
                        "code": "002216",
                        "broker_action": "none",
                    }
                ]
            ).to_csv(paper_dir / "stock_target_review_actions.csv", index=False)
            (paper_dir / "metrics.json").write_text(
                json.dumps(
                    {
                        "latest_regime": "risk_on",
                        "latest_core_weight": 0.75,
                        "latest_satellite_weight": 0.25,
                        "latest_cash_weight": 0.0,
                        "latest_date": "2026-06-15",
                        "target_holdings_path": str(paper_dir / "target_holdings.csv"),
                        "stock_target_review_actions_path": str(paper_dir / "stock_target_review_actions.csv"),
                        "stock_target_review_action_count": 1,
                    }
                ),
                encoding="utf-8",
            )

            result = run_live_preflight(
                dashboard_snapshot=dashboard_path,
                paper_account_dir=paper_dir,
                output_dir=output_dir,
            )

            self.assertEqual(result.snapshot["decision"], "blocked")
            self.assertEqual(result.snapshot["stock_target_review_action_count"], 1)
            self.assertEqual(result.snapshot["stock_target_review_blocking_action_count"], 1)
            self.assertEqual(result.snapshot["stock_target_review_monitor_action_count"], 0)
            self.assertEqual(result.snapshot["stock_target_review_unknown_action_count"], 1)
            self.assertIn("stock_target_review_blocking_action_count=1", result.snapshot["blocking_items"])

    def test_daily_pipeline_live_preflight_context_includes_shadow_review_decisions(self) -> None:
        result = SimpleNamespace(
            output_dir=Path("live_preflight"),
            snapshot_path=Path("live_preflight/live_preflight_snapshot.json"),
            report_path=Path("live_preflight/live_preflight.md"),
            checklist_path=Path("live_preflight/live_preflight_checklist.csv"),
            snapshot={
                "decision": "blocked",
                "broker_connection_status": "not_connected",
                "blocking_items": ["live_shadow_review_blocking_decision_count=1"],
                "monitor_items": ["live_shadow_review_monitor_decision_count=1"],
                "pipeline_snapshot_status": "ok",
                "pipeline_snapshot_path": "pipeline/daily_pipeline_snapshot.json",
                "active_target_count": 3,
                "stock_target_review_action_count": 0,
                "live_shadow_review_decisions_path": "review/live_shadow_review_decisions.csv",
                "live_shadow_review_decision_count": 3,
                "live_shadow_review_blocking_decision_count": 1,
                "live_shadow_review_monitor_decision_count": 1,
                "live_shadow_review_unknown_decision_count": 0,
            },
        )

        context = _live_preflight_context(result)

        self.assertEqual(context["live_preflight_live_shadow_review_decisions_path"], "review/live_shadow_review_decisions.csv")
        self.assertEqual(context["live_preflight_live_shadow_review_decision_count"], 3)
        self.assertEqual(context["live_preflight_live_shadow_review_blocking_decision_count"], 1)
        self.assertEqual(context["live_preflight_live_shadow_review_monitor_decision_count"], 1)
        self.assertEqual(context["live_preflight_live_shadow_review_unknown_decision_count"], 0)

    def test_daily_pipeline_preflight_blocked_blocks_live_shadow(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            live_holdings = root / "live_holdings.csv"
            live_prices = root / "live_prices.csv"
            live_holdings.write_text("code,name,quantity,current_price\n000001,Test,0,10\n", encoding="utf-8")
            live_prices.write_text("code,price\n000001,10.0\n", encoding="utf-8")

            class FakePaperResult:
                output_dir = root / "paper"
                ledger = pd.DataFrame()
                audit = pd.DataFrame()
                target_holdings = pd.DataFrame()
                stock_targets = pd.DataFrame()
                stock_target_review = pd.DataFrame()
                stock_target_review_actions = pd.DataFrame()
                stock_target_review_assistant = pd.DataFrame()
                stock_target_review_outcomes = pd.DataFrame()
                stock_target_review_outcomes_history = pd.DataFrame()
                stock_target_review_outcome_analysis = pd.DataFrame()
                stock_target_review_outcome_calendar = pd.DataFrame()
                stock_target_review_outcome_due = pd.DataFrame()
                monthly_returns = pd.DataFrame()
                metrics = {
                    "latest_date": "2026-06-14",
                    "latest_regime": "risk_off",
                    "latest_candidate": "core_only",
                    "latest_core_weight": 1.0,
                    "latest_satellite_weight": 0.0,
                    "latest_cash_weight": 0.0,
                    "final_equity": 100000.0,
                    "total_return": 0.0,
                    "max_drawdown": -0.02,
                    "current_drawdown": 0.0,
                    "sharpe": 1.0,
                    "audit_event_count": 0,
                    "total_estimated_fee": 0.0,
                }

                def __getattr__(self, name: str) -> object:
                    if name.endswith("_payload"):
                        return {}
                    if name.endswith("_path") or name == "report_path":
                        return self.output_dir / f"{name}.txt"
                    raise AttributeError(name)

            daily_result = SimpleNamespace(
                output_dir=root / "daily",
                report_path=root / "daily" / "daily_model_check.md",
                snapshot_path=root / "daily" / "daily_model_check_snapshot.json",
                snapshot={
                    "data_freshness_status": "fresh_enough",
                    "action_posture": "monitor",
                    "phase2_posture": "risk_on",
                    "latest_date": "2026-06-14",
                },
            )
            paper_result = FakePaperResult()
            dashboard_result = SimpleNamespace(
                output_dir=root / "dashboard",
                snapshot_path=root / "dashboard" / "latest_dashboard_snapshot.json",
                report_path=root / "dashboard" / "latest_dashboard.md",
                snapshot={
                    "dashboard_posture": "core_base_watch_allocator_gate",
                    "sentiment_state": "neutral",
                    "sentiment_freshness_status": "fresh_enough",
                    "market_cache_status": "fresh_enough",
                    "allocator_input_status": "fresh_enough",
                    "paper_freshness_status": "fresh_enough",
                    "trigger_freshness_status": "fresh_enough",
                    "model_audit_status": "ok",
                    "model_audit_walk_forward_action_items": 0,
                    "model_audit_walk_forward_resume_candidates": 0,
                    "model_audit_walk_forward_archive_review_candidates": 0,
                    "promotion_status": "ok",
                    "promotion_decision": "promote_candidate",
                    "pipeline_history_status": "ok",
                    "pipeline_history_health_state": "ok",
                    "pipeline_history_alert_count": 0,
                    "pipeline_history_latest_as_of_date": "2026-06-14",
                },
            )
            satellite_result = SimpleNamespace(
                output_dir=root / "risk",
                snapshot_path=root / "risk" / "satellite_risk_budget_snapshot.json",
                report_path=root / "risk" / "satellite_risk_budget_review.md",
                checklist_path=root / "risk" / "satellite_risk_budget_checklist.csv",
                snapshot={
                    "risk_budget_decision": "eligible_for_small_satellite_trial",
                    "risk_budget_reason": "No blockers.",
                    "next_action_stage": "monitor",
                    "recommended_satellite_budget": 0.05,
                    "selected_horizon": "12m",
                    "outcome_ready_horizon_count": 1,
                },
            )
            preflight_result = SimpleNamespace(
                output_dir=root / "preflight",
                snapshot_path=root / "preflight" / "live_shadow_preflight_snapshot.json",
                report_path=root / "preflight" / "live_shadow_preflight_report.md",
                snapshot={"status": "blocked", "blockers": ["risk exposure too high"]},
            )
            preflight_status_result = SimpleNamespace(
                output_dir=root / "live_preflight",
                snapshot_path=root / "live_preflight" / "live_preflight_snapshot.json",
                report_path=root / "live_preflight" / "live_preflight.md",
                checklist_path=root / "live_preflight" / "live_preflight_checklist.csv",
                snapshot={
                    "decision": "hold",
                    "broker_connection_status": "not_connected",
                    "blocking_items": [],
                    "monitor_items": [],
                    "pipeline_snapshot_status": "present",
                    "pipeline_snapshot_path": str(root / "pipeline_snapshot.json"),
                    "active_target_count": 0,
                    "stock_target_review_action_count": 0,
                },
            )
            alerts_result = SimpleNamespace(
                output_dir=root / "alerts",
                report_path=root / "alerts" / "alerts.md",
                json_path=root / "alerts" / "alerts.json",
                latest_report_path=root / "alerts" / "alerts_latest.md",
                payload={},
            )

            with patch("quant_etf_lab.daily_pipeline.run_daily_model_check", return_value=daily_result), patch(
                "quant_etf_lab.daily_pipeline.run_paper_account", return_value=paper_result
            ), patch(
                "quant_etf_lab.daily_pipeline.run_daily_dashboard", return_value=dashboard_result
            ), patch(
                "quant_etf_lab.daily_pipeline.run_satellite_risk_budget_review", return_value=satellite_result
            ), patch(
                "quant_etf_lab.daily_pipeline.run_live_shadow_preflight", return_value=preflight_result
            ) as preflight_mock, patch(
                "quant_etf_lab.daily_pipeline.run_live_shadow"
            ) as shadow_mock, patch(
                "quant_etf_lab.daily_pipeline.run_live_preflight",
                return_value=preflight_status_result,
            ), patch(
                "quant_etf_lab.daily_pipeline.write_daily_alerts", return_value=alerts_result
            ):
                result = run_daily_pipeline(
                    project_root=root,
                    output_dir=root / "pipeline",
                    as_of_date="2026-06-14",
                    date_stamp=False,
                    live_shadow_preflight=True,
                    live_shadow_preflight_fail_on_blocked=True,
                    live_shadow_holdings_file=live_holdings,
                    live_shadow_prices_file=live_prices,
                    live_shadow_cash=20000.0,
                    run_history_review=False,
                    history_path=None,
                )

            preflight_mock.assert_called_once()
            shadow_mock.assert_not_called()
            self.assertEqual(result.snapshot.get("live_shadow_preflight_status"), "completed")
            self.assertEqual(result.snapshot.get("live_shadow_preflight_decision"), "blocked")
            self.assertEqual(result.snapshot.get("live_shadow_status"), "blocked_by_preflight")
            self.assertEqual(result.snapshot.get("live_shadow_trade_plan_status"), "blocked_by_preflight")
            self.assertEqual(result.snapshot.get("live_shadow_broker_action"), "none")
            self.assertEqual(result.live_shadow_result, None)
            self.assertEqual(result.snapshot.get("pipeline_next_step_stage"), "review_live_shadow_preflight")
            self.assertTrue(result.snapshot.get("pipeline_user_intervention_required"))
            self.assertEqual(result.snapshot.get("pipeline_blocker"), "live_shadow_preflight_blocked")


    def test_daily_pipeline_separates_final_history_and_dashboard_refresh_failures(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            history = root / "history.csv"

            class FakePaperResult:
                output_dir = root / "paper"
                ledger = pd.DataFrame()
                audit = pd.DataFrame()
                target_holdings = pd.DataFrame()
                stock_targets = pd.DataFrame()
                stock_target_review = pd.DataFrame()
                stock_target_review_actions = pd.DataFrame()
                stock_target_review_assistant = pd.DataFrame()
                stock_target_review_outcomes = pd.DataFrame()
                stock_target_review_outcomes_history = pd.DataFrame()
                stock_target_review_outcome_analysis = pd.DataFrame()
                stock_target_review_outcome_calendar = pd.DataFrame()
                stock_target_review_outcome_due = pd.DataFrame()
                monthly_returns = pd.DataFrame()
                metrics = {
                    "latest_date": "2026-06-12",
                    "latest_regime": "risk_off",
                    "latest_candidate": "core_only",
                    "latest_core_weight": 1.0,
                    "latest_satellite_weight": 0.0,
                    "latest_cash_weight": 0.0,
                    "final_equity": 1000000.0,
                    "total_return": 0.0,
                    "max_drawdown": -0.02,
                    "current_drawdown": -0.01,
                    "sharpe": 1.0,
                    "audit_event_count": 0,
                }

                def __getattr__(self, name: str) -> object:
                    if name.endswith("_payload"):
                        return {}
                    if name.endswith("_path") or name == "report_path":
                        return self.output_dir / f"{name}.txt"
                    raise AttributeError(name)

            daily_result = SimpleNamespace(
                output_dir=root / "daily",
                report_path=root / "daily" / "daily_model_check.md",
                snapshot_path=root / "daily" / "daily_model_check_snapshot.json",
                phase2_result=None,
                snapshot={
                    "data_freshness_status": "fresh_enough",
                    "action_posture": "review_core_base_allocator_gate",
                    "phase2_posture": "core_base_allocator_gated_satellite",
                    "latest_date": "2026-06-12",
                },
            )
            paper_result = FakePaperResult()
            dashboard_snapshot = {
                "dashboard_posture": "core_base_watch_allocator_gate",
                "sentiment_state": "warm",
                "sentiment_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "model_audit_status": "ok",
                "model_audit_walk_forward_action_items": 0,
                "model_audit_walk_forward_resume_candidates": 0,
                "model_audit_walk_forward_archive_review_candidates": 0,
                "promotion_status": "ok",
                "promotion_decision": "promote_candidate",
                "pipeline_history_status": "ok",
                "pipeline_history_health_state": "ok",
                "pipeline_history_alert_count": 0,
                "pipeline_history_latest_as_of_date": "2026-06-12",
            }
            dashboard_result = SimpleNamespace(
                output_dir=root / "dashboard",
                report_path=root / "dashboard" / "latest_dashboard.md",
                snapshot_path=root / "dashboard" / "latest_dashboard_snapshot.json",
                snapshot=dashboard_snapshot,
            )
            history_review_result = SimpleNamespace(
                output_dir=root / "history_review",
                snapshot_path=root / "history_review" / "pipeline_history_review_snapshot.json",
                report_path=root / "history_review" / "pipeline_history_review.md",
                history_path=history,
                history=pd.DataFrame(),
                snapshot={
                    "health_state": "ok",
                    "alert_count": 0,
                    "latest_as_of_date": "2026-06-12",
                    "latest_alert_level": "info",
                    "latest_alert_action_stage": "monitor",
                    "latest_alert_count": 1,
                },
            )
            satellite_result = SimpleNamespace(
                output_dir=root / "risk",
                report_path=root / "risk" / "satellite_risk_budget_review.md",
                snapshot_path=root / "risk" / "satellite_risk_budget_snapshot.json",
                checklist_path=root / "risk" / "satellite_risk_budget_checklist.csv",
                snapshot={
                    "risk_budget_decision": "wait_for_outcome_samples",
                    "risk_budget_reason": "Waiting for outcome samples.",
                    "next_action_stage": "wait",
                    "recommended_satellite_budget": 0.0,
                    "selected_horizon": "",
                    "outcome_ready_horizon_count": 0,
                },
            )

            with patch("quant_etf_lab.daily_pipeline.run_daily_model_check", return_value=daily_result), patch(
                "quant_etf_lab.daily_pipeline.run_paper_account", return_value=paper_result
            ), patch(
                "quant_etf_lab.daily_pipeline.run_satellite_risk_budget_review", return_value=satellite_result
            ), patch(
                "quant_etf_lab.daily_pipeline.run_pipeline_history_review",
                side_effect=[history_review_result, history_review_result],
            ), patch(
                "quant_etf_lab.daily_pipeline.run_daily_dashboard",
                side_effect=[
                    dashboard_result,
                    dashboard_result,
                    RuntimeError("final dashboard write failed"),
                ],
            ):
                result = run_daily_pipeline(
                    project_root=root,
                    output_dir=root / "pipeline",
                    daily_check_output_dir=root / "daily",
                    paper_account_output_dir=root / "paper",
                    dashboard_output_dir=root / "dashboard",
                    history_path=history,
                    history_review_output_dir=root / "history_review",
                    satellite_risk_budget_output_dir=root / "risk",
                    as_of_date="2026-06-12",
                    date_stamp=False,
                )

            self.assertEqual(result.snapshot["history_review_alert_refresh_status"], "completed")
            self.assertIsNone(result.snapshot["history_review_alert_refresh_error"])
            self.assertEqual(result.snapshot["dashboard_alert_refresh_status"], "failed")
            self.assertIn("final dashboard write failed", result.snapshot["dashboard_alert_refresh_error"])
            self.assertEqual(result.alerts_result.payload["dashboard_alert_refresh_status"], "failed")
            self.assertEqual(result.snapshot["dashboard_history_refresh_status"], "failed")
            self.assertIn("final dashboard write failed", result.snapshot["dashboard_history_refresh_error"])
            codes = {item["code"] for item in result.alerts_result.payload["alerts"]}
            self.assertIn("dashboard_history_refresh_failed", codes)
            self.assertNotIn("history_review_alert_refresh_failed", codes)

            history_frame = pd.read_csv(history)
            self.assertEqual(history_frame.iloc[-1]["history_review_alert_refresh_status"], "completed")
            self.assertEqual(history_frame.iloc[-1]["dashboard_alert_refresh_status"], "failed")
            self.assertEqual(history_frame.iloc[-1]["dashboard_history_refresh_status"], "failed")

    def test_trading_gate_marks_weekend_as_non_trading_day(self) -> None:
        gate = build_a_share_trading_gate(
            {"local_equity_latest_date": "2026-06-12"},
            "2026-06-14",
            generated_at="2026-06-14T16:10:00",
        )

        self.assertEqual(gate["trading_day_gate_status"], "non_trading_day")
        self.assertFalse(gate["is_a_share_trading_day"])
        self.assertEqual(gate["after_close_data_status"], "not_required")
        self.assertEqual(gate["trading_day_gate_action"], "skip_new_signal_interpretation")

    def test_trading_gate_blocks_weekday_after_close_without_current_data(self) -> None:
        gate = build_a_share_trading_gate(
            {"local_equity_latest_date": "2024-04-01"},
            "2024-04-02",
            generated_at="2024-04-02T16:10:00",
        )

        self.assertEqual(gate["trading_day_gate_status"], "trading_day_data_not_ready")
        self.assertTrue(gate["is_a_share_trading_day"])
        self.assertEqual(gate["after_close_data_status"], "not_ready")
        self.assertEqual(gate["trading_day_gate_action"], "verify_trading_day_or_refresh_data")

    def test_daily_pipeline_next_step_context_prioritizes_blockers_and_review(self) -> None:
        refresh = _next_step_context(
            {
                "trading_day_gate_status": "trading_day_data_not_ready",
                "alert_level": "warning",
                "paper_stock_target_review_required_unreviewed_count": 2,
            }
        )
        self.assertEqual(refresh["pipeline_next_step_stage"], "refresh_data")
        self.assertFalse(refresh["pipeline_user_intervention_required"])

        manual = _next_step_context(
            {
                "trading_day_gate_status": "non_trading_day",
                "alert_level": "warning",
                "paper_stock_target_review_required_unreviewed_count": 2,
            }
        )
        self.assertEqual(manual["pipeline_next_step_stage"], "update_manual_review_notes")
        self.assertTrue(manual["pipeline_user_intervention_required"])

        preflight_block = _next_step_context(
            {
                "trading_day_gate_status": "trading_day_data_ready",
                "alert_level": "normal",
                "live_shadow_status": "blocked_by_preflight",
                "live_shadow_preflight_status": "completed",
                "live_shadow_preflight_decision": "blocked",
                "live_shadow_preflight_blockers_count": 2,
                "paper_stock_target_review_required_unreviewed_count": 0,
            }
        )
        self.assertEqual(preflight_block["pipeline_next_step_stage"], "review_live_shadow_preflight")
        self.assertEqual(preflight_block["pipeline_blocker"], "live_shadow_preflight_blocked")
        self.assertTrue(preflight_block["pipeline_user_intervention_required"])

        outcome = _next_step_context(
            {
                "trading_day_gate_status": "trading_day_data_ready",
                "alert_level": "normal",
                "paper_stock_target_review_required_unreviewed_count": 0,
                "paper_stock_target_review_action_count": 0,
                "paper_stock_target_review_outcome_analysis_ready_horizon_count": 1,
            }
        )
        self.assertEqual(outcome["pipeline_next_step_stage"], "review_outcome_analysis")
        self.assertFalse(outcome["pipeline_user_intervention_required"])

        chip_ready = _next_step_context(
            {
                "trading_day_gate_status": "trading_day_data_ready",
                "alert_level": "normal",
                "paper_stock_target_review_required_unreviewed_count": 0,
                "paper_stock_target_review_action_count": 0,
                "paper_stock_target_review_outcome_analysis_ready_horizon_count": 0,
                "paper_stock_target_review_outcome_due_row_count": 0,
                "chip_reversal_candidate_outcomes_readiness_status": "ready",
                "chip_reversal_candidate_outcomes_ready_horizons": ["1d"],
                "chip_reversal_candidate_outcomes_report_path": "chip_reversal_candidate_outcomes.md",
            }
        )
        self.assertEqual(chip_ready["pipeline_next_step_stage"], "review_chip_reversal_outcomes")
        self.assertIn("chip_reversal_candidate_outcomes.md", chip_ready["pipeline_next_step_action"])
        self.assertFalse(chip_ready["pipeline_user_intervention_required"])

        chip_waiting = _next_step_context(
            {
                "trading_day_gate_status": "trading_day_data_ready",
                "alert_level": "normal",
                "paper_stock_target_review_required_unreviewed_count": 0,
                "paper_stock_target_review_action_count": 0,
                "paper_stock_target_review_outcome_analysis_ready_horizon_count": 0,
                "paper_stock_target_review_outcome_due_row_count": 0,
                "chip_reversal_candidate_outcomes_readiness_status": "waiting_for_future_bar",
                "chip_reversal_candidate_outcomes_pending_horizons": ["1d", "2d"],
                "chip_reversal_candidate_outcomes_next_review_horizon": "1d",
                "chip_reversal_candidate_outcomes_next_review_reason": "future_bar_not_available",
            }
        )
        self.assertEqual(chip_waiting["pipeline_next_step_stage"], "accumulate_chip_reversal_outcome_samples")
        self.assertEqual(chip_waiting["pipeline_blocker"], "chip_reversal_outcomes_waiting_for_future_bar")
        self.assertFalse(chip_waiting["pipeline_user_intervention_required"])

        promotion = _next_step_context(
            {
                "trading_day_gate_status": "trading_day_data_ready",
                "alert_level": "info",
                "paper_stock_target_review_required_unreviewed_count": 0,
                "paper_stock_target_review_action_count": 0,
                "paper_stock_target_review_outcome_analysis_ready_horizon_count": 0,
                "paper_stock_target_review_outcome_due_row_count": 0,
                "paper_stock_target_review_outcome_analysis_status": "ready_for_review",
                "promotion_decision": "watch_candidate",
                "promotion_candidate_passes_headline_gate": True,
                "promotion_sensitivity_group_support_count": 1,
                "promotion_sensitivity_run_support_count": 2,
                "promotion_min_sensitivity_support": 2,
            }
        )
        self.assertEqual(promotion["pipeline_next_step_stage"], "monitor_allocator_promotion_evidence")
        self.assertFalse(promotion["pipeline_user_intervention_required"])

        non_trading_promotion = _next_step_context(
            {
                "trading_day_gate_status": "non_trading_day",
                "alert_level": "info",
                "promotion_decision": "watch_candidate",
                "promotion_candidate_passes_headline_gate": True,
                "promotion_sensitivity_group_support_count": 1,
                "promotion_sensitivity_run_support_count": 2,
                "promotion_min_sensitivity_support": 2,
            }
        )
        self.assertEqual(non_trading_promotion["pipeline_next_step_stage"], "wait_for_next_trading_close")

    def test_daily_alerts_flag_stale_data_drawdown_and_satellite_open(self) -> None:
        payload = build_daily_alert_payload(
            {
                "as_of_date": "2024-04-02",
                "data_freshness_status": "stale",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "wait_for_fresh_model_data",
                "paper_current_drawdown": -0.13,
                "paper_max_drawdown": -0.17,
                "trading_day_gate_status": "trading_day_data_not_ready",
                "after_close_data_status": "not_ready",
                "history_dashboard_posture_changed": True,
                "history_previous_dashboard_posture": "core_base_watch_allocator_gate",
                "history_paper_total_return_change": -0.03,
                "history_paper_satellite_weight_change": 0.2,
                "paper_latest_satellite_weight": 0.2,
            }
        )

        codes = {item["code"] for item in payload["alerts"]}
        self.assertEqual(payload["alert_level"], "critical")
        self.assertEqual(payload["action_stage"], "blocked")
        self.assertTrue(payload["action_required"])
        action_codes = {item["action_code"] for item in payload["action_items"]}
        self.assertIn("verify_trading_day_or_refresh_data", action_codes)
        self.assertIn("refresh_model_data", action_codes)
        self.assertIn("review_drawdown", action_codes)
        self.assertIn("review_satellite_activation", action_codes)
        self.assertIn("data_freshness_status_not_fresh", codes)
        self.assertIn("trading_day_data_not_ready", codes)
        self.assertIn("dashboard_wait_state", codes)
        self.assertIn("current_drawdown_critical", codes)
        self.assertIn("max_drawdown_limit_breached", codes)
        self.assertIn("daily_return_drop_warning", codes)
        self.assertIn("satellite_weight_opened", codes)

    def test_daily_alerts_include_model_audit_attention_items(self) -> None:
        payload = build_daily_alert_payload(
            {
                "as_of_date": "2024-04-02",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "trading_day_data_ready",
                "model_audit_status": "needs_attention",
                "model_audit_walk_forward_action_items": 2,
                "model_audit_walk_forward_resume_candidates": 1,
                "model_audit_walk_forward_archive_review_candidates": 1,
                "model_audit_top_action": "resume_or_finalize",
                "model_audit_report_path": "audit/model_build_audit.md",
                "model_audit_actions_path": "audit/walk_forward_run_actions.csv",
            }
        )

        codes = {item["code"] for item in payload["alerts"]}
        action_codes = {item["action_code"] for item in payload["action_items"]}
        self.assertEqual(payload["alert_level"], "warning")
        self.assertEqual(payload["action_stage"], "review_required")
        self.assertIn("model_audit_needs_attention", codes)
        self.assertIn("review_model_build_audit", action_codes)
        self.assertEqual(payload["model_audit_status"], "needs_attention")
        self.assertEqual(payload["model_audit_walk_forward_action_items"], 2)
        self.assertIn("audit/model_build_audit.md", payload["model_audit_report_path"])

    def test_daily_alerts_downgrade_resolved_future_dated_model_audit(self) -> None:
        payload = build_daily_alert_payload(
            {
                "as_of_date": "2026-06-12",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "trading_day_data_ready",
                "model_audit_status": "future_dated",
                "model_audit_walk_forward_action_items": 0,
                "model_audit_walk_forward_resume_candidates": 0,
                "model_audit_walk_forward_archive_review_candidates": 0,
                "model_audit_top_action": "",
                "model_audit_report_path": "audit/model_build_audit.md",
                "model_audit_actions_path": "audit/walk_forward_run_actions.csv",
            }
        )

        codes = {item["code"] for item in payload["alerts"]}
        action_codes = {item["action_code"] for item in payload["action_items"]}
        self.assertEqual(payload["alert_level"], "info")
        self.assertEqual(payload["action_stage"], "monitor")
        self.assertFalse(payload["action_required"])
        self.assertNotIn("model_audit_needs_attention", codes)
        self.assertNotIn("review_model_build_audit", action_codes)
        self.assertIn("model_audit_resolved_future_dated", codes)

    def test_daily_alerts_downgrade_resolved_model_audit_attention(self) -> None:
        payload = build_daily_alert_payload(
            {
                "as_of_date": "2026-06-27",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "trading_day_data_ready",
                "model_audit_status": "needs_attention",
                "model_audit_walk_forward_action_items": 0,
                "model_audit_walk_forward_resume_candidates": 0,
                "model_audit_walk_forward_archive_review_candidates": 0,
                "model_audit_top_action": "",
                "model_audit_report_path": "audit/model_build_audit.md",
                "model_audit_actions_path": "audit/walk_forward_run_actions.csv",
            }
        )

        codes = {item["code"] for item in payload["alerts"]}
        action_codes = {item["action_code"] for item in payload["action_items"]}
        self.assertEqual(payload["alert_level"], "info")
        self.assertEqual(payload["action_stage"], "monitor")
        self.assertFalse(payload["action_required"])
        self.assertNotIn("model_audit_needs_attention", codes)
        self.assertNotIn("review_model_build_audit", action_codes)
        self.assertIn("model_audit_resolved_attention", codes)

    def test_daily_alerts_flag_correlated_promotion_evidence_as_info(self) -> None:
        payload = build_daily_alert_payload(
            {
                "as_of_date": "2026-06-14",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "trading_day_data_ready",
                "promotion_status": "ok",
                "promotion_decision": "watch_candidate",
                "promotion_candidate_passes_headline_gate": True,
                "promotion_sensitivity_support_count": 1,
                "promotion_sensitivity_group_support_count": 1,
                "promotion_sensitivity_run_support_count": 2,
                "promotion_min_sensitivity_support": 2,
                "promotion_report_path": "outputs/research/allocator_promotion_grouped_20260614/allocator_promotion_review.md",
            }
        )

        codes = {item["code"] for item in payload["alerts"]}
        action_codes = {item["action_code"] for item in payload["action_items"]}
        self.assertEqual(payload["alert_level"], "info")
        self.assertEqual(payload["action_stage"], "monitor")
        self.assertFalse(payload["action_required"])
        self.assertIn("promotion_evidence_correlated", codes)
        self.assertIn("monitor_allocator_promotion_evidence", action_codes)
        self.assertEqual(payload["promotion_sensitivity_group_support_count"], 1)
        self.assertEqual(payload["promotion_sensitivity_run_support_count"], 2)

    def test_daily_alerts_include_stock_target_review_queue(self) -> None:
        payload = build_daily_alert_payload(
            {
                "as_of_date": "2024-04-02",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "trading_day_data_ready",
                "paper_stock_target_review_report_path": "paper/stock_target_review.md",
                "paper_stock_target_review_actions_report_path": "paper/stock_target_review_actions.md",
                "paper_stock_target_review_assistant_report_path": "paper/stock_target_review_assistant.md",
                "paper_stock_target_review_outcomes_report_path": "paper/stock_target_review_outcomes.md",
                "paper_stock_target_review_outcomes_history_report_path": "paper/stock_target_review_outcomes_history.md",
                "paper_stock_target_review_outcome_analysis_report_path": "paper/stock_target_review_outcome_analysis.md",
                "paper_stock_target_review_outcome_calendar_report_path": "paper/stock_target_review_outcome_calendar.md",
                "paper_stock_target_review_outcome_due_report_path": "paper/stock_target_review_outcome_due.md",
                "paper_stock_target_review_outcome_analysis_status": "sample_insufficient",
                "paper_stock_target_review_outcome_analysis_sample_warning": "Sample size is below the configured anti-overfit readiness gates.",
                "paper_stock_target_review_outcome_analysis_ready_horizon_count": 0,
                "paper_stock_target_review_outcome_calendar_next_action_date": "2024-04-03",
                "paper_stock_target_review_outcome_calendar_next_action_horizon": "1d",
                "paper_stock_target_review_outcome_due_row_count": 1,
                "paper_stock_target_review_outcome_due_pending_count": 2,
                "paper_stock_target_review_outcome_due_next_date": "2024-04-03",
                "paper_stock_target_review_outcome_due_next_horizon": "1d",
                "paper_stock_target_review_outcome_maturity_next_1d_date": "2024-04-03",
                "paper_stock_target_review_outcome_maturity_next_5d_date": "2024-04-09",
                "paper_stock_target_review_outcome_maturity_next_10d_date": "2024-04-16",
                "paper_stock_target_review_outcome_maturity_next_20d_date": "2024-04-30",
                "paper_stock_target_review_required_count": 2,
                "paper_stock_target_review_trigger_count": 1,
                "paper_stock_target_review_drawdown_count": 1,
                "paper_stock_target_review_suppressed_layer_count": 1,
                "paper_stock_target_review_watch_count": 1,
                "paper_stock_target_review_notes_path": "paper/stock_target_review_notes.csv",
                "paper_stock_target_review_required_unreviewed_count": 1,
                "paper_stock_target_review_manual_watch_count": 1,
                "paper_stock_target_review_manual_exclude_candidate_count": 1,
                "paper_stock_target_review_manual_other_status_count": 1,
            }
        )

        codes = {item["code"] for item in payload["alerts"]}
        action_codes = {item["action_code"] for item in payload["action_items"]}
        self.assertEqual(payload["alert_level"], "warning")
        self.assertEqual(payload["action_stage"], "review_required")
        self.assertTrue(payload["action_required"])
        self.assertIn("stock_target_trigger_review", codes)
        self.assertIn("stock_target_drawdown_review", codes)
        self.assertIn("stock_target_suppressed_layer_review", codes)
        self.assertIn("stock_target_watch_review", codes)
        self.assertIn("stock_target_review_notes_open", codes)
        self.assertIn("stock_target_manual_watch", codes)
        self.assertIn("stock_target_manual_exclude_candidate", codes)
        self.assertIn("stock_target_manual_other_status", codes)
        self.assertIn("stock_target_outcome_due_queue", codes)
        self.assertIn("stock_target_outcome_analysis_pending", codes)
        self.assertIn("review_stock_target_trigger", action_codes)
        self.assertIn("review_stock_target_drawdown", action_codes)
        self.assertIn("update_stock_target_review_notes", action_codes)
        self.assertIn("review_manual_watch_targets", action_codes)
        self.assertIn("review_manual_exclusion_candidates", action_codes)
        self.assertIn("normalize_manual_review_status", action_codes)
        self.assertIn("refresh_due_outcome_samples", action_codes)
        self.assertIn("wait_stock_target_outcome_samples", action_codes)
        self.assertIn("paper/stock_target_review.md", payload["stock_target_review_report_path"])
        self.assertIn("paper/stock_target_review_actions.md", payload["stock_target_review_actions_report_path"])
        self.assertIn("paper/stock_target_review_assistant.md", payload["stock_target_review_assistant_report_path"])
        self.assertIn("paper/stock_target_review_outcomes.md", payload["stock_target_review_outcomes_report_path"])
        self.assertIn("paper/stock_target_review_outcomes_history.md", payload["stock_target_review_outcomes_history_report_path"])
        self.assertIn("paper/stock_target_review_outcome_analysis.md", payload["stock_target_review_outcome_analysis_report_path"])
        self.assertIn("paper/stock_target_review_outcome_calendar.md", payload["stock_target_review_outcome_calendar_report_path"])
        self.assertIn("paper/stock_target_review_outcome_due.md", payload["stock_target_review_outcome_due_report_path"])
        self.assertEqual(payload["stock_target_review_outcome_analysis_status"], "sample_insufficient")
        self.assertEqual(payload["stock_target_review_outcome_analysis_ready_horizon_count"], 0)
        self.assertEqual(payload["stock_target_review_outcome_maturity_next_1d_date"], "2024-04-03")
        self.assertEqual(payload["stock_target_review_outcome_calendar_next_action_date"], "2024-04-03")
        self.assertEqual(payload["stock_target_review_outcome_calendar_next_action_horizon"], "1d")
        self.assertEqual(payload["stock_target_review_outcome_due_row_count"], 1)
        self.assertEqual(payload["stock_target_review_outcome_due_next_date"], "2024-04-03")
        self.assertIn("paper/stock_target_review_notes.csv", payload["stock_target_review_notes_path"])
        self.assertEqual(payload["stock_target_review_manual_watch_count"], 1)
        self.assertEqual(payload["stock_target_review_manual_exclude_candidate_count"], 1)
        self.assertEqual(payload["stock_target_review_manual_other_status_count"], 1)

    def test_daily_alerts_carry_daily_preflight_skipped_flag(self) -> None:
        payload = build_daily_alert_payload(
            {
                "as_of_date": "2024-04-02",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "trading_day_data_ready",
                "daily_preflight_skipped": True,
            }
        )

        self.assertTrue(payload["daily_preflight_skipped"])
        with tempfile.TemporaryDirectory() as tmp:
            result = write_daily_alerts(
                {
                    "as_of_date": "2024-04-02",
                    "data_freshness_status": "fresh_enough",
                    "market_cache_status": "fresh_enough",
                    "allocator_input_status": "fresh_enough",
                    "paper_freshness_status": "fresh_enough",
                    "sentiment_freshness_status": "fresh_enough",
                    "trigger_freshness_status": "fresh_enough",
                    "dashboard_posture": "core_base_watch_allocator_gate",
                    "trading_day_gate_status": "trading_day_data_ready",
                    "daily_preflight_skipped": True,
                },
                output_dir=Path(tmp),
                stock_target_review_warning_only_after_close=False,
            )
            report_text = result.report_path.read_text(encoding="utf-8")
        self.assertIn("Daily preflight skipped", report_text)

    def test_daily_alerts_downgrade_reviewed_drawdown_rows_to_monitor(self) -> None:
        payload = build_daily_alert_payload(
            {
                "as_of_date": "2026-06-12",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "trading_day_data_ready",
                "paper_stock_target_review_report_path": "paper/stock_target_review.md",
                "paper_stock_target_review_actions_report_path": "paper/stock_target_review_actions.md",
                "paper_stock_target_review_drawdown_count": 3,
                "paper_stock_target_review_required_count": 3,
                "paper_stock_target_review_required_unreviewed_count": 0,
                "paper_stock_target_review_action_count": 0,
            }
        )

        codes = {item["code"] for item in payload["alerts"]}
        action_codes = {item["action_code"] for item in payload["action_items"]}
        self.assertEqual(payload["alert_level"], "info")
        self.assertEqual(payload["action_stage"], "monitor")
        self.assertFalse(payload["action_required"])
        self.assertIn("stock_target_drawdown_review_monitored", codes)
        self.assertNotIn("stock_target_drawdown_review", codes)
        self.assertNotIn("review_stock_target_drawdown", action_codes)

    def test_daily_alerts_flag_ready_stock_target_outcome_analysis_as_info(self) -> None:
        payload = build_daily_alert_payload(
            {
                "as_of_date": "2024-04-02",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "trading_day_data_ready",
                "paper_stock_target_review_outcome_analysis_report_path": "paper/stock_target_review_outcome_analysis.md",
                "paper_stock_target_review_outcome_analysis_status": "ready_for_review",
                "paper_stock_target_review_outcome_analysis_ready_horizon_count": 2,
            }
        )

        codes = {item["code"] for item in payload["alerts"]}
        action_codes = {item["action_code"] for item in payload["action_items"]}
        self.assertEqual(payload["alert_level"], "info")
        self.assertEqual(payload["action_stage"], "monitor")
        self.assertFalse(payload["action_required"])
        self.assertIn("stock_target_outcome_analysis_ready", codes)
        self.assertIn("review_stock_target_outcomes", action_codes)
        self.assertEqual(payload["stock_target_review_outcome_analysis_ready_horizon_count"], 2)

    def test_daily_alerts_include_satellite_risk_budget_wait_state(self) -> None:
        payload = build_daily_alert_payload(
            {
                "as_of_date": "2026-06-12",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "trading_day_data_ready",
                "satellite_risk_budget_status": "completed",
                "satellite_risk_budget_decision": "wait_for_outcome_samples",
                "satellite_risk_budget_reason": "Outcome samples have not reached a ready review horizon.",
                "satellite_risk_budget_recommended_satellite_weight": 0.0,
                "satellite_risk_budget_ready_horizon_count": 0,
                "satellite_risk_budget_report_path": "risk/satellite_risk_budget_review.md",
            }
        )

        codes = {item["code"] for item in payload["alerts"]}
        action_codes = {item["action_code"] for item in payload["action_items"]}
        self.assertEqual(payload["alert_level"], "info")
        self.assertEqual(payload["action_stage"], "monitor")
        self.assertIn("satellite_risk_budget_waiting", codes)
        self.assertIn("wait_satellite_risk_budget_samples", action_codes)
        self.assertEqual(payload["satellite_risk_budget_decision"], "wait_for_outcome_samples")
        self.assertEqual(payload["satellite_risk_budget_recommended_satellite_weight"], 0.0)

    def test_daily_alerts_flag_satellite_risk_budget_trial_as_info(self) -> None:
        payload = build_daily_alert_payload(
            {
                "as_of_date": "2026-06-20",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "trading_day_data_ready",
                "satellite_risk_budget_status": "completed",
                "satellite_risk_budget_decision": "eligible_for_small_satellite_trial",
                "satellite_risk_budget_reason": "Ready outcome horizon passes the guards.",
                "satellite_risk_budget_recommended_satellite_weight": 0.05,
                "satellite_risk_budget_selected_horizon": "1d",
                "satellite_risk_budget_ready_horizon_count": 1,
                "satellite_risk_budget_report_path": "risk/satellite_risk_budget_review.md",
            }
        )

        codes = {item["code"] for item in payload["alerts"]}
        action_codes = {item["action_code"] for item in payload["action_items"]}
        self.assertEqual(payload["alert_level"], "info")
        self.assertFalse(payload["action_required"])
        self.assertIn("satellite_risk_budget_trial_eligible", codes)
        self.assertIn("review_satellite_risk_budget_trial", action_codes)
        self.assertEqual(payload["satellite_risk_budget_selected_horizon"], "1d")

    def test_daily_alerts_warn_when_satellite_risk_budget_review_failed(self) -> None:
        payload = build_daily_alert_payload(
            {
                "as_of_date": "2026-06-20",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "trading_day_data_ready",
                "satellite_risk_budget_status": "failed",
                "satellite_risk_budget_decision": "review_failed",
                "satellite_risk_budget_reason": "Missing outcome analysis JSON.",
                "satellite_risk_budget_report_path": None,
            }
        )

        codes = {item["code"] for item in payload["alerts"]}
        action_codes = {item["action_code"] for item in payload["action_items"]}
        self.assertEqual(payload["alert_level"], "warning")
        self.assertEqual(payload["action_stage"], "review_required")
        self.assertIn("satellite_risk_budget_failed", codes)
        self.assertIn("inspect_satellite_risk_budget_review", action_codes)

    def test_daily_alerts_flag_live_shadow_preflight_blocked(self) -> None:
        payload = build_daily_alert_payload(
            {
                "as_of_date": "2026-06-20",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "trading_day_data_ready",
                "live_shadow_preflight_status": "completed",
                "live_shadow_preflight_decision": "blocked",
                "live_shadow_preflight_report_path": "outputs/research/live_shadow_preflight/live_shadow_preflight_report.md",
                "live_shadow_preflight_blockers_count": 2,
                "live_shadow_status": "blocked_by_preflight",
            }
        )

        codes = {item["code"] for item in payload["alerts"]}
        action_codes = {item["action_code"] for item in payload["action_items"]}
        self.assertEqual(payload["alert_level"], "warning")
        self.assertEqual(payload["action_stage"], "review_required")
        self.assertIn("live_shadow_preflight_blocked", codes)
        self.assertIn("review_live_shadow_preflight", action_codes)
        self.assertEqual(payload["live_shadow_preflight_status"], "completed")
        self.assertEqual(payload["live_shadow_status"], "blocked_by_preflight")
        self.assertEqual(payload["live_shadow_preflight_blockers_count"], 2)

        detail = next(item["detail"] for item in payload["alerts"] if item["code"] == "live_shadow_preflight_blocked")
        self.assertIn("blocker", detail.lower())
        self.assertIn("live_shadow_preflight_report.md", detail)

    def test_daily_alerts_warn_when_dashboard_history_refresh_failed(self) -> None:
        payload = build_daily_alert_payload(
            {
                "as_of_date": "2026-06-20",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "trading_day_data_ready",
                "history_review_status": "completed",
                "history_review_report_path": "history/pipeline_history_review.md",
                "dashboard_history_refresh_status": "failed",
                "dashboard_history_refresh_error": "Permission denied writing latest_dashboard.md.",
                "dashboard_report_path": "dashboard/latest_dashboard.md",
            }
        )

        codes = {item["code"] for item in payload["alerts"]}
        action_codes = {item["action_code"] for item in payload["action_items"]}
        self.assertEqual(payload["alert_level"], "warning")
        self.assertEqual(payload["action_stage"], "review_required")
        self.assertIn("dashboard_history_refresh_failed", codes)
        self.assertIn("inspect_dashboard_history_refresh", action_codes)
        self.assertEqual(payload["dashboard_history_refresh_status"], "failed")
        self.assertIn("Permission denied", payload["dashboard_history_refresh_error"])

    def test_daily_alerts_warn_when_history_review_alert_refresh_failed(self) -> None:
        payload = build_daily_alert_payload(
            {
                "as_of_date": "2026-06-20",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "trading_day_data_ready",
                "history_review_status": "completed",
                "history_review_report_path": "history/pipeline_history_review.md",
                "history_review_alert_refresh_status": "failed",
                "history_review_alert_refresh_error": "Permission denied rewriting pipeline_history_review.md.",
            }
        )

        codes = {item["code"] for item in payload["alerts"]}
        action_codes = {item["action_code"] for item in payload["action_items"]}
        self.assertEqual(payload["alert_level"], "warning")
        self.assertEqual(payload["action_stage"], "review_required")
        self.assertIn("history_review_alert_refresh_failed", codes)
        self.assertIn("inspect_history_review_alert_refresh", action_codes)
        self.assertEqual(payload["history_review_alert_refresh_status"], "failed")
        self.assertIn("Permission denied", payload["history_review_alert_refresh_error"])

    def test_daily_alerts_warn_when_stock_market_cap_cache_unavailable_or_stale(self) -> None:
        unavailable = build_daily_alert_payload(
            {
                "as_of_date": "2024-04-02",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "trading_day_data_ready",
                "paper_stock_market_cap_cache_status": "missing",
                "paper_stock_market_cap_cache_row_count": 0,
                "paper_stock_market_cap_path": "data/processed/stock_market_cap_yi.csv",
                "paper_stock_tracking_max_market_cap_yi": 1500.0,
            }
        )
        unavailable_codes = {item["code"] for item in unavailable["alerts"]}
        unavailable_actions = {item["action_code"] for item in unavailable["action_items"]}
        self.assertEqual(unavailable["alert_level"], "warning")
        self.assertEqual(unavailable["action_stage"], "review_required")
        self.assertIn("stock_market_cap_cache_unavailable", unavailable_codes)
        self.assertIn("refresh_stock_market_cap_cache", unavailable_actions)

        stale = build_daily_alert_payload(
            {
                "as_of_date": "2024-04-02",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "trading_day_data_ready",
                "paper_latest_date": "2024-04-02",
                "paper_stock_market_cap_cache_status": "ok",
                "paper_stock_market_cap_cache_row_count": 10,
                "paper_stock_market_cap_cache_latest_snapshot_date": "2024-04-01",
                "paper_stock_market_cap_cache_updated_at": "2024-04-01T16:00:00",
                "paper_stock_market_cap_path": "data/processed/stock_market_cap_yi.csv",
                "paper_stock_tracking_max_market_cap_yi": 1500.0,
            }
        )
        stale_codes = {item["code"] for item in stale["alerts"]}
        self.assertEqual(stale["alert_level"], "warning")
        self.assertIn("stock_market_cap_cache_stale", stale_codes)
        self.assertEqual(stale["stock_market_cap_cache_latest_snapshot_date"], "2024-04-01")

    def test_daily_alerts_info_when_some_stock_targets_missing_market_cap(self) -> None:
        payload = build_daily_alert_payload(
            {
                "as_of_date": "2024-04-02",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "trading_day_data_ready",
                "paper_latest_date": "2024-04-02",
                "paper_stock_target_count": 10,
                "paper_stock_market_cap_cache_status": "ok",
                "paper_stock_market_cap_cache_row_count": 8,
                "paper_stock_market_cap_cache_latest_snapshot_date": "2024-04-02",
                "paper_stock_tracking_market_cap_missing_count": 2,
                "paper_stock_market_cap_path": "data/processed/stock_market_cap_yi.csv",
                "paper_stock_tracking_max_market_cap_yi": 1500.0,
            }
        )

        codes = {item["code"] for item in payload["alerts"]}
        action_codes = {item["action_code"] for item in payload["action_items"]}
        self.assertEqual(payload["alert_level"], "info")
        self.assertEqual(payload["action_stage"], "monitor")
        self.assertFalse(payload["action_required"])
        self.assertIn("stock_market_cap_target_missing", codes)
        self.assertIn("review_stock_market_cap_coverage", action_codes)
        self.assertEqual(payload["stock_tracking_market_cap_missing_count"], 2)

    def test_daily_alerts_can_downgrade_stock_review_on_non_trading_day(self) -> None:
        payload = build_daily_alert_payload(
            {
                "as_of_date": "2026-06-14",
                "data_freshness_status": "fresh_enough",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "trading_day_gate_status": "non_trading_day",
                "trading_day_evidence": "weekend_rule",
                "paper_stock_target_review_report_path": "paper/stock_target_review.md",
                "paper_stock_target_review_required_count": 1,
                "paper_stock_target_review_drawdown_count": 1,
                "paper_stock_target_review_required_unreviewed_count": 1,
            },
            stock_target_review_warning_only_after_close=True,
        )

        codes = {item["code"] for item in payload["alerts"]}
        self.assertEqual(payload["alert_level"], "info")
        self.assertEqual(payload["action_stage"], "monitor")
        self.assertFalse(payload["action_required"])
        self.assertIn("non_trading_day", codes)
        self.assertIn("stock_target_drawdown_review", codes)
        self.assertTrue(payload["stock_target_review_warning_only_after_close"])

    def test_pipeline_history_review_flags_watch_alerts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            history = root / "daily_pipeline_history.csv"
            pd.DataFrame(
                [
                    {
                        "generated_at": "2024-04-01T16:00:00",
                        "as_of_date": "2024-04-01",
                        "local_equity_latest_date": "2024-04-01",
                        "data_freshness_status": "fresh_enough",
                        "phase2_posture": "core_base_allocator_gated_satellite",
                        "action_posture": "review_core_base_allocator_gate",
                        "dashboard_posture": "core_base_watch_allocator_gate",
                        "market_cache_status": "fresh_enough",
                        "allocator_input_status": "fresh_enough",
                        "paper_freshness_status": "fresh_enough",
                        "sentiment_state": "warm",
                        "sentiment_freshness_status": "fresh_enough",
                        "trigger_freshness_status": "fresh_enough",
                        "paper_latest_date": "2024-04-01",
                        "paper_latest_regime": "risk_on",
                        "paper_latest_candidate": "sat20_test",
                        "paper_latest_core_weight": 0.8,
                        "paper_latest_satellite_weight": 0.2,
                        "paper_latest_cash_weight": 0.0,
                        "paper_final_equity": 1000000.0,
                        "paper_total_return": 0.0,
                        "paper_max_drawdown": -0.04,
                        "paper_current_drawdown": -0.02,
                        "paper_sharpe": 0.9,
                    },
                    {
                        "generated_at": "2024-04-02T16:00:00",
                        "as_of_date": "2024-04-02",
                        "local_equity_latest_date": "2024-04-02",
                        "data_freshness_status": "fresh_enough",
                        "phase2_posture": "core_base_allocator_gated_satellite",
                        "action_posture": "review_core_base_allocator_gate",
                        "dashboard_posture": "core_base_watch_allocator_gate",
                        "market_cache_status": "fresh_enough",
                        "allocator_input_status": "fresh_enough",
                        "paper_freshness_status": "fresh_enough",
                        "sentiment_state": "warm",
                        "sentiment_freshness_status": "fresh_enough",
                        "trigger_freshness_status": "fresh_enough",
                        "paper_latest_date": "2024-04-02",
                        "paper_latest_regime": "risk_off",
                        "paper_latest_candidate": "sat20_test",
                        "paper_latest_core_weight": 1.0,
                        "paper_latest_satellite_weight": 0.0,
                        "paper_latest_cash_weight": 0.0,
                        "paper_final_equity": 990000.0,
                        "paper_total_return": -0.01,
                        "paper_max_drawdown": -0.09,
                        "paper_current_drawdown": -0.09,
                        "paper_sharpe": 0.7,
                    },
                ]
            ).to_csv(history, index=False)

            result = run_pipeline_history_review(
                project_root=root,
                history_file=history,
                output_dir=root / "review",
                as_of_date="2024-04-03",
                lookback_runs=5,
                max_staleness_days=2,
                drawdown_watch_threshold=-0.08,
                min_sharpe_watch=0.5,
            )

            self.assertTrue(result.report_path.exists())
            self.assertTrue(result.snapshot_path.exists())
            self.assertEqual(result.snapshot["health_state"], "watch")
            self.assertEqual(result.snapshot["history_freshness_status"], "fresh_enough")
            self.assertEqual(result.snapshot["latest_as_of_date"], "2024-04-02")
            self.assertLess(result.snapshot["lookback_final_equity_delta"], 0.0)
            self.assertGreaterEqual(result.snapshot["alert_count"], 1)
            self.assertIn("Pipeline History Review", result.report_path.read_text(encoding="utf-8"))

    def test_pipeline_history_review_flags_correlated_promotion_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            history = root / "daily_pipeline_history.csv"
            pd.DataFrame(
                [
                    {
                        "generated_at": "2026-06-13T16:00:00",
                        "as_of_date": "2026-06-13",
                        "local_equity_latest_date": "2026-06-13",
                        "data_freshness_status": "fresh_enough",
                        "phase2_posture": "core_base_allocator_gated_satellite",
                        "action_posture": "review_core_base_allocator_gate",
                        "dashboard_posture": "core_base_watch_allocator_gate",
                        "market_cache_status": "fresh_enough",
                        "allocator_input_status": "fresh_enough",
                        "paper_freshness_status": "fresh_enough",
                        "sentiment_freshness_status": "fresh_enough",
                        "trigger_freshness_status": "fresh_enough",
                        "promotion_status": "ok",
                        "promotion_decision": "watch_candidate",
                        "promotion_candidate_passes_headline_gate": True,
                        "promotion_sensitivity_group_support_count": 1,
                        "promotion_sensitivity_run_support_count": 2,
                        "promotion_min_sensitivity_support": 2,
                        "paper_latest_regime": "risk_off",
                        "paper_latest_candidate": "sat20_test",
                        "paper_latest_core_weight": 1.0,
                        "paper_latest_satellite_weight": 0.0,
                        "paper_latest_cash_weight": 0.0,
                        "paper_final_equity": 1000000.0,
                        "paper_total_return": 0.1,
                        "paper_max_drawdown": -0.04,
                        "paper_current_drawdown": -0.02,
                        "paper_sharpe": 1.2,
                    }
                ]
            ).to_csv(history, index=False)

            result = run_pipeline_history_review(
                project_root=root,
                history_file=history,
                output_dir=root / "review",
                as_of_date="2026-06-14",
                lookback_runs=5,
                max_staleness_days=2,
                drawdown_watch_threshold=-0.08,
                min_sharpe_watch=0.5,
            )

            self.assertEqual(result.snapshot["health_state"], "watch")
            self.assertEqual(result.snapshot["latest_promotion_decision"], "watch_candidate")
            self.assertEqual(int(result.snapshot["latest_promotion_group_support_count"]), 1)
            self.assertEqual(int(result.snapshot["latest_promotion_run_support_count"]), 2)
            self.assertEqual(int(result.snapshot["latest_promotion_min_support"]), 2)
            self.assertTrue(any("correlated" in item for item in result.snapshot["alerts"]))
            report_text = result.report_path.read_text(encoding="utf-8")
            self.assertIn("Promotion decision", report_text)
            self.assertIn("Promotion group/run support", report_text)
            self.assertIn("1 / 2 / run 2", report_text)

    def test_pipeline_history_review_tracks_satellite_risk_budget_trial_change(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            history = root / "daily_pipeline_history.csv"
            base_row = {
                "local_equity_latest_date": "2026-06-11",
                "data_freshness_status": "fresh_enough",
                "phase2_posture": "core_base_allocator_gated_satellite",
                "action_posture": "review_core_base_allocator_gate",
                "dashboard_posture": "core_base_watch_allocator_gate",
                "market_cache_status": "fresh_enough",
                "allocator_input_status": "fresh_enough",
                "paper_freshness_status": "fresh_enough",
                "sentiment_freshness_status": "fresh_enough",
                "trigger_freshness_status": "fresh_enough",
                "paper_latest_regime": "risk_off",
                "paper_latest_candidate": "sat20_test",
                "paper_latest_core_weight": 1.0,
                "paper_latest_satellite_weight": 0.0,
                "paper_latest_cash_weight": 0.0,
                "paper_final_equity": 1000000.0,
                "paper_total_return": 0.1,
                "paper_max_drawdown": -0.04,
                "paper_current_drawdown": -0.02,
                "paper_sharpe": 1.2,
                "satellite_risk_budget_status": "completed",
                "satellite_risk_budget_reason": "Outcome samples have not reached a ready review horizon.",
                "satellite_risk_budget_selected_horizon": "",
                "satellite_risk_budget_report_path": "risk/wait.md",
                "satellite_risk_budget_snapshot_path": "risk/wait.json",
                "satellite_risk_budget_checklist_path": "risk/wait.csv",
            }
            pd.DataFrame(
                [
                    {
                        **base_row,
                        "generated_at": "2026-06-11T16:00:00",
                        "as_of_date": "2026-06-11",
                        "satellite_risk_budget_decision": "wait_for_outcome_samples",
                        "satellite_risk_budget_recommended_satellite_weight": 0.0,
                    },
                    {
                        **base_row,
                        "generated_at": "2026-06-12T16:00:00",
                        "as_of_date": "2026-06-12",
                        "satellite_risk_budget_reason": "Ready outcome horizon passes the guards.",
                        "satellite_risk_budget_decision": "eligible_for_small_satellite_trial",
                        "satellite_risk_budget_recommended_satellite_weight": 0.05,
                        "satellite_risk_budget_selected_horizon": "1d",
                        "satellite_risk_budget_report_path": "risk/trial.md",
                    },
                ]
            ).to_csv(history, index=False)

            result = run_pipeline_history_review(
                project_root=root,
                history_file=history,
                output_dir=root / "review",
                as_of_date="2026-06-12",
                lookback_runs=5,
                max_staleness_days=2,
                drawdown_watch_threshold=-0.08,
                min_sharpe_watch=0.5,
            )

            self.assertEqual(result.snapshot["health_state"], "watch")
            self.assertEqual(result.snapshot["latest_satellite_risk_budget_status"], "completed")
            self.assertEqual(
                result.snapshot["latest_satellite_risk_budget_decision"],
                "eligible_for_small_satellite_trial",
            )
            self.assertEqual(result.snapshot["previous_satellite_risk_budget_decision"], "wait_for_outcome_samples")
            self.assertTrue(result.snapshot["satellite_risk_budget_decision_changed"])
            self.assertAlmostEqual(result.snapshot["latest_satellite_risk_budget_recommended_satellite_weight"], 0.05)
            self.assertAlmostEqual(result.snapshot["satellite_risk_budget_recommended_weight_change"], 0.05)
            self.assertTrue(any("eligible for a small satellite trial" in item for item in result.snapshot["alerts"]))
            self.assertTrue(any("decision changed" in item for item in result.snapshot["alerts"]))
            report_text = result.report_path.read_text(encoding="utf-8")
            self.assertIn("Satellite Risk Budget", report_text)
            self.assertIn("eligible_for_small_satellite_trial", report_text)
            self.assertIn("5.00%", report_text)

    def test_pipeline_history_review_flags_satellite_risk_budget_failure(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            history = root / "daily_pipeline_history.csv"
            pd.DataFrame(
                [
                    {
                        "generated_at": "2026-06-12T16:00:00",
                        "as_of_date": "2026-06-12",
                        "local_equity_latest_date": "2026-06-12",
                        "data_freshness_status": "fresh_enough",
                        "phase2_posture": "core_base_allocator_gated_satellite",
                        "action_posture": "review_core_base_allocator_gate",
                        "dashboard_posture": "core_base_watch_allocator_gate",
                        "market_cache_status": "fresh_enough",
                        "allocator_input_status": "fresh_enough",
                        "paper_freshness_status": "fresh_enough",
                        "sentiment_freshness_status": "fresh_enough",
                        "trigger_freshness_status": "fresh_enough",
                        "paper_latest_regime": "risk_off",
                        "paper_latest_candidate": "sat20_test",
                        "paper_latest_core_weight": 1.0,
                        "paper_latest_satellite_weight": 0.0,
                        "paper_latest_cash_weight": 0.0,
                        "paper_final_equity": 1000000.0,
                        "paper_total_return": 0.1,
                        "paper_max_drawdown": -0.04,
                        "paper_current_drawdown": -0.02,
                        "paper_sharpe": 1.2,
                        "satellite_risk_budget_status": "failed",
                        "satellite_risk_budget_decision": "review_failed",
                        "satellite_risk_budget_reason": "Missing outcome analysis JSON.",
                        "satellite_risk_budget_recommended_satellite_weight": "",
                        "satellite_risk_budget_selected_horizon": "",
                        "satellite_risk_budget_report_path": "",
                    }
                ]
            ).to_csv(history, index=False)

            result = run_pipeline_history_review(
                project_root=root,
                history_file=history,
                output_dir=root / "review",
                as_of_date="2026-06-12",
                lookback_runs=5,
                max_staleness_days=2,
                drawdown_watch_threshold=-0.08,
                min_sharpe_watch=0.5,
            )

            self.assertEqual(result.snapshot["health_state"], "watch")
            self.assertEqual(result.snapshot["latest_satellite_risk_budget_status"], "failed")
            self.assertEqual(result.snapshot["latest_satellite_risk_budget_decision"], "review_failed")
            self.assertTrue(any("review failed" in item for item in result.snapshot["alerts"]))
            self.assertIn("Missing outcome analysis JSON", result.report_path.read_text(encoding="utf-8"))

    def test_pipeline_history_review_flags_dashboard_history_refresh_failure(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            history = root / "daily_pipeline_history.csv"
            pd.DataFrame(
                [
                    {
                        "generated_at": "2026-06-12T16:00:00",
                        "as_of_date": "2026-06-12",
                        "local_equity_latest_date": "2026-06-12",
                        "data_freshness_status": "fresh_enough",
                        "phase2_posture": "core_base_allocator_gated_satellite",
                        "action_posture": "review_core_base_allocator_gate",
                        "dashboard_posture": "core_base_watch_allocator_gate",
                        "market_cache_status": "fresh_enough",
                        "allocator_input_status": "fresh_enough",
                        "paper_freshness_status": "fresh_enough",
                        "sentiment_freshness_status": "fresh_enough",
                        "trigger_freshness_status": "fresh_enough",
                        "paper_latest_regime": "risk_off",
                        "paper_latest_candidate": "sat20_test",
                        "paper_latest_core_weight": 1.0,
                        "paper_latest_satellite_weight": 0.0,
                        "paper_latest_cash_weight": 0.0,
                        "paper_final_equity": 1000000.0,
                        "paper_total_return": 0.1,
                        "paper_max_drawdown": -0.04,
                        "paper_current_drawdown": -0.02,
                        "paper_sharpe": 1.2,
                        "dashboard_history_refresh_status": "failed",
                        "dashboard_history_refresh_error": "Permission denied writing latest_dashboard.md.",
                        "dashboard_history_refresh_report_path": "dashboard/latest_dashboard.md",
                        "dashboard_pipeline_history_status": "missing",
                        "dashboard_pipeline_history_health_state": "unknown",
                        "dashboard_pipeline_history_latest_as_of_date": "",
                    }
                ]
            ).to_csv(history, index=False)

            result = run_pipeline_history_review(
                project_root=root,
                history_file=history,
                output_dir=root / "review",
                as_of_date="2026-06-12",
                lookback_runs=5,
                max_staleness_days=2,
                drawdown_watch_threshold=-0.08,
                min_sharpe_watch=0.5,
            )

            self.assertEqual(result.snapshot["health_state"], "watch")
            self.assertEqual(result.snapshot["latest_dashboard_history_refresh_status"], "failed")
            self.assertEqual(result.snapshot["latest_dashboard_pipeline_history_health_state"], "unknown")
            self.assertEqual(result.snapshot["lookback_dashboard_history_refresh_failed_runs"], 1)
            self.assertTrue(any("Dashboard history refresh failed" in item for item in result.snapshot["alerts"]))
            report_text = result.report_path.read_text(encoding="utf-8")
            self.assertIn("Dashboard History Refresh", report_text)
            self.assertIn("Permission denied writing latest_dashboard.md", report_text)
            self.assertIn("dashboard/latest_dashboard.md", report_text)

    def test_pipeline_history_review_flags_dashboard_alert_refresh_failure(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            history = root / "daily_pipeline_history.csv"
            pd.DataFrame(
                [
                    {
                        "generated_at": "2026-06-12T16:00:00",
                        "as_of_date": "2026-06-12",
                        "local_equity_latest_date": "2026-06-12",
                        "data_freshness_status": "fresh_enough",
                        "phase2_posture": "core_base_allocator_gated_satellite",
                        "action_posture": "review_core_base_allocator_gate",
                        "dashboard_posture": "core_base_watch_allocator_gate",
                        "market_cache_status": "fresh_enough",
                        "allocator_input_status": "fresh_enough",
                        "paper_freshness_status": "fresh_enough",
                        "sentiment_freshness_status": "fresh_enough",
                        "trigger_freshness_status": "fresh_enough",
                        "paper_latest_regime": "risk_off",
                        "paper_latest_candidate": "sat20_test",
                        "paper_latest_core_weight": 1.0,
                        "paper_latest_satellite_weight": 0.0,
                        "paper_latest_cash_weight": 0.0,
                        "paper_final_equity": 1000000.0,
                        "paper_total_return": 0.1,
                        "paper_max_drawdown": -0.04,
                        "paper_current_drawdown": -0.02,
                        "paper_sharpe": 1.2,
                        "dashboard_history_refresh_status": "completed",
                        "dashboard_history_refresh_report_path": "dashboard/latest_dashboard.md",
                        "dashboard_pipeline_history_status": "ok",
                        "dashboard_pipeline_history_health_state": "ok",
                        "dashboard_pipeline_history_latest_as_of_date": "",
                        "dashboard_alert_refresh_status": "failed",
                        "dashboard_alert_refresh_error": "Permission denied reading latest_alerts.json.",
                    }
                ]
            ).to_csv(history, index=False)

            result = run_pipeline_history_review(
                project_root=root,
                history_file=history,
                output_dir=root / "review",
                as_of_date="2026-06-12",
                lookback_runs=5,
                max_staleness_days=2,
                drawdown_watch_threshold=-0.08,
                min_sharpe_watch=0.5,
            )

            self.assertEqual(result.snapshot["health_state"], "watch")
            self.assertEqual(result.snapshot["latest_dashboard_alert_refresh_status"], "failed")
            self.assertEqual(result.snapshot["latest_dashboard_alert_refresh_error"], "Permission denied reading latest_alerts.json.")
            self.assertEqual(result.snapshot["lookback_dashboard_alert_refresh_failed_runs"], 1)
            self.assertTrue(any("Dashboard alert refresh failed" in item for item in result.snapshot["alerts"]))
            report_text = result.report_path.read_text(encoding="utf-8")
            self.assertIn("Alert refresh status", report_text)
            self.assertIn("Permission denied reading latest_alerts.json.", report_text)
            self.assertIn("Dashboard alert refresh failed", report_text)

    def test_pipeline_history_review_flags_dashboard_alert_refresh_repeated_failures(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            history = root / "daily_pipeline_history.csv"
            pd.DataFrame(
                [
                    {
                        "generated_at": "2026-06-10T16:00:00",
                        "as_of_date": "2026-06-10",
                        "local_equity_latest_date": "2026-06-10",
                        "data_freshness_status": "fresh_enough",
                        "phase2_posture": "core_base_allocator_gated_satellite",
                        "action_posture": "review_core_base_allocator_gate",
                        "dashboard_posture": "core_base_watch_allocator_gate",
                        "market_cache_status": "fresh_enough",
                        "allocator_input_status": "fresh_enough",
                        "paper_freshness_status": "fresh_enough",
                        "sentiment_freshness_status": "fresh_enough",
                        "trigger_freshness_status": "fresh_enough",
                        "paper_latest_regime": "risk_off",
                        "paper_latest_core_weight": 1.0,
                        "paper_latest_satellite_weight": 0.0,
                        "paper_latest_cash_weight": 0.0,
                        "paper_final_equity": 1000000.0,
                        "paper_total_return": 0.1,
                        "paper_max_drawdown": -0.04,
                        "paper_current_drawdown": -0.02,
                        "paper_sharpe": 1.2,
                        "dashboard_alert_refresh_status": "failed",
                        "dashboard_alert_refresh_error": "Permission denied in failure window 1.",
                    },
                    {
                        "generated_at": "2026-06-11T16:00:00",
                        "as_of_date": "2026-06-11",
                        "local_equity_latest_date": "2026-06-11",
                        "data_freshness_status": "fresh_enough",
                        "phase2_posture": "core_base_allocator_gated_satellite",
                        "action_posture": "review_core_base_allocator_gate",
                        "dashboard_posture": "core_base_watch_allocator_gate",
                        "market_cache_status": "fresh_enough",
                        "allocator_input_status": "fresh_enough",
                        "paper_freshness_status": "fresh_enough",
                        "sentiment_freshness_status": "fresh_enough",
                        "trigger_freshness_status": "fresh_enough",
                        "paper_latest_regime": "risk_off",
                        "paper_latest_core_weight": 1.0,
                        "paper_latest_satellite_weight": 0.0,
                        "paper_latest_cash_weight": 0.0,
                        "paper_final_equity": 1000000.0,
                        "paper_total_return": 0.1,
                        "paper_max_drawdown": -0.04,
                        "paper_current_drawdown": -0.02,
                        "paper_sharpe": 1.2,
                        "dashboard_alert_refresh_status": "failed",
                        "dashboard_alert_refresh_error": "Permission denied in failure window 2.",
                    },
                    {
                        "generated_at": "2026-06-12T16:00:00",
                        "as_of_date": "2026-06-12",
                        "local_equity_latest_date": "2026-06-12",
                        "data_freshness_status": "fresh_enough",
                        "phase2_posture": "core_base_allocator_gated_satellite",
                        "action_posture": "review_core_base_allocator_gate",
                        "dashboard_posture": "core_base_watch_allocator_gate",
                        "market_cache_status": "fresh_enough",
                        "allocator_input_status": "fresh_enough",
                        "paper_freshness_status": "fresh_enough",
                        "sentiment_freshness_status": "fresh_enough",
                        "trigger_freshness_status": "fresh_enough",
                        "paper_latest_regime": "risk_off",
                        "paper_latest_core_weight": 1.0,
                        "paper_latest_satellite_weight": 0.0,
                        "paper_latest_cash_weight": 0.0,
                        "paper_final_equity": 1000000.0,
                        "paper_total_return": 0.1,
                        "paper_max_drawdown": -0.04,
                        "paper_current_drawdown": -0.02,
                        "paper_sharpe": 1.2,
                        "dashboard_alert_refresh_status": "completed",
                        "dashboard_alert_refresh_error": "",
                        "dashboard_history_refresh_status": "completed",
                        "dashboard_history_refresh_report_path": "dashboard/latest_dashboard.md",
                        "dashboard_pipeline_history_status": "ok",
                        "dashboard_pipeline_history_health_state": "ok",
                        "dashboard_pipeline_history_latest_as_of_date": "",
                    },
                ]
            ).to_csv(history, index=False)

            result = run_pipeline_history_review(
                project_root=root,
                history_file=history,
                output_dir=root / "review",
                as_of_date="2026-06-12",
                lookback_runs=5,
                max_staleness_days=2,
                drawdown_watch_threshold=-0.08,
                min_sharpe_watch=0.5,
            )

            self.assertEqual(result.snapshot["health_state"], "watch")
            self.assertEqual(result.snapshot["lookback_dashboard_alert_refresh_failed_runs"], 2)
            self.assertTrue(
                any(
                    "failed 2 times in the lookback window" in str(item).lower()
                    for item in result.snapshot["alerts"]
                )
            )

    def test_pipeline_history_review_flags_daily_alert_warning(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            history = root / "daily_pipeline_history.csv"
            pd.DataFrame(
                [
                    {
                        "generated_at": "2026-06-11T16:00:00",
                        "as_of_date": "2026-06-11",
                        "local_equity_latest_date": "2026-06-11",
                        "data_freshness_status": "fresh_enough",
                        "phase2_posture": "core_base_allocator_gated_satellite",
                        "action_posture": "review_core_base_allocator_gate",
                        "dashboard_posture": "core_base_watch_allocator_gate",
                        "market_cache_status": "fresh_enough",
                        "allocator_input_status": "fresh_enough",
                        "paper_freshness_status": "fresh_enough",
                        "sentiment_freshness_status": "fresh_enough",
                        "trigger_freshness_status": "fresh_enough",
                        "paper_latest_regime": "risk_off",
                        "paper_latest_candidate": "sat20_test",
                        "paper_latest_core_weight": 1.0,
                        "paper_latest_satellite_weight": 0.0,
                        "paper_latest_cash_weight": 0.0,
                        "paper_final_equity": 1000000.0,
                        "paper_total_return": 0.1,
                        "paper_max_drawdown": -0.04,
                        "paper_current_drawdown": -0.02,
                        "paper_sharpe": 1.2,
                        "alert_level": "info",
                        "alert_count": 1,
                        "alert_action_stage": "monitor",
                        "alert_action_summary": "Routine monitor.",
                        "alert_action_item_count": 1,
                        "alert_action_required": False,
                        "alert_warning_count": 0,
                        "alert_critical_count": 0,
                        "alert_info_count": 1,
                        "alerts_report_path": "alerts/previous.md",
                    },
                    {
                        "generated_at": "2026-06-12T16:00:00",
                        "as_of_date": "2026-06-12",
                        "local_equity_latest_date": "2026-06-12",
                        "data_freshness_status": "fresh_enough",
                        "phase2_posture": "core_base_allocator_gated_satellite",
                        "action_posture": "review_core_base_allocator_gate",
                        "dashboard_posture": "core_base_watch_allocator_gate",
                        "market_cache_status": "fresh_enough",
                        "allocator_input_status": "fresh_enough",
                        "paper_freshness_status": "fresh_enough",
                        "sentiment_freshness_status": "fresh_enough",
                        "trigger_freshness_status": "fresh_enough",
                        "paper_latest_regime": "risk_off",
                        "paper_latest_candidate": "sat20_test",
                        "paper_latest_core_weight": 1.0,
                        "paper_latest_satellite_weight": 0.0,
                        "paper_latest_cash_weight": 0.0,
                        "paper_final_equity": 1005000.0,
                        "paper_total_return": 0.105,
                        "paper_max_drawdown": -0.04,
                        "paper_current_drawdown": -0.02,
                        "paper_sharpe": 1.2,
                        "alert_level": "warning",
                        "alert_count": 2,
                        "alert_action_stage": "review_required",
                        "alert_action_summary": "Review model audit and stock targets.",
                        "alert_action_item_count": 2,
                        "alert_action_required": True,
                        "alert_warning_count": 2,
                        "alert_critical_count": 0,
                        "alert_info_count": 0,
                        "alerts_report_path": "alerts/latest.md",
                    },
                ]
            ).to_csv(history, index=False)

            result = run_pipeline_history_review(
                project_root=root,
                history_file=history,
                output_dir=root / "review",
                as_of_date="2026-06-12",
                lookback_runs=5,
                max_staleness_days=2,
                drawdown_watch_threshold=-0.08,
                min_sharpe_watch=0.5,
            )

            self.assertEqual(result.snapshot["health_state"], "watch")
            self.assertEqual(result.snapshot["latest_alert_level"], "warning")
            self.assertEqual(result.snapshot["latest_alert_action_stage"], "review_required")
            self.assertEqual(result.snapshot["latest_alert_count"], 2)
            self.assertEqual(result.snapshot["lookback_alert_warning_or_critical_runs"], 1)
            self.assertEqual(result.snapshot["lookback_alert_level_changes"], 1)
            self.assertTrue(any("Daily alert level is warning" in item for item in result.snapshot["alerts"]))
            report_text = result.report_path.read_text(encoding="utf-8")
            self.assertIn("Daily Alerts", report_text)
            self.assertIn("review_required", report_text)
            self.assertIn("alerts/latest.md", report_text)

    def test_phase2_status_marks_complete_and_incomplete_layers(self) -> None:
        def write_metrics(path: Path, total_return: float, max_drawdown: float, sharpe: float) -> None:
            path.mkdir(parents=True, exist_ok=True)
            payload = {"total_return": total_return, "max_drawdown": max_drawdown, "sharpe": sharpe}
            (path / "metrics.json").write_text(json.dumps(payload), encoding="utf-8")

        def write_walk_forward(path: Path, portfolio: bool = False) -> None:
            path.mkdir(parents=True, exist_ok=True)
            summary_name = "portfolio_walk_forward_summary.csv" if portfolio else "walk_forward_summary.csv"
            candidate_name = "portfolio_candidate_results.csv" if portfolio else "candidate_results.csv"
            pd.DataFrame(
                {
                    "window": ["w1", "w2"],
                    "selected_candidate": ["core_only", "sat20"],
                    "test_total_return": [0.10, -0.02],
                }
            ).to_csv(path / summary_name, index=False)
            pd.DataFrame({"candidate": ["core_only", "sat20"]}).to_csv(path / candidate_name, index=False)
            pd.DataFrame(
                {
                    "date": pd.date_range("2020-01-01", periods=4, freq="D"),
                    "window": ["w1"] * 4,
                    "stitched_equity": [100.0, 110.0, 105.0, 120.0],
                }
            ).to_csv(path / "oos_equity_stitched.csv", index=False)

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            baseline = root / "baseline"
            core = root / "core"
            satellite = root / "satellite"
            portfolio = root / "portfolio"
            allocator = root / "allocator"
            write_metrics(baseline, 0.40, -0.20, 0.8)
            write_walk_forward(core)
            satellite.mkdir()
            write_metrics(portfolio, 0.22, -0.10, 0.7)
            write_walk_forward(allocator, portfolio=True)

            result = run_phase2_status(
                project_root=root,
                output_dir=root / "phase2",
                baseline_run_dir=baseline,
                core_wf_dir=core,
                satellite_wf_dir=satellite,
                portfolio_run_dir=portfolio,
                portfolio_wf_dir=allocator,
            )

            self.assertTrue(result.components_path.exists())
            self.assertTrue(result.report_path.exists())
            statuses = dict(zip(result.components["layer"], result.components["status"]))
            self.assertEqual(statuses["baseline"], "complete")
            self.assertEqual(statuses["core"], "complete")
            self.assertEqual(statuses["satellite"], "incomplete")
            self.assertEqual(statuses["portfolio"], "complete")
            self.assertEqual(statuses["allocator"], "complete")
            self.assertIn("Phase 2 scaffold is usable", result.report_path.read_text(encoding="utf-8"))

    def test_phase2_status_marks_active_satellite_lock_as_running(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            satellite = root / "main_chinext_satellite_quality_v2_full"
            satellite.mkdir()
            seed_process_lock(
                root / "outputs" / "locks",
                "walk-forward|main_chinext_satellite_quality_v2_full",
                os.getpid(),
                "python -m quant_etf_lab walk-forward --run-id-prefix main_chinext_satellite_quality_v2_full",
            )

            result = run_phase2_status(
                project_root=root,
                output_dir=root / "phase2",
                satellite_wf_dir=satellite,
            )

            satellite_row = result.components[result.components["layer"] == "satellite"].iloc[0]
            self.assertEqual(satellite_row["status"], "running")
            self.assertIn("PID", satellite_row["note"])

    def test_process_lock_blocks_running_equivalent_command(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            lock_dir = Path(tmp)
            seed_process_lock(lock_dir, "same-command", os.getpid(), "python -m quant_etf_lab walk-forward")
            with self.assertRaises(ProcessLockError):
                with process_lock(lock_dir, "same-command", "python -m quant_etf_lab walk-forward"):
                    pass

    def test_process_lock_replaces_stale_lock(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            lock_dir = Path(tmp)
            seed_process_lock(lock_dir, "same-command", 99999999, "stale command")
            with process_lock(lock_dir, "same-command", "fresh command"):
                self.assertTrue(any(lock_dir.iterdir()))
            self.assertFalse(any(lock_dir.iterdir()))

    def test_walk_forward_lock_key_ignores_window_limit_for_same_output_run(self) -> None:
        parser = build_parser()
        smoke_args = parser.parse_args(
            [
                "walk-forward",
                "--config",
                "configs/test.yaml",
                "--output-dir",
                "outputs/walk_forward",
                "--run-id-prefix",
                "same_run",
                "--window-limit",
                "4",
            ]
        )
        resume_args = parser.parse_args(
            [
                "walk-forward",
                "--config",
                "configs/test.yaml",
                "--output-dir",
                "outputs/walk_forward",
                "--run-id-prefix",
                "same_run",
                "--window-limit",
                "9",
                "--resume",
            ]
        )
        self.assertEqual(
            _walk_forward_lock_key(smoke_args, resolve_walk_forward_options(smoke_args)),
            _walk_forward_lock_key(resume_args, resolve_walk_forward_options(resume_args)),
        )

    def test_tasklist_pid_match_is_exact(self) -> None:
        output = (
            '"python.exe","1234","Console","1","10,000 K"\n'
            '"python.exe","5123","Console","1","10,000 K"\n'
        )
        self.assertFalse(_tasklist_contains_pid(output, 123))
        self.assertTrue(_tasklist_contains_pid(output + '"python.exe","123","Console","1","10,000 K"\n', 123))

    def test_portfolio_combine_cli(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["portfolio", "combine", "--config", "configs/portfolio.yaml", "--run-id", "p1"])
        self.assertEqual(args.command, "portfolio")
        self.assertEqual(args.portfolio_command, "combine")
        self.assertEqual(args.config, "configs/portfolio.yaml")
        self.assertEqual(args.run_id, "p1")

    def test_portfolio_walk_forward_cli(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "portfolio",
                "walk-forward",
                "--config",
                "configs/portfolio.yaml",
                "--train-months",
                "12",
                "--test-months",
                "3",
                "--grid",
                "compact",
                "--score-mode",
                "capital_efficiency",
                "--score-mode-min-edge",
                "0.04",
            ]
        )
        self.assertEqual(args.command, "portfolio")
        self.assertEqual(args.portfolio_command, "walk-forward")
        self.assertEqual(args.train_months, 12)
        self.assertEqual(args.test_months, 3)
        self.assertEqual(args.grid, "compact")
        self.assertEqual(args.score_mode, "capital_efficiency")
        self.assertAlmostEqual(args.score_mode_min_edge, 0.04)

    def test_portfolio_combine_uses_previous_day_regime(self) -> None:
        dates = pd.date_range("2020-01-01", periods=5, freq="D")
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "configs").mkdir()
            (root / "data").mkdir()
            core_path = root / "core.csv"
            satellite_path = root / "satellite.csv"
            benchmark_path = root / "data" / "benchmark.csv"
            pd.DataFrame({"date": dates, "equity": [100, 100, 100, 100, 100]}).to_csv(core_path, index=False)
            pd.DataFrame({"date": dates, "equity": [100, 100, 110, 121, 133.1]}).to_csv(satellite_path, index=False)
            pd.DataFrame({"date": dates, "close": [10, 11, 12, 13, 14]}).to_csv(benchmark_path, index=False)
            config_path = root / "configs" / "portfolio.yaml"
            config_path.write_text(
                """
project:
  name: test_portfolio
  initial_cash: 1000
  output_dir: outputs
curves:
  core:
    path: core.csv
    equity_column: equity
  satellite:
    path: satellite.csv
    equity_column: equity
regime:
  benchmark_path: data/benchmark.csv
  ma_window: 2
  drop_window: 1
  risk_on_drop_threshold: -0.50
  crash_drop_threshold: -0.80
  default_regime: risk_off
weights:
  risk_on:
    core: 0.5
    satellite: 0.5
  risk_off:
    core: 1.0
    satellite: 0.0
  crash:
    core: 1.0
    satellite: 0.0
""",
                encoding="utf-8",
            )
            result = run_portfolio_combine(load_portfolio_config(config_path), run_id="test", write_outputs=False)
        self.assertEqual(float(result.equity.loc[1, "satellite_weight"]), 0.0)
        self.assertEqual(float(result.equity.loc[2, "satellite_weight"]), 0.5)
        self.assertAlmostEqual(float(result.equity.loc[2, "portfolio_return"]), 0.05)

    def test_portfolio_combine_metrics_compare_to_core(self) -> None:
        dates = pd.date_range("2020-01-01", periods=4, freq="D")
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "configs").mkdir()
            (root / "data").mkdir()
            pd.DataFrame({"date": dates, "equity": [100, 101, 102, 103]}).to_csv(root / "core.csv", index=False)
            pd.DataFrame({"date": dates, "equity": [100, 100, 120, 120]}).to_csv(root / "satellite.csv", index=False)
            pd.DataFrame({"date": dates, "close": [10, 11, 12, 13]}).to_csv(root / "data" / "benchmark.csv", index=False)
            config_path = root / "configs" / "portfolio.yaml"
            config_path.write_text(
                """
project:
  name: test_portfolio
  initial_cash: 1000
  output_dir: outputs
curves:
  core:
    path: core.csv
    equity_column: equity
  satellite:
    path: satellite.csv
    equity_column: equity
regime:
  benchmark_path: data/benchmark.csv
  ma_window: 2
  drop_window: 1
  risk_on_drop_threshold: -0.50
  crash_drop_threshold: -0.80
weights:
  risk_on:
    core: 0.75
    satellite: 0.25
  risk_off:
    core: 1.0
    satellite: 0.0
  crash:
    core: 1.0
    satellite: 0.0
""",
                encoding="utf-8",
            )
            result = run_portfolio_combine(load_portfolio_config(config_path), run_id="test", write_outputs=False)
        self.assertIn("excess_return_vs_core", result.metrics)
        self.assertIn("average_satellite_weight", result.metrics)
        self.assertGreater(result.metrics["observation_days"], 0)

    def test_portfolio_satellite_filter_uses_previous_day_health(self) -> None:
        dates = pd.date_range("2020-01-01", periods=6, freq="D")
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "configs").mkdir()
            (root / "data").mkdir()
            pd.DataFrame({"date": dates, "equity": [100, 100, 100, 100, 100, 100]}).to_csv(root / "core.csv", index=False)
            pd.DataFrame({"date": dates, "equity": [100, 110, 120, 100, 90, 130]}).to_csv(
                root / "satellite.csv",
                index=False,
            )
            pd.DataFrame({"date": dates, "close": [10, 11, 12, 13, 14, 15]}).to_csv(
                root / "data" / "benchmark.csv",
                index=False,
            )
            config_path = root / "configs" / "portfolio.yaml"
            config_path.write_text(
                """
project:
  name: test_portfolio_filter
  initial_cash: 1000
  output_dir: outputs
curves:
  core:
    path: core.csv
    equity_column: equity
  satellite:
    path: satellite.csv
    equity_column: equity
regime:
  benchmark_path: data/benchmark.csv
  ma_window: 2
  drop_window: 1
  risk_on_drop_threshold: -0.50
  crash_drop_threshold: -0.80
  default_regime: risk_off
weights:
  risk_on:
    core: 0.7
    satellite: 0.3
  risk_off:
    core: 1.0
    satellite: 0.0
  crash:
    core: 1.0
    satellite: 0.0
satellite_filter:
  enabled: true
  ma_window: 2
  momentum_window: 1
  min_momentum: -0.50
  max_drawdown: 0.15
  reduced_drawdown: 0.10
  reduced_scale: 0.5
  default_scale: 0.0
  require_above_ma: true
  reallocate_to: core
""",
                encoding="utf-8",
            )
            result = run_portfolio_combine(load_portfolio_config(config_path), run_id="test", write_outputs=False)
        self.assertEqual(float(result.equity.loc[2, "satellite_weight"]), 0.3)
        self.assertEqual(float(result.equity.loc[4, "satellite_weight"]), 0.0)
        self.assertEqual(float(result.equity.loc[4, "core_weight"]), 1.0)
        self.assertEqual(result.equity.loc[4, "satellite_effective_reason"], "satellite_drawdown_stop")
        self.assertGreater(result.metrics["satellite_filter_off_day_ratio"], 0.0)

    def test_portfolio_candidate_grid_includes_guarded_candidates(self) -> None:
        candidates = build_portfolio_candidate_grid("guarded")
        self.assertTrue(any(candidate.name == "core_only" for candidate in candidates))
        self.assertTrue(any("health_mom10_dd20" in candidate.name for candidate in candidates))
        self.assertTrue(any(candidate.satellite_filter.enabled for candidate in candidates))

    def test_portfolio_activation_grid_trains_regime_rules(self) -> None:
        candidates = build_portfolio_candidate_grid("activation")
        self.assertTrue(any(candidate.name == "core_only" for candidate in candidates))
        self.assertTrue(any("ma60_drop03" in candidate.name for candidate in candidates))
        self.assertTrue(any("ma200_drop05" in candidate.name for candidate in candidates))
        self.assertTrue(any(candidate.regime_overrides.get("ma_window") == 60 for candidate in candidates))
        self.assertTrue(any(candidate.regime_overrides.get("risk_on_drop_threshold") == -0.03 for candidate in candidates))

    def test_portfolio_activation_dd50_grid_has_high_satellite_weights(self) -> None:
        candidates = build_portfolio_candidate_grid("activation_dd50")
        self.assertTrue(any(candidate.name == "core_only" for candidate in candidates))
        self.assertTrue(any(candidate.name.startswith("sat30_") for candidate in candidates))
        self.assertTrue(any(candidate.name.startswith("sat35_") for candidate in candidates))
        self.assertTrue(any(candidate.name.startswith("sat50_") for candidate in candidates))
        self.assertTrue(any(candidate.regime_overrides.get("risk_on_drop_threshold") == -0.03 for candidate in candidates))
        self.assertTrue(any(candidate.satellite_filter.enabled is False for candidate in candidates))

    def test_portfolio_source_candidate_grid_crosses_sources_with_allocations(self) -> None:
        sources = {
            "quality": CurveConfig("satellite", Path("quality.csv"), "equity"),
            "reversal": CurveConfig("satellite", Path("reversal.csv"), "equity"),
            "blend_q70_r30": CurveConfig("satellite", Path("blend.csv"), "equity"),
        }
        candidates = build_portfolio_source_candidate_grid(sources, "activation")
        names = [candidate.name for candidate in candidates]
        self.assertEqual(names.count("core_only"), 1)
        self.assertTrue(any(name.startswith("quality__sat25_ma60_drop03") for name in names))
        self.assertTrue(any(name.startswith("reversal__sat25_ma60_drop03") for name in names))
        self.assertTrue(any(name.startswith("blend_q70_r30__sat25_ma60_drop03") for name in names))
        self.assertTrue(all(candidate.source_name in sources or candidate.source_name == "core_only" for candidate in candidates))
        capped = build_portfolio_source_candidate_grid(sources, "activation", max_satellite_weight=0.20)
        capped_names = [candidate.name for candidate in capped]
        self.assertEqual(capped_names.count("core_only"), 1)
        self.assertTrue(any(name.startswith("quality__sat20_ma60_drop03") for name in capped_names))
        self.assertFalse(any("__sat25_" in name for name in capped_names))
        source_capped = build_portfolio_source_candidate_grid(
            sources,
            "activation",
            source_max_satellite_weight={"reversal": 0.20},
        )
        source_capped_names = [candidate.name for candidate in source_capped]
        self.assertTrue(any(name.startswith("quality__sat25_ma60_drop03") for name in source_capped_names))
        self.assertFalse(any(name.startswith("reversal__sat25_") for name in source_capped_names))
        self.assertTrue(any(name.startswith("reversal__sat20_ma60_drop03") for name in source_capped_names))
        self.assertTrue(any(name.startswith("blend_q70_r30__sat25_ma60_drop03") for name in source_capped_names))

    def test_portfolio_source_candidate_grid_activation_dd50_weights(self) -> None:
        sources = {
            "quality": CurveConfig("satellite", Path("quality.csv"), "equity"),
        }
        candidates = build_portfolio_source_candidate_grid(sources, "activation_dd50")
        names = [candidate.name for candidate in candidates]
        self.assertTrue(any(name.startswith("quality__sat30_") for name in names))
        self.assertTrue(any(name.startswith("quality__sat35_") for name in names))
        self.assertTrue(any(name.startswith("quality__sat50_") for name in names))
        capped = build_portfolio_source_candidate_grid(sources, "activation_dd50", max_satellite_weight=0.35)
        capped_names = [candidate.name for candidate in capped]
        self.assertTrue(any(name.startswith("quality__sat35_") for name in capped_names))
        self.assertFalse(any(name.startswith("quality__sat50_") for name in capped_names))
        self.assertIn("core_only", names)

    def test_portfolio_walk_forward_activation_cli(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["portfolio", "walk-forward", "--grid", "activation"])
        self.assertEqual(args.grid, "activation")

    def test_portfolio_walk_forward_activation_dd50_cli(self) -> None:
        parser = build_parser()
        args = parser.parse_args(["portfolio", "walk-forward", "--grid", "activation_dd50"])
        self.assertEqual(args.grid, "activation_dd50")

    def test_portfolio_source_selection_cli(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "portfolio",
                "source-selection",
                "--grid",
                "activation",
                "--run-id-prefix",
                "source_test",
                "--default-source",
                "quality",
                "--source-switch-margin",
                "0.12",
                "--source-validation-months",
                "1",
                "--source-stability-penalty",
                "0.08",
                "--source-switch-margin-by-source",
                "pos10_risk_lean=0.02",
                "--source-switch-margin-by-source",
                "highgain_pos8=0.02",
                "--score-mode",
                "capital_efficiency",
                "--score-mode-min-edge",
                "0.04",
                "--max-satellite-weight",
                "0.20",
                "--source-max-satellite-weight",
                "reversal=0.20",
                "--source-max-satellite-weight",
                "quality=0.25",
                "--source-group",
                "highyield=highgain_pos8_dd50,highgain_dd40_protected",
            ]
        )
        self.assertEqual(args.portfolio_command, "source-selection")
        self.assertEqual(args.grid, "activation")
        self.assertEqual(args.run_id_prefix, "source_test")
        self.assertEqual(args.default_source, "quality")
        self.assertAlmostEqual(args.source_switch_margin, 0.12)
        self.assertEqual(args.source_validation_months, 1)
        self.assertAlmostEqual(args.source_stability_penalty, 0.08)
        self.assertEqual(args.source_switch_margin_by_source, ["pos10_risk_lean=0.02", "highgain_pos8=0.02"])
        self.assertEqual(args.score_mode, "capital_efficiency")
        self.assertAlmostEqual(args.score_mode_min_edge, 0.04)
        self.assertAlmostEqual(args.max_satellite_weight, 0.20)
        self.assertEqual(args.source_max_satellite_weight, ["reversal=0.20", "quality=0.25"])
        self.assertEqual(args.source_group, ["highyield=highgain_pos8_dd50,highgain_dd40_protected"])

    def test_source_group_preselection_keeps_best_source_per_group(self) -> None:
        allocation = PortfolioCandidate(
            name="sat30",
            weights={
                "risk_on": {"core": 0.7, "satellite": 0.3, "cash": 0.0},
                "risk_off": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                "crash": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
            },
            satellite_filter=SatelliteFilterConfig(),
        )
        highgain_pos8 = PortfolioSourceCandidate("highgain_pos8_dd50__sat30", "highgain_pos8_dd50", None, allocation)
        highgain_dd40 = PortfolioSourceCandidate("highgain_dd40_protected__sat30", "highgain_dd40_protected", None, allocation)
        quality = PortfolioSourceCandidate("quality__sat30", "quality", None, allocation)
        result = PortfolioResult(
            run_id="train",
            run_dir=Path("."),
            equity=pd.DataFrame(),
            allocation_events=pd.DataFrame(),
            monthly_returns=pd.DataFrame(),
            metrics={},
        )

        filtered, status = _apply_source_group_preselection(
            [(0.80, highgain_pos8, result), (0.92, highgain_dd40, result), (0.70, quality, result)],
            {"highyield": ["highgain_pos8_dd50", "highgain_dd40_protected"]},
        )

        self.assertEqual([candidate.source_name for _, candidate, _ in filtered], ["highgain_dd40_protected", "quality"])
        self.assertEqual(status["highgain_pos8_dd50"]["source_group_winner"], "highgain_dd40_protected")
        self.assertFalse(status["highgain_pos8_dd50"]["source_group_selected"])
        self.assertTrue(status["highgain_dd40_protected"]["source_group_selected"])

    def test_source_group_preselection_applies_source_specific_margin_before_filtering(self) -> None:
        allocation = PortfolioCandidate(
            name="sat30",
            weights={
                "risk_on": {"core": 0.7, "satellite": 0.3, "cash": 0.0},
                "risk_off": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                "crash": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
            },
            satellite_filter=SatelliteFilterConfig(),
        )
        highgain_pos8 = PortfolioSourceCandidate("highgain_pos8_dd50__sat30", "highgain_pos8_dd50", None, allocation)
        highyield_qfull = PortfolioSourceCandidate("highyield_qfull_dd30__sat30", "highyield_qfull_dd30", None, allocation)
        quality = PortfolioSourceCandidate("quality__sat30", "quality", None, allocation)
        result = PortfolioResult(
            run_id="train",
            run_dir=Path("."),
            equity=pd.DataFrame(),
            allocation_events=pd.DataFrame(),
            monthly_returns=pd.DataFrame(),
            metrics={},
        )

        filtered, status = _apply_source_group_preselection(
            [(1.433, highgain_pos8, result), (1.399, highyield_qfull, result), (1.100, quality, result)],
            {"highyield": ["highgain_pos8_dd50", "highyield_qfull_dd30"]},
            source_switch_margin_by_source={"highgain_pos8_dd50": 0.10},
        )

        selected_score, selected_candidate, _, raw_best = _select_source_candidate(
            filtered,
            source_switch_margin_by_source={"highgain_pos8_dd50": 0.10},
        )

        self.assertEqual(raw_best[1].source_name, "highyield_qfull_dd30")
        self.assertEqual(selected_candidate.source_name, "highyield_qfull_dd30")
        self.assertEqual(selected_score, 1.399)
        self.assertEqual(status["highgain_pos8_dd50"]["source_group_winner"], "highyield_qfull_dd30")
        self.assertFalse(status["highgain_pos8_dd50"]["source_group_selected"])
        self.assertTrue(status["highyield_qfull_dd30"]["source_group_selected"])

    def test_source_selection_stability_penalty_keeps_previous_source_when_edge_is_small(self) -> None:
        allocation = PortfolioCandidate(
            name="sat20",
            weights={
                "risk_on": {"core": 0.8, "satellite": 0.2, "cash": 0.0},
                "risk_off": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                "crash": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
            },
            satellite_filter=SatelliteFilterConfig(),
        )
        quality = PortfolioSourceCandidate("quality__sat20", "quality", None, allocation)
        reversal = PortfolioSourceCandidate("reversal__sat20", "reversal", None, allocation)
        result = PortfolioResult(
            run_id="train",
            run_dir=Path("."),
            equity=pd.DataFrame(),
            allocation_events=pd.DataFrame(),
            monthly_returns=pd.DataFrame(),
            metrics={},
        )

        selected_score, selected_candidate, _, raw_best = _select_source_candidate(
            [(1.05, reversal, result), (1.00, quality, result)],
            previous_source_name="quality",
            source_stability_penalty=0.08,
        )

        self.assertEqual(raw_best[1].source_name, "reversal")
        self.assertEqual(selected_candidate.source_name, "quality")
        self.assertEqual(selected_score, 1.00)

    def test_source_selection_source_specific_margin_blocks_fragile_source_only(self) -> None:
        allocation = PortfolioCandidate(
            name="sat10",
            weights={
                "risk_on": {"core": 0.9, "satellite": 0.1, "cash": 0.0},
                "risk_off": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                "crash": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
            },
            satellite_filter=SatelliteFilterConfig(),
        )
        pos10 = PortfolioSourceCandidate("pos10__sat10", "pos10_risk_lean", None, allocation)
        blend = PortfolioSourceCandidate("blend__sat25", "blend_q30_r70", None, allocation)
        quality = PortfolioSourceCandidate("quality__sat25", "quality", None, allocation)
        result = PortfolioResult(
            run_id="train",
            run_dir=Path("."),
            equity=pd.DataFrame(),
            allocation_events=pd.DataFrame(),
            monthly_returns=pd.DataFrame(),
            metrics={},
        )

        selected_score, selected_candidate, _, raw_best = _select_source_candidate(
            [(1.359, pos10, result), (1.347, blend, result), (1.338, quality, result)],
            source_switch_margin_by_source={"pos10_risk_lean": 0.02},
        )

        self.assertEqual(raw_best[1].source_name, "pos10_risk_lean")
        self.assertEqual(selected_candidate.source_name, "blend_q30_r70")
        self.assertEqual(selected_score, 1.347)

    def test_portfolio_source_selection_walk_forward_records_selected_source(self) -> None:
        dates = pd.date_range("2020-01-01", periods=100, freq="D")
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "configs").mkdir()
            (root / "data").mkdir()
            pd.DataFrame({"date": dates, "equity": np.full(len(dates), 100.0)}).to_csv(root / "core.csv", index=False)
            pd.DataFrame({"date": dates, "equity": np.linspace(100, 160, len(dates))}).to_csv(root / "quality.csv", index=False)
            pd.DataFrame({"date": dates, "equity": np.linspace(100, 95, len(dates))}).to_csv(root / "reversal.csv", index=False)
            pd.DataFrame({"date": dates, "close": np.linspace(10, 20, len(dates))}).to_csv(root / "data" / "benchmark.csv", index=False)
            config_path = root / "configs" / "portfolio.yaml"
            config_path.write_text(
                """
project:
  name: test_source_selection
  initial_cash: 1000
  output_dir: outputs
curves:
  core:
    path: core.csv
    equity_column: equity
  satellite:
    path: quality.csv
    equity_column: equity
regime:
  benchmark_path: data/benchmark.csv
  ma_window: 2
  drop_window: 1
  risk_on_drop_threshold: -0.50
  crash_drop_threshold: -0.80
  default_regime: risk_off
weights:
  risk_on:
    core: 0.8
    satellite: 0.2
  risk_off:
    core: 1.0
    satellite: 0.0
  crash:
    core: 1.0
    satellite: 0.0
satellite_sources:
  quality:
    path: quality.csv
    equity_column: equity
  reversal:
    path: reversal.csv
    equity_column: equity
""",
                encoding="utf-8",
            )
            config, sources = load_portfolio_source_selection_config(config_path)
            wf_dir, summary = run_portfolio_source_selection_walk_forward(
                config,
                sources,
                train_months=1,
                test_months=1,
                step_months=1,
                grid="compact",
                score_mode="capital_efficiency",
                score_mode_min_edge=0.03,
                output_dir=root / "source_selection",
                run_id_prefix="test_source_selection",
            )
            self.assertTrue((wf_dir / "portfolio_candidate_results.csv").exists())
            self.assertIn("selected_source", summary.columns)
            self.assertIn("score_mode", summary.columns)
            self.assertIn("score_mode_min_edge", summary.columns)
            self.assertIn("score_mode_gate_applied", summary.columns)
            self.assertEqual(summary.iloc[0]["score_mode"], "capital_efficiency")
            self.assertAlmostEqual(float(summary.iloc[0]["score_mode_min_edge"]), 0.03)
            self.assertEqual(summary.iloc[0]["selected_source"], "quality")

    def test_portfolio_source_selection_can_score_on_internal_validation_period(self) -> None:
        dates = pd.date_range("2020-01-01", periods=180, freq="D")
        reversal_equity = []
        for index in range(len(dates)):
            if index < 60:
                reversal_equity.append(100.0 + 2.5 * index)
            else:
                reversal_equity.append(max(80.0, 247.5 - 3.0 * (index - 60)))
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "configs").mkdir()
            (root / "data").mkdir()
            pd.DataFrame({"date": dates, "equity": np.full(len(dates), 100.0)}).to_csv(root / "core.csv", index=False)
            pd.DataFrame({"date": dates, "equity": np.linspace(100, 180, len(dates))}).to_csv(root / "quality.csv", index=False)
            pd.DataFrame({"date": dates, "equity": reversal_equity}).to_csv(root / "reversal.csv", index=False)
            pd.DataFrame({"date": dates, "close": np.linspace(10, 20, len(dates))}).to_csv(root / "data" / "benchmark.csv", index=False)
            config_path = root / "configs" / "portfolio.yaml"
            config_path.write_text(
                """
project:
  name: test_source_validation
  initial_cash: 1000
  output_dir: outputs
curves:
  core:
    path: core.csv
    equity_column: equity
  satellite:
    path: quality.csv
    equity_column: equity
regime:
  benchmark_path: data/benchmark.csv
  ma_window: 2
  drop_window: 1
  risk_on_drop_threshold: -0.50
  crash_drop_threshold: -0.80
  default_regime: risk_off
weights:
  risk_on:
    core: 0.8
    satellite: 0.2
  risk_off:
    core: 1.0
    satellite: 0.0
  crash:
    core: 1.0
    satellite: 0.0
satellite_sources:
  quality:
    path: quality.csv
    equity_column: equity
  reversal:
    path: reversal.csv
    equity_column: equity
""",
                encoding="utf-8",
            )
            config, sources = load_portfolio_source_selection_config(config_path)
            _, summary = run_portfolio_source_selection_walk_forward(
                config,
                sources,
                train_months=3,
                test_months=1,
                step_months=1,
                grid="compact",
                source_validation_months=1,
                output_dir=root / "source_selection",
                run_id_prefix="test_source_validation",
            )
            self.assertIn("validation_score", summary.columns)
            self.assertEqual(summary.iloc[0]["source_validation_months"], 1)
            self.assertEqual(summary.iloc[0]["selected_source"], "quality")

    def test_generate_portfolio_windows(self) -> None:
        windows = generate_portfolio_windows(
            pd.Timestamp("2022-01-04"),
            pd.Timestamp("2024-06-30"),
            train_months=12,
            test_months=6,
            step_months=6,
        )
        self.assertEqual(windows[0].train_start, "20220104")
        self.assertEqual(windows[0].test_start, "20230104")
        self.assertEqual(windows[0].test_end, "20230703")
        self.assertGreaterEqual(len(windows), 2)

    def test_score_portfolio_metrics_penalizes_drawdown(self) -> None:
        base = {
            "total_return": 0.20,
            "cagr": 0.08,
            "sharpe": 0.7,
            "max_drawdown": -0.12,
            "average_satellite_weight": 0.04,
        }
        high_drawdown = {**base, "max_drawdown": -0.30}
        self.assertGreater(score_portfolio_metrics(base), score_portfolio_metrics(high_drawdown))

    def test_score_portfolio_metrics_capital_efficiency_rewards_return_per_satellite_weight(self) -> None:
        efficient = {
            "total_return": 0.12,
            "excess_return_vs_core": 0.06,
            "cagr": 0.06,
            "sharpe": 0.7,
            "max_drawdown": -0.10,
            "average_satellite_weight": 0.05,
        }
        capital_heavy = {
            **efficient,
            "total_return": 0.13,
            "excess_return_vs_core": 0.065,
            "average_satellite_weight": 0.25,
        }
        self.assertGreater(
            score_portfolio_metrics(efficient, score_mode="capital_efficiency"),
            score_portfolio_metrics(capital_heavy, score_mode="capital_efficiency"),
        )

    def test_score_mode_min_edge_gate_falls_back_to_balanced_candidate_when_edge_is_small(self) -> None:
        balanced_candidate = PortfolioCandidate(
            name="balanced_sat",
            weights={
                "risk_on": {"core": 0.8, "satellite": 0.2, "cash": 0.0},
                "risk_off": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
                "crash": {"core": 1.0, "satellite": 0.0, "cash": 0.0},
            },
            satellite_filter=SatelliteFilterConfig(),
        )
        capital_candidate = PortfolioCandidate(
            name="capital_efficient_sat",
            weights=balanced_candidate.weights,
            satellite_filter=SatelliteFilterConfig(),
        )
        result = PortfolioResult(
            run_id="train",
            run_dir=Path("."),
            equity=pd.DataFrame(),
            allocation_events=pd.DataFrame(),
            monthly_returns=pd.DataFrame(),
            metrics={},
        )

        selected = (1.00, capital_candidate, result)
        balanced_selected = (0.90, balanced_candidate, result)
        gated_score, gated_candidate, _, gate_payload = portfolio_module._apply_score_mode_min_edge_gate(
            selected,
            balanced_selected,
            active_scores_by_candidate={
                "capital_efficient_sat": 1.00,
                "balanced_sat": 0.97,
            },
            score_mode="capital_efficiency",
            score_mode_min_edge=0.05,
        )

        self.assertEqual(gated_candidate.name, "balanced_sat")
        self.assertAlmostEqual(gated_score, 0.97)
        self.assertTrue(gate_payload["score_mode_gate_applied"])
        self.assertAlmostEqual(gate_payload["score_mode_edge_vs_baseline"], 0.03)

    def test_opportunity_grid_includes_growth_risk_candidates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            candidates = build_parameter_grid(make_config(Path(tmp)), grid="opportunity")
        self.assertEqual(len(candidates), 12)
        growth = [candidate for candidate in candidates if "risk_bear_growth" in candidate.name]
        self.assertTrue(growth)
        self.assertEqual(growth[0].risk_overrides["benchmark_off_exposure"], 0.65)

    def test_opportunity_v2_grid_includes_high_gain_dd40_candidates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            candidates = build_parameter_grid(make_config(Path(tmp)), grid="opportunity_v2")

        self.assertEqual(len(candidates), 32)
        self.assertTrue(any(candidate.name == "current_reb10_risk_loose" for candidate in candidates))
        self.assertTrue(any(candidate.name.startswith("breakout_momentum_reb8") for candidate in candidates))
        self.assertTrue(any(candidate.name.startswith("trend_accel_reb10") for candidate in candidates))
        self.assertTrue(any(candidate.max_positions == 12 for candidate in candidates))
        dd40 = [candidate for candidate in candidates if "risk_growth_dd40" in candidate.name]
        self.assertTrue(dd40)
        self.assertEqual(dd40[0].risk_overrides["protection_drawdown"], 0.40)
        self.assertEqual(dd40[0].risk_overrides["drawdown_levels"][-1], (0.40, 0.0))

    def test_opportunity_v2_gate_blocks_new_candidate_with_weak_validation(self) -> None:
        candidate = ParameterCandidate("strong_gain_quality_reb15_risk_growth_dd40", {"momentum": 1.0}, 15, 12, {})

        reason = _candidate_selection_gate_reason(
            candidate,
            {"total_return": -0.01, "sharpe": 0.2},
            grid="opportunity_v2",
        )

        self.assertEqual(reason, "non_current_negative_validation_return")

    def test_opportunity_v2_gate_allows_current_candidate_with_weak_validation(self) -> None:
        candidate = ParameterCandidate("current_reb10_risk_loose", {"momentum": 1.0}, 10, 10, {})

        reason = _candidate_selection_gate_reason(
            candidate,
            {"total_return": -0.10, "sharpe": -0.5},
            grid="opportunity_v2",
        )

        self.assertIsNone(reason)

    def test_stable_grid_uses_broad_universe_candidates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            candidates = build_parameter_grid(make_config(Path(tmp)), grid="stable")
        self.assertEqual(len(candidates), 12)
        self.assertTrue(any(candidate.max_positions == 15 for candidate in candidates))
        self.assertTrue(all(candidate.risk_overrides for candidate in candidates))
        self.assertTrue(any("risk_stable_strict" in candidate.name for candidate in candidates))

    def test_stable_v2_grid_uses_only_strict_risk_candidates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            candidates = build_parameter_grid(make_config(Path(tmp)), grid="stable_v2")
        self.assertEqual(len(candidates), 8)
        self.assertTrue(all(candidate.risk_overrides for candidate in candidates))
        self.assertFalse(any("risk_stable_guard" in candidate.name for candidate in candidates))
        self.assertTrue(any("risk_stable_ultra" in candidate.name for candidate in candidates))

    def test_reversal_grid_uses_factor_lab_evidence_candidates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            candidates = build_parameter_grid(make_config(Path(tmp)), grid="reversal")
        self.assertEqual(len(candidates), 8)
        focus = [candidate for candidate in candidates if candidate.name.startswith("reversal_focus_reb20")]
        self.assertTrue(focus)
        self.assertEqual(focus[0].factor_weights["reversal"], 0.55)
        self.assertEqual(focus[0].factor_weights["liquidity"], 0.10)
        self.assertTrue(any("risk_reversal_strict" in candidate.name for candidate in candidates))

    def test_risk_grid_preserves_config_extra_factor_weights(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            config = replace(
                config,
                strategy=replace(
                    config.strategy,
                    factor_weights={
                        "momentum": 0.19,
                        "trend": 0.19,
                        "reversal": 0.24,
                        "volatility": 0.24,
                        "liquidity": 0.09,
                        "chip_reversal": 0.05,
                    },
                ),
            )
            candidates = build_parameter_grid(config, grid="risk")

        self.assertTrue(candidates)
        self.assertTrue(all(candidate.factor_weights.get("chip_reversal") == 0.05 for candidate in candidates))
        self.assertTrue(all(abs(sum(candidate.factor_weights.values()) - 1.0) < 1e-9 for candidate in candidates))
        focus = [candidate for candidate in candidates if candidate.name.startswith("reversal_focus_reb20")]
        self.assertTrue(focus)
        self.assertAlmostEqual(focus[0].factor_weights["reversal"], 0.55 * 0.95)
        self.assertAlmostEqual(focus[0].factor_weights["liquidity"], 0.10 * 0.95)

    def test_satellite_grid_uses_growth_candidates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            candidates = build_parameter_grid(make_config(Path(tmp)), grid="satellite")
        self.assertEqual(len(candidates), 12)
        self.assertTrue(any("growth_momentum" in candidate.name for candidate in candidates))
        self.assertTrue(any("risk_satellite_growth" in candidate.name for candidate in candidates))
        self.assertTrue(any(candidate.max_positions == 15 for candidate in candidates))

    def test_satellite_v2_grid_includes_risk_on_only_candidates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            candidates = build_parameter_grid(make_config(Path(tmp)), grid="satellite_v2")
        self.assertEqual(len(candidates), 12)
        self.assertTrue(any("risk_satellite_guard" in candidate.name for candidate in candidates))
        self.assertTrue(any("risk_satellite_on_only" in candidate.name for candidate in candidates))
        self.assertFalse(any("risk_satellite_growth" in candidate.name for candidate in candidates))

    def test_training_score_penalizes_excess_drawdown(self) -> None:
        base = {"total_return": 0.20, "cagr": 0.08, "sharpe": 0.5, "max_drawdown": -0.20}
        high_drawdown = {**base, "max_drawdown": -0.50}
        self.assertGreater(score_training_metrics(base), score_training_metrics(high_drawdown))

    def test_capital_training_score_rewards_win_rate_and_utilization(self) -> None:
        weak = {
            "total_return": 0.18,
            "cagr": 0.06,
            "sharpe": 0.45,
            "max_drawdown": -0.20,
            "profit_factor": 1.1,
            "win_rate": 0.48,
            "average_risk_exposure": 0.35,
        }
        strong = {
            **weak,
            "win_rate": 0.67,
            "average_risk_exposure": 0.90,
        }
        self.assertGreater(
            score_training_metrics(strong, objective="capital"),
            score_training_metrics(weak, objective="capital"),
        )

    def test_capital_training_score_penalizes_excess_drawdown(self) -> None:
        base = {"total_return": 0.20, "cagr": 0.08, "sharpe": 0.55, "win_rate": 0.55, "average_risk_exposure": 0.8}
        high_drawdown = {**base, "max_drawdown": -0.55}
        self.assertGreater(
            score_training_metrics(base, objective="capital"),
            score_training_metrics(high_drawdown, objective="capital"),
        )

    def test_opportunity_quality_score_rewards_success_and_utilization(self) -> None:
        diagnostics = {"monthly_win_rate": 0.55}
        weak = {
            "total_return": 0.18,
            "cagr": 0.07,
            "sharpe": 0.55,
            "max_drawdown": -0.22,
            "profit_factor": 1.1,
            "win_rate": 0.48,
            "average_risk_exposure": 0.35,
        }
        strong = {
            **weak,
            "profit_factor": 1.5,
            "win_rate": 0.64,
            "average_risk_exposure": 0.85,
        }
        self.assertGreater(
            score_training_metrics(strong, objective="opportunity_quality", diagnostics=diagnostics),
            score_training_metrics(weak, objective="opportunity_quality", diagnostics=diagnostics),
        )

    def test_opportunity_quality_score_penalizes_excess_drawdown(self) -> None:
        diagnostics = {"monthly_win_rate": 0.60}
        base = {
            "total_return": 0.28,
            "cagr": 0.10,
            "sharpe": 0.75,
            "max_drawdown": -0.22,
            "profit_factor": 1.4,
            "win_rate": 0.58,
            "average_risk_exposure": 0.80,
        }
        high_drawdown = {**base, "max_drawdown": -0.55}
        self.assertGreater(
            score_training_metrics(base, objective="opportunity_quality", diagnostics=diagnostics),
            score_training_metrics(high_drawdown, objective="opportunity_quality", diagnostics=diagnostics),
        )

    def test_opportunity_growth_score_rewards_return_and_drawdown_limit(self) -> None:
        metrics = {
            "total_return": 0.28,
            "cagr": 0.10,
            "sharpe": 0.55,
            "max_drawdown": -0.20,
            "profit_factor": 1.6,
            "win_rate": 0.58,
            "average_risk_exposure": 0.75,
        }
        weak = {
            "total_return": 0.05,
            "cagr": 0.02,
            "sharpe": 0.15,
            "max_drawdown": -0.26,
            "profit_factor": 0.8,
            "win_rate": 0.42,
            "average_risk_exposure": 0.65,
        }
        diagnostics = {"monthly_win_rate": 0.40}
        self.assertGreater(
            score_training_metrics(metrics, objective="opportunity_growth", diagnostics=diagnostics),
            score_training_metrics(weak, objective="opportunity_growth", diagnostics=diagnostics),
        )

    def test_opportunity_q_full_score_rewards_high_return(self) -> None:
        metrics = {
            "total_return": 0.30,
            "cagr": 0.12,
            "sharpe": 0.55,
            "max_drawdown": -0.20,
            "profit_factor": 1.7,
            "win_rate": 0.58,
            "average_risk_exposure": 0.80,
        }
        weak = {
            "total_return": 0.06,
            "cagr": 0.03,
            "sharpe": 0.20,
            "max_drawdown": -0.35,
            "profit_factor": 0.9,
            "win_rate": 0.42,
            "average_risk_exposure": 0.55,
        }
        diagnostics = {
            "monthly_win_rate": 0.45,
            "annual_return_std": 0.05,
            "return_concentration": 0.35,
            "worst_month_return": -0.12,
        }
        self.assertGreater(
            score_training_metrics(metrics, objective="opportunity_q_full", diagnostics=diagnostics),
            score_training_metrics(weak, objective="opportunity_q_full", diagnostics=diagnostics),
        )

    def test_opportunity_q_full_score_penalizes_excess_drawdown(self) -> None:
        diagnostics = {"monthly_win_rate": 0.40, "annual_return_std": 0.06, "return_concentration": 0.40, "worst_month_return": -0.14}
        base = {
            "total_return": 0.22,
            "cagr": 0.08,
            "sharpe": 0.60,
            "max_drawdown": -0.24,
            "profit_factor": 1.5,
            "win_rate": 0.54,
            "average_risk_exposure": 0.80,
        }
        high_drawdown = {**base, "max_drawdown": -0.70}
        self.assertGreater(
            score_training_metrics(base, objective="opportunity_q_full", diagnostics=diagnostics),
            score_training_metrics(high_drawdown, objective="opportunity_q_full", diagnostics=diagnostics),
        )

    def test_opportunity_growth_score_penalizes_excess_drawdown(self) -> None:
        diagnostics = {"monthly_win_rate": 0.40}
        base = {
            "total_return": 0.22,
            "cagr": 0.09,
            "sharpe": 0.65,
            "max_drawdown": -0.20,
            "profit_factor": 1.2,
            "win_rate": 0.52,
            "average_risk_exposure": 0.80,
        }
        high_drawdown = {**base, "max_drawdown": -0.60}
        self.assertGreater(
            score_training_metrics(base, objective="opportunity_growth", diagnostics=diagnostics),
            score_training_metrics(high_drawdown, objective="opportunity_growth", diagnostics=diagnostics),
        )

    def test_robust_training_score_penalizes_losing_years(self) -> None:
        metrics = {
            "total_return": 0.20,
            "cagr": 0.08,
            "sharpe": 0.5,
            "max_drawdown": -0.20,
            "profit_factor": 1.2,
            "win_rate": 0.55,
        }
        stable = {"worst_year_return": 0.02, "negative_year_ratio": 0.0, "worst_year_drawdown": -0.15}
        unstable = {"worst_year_return": -0.30, "negative_year_ratio": 0.5, "worst_year_drawdown": -0.35}
        self.assertGreater(
            score_training_metrics(metrics, objective="robust", diagnostics=stable),
            score_training_metrics(metrics, objective="robust", diagnostics=unstable),
        )

    def test_stable_training_score_penalizes_concentrated_fit(self) -> None:
        metrics = {
            "total_return": 0.30,
            "cagr": 0.10,
            "sharpe": 0.8,
            "max_drawdown": -0.18,
            "profit_factor": 1.3,
            "win_rate": 0.56,
        }
        stable = {
            "worst_year_return": 0.01,
            "negative_year_ratio": 0.0,
            "worst_year_drawdown": -0.12,
            "annual_return_std": 0.08,
            "return_concentration": 0.35,
            "monthly_win_rate": 0.58,
            "worst_month_return": -0.06,
        }
        overfit_like = {
            **stable,
            "annual_return_std": 0.45,
            "return_concentration": 0.85,
            "monthly_win_rate": 0.42,
            "worst_month_return": -0.22,
        }
        self.assertGreater(
            score_training_metrics(metrics, objective="stable", diagnostics=stable),
            score_training_metrics(metrics, objective="stable", diagnostics=overfit_like),
        )

    def test_stable_training_score_penalizes_high_risk_exposure(self) -> None:
        metrics = {
            "total_return": 0.20,
            "cagr": 0.08,
            "sharpe": 0.6,
            "max_drawdown": -0.18,
            "profit_factor": 1.2,
            "win_rate": 0.55,
        }
        diagnostics = {
            "worst_year_return": 0.01,
            "negative_year_ratio": 0.0,
            "worst_year_drawdown": -0.12,
            "annual_return_std": 0.08,
            "return_concentration": 0.35,
            "monthly_win_rate": 0.58,
            "worst_month_return": -0.06,
        }
        low_exposure = {**metrics, "average_risk_exposure": 0.35}
        high_exposure = {**metrics, "average_risk_exposure": 0.80}
        self.assertGreater(
            score_training_metrics(low_exposure, objective="stable", diagnostics=diagnostics),
            score_training_metrics(high_exposure, objective="stable", diagnostics=diagnostics),
        )

    def test_satellite_training_score_rewards_return_but_penalizes_drawdown(self) -> None:
        diagnostics = {
            "worst_year_return": 0.02,
            "negative_year_ratio": 0.0,
            "worst_year_drawdown": -0.18,
            "annual_return_std": 0.10,
            "return_concentration": 0.40,
            "monthly_win_rate": 0.58,
            "worst_month_return": -0.08,
        }
        strong = {
            "total_return": 0.35,
            "cagr": 0.14,
            "sharpe": 0.9,
            "max_drawdown": -0.28,
            "profit_factor": 1.3,
            "win_rate": 0.56,
            "average_risk_exposure": 0.65,
        }
        weak = {**strong, "total_return": 0.08, "cagr": 0.03, "sharpe": 0.35}
        high_drawdown = {**strong, "max_drawdown": -0.55}
        self.assertGreater(
            score_training_metrics(strong, max_drawdown_limit=0.35, objective="satellite", diagnostics=diagnostics),
            score_training_metrics(weak, max_drawdown_limit=0.35, objective="satellite", diagnostics=diagnostics),
        )
        self.assertGreater(
            score_training_metrics(strong, max_drawdown_limit=0.35, objective="satellite", diagnostics=diagnostics),
            score_training_metrics(high_drawdown, max_drawdown_limit=0.35, objective="satellite", diagnostics=diagnostics),
        )

    def test_report_handles_empty_trade_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            run_dir = Path(tmp)
            metrics = {
                "initial_cash": 100000.0,
                "final_equity": 100000.0,
                "total_return": 0.0,
                "benchmark_return": 0.0,
                "excess_return": 0.0,
                "cagr": 0.0,
                "max_drawdown": 0.0,
                "sharpe": 0.0,
                "win_rate": 0.0,
                "payoff_ratio": 0.0,
                "profit_factor": 0.0,
            }
            (run_dir / "metrics.json").write_text(json.dumps(metrics), encoding="utf-8")
            (run_dir / "trades.csv").write_text("", encoding="utf-8")
            report_path = write_report(run_dir)
            self.assertTrue(report_path.exists())

    def test_diagnostics_market_model_recovers_alpha_beta(self) -> None:
        benchmark_returns = pd.Series([0.01, -0.02, 0.015, 0.005, -0.01, 0.02, -0.005, 0.012])
        strategy_returns = 0.001 + 2.0 * benchmark_returns
        returns = pd.DataFrame(
            {
                "strategy_return": strategy_returns,
                "benchmark_return": benchmark_returns,
                "active_return": strategy_returns - benchmark_returns,
            }
        )
        summary, covariance, correlation, pca_components, metrics = build_diagnostics(returns, hac_lags=1)
        self.assertFalse(summary.empty)
        self.assertEqual(list(covariance.columns), ["strategy_return", "benchmark_return"])
        self.assertAlmostEqual(float(correlation.loc["strategy_return", "benchmark_return"]), 1.0)
        self.assertFalse(pca_components.empty)
        self.assertAlmostEqual(float(metrics["ols_beta"]), 2.0, places=6)
        self.assertAlmostEqual(float(metrics["ols_alpha_daily"]), 0.001, places=6)
        self.assertGreater(float(metrics["ols_r_squared"]), 0.999)

    def test_run_diagnostics_writes_report_files(self) -> None:
        dates = pd.date_range("2020-01-01", periods=6, freq="D")
        strategy_equity = [100000, 101000, 100500, 102000, 103000, 102500]
        benchmark_equity = [100000, 100500, 100000, 100800, 101200, 101000]
        with tempfile.TemporaryDirectory() as tmp:
            run_dir = Path(tmp)
            pd.DataFrame({"date": dates, "equity": strategy_equity}).to_csv(
                run_dir / "equity_curve.csv",
                index=False,
            )
            pd.DataFrame({"date": dates, "benchmark_equity": benchmark_equity}).to_csv(
                run_dir / "benchmark.csv",
                index=False,
            )
            result = run_diagnostics(run_dir, hac_lags=1)
            self.assertTrue((result.output_dir / "diagnostic_metrics.json").exists())
            self.assertTrue((result.output_dir / "covariance_matrix.csv").exists())
            self.assertTrue(result.report_path.exists())
            self.assertIn("Econometric Market Model", result.report_path.read_text(encoding="utf-8"))

    def test_run_batch_diagnostics_writes_comparison(self) -> None:
        dates = pd.date_range("2020-01-01", periods=6, freq="D")
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            run_dirs = []
            for idx, multiplier in enumerate([1.0, 1.5], start=1):
                run_dir = root / f"run{idx}"
                run_dir.mkdir()
                strategy_equity = [100000, 100000 + 1000 * multiplier, 100500, 102000, 103000, 102500]
                benchmark_equity = [100000, 100500, 100000, 100800, 101200, 101000]
                pd.DataFrame({"date": dates, "equity": strategy_equity}).to_csv(run_dir / "equity_curve.csv", index=False)
                pd.DataFrame({"date": dates, "benchmark_equity": benchmark_equity}).to_csv(run_dir / "benchmark.csv", index=False)
                run_dirs.append(run_dir)
            result = run_batch_diagnostics(run_dirs, output_dir=root / "batch", hac_lags=1)
            self.assertEqual(len(result.comparison), 2)
            self.assertTrue((result.output_dir / "diagnostic_compare.csv").exists())
            self.assertTrue(result.report_path.exists())
            self.assertIn("ols_beta", set(result.comparison.columns))

    def test_find_drawdown_periods_identifies_recovery(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            run_dir = Path(tmp)
            pd.DataFrame(
                {
                    "date": pd.date_range("2020-01-01", periods=7, freq="D"),
                    "equity": [100000, 110000, 100000, 80000, 90000, 111000, 108000],
                }
            ).to_csv(run_dir / "equity_curve.csv", index=False)
            equity = load_equity_drawdown(run_dir)
        periods = find_drawdown_periods(equity, min_depth=0.10)
        self.assertEqual(len(periods), 1)
        self.assertEqual(pd.Timestamp(periods.iloc[0]["peak_date"]), pd.Timestamp("2020-01-02"))
        self.assertEqual(pd.Timestamp(periods.iloc[0]["trough_date"]), pd.Timestamp("2020-01-04"))
        self.assertTrue(bool(periods.iloc[0]["recovered"]))
        self.assertAlmostEqual(float(periods.iloc[0]["drawdown"]), -0.2727272727)

    def test_run_drawdown_attribution_writes_trade_breakdown(self) -> None:
        dates = pd.date_range("2020-01-01", periods=7, freq="D")
        with tempfile.TemporaryDirectory() as tmp:
            run_dir = Path(tmp)
            pd.DataFrame(
                {
                    "date": dates,
                    "equity": [100000, 110000, 100000, 80000, 90000, 111000, 108000],
                    "cash": [100000, 20000, 25000, 30000, 50000, 111000, 108000],
                    "positions_value": [0, 90000, 75000, 50000, 40000, 0, 0],
                    "active_positions": [0, 2, 2, 1, 1, 0, 0],
                    "risk_exposure": [1, 1, 0.7, 0.7, 0.5, 1, 1],
                }
            ).to_csv(run_dir / "equity_curve.csv", index=False)
            pd.DataFrame(
                {
                    "date": dates,
                    "benchmark_equity": [100000, 101000, 97000, 91000, 93000, 104000, 102000],
                }
            ).to_csv(run_dir / "benchmark.csv", index=False)
            pd.DataFrame(
                {
                    "date": dates,
                    "risk_exposure": [1, 1, 0.7, 0.7, 0.5, 1, 1],
                    "risk_reason": ["on", "on", "drawdown_10", "drawdown_10", "drawdown_20", "on", "on"],
                }
            ).to_csv(run_dir / "risk_curve.csv", index=False)
            pd.DataFrame(
                {
                    "date": [dates[1], dates[2], dates[3]],
                    "code": ["000001", "000001", "000002"],
                    "name": ["A", "A", "B"],
                    "side": ["BUY", "SELL", "SELL"],
                    "price": [10.0, 9.0, 8.0],
                    "quantity": [1000, 1000, 1000],
                    "gross_amount": [10000, 9000, 8000],
                    "fee": [3.0, 3.0, 3.0],
                    "realized_pnl": [None, -1000.0, -500.0],
                    "cash_after": [90000, 99000, 107000],
                    "exit_reason": [None, "stop_loss", "signal_exit"],
                }
            ).to_csv(run_dir / "trades.csv", index=False)
            result = run_drawdown_attribution(run_dir, min_depth=0.10, top_n=1)
            self.assertTrue(result.report_path.exists())
            self.assertTrue((result.output_dir / "drawdown_attribution.csv").exists())
            self.assertEqual(int(result.metrics["drawdown_episode_count"]), 1)
            self.assertAlmostEqual(float(result.attribution.iloc[0]["realized_pnl"]), -1500.0)
            self.assertEqual(set(result.losing_symbols["code"]), {"000001", "000002"})

    def test_walk_forward_rejects_nonpositive_window_limit(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            config = make_config(Path(tmp))
            with self.assertRaisesRegex(ValueError, "window_limit"):
                run_walk_forward(config, window_limit=0, skip_missing=True)


if __name__ == "__main__":
    unittest.main()
